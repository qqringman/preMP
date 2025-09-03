"""
SFTP ä¸‹è¼‰èˆ‡æ¯”è¼ƒç³»çµ± - Flask Web æ‡‰ç”¨ç¨‹å¼ (å®Œæ•´æ›´æ–°ç‰ˆ)
"""
import os
import shutil
import sys
import json
import threading
import time
import asyncio
from datetime import datetime, timedelta
from flask import Flask, render_template, request, jsonify, send_file, session, url_for, redirect
from flask_socketio import SocketIO, emit, join_room, leave_room
from werkzeug.utils import secure_filename
import pandas as pd
import config
from sftp_downloader import SFTPDownloader
from file_comparator import FileComparator
from zip_packager import ZipPackager
import utils
from flask import make_response
import utils
import openpyxl
from copy import copy
import io
from excel_handler import ExcelHandler
from metadata_manager import metadata_manager
from functools import wraps

# åˆå§‹åŒ– Flask æ‡‰ç”¨
app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'  # è«‹æ›´æ”¹ç‚ºå®‰å…¨çš„å¯†é‘°
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB æœ€å¤§æª”æ¡ˆå¤§å°

# è¨­å®š session å¯†é‘°
app.secret_key = getattr(config, 'SECRET_KEY', 'your-secret-key-change-this-in-production')

# è¨­å®š session éæœŸæ™‚é–“
app.permanent_session_lifetime = getattr(config, 'SESSION_TIMEOUT', 3600)

# åœ¨æª”æ¡ˆé–‹é ­çš„ import å€æ®µåŠ å…¥
from admin_routes import admin_bp

# åœ¨å»ºç«‹ app ä¹‹å¾Œï¼Œè¨»å†Š blueprint
app.register_blueprint(admin_bp)

# åˆå§‹åŒ– SocketIO
socketio = SocketIO(app, cors_allowed_origins="*")

# åˆå§‹åŒ– Excel è™•ç†å™¨ï¼ˆåœ¨ app åˆå§‹åŒ–å¾Œï¼‰
excel_handler = ExcelHandler()

# å„²å­˜ä¸Šå‚³æª”æ¡ˆçš„å…ƒè³‡æ–™
uploaded_excel_metadata = {}

# ç¢ºä¿å¿…è¦çš„ç›®éŒ„å­˜åœ¨
for folder in ['uploads', 'downloads', 'compare_results', 'zip_output', 'logs']:
    if not os.path.exists(folder):
        os.makedirs(folder)

# å…¨åŸŸè®Šæ•¸å­˜å„²è™•ç†é€²åº¦å’Œæ­·å²è¨˜éŒ„
processing_status = {}
recent_activities = []
recent_comparisons = []
task_results = {}  # å„²å­˜ä»»å‹™çµæœä»¥ä¾›æ¨ç´åˆ†æ

class WebProcessor:
    """Web è™•ç†å™¨é¡åˆ¥"""
    
    def __init__(self, task_id):
        self.task_id = task_id
        self.downloader = None
        self.comparator = FileComparator()
        self.packager = ZipPackager()
        self.progress = 0
        self.status = 'idle'
        self.message = ''
        self.results = {}
        self.logger = utils.setup_logger(f'WebProcessor_{task_id}')  # æ·»åŠ  logger

    def update_progress(self, progress, status, message, stats=None, files=None):
        """æ›´æ–°è™•ç†é€²åº¦"""
        self.progress = progress
        self.status = status
        self.message = message
        
        update_data = {
            'progress': progress,
            'status': status,
            'message': message,
            'results': self.results
        }
        
        if stats:
            update_data['stats'] = stats
            
        if files:
            update_data['files'] = files
            
        processing_status[self.task_id] = update_data
        
        # é€é SocketIO ç™¼é€å³æ™‚æ›´æ–°
        socketio.emit('progress_update', {
            'task_id': self.task_id,
            **update_data
        }, room=self.task_id)
        
    def process_one_step(self, excel_file, sftp_config):
        """åŸ·è¡Œä¸€æ­¥åˆ°ä½è™•ç† - ä¿®æ­£æª”æ¡ˆè³‡æ–™ä¿å­˜"""
        try:
            # æ­¥é©Ÿ 1ï¼šä¸‹è¼‰
            self.update_progress(10, 'downloading', 'æ­£åœ¨é€£æ¥ SFTP ä¼ºæœå™¨...')
            
            # ä½¿ç”¨æ–°çš„ Web ä¸‹è¼‰å™¨
            from sftp_web_downloader import SFTPWebDownloader
            
            self.downloader = SFTPWebDownloader(
                sftp_config.get('host', config.SFTP_HOST),
                sftp_config.get('port', config.SFTP_PORT),
                sftp_config.get('username', config.SFTP_USERNAME),
                sftp_config.get('password', config.SFTP_PASSWORD)
            )
            
            # è¨­å®šé€²åº¦å›èª¿
            def progress_callback(progress, status, message, stats=None, files=None):
                if stats:
                    self.update_progress(
                        int(progress * 0.4),  # ä¸‹è¼‰ä½”ç¸½é€²åº¦çš„ 40%
                        status, 
                        message, 
                        stats=stats,
                        files=files
                    )
                else:
                    self.update_progress(int(progress * 0.4), status, message)
            
            self.downloader.set_progress_callback(progress_callback)
            
            self.update_progress(20, 'downloading', 'é–‹å§‹ä¸‹è¼‰æª”æ¡ˆ...')
            
            # å»ºç«‹ä¸‹è¼‰ç›®éŒ„
            download_dir = os.path.join('downloads', self.task_id)
            
            # åŸ·è¡Œä¸‹è¼‰
            report_path = self.downloader.download_from_excel_with_progress(
                excel_file, download_dir
            )
            
            # å–å¾—çµ±è¨ˆè³‡æ–™å’Œæª”æ¡ˆåˆ—è¡¨
            download_data = self.downloader.get_download_stats()
            stats = download_data['stats']
            files = download_data['files']
            
            # å„²å­˜ä¸‹è¼‰çµæœ
            self.results['download_report'] = report_path
            self.results['stats'] = stats
            self.results['files'] = files
            self.results['download_results'] = {
                'stats': stats,
                'files': files,
                'report_path': report_path
            }
            
            self.update_progress(40, 'downloaded', 'ä¸‹è¼‰å®Œæˆï¼', stats, files)
            
            # è™•ç† Excel æª”æ¡ˆè¤‡è£½æ”¹å
            global excel_handler
            if excel_file in uploaded_excel_metadata:
                excel_metadata = uploaded_excel_metadata[excel_file]
                excel_result = excel_handler.process_download_complete(
                    self.task_id,
                    download_dir,
                    excel_metadata
                )
                if excel_result['excel_copied']:
                    self.results['excel_copied'] = True
                    self.results['excel_new_name'] = excel_result['excel_new_name']
            else:
                # ğŸ”¥ æ–°å¢ï¼šå¦‚æœæ²’æœ‰å…ƒè³‡æ–™ï¼Œç›´æ¥è¤‡è£½åŸå§‹ Excel æª”æ¡ˆåˆ°ä¸‹è¼‰ç›®éŒ„
                try:
                    import shutil
                    if os.path.exists(excel_file):
                        original_filename = os.path.basename(excel_file)
                        target_path = os.path.join(download_dir, original_filename)
                        shutil.copy2(excel_file, target_path)
                        self.logger.info(f"å·²è¤‡è£½åŸå§‹ Excel æª”æ¡ˆåˆ°ä¸‹è¼‰ç›®éŒ„: {target_path}")
                    else:
                        self.logger.warning(f"æ‰¾ä¸åˆ°åŸå§‹ Excel æª”æ¡ˆ: {excel_file}")
                except Exception as e:
                    self.logger.error(f"è¤‡è£½ Excel æª”æ¡ˆå¤±æ•—: {str(e)}")
            
            # æ­¥é©Ÿ 2ï¼šæ¯”è¼ƒ
            self.update_progress(50, 'comparing', 'æ­£åœ¨åŸ·è¡Œæ‰€æœ‰æ¯”å°...')
            compare_dir = os.path.join('compare_results', self.task_id)
            
            if not os.path.exists(compare_dir):
                os.makedirs(compare_dir)
                
            all_results = self.comparator.compare_all_scenarios(download_dir, compare_dir)
            
            self.results['compare_results'] = all_results
            self.update_progress(80, 'compared', 'æ¯”å°å®Œæˆï¼')
            
            # æ­¥é©Ÿ 3ï¼šæ‰“åŒ…
            self.update_progress(90, 'packaging', 'æ­£åœ¨æ‰“åŒ…çµæœ...')
            timestamp = utils.get_timestamp()
            zip_name = f"all_results_{timestamp}.zip"
            zip_path = os.path.join('zip_output', self.task_id, zip_name)
            
            # å»ºç«‹è‡¨æ™‚ç›®éŒ„ä¾†æ•´åˆæ‰€æœ‰å…§å®¹
            import tempfile
            import shutil
            
            with tempfile.TemporaryDirectory() as temp_dir:
                # è¤‡è£½ä¸‹è¼‰çš„æª”æ¡ˆ
                if os.path.exists(download_dir):
                    dest_downloads = os.path.join(temp_dir, 'downloads')
                    shutil.copytree(download_dir, dest_downloads)
                
                # è¤‡è£½æ¯”å°çµæœ
                if os.path.exists(compare_dir):
                    dest_compare = os.path.join(temp_dir, 'compare_results')
                    shutil.copytree(compare_dir, dest_compare)
                
                # ç¢ºä¿ ZIP è¼¸å‡ºç›®éŒ„å­˜åœ¨
                zip_output_dir = os.path.dirname(zip_path)
                if not os.path.exists(zip_output_dir):
                    os.makedirs(zip_output_dir)
                
                # æ‰“åŒ…
                zip_path = self.packager.create_zip(temp_dir, zip_path)
            
            self.results['zip_file'] = zip_path
            
            # ç¢ºä¿æœ€çµ‚æ›´æ–°åŒ…å«å®Œæ•´çš„æª”æ¡ˆè³‡æ–™
            final_update_data = {
                'progress': 100,
                'status': 'completed',
                'message': 'æ‰€æœ‰è™•ç†å®Œæˆï¼',
                'results': self.results,
                'stats': stats,
                'files': files
            }
            
            processing_status[self.task_id] = final_update_data
            
            # é€é SocketIO ç™¼é€æœ€çµ‚æ›´æ–°
            socketio.emit('progress_update', {
                'task_id': self.task_id,
                **final_update_data
            }, room=self.task_id)
            
            # è¨˜éŒ„åˆ°æœ€è¿‘æ´»å‹•
            add_activity('å®Œæˆä¸€æ­¥åˆ°ä½è™•ç†', 'success', 
                        f'ä¸‹è¼‰ {stats["downloaded"]} å€‹æª”æ¡ˆï¼Œå®Œæˆ 3 å€‹æ¯”å°æƒ…å¢ƒï¼Œä»»å‹™ {self.task_id}')
            
            # åŒæ™‚è¨˜éŒ„åˆ°æœ€è¿‘æ¯”å°è¨˜éŒ„
            add_comparison(self.task_id, 'å®Œæˆä¸€æ­¥åˆ°ä½è™•ç†', 'completed', stats["downloaded"])
            
            # å„²å­˜çµæœä¾›æ¨ç´åˆ†æä½¿ç”¨
            save_task_results(self.task_id, self.results)
            
        except Exception as e:
            self.logger.error(f"One-step processing error: {str(e)}")
            self.update_progress(0, 'error', f'è™•ç†å¤±æ•—ï¼š{str(e)}')
            
            # è¨˜éŒ„å¤±æ•—åˆ°æœ€è¿‘æ´»å‹•
            add_activity('ä¸€æ­¥åˆ°ä½è™•ç†å¤±æ•—', 'error', f'{str(e)}ï¼Œä»»å‹™ {self.task_id}')
            
            # åŒæ™‚è¨˜éŒ„å¤±æ•—åˆ°æœ€è¿‘æ¯”å°è¨˜éŒ„
            add_comparison(self.task_id, 'ä¸€æ­¥åˆ°ä½è™•ç†å¤±æ•—', 'error', 0)
            raise
            
    def process_download(self, excel_file, sftp_config, options):
        """åŸ·è¡Œä¸‹è¼‰è™•ç† - åŒ…å« Excel æª”æ¡ˆè¤‡è£½æ”¹ååŠŸèƒ½"""
        try:
            self.update_progress(10, 'downloading', 'æ­£åœ¨é€£æ¥ SFTP ä¼ºæœå™¨...')
            
            # ä½¿ç”¨æ–°çš„ Web ä¸‹è¼‰å™¨
            from sftp_web_downloader import SFTPWebDownloader
            
            self.downloader = SFTPWebDownloader(
                sftp_config.get('host', config.SFTP_HOST),
                sftp_config.get('port', config.SFTP_PORT),
                sftp_config.get('username', config.SFTP_USERNAME),
                sftp_config.get('password', config.SFTP_PASSWORD)
            )
            
            # è¨­å®šé€²åº¦å›èª¿
            def progress_callback(progress, status, message, stats=None, files=None):
                if stats:
                    self.update_progress(
                        int(progress),
                        status, 
                        message, 
                        stats=stats,
                        files=files
                    )
                else:
                    self.update_progress(int(progress), status, message)
            
            self.downloader.set_progress_callback(progress_callback)
            
            self.update_progress(20, 'downloading', 'é–‹å§‹ä¸‹è¼‰æª”æ¡ˆ...')
            
            # å»ºç«‹ä¸‹è¼‰ç›®éŒ„
            download_dir = os.path.join('downloads', self.task_id)
            
            try:
                # åŸ·è¡Œä¸‹è¼‰
                report_path = self.downloader.download_from_excel_with_progress(
                    excel_file, download_dir
                )
                
                # å–å¾—çµ±è¨ˆè³‡æ–™å’Œæª”æ¡ˆåˆ—è¡¨
                download_data = self.downloader.get_download_stats()
                stats = download_data['stats']
                files = download_data['files']
                
                # ===== è™•ç† Excel æª”æ¡ˆè¤‡è£½æ”¹å =====
                self.logger.info("=" * 60)
                self.logger.info("é–‹å§‹è™•ç† Excel æª”æ¡ˆæ”¹å")
                self.logger.info(f"  Task ID: {self.task_id}")
                self.logger.info(f"  ä¸‹è¼‰è³‡æ–™å¤¾: {download_dir}")
                self.logger.info(f"  Excel æª”æ¡ˆè·¯å¾‘: {excel_file}")
                
                # é‡æ–°æª¢æŸ¥ Excel æª”æ¡ˆçš„æ¬„ä½ï¼ˆé€™æœƒåŒ…å« filepathï¼‰
                global excel_handler
                excel_check_result = excel_handler.check_excel_columns(excel_file)
                
                # ç¢ºä¿ filepath å­˜åœ¨ï¼ˆå‘å¾Œç›¸å®¹ï¼‰
                if 'filepath' not in excel_check_result or not excel_check_result['filepath']:
                    excel_check_result['filepath'] = excel_file
                    self.logger.info(f"  è£œå…… filepath: {excel_file}")
                
                # è™•ç† Excel æª”æ¡ˆè¤‡è£½æ”¹å
                excel_result = excel_handler.process_download_complete(
                    self.task_id,
                    download_dir,
                    excel_check_result
                )
                
                # ç”Ÿæˆè³‡æ–™å¤¾çµæ§‹
                folder_structure = self._generate_folder_structure(download_dir, report_path)
                
                # å„²å­˜çµæœ
                self.results['download_report'] = report_path
                self.results['stats'] = stats
                self.results['files'] = files
                self.results['folder_structure'] = folder_structure
                
                # å¦‚æœæœ‰ Excel è¤‡è£½çµæœï¼ŒåŠ å…¥åˆ°çµæœä¸­
                if excel_result['excel_copied']:
                    self.results['excel_copied'] = True
                    self.results['excel_new_name'] = excel_result['excel_new_name']
                    # æ›´æ–°è¨Šæ¯
                    self.update_progress(95, 'downloading', 
                        f'Excel æª”æ¡ˆå·²å¦å­˜ç‚º: {excel_result["excel_new_name"]}')
                    self.logger.info(f"Excel æª”æ¡ˆè¤‡è£½æˆåŠŸ: {excel_result['excel_new_name']}")
                else:
                    self.logger.info(f"{excel_result['message']}")
                
                self.logger.info("=" * 60)
                
                # ç¢ºä¿æœ€çµ‚çµ±è¨ˆæ­£ç¢º
                self.update_progress(100, 'completed', 'ä¸‹è¼‰å®Œæˆï¼', stats, files)
                
                # è¨˜éŒ„åˆ°æœ€è¿‘æ´»å‹•
                add_activity('å®Œæˆæª”æ¡ˆä¸‹è¼‰', 'success', 
                            f'æˆåŠŸä¸‹è¼‰ {stats["downloaded"]} å€‹æª”æ¡ˆï¼Œä»»å‹™ {self.task_id}')
                
                # åŒæ™‚è¨˜éŒ„åˆ°æœ€è¿‘æ¯”å°è¨˜éŒ„
                add_comparison(self.task_id, 'å®Œæˆæª”æ¡ˆä¸‹è¼‰', 'completed', stats["downloaded"])
                
            except Exception as e:
                error_msg = str(e)
                current_stats = self.downloader.get_download_stats()
                self.update_progress(0, 'error', f'ä¸‹è¼‰å¤±æ•—ï¼š{error_msg}', 
                                stats=current_stats['stats'])
                
                # è¨˜éŒ„å¤±æ•—åˆ°æœ€è¿‘æ´»å‹•
                add_activity('æª”æ¡ˆä¸‹è¼‰å¤±æ•—', 'error', f'{error_msg}ï¼Œä»»å‹™ {self.task_id}')
                
                # åŒæ™‚è¨˜éŒ„å¤±æ•—åˆ°æœ€è¿‘æ¯”å°è¨˜éŒ„
                add_comparison(self.task_id, 'æª”æ¡ˆä¸‹è¼‰å¤±æ•—', 'error', 0)
                raise
                
        except Exception as e:
            error_msg = str(e)
            self.update_progress(0, 'error', f'è™•ç†å¤±æ•—ï¼š{error_msg}')
            
            # è¨˜éŒ„å¤±æ•—åˆ°æœ€è¿‘æ´»å‹•
            add_activity('æª”æ¡ˆä¸‹è¼‰å¤±æ•—', 'error', f'{error_msg}ï¼Œä»»å‹™ {self.task_id}')
            
            # åŒæ™‚è¨˜éŒ„å¤±æ•—åˆ°æœ€è¿‘æ¯”å°è¨˜éŒ„
            add_comparison(self.task_id, 'æª”æ¡ˆä¸‹è¼‰å¤±æ•—', 'error', 0)
            raise

    def _handle_excel_copy_rename(self, excel_file, download_dir):
        """
        è™•ç† Excel æª”æ¡ˆçš„è¤‡è£½å’Œæ”¹å
        
        Args:
            excel_file: Excel æª”æ¡ˆè·¯å¾‘
            download_dir: ä¸‹è¼‰ç›®éŒ„
            
        Returns:
            è™•ç†çµæœå­—å…¸
        """
        result = {
            'excel_copied': False,
            'excel_new_name': None
        }
        
        try:
            # å¾å…¨åŸŸè®Šæ•¸ç²å– Excel å…ƒè³‡æ–™
            global uploaded_excel_metadata, excel_handler
            
            if excel_file in uploaded_excel_metadata:
                excel_metadata = uploaded_excel_metadata[excel_file]
                
                self.logger.info("=" * 60)
                self.logger.info(f"é–‹å§‹è™•ç† Excel æª”æ¡ˆæ”¹å")
                self.logger.info(f"  Task ID: {self.task_id}")
                self.logger.info(f"  ä¸‹è¼‰è³‡æ–™å¤¾: {download_dir}")
                
                # ä½¿ç”¨ ExcelHandler è™•ç†
                process_result = excel_handler.process_download_complete(
                    self.task_id,
                    download_dir,
                    excel_metadata
                )
                
                result.update(process_result)
                
                if result['excel_copied']:
                    self.logger.info(f"âœ… Excel æª”æ¡ˆå·²æˆåŠŸè¤‡è£½ä¸¦æ”¹å: {result['excel_new_name']}")
                    
                    # åŠ å…¥æ´»å‹•è¨˜éŒ„
                    add_activity('Excel æª”æ¡ˆè™•ç†', 'success', 
                            f"æª”æ¡ˆå·²å¦å­˜ç‚º: {result['excel_new_name']}")
                else:
                    self.logger.info(f"â„¹ï¸ {process_result.get('message', 'ä¸éœ€è¦è™•ç† Excel æª”æ¡ˆ')}")
                    
            else:
                self.logger.info(f"æ²’æœ‰æ‰¾åˆ° Excel å…ƒè³‡æ–™: {excel_file}")
                
        except Exception as e:
            self.logger.error(f"è™•ç† Excel æª”æ¡ˆæ™‚ç™¼ç”ŸéŒ¯èª¤: {str(e)}")
            # ä¸è¦å› ç‚º Excel è™•ç†å¤±æ•—è€Œä¸­æ–·æ•´å€‹ä¸‹è¼‰æµç¨‹
            
        return result
        
    def _generate_simple_folder_structure(self, download_dir):
        """ç”Ÿæˆç°¡å–®çš„è³‡æ–™å¤¾çµæ§‹"""
        structure = {}
        
        try:
            for item in os.listdir(download_dir):
                item_path = os.path.join(download_dir, item)
                if os.path.isdir(item_path):
                    # éè¿´è™•ç†å­ç›®éŒ„
                    structure[item] = {}
                    for subitem in os.listdir(item_path):
                        structure[item][subitem] = subitem
                else:
                    # æª”æ¡ˆ
                    if not item.endswith('.xlsx'):  # æ’é™¤å ±å‘Šæª”æ¡ˆ
                        structure[item] = item
        except Exception as e:
            print(f"Error generating folder structure: {e}")
            
        return structure
        
    def _generate_folder_structure(self, download_dir, report_path):
        """ç”Ÿæˆè³‡æ–™å¤¾çµæ§‹"""
        folder_structure = {}
        
        try:
            # ç¢ºä¿ç›®éŒ„å­˜åœ¨
            if not os.path.exists(download_dir):
                return folder_structure
                
            # éè¿´å»ºç«‹è³‡æ–™å¤¾çµæ§‹
            def build_tree(path, tree_dict):
                try:
                    items = os.listdir(path)
                    for item in sorted(items):
                        item_path = os.path.join(path, item)
                        
                        # è·³éå ±å‘Šæª”æ¡ˆ
                        if report_path and item_path == report_path:
                            continue
                            
                        if os.path.isdir(item_path):
                            # è³‡æ–™å¤¾
                            tree_dict[item] = {}
                            build_tree(item_path, tree_dict[item])
                        else:
                            # æª”æ¡ˆ
                            tree_dict[item] = os.path.relpath(item_path, os.getcwd())
                except Exception as e:
                    self.logger.error(f"Error building tree for {path}: {e}")
                    
            build_tree(download_dir, folder_structure)
                        
        except Exception as e:
            self.logger.error(f"Error generating folder structure: {e}")
            print(f"Error generating folder structure: {e}")
            
        return folder_structure
        
    def _get_download_stats(self, report_path, download_dir):
        """å¾ä¸‹è¼‰å ±å‘Šæˆ–ç›®éŒ„ä¸­ç²å–çµ±è¨ˆè³‡æ–™"""
        stats = {
            'total': 0,
            'downloaded': 0,
            'skipped': 0,
            'failed': 0
        }
        
        # å˜—è©¦å¾å ±å‘Šæª”æ¡ˆè®€å–çµ±è¨ˆ
        if report_path and os.path.exists(report_path):
            try:
                import pandas as pd
                df = pd.read_excel(report_path)
                stats['total'] = len(df)
                
                # æª¢æŸ¥å¯èƒ½çš„ç‹€æ…‹æ¬„ä½åç¨±
                status_columns = ['status', 'Status', 'ç‹€æ…‹', 'download_status', 'result']
                status_column = None
                
                for col in status_columns:
                    if col in df.columns:
                        status_column = col
                        break
                
                if status_column:
                    # çµ±è¨ˆå„ç¨®ç‹€æ…‹
                    for status in df[status_column].unique():
                        status_lower = str(status).lower()
                        if 'download' in status_lower or 'success' in status_lower or 'æˆåŠŸ' in status:
                            stats['downloaded'] += len(df[df[status_column] == status])
                        elif 'skip' in status_lower or 'è·³é' in status:
                            stats['skipped'] += len(df[df[status_column] == status])
                        elif 'fail' in status_lower or 'error' in status_lower or 'å¤±æ•—' in status:
                            stats['failed'] += len(df[df[status_column] == status])
                else:
                    # å¦‚æœæ²’æœ‰ç‹€æ…‹æ¬„ä½ï¼Œè¨ˆç®—ä¸‹è¼‰ç›®éŒ„ä¸­çš„æª”æ¡ˆæ•¸
                    file_count = 0
                    for root, dirs, files in os.walk(download_dir):
                        # æ’é™¤å ±å‘Šæª”æ¡ˆ
                        files = [f for f in files if not f.endswith('_report.xlsx')]
                        file_count += len(files)
                    stats['downloaded'] = file_count
                    
            except Exception as e:
                print(f"Error reading report: {e}")
                # å¦‚æœç„¡æ³•è®€å–å ±å‘Šï¼Œä½¿ç”¨é è¨­å€¼
                stats['downloaded'] = stats['total']
        
        return stats
                    
    def process_comparison(self, source_dir, scenarios):
        """åŸ·è¡Œæ¯”å°è™•ç†"""
        
        # åœ¨æ–¹æ³•é–‹å§‹å°± import æ‰€éœ€çš„æ¨¡çµ„
        import pandas as pd
        from file_comparator import FileComparator
        
        try:
            self.update_progress(10, 'comparing', 'æ­£åœ¨æº–å‚™æ¯”å°...')
            
            compare_dir = os.path.join('compare_results', self.task_id)
            
            if not os.path.exists(compare_dir):
                os.makedirs(compare_dir, exist_ok=True)
            
            if scenarios == 'all':
                self.update_progress(30, 'comparing', 'æ­£åœ¨åŸ·è¡Œæ‰€æœ‰æ¯”å°æƒ…å¢ƒ...')
                all_results = self.comparator.compare_all_scenarios(source_dir, compare_dir)
                
                # ç›´æ¥ä½¿ç”¨ FileComparator è¿”å›çš„çµæœ
                self.results['compare_results'] = all_results
                
                # å¦‚æœæœ‰ç¸½æ‘˜è¦å ±å‘Šï¼Œä¿å­˜è·¯å¾‘
                if 'summary_report' in all_results:
                    self.results['summary_report'] = all_results['summary_report']
            
            else:
                # å–®ä¸€æƒ…å¢ƒæ¯”å°
                self.update_progress(30, 'comparing', f'æ­£åœ¨åŸ·è¡Œ {scenarios} æ¯”å°æƒ…å¢ƒ...')
                
                # ğŸ”¥ é—œéµä¿®æ­£ï¼šæ¨¡ä»¿ compare_all_scenarios çš„é‚è¼¯
                try:
                    # è¼‰å…¥ mapping tables ä¸¦ä¿å­˜åˆ° comparator
                    self.logger.info(f"ğŸ” æº–å‚™è¼‰å…¥ mapping tablesï¼Œsource_dir: {source_dir}")
                    self.comparator.mapping_tables = self.comparator._load_mapping_tables(source_dir)
                    self.logger.info(f"ğŸ“Š è¼‰å…¥çµæœ: {len(self.comparator.mapping_tables) if self.comparator.mapping_tables else 0} å€‹ mapping tables")
                    
                    # ğŸ”¥ é—œéµï¼šå¦‚æœæœ‰ mapping tableï¼Œä½¿ç”¨ _compare_with_mapping çš„é‚è¼¯
                    if self.comparator.mapping_tables:
                        self.logger.info("âœ… æ‰¾åˆ° mapping tablesï¼Œä½¿ç”¨ mapping é‚è¼¯")
                        
                        # ä½¿ç”¨ compare_with_mappingï¼Œç„¶å¾Œç¯©é¸ç‰¹å®šæƒ…å¢ƒçš„çµæœ
                        all_results = self.comparator._compare_with_mapping(source_dir, compare_dir)
                        
                        # å¾å®Œæ•´çµæœä¸­æå–å–®ä¸€æƒ…å¢ƒ
                        scenario_map = {
                            'master_vs_premp': 'master_vs_premp',
                            'premp_vs_wave': 'premp_vs_wave',
                            'wave_vs_backup': 'wave_vs_backup'
                        }
                        
                        if scenarios in scenario_map:
                            scenario_key = scenario_map[scenarios]
                            
                            # æå–å–®ä¸€æƒ…å¢ƒçµæœ
                            if scenario_key in all_results:
                                single_scenario_result = {scenario_key: all_results[scenario_key]}
                                self.results['compare_results'] = single_scenario_result
                                
                                # å°‹æ‰¾å°æ‡‰çš„å ±è¡¨æª”æ¡ˆ
                                scenario_report_path = os.path.join(compare_dir, scenario_key, 'all_scenarios_compare.xlsx')
                                if os.path.exists(scenario_report_path):
                                    self.results['summary_report'] = scenario_report_path
                                elif 'summary_report' in all_results:
                                    self.results['summary_report'] = all_results['summary_report']
                                    
                                self.logger.info(f"âœ… æˆåŠŸä½¿ç”¨ mapping é‚è¼¯è™•ç† {scenario_key}")
                            else:
                                raise Exception(f"åœ¨çµæœä¸­æ‰¾ä¸åˆ°æƒ…å¢ƒ: {scenario_key}")
                        else:
                            raise Exception(f"ä¸æ”¯æ´çš„æƒ…å¢ƒ: {scenarios}")
                            
                    else:
                        # æ²’æœ‰ mapping tableï¼Œä½¿ç”¨åŸæœ‰é‚è¼¯
                        self.logger.warning("âš ï¸ æ²’æœ‰è¼‰å…¥ mapping tableï¼Œä½¿ç”¨åŸæœ‰é‚è¼¯ï¼ˆå¯èƒ½æœƒå¤±æ•—ï¼‰")
                        success = self._process_single_scenario_without_mapping(source_dir, compare_dir, scenarios)
                        if not success:
                            raise Exception("æ²’æœ‰ mapping table çš„æƒ…æ³ä¸‹è™•ç†å¤±æ•—")
                        
                except Exception as e:
                    self.logger.error(f"âŒ å–®ä¸€æƒ…å¢ƒæ¯”å°å¤±æ•—: {str(e)}")
                    # è¨­å®šå¤±æ•—çš„çµæœçµæ§‹
                    scenario_key = {'master_vs_premp': 'master_vs_premp', 'premp_vs_wave': 'premp_vs_wave', 'wave_vs_backup': 'wave_vs_backup'}.get(scenarios, scenarios)
                    self.results['compare_results'] = {
                        scenario_key: {
                            'success': 0,
                            'failed': 0,
                            'modules': [],
                            'failed_modules': []
                        }
                    }
                    raise
            
            self.update_progress(100, 'completed', 'æ¯”å°å®Œæˆï¼')
            
            # è¨˜éŒ„æ¯”å°ï¼Œè¨ˆç®—æ­£ç¢ºçš„ç¸½æ¨¡çµ„æ•¸
            total_modules = 0
            total_failed = 0
            
            # ç¢ºä¿ compare_results å­˜åœ¨
            if 'compare_results' in self.results:
                for scenario_data in self.results['compare_results'].values():
                    if isinstance(scenario_data, dict):
                        total_modules += scenario_data.get('success', 0)
                        total_failed += scenario_data.get('failed', 0)
            
            add_comparison(self.task_id, scenarios, 'completed', total_modules)
            
            # å„²å­˜çµæœ
            save_task_results(self.task_id, self.results)
            
        except Exception as e:
            self.logger.error(f"Comparison error: {str(e)}")
            self.update_progress(0, 'error', f'æ¯”å°å¤±æ•—ï¼š{str(e)}')
            
            # ç¢ºä¿å³ä½¿å‡ºéŒ¯ä¹Ÿæœ‰åŸºæœ¬çš„çµæ§‹ï¼Œé¿å…å¾ŒçºŒ KeyError
            if 'compare_results' not in self.results:
                self.results['compare_results'] = {}
                
            raise

    def _process_single_scenario_without_mapping(self, source_dir, compare_dir, scenarios):
        """æ²’æœ‰ mapping table æ™‚çš„å¾Œå‚™è™•ç†æ–¹å¼"""
        # é€™è£¡å¯ä»¥å¯¦ä½œåŸæœ‰çš„é‚è¼¯ï¼Œä½†é€šå¸¸åœ¨æœ‰ mapping table çš„æƒ…æ³ä¸‹ä¸æœƒç”¨åˆ°
        self.logger.warning("ä½¿ç”¨å¾Œå‚™é‚è¼¯è™•ç†å–®ä¸€æƒ…å¢ƒï¼ˆé€šå¸¸æœƒå¤±æ•—ï¼‰")
        return False

    def _get_scenario_display_name(self, scenario):
        """å–å¾—æƒ…å¢ƒçš„é¡¯ç¤ºåç¨±"""
        name_map = {
            'master_vs_premp': 'Master vs PreMP',
            'premp_vs_wave': 'PreMP vs Wave',
            'wave_vs_backup': 'Wave vs Backup'
        }
        return name_map.get(scenario, scenario)

# è¼”åŠ©å‡½æ•¸
def add_activity(action, status, details=None):
    """æ·»åŠ æ´»å‹•è¨˜éŒ„"""
    activity = {
        'timestamp': datetime.now(),
        'action': action,
        'status': status,
        'details': details
    }
    recent_activities.insert(0, activity)
    # åªä¿ç•™æœ€è¿‘ 20 ç­†
    if len(recent_activities) > 20:
        recent_activities.pop()

def add_comparison(task_id, scenario, status, modules):
    """æ·»åŠ æ¯”å°è¨˜éŒ„"""
    comparison = {
        'id': task_id,  # ä½¿ç”¨å‚³å…¥çš„ task_id
        'task_id': task_id,  # æ˜ç¢ºä¿å­˜ task_id
        'timestamp': datetime.now(),
        'scenario': scenario,
        'status': status,
        'modules': modules,
        'duration': '< 1 åˆ†é˜'
    }
    recent_comparisons.insert(0, comparison)
    # åªä¿ç•™æœ€è¿‘ 10 ç­†
    if len(recent_comparisons) > 10:
        recent_comparisons.pop()

def save_task_results(task_id, results):
    """å„²å­˜ä»»å‹™çµæœä¾›æ¨ç´åˆ†æä½¿ç”¨"""
    # ç¢ºä¿å„²å­˜å®Œæ•´çš„çµæœ
    task_results[task_id] = {
        'results': results,
        'summary_report': results.get('summary_report', ''),
        'compare_results': results.get('compare_results', {}),
        'timestamp': datetime.now()
    }
    
    # åŒæ™‚æ›´æ–° processing_status ä»¥ç¢ºä¿è³‡æ–™æŒä¹…æ€§
    if task_id in processing_status:
        processing_status[task_id]['results'] = results

def global_login_required(f):
    """å…¨åŸŸç™»å…¥æª¢æŸ¥è£é£¾å™¨"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # å¦‚æœæ²’æœ‰å•Ÿç”¨ç™»å…¥åŠŸèƒ½ï¼Œç›´æ¥é€šé
        if not getattr(config, 'ENABLE_LOGIN', False):
            return f(*args, **kwargs)
        
        # å¦‚æœæ˜¯å…¨åŸŸç™»å…¥æ¨¡å¼
        if getattr(config, 'LOGIN_MODE', 'admin_only') == 'global':
            # æª¢æŸ¥ç™»å…¥ç‹€æ…‹
            if 'logged_in' not in session or not session['logged_in']:
                if request.is_json:
                    return jsonify({'error': 'è«‹å…ˆç™»å…¥', 'redirect': '/login'}), 401
                return redirect(url_for('admin.login_page', redirect=request.url))
        
        return f(*args, **kwargs)
    return decorated_function

# è·¯ç”±å®šç¾©
@app.route('/')
@global_login_required
def index():
    """é¦–é """
    return render_template('index.html')

@app.route('/compare')
@global_login_required
def compare_page():
    """æ¯”è¼ƒé é¢"""
    return render_template('compare.html')

@app.route('/one-step')
@global_login_required
def one_step_page():
    """ä¸€æ­¥åˆ°ä½é é¢"""
    return render_template('one_step.html')

@app.route('/results/<task_id>')
@global_login_required
def results_page(task_id):
    """çµæœé é¢"""
    return render_template('results.html', task_id=task_id)

# API ç«¯é»
@app.route('/api/upload', methods=['POST'])
def upload_file():
    """ä¸Šå‚³æª”æ¡ˆ API - æ”¯æ´ Excel å’Œ CSVï¼Œä¸¦æª¢æŸ¥æ¬„ä½"""
    if 'file' not in request.files:
        return jsonify({'error': 'æ²’æœ‰æª”æ¡ˆ'}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'æ²’æœ‰é¸æ“‡æª”æ¡ˆ'}), 400
    
    # æ”¯æ´çš„æª”æ¡ˆæ ¼å¼
    allowed_extensions = {'.xlsx', '.xls', '.csv'}
    file_ext = os.path.splitext(file.filename)[1].lower()
    
    if file and file_ext in allowed_extensions:
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        # ç¢ºä¿ä¸Šå‚³ç›®éŒ„å­˜åœ¨
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        
        file.save(filepath)
        
        # æª¢æŸ¥ Excel æ¬„ä½
        excel_metadata = {
            'original_name': file.filename,
            'filepath': filepath,
            'has_sftp_columns': False,
            'root_folder': None
        }
        
        # å¦‚æœæ˜¯ Excel æˆ– CSV æª”æ¡ˆï¼Œæª¢æŸ¥æ¬„ä½
        if file_ext in ['.xlsx', '.xls', '.csv']:
            try:
                check_result = excel_handler.check_excel_columns(filepath)
                excel_metadata.update(check_result)
                app.logger.info(f"Excel æ¬„ä½æª¢æŸ¥çµæœ: {check_result}")
            except Exception as e:
                app.logger.warning(f"æª¢æŸ¥ Excel æ¬„ä½æ™‚ç™¼ç”ŸéŒ¯èª¤: {str(e)}")
        
        # ä½¿ç”¨å…ƒè³‡æ–™ç®¡ç†å™¨å„²å­˜
        metadata_manager.store_metadata(filepath, excel_metadata)
        
        app.logger.info(f'æª”æ¡ˆä¸Šå‚³: {filename} (é¡å‹: {file_ext})')
        
        return jsonify({
            'filename': filename, 
            'filepath': filepath,
            'file_type': file_ext[1:],
            'excel_metadata': excel_metadata
        })
    
    return jsonify({
        'error': f'åªæ”¯æ´ Excel (.xlsx, .xls) å’Œ CSV (.csv) æª”æ¡ˆï¼Œæ‚¨ä¸Šå‚³çš„æ˜¯ {file_ext} æª”æ¡ˆ'
    }), 400

@app.route('/api/test-connection', methods=['POST'])
def test_connection():
    """æ¸¬è©¦ SFTP é€£ç·š"""
    data = request.json
    try:
        downloader = SFTPDownloader(
            data.get('host', config.SFTP_HOST),
            data.get('port', config.SFTP_PORT),
            data.get('username', config.SFTP_USERNAME),
            data.get('password', config.SFTP_PASSWORD)
        )
        if downloader.test_connection():
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'message': 'é€£ç·šå¤±æ•—'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/one-step', methods=['POST'])
def process_one_step():
    """ä¸€æ­¥åˆ°ä½è™•ç† API"""
    data = request.json
    excel_file = data.get('excel_file')
    sftp_config = data.get('sftp_config', {})
    
    if not excel_file:
        return jsonify({'error': 'ç¼ºå°‘ Excel æª”æ¡ˆ'}), 400
        
    # ç”Ÿæˆä»»å‹™ ID
    task_id = f"task_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.urandom(4).hex()}"
    
    # åœ¨èƒŒæ™¯åŸ·è¡Œè™•ç†
    processor = WebProcessor(task_id)
    thread = threading.Thread(
        target=processor.process_one_step,
        args=(excel_file, sftp_config)
    )
    thread.start()
    
    return jsonify({'task_id': task_id})

@app.route('/api/download', methods=['POST'])
def process_download():
    """ä¸‹è¼‰è™•ç† API - åŒ…å« Excel å…ƒè³‡æ–™è™•ç†"""
    try:
        data = request.json
        
        # é©—è­‰å¿…è¦åƒæ•¸
        excel_file = data.get('excel_file')
        if not excel_file:
            return jsonify({'error': 'ç¼ºå°‘ Excel æª”æ¡ˆ'}), 400
            
        sftp_config = data.get('sftp_config', {})
        options = data.get('options', {})
        
        # ç²å– Excel å…ƒè³‡æ–™ï¼ˆå¾å‰ç«¯å‚³ä¾†æˆ–å¾å…¨åŸŸè®Šæ•¸å–å¾—ï¼‰
        excel_metadata = data.get('excel_metadata')
        if not excel_metadata and excel_file in uploaded_excel_metadata:
            excel_metadata = uploaded_excel_metadata[excel_file]
        
        # ç”Ÿæˆä»»å‹™ ID
        task_id = f"task_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.urandom(4).hex()}"
        
        # å¦‚æœæœ‰å…ƒè³‡æ–™ï¼Œå„²å­˜åˆ°å…¨åŸŸè®Šæ•¸ä¸­
        if excel_metadata:
            uploaded_excel_metadata[excel_file] = excel_metadata
        
        # åœ¨èƒŒæ™¯åŸ·è¡Œè™•ç†
        processor = WebProcessor(task_id)
        thread = threading.Thread(
            target=processor.process_download,
            args=(excel_file, sftp_config, options)
        )
        thread.start()
        
        return jsonify({'task_id': task_id})
        
    except Exception as e:
        print(f"Download API error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/compare', methods=['POST'])
def process_compare():
    """æ¯”è¼ƒè™•ç† API"""
    data = request.json
    source_dir = data.get('source_dir')
    scenarios = data.get('scenarios', 'all')
    
    if not source_dir:
        return jsonify({'error': 'ç¼ºå°‘ä¾†æºç›®éŒ„'}), 400
        
    # ç”Ÿæˆä»»å‹™ ID
    task_id = f"task_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.urandom(4).hex()}"
    
    # åœ¨èƒŒæ™¯åŸ·è¡Œè™•ç†
    processor = WebProcessor(task_id)
    thread = threading.Thread(
        target=processor.process_comparison,
        args=(source_dir, scenarios)
    )
    thread.start()
    
    return jsonify({'task_id': task_id})

@app.route('/api/status/<task_id>')
def get_status(task_id):
    """å–å¾—ä»»å‹™ç‹€æ…‹ API - å¢å¼·ç‰ˆï¼Œæ”¯æ´å¾æ–‡ä»¶ç³»çµ±æ¢å¾©ä»»å‹™ç‹€æ…‹"""
    
    # 1. é¦–å…ˆæª¢æŸ¥è¨˜æ†¶é«”ä¸­çš„ç‹€æ…‹
    if task_id in processing_status:
        return jsonify(processing_status[task_id])
    
    # 2. å¦‚æœè¨˜æ†¶é«”ä¸­æ²’æœ‰ï¼Œå˜—è©¦å¾æ–‡ä»¶ç³»çµ±æ¢å¾©ä»»å‹™ç‹€æ…‹
    try:
        task_status = recover_task_status_from_filesystem(task_id)
        if task_status:
            # å°‡æ¢å¾©çš„ç‹€æ…‹å­˜å›è¨˜æ†¶é«”ä¸­ï¼Œä¾›å¾ŒçºŒä½¿ç”¨
            processing_status[task_id] = task_status
            return jsonify(task_status)
    except Exception as e:
        app.logger.error(f'Error recovering task status for {task_id}: {str(e)}')
    
    # 3. å¦‚æœéƒ½æ‰¾ä¸åˆ°ï¼Œè¿”å› not_found ç‹€æ…‹
    return jsonify({
        'progress': 0,
        'status': 'not_found',
        'message': 'æ‰¾ä¸åˆ°ä»»å‹™',
        'task_id': task_id
    })

def recover_task_status_from_filesystem(task_id):
    """å¾æ–‡ä»¶ç³»çµ±æ¢å¾©ä»»å‹™ç‹€æ…‹ - å¢å¼·ä¸€æ­¥åˆ°ä½æ”¯æ´"""
    if not task_id.startswith('task_'):
        return None
    
    # æª¢æŸ¥ä¸‹è¼‰ç›®éŒ„
    download_dir = os.path.join('downloads', task_id)
    compare_dir = os.path.join('compare_results', task_id)
    
    task_info = {
        'task_id': task_id,
        'progress': 100,
        'status': 'completed',
        'message': 'ä»»å‹™å·²å®Œæˆï¼ˆå¾æ–‡ä»¶ç³»çµ±æ¢å¾©ï¼‰',
        'results': {}
    }
    
    # æª¢æŸ¥æ˜¯å¦æœ‰ä¸‹è¼‰çµæœ
    if os.path.exists(download_dir):
        app.logger.info(f'Found download directory for task {task_id}')
        
        # çµ±è¨ˆä¸‹è¼‰çš„æ–‡ä»¶
        download_stats = analyze_download_directory(download_dir)
        
        # ä¿å­˜åˆ°å¤šå€‹ä½ç½®ç¢ºä¿ç›¸å®¹æ€§
        task_info['results']['stats'] = download_stats['stats']
        task_info['results']['files'] = download_stats['files']
        task_info['results']['download_results'] = {
            'stats': download_stats['stats'],
            'files': download_stats['files']
        }
        
        # åŒæ™‚åœ¨é ‚å±¤ä¿å­˜çµ±è¨ˆè³‡æ–™ï¼ˆä¾›ä¸‹è¼‰é é¢ä½¿ç”¨ï¼‰
        task_info['stats'] = download_stats['stats']
        task_info['files'] = download_stats['files']
        
        # æŸ¥æ‰¾ä¸‹è¼‰å ±å‘Š
        report_files = [f for f in os.listdir(download_dir) if f.endswith('_report.xlsx')]
        if report_files:
            task_info['results']['download_report'] = os.path.join(download_dir, report_files[0])
        
        # ç”Ÿæˆæ–‡ä»¶å¤¾çµæ§‹
        task_info['results']['folder_structure'] = generate_folder_structure_from_directory(download_dir)
        
        # å¦‚æœåªæœ‰ä¸‹è¼‰ï¼Œæ²’æœ‰æ¯”è¼ƒï¼Œå°±è¿”å›ä¸‹è¼‰å®Œæˆç‹€æ…‹
        if not os.path.exists(compare_dir):
            task_info['message'] = f'ä¸‹è¼‰å®Œæˆï¼Œå…± {download_stats["stats"]["downloaded"]} å€‹æ–‡ä»¶'
            return task_info
    
    # æª¢æŸ¥æ˜¯å¦æœ‰æ¯”è¼ƒçµæœ
    if os.path.exists(compare_dir):
        app.logger.info(f'Found compare directory for task {task_id}')
        
        # åˆ†ææ¯”è¼ƒçµæœ
        compare_stats = analyze_compare_directory(compare_dir)
        task_info['results']['compare_results'] = compare_stats
        
        # æŸ¥æ‰¾ç¸½æ‘˜è¦å ±å‘Š
        summary_files = [f for f in os.listdir(compare_dir) 
                        if f in ['all_scenarios_summary.xlsx', 'all_scenarios_compare.xlsx']]
        if summary_files:
            task_info['results']['summary_report'] = os.path.join(compare_dir, summary_files[0])
        
        # å¦‚æœæœ‰ä¸‹è¼‰å’Œæ¯”å°çµæœï¼Œé€™æ˜¯ä¸€æ­¥åˆ°ä½ä»»å‹™
        if os.path.exists(download_dir):
            total_files = task_info.get('stats', {}).get('downloaded', 0)
            total_scenarios = len([d for d in os.listdir(compare_dir) 
                                 if os.path.isdir(os.path.join(compare_dir, d))])
            task_info['message'] = f'ä¸€æ­¥åˆ°ä½è™•ç†å®Œæˆï¼šä¸‹è¼‰ {total_files} å€‹æª”æ¡ˆï¼Œè™•ç† {total_scenarios} å€‹æ¯”è¼ƒæƒ…å¢ƒ'
        else:
            # åªæœ‰æ¯”è¼ƒçµæœ
            total_scenarios = len([d for d in os.listdir(compare_dir) 
                                 if os.path.isdir(os.path.join(compare_dir, d))])
            task_info['message'] = f'æ¯”è¼ƒå®Œæˆï¼Œè™•ç†äº† {total_scenarios} å€‹æ¯”è¼ƒæƒ…å¢ƒ'
    
    # å¦‚æœæ²’æœ‰æ‰¾åˆ°ä»»ä½•ç›¸é—œç›®éŒ„ï¼Œè¿”å› None
    if not os.path.exists(download_dir) and not os.path.exists(compare_dir):
        return None
    
    return task_info

def generate_folder_structure_from_directory(directory):
    """å¾ç›®éŒ„ç”Ÿæˆæ–‡ä»¶å¤¾çµæ§‹"""
    structure = {}
    
    try:
        for root, dirs, files in os.walk(directory):
            # ç²å–ç›¸å°è·¯å¾‘
            rel_path = os.path.relpath(root, directory)
            if rel_path == '.':
                current_level = structure
            else:
                # æ§‹å»ºåµŒå¥—çµæ§‹
                path_parts = rel_path.split(os.sep)
                current_level = structure
                for part in path_parts:
                    if part not in current_level:
                        current_level[part] = {}
                    current_level = current_level[part]
            
            # æ·»åŠ æ–‡ä»¶
            for file in files:
                if not file.endswith('_report.xlsx'):  # è·³éå ±å‘Šæ–‡ä»¶
                    file_path = os.path.join(root, file)
                    current_level[file] = os.path.relpath(file_path, os.getcwd())
    
    except Exception as e:
        app.logger.error(f'Error generating folder structure: {str(e)}')
    
    return structure

@app.route('/api/check-task-exists/<task_id>')
def check_task_exists(task_id):
    """æª¢æŸ¥ä»»å‹™æ˜¯å¦å­˜åœ¨æ–¼æ–‡ä»¶ç³»çµ±ä¸­"""
    try:
        download_dir = os.path.join('downloads', task_id)
        compare_dir = os.path.join('compare_results', task_id)
        
        exists_info = {
            'task_id': task_id,
            'exists': False,
            'has_download': False,
            'has_compare': False,
            'paths': []
        }
        
        if os.path.exists(download_dir):
            exists_info['has_download'] = True
            exists_info['exists'] = True
            exists_info['paths'].append(f'downloads/{task_id}')
        
        if os.path.exists(compare_dir):
            exists_info['has_compare'] = True
            exists_info['exists'] = True  
            exists_info['paths'].append(f'compare_results/{task_id}')
        
        return jsonify(exists_info)
        
    except Exception as e:
        return jsonify({
            'task_id': task_id,
            'exists': False,
            'error': str(e)
        })

def analyze_compare_directory(compare_dir):
    """åˆ†ææ¯”è¼ƒçµæœç›®éŒ„"""
    compare_results = {}
    
    try:
        # æª¢æŸ¥å„å€‹æ¯”è¼ƒæƒ…å¢ƒç›®éŒ„
        scenarios = ['master_vs_premp', 'premp_vs_wave', 'wave_vs_backup']
        
        for scenario in scenarios:
            scenario_dir = os.path.join(compare_dir, scenario)
            if os.path.exists(scenario_dir):
                # çµ±è¨ˆè©²æƒ…å¢ƒä¸‹çš„æ¨¡çµ„æ•¸é‡
                module_count = count_modules_in_scenario(scenario_dir)
                compare_results[scenario] = {
                    'success': module_count,
                    'failed': 0,  # å¾æ–‡ä»¶ç³»çµ±æ¢å¾©æ™‚å¾ˆé›£ç¢ºå®šå¤±æ•—æ•¸é‡
                    'modules': [],  # å¯ä»¥é€²ä¸€æ­¥å¯¦ç¾è©³ç´°æ¨¡çµ„åˆ—è¡¨
                    'failed_modules': []
                }
    
    except Exception as e:
        app.logger.error(f'Error analyzing compare directory {compare_dir}: {str(e)}')
    
    return compare_results

def count_modules_in_scenario(scenario_dir):
    """è¨ˆç®—æƒ…å¢ƒç›®éŒ„ä¸‹çš„æ¨¡çµ„æ•¸é‡"""
    try:
        # è¨ˆç®—æœ‰æ¯”è¼ƒå ±å‘Šçš„æ¨¡çµ„æ•¸é‡
        module_count = 0
        for root, dirs, files in os.walk(scenario_dir):
            # è¨ˆç®—xlsxæ–‡ä»¶æ•¸é‡ä½œç‚ºæ¨¡çµ„æ•¸é‡çš„åƒè€ƒ
            xlsx_files = [f for f in files if f.endswith('.xlsx') and not f.startswith('all_')]
            module_count += len(xlsx_files)
        return module_count
    except Exception:
        return 0

def analyze_download_directory(download_dir):
    """åˆ†æä¸‹è¼‰ç›®éŒ„ï¼Œçµ±è¨ˆæ–‡ä»¶ä¿¡æ¯"""
    stats = {'total': 0, 'downloaded': 0, 'skipped': 0, 'failed': 0}
    files = {'downloaded': [], 'skipped': [], 'failed': []}
    
    try:
        # éæ­·ä¸‹è¼‰ç›®éŒ„
        for root, dirs, file_list in os.walk(download_dir):
            for file in file_list:
                # è·³éå ±å‘Šæ–‡ä»¶
                if file.endswith('_report.xlsx'):
                    continue
                
                file_path = os.path.join(root, file)
                rel_path = os.path.relpath(file_path, download_dir)
                
                # æ§‹å»ºæ–‡ä»¶ä¿¡æ¯
                file_info = {
                    'name': file,
                    'path': file_path,
                    'ftp_path': rel_path,  # ä½¿ç”¨ç›¸å°è·¯å¾‘ä½œç‚ºFTPè·¯å¾‘
                    'size': os.path.getsize(file_path) if os.path.exists(file_path) else 0
                }
                
                files['downloaded'].append(file_info)
                stats['downloaded'] += 1
        
        stats['total'] = stats['downloaded']
        
        # å˜—è©¦å¾å ±å‘Šæ–‡ä»¶ä¸­ç²å–æ›´æº–ç¢ºçš„çµ±è¨ˆ
        report_files = [f for f in os.listdir(download_dir) if f.endswith('_report.xlsx')]
        if report_files:
            try:
                report_path = os.path.join(download_dir, report_files[0])
                import pandas as pd
                df = pd.read_excel(report_path)
                
                # å¾å ±å‘Šä¸­ç²å–æ›´æº–ç¢ºçš„çµ±è¨ˆ
                if 'status' in df.columns:
                    status_counts = df['status'].value_counts()
                    stats['downloaded'] = status_counts.get('downloaded', 0)
                    stats['skipped'] = status_counts.get('skipped', 0)
                    stats['failed'] = status_counts.get('failed', 0)
                    stats['total'] = len(df)
                    
            except Exception as e:
                app.logger.warning(f'Could not read download report: {str(e)}')
        
    except Exception as e:
        app.logger.error(f'Error analyzing download directory {download_dir}: {str(e)}')
    
    return {'stats': stats, 'files': files}

@app.route('/api/list-directories')
def list_directories():
    """åˆ—å‡ºå¯ç”¨çš„ç›®éŒ„ API"""
    directories = []
    
    # åªæª¢æŸ¥ downloads ç›®éŒ„ä¸‹çš„ task_ é–‹é ­çš„è³‡æ–™å¤¾
    download_base_dir = config.DEFAULT_OUTPUT_DIR  # ä½¿ç”¨ config ä¸­çš„é è¨­ç›®éŒ„
    item_path = download_base_dir
    if os.path.isdir(item_path):
        directories.append({
            'path': item_path,
            'name': f'{os.path.basename(download_base_dir)}',  # é¡¯ç¤ºç›¸å°è·¯å¾‘
            'type': 'download'
        })
    if os.path.exists(download_base_dir):
        for item in os.listdir(download_base_dir):
            # åªåˆ—å‡º task_ é–‹é ­çš„è³‡æ–™å¤¾
            if item.startswith('task_'):
                item_path = os.path.join(download_base_dir, item)
                if os.path.isdir(item_path):
                    directories.append({
                        'path': item_path,
                        'name': f'{os.path.basename(download_base_dir)}/{item}',  # é¡¯ç¤ºç›¸å°è·¯å¾‘
                        'type': 'download'
                    })
    
    # æŒ‰æ™‚é–“æ’åºï¼ˆæœ€æ–°çš„åœ¨å‰ï¼‰
    directories.sort(key=lambda x: os.path.getmtime(x['path']), reverse=True)
    
    return jsonify(directories)

@app.route('/api/recent-activities')
def get_recent_activities():
    """å–å¾—æœ€è¿‘æ´»å‹• - åŒ…å«å¾æª”æ¡ˆç³»çµ±æ¨æ–·çš„æ´»å‹•"""
    try:
        activities = []
        
        # 1. å¾è¨˜æ†¶é«”ä¸­çš„æ´»å‹•è¨˜éŒ„
        for activity in recent_activities[:10]:
            activities.append({
                'timestamp': activity['timestamp'].isoformat(),
                'action': activity['action'],
                'status': activity['status'],
                'details': activity['details']
            })
        
        # 2. å¦‚æœè¨˜æ†¶é«”ä¸­æ²’æœ‰è¶³å¤ çš„æ´»å‹•ï¼Œå¾æª”æ¡ˆç³»çµ±æ¨æ–·
        if len(activities) < 5:
            inferred_activities = infer_activities_from_filesystem()
            activities.extend(inferred_activities)
        
        # 3. æŒ‰æ™‚é–“æ’åºä¸¦åªå–å‰10ç­†
        activities = sorted(activities, key=lambda x: x['timestamp'], reverse=True)[:10]
        
        # 4. å¦‚æœé‚„æ˜¯æ²’æœ‰è³‡æ–™ï¼Œè¿”å›ç©ºåˆ—è¡¨ï¼ˆå‰ç«¯æœƒé¡¯ç¤ºå‹å¥½è¨Šæ¯ï¼‰
        return jsonify(activities)
        
    except Exception as e:
        app.logger.error(f'Get recent activities error: {e}')
        return jsonify([])

def infer_activities_from_filesystem():
    """å¾æª”æ¡ˆç³»çµ±æ¨æ–·æ­·å²æ´»å‹•"""
    activities = []
    
    try:
        # å¾ä¸‹è¼‰ç›®éŒ„æ¨æ–·æ´»å‹•
        downloads_dir = 'downloads'
        if os.path.exists(downloads_dir):
            for item in os.listdir(downloads_dir):
                if item.startswith('task_') and os.path.isdir(os.path.join(downloads_dir, item)):
                    item_path = os.path.join(downloads_dir, item)
                    mtime = os.path.getmtime(item_path)
                    timestamp = datetime.fromtimestamp(mtime)
                    
                    # çµ±è¨ˆè©²ä»»å‹™çš„æª”æ¡ˆæ•¸é‡
                    file_count = 0
                    for root, dirs, files in os.walk(item_path):
                        file_count += len([f for f in files if not f.endswith('_report.xlsx')])
                    
                    activities.append({
                        'timestamp': timestamp.isoformat(),
                        'action': 'å®Œæˆæª”æ¡ˆä¸‹è¼‰',
                        'status': 'success',
                        'details': f'{file_count} å€‹æª”æ¡ˆï¼Œä»»å‹™ {item}'
                    })
        
        # å¾æ¯”å°çµæœç›®éŒ„æ¨æ–·æ´»å‹•
        compare_dir = 'compare_results'
        if os.path.exists(compare_dir):
            for item in os.listdir(compare_dir):
                if item.startswith('task_') and os.path.isdir(os.path.join(compare_dir, item)):
                    item_path = os.path.join(compare_dir, item)
                    mtime = os.path.getmtime(item_path)
                    timestamp = datetime.fromtimestamp(mtime)
                    
                    # æª¢æŸ¥æ¯”å°æƒ…å¢ƒ
                    scenarios = []
                    for scenario in ['master_vs_premp', 'premp_vs_wave', 'wave_vs_backup']:
                        scenario_path = os.path.join(item_path, scenario)
                        if os.path.exists(scenario_path):
                            scenarios.append(scenario)
                    
                    scenario_text = f"{len(scenarios)} å€‹æ¯”å°æƒ…å¢ƒ" if scenarios else "æ¯”å°è™•ç†"
                    
                    activities.append({
                        'timestamp': timestamp.isoformat(),
                        'action': 'å®Œæˆæ¯”å°åˆ†æ',
                        'status': 'success',
                        'details': f'{scenario_text}ï¼Œä»»å‹™ {item}'
                    })
    
    except Exception as e:
        app.logger.warning(f'Error inferring activities: {e}')
    
    return activities

@app.route('/download')
@global_login_required
def download_page():
    """ä¸‹è¼‰é é¢ - æ”¯æ´ task_id åƒæ•¸"""
    # ç²å– task_id åƒæ•¸
    task_id = request.args.get('task_id')
    
    # å¦‚æœæœ‰ task_idï¼Œå‚³éçµ¦æ¨¡æ¿
    return render_template('download.html', task_id=task_id)
    
@app.route('/api/recent-comparisons')
def get_recent_comparisons():
    """å–å¾—æœ€è¿‘æ¯”å°è¨˜éŒ„"""
    return jsonify([{
        'id': comp.get('task_id', comp.get('id', '')),  # å„ªå…ˆä½¿ç”¨ task_id
        'task_id': comp.get('task_id', comp.get('id', '')),  # ç¢ºä¿æœ‰ task_id
        'timestamp': comp['timestamp'].isoformat(),
        'scenario': comp['scenario'],
        'status': comp['status'],
        'modules': comp['modules'],
        'duration': comp.get('duration', '< 1 åˆ†é˜')
    } for comp in recent_comparisons[:10]])

@app.route('/api/statistics')
def get_statistics():
    """å–å¾—çœŸå¯¦çµ±è¨ˆè³‡æ–™"""
    try:
        # è¨ˆç®—ç¸½è™•ç†æ•¸
        total_processed = calculate_total_processed()
        
        # è¨ˆç®—ä»Šæ—¥è™•ç†æ•¸
        today_processed = calculate_today_processed()
        
        # è¨ˆç®—æˆåŠŸç‡
        success_rate = calculate_success_rate()
        
        return jsonify({
            'total': total_processed,
            'today': today_processed,
            'successRate': success_rate
        })
        
    except Exception as e:
        app.logger.error(f'Calculate statistics error: {e}')
        # å¦‚æœè¨ˆç®—å¤±æ•—ï¼Œè¿”å› 0 è€Œä¸æ˜¯å‡è³‡æ–™
        return jsonify({
            'total': 0,
            'today': 0,
            'successRate': 0
        })

def calculate_success_rate():
    """è¨ˆç®—æˆåŠŸç‡"""
    if not processing_status and not recent_activities and not recent_comparisons:
        return 0
    
    total_tasks = 0
    successful_tasks = 0
    
    # å¾è™•ç†ç‹€æ…‹è¨ˆç®—
    for task_id, status in processing_status.items():
        if task_id.startswith('task_'):
            total_tasks += 1
            if status.get('status') == 'completed':
                successful_tasks += 1
    
    # å¾æ¯”å°è¨˜éŒ„è¨ˆç®—
    for comparison in recent_comparisons:
        total_tasks += 1
        if comparison.get('status') == 'completed':
            successful_tasks += 1
    
    # å¾æ´»å‹•è¨˜éŒ„è¨ˆç®—ï¼ˆåªè¨ˆç®—æ˜ç¢ºæ¨™ç¤ºæˆåŠŸ/å¤±æ•—çš„æ´»å‹•ï¼‰
    for activity in recent_activities:
        if activity.get('status') in ['success', 'error']:
            total_tasks += 1
            if activity.get('status') == 'success':
                successful_tasks += 1
    
    # é¿å…é™¤ä»¥é›¶
    if total_tasks == 0:
        return 0
    
    # è¨ˆç®—ç™¾åˆ†æ¯”ä¸¦å››æ¨äº”å…¥
    success_rate = (successful_tasks / total_tasks) * 100
    return round(success_rate, 1)

@app.route('/api/detailed-statistics')
def get_detailed_statistics():
    """å–å¾—è©³ç´°çµ±è¨ˆè³‡æ–™"""
    try:
        # åŸºæœ¬çµ±è¨ˆ
        total = calculate_total_processed()
        today = calculate_today_processed()
        success_rate = calculate_success_rate()
        
        # é¡å¤–çµ±è¨ˆ
        download_tasks = count_download_tasks()
        compare_tasks = count_compare_tasks()
        failed_tasks = count_failed_tasks()
        
        # æœ¬é€±çµ±è¨ˆ
        week_processed = calculate_week_processed()
        
        # å¹³å‡è™•ç†æ™‚é–“ï¼ˆå¦‚æœæœ‰è¨˜éŒ„çš„è©±ï¼‰
        avg_processing_time = calculate_average_processing_time()
        
        return jsonify({
            'basic': {
                'total': total,
                'today': today,
                'successRate': success_rate
            },
            'breakdown': {
                'downloadTasks': download_tasks,
                'compareTasks': compare_tasks,
                'failedTasks': failed_tasks
            },
            'temporal': {
                'weekProcessed': week_processed,
                'avgProcessingTime': avg_processing_time
            }
        })
        
    except Exception as e:
        app.logger.error(f'Detailed statistics error: {e}')
        return jsonify({'error': str(e)}), 500

def calculate_average_processing_time():
    """è¨ˆç®—å¹³å‡è™•ç†æ™‚é–“ï¼ˆåˆ†é˜ï¼‰"""
    # é€™å€‹éœ€è¦ä½ åœ¨è™•ç†éç¨‹ä¸­è¨˜éŒ„é–‹å§‹å’ŒçµæŸæ™‚é–“
    # ç›®å‰å…ˆè¿”å›ä¼°è¨ˆå€¼
    try:
        # å¯ä»¥å¾ processing_status æˆ–å…¶ä»–åœ°æ–¹ç²å–å¯¦éš›è™•ç†æ™‚é–“
        # é€™è£¡æä¾›ä¸€å€‹ç°¡å–®çš„ä¼°ç®—
        total_tasks = len(processing_status)
        if total_tasks > 0:
            # å‡è¨­å¹³å‡æ¯å€‹ä»»å‹™ 5 åˆ†é˜ï¼ˆä½ å¯ä»¥æ ¹æ“šå¯¦éš›æƒ…æ³èª¿æ•´ï¼‰
            return 5.0
        return 0.0
    except Exception:
        return 0.0

def calculate_week_processed():
    """è¨ˆç®—æœ¬é€±è™•ç†æ•¸"""
    from datetime import datetime, timedelta
    
    # ç²å–æœ¬é€±çš„é–‹å§‹æ—¥æœŸï¼ˆé€±ä¸€ï¼‰
    today = datetime.now()
    week_start = today - timedelta(days=today.weekday())
    week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
    
    week_count = 0
    
    # å¾æ´»å‹•è¨˜éŒ„è¨ˆç®—
    for activity in recent_activities:
        try:
            if activity['timestamp'] >= week_start:
                week_count += 1
        except Exception:
            continue
    
    # å¾æ¯”å°è¨˜éŒ„è¨ˆç®—
    for comparison in recent_comparisons:
        try:
            if comparison['timestamp'] >= week_start:
                week_count += 1
        except Exception:
            continue
            
    return week_count

def count_failed_tasks():
    """çµ±è¨ˆå¤±æ•—ä»»å‹™æ•¸é‡"""
    failed_count = 0
    
    # å¾è™•ç†ç‹€æ…‹çµ±è¨ˆ
    for status in processing_status.values():
        if status.get('status') == 'error':
            failed_count += 1
    
    # å¾æ´»å‹•è¨˜éŒ„çµ±è¨ˆ
    for activity in recent_activities:
        if activity.get('status') == 'error':
            failed_count += 1
            
    return failed_count

def count_compare_tasks():
    """çµ±è¨ˆæ¯”å°ä»»å‹™æ•¸é‡"""
    count = 0
    compare_dir = 'compare_results'
    if os.path.exists(compare_dir):
        try:
            count = len([d for d in os.listdir(compare_dir) 
                        if d.startswith('task_') and os.path.isdir(os.path.join(compare_dir, d))])
        except Exception:
            pass
    return count

def count_download_tasks():
    """çµ±è¨ˆä¸‹è¼‰ä»»å‹™æ•¸é‡"""
    count = 0
    downloads_dir = 'downloads'
    if os.path.exists(downloads_dir):
        try:
            count = len([d for d in os.listdir(downloads_dir) 
                        if d.startswith('task_') and os.path.isdir(os.path.join(downloads_dir, d))])
        except Exception:
            pass
    return count

def calculate_today_processed():
    """è¨ˆç®—ä»Šæ—¥è™•ç†æ•¸"""
    from datetime import datetime, timedelta
    
    today = datetime.now().date()
    today_count = 0
    
    # å¾æ´»å‹•è¨˜éŒ„è¨ˆç®—ä»Šæ—¥è™•ç†æ•¸
    for activity in recent_activities:
        try:
            # activity['timestamp'] æ˜¯ datetime ç‰©ä»¶
            activity_date = activity['timestamp'].date()
            if activity_date == today:
                today_count += 1
        except Exception as e:
            app.logger.warning(f'Error parsing activity timestamp: {e}')
            continue
    
    # å¾æ¯”å°è¨˜éŒ„è¨ˆç®—ä»Šæ—¥è™•ç†æ•¸
    for comparison in recent_comparisons:
        try:
            # comparison['timestamp'] æ˜¯ datetime ç‰©ä»¶
            comparison_date = comparison['timestamp'].date()
            if comparison_date == today:
                today_count += 1
        except Exception as e:
            app.logger.warning(f'Error parsing comparison timestamp: {e}')
            continue
    
    # å¾ä»»å‹™ç›®éŒ„çš„ä¿®æ”¹æ™‚é–“è¨ˆç®—
    try:
        for dir_name in ['downloads', 'compare_results']:
            if os.path.exists(dir_name):
                for item in os.listdir(dir_name):
                    if item.startswith('task_'):
                        item_path = os.path.join(dir_name, item)
                        if os.path.isdir(item_path):
                            # ç²å–ç›®éŒ„çš„ä¿®æ”¹æ™‚é–“
                            mtime = os.path.getmtime(item_path)
                            mdate = datetime.fromtimestamp(mtime).date()
                            if mdate == today:
                                today_count += 1
    except Exception as e:
        app.logger.warning(f'Error counting today tasks from directories: {e}')
    
    return today_count

def calculate_total_processed():
    """è¨ˆç®—ç¸½è™•ç†æ•¸"""
    total = 0
    
    # æ–¹æ³•1ï¼šå¾æ´»å‹•è¨˜éŒ„è¨ˆç®—
    total += len(recent_activities)
    
    # æ–¹æ³•2ï¼šå¾æ¯”å°è¨˜éŒ„è¨ˆç®—
    total += len(recent_comparisons)
    
    # æ–¹æ³•3ï¼šå¾æª”æ¡ˆç³»çµ±è¨ˆç®—ï¼ˆçµ±è¨ˆå·²å®Œæˆçš„ä»»å‹™ç›®éŒ„ï¼‰
    try:
        # çµ±è¨ˆ downloads ç›®éŒ„ä¸‹çš„ä»»å‹™è³‡æ–™å¤¾
        downloads_dir = 'downloads'
        if os.path.exists(downloads_dir):
            download_tasks = [d for d in os.listdir(downloads_dir) 
                            if d.startswith('task_') and os.path.isdir(os.path.join(downloads_dir, d))]
            total += len(download_tasks)
        
        # çµ±è¨ˆ compare_results ç›®éŒ„ä¸‹çš„ä»»å‹™è³‡æ–™å¤¾
        compare_dir = 'compare_results'
        if os.path.exists(compare_dir):
            compare_tasks = [d for d in os.listdir(compare_dir) 
                           if d.startswith('task_') and os.path.isdir(os.path.join(compare_dir, d))]
            total += len(compare_tasks)
            
    except Exception as e:
        app.logger.warning(f'Error counting task directories: {e}')
    
    # å»é‡è¤‡ï¼ˆå› ç‚ºåŒä¸€å€‹ä»»å‹™å¯èƒ½åŒæ™‚æœ‰ä¸‹è¼‰å’Œæ¯”å°ï¼‰
    unique_tasks = set()
    
    # å¾è™•ç†ç‹€æ…‹ç²å–å”¯ä¸€ä»»å‹™
    for task_id in processing_status.keys():
        if task_id.startswith('task_'):
            unique_tasks.add(task_id)
    
    # å¾ç›®éŒ„ç²å–å”¯ä¸€ä»»å‹™
    try:
        for dir_name in ['downloads', 'compare_results']:
            if os.path.exists(dir_name):
                for item in os.listdir(dir_name):
                    if item.startswith('task_') and os.path.isdir(os.path.join(dir_name, item)):
                        unique_tasks.add(item)
    except Exception:
        pass
    
    return len(unique_tasks) if unique_tasks else max(total, 0)

@app.route('/api/pivot-data/<task_id>')
def get_pivot_data(task_id):
    """å–å¾—æ¨ç´åˆ†æè³‡æ–™ API - æ”¯æ´æŒ‰æƒ…å¢ƒæŸ¥æ‰¾"""
    try:
        app.logger.info(f'Getting pivot data for task: {task_id}')
        
        # æª¢æŸ¥æ˜¯å¦æœ‰æƒ…å¢ƒåƒæ•¸
        scenario = request.args.get('scenario', 'all')
        app.logger.info(f'Requested scenario: {scenario}')
        
        # æŸ¥æ‰¾ä»»å‹™çµæœ
        summary_report_path = None
        
        # 1. æ ¹æ“šæ‚¨æä¾›çš„å¯¦éš›è·¯å¾‘çµæ§‹é€²è¡Œç²¾ç¢ºæ˜ å°„
        if scenario == 'all':
            # å…¨éƒ¨æƒ…å¢ƒä½¿ç”¨ all_scenarios_summary.xlsx
            summary_report_path = os.path.join('compare_results', task_id, 'all_scenarios_summary.xlsx')
            if not os.path.exists(summary_report_path):
                # å‚™é¸è·¯å¾‘
                alt_paths = [
                    os.path.join('compare_results', task_id, 'all_scenarios_compare.xlsx'),
                    os.path.join('compare_results', task_id, 'all_compare.xlsx')
                ]
                for path in alt_paths:
                    if os.path.exists(path):
                        summary_report_path = path
                        break
        else:
            # æ ¹æ“šå¯¦éš›çš„è³‡æ–™å¤¾çµæ§‹ï¼Œä¸éœ€è¦æ˜ å°„ï¼
            # å‰ç«¯çš„ scenario åç¨±å°±æ˜¯å¯¦éš›çš„è³‡æ–™å¤¾åç¨±
            folder_name = scenario  # ç›´æ¥ä½¿ç”¨ scenario ä½œç‚ºè³‡æ–™å¤¾åç¨±
            
            # æ§‹å»ºç²¾ç¢ºè·¯å¾‘
            summary_report_path = os.path.join('compare_results', task_id, folder_name, 'all_scenarios_compare.xlsx')
            app.logger.info(f'Checking path: {summary_report_path}')
            
            if not os.path.exists(summary_report_path):
                app.logger.warning(f'File not found at primary path: {summary_report_path}')
                
                # å˜—è©¦å…¶ä»–å¯èƒ½çš„æª”å
                alt_names = ['all_compare.xlsx', f'{folder_name}_compare.xlsx']
                for alt_name in alt_names:
                    alt_path = os.path.join('compare_results', task_id, folder_name, alt_name)
                    app.logger.info(f'Trying alternative path: {alt_path}')
                    if os.path.exists(alt_path):
                        summary_report_path = alt_path
                        app.logger.info(f'Found alternative file: {alt_path}')
                        break
        
        # 2. å¦‚æœé‚„æ˜¯æ‰¾ä¸åˆ°ï¼Œæä¾›è©³ç´°çš„éŒ¯èª¤è¨Šæ¯
        if not summary_report_path or not os.path.exists(summary_report_path):
            app.logger.error(f'Could not find report for scenario: {scenario}')
            
            # åˆ—å‡ºå¯¦éš›å­˜åœ¨çš„æª”æ¡ˆçµæ§‹ä¾›é™¤éŒ¯
            compare_dir = os.path.join('compare_results', task_id)
            available_files = []
            
            if os.path.exists(compare_dir):
                app.logger.info(f'Listing directory structure for {compare_dir}:')
                for root, dirs, files in os.walk(compare_dir):
                    for file in files:
                        if file.endswith('.xlsx'):
                            rel_path = os.path.relpath(os.path.join(root, file), compare_dir)
                            available_files.append(rel_path)
                            app.logger.info(f'  Found: {rel_path}')
            
            # è¿”å›æ›´è©³ç´°çš„éŒ¯èª¤è¨Šæ¯
            error_msg = f'æ‰¾ä¸åˆ°è³‡æ–™æª”æ¡ˆï¼š{scenario}'
            
            return jsonify({
                'error': error_msg,
                'available_files': available_files
            }), 404
        
        # 3. è®€å–ä¸¦è¿”å›è³‡æ–™
        try:
            app.logger.info(f'Reading Excel file: {summary_report_path}')
            
            # è®€å– Excel æª”æ¡ˆçš„æ‰€æœ‰å·¥ä½œè¡¨
            excel_data = pd.read_excel(summary_report_path, sheet_name=None)
            
            # åˆ—å‡ºæ‰€æœ‰å·¥ä½œè¡¨åç¨±
            app.logger.info(f'Available sheets: {list(excel_data.keys())}')
            
            # è½‰æ›ç‚º JSON æ ¼å¼
            pivot_data = {}
            for sheet_name, df in excel_data.items():
                app.logger.info(f'Processing sheet: {sheet_name} with {len(df)} rows')
                
                # è™•ç† NaN å€¼
                df = df.fillna('')
                
                # å°‡æ—¥æœŸè½‰æ›ç‚ºå­—ä¸²
                for col in df.columns:
                    if df[col].dtype == 'datetime64[ns]':
                        df[col] = df[col].astype(str)
                    elif df[col].dtype == 'object':
                        # ç¢ºä¿æ‰€æœ‰å€¼éƒ½æ˜¯å¯åºåˆ—åŒ–çš„
                        df[col] = df[col].apply(lambda x: str(x) if pd.notna(x) else '')
                
                # å°‡ DataFrame è½‰æ›ç‚ºè¨˜éŒ„æ ¼å¼
                pivot_data[sheet_name] = {
                    'columns': df.columns.tolist(),
                    'data': df.to_dict('records')
                }
            
            app.logger.info(f'Successfully loaded {len(pivot_data)} sheets')
            return jsonify(pivot_data)
            
        except Exception as e:
            app.logger.error(f'Error reading Excel file: {e}')
            import traceback
            app.logger.error(traceback.format_exc())
            return jsonify({'error': f'è®€å–è³‡æ–™å¤±æ•—ï¼š{str(e)}'}), 500
        
    except Exception as e:
        app.logger.error(f'Get pivot data error: {e}')
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

def row_belongs_to_scenario(row, scenario):
    """åˆ¤æ–·è³‡æ–™è¡Œæ˜¯å¦å±¬æ–¼ç‰¹å®šæƒ…å¢ƒ"""
    base_folder = str(row.get('base_folder', ''))
    compare_folder = str(row.get('compare_folder', ''))
    
    if scenario == 'master_vs_premp':
        # Master è³‡æ–™å¤¾ï¼ˆç„¡å¾Œç¶´ï¼‰vs PreMP è³‡æ–™å¤¾
        return (not any(base_folder.endswith(suffix) for suffix in ['-premp', '-wave', '-wave.backup', '-mp']) and
                (compare_folder.endswith('-premp') or compare_folder.endswith('-pre-mp')))
    
    elif scenario == 'premp_vs_wave':
        # PreMP è³‡æ–™å¤¾ vs Wave/MP è³‡æ–™å¤¾
        return ((base_folder.endswith('-premp') or base_folder.endswith('-pre-mp')) and 
                (compare_folder.endswith('-wave') or compare_folder.endswith('-mp')) and 
                not compare_folder.endswith('-wave.backup') and
                not compare_folder.endswith('-mpbackup'))
    
    elif scenario == 'wave_vs_backup':
        # Wave/MP è³‡æ–™å¤¾ vs Backup è³‡æ–™å¤¾
        return ((base_folder.endswith('-wave') or base_folder.endswith('-mp')) and 
                not base_folder.endswith('-wave.backup') and
                not base_folder.endswith('-mpbackup') and
                (compare_folder.endswith('-wave.backup') or 
                 compare_folder.endswith('-wavebackup') or
                 compare_folder.endswith('-mpbackup')))
    
    return False

def row_belongs_to_scenario(row, scenario):
    """åˆ¤æ–·è³‡æ–™è¡Œæ˜¯å¦å±¬æ–¼ç‰¹å®šæƒ…å¢ƒï¼ˆç¨ç«‹å‡½æ•¸ï¼Œä¸æ˜¯æ–¹æ³•ï¼‰"""
    # æ ¹æ“šè³‡æ–™å¤¾åç¨±åˆ¤æ–·
    base_folder = str(row.get('base_folder', ''))
    compare_folder = str(row.get('compare_folder', ''))
    
    if scenario == 'master_vs_premp':
        # Master è³‡æ–™å¤¾ï¼ˆç„¡å¾Œç¶´ï¼‰vs PreMP è³‡æ–™å¤¾
        return (not any(base_folder.endswith(suffix) for suffix in ['-premp', '-wave', '-wave.backup']) and
                compare_folder.endswith('-premp'))
    
    elif scenario == 'premp_vs_wave':
        # PreMP è³‡æ–™å¤¾ vs Wave è³‡æ–™å¤¾
        return (base_folder.endswith('-premp') and 
                compare_folder.endswith('-wave') and 
                not compare_folder.endswith('-wave.backup'))
    
    elif scenario == 'wave_vs_backup':
        # Wave è³‡æ–™å¤¾ vs Wave.backup è³‡æ–™å¤¾
        return (base_folder.endswith('-wave') and 
                not base_folder.endswith('-wave.backup') and
                compare_folder.endswith('-wave.backup'))
    
    return False

def _row_belongs_to_scenario(self, row, scenario):
    """åˆ¤æ–·è³‡æ–™è¡Œæ˜¯å¦å±¬æ–¼ç‰¹å®šæƒ…å¢ƒ"""
    # æ ¹æ“šè³‡æ–™å¤¾åç¨±åˆ¤æ–·
    base_folder = str(row.get('base_folder', ''))
    compare_folder = str(row.get('compare_folder', ''))
    
    if scenario == 'master_vs_premp':
        # Master è³‡æ–™å¤¾ï¼ˆç„¡å¾Œç¶´ï¼‰vs PreMP è³‡æ–™å¤¾
        return (not any(base_folder.endswith(suffix) for suffix in ['-premp', '-wave', '-wave.backup']) and
                base_folder.endswith('-premp'))
    
    elif scenario == 'premp_vs_wave':
        # PreMP è³‡æ–™å¤¾ vs Wave è³‡æ–™å¤¾
        return (base_folder.endswith('-premp') and 
                compare_folder.endswith('-wave') and 
                not compare_folder.endswith('-wave.backup'))
    
    elif scenario == 'wave_vs_backup':
        # Wave è³‡æ–™å¤¾ vs Wave.backup è³‡æ–™å¤¾
        return (compare_folder.endswith('-wave') and 
                not compare_folder.endswith('-wave.backup') and
                compare_folder.endswith('-wave.backup'))
    
    return False

def get_mock_pivot_data():
    """å–å¾—æ¨¡æ“¬çš„æ¨ç´åˆ†æè³‡æ–™"""
    return {
        'revision_diff': {
            'columns': ['SN', 'module', 'name', 'path', 'base_short', 'base_revision', 
                       'compare_short', 'compare_revision', 'has_wave'],
            'data': [
                {
                    'SN': 1,
                    'module': 'bootcode',
                    'name': 'realtek/mac7p/bootcode',
                    'path': 'bootcode/src',
                    'base_short': 'abc123d',
                    'base_revision': 'abc123def456...',
                    'compare_short': 'def456g',
                    'compare_revision': 'def456ghi789...',
                    'has_wave': 'N'
                },
                {
                    'SN': 2,
                    'module': 'emcu',
                    'name': 'realtek/emcu',
                    'path': 'emcu/src',
                    'base_short': 'aaa111b',
                    'base_revision': 'aaa111bbb222...',
                    'compare_short': 'ccc333d',
                    'compare_revision': 'ccc333ddd444...',
                    'has_wave': 'Y'
                }
            ]
        },
        'branch_error': {
            'columns': ['SN', 'module', 'name', 'problem', 'has_wave'],
            'data': [
                {
                    'SN': 1,
                    'module': 'bootcode',
                    'name': 'realtek/bootcode',
                    'problem': 'æ²’æ”¹æˆ premp',
                    'has_wave': 'N'
                }
            ]
        }
    }

# è·¯å¾‘å»ºè­° API - æ–°å¢
@app.route('/api/path-suggestions')
def get_path_suggestions():
    """ç²å–è·¯å¾‘å»ºè­° API"""
    path = request.args.get('path', '')
    
    if not path:
        return jsonify({'directories': [], 'files': []})
    
    try:
        # å˜—è©¦ç²å–çœŸå¯¦è·¯å¾‘å»ºè­°
        suggestions = {'directories': [], 'files': []}
        
        # æª¢æŸ¥è·¯å¾‘æ˜¯å¦å­˜åœ¨
        if os.path.exists(path):
            # å¦‚æœè·¯å¾‘å­˜åœ¨ï¼Œåˆ—å‡ºå­ç›®éŒ„å’Œæª”æ¡ˆ
            try:
                for item in os.listdir(path):
                    item_path = os.path.join(path, item)
                    if os.path.isdir(item_path):
                        suggestions['directories'].append({
                            'name': item,
                            'path': item_path
                        })
                    elif item.endswith('.xlsx'):
                        suggestions['files'].append({
                            'name': item,
                            'path': item_path,
                            'size': os.path.getsize(item_path)
                        })
            except PermissionError:
                pass
        else:
            # å¦‚æœè·¯å¾‘ä¸å­˜åœ¨ï¼Œå˜—è©¦å°‹æ‰¾ç›¸ä¼¼çš„è·¯å¾‘
            parent_path = os.path.dirname(path)
            if os.path.exists(parent_path):
                base_name = os.path.basename(path).lower()
                try:
                    for item in os.listdir(parent_path):
                        if base_name in item.lower():
                            item_path = os.path.join(parent_path, item)
                            if os.path.isdir(item_path):
                                suggestions['directories'].append({
                                    'name': item,
                                    'path': item_path
                                })
                            elif item.endswith('.xlsx'):
                                suggestions['files'].append({
                                    'name': item,
                                    'path': item_path,
                                    'size': os.path.getsize(item_path)
                                })
                except PermissionError:
                    pass
            
        # å¦‚æœæ²’æœ‰æ‰¾åˆ°å»ºè­°ï¼Œæä¾›å¸¸ç”¨è·¯å¾‘
        if not suggestions['directories'] and not suggestions['files']:
            from config import COMMON_PATHS
            for common_path in COMMON_PATHS:
                if path.lower() in common_path.lower():
                    if os.path.exists(common_path):
                        name = common_path.split('/')[-1] or common_path
                        suggestions['directories'].append({
                            'name': name,
                            'path': common_path
                        })
        
        # é™åˆ¶å»ºè­°æ•¸é‡
        suggestions['directories'] = suggestions['directories'][:10]
        suggestions['files'] = suggestions['files'][:10]
        
        return jsonify(suggestions)
        
    except Exception as e:
        print(f"Path suggestions error: {e}")
        return jsonify({'directories': [], 'files': []})

# ç€è¦½ä¼ºæœå™¨ API - æ›´æ–°
@app.route('/api/browse-server')
def browse_server():
    path = request.args.get('path', '/')
    
    try:
        items = os.listdir(path)
        folders = []
        files = []
        
        for item in items:
            item_path = os.path.join(path, item)
            if os.path.isdir(item_path):
                folders.append({
                    'name': item,
                    'path': item_path
                })
            elif os.path.isfile(item_path):
                # æª¢æŸ¥æ˜¯å¦ç‚ºæ”¯æ´çš„æª”æ¡ˆé¡å‹
                if item.lower().endswith(('.xlsx', '.xls', '.csv')):
                    files.append({
                        'name': item,
                        'path': item_path,
                        'size': os.path.getsize(item_path)
                    })
        
        return jsonify({
            'folders': folders,
            'files': files
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# éåŒæ­¥ä¸‹è¼‰æ”¯æ´
@app.route('/api/prepare-download/<task_id>', methods=['POST'])
def prepare_download(task_id):
    """æº–å‚™å¤§æª”æ¡ˆä¸‹è¼‰"""
    data = request.json
    format_type = data.get('format', 'zip')
    
    # ç”Ÿæˆä¸‹è¼‰ä»»å‹™ ID
    download_task_id = f"download_{task_id}_{format_type}"
    
    # åœ¨èƒŒæ™¯æº–å‚™æª”æ¡ˆ
    def prepare_file():
        time.sleep(3)  # æ¨¡æ“¬æª”æ¡ˆæº–å‚™æ™‚é–“
        # å¯¦éš›æ‡‰ç”¨ä¸­é€™è£¡æ‡‰è©²ç”ŸæˆçœŸå¯¦çš„æª”æ¡ˆ
        download_url = f"/api/download-ready/{download_task_id}"
        processing_status[download_task_id] = {
            'ready': True,
            'download_url': download_url
        }
    
    thread = threading.Thread(target=prepare_file)
    thread.start()
    
    return jsonify({
        'task_id': download_task_id,
        'ready': False
    })

@app.route('/api/download-status/<task_id>')
def download_status(task_id):
    """æª¢æŸ¥ä¸‹è¼‰æº–å‚™ç‹€æ…‹"""
    status = processing_status.get(task_id, {'ready': False})
    return jsonify(status)

@app.route('/api/download-ready/<task_id>')
def download_ready(task_id):
    """ä¸‹è¼‰æº–å‚™å¥½çš„æª”æ¡ˆ"""
    # å¯¦éš›æ‡‰ç”¨ä¸­é€™è£¡æ‡‰è©²è¿”å›çœŸå¯¦çš„æª”æ¡ˆ
    # é€™è£¡è¿”å›ä¸€å€‹æ¨¡æ“¬çš„æª”æ¡ˆ
    return send_file('README.md', as_attachment=True, download_name='result.zip')

@app.route('/api/export-excel/<task_id>')
def export_excel(task_id):
    """åŒ¯å‡º Excel API - æ ¹æ“šç•¶å‰æƒ…å¢ƒä¸‹è¼‰å°æ‡‰çš„æª”æ¡ˆ"""
    try:
        # ç²å–æƒ…å¢ƒåƒæ•¸ï¼ˆå¾å‰ç«¯å‚³ä¾†ï¼‰
        scenario = request.args.get('scenario', 'all')
        app.logger.info(f'åŒ¯å‡ºExcel: task_id={task_id}, scenario={scenario}')
        
        # æ ¹æ“šæƒ…å¢ƒæ±ºå®šè¦ä¸‹è¼‰çš„æª”æ¡ˆè·¯å¾‘ï¼ˆèˆ‡ get_pivot_data é‚è¼¯ä¸€è‡´ï¼‰
        summary_report_path = None
        
        if scenario == 'all':
            # å…¨éƒ¨æƒ…å¢ƒä½¿ç”¨ all_scenarios_summary.xlsx
            summary_report_path = os.path.join('compare_results', task_id, 'all_scenarios_summary.xlsx')
            if not os.path.exists(summary_report_path):
                # å‚™é¸è·¯å¾‘
                alt_paths = [
                    os.path.join('compare_results', task_id, 'all_scenarios_compare.xlsx'),
                    os.path.join('compare_results', task_id, 'all_compare.xlsx')
                ]
                for path in alt_paths:
                    if os.path.exists(path):
                        summary_report_path = path
                        break
        else:
            # ç‰¹å®šæƒ…å¢ƒä½¿ç”¨å°æ‡‰è³‡æ–™å¤¾ä¸‹çš„ all_scenarios_compare.xlsx
            folder_name = scenario  # ç›´æ¥ä½¿ç”¨ scenario ä½œç‚ºè³‡æ–™å¤¾åç¨±
            summary_report_path = os.path.join('compare_results', task_id, folder_name, 'all_scenarios_compare.xlsx')
            
            if not os.path.exists(summary_report_path):
                # å˜—è©¦å…¶ä»–å¯èƒ½çš„æª”å
                alt_names = ['all_compare.xlsx', f'{folder_name}_compare.xlsx']
                for alt_name in alt_names:
                    alt_path = os.path.join('compare_results', task_id, folder_name, alt_name)
                    if os.path.exists(alt_path):
                        summary_report_path = alt_path
                        break
        
        # æª¢æŸ¥æª”æ¡ˆæ˜¯å¦å­˜åœ¨
        if not summary_report_path or not os.path.exists(summary_report_path):
            app.logger.error(f'æ‰¾ä¸åˆ°Excelæª”æ¡ˆ: scenario={scenario}, path={summary_report_path}')
            return jsonify({'error': f'æ‰¾ä¸åˆ° {scenario} æƒ…å¢ƒçš„å ±è¡¨æª”æ¡ˆ'}), 404
        
        # ç”Ÿæˆæª”æ¡ˆåç¨±ï¼ˆåŒ…å«æƒ…å¢ƒè³‡è¨Šï¼‰
        if scenario == 'all':
            filename = f'å®Œæ•´å ±è¡¨_{task_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        else:
            scenario_names = {
                'master_vs_premp': 'Master_vs_PreMP',
                'premp_vs_wave': 'PreMP_vs_Wave', 
                'wave_vs_backup': 'Wave_vs_Backup'
            }
            scenario_display = scenario_names.get(scenario, scenario)
            filename = f'{scenario_display}å ±è¡¨_{task_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        app.logger.info(f'åŒ¯å‡ºæª”æ¡ˆ: {summary_report_path} -> {filename}')
        
        return send_file(
            summary_report_path, 
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        app.logger.error(f'Export Excel error: {e}')
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-zip/<task_id>')
def export_zip(task_id):
    """åŒ¯å‡º ZIP æª”æ¡ˆ API - åŒ…å«æ‰€æœ‰çµæœ"""
    try:
        import tempfile
        import shutil
        import zipfile
        import io
        
        # ä½¿ç”¨ BytesIO è€Œä¸æ˜¯è‡¨æ™‚æª”æ¡ˆ
        zip_buffer = io.BytesIO()
        
        # å‰µå»º ZIP
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
            files_added = False
            
            # æ·»åŠ ä¸‹è¼‰çš„æª”æ¡ˆ
            download_dir = os.path.join('downloads', task_id)
            if os.path.exists(download_dir):
                for root, dirs, files in os.walk(download_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arc_name = os.path.join('downloads', os.path.relpath(file_path, download_dir))
                        zipf.write(file_path, arc_name)
                        files_added = True
                app.logger.info(f'Added downloads directory: {download_dir}')
            
            # æ·»åŠ æ¯”å°çµæœ
            compare_dir = os.path.join('compare_results', task_id)
            if os.path.exists(compare_dir):
                for root, dirs, files in os.walk(compare_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arc_name = os.path.join('compare_results', os.path.relpath(file_path, compare_dir))
                        zipf.write(file_path, arc_name)
                        files_added = True
                app.logger.info(f'Added compare results directory: {compare_dir}')
            
            # å¦‚æœæ²’æœ‰æª”æ¡ˆï¼Œè‡³å°‘æ·»åŠ ä¸€å€‹ README
            if not files_added:
                readme_content = f'ä»»å‹™ ID: {task_id}\n'
                readme_content += f'å‰µå»ºæ™‚é–“: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n'
                readme_content += 'æš«ç„¡æ¯”å°çµæœ\n'
                zipf.writestr('README.txt', readme_content.encode('utf-8'))
        
        # é‡ç½® buffer ä½ç½®
        zip_buffer.seek(0)
        
        # ç”Ÿæˆæª”æ¡ˆåç¨±
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        zip_filename = f'results_{task_id}_{timestamp}.zip'
        
        # è¿”å›æª”æ¡ˆ
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=zip_filename
        )
        
    except Exception as e:
        app.logger.error(f'Export ZIP error: {e}')
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
    
def create_mock_excel(task_id):
    """å‰µå»ºæ¨¡æ“¬çš„ Excel æª”æ¡ˆ"""
    import tempfile
    
    # å‰µå»ºè‡¨æ™‚æª”æ¡ˆ
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    
    # å‰µå»ºæ¨¡æ“¬è³‡æ–™
    data = get_mock_pivot_data()
    
    with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
        for sheet_name, sheet_data in data.items():
            df = pd.DataFrame(sheet_data['data'])
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    return temp_file.name

@app.route('/api/preview-file')
def preview_file():
    """é è¦½æª”æ¡ˆå…§å®¹ API"""
    file_path = request.args.get('path')
    
    if not file_path:
        return jsonify({'error': 'ç¼ºå°‘æª”æ¡ˆè·¯å¾‘'}), 400
        
    try:
        # å»ºæ§‹çœŸå¯¦çš„æª”æ¡ˆè·¯å¾‘
        full_path = os.path.join(os.getcwd(), file_path)
        
        # å®‰å…¨æ€§æª¢æŸ¥
        allowed_dirs = ['downloads', 'compare_results']
        path_parts = file_path.split('/')
        
        if len(path_parts) > 0 and path_parts[0] not in allowed_dirs:
            return jsonify({'error': 'ç„¡æ•ˆçš„æª”æ¡ˆè·¯å¾‘'}), 403
            
        # æª¢æŸ¥æª”æ¡ˆæ˜¯å¦å­˜åœ¨
        if not os.path.exists(full_path):
            return jsonify({'error': 'æª”æ¡ˆä¸å­˜åœ¨'}), 404
            
        # è®€å–æª”æ¡ˆå…§å®¹ï¼ˆé™åˆ¶å¤§å°ï¼‰
        max_size = 1024 * 1024  # 1MB
        file_size = os.path.getsize(full_path)
        
        if file_size > max_size:
            return jsonify({'content': f'æª”æ¡ˆå¤ªå¤§ ({file_size} bytes)ï¼Œç„¡æ³•é è¦½'})
            
        # åˆ¤æ–·æª”æ¡ˆé¡å‹
        file_ext = os.path.splitext(full_path)[1].lower()
        
        # è®€å–æª”æ¡ˆ
        try:
            with open(full_path, 'r', encoding='utf-8') as f:
                content = f.read()
                
            # é‡å°XMLç‰¹åˆ¥è™•ç†
            if file_ext == '.xml':
                import xml.dom.minidom
                try:
                    # æ ¼å¼åŒ–XML
                    dom = xml.dom.minidom.parseString(content)
                    content = dom.toprettyxml(indent="  ")
                    # ç§»é™¤å¤šé¤˜çš„ç©ºç™½è¡Œ
                    content = '\n'.join([line for line in content.split('\n') if line.strip()])
                except:
                    # å¦‚æœè§£æå¤±æ•—ï¼Œè¿”å›åŸå§‹å…§å®¹
                    pass
                    
        except UnicodeDecodeError:
            # å˜—è©¦å…¶ä»–ç·¨ç¢¼
            try:
                with open(full_path, 'r', encoding='big5') as f:
                    content = f.read()
            except:
                content = 'ç„¡æ³•è§£ç¢¼æª”æ¡ˆå…§å®¹ï¼ˆå¯èƒ½æ˜¯äºŒé€²ä½æª”æ¡ˆï¼‰'
                
        return jsonify({
            'content': content,
            'type': file_ext[1:] if file_ext else 'txt'
        })
        
    except Exception as e:
        print(f"Preview file error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-file')
def download_single_file():
    """ä¸‹è¼‰å–®ä¸€æª”æ¡ˆ API"""
    file_path = request.args.get('path')
    
    if not file_path:
        return jsonify({'error': 'ç¼ºå°‘æª”æ¡ˆè·¯å¾‘'}), 400
        
    try:
        # å»ºæ§‹çœŸå¯¦çš„æª”æ¡ˆè·¯å¾‘
        full_path = os.path.join(os.getcwd(), file_path)
        
        # å®‰å…¨æ€§æª¢æŸ¥ - ç¢ºä¿è·¯å¾‘åœ¨å…è¨±çš„ç›®éŒ„å…§
        allowed_dirs = ['downloads', 'compare_results', 'zip_output']
        path_parts = file_path.split('/')
        
        if len(path_parts) > 0 and path_parts[0] not in allowed_dirs:
            return jsonify({'error': 'ç„¡æ•ˆçš„æª”æ¡ˆè·¯å¾‘'}), 403
            
        # æª¢æŸ¥æª”æ¡ˆæ˜¯å¦å­˜åœ¨
        if not os.path.exists(full_path):
            return jsonify({'error': 'æª”æ¡ˆä¸å­˜åœ¨'}), 404
            
        # è¿”å›çœŸå¯¦æª”æ¡ˆ
        return send_file(
            full_path, 
            as_attachment=True,
            download_name=os.path.basename(file_path)
        )
        
    except Exception as e:
        print(f"Download file error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-report/<task_id>')
def download_report(task_id):
    """ä¸‹è¼‰å ±è¡¨ API"""
    try:
        # æŸ¥æ‰¾ä»»å‹™çš„ä¸‹è¼‰å ±å‘Š
        if task_id in processing_status:
            task_data = processing_status[task_id]
            report_path = task_data.get('results', {}).get('download_report')
            
            if report_path and os.path.exists(report_path):
                return send_file(report_path, as_attachment=True, 
                               download_name=f'download_report_{task_id}.xlsx')
        
        # å¦‚æœæ²’æœ‰æ‰¾åˆ°ï¼Œå˜—è©¦åœ¨ä¸‹è¼‰ç›®éŒ„ä¸­æŸ¥æ‰¾
        download_dir = os.path.join('downloads', task_id)
        report_path = os.path.join(download_dir, 'download_report.xlsx')
        
        if os.path.exists(report_path):
            return send_file(report_path, as_attachment=True,
                           download_name=f'download_report_{task_id}.xlsx')
        
        return jsonify({'error': 'æ‰¾ä¸åˆ°ä¸‹è¼‰å ±è¡¨'}), 404
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# SocketIO äº‹ä»¶è™•ç†
@socketio.on('connect')
def handle_connect():
    """è™•ç†é€£ç·šäº‹ä»¶"""
    print(f'Client connected: {request.sid}')

@socketio.on('disconnect')
def handle_disconnect():
    """è™•ç†æ–·ç·šäº‹ä»¶"""
    print(f'Client disconnected: {request.sid}')

@socketio.on('join_task')
def handle_join_task(data):
    """åŠ å…¥ä»»å‹™æˆ¿é–“"""
    task_id = data.get('task_id')
    if task_id:
        join_room(task_id)
        emit('joined', {'task_id': task_id})

# éŒ¯èª¤è™•ç†
@app.errorhandler(404)
def not_found(error):
    """404 éŒ¯èª¤è™•ç†"""
    return jsonify({'error': 'Not found'}), 404

@app.errorhandler(500)
def internal_error(error):
    """500 éŒ¯èª¤è™•ç†"""
    return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/list-folders')
async def list_folders():
    path = request.args.get('path', '/home/vince_lin/ai/preMP/downloads')
    
    try:
        folders = []
        
        # ä½¿ç”¨ SFTP æˆ–æœ¬åœ°æª”æ¡ˆç³»çµ±åˆ—å‡ºè³‡æ–™å¤¾
        if os.path.exists(path) and os.path.isdir(path):
            for item in os.listdir(path):
                item_path = os.path.join(path, item)
                if os.path.isdir(item_path):
                    folders.append({
                        'name': item,
                        'path': item_path
                    })
        
        # æŒ‰åç¨±æ’åº
        folders.sort(key=lambda x: x['name'])
        
        return jsonify({'folders': folders})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/browse-directory')
async def browse_directory():
    path = request.args.get('path', '/home/vince_lin/ai/preMP')
    
    try:
        folders = []
        
        # æª¢æŸ¥è·¯å¾‘æ˜¯å¦å­˜åœ¨
        if not os.path.exists(path) or not os.path.isdir(path):
            return jsonify({'error': 'è·¯å¾‘ä¸å­˜åœ¨æˆ–ä¸æ˜¯ç›®éŒ„'})
        
        # åˆ—å‡ºæ‰€æœ‰å­ç›®éŒ„
        for item in sorted(os.listdir(path)):
            item_path = os.path.join(path, item)
            if os.path.isdir(item_path):
                folders.append({
                    'name': item,
                    'path': item_path
                })
        
        return jsonify({
            'path': path,
            'folders': folders
        })
        
    except PermissionError:
        return jsonify({'error': 'æ²’æœ‰æ¬Šé™è¨ªå•æ­¤ç›®éŒ„'})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/list-export-tasks')
def list_export_tasks():
    """åˆ—å‡ºå¯åŒ¯å‡ºçš„ä»»å‹™ API"""
    try:
        tasks = []
        
        # å¾æ¯”å°çµæœç›®éŒ„ç²å–ä»»å‹™
        if os.path.exists('compare_results'):
            for task_dir in os.listdir('compare_results'):
                task_path = os.path.join('compare_results', task_dir)
                if os.path.isdir(task_path):
                    # æª¢æŸ¥æ˜¯å¦æœ‰å ±å‘Šæª”æ¡ˆ
                    has_report = False
                    for file in os.listdir(task_path):
                        if file.endswith('.xlsx'):
                            has_report = True
                            break
                    
                    if has_report:
                        # ç²å–ä¿®æ”¹æ™‚é–“
                        timestamp = os.path.getmtime(task_path)
                        tasks.append({
                            'id': task_dir,
                            'name': f'æ¯”å°ä»»å‹™ {task_dir}',
                            'timestamp': timestamp * 1000,  # è½‰æ›ç‚ºæ¯«ç§’
                            'type': 'compare'
                        })
        
        # å¾è™•ç†ç‹€æ…‹ç²å–ä»»å‹™
        for task_id, status in processing_status.items():
            if status.get('status') == 'completed':
                tasks.append({
                    'id': task_id,
                    'name': f'è™•ç†ä»»å‹™ {task_id}',
                    'timestamp': time.time() * 1000,
                    'type': 'process'
                })
        
        # æŒ‰æ™‚é–“æ’åºï¼ˆæœ€æ–°çš„åœ¨å‰ï¼‰
        tasks.sort(key=lambda x: x['timestamp'], reverse=True)
        
        return jsonify({'tasks': tasks})
        
    except Exception as e:
        app.logger.error(f'List export tasks error: {e}')
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-sheet/<task_id>/<sheet_name>')
def export_sheet(task_id, sheet_name):
    """åŒ¯å‡ºå–®ä¸€è³‡æ–™è¡¨ API"""
    try:
        format_type = request.args.get('format', 'excel')
        
        # ç²å–è³‡æ–™
        data = get_task_data(task_id)
        if not data or sheet_name not in data:
            return jsonify({'error': 'æ‰¾ä¸åˆ°è³‡æ–™è¡¨'}), 404
        
        sheet_data = data[sheet_name]
        
        if format_type == 'excel':
            # å‰µå»º Excel æª”æ¡ˆ
            import tempfile
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            
            df = pd.DataFrame(sheet_data['data'])
            with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # ç¢ºä¿æª”æ¡ˆé—œé–‰
            temp_file.close()
            
            return send_file(
                temp_file.name,
                as_attachment=True,
                download_name=f'{sheet_name}_{task_id}.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        return jsonify({'error': 'ä¸æ”¯æ´çš„æ ¼å¼'}), 400
        
    except Exception as e:
        app.logger.error(f'Export sheet error: {e}')
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-pdf/<task_id>')
def export_pdf(task_id):
    """åŒ¯å‡º PDF å ±å‘Š API"""
    try:
        # é€™è£¡å¯ä»¥æ•´åˆ PDF ç”ŸæˆåŠŸèƒ½
        # ç›®å‰å…ˆè¿”å›éŒ¯èª¤è¨Šæ¯
        return jsonify({'error': 'PDF åŒ¯å‡ºåŠŸèƒ½é–‹ç™¼ä¸­'}), 501
        
    except Exception as e:
        app.logger.error(f'Export PDF error: {e}')
        return jsonify({'error': str(e)}), 500

def get_task_data(task_id):
    """ç²å–ä»»å‹™è³‡æ–™çš„è¼”åŠ©å‡½æ•¸"""
    try:
        # 1. å¾è™•ç†ç‹€æ…‹ä¸­æŸ¥æ‰¾
        if task_id in processing_status:
            task_data = processing_status[task_id]
            results = task_data.get('results', {})
            
            if 'summary_report' in results:
                summary_report_path = results['summary_report']
                if os.path.exists(summary_report_path):
                    excel_data = pd.read_excel(summary_report_path, sheet_name=None)
                    data = {}
                    for sheet_name, df in excel_data.items():
                        df = df.fillna('')
                        data[sheet_name] = {
                            'columns': df.columns.tolist(),
                            'data': df.to_dict('records')
                        }
                    return data
        
        # 2. å¾æ¯”å°çµæœç›®éŒ„ä¸­æŸ¥æ‰¾
        compare_dir = os.path.join('compare_results', task_id)
        if os.path.exists(compare_dir):
            for file in os.listdir(compare_dir):
                if file == 'all_compare.xlsx' or file.endswith('_summary.xlsx'):
                    file_path = os.path.join(compare_dir, file)
                    excel_data = pd.read_excel(file_path, sheet_name=None)
                    data = {}
                    for sheet_name, df in excel_data.items():
                        df = df.fillna('')
                        data[sheet_name] = {
                            'columns': df.columns.tolist(),
                            'data': df.to_dict('records')
                        }
                    return data
        
        return None
        
    except Exception as e:
        app.logger.error(f'Get task data error: {e}')
        return None


@app.route('/api/export-html/<task_id>')
def export_html(task_id):
    """åŒ¯å‡º HTML å ±å‘Š API"""
    try:
        # ç²å–ä»»å‹™è³‡æ–™
        data = get_task_data(task_id)
        if not data:
            return jsonify({'error': 'æ‰¾ä¸åˆ°ä»»å‹™è³‡æ–™'}), 404
        
        # ç”Ÿæˆ HTML å ±å‘Š
        html_content = generate_html_report(task_id, data)
        
        # è¿”å› HTML æª”æ¡ˆ
        response = make_response(html_content)
        response.headers['Content-Type'] = 'text/html; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename=report_{task_id}.html'
        
        return response
        
    except Exception as e:
        app.logger.error(f'Export HTML error: {e}')
        return jsonify({'error': str(e)}), 500

def generate_html_report(task_id, data):
    """ç”Ÿæˆ HTML å ±å‘Šå…§å®¹"""
    # ç²å–å…§åµŒçš„ CSS æ¨£å¼
    css_content = get_embedded_report_styles()
    
    # ç”Ÿæˆè³‡æ–™è¡¨æ ¼ HTML
    tables_html = ""
    ordered_sheets = ['revision_diff', 'branch_error', 'lost_project', 'version_diff', 'ç„¡æ³•æ¯”å°']
    
    for sheet_name in ordered_sheets:
        if sheet_name in data:
            sheet_data = data[sheet_name]
            tables_html += f"""
            <div class="sheet-section">
                <h2 class="sheet-title">
                    <i class="fas fa-table"></i> {sheet_name}
                </h2>
                <div class="table-container">
                    <div class="table-wrapper">
                        {generate_table_html(sheet_data)}
                    </div>
                </div>
            </div>
            """
    
    html = f"""
    <!DOCTYPE html>
    <html lang="zh-TW">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>æ¯”å°å ±å‘Š - {task_id}</title>
        <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css">
        <style>{css_content}</style>
    </head>
    <body>
        <div class="container">
            <div class="page-header">
                <h1 class="page-title">
                    <i class="fas fa-chart-line"></i> æ¯”å°çµæœå ±å‘Š
                </h1>
                <p class="page-subtitle">ä»»å‹™ ID: {task_id}</p>
                <p class="page-subtitle">ç”Ÿæˆæ™‚é–“: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
            
            <div class="report-content">
                {tables_html}
            </div>
        </div>
    </body>
    </html>
    """
    return html

def generate_html_report(task_id, data):
    """ç”Ÿæˆ HTML å ±å‘Šå…§å®¹"""
    html = f"""
    <!DOCTYPE html>
    <html lang="zh-TW">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>æ¯”å°å ±å‘Š - {task_id}</title>
        <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css">
        <style>
            {get_export_styles()}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="page-header">
                <h1 class="page-title">
                    <i class="fas fa-chart-line"></i> æ¯”å°çµæœå ±å‘Š
                </h1>
                <p class="page-subtitle">ä»»å‹™ ID: {task_id}</p>
            </div>
            
            <div class="report-content">
                {generate_data_tables(data)}
            </div>
        </div>
    </body>
    </html>
    """
    return html

def get_export_styles():
    """ç²å–åŒ¯å‡º HTML çš„æ¨£å¼ï¼ˆèˆ‡ get_embedded_report_styles ä¸åŒï¼Œé€™æ˜¯æ›´ç°¡æ½”çš„ç‰ˆæœ¬ï¼‰"""
    return """
    * {
        margin: 0;
        padding: 0;
        box-sizing: border-box;
    }
    
    body {
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
        background: #FAFBFD;
        color: #1A237E;
        line-height: 1.6;
    }
    
    .container {
        max-width: 1400px;
        margin: 0 auto;
        padding: 20px;
    }
    
    /* é é¢æ¨™é¡Œ - èˆ‡ download.css é¢¨æ ¼ä¸€è‡´ */
    .page-header {
        text-align: center;
        margin-bottom: 48px;
        padding: 48px 0;
        position: relative;
        background: white;
        border-radius: 20px;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.06);
    }
    
    .page-header::before {
        content: "";
        position: absolute;
        top: 0;
        left: 50%;
        transform: translateX(-50%);
        width: 120%;
        height: 100%;
        background: radial-gradient(ellipse at center, rgba(33, 150, 243, 0.03) 0%, transparent 70%);
        z-index: -1;
    }
    
    .page-title {
        font-size: 2.75rem;
        font-weight: 600;
        color: #1A237E;
        margin-bottom: 8px;
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 16px;
        letter-spacing: -0.02em;
        text-shadow: 0 2px 4px rgba(0, 33, 71, 0.05);
    }
    
    .page-title i {
        font-size: 2.5rem;
        color: #2196F3;
        filter: drop-shadow(0 2px 4px rgba(33, 150, 243, 0.15));
    }
    
    .page-subtitle {
        font-size: 1.125rem;
        color: #5C6BC0;
        font-weight: 400;
    }
    
    /* å ±è¡¨å…§å®¹å€å¡Š */
    .report-content {
        margin-top: 32px;
    }
    
    .sheet-section {
        background: white;
        border-radius: 20px;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.06);
        margin-bottom: 32px;
        overflow: hidden;
        border: 1px solid #EDF2FC;
        transition: all 0.3s ease;
    }
    
    .sheet-section:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.08);
    }
    
    .sheet-title {
        background: #2196F3;
        color: white;
        padding: 24px 32px;
        margin: 0;
        font-size: 1.375rem;
        font-weight: 600;
        display: flex;
        align-items: center;
        gap: 12px;
        position: relative;
        overflow: hidden;
    }
    
    .sheet-title::before {
        content: "";
        position: absolute;
        top: -100%;
        right: -25%;
        width: 70%;
        height: 300%;
        background: linear-gradient(120deg, transparent 0%, rgba(255, 255, 255, 0.4) 50%, transparent 100%);
        transform: rotate(35deg);
        opacity: 0.5;
    }
    
    .sheet-title i {
        font-size: 1.25rem;
    }
    
    /* è¡¨æ ¼å®¹å™¨ */
    .table-container {
        overflow-x: auto;
        background: #FAFBFD;
    }
    
    .table-wrapper {
        min-width: 100%;
        overflow-x: auto;
    }
    
    /* è¡¨æ ¼æ¨£å¼ */
    .data-table {
        width: 100%;
        border-collapse: collapse;
        font-size: 0.9rem;
        background: white;
    }
    
    .data-table th {
        background: #E3F2FD;
        color: #1976D2;
        padding: 14px 20px;
        text-align: left;
        font-weight: 600;
        border-bottom: 2px solid #BBDEFB;
        white-space: nowrap;
        position: sticky;
        top: 0;
        z-index: 10;
    }
    
    .data-table td {
        padding: 12px 20px;
        border-bottom: 1px solid #EDF2FC;
        color: #424242;
    }
    
    .data-table tr:hover td {
        background: #F5F9FF;
    }
    
    .data-table tr:nth-child(even) td {
        background: #FAFBFD;
    }
    
    .data-table tr:nth-child(even):hover td {
        background: #F0F7FF;
    }
    
    /* ç‰¹æ®Šæ¬„ä½æ¨£å¼ */
    .highlight-red {
        color: #F44336 !important;
        font-weight: 600;
    }
    
    .highlight-blue {
        color: #2196F3 !important;
    }
    
    /* å¾½ç« æ¨£å¼ */
    .badge {
        display: inline-block;
        padding: 4px 12px;
        border-radius: 9999px;
        font-size: 0.8125rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.025em;
    }
    
    .badge-success {
        background: #E8F5E9;
        color: #4CAF50;
        border: 1px solid #4CAF50;
    }
    
    .badge-danger {
        background: #FFEBEE;
        color: #F44336;
        border: 1px solid #F44336;
    }
    
    .badge-warning {
        background: #FFF3E0;
        color: #FF9800;
        border: 1px solid #FF9800;
    }
    
    .badge-info {
        background: #E3F2FD;
        color: #2196F3;
        border: 1px solid #2196F3;
    }
    
    /* é€£çµæ¨£å¼ */
    .link {
        color: #2196F3;
        text-decoration: none;
        font-weight: 500;
        transition: all 0.2s ease;
    }
    
    .link:hover {
        color: #1976D2;
        text-decoration: underline;
    }
    
    .link i {
        margin-left: 4px;
        font-size: 0.875em;
    }
    
    /* ç©ºè³‡æ–™æç¤º */
    .no-data {
        text-align: center;
        padding: 60px;
        color: #9FA8DA;
        font-size: 1.125rem;
    }
    
    .no-data i {
        font-size: 3rem;
        margin-bottom: 16px;
        opacity: 0.5;
    }
    
    /* åˆ—å°æ¨£å¼ */
    @media print {
        body {
            background: white;
        }
        
        .page-header {
            box-shadow: none;
            border: 1px solid #E0E0E0;
            page-break-after: avoid;
        }
        
        .sheet-section {
            box-shadow: none;
            border: 1px solid #E0E0E0;
            page-break-inside: avoid;
            margin-bottom: 20px;
        }
        
        .sheet-section:hover {
            transform: none;
        }
        
        .table-container {
            overflow: visible;
        }
        
        .data-table {
            font-size: 0.8rem;
        }
        
        .data-table th,
        .data-table td {
            padding: 8px 12px;
        }
    }
    
    /* éŸ¿æ‡‰å¼è¨­è¨ˆ */
    @media (max-width: 768px) {
        .container {
            padding: 10px;
        }
        
        .page-header {
            padding: 24px 16px;
        }
        
        .page-title {
            font-size: 2rem;
        }
        
        .sheet-title {
            padding: 16px 20px;
            font-size: 1.125rem;
        }
        
        .data-table {
            font-size: 0.8rem;
        }
        
        .data-table th,
        .data-table td {
            padding: 8px 12px;
        }
    }
    """

def generate_data_tables(data):
    """ç”Ÿæˆæ‰€æœ‰è³‡æ–™è¡¨çš„ HTML"""
    if not data:
        return '<div class="no-data"><i class="fas fa-inbox"></i><p>ç„¡è³‡æ–™</p></div>'
    
    tables_html = ""
    
    # æŒ‰ç…§å›ºå®šé †åºè™•ç†è³‡æ–™è¡¨
    sheet_order = ['revision_diff', 'branch_error', 'lost_project', 'version_diff', 'ç„¡æ³•æ¯”å°']
    
    # å…ˆè™•ç†å·²å®šç¾©é †åºçš„è³‡æ–™è¡¨
    for sheet_name in sheet_order:
        if sheet_name in data:
            sheet_data = data[sheet_name]
            icon = get_sheet_icon(sheet_name)
            tables_html += f"""
            <div class="sheet-section">
                <h2 class="sheet-title">
                    <i class="fas {icon}"></i> {format_sheet_name(sheet_name)}
                </h2>
                <div class="table-container">
                    <div class="table-wrapper">
                        {generate_single_table_html(sheet_name, sheet_data)}
                    </div>
                </div>
            </div>
            """
    
    # å†è™•ç†å…¶ä»–æœªå®šç¾©çš„è³‡æ–™è¡¨
    for sheet_name, sheet_data in data.items():
        if sheet_name not in sheet_order:
            icon = 'fa-table'
            tables_html += f"""
            <div class="sheet-section">
                <h2 class="sheet-title">
                    <i class="fas {icon}"></i> {format_sheet_name(sheet_name)}
                </h2>
                <div class="table-container">
                    <div class="table-wrapper">
                        {generate_single_table_html(sheet_name, sheet_data)}
                    </div>
                </div>
            </div>
            """
    
    return tables_html

def generate_single_table_html(sheet_name, sheet_data):
    """ç”Ÿæˆå–®ä¸€è¡¨æ ¼çš„ HTML"""
    if not sheet_data or 'data' not in sheet_data:
        return '<div class="no-data"><i class="fas fa-inbox"></i><p>æ­¤è³‡æ–™è¡¨ç„¡å…§å®¹</p></div>'
    
    columns = sheet_data.get('columns', [])
    data = sheet_data.get('data', [])
    
    if not columns or not data:
        return '<div class="no-data"><i class="fas fa-inbox"></i><p>æ­¤è³‡æ–™è¡¨ç„¡å…§å®¹</p></div>'
    
    # é–‹å§‹å»ºç«‹è¡¨æ ¼
    html = '<table class="data-table">'
    
    # è¡¨é ­
    html += '<thead><tr>'
    for col in columns:
        # ç‰¹æ®Šè™•ç†æŸäº›æ¬„ä½çš„æ¨™é¡Œæ¨£å¼
        header_class = ''
        if col in ['problem', 'å•é¡Œ', 'base_short', 'compare_short', 'base_revision', 'compare_revision']:
            header_class = 'highlight-header'
        html += f'<th class="{header_class}">{col}</th>'
    html += '</tr></thead>'
    
    # è¡¨èº«
    html += '<tbody>'
    for row_idx, row in enumerate(data):
        html += '<tr>'
        for col in columns:
            value = row.get(col, '')
            cell_class = ''
            formatted_value = str(value) if value is not None else ''
            
            # ç‰¹æ®Šè™•ç†æŸäº›æ¬„ä½
            if col in ['problem', 'å•é¡Œ'] and formatted_value:
                cell_class = 'highlight-red'
            elif col == 'ç‹€æ…‹':
                if formatted_value == 'æ–°å¢':
                    formatted_value = f'<span class="badge badge-success">{formatted_value}</span>'
                elif formatted_value == 'åˆªé™¤':
                    formatted_value = f'<span class="badge badge-danger">{formatted_value}</span>'
            elif col == 'has_wave':
                if formatted_value == 'Y':
                    formatted_value = f'<span class="badge badge-info">Y</span>'
                elif formatted_value == 'N':
                    formatted_value = f'<span class="badge badge-warning">N</span>'
            elif col in ['base_link', 'compare_link', 'link'] and formatted_value:
                # è™•ç†é€£çµ
                if formatted_value.startswith('http'):
                    short_url = formatted_value.split('/')[-1][:30] + '...' if len(formatted_value) > 50 else formatted_value
                    formatted_value = f'<a href="{formatted_value}" target="_blank" class="link" title="{formatted_value}">{short_url} <i class="fas fa-external-link-alt"></i></a>'
            elif col in ['base_short', 'compare_short', 'base_revision', 'compare_revision']:
                # Hash å€¼ç”¨ç­‰å¯¬å­—é«”
                formatted_value = f'<code>{formatted_value}</code>'
            
            html += f'<td class="{cell_class}">{formatted_value}</td>'
        html += '</tr>'
    html += '</tbody>'
    
    html += '</table>'
    return html

def get_sheet_icon(sheet_name):
    """æ ¹æ“šè³‡æ–™è¡¨åç¨±è¿”å›å°æ‡‰çš„åœ–æ¨™"""
    icon_map = {
        'revision_diff': 'fa-code-branch',
        'branch_error': 'fa-exclamation-triangle',
        'lost_project': 'fa-folder-minus',
        'version_diff': 'fa-file-alt',
        'ç„¡æ³•æ¯”å°': 'fa-times-circle'
    }
    return icon_map.get(sheet_name, 'fa-table')

def format_sheet_name(sheet_name):
    """æ ¼å¼åŒ–è³‡æ–™è¡¨åç¨±ç‚ºæ›´å‹å¥½çš„é¡¯ç¤ºåç¨±"""
    name_map = {
        'revision_diff': 'Revision å·®ç•°',
        'branch_error': 'åˆ†æ”¯éŒ¯èª¤',
        'lost_project': 'æ–°å¢/åˆªé™¤å°ˆæ¡ˆ',
        'version_diff': 'ç‰ˆæœ¬æª”æ¡ˆå·®ç•°',
        'ç„¡æ³•æ¯”å°': 'ç„¡æ³•æ¯”å°çš„æ¨¡çµ„'
    }
    return name_map.get(sheet_name, sheet_name)

def get_embedded_report_styles():
    """ç²å–å…§åµŒçš„å ±å‘Šæ¨£å¼"""
    return """
    body {
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
        background: #F5F7FA;
        margin: 0;
        padding: 0;
        color: #1A237E;
    }
    
    .container {
        max-width: 1400px;
        margin: 0 auto;
        padding: 20px;
    }
    
    .page-header {
        text-align: center;
        margin-bottom: 48px;
        padding: 48px 0;
        background: white;
        border-radius: 16px;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.06);
    }
    
    .page-title {
        font-size: 2.75rem;
        font-weight: 600;
        color: #1A237E;
        margin-bottom: 8px;
    }
    
    .page-subtitle {
        font-size: 1.125rem;
        color: #5C6BC0;
        margin: 0;
    }
    
    .sheet-section {
        background: white;
        border-radius: 16px;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.06);
        margin-bottom: 24px;
        overflow: hidden;
    }
    
    .sheet-title {
        background: #2196F3;
        color: white;
        padding: 20px 32px;
        margin: 0;
        font-size: 1.5rem;
        font-weight: 600;
    }
    
    .table-container {
        padding: 0;
        overflow-x: auto;
    }
    
    .data-table {
        width: 100%;
        border-collapse: collapse;
        font-size: 0.9rem;
    }
    
    .data-table th {
        background: #E3F2FD;
        color: #1976D2;
        padding: 14px 20px;
        text-align: left;
        font-weight: 600;
        border-bottom: 2px solid #BBDEFB;
        position: sticky;
        top: 0;
    }
    
    .data-table td {
        padding: 12px 20px;
        border-bottom: 1px solid #E8EAF6;
    }
    
    .data-table tr:hover td {
        background: #F5F7FA;
    }
    
    .data-table td.highlight-red {
        color: #F44336;
        font-weight: 600;
    }
    
    .badge {
        display: inline-block;
        padding: 4px 12px;
        border-radius: 12px;
        font-size: 0.8125rem;
        font-weight: 600;
    }
    
    .badge-success {
        background: #E8F5E9;
        color: #4CAF50;
    }
    
    .badge-danger {
        background: #FFEBEE;
        color: #F44336;
    }
    
    .link {
        color: #2196F3;
        text-decoration: none;
    }
    
    .link:hover {
        text-decoration: underline;
    }
    
    @media print {
        .page-header {
            box-shadow: none;
            border: 1px solid #E0E0E0;
        }
        
        .sheet-section {
            page-break-inside: avoid;
            box-shadow: none;
            border: 1px solid #E0E0E0;
        }
    }
    """

def generate_table_html(sheet_data):
    """ç”Ÿæˆè¡¨æ ¼ HTML"""
    if not sheet_data or 'data' not in sheet_data:
        return '<p>ç„¡è³‡æ–™</p>'
    
    columns = sheet_data.get('columns', [])
    data = sheet_data.get('data', [])
    
    if not columns or not data:
        return '<p>ç„¡è³‡æ–™</p>'
    
    # é–‹å§‹å»ºç«‹è¡¨æ ¼
    html = '<table class="data-table">'
    
    # è¡¨é ­
    html += '<thead><tr>'
    for col in columns:
        html += f'<th>{col}</th>'
    html += '</tr></thead>'
    
    # è¡¨èº«
    html += '<tbody>'
    for row in data:
        html += '<tr>'
        for col in columns:
            value = row.get(col, '')
            cell_class = ''
            
            # ç‰¹æ®Šè™•ç†æŸäº›æ¬„ä½
            if col in ['problem', 'å•é¡Œ'] and value:
                cell_class = 'highlight-red'
            elif col == 'ç‹€æ…‹':
                if value == 'æ–°å¢':
                    value = f'<span class="badge badge-success">{value}</span>'
                elif value == 'åˆªé™¤':
                    value = f'<span class="badge badge-danger">{value}</span>'
            elif col in ['base_link', 'compare_link', 'link'] and value:
                # è™•ç†é€£çµ
                if value.startswith('http'):
                    value = f'<a href="{value}" target="_blank" class="link">{value} <i class="fas fa-external-link-alt"></i></a>'
            
            html += f'<td class="{cell_class}">{value}</td>'
        html += '</tr>'
    html += '</tbody>'
    
    html += '</table>'
    return html

@app.route('/api/export-excel-single/<task_id>/<sheet_name>')
def export_excel_single_sheet(task_id, sheet_name):
    """
    åŒ¯å‡ºåŸå§‹ Excel æª”æ¡ˆä¸­çš„å–®ä¸€è³‡æ–™è¡¨
    ä¿ç•™åŸå§‹æ ¼å¼ï¼Œåªç§»é™¤å…¶ä»–è³‡æ–™è¡¨
    """
    try:
        # æŸ¥æ‰¾åŸå§‹çš„æ¯”å°çµæœæª”æ¡ˆ
        excel_path = None
        
        # 1. å¾æ¯”å°çµæœç›®éŒ„æŸ¥æ‰¾
        compare_dir = os.path.join('compare_results', task_id)
        if os.path.exists(compare_dir):
            # å„ªå…ˆæŸ¥æ‰¾ all_scenarios_compare.xlsx æˆ– all_compare.xlsx
            priority_files = ['all_scenarios_compare.xlsx', 'all_compare.xlsx']
            for filename in priority_files:
                file_path = os.path.join(compare_dir, filename)
                if os.path.exists(file_path):
                    excel_path = file_path
                    break
            
            # å¦‚æœæ²’æ‰¾åˆ°ï¼ŒæŸ¥æ‰¾ä»»ä½• xlsx æª”æ¡ˆ
            if not excel_path:
                for file in os.listdir(compare_dir):
                    if file.endswith('.xlsx') and not file.startswith('~'):
                        excel_path = os.path.join(compare_dir, file)
                        break
        
        # 2. å¾è™•ç†ç‹€æ…‹ä¸­æŸ¥æ‰¾
        if not excel_path and task_id in processing_status:
            task_data = processing_status[task_id]
            results = task_data.get('results', {})
            if 'summary_report' in results and os.path.exists(results['summary_report']):
                excel_path = results['summary_report']
            elif 'compare_results' in results:
                compare_results = results['compare_results']
                if isinstance(compare_results, dict) and 'summary_report' in compare_results:
                    if os.path.exists(compare_results['summary_report']):
                        excel_path = compare_results['summary_report']
        
        if not excel_path or not os.path.exists(excel_path):
            app.logger.error(f'æ‰¾ä¸åˆ° Excel æª”æ¡ˆ: task_id={task_id}')
            return jsonify({'error': 'æ‰¾ä¸åˆ°åŸå§‹æª”æ¡ˆ'}), 404
        
        # è®€å–åŸå§‹ Excel æª”æ¡ˆ
        wb = openpyxl.load_workbook(excel_path, data_only=False, keep_vba=False)
        
        # æª¢æŸ¥è³‡æ–™è¡¨æ˜¯å¦å­˜åœ¨
        if sheet_name not in wb.sheetnames:
            app.logger.error(f'æ‰¾ä¸åˆ°è³‡æ–™è¡¨: {sheet_name} in {excel_path}')
            return jsonify({'error': f'æ‰¾ä¸åˆ°è³‡æ–™è¡¨: {sheet_name}'}), 404
        
        # å‰µå»ºæ–°çš„å·¥ä½œç°¿ï¼ŒåªåŒ…å«æŒ‡å®šçš„è³‡æ–™è¡¨
        new_wb = openpyxl.Workbook()
        
        # ç§»é™¤é è¨­çš„å·¥ä½œè¡¨
        default_sheet = new_wb.active
        new_wb.remove(default_sheet)
        
        # è¤‡è£½æŒ‡å®šçš„å·¥ä½œè¡¨ï¼ˆåŒ…å«æ ¼å¼ï¼‰
        source_sheet = wb[sheet_name]
        target_sheet = new_wb.create_sheet(title=sheet_name)
        
        # è¤‡è£½å„²å­˜æ ¼è³‡æ–™å’Œæ ¼å¼
        for row in source_sheet.iter_rows():
            for cell in row:
                target_cell = target_sheet.cell(
                    row=cell.row, 
                    column=cell.column, 
                    value=cell.value
                )
                
                # è¤‡è£½æ ¼å¼
                if cell.has_style:
                    target_cell.font = copy(cell.font)
                    target_cell.fill = copy(cell.fill)
                    target_cell.border = copy(cell.border)
                    target_cell.alignment = copy(cell.alignment)
                    target_cell.number_format = cell.number_format
                    target_cell.protection = copy(cell.protection)
        
        # è¤‡è£½åˆ—å¯¬
        for column_cells in source_sheet.columns:
            column_letter = column_cells[0].column_letter
            if source_sheet.column_dimensions[column_letter].width:
                target_sheet.column_dimensions[column_letter].width = \
                    source_sheet.column_dimensions[column_letter].width
        
        # è¤‡è£½è¡Œé«˜
        for row_cells in source_sheet.rows:
            row_number = row_cells[0].row
            if source_sheet.row_dimensions[row_number].height:
                target_sheet.row_dimensions[row_number].height = \
                    source_sheet.row_dimensions[row_number].height
        
        # è¤‡è£½åˆä½µå„²å­˜æ ¼
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))
        
        # è¤‡è£½è‡ªå‹•ç¯©é¸
        if source_sheet.auto_filter.ref:
            target_sheet.auto_filter.ref = source_sheet.auto_filter.ref
        
        # å„²å­˜åˆ°è¨˜æ†¶é«”
        output = io.BytesIO()
        new_wb.save(output)
        output.seek(0)
        
        # ç”Ÿæˆæª”æ¡ˆåç¨±
        filename = f"{sheet_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # å›å‚³æª”æ¡ˆ
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        app.logger.error(f"åŒ¯å‡ºå–®ä¸€è³‡æ–™è¡¨éŒ¯èª¤: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/check-excel-columns', methods=['POST'])
def check_excel_columns():
    """æª¢æŸ¥ Excel æª”æ¡ˆçš„æ¬„ä½"""
    try:
        data = request.json
        filepath = data.get('filepath')
        
        if not filepath or not os.path.exists(filepath):
            return jsonify({'error': 'æª”æ¡ˆä¸å­˜åœ¨'}), 404
        
        # ä½¿ç”¨ ExcelHandler æª¢æŸ¥æ¬„ä½
        result = excel_handler.check_excel_columns(filepath)
        
        app.logger.info(f"æª¢æŸ¥ Excel æ¬„ä½çµæœ: {result}")
        
        return jsonify(result)
        
    except Exception as e:
        app.logger.error(f"æª¢æŸ¥ Excel æ¬„ä½å¤±æ•—: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-complete/<task_id>', methods=['POST'])
def handle_download_complete(task_id):
    """è™•ç†ä¸‹è¼‰å®Œæˆäº‹ä»¶ - ç”¨æ–¼æ‰‹å‹•è§¸ç™¼ Excel æª”æ¡ˆè™•ç†"""
    try:
        data = request.json
        excel_file = data.get('excel_file')
        
        # ç²å–ä¸‹è¼‰è³‡æ–™å¤¾è·¯å¾‘
        download_folder = os.path.join('downloads', task_id)
        
        if not os.path.exists(download_folder):
            return jsonify({'error': 'ä¸‹è¼‰è³‡æ–™å¤¾ä¸å­˜åœ¨'}), 404
        
        # ç²å– Excel å…ƒè³‡æ–™
        excel_metadata = uploaded_excel_metadata.get(excel_file)
        
        if not excel_metadata:
            return jsonify({
                'excel_copied': False,
                'message': 'æ²’æœ‰ Excel å…ƒè³‡æ–™'
            })
        
        # è™•ç† Excel æª”æ¡ˆ
        excel_result = excel_handler.process_download_complete(
            task_id,
            download_folder,
            excel_metadata
        )
        
        # æ›´æ–°ä»»å‹™çµæœ
        if task_id in processing_status:
            processing_status[task_id]['results'].update(excel_result)
        
        return jsonify(excel_result)
        
    except Exception as e:
        app.logger.error(f"è™•ç†ä¸‹è¼‰å®Œæˆäº‹ä»¶å¤±æ•—: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/copy-excel-to-results', methods=['POST'])
def copy_excel_to_results():
    """è¤‡è£½ Excel æª”æ¡ˆåˆ°çµæœè³‡æ–™å¤¾ API"""
    try:
        data = request.json
        task_id = data.get('task_id')
        original_filepath = data.get('original_filepath')
        new_filename = data.get('new_filename')
        
        if not all([task_id, original_filepath]):
            return jsonify({'error': 'ç¼ºå°‘å¿…è¦åƒæ•¸'}), 400
        
        # ç²å–ä¸‹è¼‰è³‡æ–™å¤¾
        download_folder = os.path.join('downloads', task_id)
        
        if not os.path.exists(download_folder):
            os.makedirs(download_folder, exist_ok=True)
        
        # å¦‚æœæ²’æœ‰æŒ‡å®šæ–°æª”åï¼Œå¾å…ƒè³‡æ–™æ±ºå®š
        if not new_filename:
            excel_metadata = uploaded_excel_metadata.get(original_filepath, {})
            root_folder = excel_metadata.get('root_folder')
            
            if root_folder == 'DailyBuild':
                new_filename = 'DailyBuild_mapping.xlsx'
            elif root_folder in ['/DailyBuild/PrebuildFW', 'PrebuildFW']:
                new_filename = 'PrebuildFW_mapping.xlsx'
            else:
                new_filename = os.path.basename(original_filepath)
        
        # è¤‡è£½æª”æ¡ˆ
        target_path = os.path.join(download_folder, new_filename)
        shutil.copy2(original_filepath, target_path)
        
        app.logger.info(f"Excel æª”æ¡ˆå·²è¤‡è£½: {original_filepath} -> {target_path}")
        
        return jsonify({
            'success': True,
            'new_path': target_path,
            'new_filename': new_filename
        })
        
    except Exception as e:
        app.logger.error(f"è¤‡è£½ Excel æª”æ¡ˆå¤±æ•—: {str(e)}")
        return jsonify({'error': str(e)}), 500

# åœ¨ web_app.py çš„ API è·¯ç”±éƒ¨åˆ†æ–°å¢
@app.route('/api/results-structure/<task_id>')
def get_results_structure(task_id):
    """ç²å–çµæœæª”æ¡ˆçµæ§‹ API"""
    try:
        structure = {
            'task_id': task_id,
            'scenarios': {}
        }
        
        # æª¢æŸ¥æ¯”å°çµæœç›®éŒ„
        compare_dir = os.path.join('compare_results', task_id)
        
        if not os.path.exists(compare_dir):
            return jsonify({'error': 'æ‰¾ä¸åˆ°çµæœç›®éŒ„'}), 404
        
        # æª¢æŸ¥å„æƒ…å¢ƒå­ç›®éŒ„
        scenarios = ['master_vs_premp', 'premp_vs_wave', 'wave_vs_backup']
        
        for scenario in scenarios:
            scenario_dir = os.path.join(compare_dir, scenario)
            if os.path.exists(scenario_dir):
                scenario_data = {
                    'path': scenario_dir,
                    'files': []
                }
                
                # åˆ—å‡ºè©²æƒ…å¢ƒç›®éŒ„ä¸‹çš„æ‰€æœ‰æª”æ¡ˆ
                for file in os.listdir(scenario_dir):
                    if file.endswith('.xlsx'):
                        file_path = os.path.join(scenario_dir, file)
                        file_stat = os.stat(file_path)
                        scenario_data['files'].append({
                            'name': file,
                            'path': file_path,
                            'size': file_stat.st_size,
                            'modified': file_stat.st_mtime
                        })
                
                structure['scenarios'][scenario] = scenario_data
        
        return jsonify(structure)
        
    except Exception as e:
        app.logger.error(f'Get results structure error: {e}')
        return jsonify({'error': str(e)}), 500

@app.route('/api/check-scenarios/<task_id>')
def check_scenarios(task_id):
    """æª¢æŸ¥æƒ…å¢ƒæª”æ¡ˆæ˜¯å¦å­˜åœ¨"""
    base_path = f"compare_results/{task_id}"
    
    files_to_check = {
        'all': f"{base_path}/all_scenarios_summary.xlsx",
        'master_vs_premp': f"{base_path}/master_vs_premp/all_scenarios_compare.xlsx", 
        'premp_vs_wave': f"{base_path}/premp_vs_wave/all_scenarios_compare.xlsx",
        'wave_vs_backup': f"{base_path}/wave_vs_backup/all_scenarios_compare.xlsx"
    }
    
    scenario_status = {}
    for scenario, file_path in files_to_check.items():
        scenario_status[scenario] = os.path.exists(file_path)
    
    return jsonify(scenario_status)
    
if __name__ == '__main__':
    # é–‹ç™¼æ¨¡å¼åŸ·è¡Œ
    socketio.run(app, debug=True, host='0.0.0.0', port=5000)