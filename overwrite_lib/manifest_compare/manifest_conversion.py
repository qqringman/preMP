#!/usr/bin/env python3
"""
Manifest 比較工具 - 純比較版本
專門用於比較兩個 manifest.xml 檔案，不執行任何轉換
支援本地檔案與 Gerrit manifest 比較
"""

import os
import sys
import xml.etree.ElementTree as ET
import pandas as pd
from typing import Dict, List, Tuple, Optional, Any
import argparse
from datetime import datetime
import logging
import tempfile
import subprocess
import shutil
import re

# 添加專案路徑
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

import config
from excel_handler import ExcelHandler
from gerrit_manager import GerritManager
import utils

# 設定日誌
logger = utils.setup_logger(__name__)

class ManifestComparator:
    """Manifest 比較器 - 純比較版本，不執行轉換"""
    
    def __init__(self):
        """初始化方法"""
        self.excel_handler = ExcelHandler()
        self.gerrit_manager = GerritManager()
        self.logger = logger
        
        # 檔案路徑記錄
        self.local_file_path = None
        self.gerrit_file_path = None
        self.expanded_file_path = None
        self.use_expanded = False
        
        # 🔥 使用 config.py 動態生成 Gerrit 基礎 URL
        base_path = config.get_gerrit_manifest_base_path()
        self.gerrit_base_url = f"https://mm2sd.rtkbf.com/gerrit/plugins/gitiles/realtek/android/manifest/+/refs/heads/{base_path}"
        
        self.logger.info(f"🔧 初始化 ManifestComparator")
        self.logger.info(f"   當前 Android 版本: {config.get_current_android_version()}")
        self.logger.info(f"   Gerrit 基礎 URL: {self.gerrit_base_url}")
        
        # 🔥 使用 config.py 中的函數直接生成 Gerrit 檔案 URL 映射
        self.gerrit_urls = {
            'master': {
                'filename': 'atv-google-refplus.xml',
                'url': config.get_master_manifest_url()
            },
            'premp': {
                'filename': 'atv-google-refplus-premp.xml',
                'url': config.get_premp_manifest_url()
            },
            'mp': {
                'filename': 'atv-google-refplus-wave.xml',
                'url': config.get_mp_manifest_url()
            },
            'mp_backup': {
                'filename': 'atv-google-refplus-wave-backup.xml',
                'url': config.get_mp_backup_manifest_url()
            }
        }
        
        # 記錄所有生成的 URL 供調試
        self.logger.debug(f"📋 Gerrit URL 映射:")
        for key, value in self.gerrit_urls.items():
            self.logger.debug(f"   {key.upper()}: {value['url']}")
    
    def compare_local_with_gerrit(self, local_file: str, gerrit_type: str, output_file: str) -> bool:
        """比較本地檔案與 Gerrit manifest 檔案"""
        try:
            self.logger.info("=" * 80)
            self.logger.info(f"開始執行本地檔案與 {gerrit_type.upper()} 比較（純比較版本）")
            self.logger.info("=" * 80)
            
            # 確保輸出資料夾存在
            output_folder = os.path.dirname(output_file)
            if not output_folder:
                output_folder = "."
            utils.ensure_dir(output_folder)
            
            # 步驟 1: 複製本地檔案到 output 目錄
            self.logger.info("\n📋 步驟 1: 複製本地檔案到輸出目錄")
            self.local_file_path = self._copy_local_file_to_output(local_file, output_folder)
            
            # 步驟 2: 從 Gerrit 下載檔案
            self.logger.info(f"\n⬇️ 步驟 2: 從 Gerrit 下載 {gerrit_type.upper()} manifest")
            self.gerrit_file_path = self._download_gerrit_file(gerrit_type, output_folder)
            
            if not self.gerrit_file_path:
                self.logger.error(f"❌ 無法下載 {gerrit_type.upper()} manifest")
                return False
            
            # 步驟 3: 檢查並處理 Gerrit 檔案的 include 展開
            self.logger.info(f"\n🔍 步驟 3: 檢查 {gerrit_type.upper()} manifest 是否需要展開")
            actual_gerrit_file = self._handle_gerrit_include_expansion(self.gerrit_file_path, output_folder)
            
            # 步驟 4: 讀取檔案內容並執行比較分析
            self.logger.info(f"\n📊 步驟 4: 執行比較分析")
            
            with open(self.local_file_path, 'r', encoding='utf-8') as f:
                local_content = f.read()
            
            with open(actual_gerrit_file, 'r', encoding='utf-8') as f:
                gerrit_content = f.read()
            
            # 執行差異分析
            diff_analysis = self._analyze_differences(
                local_content, gerrit_content, f"local_vs_{gerrit_type}"
            )
            
            # 步驟 5: 生成 Excel 報告
            self.logger.info(f"\n📄 步驟 5: 生成 Excel 報告")
            
            success = self._generate_excel_report(
                f"local_vs_{gerrit_type}", self.local_file_path, actual_gerrit_file,
                diff_analysis, output_folder, os.path.basename(output_file),
                True, True, self.expanded_file_path, self.use_expanded
            )
            
            # 步驟 6: 顯示結果統計
            self._show_comparison_results(f"local_vs_{gerrit_type}", diff_analysis)
            
            return success
            
        except Exception as e:
            self.logger.error(f"本地檔案與 {gerrit_type.upper()} 比較執行失敗: {str(e)}")
            import traceback
            self.logger.error(f"詳細錯誤: {traceback.format_exc()}")
            return False
    
    def compare_local_files(self, file1: str, file2: str, output_file: str) -> bool:
        """比較兩個本地檔案"""
        try:
            self.logger.info("=" * 80)
            self.logger.info(f"開始執行本地檔案比較（純比較版本）")
            self.logger.info("=" * 80)
            
            # 確保輸出資料夾存在
            output_folder = os.path.dirname(output_file)
            if not output_folder:
                output_folder = "."
            utils.ensure_dir(output_folder)
            
            # 複製檔案到輸出目錄
            self.logger.info("\n📋 複製檔案到輸出目錄")
            file1_dest = self._copy_local_file_to_output(file1, output_folder)
            file2_dest = self._copy_local_file_to_output(file2, output_folder)
            
            # 讀取檔案內容
            with open(file1_dest, 'r', encoding='utf-8') as f:
                content1 = f.read()
            
            with open(file2_dest, 'r', encoding='utf-8') as f:
                content2 = f.read()
            
            self.logger.info(f"✅ 檔案讀取完成:")
            self.logger.info(f"   檔案1 內容長度: {len(content1)}")
            self.logger.info(f"   檔案2 內容長度: {len(content2)}")
            
            # 執行差異分析
            self.logger.info("\n🔍 開始執行差異分析")
            diff_analysis = self._analyze_differences(content1, content2, "local_vs_local")
            
            # 生成 Excel 報告
            success = self._generate_excel_report(
                "local_vs_local", file1_dest, file2_dest,
                diff_analysis, output_folder, os.path.basename(output_file),
                True, True, None, False
            )
            
            self._show_comparison_results("local_vs_local", diff_analysis)
            
            return success
            
        except Exception as e:
            self.logger.error(f"本地檔案比較執行失敗: {str(e)}")
            import traceback
            self.logger.error(f"詳細錯誤: {traceback.format_exc()}")
            return False

    # ===============================
    # ===== 核心比較分析方法 =====
    # ===============================

    def _analyze_differences(self, source_content: str, target_content: str, comparison_type: str) -> Dict[str, Any]:
        """分析兩個檔案的差異"""
        try:
            self.logger.info(f"🔍 開始分析差異:")
            self.logger.info(f"   比較類型: {comparison_type}")
            self.logger.info(f"   來源內容長度: {len(source_content)}")
            self.logger.info(f"   目標內容長度: {len(target_content)}")
            
            # 解析兩個檔案的專案
            source_projects = self._extract_projects_with_line_numbers(source_content)
            target_projects = self._extract_projects_with_line_numbers(target_content)
            
            if not source_projects:
                self.logger.error("❌ 來源檔案解析失敗")
                return self._create_empty_analysis()
            
            if not target_projects:
                self.logger.error("❌ 目標檔案解析失敗")
                return self._create_empty_analysis()
            
            self.logger.info(f"✅ 檔案解析完成:")
            self.logger.info(f"   來源檔案專案數: {len(source_projects)}")
            self.logger.info(f"   目標檔案專案數: {len(target_projects)}")
            
            # 創建項目資訊列表
            project_info_list = self._create_project_info_list(source_projects, target_projects)
            
            # 進行差異比較
            differences = self._compare_projects(source_projects, target_projects, comparison_type)
            
            # 統計摘要
            actual_differences_count = sum(1 for diff in differences if diff['comparison_status'] != '✔️ 相同')
            identical_count = len(differences) - actual_differences_count
            
            summary = {
                'source_count': len(source_projects),
                'target_count': len(target_projects),
                'total_compared': len(project_info_list),
                'differences_count': actual_differences_count,
                'identical_count': identical_count
            }
            
            analysis = {
                'has_target': True,
                'source_projects': source_projects,
                'target_projects': target_projects,
                'project_info_list': project_info_list,
                'differences': differences,
                'summary': summary
            }
            
            self.logger.info(f"📊 差異分析完成:")
            self.logger.info(f"   總比較項目: {summary['total_compared']}")
            self.logger.info(f"   差異項目: {summary['differences_count']}")
            self.logger.info(f"   相同項目: {summary['identical_count']}")
            
            return analysis
            
        except Exception as e:
            self.logger.error(f"差異分析失敗: {str(e)}")
            import traceback
            self.logger.error(f"詳細錯誤: {traceback.format_exc()}")
            return self._create_empty_analysis()

    def _create_empty_analysis(self) -> Dict[str, Any]:
        """創建空的分析結果"""
        return {
            'has_target': False,
            'source_projects': [],
            'target_projects': [],
            'project_info_list': [],
            'differences': [],
            'summary': {
                'source_count': 0,
                'target_count': 0,
                'total_compared': 0,
                'differences_count': 0,
                'identical_count': 0
            }
        }

    def _create_project_info_list(self, source_projects: List[Dict], target_projects: List[Dict]) -> List[Dict]:
        """創建專案資訊列表（用於比較後專案頁籤）"""
        try:
            # 建立目標專案索引
            target_index = {}
            for proj in target_projects:
                key = f"{proj['name']}|||{proj['path']}"
                target_index[key] = proj
            
            project_info_list = []
            
            # 處理來源檔案的專案
            for i, source_proj in enumerate(source_projects, 1):
                key = f"{source_proj['name']}|||{source_proj['path']}"
                target_proj = target_index.get(key)
                
                if target_proj:
                    # 專案在兩個檔案中都存在
                    source_rev = source_proj['revision']
                    target_rev = target_proj['revision']
                    
                    is_same = (source_rev == target_rev)
                    
                    if is_same:
                        description = f"兩檔案版本相同: {source_rev}"
                    else:
                        description = f"來源檔案: {source_rev} → 目標檔案: {target_rev}"
                    
                    # 🔥 修正：確保使用目標專案的 dest-branch（如果存在），否則使用來源專案的
                    dest_branch = target_proj.get('dest-branch', '') or source_proj.get('dest-branch', '')
                    
                    project_info = {
                        'SN': i,
                        'name': source_proj['name'],
                        'path': source_proj['path'],
                        'source_revision': source_rev,
                        'target_revision': target_rev,
                        'revision_equal': '',  # 將用 Excel 公式填充
                        'description': description,
                        'upstream': target_proj['upstream'],
                        'dest-branch': dest_branch,  # 🔥 重要：確保包含 dest-branch
                        'groups': target_proj['groups'],
                        'clone-depth': target_proj['clone-depth'],
                        'remote': target_proj['remote'],
                        'found_in_target': True
                    }
                else:
                    # 專案只在來源檔案中存在
                    project_info = {
                        'SN': i,
                        'name': source_proj['name'],
                        'path': source_proj['path'],
                        'source_revision': source_proj['revision'],
                        'target_revision': 'N/A (專案不存在)',
                        'revision_equal': '',
                        'description': '專案僅存在於來源檔案',
                        'upstream': source_proj['upstream'],
                        'dest-branch': source_proj['dest-branch'],  # 🔥 重要：確保包含 dest-branch
                        'groups': source_proj['groups'],
                        'clone-depth': source_proj['clone-depth'],
                        'remote': source_proj['remote'],
                        'found_in_target': False
                    }
                
                project_info_list.append(project_info)
            
            self.logger.info(f"✅ 創建專案資訊列表完成: {len(project_info_list)} 個專案（包含 dest-branch）")
            return project_info_list
            
        except Exception as e:
            self.logger.error(f"創建專案資訊列表失敗: {str(e)}")
            import traceback
            self.logger.error(f"詳細錯誤: {traceback.format_exc()}")
            return []

    def _format_summary_sheet_in_context(self, worksheet, comparison_type: str, target_file_path: str):
        """在 ExcelWriter context 內格式化比較摘要頁籤"""
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            
            # 定義背景色
            light_blue_fill = PatternFill(start_color="E8F0FF", end_color="E8F0FF", fill_type="solid")
            light_red_fill = PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid")
            light_green_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            black_font = Font(color="000000", bold=False)
            
            # 為數據行設定背景色
            for col in range(1, worksheet.max_column + 1):
                header_value = str(worksheet.cell(row=1, column=col).value) if worksheet.cell(row=1, column=col).value else ''
                data_cell = worksheet.cell(row=2, column=col)
                
                # 根據欄位類型設定背景色
                if header_value in ['📊 來源檔案專案數', '🎯 目標檔案專案數', '📋 總比較專案數']:
                    data_cell.fill = light_blue_fill
                    data_cell.font = black_font
                elif header_value in ['❌ 差異專案數']:
                    data_cell.fill = light_red_fill
                    data_cell.font = black_font
                elif header_value in ['✔️ 相同專案數', '📈 相同率']:
                    data_cell.fill = light_green_fill
                    data_cell.font = black_font
                
                # 🔥 為目標檔案添加超連結
                if header_value == '目標檔案' and comparison_type != "local_vs_local":
                    filename = str(data_cell.value) if data_cell.value else ''
                    if filename and filename not in ['', 'N/A']:
                        # 移除 gerrit_ 前綴來生成正確的連結
                        clean_filename = filename.replace('gerrit_', '') if filename.startswith('gerrit_') else filename
                        gerrit_url = self._generate_gerrit_manifest_link(clean_filename)
                        self._add_hyperlink_formula_to_cell_in_context(worksheet, 2, col, gerrit_url, filename)
                        self.logger.info(f"✅ 已為比較摘要的目標檔案添加超連結: {filename}")
            
        except Exception as e:
            self.logger.error(f"格式化比較摘要頁籤失敗: {str(e)}")

    def _compare_projects(self, source_projects: List[Dict], target_projects: List[Dict], comparison_type: str) -> List[Dict]:
        """比較專案並生成差異列表（包含相同和差異的專案）"""
        try:
            differences = []
            
            # 判斷比較模式
            is_local_comparison = (comparison_type == "local_vs_local")
            
            # 建立目標專案索引
            target_index = {}
            for proj in target_projects:
                key = f"{proj['name']}|||{proj['path']}"
                target_index[key] = proj
            
            # 獲得檔案名稱
            source_filename, target_filename = self._get_filenames_for_comparison(comparison_type)
            
            # 比較每個來源專案
            for source_proj in source_projects:
                key = f"{source_proj['name']}|||{source_proj['path']}"
                target_proj = target_index.get(key)
                
                if target_proj:
                    # 專案在兩個檔案中都存在，比較屬性
                    diff_details = self._get_detailed_differences_between_projects(source_proj, target_proj)
                    
                    if diff_details:
                        # 有差異
                        comparison_status = '❌ 不同'
                        comparison_result = self._format_difference_summary(diff_details)
                        status_color = 'red'
                    else:
                        # 完全相同
                        comparison_status = '✔️ 相同'
                        comparison_result = '兩檔案中此專案的所有屬性完全一致'
                        status_color = 'green'
                    
                    # 根據比較模式設定欄位名稱
                    if is_local_comparison:
                        target_prefix = 'compare_'
                    else:
                        target_prefix = 'gerrit_'
                    
                    difference = {
                        'SN': len(differences) + 1,
                        'source_file': source_filename,
                        'content': source_proj.get('full_line', ''),
                        'name': source_proj['name'],
                        'path': source_proj['path'],
                        'revision': source_proj['revision'],
                        'upstream': source_proj['upstream'],
                        'dest-branch': source_proj['dest-branch'],
                        'groups': source_proj['groups'],
                        'clone-depth': source_proj['clone-depth'],
                        'remote': source_proj['remote'],
                        'source_link': self._generate_source_link(source_proj['name'], source_proj['revision'], source_proj['remote']),
                        f'{target_prefix}source_file': target_filename,
                        f'{target_prefix}content': target_proj.get('full_line', ''),
                        f'{target_prefix}name': target_proj['name'],
                        f'{target_prefix}path': target_proj['path'],
                        f'{target_prefix}revision': target_proj['revision'],
                        f'{target_prefix}upstream': target_proj['upstream'],
                        f'{target_prefix}dest-branch': target_proj['dest-branch'],
                        f'{target_prefix}groups': target_proj['groups'],
                        f'{target_prefix}clone-depth': target_proj['clone-depth'],
                        f'{target_prefix}remote': target_proj['remote'],
                        f'{target_prefix}source_link': self._generate_source_link(target_proj['name'], target_proj['revision'], target_proj['remote']),
                        'comparison_status': comparison_status,
                        'comparison_result': comparison_result,
                        'status_color': status_color
                    }
                    
                    # 🔥 修正：所有專案都加入列表（包含相同的）
                    differences.append(difference)
                else:
                    # 專案只在來源檔案中存在
                    difference = {
                        'SN': len(differences) + 1,
                        'source_file': source_filename,
                        'content': source_proj.get('full_line', ''),
                        'name': source_proj['name'],
                        'path': source_proj['path'],
                        'revision': source_proj['revision'],
                        'upstream': source_proj['upstream'],
                        'dest-branch': source_proj['dest-branch'],
                        'groups': source_proj['groups'],
                        'clone-depth': source_proj['clone-depth'],
                        'remote': source_proj['remote'],
                        'source_link': self._generate_source_link(source_proj['name'], source_proj['revision'], source_proj['remote']),
                        f'{target_prefix}source_file': target_filename,
                        f'{target_prefix}content': 'N/A (專案不存在)',
                        f'{target_prefix}name': '',
                        f'{target_prefix}path': '',
                        f'{target_prefix}revision': '',
                        f'{target_prefix}upstream': '',
                        f'{target_prefix}dest-branch': '',
                        f'{target_prefix}groups': '',
                        f'{target_prefix}clone-depth': '',
                        f'{target_prefix}remote': '',
                        f'{target_prefix}source_link': '',
                        'comparison_status': '➕ 新增',
                        'comparison_result': '專案僅存在於來源檔案，目標檔案無此專案',
                        'status_color': 'orange'
                    }
                    differences.append(difference)
            
            # 檢查目標檔案中存在但來源檔案不存在的專案
            source_keys = set(f"{proj['name']}|||{proj['path']}" for proj in source_projects)
            
            for key, target_proj in target_index.items():
                if key not in source_keys:
                    difference = {
                        'SN': len(differences) + 1,
                        'source_file': source_filename,
                        'content': 'N/A (專案不存在)',
                        'name': '',
                        'path': '',
                        'revision': '',
                        'upstream': '',
                        'dest-branch': '',
                        'groups': '',
                        'clone-depth': '',
                        'remote': '',
                        'source_link': '',
                        f'{target_prefix}source_file': target_filename,
                        f'{target_prefix}content': target_proj.get('full_line', ''),
                        f'{target_prefix}name': target_proj['name'],
                        f'{target_prefix}path': target_proj['path'],
                        f'{target_prefix}revision': target_proj['revision'],
                        f'{target_prefix}upstream': target_proj['upstream'],
                        f'{target_prefix}dest-branch': target_proj['dest-branch'],
                        f'{target_prefix}groups': target_proj['groups'],
                        f'{target_prefix}clone-depth': target_proj['clone-depth'],
                        f'{target_prefix}remote': target_proj['remote'],
                        f'{target_prefix}source_link': self._generate_source_link(target_proj['name'], target_proj['revision'], target_proj['remote']),
                        'comparison_status': '❓ 無此專案',
                        'comparison_result': '專案僅存在於目標檔案，來源檔案無此專案',
                        'status_color': 'blue'
                    }
                    differences.append(difference)
            
            self.logger.info(f"📊 專案比較完成:")
            self.logger.info(f"   總差異記錄數: {len(differences)}")
            
            return differences
            
        except Exception as e:
            self.logger.error(f"比較專案失敗: {str(e)}")
            return []

    def _get_filenames_for_comparison(self, comparison_type: str) -> Tuple[str, str]:
        """根據比較類型獲得檔案名稱"""
        if comparison_type == "local_vs_local":
            return "本地檔案1", "本地檔案2"
        elif comparison_type == "local_vs_master":
            return "本地檔案", "atv-google-refplus.xml"
        elif comparison_type == "local_vs_premp":
            return "本地檔案", "atv-google-refplus-premp.xml"
        elif comparison_type == "local_vs_mp":
            return "本地檔案", "atv-google-refplus-wave.xml"
        elif comparison_type == "local_vs_mp_backup":
            return "本地檔案", "atv-google-refplus-wave-backup.xml"
        else:
            return "來源檔案", "目標檔案"

    def _extract_projects_with_line_numbers(self, xml_content: str) -> List[Dict[str, Any]]:
        """提取專案資訊並記錄行號"""
        try:
            if not xml_content or not xml_content.strip():
                self.logger.warning("XML 內容為空")
                return []
            
            projects = []
            lines = xml_content.split('\n')
            
            root = ET.fromstring(xml_content)
            
            # 讀取 default 資訊
            default_remote = ''
            default_revision = ''
            default_element = root.find('default')
            if default_element is not None:
                default_remote = default_element.get('remote', '')
                default_revision = default_element.get('revision', '')
            
            # 🔥 新增調試信息
            dest_branch_count = 0

            for project in root.findall('project'):
                project_name = project.get('name', '')
                
                line_number, full_line = self._find_project_line_and_content(lines, project_name)
                
                # 確保所有欄位都用空字串，不用 N/A
                project_revision = project.get('revision', '')
                project_remote = project.get('remote', '')
                project_dest_branch = project.get('dest-branch', '')  # 🔥 重要：提取 dest-branch
                
                # 🔥 調試：統計有 dest-branch 的專案
                if project_dest_branch:
                    dest_branch_count += 1
                
                final_revision = project_revision if project_revision else (default_revision if default_revision else '')
                final_remote = project_remote if project_remote else (default_remote if default_remote else '')
                
                project_info = {
                    'line_number': line_number,
                    'name': project.get('name', ''),
                    'path': project.get('path', ''),
                    'revision': final_revision,
                    'upstream': project.get('upstream', ''),
                    'dest-branch': project.get('dest-branch', ''),
                    'groups': project.get('groups', ''),
                    'clone-depth': project.get('clone-depth', ''),
                    'remote': final_remote,
                    'full_line': full_line
                }
                projects.append(project_info)
            
            self.logger.info(f"✅ 提取了 {len(projects)} 個專案")
            return projects
            
        except Exception as e:
            self.logger.error(f"提取專案資訊失敗: {str(e)}")
            return []

    def _find_project_line_and_content(self, lines: List[str], project_name: str) -> tuple:
        """尋找專案在 XML 中的行號和完整內容"""
        line_number = 0
        full_content = ""
        
        try:
            import re
            
            for i, line in enumerate(lines, 1):
                stripped_line = line.strip()
                
                # 檢查是否包含該專案名稱
                if f'name="{project_name}"' in line:
                    line_number = i
                    
                    # 使用正規表達式只抓取 project 標籤本身
                    if stripped_line.startswith('<project') and stripped_line.endswith('/>'):
                        # 單行 project 標籤
                        full_content = stripped_line
                    elif stripped_line.startswith('<project'):
                        # 多行 project 標籤，需要找到結束位置
                        if '/>' in stripped_line:
                            # project 標籤在同一行結束
                            project_match = re.search(r'<project[^>]*/?>', stripped_line)
                            if project_match:
                                full_content = project_match.group(0)
                        else:
                            # project 標籤跨多行，找到 > 結束
                            project_content = stripped_line
                            for j in range(i, len(lines)):
                                next_line = lines[j].strip()
                                if j > i - 1:
                                    project_content += " " + next_line
                                
                                if next_line.endswith('>'):
                                    project_match = re.search(r'<project[^>]*>', project_content)
                                    if project_match:
                                        full_content = project_match.group(0)
                                    break
                    else:
                        # 如果不是以<project開始，往前找
                        for k in range(i-2, -1, -1):
                            prev_line = lines[k].strip()
                            if prev_line.startswith('<project'):
                                # 組合完整內容，然後用正規表達式提取
                                combined_content = prev_line
                                for m in range(k+1, i):
                                    combined_content += " " + lines[m].strip()
                                combined_content += " " + stripped_line
                                
                                project_match = re.search(r'<project[^>]*/?>', combined_content)
                                if project_match:
                                    full_content = project_match.group(0)
                                    line_number = k + 1
                                break
                    
                    break
            
            # 清理多餘的空格
            full_content = ' '.join(full_content.split())
            
            return line_number, full_content
            
        except Exception as e:
            self.logger.error(f"尋找專案行失敗 {project_name}: {str(e)}")
            return 0, f"<project name=\"{project_name}\" ... />"

    def _get_detailed_differences_between_projects(self, source_proj: Dict, target_proj: Dict) -> List[Dict]:
        """取得兩個專案之間的詳細差異列表"""
        differences = []
        
        try:
            # 要比較的屬性列表
            attrs_to_compare = ['name', 'path', 'revision', 'upstream', 'dest-branch', 'groups', 'clone-depth', 'remote']
            
            # 逐一比較每個屬性
            for attr in attrs_to_compare:
                source_val = source_proj.get(attr, '').strip()
                target_val = target_proj.get(attr, '').strip()
                
                # 如果不同，記錄差異
                if source_val != target_val:
                    diff_info = {
                        'attribute': attr,
                        'source_value': source_val,
                        'target_value': target_val
                    }
                    differences.append(diff_info)
            
            return differences
            
        except Exception as e:
            self.logger.error(f"取得專案間詳細差異失敗: {str(e)}")
            return []

    def _format_difference_summary(self, diff_details: List[Dict]) -> str:
        """格式化差異摘要"""
        try:
            if not diff_details:
                return "無差異"
            
            # 按屬性重要性排序
            attr_priority = {'revision': 1, 'name': 2, 'path': 3, 'upstream': 4, 'dest-branch': 5, 
                            'groups': 6, 'clone-depth': 7, 'remote': 8}
            
            diff_details.sort(key=lambda x: attr_priority.get(x['attribute'], 99))
            
            # 格式化差異說明
            diff_parts = []
            for diff in diff_details[:3]:  # 最多顯示前3個差異
                attr = diff['attribute']
                source_val = diff['source_value'] or '(空)'
                target_val = diff['target_value'] or '(空)'
                
                # 特殊處理不同屬性的顯示
                if attr == 'revision':
                    diff_parts.append(f"版本號[{source_val} ≠ {target_val}]")
                elif attr == 'upstream':
                    diff_parts.append(f"上游分支[{source_val} ≠ {target_val}]")
                elif attr == 'dest-branch':
                    diff_parts.append(f"目標分支[{source_val} ≠ {target_val}]")
                elif attr == 'groups':
                    diff_parts.append(f"群組[{source_val} ≠ {target_val}]")
                elif attr == 'clone-depth':
                    diff_parts.append(f"克隆深度[{source_val} ≠ {target_val}]")
                elif attr == 'remote':
                    diff_parts.append(f"遠端[{source_val} ≠ {target_val}]")
                else:
                    diff_parts.append(f"{attr}[{source_val} ≠ {target_val}]")
            
            # 如果差異超過3個，加上省略號
            if len(diff_details) > 3:
                diff_parts.append(f"等{len(diff_details)}項差異")
            
            return "、".join(diff_parts)
            
        except Exception as e:
            self.logger.error(f"格式化差異摘要失敗: {str(e)}")
            return "差異格式化失敗"

    def _generate_source_link(self, project_name: str, revision: str, remote: str = '') -> str:
        """根據專案名稱、revision 和 remote 生成 gerrit source link"""
        try:
            if not project_name or not revision:
                return ''
            
            # 根據 remote 決定 base URL
            if remote == 'rtk-prebuilt':
                base_url = "https://mm2sd-git2.rtkbf.com/gerrit/plugins/gitiles"
            else:  # rtk 或空值
                base_url = "https://mm2sd.rtkbf.com/gerrit/plugins/gitiles"
            
            # 檢查 revision 是否為 hash (40 字符的十六進制)
            if len(revision) == 40 and all(c in '0123456789abcdefABCDEF' for c in revision):
                # Hash 格式
                return f"{base_url}/{project_name}/+/{revision}"
            
            # 檢查是否為 tag 格式
            elif revision.startswith('refs/tags/'):
                return f"{base_url}/{project_name}/+/{revision}"
            
            # 檢查是否為完整的 branch 路徑
            elif revision.startswith('refs/heads/'):
                return f"{base_url}/{project_name}/+/{revision}"
            
            # 其他情況假設為 branch name，加上 refs/heads/ 前綴
            else:
                return f"{base_url}/{project_name}/+/refs/heads/{revision}"
                
        except Exception as e:
            self.logger.error(f"生成 source link 失敗: {str(e)}")
            return ''

    # ===============================
    # ===== Excel 報告生成方法 =====
    # ===============================

    def _generate_excel_report(self, comparison_type: str, source_file_path: str, target_file_path: str,
                              diff_analysis: Dict, output_folder: str, excel_filename: str,
                              source_download_success: bool, target_download_success: bool,
                              expanded_file_path: Optional[str] = None, use_expanded: bool = False) -> bool:
        """生成 Excel 報告 - 所有格式化在 ExcelWriter context 內完成"""
        try:
            excel_file = os.path.join(output_folder, excel_filename)
            
            # 判斷比較模式
            is_local_comparison = (comparison_type == "local_vs_local")
            
            self.logger.info(f"📋 開始生成 Excel 報告:")
            self.logger.info(f"   比較類型: {comparison_type}")
            self.logger.info(f"   是否本地比較: {is_local_comparison}")
            self.logger.info(f"   輸出檔案: {excel_file}")
            
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                # 頁籤 1: 比較摘要
                self._create_summary_sheet(writer, comparison_type, source_file_path, target_file_path, 
                                         diff_analysis, source_download_success, target_download_success,
                                         expanded_file_path, use_expanded)
                
                # 頁籤 2: 比較後專案
                self._create_comparison_projects_sheet(writer, diff_analysis, is_local_comparison)
                
                # 頁籤 3: 差異明細（包含所有專案的比較結果）
                if diff_analysis['differences']:
                    self._create_differences_sheet(writer, diff_analysis, is_local_comparison)
                else:
                    self.logger.info("沒有專案比較結果，跳過差異明細頁籤")
                
                # 頁籤 4: 來源檔案 manifest
                self._create_raw_manifest_sheet(writer, source_file_path, "來源檔案 manifest")
                
                # 頁籤 5: 目標檔案 manifest
                self._create_raw_manifest_sheet(writer, target_file_path, "目標檔案 manifest")
                
                # 🔥 在 ExcelWriter context 內完成所有格式化
                self.logger.info("🎨 開始在 ExcelWriter context 內進行格式化")
                
                # 格式化所有工作表
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    self._format_worksheet_in_context(worksheet, sheet_name, is_local_comparison)
                
                # 應用比較模式特殊處理
                self._apply_comparison_mode_fixes_in_context(writer, is_local_comparison, 
                                                           source_file_path, target_file_path, comparison_type)
            
            self.logger.info(f"✅ Excel 報告生成成功: {excel_file}")
            return True
            
        except Exception as e:
            self.logger.error(f"生成 Excel 報告失敗: {str(e)}")
            import traceback
            self.logger.error(f"詳細錯誤: {traceback.format_exc()}")
            return False

    def _create_summary_sheet(self, writer, comparison_type: str, source_file_path: str, target_file_path: str,
                             diff_analysis: Dict, source_download_success: bool, target_download_success: bool,
                             expanded_file_path: Optional[str], use_expanded: bool):
        """創建比較摘要頁籤"""
        try:
            summary = diff_analysis['summary']
            
            # 根據比較類型設定描述
            if comparison_type == "local_vs_local":
                type_description = "本地檔案比較"
                source_description = os.path.basename(source_file_path)
                target_description = os.path.basename(target_file_path)
                download_status = "N/A (本地檔案)"
                include_status = "否"
                expanded_status = "否"
            else:
                type_description = comparison_type.replace("local_vs_", "本地檔案與 ").upper() + " 比較"
                source_description = os.path.basename(source_file_path)
                target_description = os.path.basename(target_file_path)
                download_status = "成功" if target_download_success else "失敗"
                include_status = "是" if expanded_file_path else "否"
                expanded_status = "是" if use_expanded else "否"
            
            summary_data = [{
                'SN': 1,
                '比較類型': type_description,
                '來源檔案': source_description,
                '目標檔案': target_description,
                '目標檔案下載狀態': download_status,
                '目標檔案包含 include': include_status,
                '使用展開檔案': expanded_status,
                '📊 來源檔案專案數': summary['source_count'],
                '🎯 目標檔案專案數': summary['target_count'],
                '📋 總比較專案數': summary['total_compared'],
                '❌ 差異專案數': summary['differences_count'],
                '✔️ 相同專案數': summary['identical_count'],
                '📈 相同率': f"{(summary['identical_count'] / max(summary['total_compared'], 1) * 100):.1f}%" if summary['total_compared'] > 0 else "N/A"
            }]
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='比較摘要', index=False)
            
            # 🔥 新增：為比較摘要頁籤添加背景色和超連結
            self._format_summary_sheet_in_context(writer.sheets['比較摘要'], comparison_type, target_file_path)
            
        except Exception as e:
            self.logger.error(f"創建比較摘要頁籤失敗: {str(e)}")

    def _create_comparison_projects_sheet(self, writer, diff_analysis: Dict, is_local_comparison: bool):
        """創建比較後專案頁籤"""
        try:
            if not diff_analysis['project_info_list']:
                return
            
            df_projects = pd.DataFrame(diff_analysis['project_info_list'])
            
            self.logger.info(f"🔍 比較後專案原始欄位: {list(df_projects.columns)}")
            
            # 🔥 新增：為每一行添加 type 欄位（基於 Dest-Branch，空值則用 revision）
            type_values = []
            for _, row in df_projects.iterrows():
                dest_branch = row.get('dest-branch', '')
                # 🔥 新增：如果 dest-branch 為空，則使用 source_revision
                revision = row.get('source_revision', '') if not dest_branch else ''
                type_value = self._determine_dest_branch_type(dest_branch, revision)
                type_values.append(type_value)
                
            # 🔥 重要：先添加 type 欄位到 DataFrame
            df_projects['type'] = type_values
            
            # 重新命名欄位以符合比較模式
            column_mapping = {
                'SN': 'SN',
                'name': '專案名稱',
                'path': '專案路徑',
                'source_revision': '來源 Revision',
                'target_revision': '目標 Revision',
                'revision_equal': 'Revision 是否相等',
                'type': 'type',  # 🔥 重要：保持 type 欄位
                'description': '比較說明',
                'upstream': 'Upstream',
                'dest-branch': 'Dest-Branch',
                'groups': 'Groups',
                'clone-depth': 'Clone-Depth',
                'remote': 'Remote'
            }
            
            df_projects = df_projects.rename(columns=column_mapping)
            
            # 🔥 在 "Revision 是否相等" 右邊安排 "type" 欄位的順序
            # 定義最終欄位順序
            final_column_order = [
                'SN', '專案名稱', '專案路徑', 
                '來源 Revision', '目標 Revision', 'Revision 是否相等', 'type',  # type 在 Revision 是否相等 右邊
                '比較說明', 'Upstream', 'Dest-Branch', 'Groups', 'Clone-Depth', 'Remote'
            ]
            
            # 只保留存在的欄位
            available_columns = [col for col in final_column_order if col in df_projects.columns]
            df_projects = df_projects[available_columns]
            
            self.logger.info(f"✅ 比較後專案最終欄位: {list(df_projects.columns)}")
            
            df_projects.to_excel(writer, sheet_name='比較後專案', index=False)
            
            self.logger.info(f"✅ 比較後專案頁籤已創建，包含 type 欄位: {len(df_projects)} 個專案")
            
        except Exception as e:
            self.logger.error(f"創建比較後專案頁籤失敗: {str(e)}")
            import traceback
            self.logger.error(f"詳細錯誤: {traceback.format_exc()}")

    def _create_differences_sheet(self, writer, diff_analysis: Dict, is_local_comparison: bool):
        """創建差異明細頁籤（顯示所有專案的比較結果）"""
        try:
            if not diff_analysis['differences']:
                return
            
            df_differences = pd.DataFrame(diff_analysis['differences'])
            
            # 🔥 新增：為每一行添加 type 欄位（基於來源檔案的 dest-branch，空值則用 revision）
            type_values = []
            for _, row in df_differences.iterrows():
                # 優先使用來源檔案的 dest-branch，如果沒有則使用目標檔案的
                dest_branch = ''
                revision = ''
                
                # 從 differences 數據中獲取 dest-branch
                if 'dest-branch' in row and pd.notna(row['dest-branch']):
                    dest_branch = str(row['dest-branch'])
                elif is_local_comparison and 'compare_dest-branch' in row and pd.notna(row['compare_dest-branch']):
                    dest_branch = str(row['compare_dest-branch'])
                elif not is_local_comparison and 'gerrit_dest-branch' in row and pd.notna(row['gerrit_dest-branch']):
                    dest_branch = str(row['gerrit_dest-branch'])
                
                # 🔥 新增：如果 dest-branch 為空，則使用 revision
                if not dest_branch:
                    if 'revision' in row and pd.notna(row['revision']):
                        revision = str(row['revision'])
                    elif is_local_comparison and 'compare_revision' in row and pd.notna(row['compare_revision']):
                        revision = str(row['compare_revision'])
                    elif not is_local_comparison and 'gerrit_revision' in row and pd.notna(row['gerrit_revision']):
                        revision = str(row['gerrit_revision'])
                
                type_value = self._determine_dest_branch_type(dest_branch, revision)
                type_values.append(type_value)
            
            # 根據比較模式調整欄位順序
            if is_local_comparison:
                # 本地比較模式的欄位順序
                column_order = [
                    'SN', 'comparison_status', 'comparison_result', 'type',  # 🔥 在 comparison_result 後插入 type
                    'source_file', 'content', 'name', 'path', 'revision',
                    'upstream', 'dest-branch', 'groups', 'clone-depth', 'remote', 'source_link',
                    'compare_source_file', 'compare_content', 'compare_name', 'compare_path', 'compare_revision',
                    'compare_upstream', 'compare_dest-branch', 'compare_groups', 'compare_clone-depth', 'compare_remote', 'compare_source_link'
                ]
            else:
                # Gerrit 比較模式的欄位順序
                column_order = [
                    'SN', 'comparison_status', 'comparison_result', 'type',  # 🔥 在 comparison_result 後插入 type
                    'source_file', 'content', 'name', 'path', 'revision',
                    'upstream', 'dest-branch', 'groups', 'clone-depth', 'remote', 'source_link',
                    'gerrit_source_file', 'gerrit_content', 'gerrit_name', 'gerrit_path', 'gerrit_revision',
                    'gerrit_upstream', 'gerrit_dest-branch', 'gerrit_groups', 'gerrit_clone-depth', 'gerrit_remote', 'gerrit_source_link'
                ]
            
            # 添加 type 欄位到 DataFrame
            df_differences['type'] = type_values
            
            # 只保留存在的欄位
            available_columns = [col for col in column_order if col in df_differences.columns]
            df_differences = df_differences[available_columns]
            
            df_differences.to_excel(writer, sheet_name='比較專案內容差異明細', index=False)
            
            self.logger.info(f"✅ 差異明細頁籤已創建，包含 type 欄位: {len(df_differences)} 個專案")
            
        except Exception as e:
            self.logger.error(f"創建差異明細頁籤失敗: {str(e)}")
            import traceback
            self.logger.error(f"詳細錯誤: {traceback.format_exc()}")

    def _create_raw_manifest_sheet(self, writer, file_path: str, sheet_name: str):
        """創建原始 manifest 頁籤"""
        try:
            if not file_path or not os.path.exists(file_path):
                self.logger.warning(f"檔案不存在，跳過 {sheet_name}: {file_path}")
                return
            
            # 直接從檔案重新解析
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            projects = self._extract_projects_with_line_numbers(content)
            
            raw_data = []
            for i, proj in enumerate(projects, 1):
                source_link = self._generate_source_link(proj['name'], proj['revision'], proj['remote'])
                
                # 🔥 新增：判斷 type（基於 dest-branch，空值則用 revision）
                type_value = self._determine_dest_branch_type(proj['dest-branch'], proj['revision'])
                
                raw_data.append({
                    'SN': i,
                    'source_file': os.path.basename(file_path),
                    'name': proj['name'],
                    'path': proj['path'],
                    'revision': proj['revision'],
                    'type': type_value,  # 🔥 在 revision 右邊插入 type
                    'upstream': proj['upstream'],
                    'dest-branch': proj['dest-branch'],
                    'groups': proj['groups'],
                    'clone-depth': proj['clone-depth'],
                    'remote': proj['remote'],
                    'source_link': source_link
                })
            
            if raw_data:
                df_raw = pd.DataFrame(raw_data)
                
                # 🔥 確保欄位順序正確（type 在 revision 右邊）
                column_order = [
                    'SN', 'source_file', 'name', 'path', 'revision', 'type',  # type 在 revision 右邊
                    'upstream', 'dest-branch', 'groups', 'clone-depth', 'remote', 'source_link'
                ]
                
                # 只保留存在的欄位
                available_columns = [col for col in column_order if col in df_raw.columns]
                df_raw = df_raw[available_columns]
                
                df_raw.to_excel(writer, sheet_name=sheet_name, index=False)
                self.logger.info(f"✅ {sheet_name} 頁籤已創建，包含 type 欄位: {len(raw_data)} 個專案")
            
        except Exception as e:
            self.logger.error(f"創建 {sheet_name} 頁籤失敗: {str(e)}")
            import traceback
            self.logger.error(f"詳細錯誤: {traceback.format_exc()}")

    def _format_worksheet_in_context(self, worksheet, sheet_name: str, is_local_comparison: bool):
        """在 ExcelWriter context 內格式化工作表"""
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
            
            # 設定 Excel 頁籤標籤顏色
            if sheet_name in ['比較摘要']:
                worksheet.sheet_properties.tabColor = "ADD8E6"  # Light Blue
            elif '差異明細' in sheet_name:
                worksheet.sheet_properties.tabColor = "FFB6C1"  # Light Pink
            else:
                worksheet.sheet_properties.tabColor = "90EE90"  # Light Green
            
            # 顏色定義
            blue_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            red_fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
            orange_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
            purple_fill = PatternFill(start_color="8A2BE2", end_color="8A2BE2", fill_type="solid")
            # 🔥 重點修正：使用正確的藍色 RGB(0, 112, 192)
            link_blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            # 🔥 新增：type 欄位表頭的綠色
            type_header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            
            white_font = Font(color="FFFFFF", bold=True)
            
            # 🔥 定義 link 欄位（這些欄位必須使用新藍色）
            link_fields = ["source_link", "gerrit_source_link", "compare_source_link"]
            
            # 🔥 儲存 type 欄位的位置，稍後設定內容顏色
            type_column = None
            
            # 設定表頭格式和欄寬
            for col_num, cell in enumerate(worksheet[1], 1):
                col_letter = get_column_letter(col_num)
                header_value = str(cell.value) if cell.value else ''
                
                # 🔥 第一優先級：type 欄位設定為綠底白字
                if header_value == 'type':
                    cell.fill = type_header_fill
                    cell.font = white_font
                    worksheet.column_dimensions[col_letter].width = 15
                    type_column = col_num  # 記錄 type 欄位位置
                    self.logger.info(f"🟢 設定 type 欄位表頭: 綠底白字 ({col_letter}欄)")
                
                # 🔥 第二優先級：link 欄位設定為新藍色
                elif header_value in link_fields:
                    cell.fill = link_blue_fill
                    cell.font = white_font
                    worksheet.column_dimensions[col_letter].width = 60
                    self.logger.info(f"🔵 設定新藍色: '{header_value}' ({col_letter}欄)")
                
                # revision 相關欄位
                elif 'revision' in header_value.lower() or header_value == 'Revision 是否相等':
                    cell.fill = red_fill
                    cell.font = white_font
                    if header_value == 'Revision 是否相等':
                        worksheet.column_dimensions[col_letter].width = 20
                    else:
                        worksheet.column_dimensions[col_letter].width = 40
                
                # 檔案相關欄位
                elif header_value in ['source_file', 'gerrit_source_file', 'compare_source_file', '來源檔案', '目標檔案']:
                    cell.fill = purple_fill
                    cell.font = white_font
                    worksheet.column_dimensions[col_letter].width = 25
                
                # 比較狀態相關欄位
                elif header_value in ['comparison_status', 'comparison_result', '比較說明']:
                    cell.fill = orange_fill
                    cell.font = white_font
                    if header_value == 'comparison_result' or header_value == '比較說明':
                        worksheet.column_dimensions[col_letter].width = 50
                    else:
                        worksheet.column_dimensions[col_letter].width = 20
                
                # Gerrit 相關欄位（非 link 的）
                elif header_value.startswith('gerrit_') and header_value not in link_fields:
                    cell.fill = green_fill
                    cell.font = white_font
                
                # Compare 相關欄位（本地比較，非 link 的）
                elif header_value.startswith('compare_') and header_value not in link_fields:
                    cell.fill = purple_fill if is_local_comparison else green_fill
                    cell.font = white_font
                
                # 其他欄位
                else:
                    cell.fill = blue_header_fill
                    cell.font = white_font
                
                # 設定對齊
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 設定其他欄位的寬度
                if header_value not in link_fields and 'revision' not in header_value.lower() and header_value != 'type':
                    if header_value == 'SN':
                        worksheet.column_dimensions[col_letter].width = 8
                    elif header_value in ['專案名稱', 'name', 'gerrit_name', 'compare_name']:
                        worksheet.column_dimensions[col_letter].width = 30
                    elif header_value in ['專案路徑', 'path', 'gerrit_path', 'compare_path']:
                        worksheet.column_dimensions[col_letter].width = 35
                    elif 'content' in header_value:
                        worksheet.column_dimensions[col_letter].width = 80
                    elif header_value in ['Groups', 'groups', 'gerrit_groups', 'compare_groups']:
                        worksheet.column_dimensions[col_letter].width = 40
                    elif header_value in ['Upstream', 'upstream', 'gerrit_upstream', 'compare_upstream',
                                        'Dest-Branch', 'dest-branch', 'gerrit_dest-branch', 'compare_dest-branch']:
                        worksheet.column_dimensions[col_letter].width = 25
                    elif header_value in ['Clone-Depth', 'clone-depth', 'gerrit_clone-depth', 'compare_clone-depth']:
                        worksheet.column_dimensions[col_letter].width = 15
                    elif header_value in ['Remote', 'remote', 'gerrit_remote', 'compare_remote']:
                        worksheet.column_dimensions[col_letter].width = 15
                    else:
                        worksheet.column_dimensions[col_letter].width = 20
            
            # 🔥 新增：為 type 欄位的內容設定顏色
            if type_column:
                self._set_type_column_colors_in_context(worksheet, type_column)
            
            # 🔥 新增：為所有 SN 欄位的內容設定置中對齊
            self._set_sn_column_center_alignment(worksheet)
            
            self.logger.info(f"✅ {sheet_name} 格式化完成")
            
        except Exception as e:
            self.logger.error(f"格式化工作表失敗 {sheet_name}: {str(e)}")

    def _set_type_column_colors_in_context(self, worksheet, type_column: int):
        """在 ExcelWriter context 內為 type 欄位的內容設定顏色"""
        try:
            from openpyxl.styles import Alignment
            
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # 為 type 欄位的所有內容設定顏色和對齊
            for row_num in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_num, column=type_column)
                type_value = str(cell.value) if cell.value else ''
                
                # 設定置中對齊
                cell.alignment = center_alignment
                
                # 根據 type 值設定顏色（只設定文字顏色）
                if type_value:
                    fill, font = self._get_type_color_style(type_value)
                    # 🔥 修正：只設定字體，不設定背景色
                    if font:
                        cell.font = font
            
            self.logger.info(f"✅ type 欄位內容文字顏色已設定完成 ({worksheet.max_row - 1} 行)")
            
        except Exception as e:
            self.logger.error(f"設定 type 欄位內容顏色失敗: {str(e)}")
            
    def _set_sn_column_center_alignment(self, worksheet):
        """為 SN 欄位的所有內容設定置中對齊"""
        try:
            from openpyxl.styles import Alignment
            
            # 找到 SN 欄位
            sn_col = None
            for col_num, cell in enumerate(worksheet[1], 1):
                header_value = str(cell.value) if cell.value else ''
                if header_value == 'SN':
                    sn_col = col_num
                    break
            
            if sn_col:
                center_alignment = Alignment(horizontal='center', vertical='center')
                
                # 為 SN 欄位的所有內容設定置中對齊
                for row_num in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_num, column=sn_col)
                    cell.alignment = center_alignment
                
                self.logger.debug(f"✅ SN 欄位內容已設定置中對齊 ({worksheet.max_row - 1} 行)")
            
        except Exception as e:
            self.logger.error(f"設定 SN 欄位置中對齊失敗: {str(e)}")

    def _apply_comparison_mode_fixes_in_context(self, writer, is_local_comparison: bool, 
                                              source_file_path: str, target_file_path: str, 
                                              comparison_type: str):
        """在 ExcelWriter context 內應用比較模式的修正"""
        try:
            self.logger.info("🔧 在 ExcelWriter context 內應用比較模式修正")
            
            # 為比較後專案頁籤添加 Excel 公式
            if '比較後專案' in writer.sheets:
                self._add_revision_comparison_formula_in_context(writer.sheets['比較後專案'])
            
            # 為差異明細頁籤設定行背景色
            if '比較專案內容差異明細' in writer.sheets:
                self._set_comparison_row_colors_in_context(writer.sheets['比較專案內容差異明細'])
            
            # 為所有頁籤添加超連結
            for sheet_name, worksheet in writer.sheets.items():
                self._add_hyperlinks_in_context(worksheet, sheet_name)
            
            self.logger.info("✅ 比較模式修正在 ExcelWriter context 內完成")
            
        except Exception as e:
            self.logger.error(f"在 ExcelWriter context 內應用比較模式修正失敗: {str(e)}")

    def _add_revision_comparison_formula_in_context(self, worksheet):
        """在 ExcelWriter context 內為比較後專案頁籤添加 Revision 比較公式"""
        try:
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, Alignment
            from openpyxl.formatting.rule import FormulaRule
            
            # 找到相關欄位的位置
            source_revision_col = None
            target_revision_col = None
            comparison_col = None
            
            for col_num, cell in enumerate(worksheet[1], 1):
                header_value = str(cell.value) if cell.value else ''
                if header_value == '來源 Revision':
                    source_revision_col = col_num
                elif header_value == '目標 Revision':
                    target_revision_col = col_num
                elif header_value == 'Revision 是否相等':
                    comparison_col = col_num
            
            if not all([source_revision_col, target_revision_col, comparison_col]):
                self.logger.warning("無法找到所需的 Revision 欄位")
                return
            
            # 取得欄位字母
            source_col_letter = get_column_letter(source_revision_col)
            target_col_letter = get_column_letter(target_revision_col)
            comparison_col_letter = get_column_letter(comparison_col)
            
            # 🔥 新增：定義置中對齊
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # 添加 Excel 公式並設定置中對齊
            for row_num in range(2, worksheet.max_row + 1):
                formula = f'=IF({source_col_letter}{row_num}={target_col_letter}{row_num},"Y","N")'
                cell = worksheet[f"{comparison_col_letter}{row_num}"]
                cell.value = formula
                # 🔥 重要：設定置中對齊
                cell.alignment = center_alignment
            
            # 設定條件格式
            green_font = Font(color="00B050", bold=True)
            red_font = Font(color="FF0000", bold=True)
            
            # 條件格式範圍
            range_string = f"{comparison_col_letter}2:{comparison_col_letter}{worksheet.max_row}"
            
            # 為 "Y" 值設定綠色字體
            green_rule = FormulaRule(
                formula=[f'${comparison_col_letter}2="Y"'],
                font=green_font
            )
            
            # 為 "N" 值設定紅色字體
            red_rule = FormulaRule(
                formula=[f'${comparison_col_letter}2="N"'],
                font=red_font
            )
            
            # 添加條件格式規則
            worksheet.conditional_formatting.add(range_string, green_rule)
            worksheet.conditional_formatting.add(range_string, red_rule)
            
            self.logger.info("✅ 已在 context 內添加 Revision 比較公式和條件格式（含置中對齊）")
            
        except Exception as e:
            self.logger.error(f"在 context 內添加 Revision 比較公式失敗: {str(e)}")

    def _set_comparison_row_colors_in_context(self, worksheet):
        """在 ExcelWriter context 內設定比較狀態的行背景色"""
        try:
            from openpyxl.styles import PatternFill
            
            # 找到 comparison_status 欄位
            status_col = None
            for col_num, cell in enumerate(worksheet[1], 1):
                header_value = str(cell.value) if cell.value else ''
                if header_value == 'comparison_status':
                    status_col = col_num
                    break
            
            if not status_col:
                self.logger.debug("找不到 comparison_status 欄位，跳過行顏色設定")
                return
            
            # 狀態顏色配置
            status_colors = {
                '✔️ 相同': PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid"),      # 淺綠底
                '❌ 不同': PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid"),      # 淺紅底
                '➕ 新增': PatternFill(start_color="E8F0FF", end_color="E8F0FF", fill_type="solid"),      # 淺藍底
                '❓ 無此專案': PatternFill(start_color="FFE8CC", end_color="FFE8CC", fill_type="solid"),  # 淺橘底
            }
            
            # 設定每一行的背景色
            applied_count = 0
            for row_num in range(2, worksheet.max_row + 1):
                status_cell = worksheet.cell(row=row_num, column=status_col)
                status_value = str(status_cell.value) if status_cell.value else ''
                
                for status_pattern, fill_color in status_colors.items():
                    if status_pattern in status_value:
                        # 設定整行的背景色
                        for col in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row_num, column=col).fill = fill_color
                        applied_count += 1
                        break
            
            self.logger.info(f"✅ 已在 context 內設定 {applied_count} 行的背景色")
            
        except Exception as e:
            self.logger.error(f"在 context 內設定比較狀態行顏色失敗: {str(e)}")

    def _add_hyperlinks_in_context(self, worksheet, sheet_name: str):
        """在 ExcelWriter context 內為頁籤添加超連結"""
        try:
            # 處理 source_link 相關欄位
            link_columns = []
            
            # 找到所有 link 欄位
            for col_num, cell in enumerate(worksheet[1], 1):
                header_value = str(cell.value) if cell.value else ''
                if header_value in ['source_link', 'gerrit_source_link', 'compare_source_link']:
                    link_columns.append((col_num, header_value))
            
            # 為每個 link 欄位添加超連結
            for col_num, header_value in link_columns:
                for row_num in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    url = str(cell.value) if cell.value else ''
                    
                    if url and url != '' and 'http' in url:
                        self._add_hyperlink_formula_to_cell_in_context(worksheet, row_num, col_num, url)
            
            # 為檔案名稱欄位添加 Gerrit 連結
            for col_num, cell in enumerate(worksheet[1], 1):
                header_value = str(cell.value) if cell.value else ''
                if header_value == 'source_file' and sheet_name in ['來源檔案 manifest', '目標檔案 manifest']:
                    for row_num in range(2, worksheet.max_row + 1):
                        filename_cell = worksheet.cell(row=row_num, column=col_num)
                        filename = str(filename_cell.value) if filename_cell.value else ''
                        
                        if filename and filename not in ['', 'N/A']:
                            gerrit_url = self._generate_gerrit_manifest_link(filename)
                            self._add_hyperlink_formula_to_cell_in_context(worksheet, row_num, col_num, gerrit_url, filename)
            
            if link_columns:
                self.logger.info(f"✅ 已在 context 內為 {sheet_name} 添加 {len(link_columns)} 種類型的超連結")
            
        except Exception as e:
            self.logger.error(f"在 context 內添加 {sheet_name} 超連結失敗: {str(e)}")

    def _add_hyperlink_formula_to_cell_in_context(self, worksheet, row: int, col: int, url: str, display_text: str = None):
        """在 ExcelWriter context 內為單元格添加 HYPERLINK 函數格式的超連結"""
        try:
            from openpyxl.styles import Font
            
            cell = worksheet.cell(row=row, column=col)
            
            if not url or url == 'N/A':
                return
            
            # 清理 URL 中的特殊字符
            clean_url = str(url).replace('"', '""')
            
            if not display_text or display_text == url:
                cell.value = f'=HYPERLINK("{clean_url}")'
            else:
                clean_display_text = str(display_text).replace('"', '""')
                cell.value = f'=HYPERLINK("{clean_url}","{clean_display_text}")'
            
            # 設定藍色超連結樣式
            cell.font = Font(color="0000FF", underline="single")
            
        except Exception as e:
            self.logger.error(f"在 context 內添加 HYPERLINK 函數失敗: {str(e)}")

    def _generate_gerrit_manifest_link(self, filename: str) -> str:
        """生成 Gerrit manifest 檔案的連結"""
        try:
            if not filename or filename == '無':
                return '無'
            
            # 移除 gerrit_ 前綴（如果有的話）
            clean_filename = filename.replace('gerrit_', '') if filename.startswith('gerrit_') else filename
            
            # 構建 Gerrit 連結
            gerrit_link = f"{self.gerrit_base_url}/{clean_filename}"
            
            return gerrit_link
            
        except Exception as e:
            self.logger.error(f"生成 Gerrit 連結失敗: {str(e)}")
            return filename

    # ===============================
    # ===== 輔助方法 =====
    # ===============================

    def _copy_local_file_to_output(self, local_file: str, output_folder: str, 
                                custom_name: Optional[str] = None) -> str:
        """複製本地檔案到輸出目錄"""
        try:
            if custom_name:
                dest_name = custom_name
            else:
                dest_name = os.path.basename(local_file)
            
            dest_path = os.path.join(output_folder, dest_name)
            shutil.copy2(local_file, dest_path)
            
            self.logger.info(f"✅ 複製本地檔案: {dest_name}")
            return dest_path
            
        except Exception as e:
            self.logger.error(f"複製本地檔案失敗: {str(e)}")
            raise

    def _download_gerrit_file(self, gerrit_type: str, output_folder: str) -> Optional[str]:
        """從 Gerrit 下載檔案到輸出目錄"""
        try:
            if gerrit_type not in self.gerrit_urls:
                self.logger.error(f"不支援的 Gerrit 類型: {gerrit_type}")
                return None
            
            config = self.gerrit_urls[gerrit_type]
            gerrit_filename = f"gerrit_{config['filename']}"
            gerrit_path = os.path.join(output_folder, gerrit_filename)
            
            self.logger.info(f"下載 {gerrit_type.upper()} manifest: {config['filename']}")
            self.logger.info(f"URL: {config['url']}")
            self.logger.info(f"保存為: {gerrit_filename}")
            
            # 使用 gerrit_manager 下載檔案
            success = self.gerrit_manager.download_file_from_link(config['url'], gerrit_path)
            
            if success and os.path.exists(gerrit_path):
                file_size = os.path.getsize(gerrit_path)
                self.logger.info(f"✅ 成功下載 {gerrit_type.upper()} manifest: {file_size} bytes")
                return gerrit_path
            else:
                self.logger.error(f"❌ 下載 {gerrit_type.upper()} manifest 失敗")
                return None
                
        except Exception as e:
            self.logger.error(f"下載 Gerrit 檔案異常: {str(e)}")
            return None

    def _handle_gerrit_include_expansion(self, gerrit_file_path: str, output_folder: str) -> str:
        """處理 Gerrit manifest 的 include 展開 - 使用 repo 命令"""
        try:
            self.logger.info("🔍 檢查 Gerrit manifest 是否需要展開")
            
            # 讀取 Gerrit 檔案內容
            with open(gerrit_file_path, 'r', encoding='utf-8') as f:
                gerrit_content = f.read()
            
            # 檢查 include 標籤
            if not self._has_include_tags(gerrit_content):
                self.logger.info("ℹ️ 未檢測到 include 標籤，使用原始檔案")
                return gerrit_file_path
            
            self.logger.info("🔍 檢測到 include 標籤，開始使用 repo 命令展開 manifest...")
            
            # 🔥 使用 repo 命令展開
            expanded_content, expanded_file_path = self._expand_manifest_with_repo(gerrit_file_path, output_folder)
            
            if expanded_content and expanded_file_path:
                # 設定展開檔案路徑
                self.expanded_file_path = expanded_file_path
                self.use_expanded = True
                
                self.logger.info(f"✅ 成功使用 repo 命令展開 manifest: {os.path.basename(expanded_file_path)}")
                return expanded_file_path
            else:
                self.logger.warning("⚠️ repo 命令展開失敗，使用原始檔案")
                return gerrit_file_path
                
        except Exception as e:
            self.logger.error(f"處理 include 展開時發生錯誤: {str(e)}")
            import traceback
            self.logger.error(f"詳細錯誤: {traceback.format_exc()}")
            self.logger.warning("⚠️ 使用原始檔案繼續執行")
            return gerrit_file_path

    def _expand_manifest_with_repo(self, gerrit_file_path: str, output_folder: str) -> tuple:
        """使用 repo 命令展開包含 include 的 manifest"""
        import subprocess
        import tempfile
        import shutil
        
        try:
            # 從檔案路徑取得檔案名
            source_filename = os.path.basename(gerrit_file_path).replace('gerrit_', '')
            repo_url = "ssh://mm2sd.rtkbf.com:29418/realtek/android/manifest"
            # 🔥 使用動態分支
            branch = config.get_default_android_master_branch()
            
            # 生成展開檔案名稱
            expanded_filename = f"gerrit_{source_filename.replace('.xml', '_expanded.xml')}"
            final_expanded_path = os.path.abspath(os.path.join(output_folder, expanded_filename))
            
            self.logger.info(f"🎯 準備使用 repo 命令展開 manifest...")
            self.logger.info(f"🎯 源檔案: {source_filename}")
            self.logger.info(f"🎯 使用分支: {branch}")
            self.logger.info(f"🎯 展開檔案名: {expanded_filename}")
            self.logger.info(f"🎯 目標路徑: {final_expanded_path}")
            
            # 確保輸出資料夾存在
            utils.ensure_dir(output_folder)
            
            # 檢查 repo 命令是否可用
            try:
                repo_check = subprocess.run(
                    ["repo", "--version"], 
                    capture_output=True, 
                    text=True, 
                    timeout=10
                )
                if repo_check.returncode == 0:
                    self.logger.info(f"✅ repo 工具可用: {repo_check.stdout.strip()}")
                else:
                    self.logger.error(f"❌ repo 工具檢查失敗: {repo_check.stderr}")
                    return None, None
            except FileNotFoundError:
                self.logger.error("❌ repo 命令未找到，請確認已安裝 repo 工具")
                return None, None
            except Exception as e:
                self.logger.error(f"❌ repo 工具檢查異常: {str(e)}")
                return None, None
            
            # 建立臨時工作目錄
            temp_work_dir = tempfile.mkdtemp(prefix='repo_expand_manifest_')
            self.logger.info(f"📁 建立臨時工作目錄: {temp_work_dir}")
            
            original_cwd = os.getcwd()
            
            try:
                # 切換到臨時目錄
                os.chdir(temp_work_dir)
                self.logger.info(f"📂 切換到臨時目錄: {temp_work_dir}")
                
                # 步驟 1: repo init
                self.logger.info(f"📄 執行 repo init...")
                init_cmd = [
                    "repo", "init", 
                    "-u", repo_url,
                    "-b", branch,
                    "-m", source_filename
                ]
                
                self.logger.info(f"🎯 Init 指令: {' '.join(init_cmd)}")
                
                init_result = subprocess.run(
                    init_cmd,
                    capture_output=True,
                    text=True,
                    timeout=120
                )
                
                self.logger.info(f"🔍 repo init 返回碼: {init_result.returncode}")
                if init_result.stdout:
                    self.logger.info(f"🔍 repo init stdout: {init_result.stdout}")
                if init_result.stderr:
                    self.logger.info(f"🔍 repo init stderr: {init_result.stderr}")
                
                if init_result.returncode != 0:
                    self.logger.error(f"❌ repo init 失敗 (返回碼: {init_result.returncode})")
                    return None, None
                
                self.logger.info("✅ repo init 成功")
                
                # 檢查 .repo 目錄
                repo_dir = os.path.join(temp_work_dir, ".repo")
                if not os.path.exists(repo_dir):
                    self.logger.error(f"❌ .repo 目錄不存在: {repo_dir}")
                    return None, None
                
                # 步驟 2: repo manifest 展開
                self.logger.info(f"📄 執行 repo manifest 展開...")
                
                manifest_cmd = ["repo", "manifest"]
                self.logger.info(f"🎯 Manifest 指令: {' '.join(manifest_cmd)}")
                
                manifest_result = subprocess.run(
                    manifest_cmd,
                    capture_output=True,
                    text=True,
                    timeout=60
                )
                
                self.logger.info(f"🔍 repo manifest 返回碼: {manifest_result.returncode}")
                if manifest_result.stderr:
                    self.logger.info(f"🔍 repo manifest stderr: {manifest_result.stderr}")
                
                if manifest_result.returncode != 0:
                    self.logger.error(f"❌ repo manifest 失敗 (返回碼: {manifest_result.returncode})")
                    return None, None
                
                expanded_content = manifest_result.stdout
                
                if not expanded_content.strip():
                    self.logger.error("❌ repo manifest 返回空內容")
                    return None, None
                
                # 分析展開內容
                project_count = expanded_content.count('<project ')
                include_count = expanded_content.count('<include ')
                self.logger.info(f"✅ repo manifest 成功:")
                self.logger.info(f"   📊 內容長度: {len(expanded_content)} 字符")
                self.logger.info(f"   📊 Project 數量: {project_count}")
                self.logger.info(f"   📊 Include 數量: {include_count}")
                
                # 步驟 3: 保存展開檔案
                self.logger.info(f"📁 保存展開檔案到: {final_expanded_path}")
                
                with open(final_expanded_path, 'w', encoding='utf-8') as f:
                    f.write(expanded_content)
                
                # 驗證保存結果
                if os.path.exists(final_expanded_path):
                    file_size = os.path.getsize(final_expanded_path)
                    self.logger.info(f"✅ 展開檔案保存成功:")
                    self.logger.info(f"   📁 檔案路徑: {final_expanded_path}")
                    self.logger.info(f"   📊 檔案大小: {file_size} bytes")
                    self.logger.info(f"   📊 專案數量: {project_count}")
                    
                    return expanded_content, final_expanded_path
                else:
                    self.logger.error(f"❌ 展開檔案保存失敗: {final_expanded_path}")
                    return None, None
                    
            finally:
                # 恢復原始工作目錄
                os.chdir(original_cwd)
                self.logger.info(f"📂 恢復原始工作目錄: {original_cwd}")
                
                # 清理臨時目錄
                try:
                    shutil.rmtree(temp_work_dir)
                    self.logger.info(f"🗑️ 清理臨時目錄成功: {temp_work_dir}")
                except Exception as e:
                    self.logger.warning(f"⚠️ 清理臨時目錄失敗: {str(e)}")
            
        except subprocess.TimeoutExpired:
            self.logger.error("❌ repo 命令執行超時")
            return None, None
        except Exception as e:
            self.logger.error(f"❌ 使用 repo 展開 manifest 失敗: {str(e)}")
            import traceback
            self.logger.error(f"❌ 錯誤詳情: {traceback.format_exc()}")
            return None, None
            
    def _has_include_tags(self, xml_content: str) -> bool:
        """檢查 XML 內容是否包含 include 標籤"""
        try:
            import re
            
            # 使用正規表達式檢查 include 標籤
            include_pattern = r'<include\s+name\s*=\s*["\'][^"\']*["\'][^>]*/?>'
            matches = re.findall(include_pattern, xml_content, re.IGNORECASE)
            
            if matches:
                self.logger.info(f"🔍 發現 {len(matches)} 個 include 標籤:")
                for i, match in enumerate(matches, 1):
                    self.logger.info(f"  {i}. {match}")
                return True
            else:
                self.logger.info("ℹ️ 未發現 include 標籤")
                return False
                
        except Exception as e:
            self.logger.error(f"檢查 include 標籤時發生錯誤: {str(e)}")
            return False

    def _show_comparison_results(self, comparison_type: str, diff_analysis: Dict):
        """顯示比較結果統計"""
        self.logger.info(f"\n📈 {comparison_type} 比較結果統計:")
        self.logger.info(f"  🔧 處理邏輯: 純比較版本（不執行轉換）")
        self.logger.info(f"  📋 Excel 格式: 比較模式專用版本")
        self.logger.info(f"  📄 處理模式: 純比對（無轉換頁籤）")
        self.logger.info(f"  🎨 格式化: 所有格式化在 ExcelWriter context 內完成")
        
        summary = diff_analysis.get('summary', {})
        self.logger.info(f"\n📊 統計摘要:")
        self.logger.info(f"  來源檔案專案數: {summary.get('source_count', 0)}")
        self.logger.info(f"  目標檔案專案數: {summary.get('target_count', 0)}")
        self.logger.info(f"  總比較項目數: {summary.get('total_compared', 0)}")
        self.logger.info(f"  差異項目數: {summary.get('differences_count', 0)}")
        self.logger.info(f"  相同項目數: {summary.get('identical_count', 0)}")
        
        if self.use_expanded:
            self.logger.info(f"  🔍 特殊處理: Gerrit include 標籤已自動展開")
            self.logger.info(f"  📄 展開檔案: {os.path.basename(self.expanded_file_path) if self.expanded_file_path else 'N/A'}")
        
        self.logger.info("=" * 80)

    def _determine_dest_branch_type(self, dest_branch: str, revision: str = '') -> str:
        """判斷 dest-branch 或 revision 的類型"""
        try:
            # 🔥 修改邏輯：優先看 dest-branch，如果是空值則看 revision
            value_to_check = ''
            
            if dest_branch and dest_branch.strip():
                value_to_check = dest_branch.strip()
                self.logger.debug(f"使用 dest-branch 判斷: '{value_to_check}'")
            elif revision and revision.strip():
                value_to_check = revision.strip()
                self.logger.debug(f"dest-branch 為空，使用 revision 判斷: '{value_to_check}'")
            else:
                return ''
            
            # 檢查是否為 tag 形式 (以 refs/tags/ 開頭)
            if value_to_check.startswith('refs/tags/'):
                return 'Tag'
            
            # 檢查是否為 hash 形式 (40 字符的十六進制)
            if len(value_to_check) == 40 and all(c in '0123456789abcdefABCDEF' for c in value_to_check):
                return 'Hash'
            
            # 檢查是否為 branch 形式 (包含 / 但不是 refs/tags/)
            if '/' in value_to_check and not value_to_check.startswith('refs/'):
                return 'Branch'
            
            # 檢查是否為完整的 branch 路徑 (refs/heads/)
            if value_to_check.startswith('refs/heads/'):
                return 'Branch'
            
            # 其他情況默認為 Branch
            return 'Branch'
            
        except Exception as e:
            self.logger.error(f"判斷 dest-branch/revision 類型失敗: {str(e)}")
            return ''

    def _get_type_color_style(self, type_value: str):
        """根據 type 值獲取對應的樣式"""
        try:
            from openpyxl.styles import Font
            
            # 🔥 修正：只設定文字顏色，不設定背景色
            if type_value == 'Branch':
                return None, Font(color="8A2BE2", bold=True)  # 紫色文字
            elif type_value == 'Tag':
                return None, Font(color="8B0000", bold=True)  # 深紅色文字
            elif type_value == 'Hash':
                return None, Font(color="0000FF", bold=True)  # 藍色文字
            else:
                return None, Font(color="000000", bold=False)  # 黑色文字
                
        except Exception as e:
            self.logger.error(f"獲取 type 顏色樣式失敗: {str(e)}")
            from openpyxl.styles import Font
            return None, Font(color="000000", bold=False)
            
def main():
    """主函數"""
    parser = argparse.ArgumentParser(description='Manifest 比較工具 - 純比較版本')
    parser.add_argument('local_file', help='本地 manifest.xml 檔案路徑')
    parser.add_argument('-g', '--gerrit-type', 
                       choices=['master', 'premp', 'mp', 'mp_backup'],
                       help='Gerrit 檔案類型 (與本地檔案比較)')
    parser.add_argument('-t', '--target-file', help='目標檔案路徑 (本地檔案比較)')
    parser.add_argument('-o', '--output', default='manifest_comparison_report.xlsx',
                       help='輸出 Excel 檔案名稱')
    
    args = parser.parse_args()
    
    # 確保輸出目錄存在
    output_dir = os.path.dirname(args.output) or '.'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # 執行比較
    comparator = ManifestComparator()
    
    if args.gerrit_type:
        # 本地檔案與 Gerrit 比較
        success = comparator.compare_local_with_gerrit(args.local_file, args.gerrit_type, args.output)
        comparison_desc = f"本地檔案與 {args.gerrit_type.upper()}"
    elif args.target_file:
        # 本地檔案比較
        success = comparator.compare_local_files(args.local_file, args.target_file, args.output)
        comparison_desc = "本地檔案"
    else:
        print("❌ 請指定 --gerrit-type 或 --target-file")
        sys.exit(1)
    
    print(f"\n{'='*60}")
    if success:
        print(f"✅ {comparison_desc} 比較完成！")
        print(f"📊 報告檔案: {args.output}")
        print(f"🔧 使用邏輯: 純比較版本（不執行轉換）")
        print(f"📋 Excel 格式: 比較模式專用")
        print(f"🎨 格式化: 所有格式化在 ExcelWriter context 內完成")
        print(f"🔵 link 欄位顏色: RGB(0, 112, 192)")
        if args.gerrit_type:
            print(f"🔍 include 處理: 自動檢測 Gerrit 檔案")
    else:
        print(f"❌ {comparison_desc} 比較失敗")
        print(f"📄 請檢查日誌了解詳細錯誤")
    print(f"{'='*60}")
    
    # 返回狀態碼
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()