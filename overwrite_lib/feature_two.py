"""
功能二：透過 manifest.xml 建立分支映射表 - 修正版 (統一建立分支報告格式)
🔥 修正：統一建立分支與不建立分支時的報告格式，確保表頭顏色和公式一致
🔥 修正：Branch 建立狀態頁籤的欄位名稱改為小寫，與專案列表頁籤一致
🔥 新增：revision, target_branch_revision 改為紅底白字
🔥 新增：target_manifest 連結藍色字體
🔥 新增：所有欄位寬度自動調適
"""
import os
import xml.etree.ElementTree as ET
import pandas as pd
from typing import Dict, List, Any, Optional, Tuple
import utils
import sys
import re

# 加入上一層目錄到路徑
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)
    
# 載入模組 (處理 import 路徑)
try:
    from excel_handler import ExcelHandler
except ImportError:
    # 如果無法直接導入，可能路徑不對，嘗試處理
    import sys
    current_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(current_dir)
    if parent_dir not in sys.path:
        sys.path.insert(0, parent_dir)
    from excel_handler import ExcelHandler

from gerrit_manager import GerritManager
import config

logger = utils.setup_logger(__name__)

class FeatureTwo:
    """功能二：建立分支映射表 - 修正版 (統一建立分支報告格式)"""
    
    def __init__(self):
        self.logger = logger
        self.excel_handler = ExcelHandler()
        self.gerrit_manager = GerritManager()

        # 🔥 從 config 取得當前 Android 版本
        self.current_android_version = config.get_current_android_version()
        self.logger.info(f"使用 Android 版本: {self.current_android_version}")
            
    def process(self, input_file: str, process_type: str, output_filename: str, 
            remove_duplicates: bool, create_branches: bool, check_branch_exists: bool,
            output_folder: str = './output', force_update_branches: bool = False) -> bool:
        """
        處理功能二的主要邏輯 - 修正版（統一報告格式 + 保留 manifest 檔案）
        """
        try:
            self.logger.info("=== 開始執行功能二：建立分支映射表 ===")
            self.logger.info(f"🔥 使用 Android 版本: {self.current_android_version}")
            self.logger.info(f"輸入檔案: {input_file}")
            self.logger.info(f"處理類型: {process_type}")
            self.logger.info(f"輸出檔案: {output_filename}")
            self.logger.info(f"去除重複: {remove_duplicates}")
            self.logger.info(f"建立分支: {create_branches}")
            self.logger.info(f"檢查分支存在性: {check_branch_exists}")
            self.logger.info(f"輸出資料夾: {output_folder}")
            self.logger.info(f"🆕 強制更新分支: {force_update_branches}")
            
            # 確保輸出資料夾存在
            utils.ensure_dir(output_folder)
            
            # 🔥 新增：保存檔案列表，用於最終檢查
            saved_files = []
            
            # 🔥 步驟 0: 保存原始 manifest 檔案
            original_manifest_path = self._save_original_manifest_file(input_file, output_folder)
            if original_manifest_path:
                saved_files.append(original_manifest_path)
            
            # 步驟 0.5: 提取來源 manifest 檔名
            source_manifest_name = self._extract_manifest_filename(input_file)
            
            # 步驟 1: 解析 manifest.xml
            projects = self._parse_manifest(input_file)
            if not projects:
                self.logger.error("無法解析 manifest.xml 或檔案為空")
                return False
            
            self.logger.info(f"成功解析 {len(projects)} 個專案")
            
            # 步驟 2: 轉換專案（使用新的邏輯）
            converted_projects = self._convert_projects(projects, process_type, check_branch_exists, source_manifest_name)
            
            # 步驟 3: 添加連結資訊
            projects_with_links = self._add_links_to_projects(converted_projects)
            
            # 步驟 4: 處理重複資料
            unique_projects, duplicate_projects = self._handle_duplicates(projects_with_links, remove_duplicates)
            
            # 步驟 4.5: 重新編號 SN（避免跳號）
            unique_projects = self._renumber_projects(unique_projects)
            duplicate_projects = self._renumber_projects(duplicate_projects)
            
            self.logger.info(f"處理完成: {len(unique_projects)} 個專案, {len(duplicate_projects)} 個重複")
            
            # 🔥 步驟 4.7: 生成轉換後的 manifest 檔案
            converted_manifest_path = self._generate_converted_manifest(
                unique_projects, input_file, output_folder, process_type
            )
            if converted_manifest_path:
                saved_files.append(converted_manifest_path)
            
            # 步驟 5: 統一生成基本 Excel 報告（無論是否建立分支都使用相同邏輯）
            self._write_excel_unified_basic(unique_projects, duplicate_projects, output_filename, output_folder)
            
            # 步驟 6: 如果選擇建立分支，執行分支建立並添加狀態頁籤
            if create_branches:
                self.logger.info("🚀 開始執行分支建立流程...")
                branch_results = self._create_branches(unique_projects, output_filename, output_folder, force_update_branches)
                # 添加分支建立狀態頁籤
                self._add_branch_status_sheet_with_revision(output_filename, output_folder, branch_results)
                self.logger.info("✅ 分支建立流程完成")
            else:
                self.logger.info("⏭️ 跳過分支建立流程")
            
            # 🔥 步驟 7: 最終檔案檢查報告
            excel_path = os.path.join(output_folder, output_filename)
            saved_files.append(excel_path)
            self._final_file_report(output_folder, saved_files)
            
            self.logger.info(f"=== 功能二執行完成，Excel 檔案：{excel_path} ===")
            return True
            
        except Exception as e:
            self.logger.error(f"功能二執行失敗: {str(e)}")
            return False

    def _extract_manifest_filename(self, input_file: str) -> str:
        """
        🔥 新方法：從輸入檔案路徑提取 manifest 檔名
        """
        try:
            import os
            filename = os.path.basename(input_file)
            self.logger.info(f"提取來源 manifest 檔名: {filename}")
            return filename
        except Exception as e:
            self.logger.error(f"提取 manifest 檔名失敗: {str(e)}")
            return "manifest.xml"

    def _add_branch_status_sheet_with_revision(self, output_file: str, output_folder: str, branch_results: List[Dict]):
        """
        🔥 修正方法：添加分支建立狀態頁籤 - 包含 title 欄位，不影響原有邏輯
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            full_output_path = os.path.join(output_folder, output_file)
            
            self.logger.info(f"🔧 使用 openpyxl 添加分支狀態頁籤（保留公式）: {full_output_path}")
            
            # 🔥 使用 openpyxl 載入現有工作簿（保留公式）
            workbook = load_workbook(full_output_path)
            
            # 🔥 加入分支建立狀態頁籤
            if branch_results:
                # 🔥 為分支結果添加 title 相關欄位（如果需要）
                enhanced_branch_results = []
                for result in branch_results:
                    enhanced_result = result.copy()
                    # 🔥 如果原始結果中沒有這些欄位，添加空值
                    if 'branch_revision' not in enhanced_result:
                        enhanced_result['branch_revision'] = '-'  # 分支建立狀態通常不需要此資訊
                    if 'title' not in enhanced_result:
                        enhanced_result['title'] = '-'  # 🔥 新增
                    if 'target_title' not in enhanced_result:
                        enhanced_result['target_title'] = '-'  # 🔥 新增
                    enhanced_branch_results.append(enhanced_result)
                
                df_branch = pd.DataFrame(enhanced_branch_results)
                
                # 🔥 調整欄位順序，添加 title 相關欄位
                column_order = [
                    'SN', 'Project', 'revision', 'branch_revision', 'title',  # 🔥 新增 title
                    'target_manifest',      # 🔥 紫底白字
                    'target_branch',        # 🔥 改為小寫，綠底白字
                    'target_type',          # 🔥 改為小寫，綠底白字
                    'target_branch_link',   # 🔥 綠底白字
                    'target_branch_revision', 'target_title',  # 🔥 新增 target_title
                    'Status', 'Message', 'Already_Exists', 'Force_Update',
                    'Remote', 'Gerrit_Server'
                ]
                
                # 🔥 映射原始欄位名稱到新的欄位名稱
                if 'Target_Branch' in df_branch.columns:
                    df_branch = df_branch.rename(columns={'Target_Branch': 'target_branch'})
                if 'Target_Type' in df_branch.columns:
                    df_branch = df_branch.rename(columns={'Target_Type': 'target_type'})
                if 'Revision' in df_branch.columns:
                    df_branch = df_branch.rename(columns={'Revision': 'target_branch_revision'})
                
                # 只保留存在的欄位
                column_order = [col for col in column_order if col in df_branch.columns]
                df_branch = df_branch[column_order]
            else:
                # 🔥 空的 DataFrame 結構（包含 title 相關欄位）
                df_branch = pd.DataFrame(columns=[
                    'SN', 'Project', 'revision', 'branch_revision', 'title', 'target_manifest', 
                    'target_branch', 'target_type', 'target_branch_link', 
                    'target_branch_revision', 'target_title',  # 🔥 新增 target_title
                    'Status', 'Message', 'Already_Exists', 'Force_Update',
                    'Remote', 'Gerrit_Server'
                ])
            
            # 🔥 創建新的工作表
            branch_sheet = workbook.create_sheet('Branch 建立狀態')
            
            # 🔥 寫入資料到新工作表
            for r in dataframe_to_rows(df_branch, index=False, header=True):
                branch_sheet.append(r)
            
            # 🔥 格式化新的分支狀態頁籤
            self._format_branch_status_sheet_in_workbook_with_titles(workbook, 'Branch 建立狀態')
            
            # 🔥 保存工作簿（保留原有公式）
            workbook.save(full_output_path)
            workbook.close()
            
            self.logger.info("✅ 成功加入包含 title 欄位的分支建立狀態頁籤（保留公式）")
            
        except Exception as e:
            self.logger.error(f"加入分支狀態頁籤失敗: {str(e)}")

    def _format_branch_status_sheet_in_workbook_with_titles(self, workbook, sheet_name):
        """
        🔥 新方法：在 workbook 中格式化分支建立狀態頁籤 - 包含 title 欄位支援
        """
        try:
            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils import get_column_letter
            
            if sheet_name not in workbook.sheetnames:
                self.logger.warning(f"⚠️ 工作表 '{sheet_name}' 不存在")
                return
                
            worksheet = workbook[sheet_name]
            
            # 定義顏色
            blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            purple_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
            red_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # 🔥 新增：黃底
            white_font = Font(color="FFFFFF", bold=True)
            black_font = Font(color="000000", bold=True)
            
            # 基本格式化
            self.excel_handler._format_worksheet(worksheet)
            
            # 分支建立狀態頁籤：特殊格式（包含 title 欄位）
            self._format_branch_status_sheet_with_titles(worksheet, green_fill, purple_fill, orange_fill, red_fill, yellow_fill, white_font)
            
            # 🔥 自動調適欄位寬度
            self._auto_adjust_column_widths(worksheet)
            
            self.logger.info(f"✅ 已格式化包含 title 的分支狀態頁籤: {sheet_name}")
            
        except Exception as e:
            self.logger.error(f"格式化分支狀態頁籤失敗: {str(e)}")

    def _format_branch_status_sheet_with_titles(self, worksheet, green_fill, purple_fill, orange_fill, red_fill, yellow_fill, white_font):
        """
        🔥 新方法：格式化分支建立狀態頁籤 - 包含 title 欄位格式
        """
        try:
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            
            # 內容樣式
            green_font = Font(color="00B050", bold=True)
            red_content_font = Font(color="FF0000", bold=True)
            blue_font = Font(color="0070C0", bold=True)
            purple_font = Font(color="7030A0", bold=True)
            orange_font = Font(color="FFC000", bold=True)
            black_font = Font(color="000000")
            
            # 🔥 綠底白字欄位（與專案列表頁籤一致）
            green_header_columns = [
                'target_branch', 'target_type', 'target_branch_link'
            ]
            
            # 🔥 紫底白字欄位
            purple_header_columns = ['Remote', 'Gerrit_Server', 'target_manifest']
            
            # 🔥 深紅底白字欄位（revision 相關）
            red_header_columns = ['revision', 'branch_revision', 'target_branch_revision']
            
            # 🔥 黃底白字欄位（title 相關）
            yellow_header_columns = ['title', 'target_title']
            
            # 🔥 橘底白字欄位
            orange_header_columns = ['Force_Update']
            
            # 狀態顏色設定
            status_colors = {
                '成功': {'fill': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                        'font': Font(color="006100", bold=True)},
                '失敗': {'fill': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                        'font': Font(color="9C0006", bold=True)},
                '跳過': {'fill': PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
                        'font': Font(color="0070C0", bold=True)}
            }
            
            # 找到所有欄位位置並設定格式
            for col_num, cell in enumerate(worksheet[1], 1):
                col_letter = get_column_letter(col_num)
                header_value = str(cell.value) if cell.value else ''
                
                # 🔥 綠底白字標頭
                if header_value in green_header_columns:
                    cell.fill = green_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                    
                    # 🔥 設定內容格式
                    if header_value == 'target_type':
                        for row_num in range(2, worksheet.max_row + 1):
                            content_cell = worksheet[f"{col_letter}{row_num}"]
                            if content_cell.value == 'Tag':
                                content_cell.font = blue_font
                            elif content_cell.value == 'Branch':
                                content_cell.font = purple_font
                            else:
                                content_cell.font = black_font
                                
                    elif header_value == 'target_branch_link':
                        # 設定較寬的欄寬
                        worksheet.column_dimensions[col_letter].width = 60
                        
                    else:
                        # 其他綠底欄位使用黑字內容
                        for row_num in range(2, worksheet.max_row + 1):
                            content_cell = worksheet[f"{col_letter}{row_num}"]
                            content_cell.font = black_font
                
                # 🔥 深紅底白字標頭（revision 相關欄位）
                elif header_value in red_header_columns:
                    cell.fill = red_fill
                    cell.font = white_font
                    
                    # revision 相關欄位內容使用黑字
                    for row_num in range(2, worksheet.max_row + 1):
                        content_cell = worksheet[f"{col_letter}{row_num}"]
                        content_cell.font = black_font
                
                # 🔥 新增：黃底白字標頭（title 相關欄位）
                elif header_value in yellow_header_columns:
                    cell.fill = yellow_fill
                    cell.font = white_font
                    
                    # title 相關欄位內容使用黑字
                    for row_num in range(2, worksheet.max_row + 1):
                        content_cell = worksheet[f"{col_letter}{row_num}"]
                        content_cell.font = black_font
                    
                    # 設定較寬欄寬
                    worksheet.column_dimensions[col_letter].width = 50
                
                # 🔥 紫底白字標頭
                elif header_value in purple_header_columns:
                    cell.fill = purple_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                    
                    if header_value == 'Remote':
                        # Remote 欄位：rtk-prebuilt 用紫字
                        for row_num in range(2, worksheet.max_row + 1):
                            content_cell = worksheet[f"{col_letter}{row_num}"]
                            if content_cell.value == 'rtk-prebuilt':
                                content_cell.font = purple_font
                            else:
                                content_cell.font = black_font
                                
                    elif header_value == 'Gerrit_Server':
                        # 設定較寬欄寬
                        worksheet.column_dimensions[col_letter].width = 40
                        # mm2sd-git2 用紫字
                        for row_num in range(2, worksheet.max_row + 1):
                            content_cell = worksheet[f"{col_letter}{row_num}"]
                            if 'mm2sd-git2' in str(content_cell.value):
                                content_cell.font = purple_font
                            else:
                                content_cell.font = black_font
                                
                    elif header_value == 'target_manifest':
                        # 設定較寬欄寬
                        worksheet.column_dimensions[col_letter].width = 50
                        # HYPERLINK 用藍色連結字體
                        blue_link_font = Font(color="0070C0", underline="single")
                        for row_num in range(2, worksheet.max_row + 1):
                            content_cell = worksheet[f"{col_letter}{row_num}"]
                            if content_cell.value and str(content_cell.value).startswith('=HYPERLINK'):
                                content_cell.font = blue_link_font
                            else:
                                content_cell.font = black_font
                
                # 🔥 橘底白字標頭
                elif header_value in orange_header_columns:
                    cell.fill = orange_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                    
                    # Force_Update 欄位："是" 用橘字
                    for row_num in range(2, worksheet.max_row + 1):
                        content_cell = worksheet[f"{col_letter}{row_num}"]
                        if content_cell.value == '是':
                            content_cell.font = orange_font
                        else:
                            content_cell.font = black_font
                
                # 🔥 Status 欄位特殊格式
                elif header_value == 'Status':
                    for row_num in range(2, worksheet.max_row + 1):
                        content_cell = worksheet[f"{col_letter}{row_num}"]
                        status = str(content_cell.value) if content_cell.value else ''
                        
                        if status in status_colors:
                            content_cell.fill = status_colors[status]['fill']
                            content_cell.font = status_colors[status]['font']
            
            self.logger.info("✅ 已設定包含 title 欄位的分支建立狀態頁籤格式")
            
        except Exception as e:
            self.logger.error(f"格式化分支建立狀態頁籤失敗: {str(e)}")
                        
    def _format_branch_status_sheet_in_workbook(self, workbook, sheet_name):
        """
        🔥 新方法：在 workbook 中格式化分支建立狀態頁籤
        """
        try:
            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils import get_column_letter
            
            if sheet_name not in workbook.sheetnames:
                self.logger.warning(f"⚠️ 工作表 '{sheet_name}' 不存在")
                return
                
            worksheet = workbook[sheet_name]
            
            # 定義顏色
            blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            purple_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
            red_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")  # 🔥 改為RGB(192,80,77)的深紅色
            white_font = Font(color="FFFFFF", bold=True)
            black_font = Font(color="000000", bold=True)  # 🔥 深紅底用白字
            
            # 基本格式化
            self.excel_handler._format_worksheet(worksheet)
            
            # 分支建立狀態頁籤：特殊格式（包含 revision 欄位）
            self._format_branch_status_sheet_with_revision(worksheet, green_fill, purple_fill, orange_fill, red_fill, white_font)
            
            # 🔥 自動調適欄位寬度
            self._auto_adjust_column_widths(worksheet)
            
            self.logger.info(f"✅ 已格式化分支狀態頁籤: {sheet_name}")
            
        except Exception as e:
            self.logger.error(f"格式化分支狀態頁籤失敗: {str(e)}")

    def _format_branch_status_sheet_with_revision(self, worksheet, green_fill, purple_fill, orange_fill, red_fill, white_font):
        """
        🔥 修正方法：格式化分支建立狀態頁籤 - 包含 branch_revision 欄位格式
        """
        try:
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            
            # 內容樣式
            green_font = Font(color="00B050", bold=True)
            red_content_font = Font(color="FF0000", bold=True)
            blue_font = Font(color="0070C0", bold=True)
            purple_font = Font(color="7030A0", bold=True)
            orange_font = Font(color="FFC000", bold=True)
            black_font = Font(color="000000")
            
            # 🔥 綠底白字欄位（與專案列表頁籤一致）
            green_header_columns = [
                'target_branch', 'target_type', 'target_branch_link'
            ]
            
            # 🔥 紫底白字欄位
            purple_header_columns = ['Remote', 'Gerrit_Server', 'target_manifest']
            
            # 🔥 深紅底白字欄位（revision 相關，包含新的 branch_revision）
            red_header_columns = ['revision', 'branch_revision', 'target_branch_revision']
            
            # 🔥 橘底白字欄位
            orange_header_columns = ['Force_Update']
            
            # 狀態顏色設定
            status_colors = {
                '成功': {'fill': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                        'font': Font(color="006100", bold=True)},
                '失敗': {'fill': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                        'font': Font(color="9C0006", bold=True)},
                '跳過': {'fill': PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
                        'font': Font(color="0070C0", bold=True)}
            }
            
            # 找到所有欄位位置並設定格式
            for col_num, cell in enumerate(worksheet[1], 1):
                col_letter = get_column_letter(col_num)
                header_value = str(cell.value) if cell.value else ''
                
                # 🔥 綠底白字標頭（與專案列表一致）
                if header_value in green_header_columns:
                    cell.fill = green_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                    
                    # 🔥 設定內容格式
                    if header_value == 'target_type':
                        for row_num in range(2, worksheet.max_row + 1):
                            content_cell = worksheet[f"{col_letter}{row_num}"]
                            if content_cell.value == 'Tag':
                                content_cell.font = blue_font
                            elif content_cell.value == 'Branch':
                                content_cell.font = purple_font
                            else:
                                content_cell.font = black_font
                                
                    elif header_value == 'target_branch_link':
                        # 設定較寬的欄寬
                        worksheet.column_dimensions[col_letter].width = 60
                        
                    else:
                        # 其他綠底欄位使用黑字內容
                        for row_num in range(2, worksheet.max_row + 1):
                            content_cell = worksheet[f"{col_letter}{row_num}"]
                            content_cell.font = black_font
                
                # 🔥 深紅底白字標頭（revision 相關欄位，包含新的 branch_revision）
                elif header_value in red_header_columns:
                    cell.fill = red_fill
                    cell.font = white_font
                    
                    # revision 相關欄位內容使用黑字
                    for row_num in range(2, worksheet.max_row + 1):
                        content_cell = worksheet[f"{col_letter}{row_num}"]
                        content_cell.font = black_font
                
                # 🔥 紫底白字標頭
                elif header_value in purple_header_columns:
                    cell.fill = purple_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                    
                    if header_value == 'Remote':
                        # Remote 欄位：rtk-prebuilt 用紫字
                        for row_num in range(2, worksheet.max_row + 1):
                            content_cell = worksheet[f"{col_letter}{row_num}"]
                            if content_cell.value == 'rtk-prebuilt':
                                content_cell.font = purple_font
                            else:
                                content_cell.font = black_font
                                
                    elif header_value == 'Gerrit_Server':
                        # 設定較寬欄寬
                        worksheet.column_dimensions[col_letter].width = 40
                        # mm2sd-git2 用紫字
                        for row_num in range(2, worksheet.max_row + 1):
                            content_cell = worksheet[f"{col_letter}{row_num}"]
                            if 'mm2sd-git2' in str(content_cell.value):
                                content_cell.font = purple_font
                            else:
                                content_cell.font = black_font
                                
                    elif header_value == 'target_manifest':
                        # 設定較寬欄寬
                        worksheet.column_dimensions[col_letter].width = 50
                        # HYPERLINK 用藍色連結字體
                        blue_link_font = Font(color="0070C0", underline="single")
                        for row_num in range(2, worksheet.max_row + 1):
                            content_cell = worksheet[f"{col_letter}{row_num}"]
                            if content_cell.value and str(content_cell.value).startswith('=HYPERLINK'):
                                content_cell.font = blue_link_font
                            else:
                                content_cell.font = black_font
                
                # 🔥 橘底白字標頭
                elif header_value in orange_header_columns:
                    cell.fill = orange_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                    
                    # Force_Update 欄位："是" 用橘字
                    for row_num in range(2, worksheet.max_row + 1):
                        content_cell = worksheet[f"{col_letter}{row_num}"]
                        if content_cell.value == '是':
                            content_cell.font = orange_font
                        else:
                            content_cell.font = black_font
                
                # 🔥 Status 欄位特殊格式
                elif header_value == 'Status':
                    for row_num in range(2, worksheet.max_row + 1):
                        content_cell = worksheet[f"{col_letter}{row_num}"]
                        status = str(content_cell.value) if content_cell.value else ''
                        
                        if status in status_colors:
                            content_cell.fill = status_colors[status]['fill']
                            content_cell.font = status_colors[status]['font']
            
            self.logger.info("✅ 已設定包含 branch_revision 欄位的分支建立狀態頁籤格式")
            
        except Exception as e:
            self.logger.error(f"格式化分支建立狀態頁籤失敗: {str(e)}")

    # 修改 Excel 輸出的欄位順序，在 revision 右方添加 branch_revision
    def _write_excel_unified_basic(self, projects: List[Dict], duplicate_projects: List[Dict], 
                      output_file: str, output_folder: str = None):
        """
        🔥 修正方法：統一的基本 Excel 寫入 - 添加 title, target_title, title_diff 欄位
        """
        try:
            # 處理輸出檔案路徑
            if not output_file:
                raise ValueError("輸出檔案名稱不能為空")
            
            if output_folder:
                utils.ensure_dir(output_folder)
                full_output_path = os.path.join(output_folder, output_file)
            else:
                output_dir = os.path.dirname(output_file)
                if not output_dir:
                    output_file = os.path.join('.', output_file)
                    output_dir = '.'
                utils.ensure_dir(output_dir)
                full_output_path = output_file
            
            self.logger.info(f"寫入統一格式 Excel 檔案: {full_output_path}")
            
            with pd.ExcelWriter(full_output_path, engine='openpyxl') as writer:
                # 🔥 頁籤 1: 專案列表（添加 title, target_title, title_diff 欄位）
                if projects:
                    # 🔥 重要：移除任何可能存在的公式欄位值
                    clean_projects = []
                    for project in projects:
                        clean_project = project.copy()
                        # 強制移除公式欄位，避免覆蓋公式
                        if 'revision_diff' in clean_project:
                            del clean_project['revision_diff']
                        if 'title_diff' in clean_project:  # 🔥 新增：移除 title_diff
                            del clean_project['title_diff']
                        clean_projects.append(clean_project)
 
                    df_main = pd.DataFrame(clean_projects)
                    
                    # 🔥 修改欄位順序：添加 title, target_title, title_diff
                    main_column_order = [
                        'SN', 'source_manifest', 'name', 'path', 
                        'revision', 'branch_revision', 'title',  # 🔥 在 branch_revision 後加 title
                        'upstream', 'dest-branch',
                        'target_manifest', 'target_branch', 'target_type', 'target_branch_exists', 
                        'target_branch_revision', 'target_title', 'revision_diff', 'title_diff',  # 🔥 在 target_branch_revision 後加 target_title，在 revision_diff 後加 title_diff
                        'target_branch_link', 'branch_link',
                        'target_open_project_link', 'open_project_link',
                        'groups', 'clone-depth', 'remote'
                    ]
                    
                    # 添加其他可能存在的欄位
                    excluded_columns = ['effective_revision']
                    for col in df_main.columns:
                        if col not in main_column_order and col not in excluded_columns:
                            main_column_order.append(col)
                    
                    # 只保留存在的欄位
                    main_column_order = [col for col in main_column_order if col in df_main.columns]
                    df_main = df_main[main_column_order]
                    
                    # 🔥 關鍵修正：確保 revision_diff 和 title_diff 欄位存在且在正確位置
                    if 'revision_diff' not in df_main.columns:
                        # 在 target_title 後面插入空的 revision_diff 欄位
                        if 'target_title' in df_main.columns:
                            target_title_idx = df_main.columns.get_loc('target_title')
                            # 在 target_title 後插入
                            df_main.insert(target_title_idx + 1, 'revision_diff', None)
                        else:
                            # 如果找不到 target_title，則在最後添加
                            df_main['revision_diff'] = None
                    
                    # 🔥 新增：確保 title_diff 欄位存在
                    if 'title_diff' not in df_main.columns:
                        # 在 revision_diff 後面插入空的 title_diff 欄位
                        if 'revision_diff' in df_main.columns:
                            revision_diff_idx = df_main.columns.get_loc('revision_diff')
                            # 在 revision_diff 後插入
                            df_main.insert(revision_diff_idx + 1, 'title_diff', None)
                        else:
                            # 如果找不到 revision_diff，則在最後添加
                            df_main['title_diff'] = None
                else:
                    # 空的 DataFrame 結構（確保包含所有必要欄位）
                    df_main = pd.DataFrame(columns=[
                        'SN', 'source_manifest', 'name', 'path', 
                        'revision', 'branch_revision', 'title',  # 🔥 新增 title
                        'upstream', 'dest-branch',
                        'target_manifest', 'target_branch', 'target_type', 'target_branch_exists', 
                        'target_branch_revision', 'target_title', 'revision_diff', 'title_diff',  # 🔥 新增 target_title, title_diff
                        'target_branch_link', 'branch_link',
                        'target_open_project_link', 'open_project_link',
                        'groups', 'clone-depth', 'remote'
                    ])
                
                df_main.to_excel(writer, sheet_name='專案列表', index=False)
                self.logger.info(f"✅ 專案列表頁籤寫入完成，共 {len(projects)} 筆資料")
                
                # 🔥 頁籤 2: 重複專案（同樣處理）
                if duplicate_projects:
                    # 🔥 重要：移除任何可能存在的公式欄位值
                    clean_duplicates = []
                    for project in duplicate_projects:
                        clean_project = project.copy()
                        # 強制移除公式欄位，避免覆蓋公式
                        if 'revision_diff' in clean_project:
                            del clean_project['revision_diff']
                        if 'title_diff' in clean_project:  # 🔥 新增
                            del clean_project['title_diff']
                        clean_duplicates.append(clean_project)
                    
                    df_dup = pd.DataFrame(clean_duplicates)
                    
                    # 🔥 重複頁籤使用相同的欄位順序
                    dup_column_order = [
                        'SN', 'source_manifest', 'name', 'path', 
                        'revision', 'branch_revision', 'title',  # 🔥 新增 title
                        'upstream', 'dest-branch',
                        'target_manifest', 'target_branch', 'target_type', 'target_branch_exists', 
                        'target_branch_revision', 'target_title', 'revision_diff', 'title_diff',  # 🔥 新增 target_title, title_diff
                        'target_branch_link', 'branch_link',
                        'groups', 'clone-depth', 'remote'
                    ]
                    
                    # 添加其他欄位
                    excluded_columns = ['effective_revision']
                    for col in df_dup.columns:
                        if col not in dup_column_order and col not in excluded_columns:
                            dup_column_order.append(col)
                    
                    dup_column_order = [col for col in dup_column_order if col in df_dup.columns]
                    df_dup = df_dup[dup_column_order]
                    
                    # 🔥 關鍵修正：確保重複頁籤的公式欄位也存在
                    if 'revision_diff' not in df_dup.columns:
                        if 'target_title' in df_dup.columns:
                            target_title_idx = df_dup.columns.get_loc('target_title')
                            df_dup.insert(target_title_idx + 1, 'revision_diff', None)
                        else:
                            df_dup['revision_diff'] = None
                    
                    # 🔥 新增：確保 title_diff 欄位存在
                    if 'title_diff' not in df_dup.columns:
                        if 'revision_diff' in df_dup.columns:
                            revision_diff_idx = df_dup.columns.get_loc('revision_diff')
                            df_dup.insert(revision_diff_idx + 1, 'title_diff', None)
                        else:
                            df_dup['title_diff'] = None
                    
                    df_dup.to_excel(writer, sheet_name='重覆', index=False)
                    self.logger.info(f"建立 '重覆' 頁籤，共 {len(duplicate_projects)} 筆資料")
                
                self.logger.info("📋 DataFrame 寫入完成，開始設定公式...")
                
            # 🔥 重要：在 writer 關閉後重新開啟來設定公式
            self._add_formulas_to_existing_excel(full_output_path)
            
            # 🔥 格式化 Excel
            self._format_existing_excel(full_output_path)
            
            self.logger.info(f"✅ 統一格式 Excel 檔案寫入完成: {full_output_path}")
            
        except Exception as e:
            self.logger.error(f"寫入統一格式 Excel 檔案失敗: {str(e)}")
            raise

    def _format_existing_excel(self, excel_path: str):
        """
        🔥 新方法：格式化現有 Excel 檔案 - 添加 title 欄位格式支援
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font
            
            self.logger.info(f"🎨 開始格式化 Excel 檔案: {excel_path}")
            
            # 載入現有的 Excel 檔案
            workbook = load_workbook(excel_path)
            
            # 定義顏色
            blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")    # 藍底
            green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")   # 綠底
            orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")  # 橘底
            purple_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")  # 🔥 紫底
            red_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")     # 🔥 改為RGB(192,80,77)的深紅色
            yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # 🔥 新增：黃底（用於 title 欄位）
            white_font = Font(color="FFFFFF", bold=True)  # 白字
            black_font = Font(color="000000", bold=True)  # 🔥 紅底用白字
            
            for sheet_name in workbook.sheetnames:
                if sheet_name in ['專案列表', '重覆']:
                    worksheet = workbook[sheet_name]
                    
                    # 基本格式化
                    self.excel_handler._format_worksheet(worksheet)
                    
                    # 統一格式化連結欄位
                    self._format_link_columns_unified(worksheet, blue_fill, green_fill, white_font)
                    
                    # 統一格式化 revision_diff 和 title_diff 欄位
                    self._format_diff_columns_unified(worksheet, orange_fill, white_font)
                    
                    # 統一格式化目標分支欄位
                    self._format_target_branch_columns_unified(worksheet, green_fill, white_font)
                    
                    # 🔥 格式化 revision 相關欄位（深紅底白字）
                    self._format_revision_columns_unified(worksheet, red_fill, white_font)
                    
                    # 🔥 格式化 manifest 相關欄位（紫底白字）
                    self._format_manifest_columns_unified(worksheet, purple_fill, white_font)
                    
                    # 🔥 新增：格式化 title 相關欄位（黃底白字）
                    self._format_title_columns_unified(worksheet, yellow_fill, white_font)
                    
                    # 🔥 自動調適欄位寬度
                    self._auto_adjust_column_widths(worksheet)
            
            # 保存檔案
            workbook.save(excel_path)
            workbook.close()
            self.logger.info(f"✅ 格式化完成並已保存: {excel_path}")
            
        except Exception as e:
            self.logger.error(f"格式化 Excel 失敗: {str(e)}")
            import traceback
            self.logger.error(f"錯誤詳情: {traceback.format_exc()}")

    def _format_diff_columns_unified(self, worksheet, orange_fill, white_font):
        """
        🔥 修正方法：格式化 diff 欄位為橘底白字，N綠字/Y紅字，並置中對齊 - 支援 title_diff
        """
        try:
            from openpyxl.styles import Font, Alignment  # 🔥 加入 Alignment
            from openpyxl.utils import get_column_letter
            from openpyxl.formatting.rule import CellIsRule
            
            # 內容樣式
            green_font = Font(color="00B050", bold=True)  # N 的綠字
            red_font = Font(color="FF0000", bold=True)    # Y 的紅字
            
            # 🔥 新增：置中對齊設定
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # 🔥 diff 欄位列表（包含新的 title_diff）
            diff_columns = ['revision_diff', 'title_diff']
            
            for diff_column in diff_columns:
                # 找到 diff 欄位的位置
                diff_col = None
                for col_num, cell in enumerate(worksheet[1], 1):
                    header_value = str(cell.value) if cell.value else ''
                    if header_value == diff_column:
                        diff_col = col_num
                        break
                
                if diff_col:
                    col_letter = get_column_letter(diff_col)
                    
                    # 🔥 格式化標題（橘底白字 + 置中）
                    header_cell = worksheet[f"{col_letter}1"]
                    header_cell.fill = orange_fill
                    header_cell.font = white_font
                    header_cell.alignment = center_alignment  # 🔥 新增：標題置中
                    
                    # 設定欄寬
                    worksheet.column_dimensions[col_letter].width = 13.71  # 🔥 精確設定為 13.71
                    
                    # 🔥 新增：為所有資料欄位設定置中對齊
                    for row_num in range(2, worksheet.max_row + 1):
                        cell = worksheet[f"{col_letter}{row_num}"]
                        cell.alignment = center_alignment  # 🔥 關鍵修復：資料置中
                    
                    # 定義資料範圍
                    data_range = f"{col_letter}2:{col_letter}{worksheet.max_row}"
                    
                    # 條件格式規則 1: 當值為 "N" 時使用綠字（相同）
                    rule_n = CellIsRule(
                        operator='equal',
                        formula=['"N"'],
                        font=green_font
                    )
                    worksheet.conditional_formatting.add(data_range, rule_n)
                    
                    # 條件格式規則 2: 當值為 "Y" 時使用紅字（不同或空值）
                    rule_y = CellIsRule(
                        operator='equal',
                        formula=['"Y"'],
                        font=red_font
                    )
                    worksheet.conditional_formatting.add(data_range, rule_y)
                    
                    self.logger.info(f"✅ 已設定 {diff_column} 欄位格式：標題橘底白字，N綠字/Y紅字，全部置中對齊")
                    
        except Exception as e:
            self.logger.error(f"格式化 diff 欄位失敗: {str(e)}")
            
    def _format_title_columns_unified(self, worksheet, yellow_fill, white_font):
        """
        🔥 新方法：格式化 title 相關欄位為黃底白字
        """
        try:
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
            
            black_font = Font(color="000000")         # 一般內容用黑字
            
            # 🔥 需要黃底白字的 title 欄位
            title_columns = ['title', 'target_title']
            
            # 找到 title 欄位的位置
            for col_num, cell in enumerate(worksheet[1], 1):
                header_value = str(cell.value) if cell.value else ''
                
                if header_value in title_columns:
                    col_letter = get_column_letter(col_num)
                    
                    # 🔥 設定標頭為黃底白字
                    cell.fill = yellow_fill
                    cell.font = white_font
                    
                    # 🔥 設定內容格式（黑字）
                    for row_num in range(2, worksheet.max_row + 1):
                        content_cell = worksheet[f"{col_letter}{row_num}"]
                        content_cell.font = black_font
                    
                    # 設定欄寬（title 內容可能較長）
                    if header_value == 'title':
                        worksheet.column_dimensions[col_letter].width = 50  # title 較寬
                    elif header_value == 'target_title':
                        worksheet.column_dimensions[col_letter].width = 50  # target_title 也較寬
            
            self.logger.info("✅ 已設定 title 欄位為黃底白字")
            
        except Exception as e:
            self.logger.error(f"格式化 title 欄位失敗: {str(e)}")
            
    # 修改格式化邏輯，讓 branch_revision 也使用深紅底白字
    def _format_revision_columns_unified(self, worksheet, red_fill, white_font):
        """
        🔥 修正方法：格式化 revision 相關欄位為深紅底白字（包含新的 branch_revision）
        """
        try:
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
            
            content_font = Font(color="000000")           # 🔥 內容用黑字
            
            # 🔥 需要深紅底白字的 revision 欄位（新增 branch_revision）
            revision_columns = ['revision', 'branch_revision', 'target_branch_revision']
            
            # 找到 revision 欄位的位置
            for col_num, cell in enumerate(worksheet[1], 1):
                header_value = str(cell.value) if cell.value else ''
                
                if header_value in revision_columns:
                    col_letter = get_column_letter(col_num)
                    
                    # 🔥 設定標頭為深紅底白字
                    cell.fill = red_fill
                    cell.font = white_font
                    
                    # 🔥 設定內容為黑字
                    for row_num in range(2, worksheet.max_row + 1):
                        content_cell = worksheet[f"{col_letter}{row_num}"]
                        content_cell.font = content_font
            
            self.logger.info("✅ 已設定 revision 欄位為深紅底白字（包含 branch_revision）")
            
        except Exception as e:
            self.logger.error(f"格式化 revision 欄位失敗: {str(e)}")

    def _format_manifest_columns_unified(self, worksheet, purple_fill, white_font):
        """
        🔥 新方法：格式化 manifest 相關欄位為紫底白字
        """
        try:
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
            
            black_font = Font(color="000000")         # 一般內容用黑字
            blue_link_font = Font(color="0070C0", underline="single")  # HYPERLINK 用藍色連結
            
            # 🔥 需要紫底白字的 manifest 欄位
            manifest_columns = ['source_manifest', 'target_manifest']
            
            # 找到 manifest 欄位的位置
            for col_num, cell in enumerate(worksheet[1], 1):
                header_value = str(cell.value) if cell.value else ''
                
                if header_value in manifest_columns:
                    col_letter = get_column_letter(col_num)
                    
                    # 🔥 設定標頭為紫底白字
                    cell.fill = purple_fill
                    cell.font = white_font
                    
                    # 🔥 設定內容格式
                    for row_num in range(2, worksheet.max_row + 1):
                        content_cell = worksheet[f"{col_letter}{row_num}"]
                        if content_cell.value:
                            # 如果是 HYPERLINK 函數，設定為藍色連結
                            if str(content_cell.value).startswith('=HYPERLINK'):
                                content_cell.font = blue_link_font
                            else:
                                # 一般文字用黑字
                                content_cell.font = black_font
                    
                    # 設定欄寬
                    if header_value == 'target_manifest':
                        worksheet.column_dimensions[col_letter].width = 50  # target_manifest 較寬
                    else:
                        worksheet.column_dimensions[col_letter].width = 30  # source_manifest 適中
            
            self.logger.info("✅ 已設定 manifest 欄位為紫底白字")
            
        except Exception as e:
            self.logger.error(f"格式化 manifest 欄位失敗: {str(e)}")

    def _format_target_manifest_column(self, worksheet, blue_fill, white_font):
        """
        🔥 新方法：格式化 target_manifest 欄位為藍色連結
        """
        try:
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
            
            blue_font = Font(color="0070C0", underline="single")  # 藍色連結字體
            
            # 找到 target_manifest 欄位的位置
            target_manifest_col = None
            for col_num, cell in enumerate(worksheet[1], 1):
                header_value = str(cell.value) if cell.value else ''
                if header_value == 'target_manifest':
                    target_manifest_col = col_num
                    break
            
            if target_manifest_col:
                col_letter = get_column_letter(target_manifest_col)
                
                # 🔥 設定標頭（可以是藍底白字或保持預設）
                header_cell = worksheet[f"{col_letter}1"]
                # header_cell.fill = blue_fill  # 如果要設定標頭背景色
                # header_cell.font = white_font
                
                # 🔥 設定內容為藍色連結
                for row_num in range(2, worksheet.max_row + 1):
                    content_cell = worksheet[f"{col_letter}{row_num}"]
                    # 只有當內容包含 HYPERLINK 函數時才設定藍色字體
                    if content_cell.value and str(content_cell.value).startswith('=HYPERLINK'):
                        content_cell.font = blue_font
                
                # 設定較寬的欄寬
                worksheet.column_dimensions[col_letter].width = 60
                
                self.logger.info("✅ 已設定 target_manifest 欄位為藍色連結")
            
        except Exception as e:
            self.logger.error(f"格式化 target_manifest 欄位失敗: {str(e)}")

    def _auto_adjust_column_widths(self, worksheet):
        """
        🔥 新方法：自動調適所有欄位寬度
        確保最小寬度不小於表頭文字寬度
        """
        try:
            from openpyxl.utils import get_column_letter
            
            for col_num in range(1, worksheet.max_column + 1):
                col_letter = get_column_letter(col_num)
                
                # 獲取表頭文字長度
                header_cell = worksheet[f"{col_letter}1"]
                header_text = str(header_cell.value) if header_cell.value else ''
                min_width = max(len(header_text) + 2, 8)  # 表頭文字寬度 + 緩衝，最小8個字符
                
                # 計算欄位內容的最大寬度
                max_content_width = min_width
                
                # 檢查前100行的內容（避免處理時間過長）
                check_rows = min(worksheet.max_row, 100)
                
                for row_num in range(1, check_rows + 1):
                    cell = worksheet[f"{col_letter}{row_num}"]
                    if cell.value:
                        # 處理 HYPERLINK 函數的特殊情況
                        cell_text = str(cell.value)
                        if cell_text.startswith('=HYPERLINK'):
                            # 估算超連結顯示文字的長度
                            if '","' in cell_text:
                                display_text = cell_text.split('","')[1].rstrip('")')
                                content_width = len(display_text)
                            else:
                                content_width = 30  # 預設超連結寬度
                        else:
                            content_width = len(cell_text)
                        
                        max_content_width = max(max_content_width, content_width)
                
                # 設定欄位寬度，考慮一些特殊欄位的最小寬度
                if header_text in ['target_branch_link', 'branch_link', 'target_manifest']:
                    # 連結欄位設定較寬
                    final_width = max(max_content_width + 2, 50)
                elif header_text in ['revision', 'target_branch_revision']:
                    # revision 欄位設定適中寬度
                    final_width = max(max_content_width + 2, 25)
                elif header_text in ['revision_diff', 'title_diff']:
                    # revision_diff 欄位固定寬度並置中
                    final_width = 13.71  # 🔥 精確設定為 13.71
                else:
                    # 一般欄位
                    final_width = max(max_content_width + 2, min_width)
                
                # 設定最大寬度限制（避免過寬）
                final_width = min(final_width, 80)
                
                worksheet.column_dimensions[col_letter].width = final_width
                
                self.logger.debug(f"欄位 {header_text} ({col_letter}): 設定寬度 {final_width}")
            
            self.logger.info("✅ 已完成所有欄位寬度自動調適")
            
        except Exception as e:
            self.logger.error(f"自動調適欄位寬度失敗: {str(e)}")
            
    def _add_formulas_to_existing_excel(self, excel_path: str):
        """
        🔥 修正版：在現有 Excel 檔案中添加公式 - 支援 hash 判斷的比較邏輯 + title_diff 公式
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            
            self.logger.info(f"🔧 開始為 Excel 檔案添加公式: {excel_path}")
            
            # 載入現有的 Excel 檔案
            workbook = load_workbook(excel_path)
            
            for sheet_name in ['專案列表', '重覆']:
                if sheet_name not in workbook.sheetnames:
                    self.logger.warning(f"⚠️ 工作表 '{sheet_name}' 不存在")
                    continue
                    
                worksheet = workbook[sheet_name]
                self.logger.info(f"🔧 開始為 '{sheet_name}' 頁籤設定公式...")
                
                # 找到各欄位的位置
                revision_col = None
                branch_revision_col = None
                title_col = None  # 🔥 新增
                target_revision_col = None
                target_title_col = None  # 🔥 新增
                revision_diff_col = None
                title_diff_col = None  # 🔥 新增
                
                # 打印所有標頭以便調試
                headers = []
                for col_num, cell in enumerate(worksheet[1], 1):
                    header = str(cell.value) if cell.value else ''
                    headers.append(f"{get_column_letter(col_num)}:{header}")
                    
                    if header == 'revision':
                        revision_col = col_num
                        self.logger.debug(f"找到 revision 欄位: {get_column_letter(col_num)} (第{col_num}欄)")
                    elif header == 'branch_revision':
                        branch_revision_col = col_num
                        self.logger.debug(f"找到 branch_revision 欄位: {get_column_letter(col_num)} (第{col_num}欄)")
                    elif header == 'title':  # 🔥 新增
                        title_col = col_num
                        self.logger.debug(f"找到 title 欄位: {get_column_letter(col_num)} (第{col_num}欄)")
                    elif header == 'target_branch_revision':
                        target_revision_col = col_num
                        self.logger.debug(f"找到 target_branch_revision 欄位: {get_column_letter(col_num)} (第{col_num}欄)")
                    elif header == 'target_title':  # 🔥 新增
                        target_title_col = col_num
                        self.logger.debug(f"找到 target_title 欄位: {get_column_letter(col_num)} (第{col_num}欄)")
                    elif header == 'revision_diff':
                        revision_diff_col = col_num
                        self.logger.debug(f"找到 revision_diff 欄位: {get_column_letter(col_num)} (第{col_num}欄)")
                    elif header == 'title_diff':  # 🔥 新增
                        title_diff_col = col_num
                        self.logger.debug(f"找到 title_diff 欄位: {get_column_letter(col_num)} (第{col_num}欄)")
                
                self.logger.debug(f"'{sheet_name}' 所有標頭: {', '.join(headers)}")
                
                # 🔥 設定 revision_diff 公式（需要所有四個欄位）
                if revision_col and branch_revision_col and target_revision_col and revision_diff_col:
                    revision_letter = get_column_letter(revision_col)
                    branch_revision_letter = get_column_letter(branch_revision_col)
                    target_letter = get_column_letter(target_revision_col)
                    diff_letter = get_column_letter(revision_diff_col)
                    
                    self.logger.info(f"📍 revision_diff 欄位對應: revision={revision_letter}, branch_revision={branch_revision_letter}, target_branch_revision={target_letter}, revision_diff={diff_letter}")
                    
                    # 🔥 為每一行設定 revision_diff 公式（從第2行開始到最後一行）
                    revision_diff_formula_count = 0
                    for row_num in range(2, worksheet.max_row + 1):
                        # 🔥 新邏輯：程式判斷該用哪個欄位，然後產生簡單公式
                        revision_cell = worksheet[f"{revision_letter}{row_num}"]
                        revision_value = str(revision_cell.value) if revision_cell.value else ''
                        
                        # 判斷是否為 hash（40字符長度）
                        is_hash = self._is_revision_hash(revision_value)
                        
                        if is_hash:
                            # 如果是 hash，用 revision 欄位比較
                            compare_letter = revision_letter
                            self.logger.debug(f"第{row_num}行使用 revision 欄位比較（hash）")
                        else:
                            # 如果不是 hash，用 branch_revision 欄位比較
                            compare_letter = branch_revision_letter
                            self.logger.debug(f"第{row_num}行使用 branch_revision 欄位比較（非hash）")
                        
                        # 產生簡單的比較公式
                        formula = (
                            f'=IF(OR({target_letter}{row_num}="-", '
                            f'{target_letter}{row_num}="", '
                            f'{compare_letter}{row_num}=""), '
                            f'"Y", '
                            f'IF({compare_letter}{row_num}={target_letter}{row_num}, '
                            f'"N", "Y"))'
                        )
                        
                        # 設定公式到儲存格
                        cell = worksheet[f"{diff_letter}{row_num}"]
                        cell.value = formula
                        revision_diff_formula_count += 1
                    
                    self.logger.info(f"✅ 已為 '{sheet_name}' 頁籤設定 {revision_diff_formula_count} 個 revision_diff 公式")
                
                else:
                    missing_cols = []
                    if not revision_col:
                        missing_cols.append("revision")
                    if not branch_revision_col:
                        missing_cols.append("branch_revision")
                    if not target_revision_col:
                        missing_cols.append("target_branch_revision")
                    if not revision_diff_col:
                        missing_cols.append("revision_diff")
                        
                    self.logger.error(f"❌ 無法為 '{sheet_name}' 頁籤設定 revision_diff 公式，缺少欄位: {', '.join(missing_cols)}")
                
                # 🔥 新增：設定 title_diff 公式
                if title_col and target_title_col and title_diff_col:
                    title_letter = get_column_letter(title_col)
                    target_title_letter = get_column_letter(target_title_col)
                    title_diff_letter = get_column_letter(title_diff_col)
                    
                    self.logger.info(f"📍 title_diff 欄位對應: title={title_letter}, target_title={target_title_letter}, title_diff={title_diff_letter}")
                    
                    # 🔥 為每一行設定 title_diff 公式（從第2行開始到最後一行）
                    title_diff_formula_count = 0
                    for row_num in range(2, worksheet.max_row + 1):
                        # 🔥 title_diff 公式：比較 title 和 target_title
                        formula = (
                            f'=IF(OR({title_letter}{row_num}="-", '
                            f'{title_letter}{row_num}="", '
                            f'{target_title_letter}{row_num}="-", '
                            f'{target_title_letter}{row_num}=""), '
                            f'"Y", '
                            f'IF({title_letter}{row_num}={target_title_letter}{row_num}, '
                            f'"N", "Y"))'
                        )
                        
                        # 設定公式到儲存格
                        cell = worksheet[f"{title_diff_letter}{row_num}"]
                        cell.value = formula
                        title_diff_formula_count += 1
                    
                    self.logger.info(f"✅ 已為 '{sheet_name}' 頁籤設定 {title_diff_formula_count} 個 title_diff 公式")
                    
                else:
                    missing_cols = []
                    if not title_col:
                        missing_cols.append("title")
                    if not target_title_col:
                        missing_cols.append("target_title")
                    if not title_diff_col:
                        missing_cols.append("title_diff")
                        
                    self.logger.error(f"❌ 無法為 '{sheet_name}' 頁籤設定 title_diff 公式，缺少欄位: {', '.join(missing_cols)}")
            
            # 保存檔案
            workbook.save(excel_path)
            workbook.close()
            self.logger.info(f"✅ 公式設定完成並已保存: {excel_path}")
            
        except Exception as e:
            self.logger.error(f"添加公式失敗: {str(e)}")
            import traceback
            self.logger.error(f"錯誤詳情: {traceback.format_exc()}")

    # ============================================
    # 🔥 其他原有方法保持不變
    # ============================================

    def _format_link_columns_unified(self, worksheet, blue_fill, green_fill, white_font):
        """
        🔥 新方法：統一格式化連結欄位 - 支援 HYPERLINK 函數
        確保O欄的branch_link樣式與之前S欄的branch_link完全一致
        """
        try:
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font
            
            blue_link_font = Font(color="0070C0", underline="single")  # 藍色連結字體
            green_link_font = Font(color="0070C0", underline="single")  # 連結統一用藍色
            
            # 找到連結欄位的位置
            link_columns = {}
            for col_num, cell in enumerate(worksheet[1], 1):
                header_value = str(cell.value) if cell.value else ''
                
                if header_value == 'branch_link':
                    link_columns['branch_link'] = col_num
                elif header_value == 'target_branch_link':
                    link_columns['target_branch_link'] = col_num
                elif header_value == 'target_open_project_link':  # 🔥 新增
                    link_columns['target_open_project_link'] = col_num
                elif header_value == 'open_project_link':  # 🔥 新增
                    link_columns['open_project_link'] = col_num
            
            # 格式化 branch_link 欄位 (藍底白字，內容藍色連結)
            if 'branch_link' in link_columns:
                col_num = link_columns['branch_link']
                col_letter = get_column_letter(col_num)
                
                header_cell = worksheet[f"{col_letter}1"]
                header_cell.fill = blue_fill
                header_cell.font = white_font
                
                # 🔥 設定 HYPERLINK 內容為藍色連結
                for row_num in range(2, worksheet.max_row + 1):
                    content_cell = worksheet[f"{col_letter}{row_num}"]
                    if content_cell.value and str(content_cell.value).startswith('=HYPERLINK'):
                        content_cell.font = blue_link_font
                
                # 調整欄寬
                worksheet.column_dimensions[col_letter].width = 60
                
            # 格式化 target_branch_link 欄位 (綠底白字，內容藍色連結)
            if 'target_branch_link' in link_columns:
                col_num = link_columns['target_branch_link']
                col_letter = get_column_letter(col_num)
                
                header_cell = worksheet[f"{col_letter}1"]
                header_cell.fill = green_fill
                header_cell.font = white_font
                
                # 🔥 設定 HYPERLINK 內容為藍色連結
                for row_num in range(2, worksheet.max_row + 1):
                    content_cell = worksheet[f"{col_letter}{row_num}"]
                    if content_cell.value and str(content_cell.value).startswith('=HYPERLINK'):
                        content_cell.font = green_link_font
                
                # 調整欄寬
                worksheet.column_dimensions[col_letter].width = 60

            # 🔥 格式化 target_open_project_link 欄位 (綠底白字，與 target_branch_link 一致)
            if 'target_open_project_link' in link_columns:
                col_num = link_columns['target_open_project_link']
                col_letter = get_column_letter(col_num)
                
                header_cell = worksheet[f"{col_letter}1"]
                header_cell.fill = green_fill
                header_cell.font = white_font
                
                # 🔥 設定 HYPERLINK 內容為藍色連結
                for row_num in range(2, worksheet.max_row + 1):
                    content_cell = worksheet[f"{col_letter}{row_num}"]
                    if content_cell.value and str(content_cell.value).startswith('=HYPERLINK'):
                        content_cell.font = green_link_font
                
                # 調整欄寬
                worksheet.column_dimensions[col_letter].width = 60

            # 🔥 格式化 open_project_link 欄位 (藍底白字，與 branch_link 一致)
            if 'open_project_link' in link_columns:
                col_num = link_columns['open_project_link']
                col_letter = get_column_letter(col_num)
                
                header_cell = worksheet[f"{col_letter}1"]
                header_cell.fill = blue_fill
                header_cell.font = white_font
                
                # 🔥 設定 HYPERLINK 內容為藍色連結
                for row_num in range(2, worksheet.max_row + 1):
                    content_cell = worksheet[f"{col_letter}{row_num}"]
                    if content_cell.value and str(content_cell.value).startswith('=HYPERLINK'):
                        content_cell.font = blue_link_font
                
                # 調整欄寬
                worksheet.column_dimensions[col_letter].width = 60
                                
            self.logger.info("已完成統一連結欄位格式化（支援 HYPERLINK，確保branch_link樣式一致）")
            
        except Exception as e:
            self.logger.error(f"格式化連結欄位失敗: {str(e)}")

    def _format_revision_diff_column_unified(self, worksheet, orange_fill, white_font):
        """
        🔥 修復版：格式化 revision_diff 欄位為橘底白字，N綠字/Y紅字，並置中對齊
        """
        try:
            from openpyxl.styles import Font, Alignment  # 🔥 加入 Alignment
            from openpyxl.utils import get_column_letter
            from openpyxl.formatting.rule import CellIsRule
            
            # 內容樣式
            green_font = Font(color="00B050", bold=True)  # N 的綠字
            red_font = Font(color="FF0000", bold=True)    # Y 的紅字
            
            # 🔥 新增：置中對齊設定
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # 找到 revision_diff 欄位的位置
            revision_diff_col = None
            for col_num, cell in enumerate(worksheet[1], 1):
                header_value = str(cell.value) if cell.value else ''
                if header_value == 'revision_diff':
                    revision_diff_col = col_num
                    break
            
            if revision_diff_col:
                col_letter = get_column_letter(revision_diff_col)
                
                # 🔥 格式化標題（橘底白字 + 置中）
                header_cell = worksheet[f"{col_letter}1"]
                header_cell.fill = orange_fill
                header_cell.font = white_font
                header_cell.alignment = center_alignment  # 🔥 新增：標題置中
                
                # 設定欄寬
                worksheet.column_dimensions[col_letter].width = 13.71  # 🔥 精確設定為 13.71
                
                # 🔥 新增：為所有資料欄位設定置中對齊
                for row_num in range(2, worksheet.max_row + 1):
                    cell = worksheet[f"{col_letter}{row_num}"]
                    cell.alignment = center_alignment  # 🔥 關鍵修復：資料置中
                
                # 定義資料範圍
                data_range = f"{col_letter}2:{col_letter}{worksheet.max_row}"
                
                # 條件格式規則 1: 當值為 "N" 時使用綠字（相同）
                rule_n = CellIsRule(
                    operator='equal',
                    formula=['"N"'],
                    font=green_font
                )
                worksheet.conditional_formatting.add(data_range, rule_n)
                
                # 條件格式規則 2: 當值為 "Y" 時使用紅字（不同或空值）
                rule_y = CellIsRule(
                    operator='equal',
                    formula=['"Y"'],
                    font=red_font
                )
                worksheet.conditional_formatting.add(data_range, rule_y)
                
                self.logger.info("✅ 已設定統一 revision_diff 欄位格式：標題橘底白字，N綠字/Y紅字，全部置中對齊")
                
        except Exception as e:
            self.logger.error(f"格式化 revision_diff 欄位失敗: {str(e)}")

    def _format_target_branch_columns_unified(self, worksheet, green_fill, white_font):
        """
        🔥 新方法：統一格式化目標分支相關欄位
        確保所有目標分支欄位都有綠底白字，target_branch_exists 內容置中
        """
        try:
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
            
            # 內容樣式
            green_font = Font(color="00B050", bold=True)  # Y 的綠字
            red_font = Font(color="FF0000", bold=True)    # N 的紅字
            blue_font = Font(color="0070C0", bold=True)   # Tag 的藍字
            purple_font = Font(color="7030A0", bold=True) # Branch 的紫字
            black_font = Font(color="000000")             # 一般文字
            center_alignment = Alignment(horizontal='center', vertical='center')  # 🔥 置中對齊
            
            # 🔥 所有需要綠底白字的目標分支欄位
            target_green_columns = [
                'target_branch', 'target_type', 'target_branch_exists', 
                'target_branch_link'
            ]
            
            # 找到目標欄位的位置
            target_columns = {}
            for col_num, cell in enumerate(worksheet[1], 1):
                header_value = str(cell.value) if cell.value else ''
                if header_value in target_green_columns:
                    target_columns[header_value] = col_num
            
            # 🔥 統一設定綠底白字標頭
            for col_name, col_num in target_columns.items():
                col_letter = get_column_letter(col_num)
                header_cell = worksheet[f"{col_letter}1"]
                header_cell.fill = green_fill
                header_cell.font = white_font
                
                # 🔥 根據欄位類型設定內容格式
                if col_name == 'target_type':
                    # target_type 欄位：Tag用藍字，Branch用紫字
                    for row_num in range(2, worksheet.max_row + 1):
                        cell = worksheet[f"{col_letter}{row_num}"]
                        if cell.value == 'Tag':
                            cell.font = blue_font
                        elif cell.value == 'Branch':
                            cell.font = purple_font
                        else:
                            cell.font = black_font
                            
                elif col_name == 'target_branch_exists':
                    # 🔥 target_branch_exists 欄位：Y用綠字，N用紅字，並設定置中對齊
                    for row_num in range(2, worksheet.max_row + 1):
                        cell = worksheet[f"{col_letter}{row_num}"]
                        cell.alignment = center_alignment  # 🔥 設定置中對齊
                        if cell.value == 'Y':
                            cell.font = green_font
                        elif cell.value == 'N':
                            cell.font = red_font
                        else:
                            cell.font = black_font
                            
                elif col_name == 'target_branch_link':
                    # target_branch_link 欄位：設定較寬欄寬
                    worksheet.column_dimensions[col_letter].width = 60
                    
                else:
                    # 其他欄位：使用黑字
                    for row_num in range(2, worksheet.max_row + 1):
                        cell = worksheet[f"{col_letter}{row_num}"]
                        cell.font = black_font
            
            self.logger.info("已設定統一的目標分支欄位格式：全部綠底白字標頭，target_branch_exists置中")
                
        except Exception as e:
            self.logger.error(f"格式化目標分支欄位失敗: {str(e)}")

    # ============================================
    # 以下其他方法保持原狀不變...
    # ============================================

    def _is_revision_hash(self, revision: str) -> bool:
        """判斷 revision 是否為 commit hash"""
        if not revision:
            return False
        
        revision = revision.strip()
        
        # Hash 特徵：40 字符的十六進制字符串
        if len(revision) == 40 and all(c in '0123456789abcdefABCDEF' for c in revision):
            return True
        
        # Branch name 特徵：包含斜線和可讀名稱
        if '/' in revision and any(c.isalpha() for c in revision):
            return False
        
        # 其他情況當作 branch name 處理
        return False

    def _get_effective_revision_for_conversion(self, project: Dict) -> str:
        """取得用於轉換的有效 revision"""
        revision = project.get('revision', '')
        upstream = project.get('upstream', '')
        remote = project.get('remote', '')
        
        # 如果 revision 是 hash，使用 upstream
        if self._is_revision_hash(revision):
            if upstream:
                self.logger.debug(f"專案 {project.get('name', '')} revision 是 hash，使用 upstream: {upstream}")
                return upstream
            else:
                self.logger.warning(f"專案 {project.get('name', '')} revision 是 hash 但沒有 upstream")
                return ''
        
        # 如果 revision 是 branch name，直接使用
        if revision:
            self.logger.debug(f"專案 {project.get('name', '')} revision 是 branch name: {revision}")
            return revision
        
        # 如果沒有 revision，返回空字串
        self.logger.debug(f"專案 {project.get('name', '')} 沒有 revision")
        return ''

    def _get_effective_revision_for_link(self, project: Dict) -> str:
        """取得用於建立連結的有效 revision"""
        return self._get_effective_revision_for_conversion(project)

    def _should_skip_revision_conversion(self, revision: str) -> bool:
        """判斷是否應該跳過 revision 轉換"""
        if not revision:
            return True
        
        # 跳過 Google 開頭的項目
        if revision.startswith('google/'):
            return True
        
        # 跳過 refs/tags/
        if revision.startswith('refs/tags/'):
            return True
        
        return False

    def _smart_conversion_fallback(self, revision: str) -> str:
        """智能轉換備案"""
        # 如果包含 mp.google-refplus，嘗試替換為 premp.google-refplus
        if 'mp.google-refplus' in revision:
            result = revision.replace('mp.google-refplus', 'premp.google-refplus')
            self.logger.debug(f"智能替換 mp→premp: {revision} → {result}")
            return result
        
        # 如果是 master 但沒有匹配到特定規則，使用預設轉換
        if '/master' in revision and 'realtek/' in revision:
            import re
            android_match = re.search(r'android-(\d+)', revision)
            if android_match:
                android_ver = android_match.group(1)
                result = f'realtek/android-{android_ver}/premp.google-refplus'
                self.logger.debug(f"智能Android版本轉換: {revision} → {result}")
                return result
            else:
                # 🔥 使用當前配置的 Android 版本而非硬編碼
                result = config.get_default_premp_branch()
                return result
        
        # 如果完全沒有匹配，返回預設值
        result = config.get_default_premp_branch()
        return result

    def _convert_master_to_premp(self, revision: str) -> str:
        """
        master → premp 轉換規則 - 使用動態 Android 版本，動態 kernel 版本匹配
        """
        if not revision:
            return revision
        
        original_revision = revision.strip()
        
        # 🆕 新增：Google wave 版本遞減轉換 (wave n → wave n-1)
        import re
        google_wave_pattern = r'google/u-tv-keystone-rtk-refplus-wave(\d+)-release'
        match = re.match(google_wave_pattern, original_revision)
        if match:
            wave_num = int(match.group(1))
            if wave_num > 1:  # 確保不會變成 wave0
                new_wave_num = wave_num - 1
                result = f'google/u-tv-keystone-rtk-refplus-wave{new_wave_num}-release'
                self.logger.debug(f"Google wave 版本遞減轉換: {original_revision} → {result}")
                return result
        
        # 跳過 Google 開頭的項目（除了上面已處理的 wave 版本）
        if original_revision.startswith('google/'):
            self.logger.debug(f"跳過 Google 項目: {original_revision}")
            return original_revision
        
        # 跳過特殊項目
        if self._should_skip_revision_conversion(original_revision):
            return original_revision
        
        # 精確匹配轉換規則 - 使用動態版本（移除預定義 kernel 版本）
        exact_mappings = {
            # 基本 master 分支轉換
            'realtek/master': config.get_default_premp_branch(),
            'realtek/gaia': config.get_default_premp_branch(),
            'realtek/gki/master': config.get_default_premp_branch(),
            
            # Android master 分支
            config.get_default_android_master_branch(): config.get_default_premp_branch(),
            
            # mp.google-refplus 轉換
            'realtek/mp.google-refplus': config.get_default_premp_branch(),
            config.get_android_path('realtek/android-{android_version}/mp.google-refplus'): config.get_default_premp_branch(),
        }
        
        # 檢查精確匹配
        if original_revision in exact_mappings:
            result = exact_mappings[original_revision]
            self.logger.debug(f"精確匹配轉換: {original_revision} → {result}")
            return result
        
        # 模式匹配轉換規則 - 完全使用正則表達式動態匹配
        import re
        
        # vX.X.X 版本轉換 - 保留版本號
        pattern_version = r'realtek/(v\d+\.\d+(?:\.\d+)?)/master$'
        match_version = re.match(pattern_version, original_revision)
        if match_version:
            version = match_version.group(1)
            result = f'realtek/{version}/premp.google-refplus'
            self.logger.debug(f"版本格式轉換: {original_revision} → {result}")
            return result
        
        # 新增規則: vX.X.X/mp.google-refplus → vX.X.X/premp.google-refplus.upgrade-{prev_version}
        pattern_version_mp = r'realtek/(v\d+\.\d+(?:\.\d+)?)/mp\.google-refplus$'
        match_version_mp = re.match(pattern_version_mp, original_revision)
        if match_version_mp:
            version = match_version_mp.group(1)
            upgrade_ver = config.get_current_android_prev_version()
            result = f'realtek/{version}/premp.google-refplus.upgrade-{upgrade_ver}'
            self.logger.debug(f"版本 mp 格式轉換: {original_revision} → {result}")
            return result
        
        # 規則 1: mp.google-refplus.upgrade-11.rtdXXXX → premp.google-refplus.upgrade-11.rtdXXXX
        pattern1 = r'realtek/android-(\d+)/mp\.google-refplus\.upgrade-(\d+)\.(rtd\w+)'
        match1 = re.match(pattern1, original_revision)
        if match1:
            android_ver, upgrade_ver, rtd_chip = match1.groups()
            if android_ver == config.get_current_android_version():
                result = config.get_premp_branch_with_upgrade(upgrade_ver, rtd_chip)
            else:
                # 如果是不同的 Android 版本，保持原版本
                result = f'realtek/android-{android_ver}/premp.google-refplus.upgrade-{upgrade_ver}.{rtd_chip}'
            self.logger.debug(f"模式1轉換: {original_revision} → {result}")
            return result
        
        # 規則 2: mp.google-refplus.upgrade-11 → premp.google-refplus.upgrade-11
        pattern2 = r'realtek/android-(\d+)/mp\.google-refplus\.upgrade-(\d+)$'
        match2 = re.match(pattern2, original_revision)
        if match2:
            android_ver, upgrade_ver = match2.groups()
            if android_ver == config.get_current_android_version():
                result = config.get_premp_branch_with_upgrade(upgrade_ver)
            else:
                result = f'realtek/android-{android_ver}/premp.google-refplus.upgrade-{upgrade_ver}'
            self.logger.debug(f"模式2轉換: {original_revision} → {result}")
            return result
        
        # 規則 3: linux-X.X/master → linux-X.X/android-{current_version}/premp.google-refplus（完全動態）
        pattern3 = r'realtek/linux-([\d.]+)/master$'
        match3 = re.match(pattern3, original_revision)
        if match3:
            linux_ver = match3.group(1)
            result = config.get_linux_android_path(
                linux_ver, 'realtek/linux-{linux_ver}/android-{android_version}/premp.google-refplus'
            )
            self.logger.debug(f"模式3轉換（動態 kernel 版本）: {original_revision} → {result}")
            return result
        
        # 規則 4: linux-X.X/android-Y/master → linux-X.X/android-{current_version}/premp.google-refplus（完全動態）
        pattern4 = r'realtek/linux-([\d.]+)/android-(\d+)/master$'
        match4 = re.match(pattern4, original_revision)
        if match4:
            linux_ver, android_ver = match4.groups()
            # 自動升級到當前 Android 版本
            result = config.get_linux_android_path(
                linux_ver, 'realtek/linux-{linux_ver}/android-{android_version}/premp.google-refplus'
            )
            self.logger.debug(f"模式4轉換（動態 kernel，升級 Android）: {original_revision} → {result}")
            return result
        
        # 規則 5: linux-X.X/android-Y/mp.google-refplus → linux-X.X/android-{current_version}/premp.google-refplus（完全動態）
        pattern5 = r'realtek/linux-([\d.]+)/android-(\d+)/mp\.google-refplus$'
        match5 = re.match(pattern5, original_revision)
        if match5:
            linux_ver, android_ver = match5.groups()
            result = config.get_linux_android_path(
                linux_ver, 'realtek/linux-{linux_ver}/android-{android_version}/premp.google-refplus'
            )
            self.logger.debug(f"模式5轉換（動態 kernel）: {original_revision} → {result}")
            return result
        
        # 規則 6: linux-X.X/android-Y/mp.google-refplus.rtdXXXX → linux-X.X/android-{current_version}/premp.google-refplus.rtdXXXX（完全動態）
        pattern6 = r'realtek/linux-([\d.]+)/android-(\d+)/mp\.google-refplus\.(rtd\w+)'
        match6 = re.match(pattern6, original_revision)
        if match6:
            linux_ver, android_ver, rtd_chip = match6.groups()
            base_path = config.get_linux_android_path(
                linux_ver, 'realtek/linux-{linux_ver}/android-{android_version}/premp.google-refplus'
            )
            result = f"{base_path}.{rtd_chip}"
            self.logger.debug(f"模式6轉換（動態 kernel）: {original_revision} → {result}")
            return result
        
        # 規則 7: android-Y/mp.google-refplus → android-{current_version}/premp.google-refplus
        pattern7 = r'realtek/android-(\d+)/mp\.google-refplus$'
        match7 = re.match(pattern7, original_revision)
        if match7:
            android_ver = match7.group(1)
            result = config.get_default_premp_branch()
            self.logger.debug(f"模式7轉換（升級到當前版本）: {original_revision} → {result}")
            return result
        
        # 規則 8: android-Y/mp.google-refplus.rtdXXXX → android-{current_version}/premp.google-refplus.rtdXXXX
        pattern8 = r'realtek/android-(\d+)/mp\.google-refplus\.(rtd\w+)'
        match8 = re.match(pattern8, original_revision)
        if match8:
            android_ver, rtd_chip = match8.groups()
            result = config.get_premp_branch_with_chip(rtd_chip)
            self.logger.debug(f"模式8轉換（升級到當前版本）: {original_revision} → {result}")
            return result
        
        # 規則 9: 晶片特定的 master 分支 → premp.google-refplus.rtdXXXX（使用當前 Android 版本）
        for chip, rtd_model in config.CHIP_TO_RTD_MAPPING.items():
            self.logger.debug(f"檢查晶片規則: {chip} -> {rtd_model}")
            if f'realtek/{chip}/master' == original_revision:
                result = config.get_premp_branch_with_chip(rtd_model)
                self.logger.info(f"晶片轉換匹配: {original_revision} → {result}")
                return result
            else:
                self.logger.debug(f"不匹配: 'realtek/{chip}/master' != '{original_revision}'")
        
        # 智能轉換備案
        smart_result = self._smart_conversion_fallback(original_revision)
        self.logger.debug(f"智能轉換: {original_revision} → {smart_result}")
        return smart_result  

    def _convert_premp_to_mp(self, revision: str) -> str:
        """premp → mp 轉換規則"""
        return revision.replace('premp.google-refplus', 'mp.google-refplus.wave')
    
    def _convert_mp_to_mpbackup(self, revision: str) -> str:
        """mp → mpbackup 轉換規則"""
        return revision.replace('mp.google-refplus.wave', 'mp.google-refplus.wave.backup')

    def _convert_master_to_wave(self, revision: str) -> str:
        """
        master → wave 轉換規則
        相當於 master → premp → wave 的鏈式轉換
        """
        if not revision:
            return revision
        
        # 第一步：master → premp
        premp_revision = self._convert_master_to_premp(revision)
        
        # 第二步：premp → wave
        wave_revision = self._convert_premp_to_mp(premp_revision)
        
        return wave_revision

    def _convert_master_to_wavebackup(self, revision: str) -> str:
        """
        master → wave.backup 轉換規則
        相當於 master → premp → wave → wave.backup 的鏈式轉換
        """
        if not revision:
            return revision
        
        # 第一步：master → premp
        premp_revision = self._convert_master_to_premp(revision)
        
        # 第二步：premp → wave  
        wave_revision = self._convert_premp_to_mp(premp_revision)
        
        # 第三步：wave → wave.backup
        backup_revision = self._convert_mp_to_mpbackup(wave_revision)
        
        return backup_revision
        
    def _convert_projects(self, projects: List[Dict], process_type: str, check_branch_exists: bool = False, 
                 source_manifest_name: str = '', is_tvconfig: bool = False) -> List[Dict]:
        """
        轉換專案的分支名稱 - 修正版（🔥 新增跳過邏輯和 tvconfig 支援 + title 查詢）
        """
        # 🔥 新增：設置實例變量供自定義轉換規則使用
        self._current_projects = projects
        import copy  # 加入這行
        converted_projects = []
        tag_count = 0
        branch_count = 0
        hash_revision_count = 0
        branch_revision_count = 0
        branch_revision_query_count = 0  # 記錄查詢 branch revision 的次數
        skipped_projects_count = 0  # 🔥 新增：跳過的專案計數
        title_query_count = 0  # 🔥 新增：記錄 title 查詢次數
        target_title_query_count = 0  # 🔥 新增：記錄 target_title 查詢次數
        
        self.logger.info(f"🔄 開始轉換專案分支，處理類型: {process_type}")
        if is_tvconfig:
            self.logger.info(f"🎯 Tvconfig 模式：使用 TVCONFIG_SKIP_PROJECTS 配置")
        else:
            self.logger.info(f"🎯 一般模式：使用 FEATURE_TWO_SKIP_PROJECTS 配置")
        
        for i, project in enumerate(projects, 1):
            converted_project = copy.deepcopy(project)
            converted_project['SN'] = i
            
            # 🔥 取得專案名稱
            project_name = project.get('name', '')
            
            # 🔥 新增 source_manifest 欄位
            converted_project['source_manifest'] = source_manifest_name
            
            # 🔥 只在沒有 remote 時才自動偵測，否則保留原始值
            original_remote = project.get('remote', '')
            if not original_remote:
                auto_remote = self._auto_detect_remote(project)
                converted_project['remote'] = auto_remote
                self.logger.debug(f"專案 {project_name} 自動偵測 remote: {auto_remote}")
            else:
                converted_project['remote'] = original_remote
                self.logger.debug(f"專案 {project_name} 保留原始 remote: {original_remote}")
            
            # 使用新邏輯取得用於轉換的 revision
            effective_revision = self._get_effective_revision_for_conversion(converted_project)
            
            # 統計 revision 類型
            original_revision = project.get('revision', '')
            if self._is_revision_hash(original_revision):
                hash_revision_count += 1
            elif original_revision:
                branch_revision_count += 1
            
            # 🔥 新增：如果 original_revision 不是 hash，查詢對應的 branch revision
            branch_revision_value = self._get_branch_revision_if_needed(
                project_name, original_revision, converted_project['remote']
            )
            converted_project['branch_revision'] = branch_revision_value
            
            if branch_revision_value and branch_revision_value != '-':
                branch_revision_query_count += 1
            
            # 如果沒有有效的 revision，跳過轉換
            if not effective_revision:
                target_branch = ''
                self.logger.debug(f"專案 {project_name} 沒有有效的 revision，跳過轉換")
            else:
                # 🔥 修改：根據處理類型進行轉換，傳遞專案名稱和 tvconfig 標記
                target_branch = self._convert_revision_by_type(
                    effective_revision, process_type, project_name, is_tvconfig
                )
                
                # 🔥 檢查是否發生了轉換
                if target_branch != effective_revision:
                    # 🔥 檢查是否被跳過
                    if self._should_skip_project_conversion(project_name, process_type, is_tvconfig):
                        skipped_projects_count += 1
                        self.logger.debug(f"專案 {project_name} 已跳過轉換：{effective_revision} (保持不變)")
                    else:
                        self.logger.debug(f"專案 {project_name} 轉換: {effective_revision} → {target_branch}")
            
            converted_project['target_branch'] = target_branch
            converted_project['effective_revision'] = effective_revision
            
            # 判斷目標是 Tag 還是 Branch
            is_tag = self._is_tag_reference(target_branch)
            converted_project['target_type'] = 'Tag' if is_tag else 'Branch'
            
            if is_tag:
                tag_count += 1
            else:
                branch_count += 1
            
            # 🔥 修正：根據參數決定是否檢查存在性，使用最終確定的 remote
            if check_branch_exists and target_branch:
                final_remote = converted_project['remote']
                
                if is_tag:
                    exists_info = self._check_target_tag_exists(project_name, target_branch, final_remote)
                else:
                    # 🔥 修正：直接傳入確定的 remote，不再測試兩種可能性
                    exists_info = self._check_target_branch_exists(project_name, target_branch, final_remote)
                
                converted_project['target_branch_exists'] = exists_info['exists_status']
                converted_project['target_branch_revision'] = exists_info['revision']
                
                # 🔥 記錄分支檢查結果
                if exists_info['exists_status'] == 'Y':
                    self.logger.debug(f"✅ 專案 {project_name} 分支檢查成功:")
                    self.logger.debug(f"  目標分支: {target_branch}")
                    self.logger.debug(f"  使用 remote: {final_remote}")
                    self.logger.debug(f"  分支 revision: {exists_info['revision']}")
                else:
                    self.logger.debug(f"❌ 專案 {project_name} 分支檢查失敗:")
                    self.logger.debug(f"  目標分支: {target_branch}")
                    self.logger.debug(f"  使用 remote: {final_remote}")
                    
            else:
                converted_project['target_branch_exists'] = '-'
                converted_project['target_branch_revision'] = '-'
            
            # 🔥 新增：查詢 commit titles
            final_remote = converted_project['remote']
            
            # 查詢 branch_revision 的 title
            branch_revision_for_title = converted_project.get('branch_revision', '-')
            if branch_revision_for_title and branch_revision_for_title != '-':
                title = self._get_commit_title(project_name, branch_revision_for_title, final_remote)
                converted_project['title'] = title
                if title != '-':
                    title_query_count += 1
                    self.logger.debug(f"✅ 查詢到 title: {project_name}/{branch_revision_for_title[:8]} -> {title[:30]}...")
            else:
                converted_project['title'] = '-'
                
            # 查詢 target_branch_revision 的 target_title
            target_branch_revision_for_title = converted_project.get('target_branch_revision', '-')
            if target_branch_revision_for_title and target_branch_revision_for_title != '-':
                target_title = self._get_commit_title(project_name, target_branch_revision_for_title, final_remote)
                converted_project['target_title'] = target_title
                if target_title != '-':
                    target_title_query_count += 1
                    self.logger.debug(f"✅ 查詢到 target_title: {project_name}/{target_branch_revision_for_title[:8]} -> {target_title[:30]}...")
            else:
                converted_project['target_title'] = '-'
            
            converted_projects.append(converted_project)
            
            # 每100個項目顯示進度
            if check_branch_exists and i % 100 == 0:
                self.logger.info(f"已處理 {i}/{len(projects)} 個專案的存在性檢查")
        
        self.logger.info(f"轉換完成 - Branch: {branch_count}, Tag: {tag_count}")
        if skipped_projects_count > 0:
            self.logger.info(f"🚫 跳過轉換的專案: {skipped_projects_count} 個")
        
        self.logger.info(f"📊 Revision 類型統計:")
        self.logger.info(f"  - 🔸 Hash revision: {hash_revision_count} 個")
        self.logger.info(f"  - 🔹 Branch revision: {branch_revision_count} 個")
        self.logger.info(f"  - 🔍 Branch revision 查詢: {branch_revision_query_count} 個")
        
        # 🔥 新增：Title 查詢統計
        self.logger.info(f"📊 Title 查詢統計:")
        self.logger.info(f"  - 📝 Title 查詢成功: {title_query_count} 個")
        self.logger.info(f"  - 📝 Target Title 查詢成功: {target_title_query_count} 個")
        
        # 🔥 統計 remote 分布
        remote_stats = {}
        auto_detected_count = 0
        
        for proj in converted_projects:
            remote = proj.get('remote', 'unknown')
            remote_stats[remote] = remote_stats.get(remote, 0) + 1
            
            # 統計自動偵測的數量
            original_project = next((p for p in projects if p.get('name') == proj.get('name', '')), {})
            if not original_project.get('remote', ''):
                auto_detected_count += 1
        
        self.logger.info(f"📊 Remote 分布統計:")
        for remote, count in remote_stats.items():
            self.logger.info(f"  - {remote}: {count} 個專案")
        
        if auto_detected_count > 0:
            self.logger.info(f"📊 自動偵測 remote: {auto_detected_count} 個專案")
        
        # 🔥 分支檢查統計
        if check_branch_exists:
            branch_check_stats = {'Y': 0, 'N': 0, '-': 0}
            for proj in converted_projects:
                status = proj.get('target_branch_exists', '-')
                branch_check_stats[status] = branch_check_stats.get(status, 0) + 1
            
            self.logger.info(f"📊 分支檢查統計:")
            self.logger.info(f"  - ✅ 分支存在: {branch_check_stats['Y']} 個")
            self.logger.info(f"  - ❌ 分支不存在: {branch_check_stats['N']} 個")
            self.logger.info(f"  - ⏭️ 未檢查: {branch_check_stats['-']} 個")
        
        return converted_projects

    def _get_commit_title(self, project_name: str, commit_hash: str, remote: str = '') -> str:
        """
        🔥 修改方法：查詢 gerrit commit 的 title - 使用 GerritManager 的新方法
        
        Args:
            project_name: 專案名稱
            commit_hash: commit hash
            remote: remote 類型
            
        Returns:
            commit title 或 '-' (如果查詢失敗)
        """
        try:
            if not project_name or not commit_hash or commit_hash == '-':
                return '-'
            
            # 如果不是有效的 hash，直接返回
            if not self._is_revision_hash(commit_hash):
                self.logger.debug(f"跳過非 hash commit title 查詢: {project_name}/{commit_hash}")
                return '-'
            
            # 🔥 修改：根據 remote 選擇正確的 GerritManager 實例，然後使用其 get_commit_title 方法
            if remote == 'rtk-prebuilt':
                temp_gerrit = self._get_prebuilt_gerrit_manager()
            else:
                temp_gerrit = self.gerrit_manager
            
            # 🔥 使用 GerritManager 的新方法
            title = temp_gerrit.get_commit_title(project_name, commit_hash)
            
            if title:
                self.logger.debug(f"✅ 查詢到 commit title: {project_name}/{commit_hash[:8]} -> {title[:50]}...")
                return title
            else:
                self.logger.debug(f"❌ 無法查詢到 commit title: {project_name}/{commit_hash[:8]}")
                return '-'
                
        except Exception as e:
            self.logger.debug(f"❌ 查詢 commit title 異常: {project_name}/{commit_hash[:8] if commit_hash else 'N/A'} - {str(e)}")
            return '-'

    def _get_commit_title_batch(self, commit_requests: list) -> dict:
        """
        🔥 修改方法：批量查詢 commit titles - 使用 GerritManager 的批量方法
        
        Args:
            commit_requests: [(project_name, commit_hash, remote), ...] 的列表
            
        Returns:
            {(project_name, commit_hash): title} 的字典
        """
        results = {}
        
        try:
            # 🔥 按 remote 分組，使用對應的 GerritManager 實例
            rtk_requests = []
            prebuilt_requests = []
            
            for project_name, commit_hash, remote in commit_requests:
                if remote == 'rtk-prebuilt':
                    prebuilt_requests.append((project_name, commit_hash))
                else:
                    rtk_requests.append((project_name, commit_hash))
            
            # 🔥 處理 rtk 請求
            if rtk_requests:
                rtk_results = self.gerrit_manager.batch_get_commit_titles(rtk_requests)
                for (project_name, commit_hash), title in rtk_results.items():
                    key = (project_name, commit_hash)
                    results[key] = title if title else '-'
            
            # 🔥 處理 rtk-prebuilt 請求  
            if prebuilt_requests:
                temp_gerrit = self._get_prebuilt_gerrit_manager()
                prebuilt_results = temp_gerrit.batch_get_commit_titles(prebuilt_requests)
                for (project_name, commit_hash), title in prebuilt_results.items():
                    key = (project_name, commit_hash)
                    results[key] = title if title else '-'
            
            return results
            
        except Exception as e:
            self.logger.error(f"批量查詢 commit titles 失敗: {str(e)}")
            # 如果批量失敗，回退到單個查詢
            for project_name, commit_hash, remote in commit_requests:
                key = (project_name, commit_hash)
                results[key] = self._get_commit_title(project_name, commit_hash, remote)
            
            return results
            
    def _get_branch_revision_if_needed(self, project_name: str, revision: str, remote: str = '') -> str:
        """
        取得 revision 對應的實際 hash 值
        
        Args:
            project_name: 專案名稱
            revision: 原始 revision
            remote: remote 類型
            
        Returns:
            如果 revision 是 hash，返回 revision 本身
            如果 revision 是 branch，返回查詢到的 hash 或 '-'（如果查詢失敗）
        """
        try:
            if not project_name or not revision:
                return '-'
            
            # 如果 revision 已經是 hash，直接返回它的值
            if self._is_revision_hash(revision):
                self.logger.debug(f"專案 {project_name} revision 已是 hash，直接使用: {revision[:8]}...")
                return revision
            
            # 如果是 branch name，查詢對應的 hash
            self.logger.debug(f"專案 {project_name} revision 是 branch，查詢實際 hash: {revision}")

            # 使用增強版查詢方法
            branch_info = self._query_branch_direct_enhanced(project_name, revision, remote)
            
            if branch_info['exists'] and branch_info['revision']:
                actual_hash = branch_info['revision']
                self.logger.debug(f"查詢到 {project_name}/{revision} 的實際 hash: {actual_hash[:8]}...")
                return actual_hash
            else:
                error_msg = branch_info.get('error', '未知錯誤')
                self.logger.debug(f"無法查詢 {project_name}/{revision} 的 hash: {error_msg}")
                return '-'
                    
        except Exception as e:
            self.logger.debug(f"查詢 {project_name}/{revision} branch revision 失敗: {str(e)}")
            return '-'
            
    def _convert_revision_by_type(self, revision: str, process_type: str, project_name: str = '', is_tvconfig: bool = False) -> str:
        """根據處理類型轉換 revision - 修正版：正確的處理類型 + 自定義轉換規則"""
        try:
            if not revision:
                return ''
            
            # 🔥 檢查是否應該跳過轉換
            if project_name and self._should_skip_project_conversion(project_name, process_type, is_tvconfig):
                self.logger.debug(f"跳過專案 {project_name} 的轉換，保持原 revision: {revision}")
                return revision
            
            # 🆕 新增：檢查自定義轉換規則（支援陣列格式）
            if project_name:
                import re
                
                # 🔥 使用 FEATURE_TWO_CUSTOM_CONVERSIONS（與 feature_three.py 分開）
                custom_rules = getattr(config, 'FEATURE_TWO_CUSTOM_CONVERSIONS', {}).get(process_type, {})
                
                # 檢查是否有匹配的規則
                for pattern, rule_config in custom_rules.items():
                    try:
                        # 先檢查 name 是否匹配
                        name_matches = False
                        try:
                            name_matches = bool(re.search(pattern, project_name))
                        except re.error:
                            name_matches = pattern in project_name
                        
                        if not name_matches:
                            continue
                        
                        # 🆕 支援三種配置格式
                        if isinstance(rule_config, list):
                            # 陣列格式：同一個 name pattern 對應多個規則
                            for rule_item in rule_config:
                                if not isinstance(rule_item, dict):
                                    continue
                                    
                                target_branch = rule_item.get('target', '')
                                path_pattern = rule_item.get('path_pattern', '')
                                
                                if not target_branch:
                                    continue
                                
                                # 檢查 path 條件
                                if path_pattern:
                                    project_path = self._get_project_path_for_conversion(project_name, process_type)
                                    if not project_path:
                                        continue
                                    
                                    # 檢查 path 是否匹配
                                    path_matches = False
                                    try:
                                        path_matches = bool(re.search(path_pattern, project_path))
                                    except re.error:
                                        path_matches = path_pattern in project_path
                                    
                                    if not path_matches:
                                        continue
                                    
                                    self.logger.info(f"🎯 Feature Two 使用自定義轉換規則（陣列格式 - name + path）: {project_name}")
                                    self.logger.info(f"   name 模式: '{pattern}' ✓")
                                    self.logger.info(f"   path 模式: '{path_pattern}' ✓ (path: {project_path})")
                                    self.logger.info(f"   目標: '{target_branch}'")
                                    return target_branch
                                else:
                                    # 沒有 path 限制，直接使用
                                    self.logger.info(f"🎯 Feature Two 使用自定義轉換規則（陣列格式 - 僅 name）: {project_name}")
                                    self.logger.info(f"   name 模式: '{pattern}' ✓")
                                    self.logger.info(f"   目標: '{target_branch}'")
                                    return target_branch
                                    
                        elif isinstance(rule_config, dict):
                            # 字典格式：單一規則
                            target_branch = rule_config.get('target', '')
                            path_pattern = rule_config.get('path_pattern', '')
                            
                            if not target_branch:
                                continue
                            
                            if path_pattern:
                                project_path = self._get_project_path_for_conversion(project_name, process_type)
                                if not project_path:
                                    continue
                                
                                path_matches = False
                                try:
                                    path_matches = bool(re.search(path_pattern, project_path))
                                except re.error:
                                    path_matches = path_pattern in project_path
                                
                                if not path_matches:
                                    continue
                                
                                self.logger.info(f"🎯 Feature Two 使用自定義轉換規則（字典格式 - name + path）: {project_name}")
                                self.logger.info(f"   name 模式: '{pattern}' ✓")
                                self.logger.info(f"   path 模式: '{path_pattern}' ✓ (path: {project_path})")
                                self.logger.info(f"   目標: '{target_branch}'")
                            else:
                                self.logger.info(f"🎯 Feature Two 使用自定義轉換規則（字典格式 - 僅 name）: {project_name}")
                                self.logger.info(f"   name 模式: '{pattern}' ✓")
                                self.logger.info(f"   目標: '{target_branch}'")
                            
                            return target_branch
                            
                        else:
                            # 簡單格式：直接是 target branch 字符串
                            target_branch = str(rule_config)
                            self.logger.info(f"🎯 Feature Two 使用自定義轉換規則（簡單格式）: {project_name}")
                            self.logger.info(f"   模式: '{pattern}' → 目標: '{target_branch}'")
                            return target_branch
                            
                    except Exception as e:
                        self.logger.error(f"處理自定義轉換規則 '{pattern}' 時發生錯誤: {str(e)}")
                        continue
            
            # 🔥 如果是 Tag 參考，直接返回不做轉換（原本邏輯）
            if self._is_tag_reference(revision):
                self.logger.debug(f"檢測到 Tag 參考，保持原樣: {revision}")
                return revision
            
            # 🔥 標準轉換邏輯（完全保持原樣，不影響現有功能）
            if process_type == 'master_vs_premp':  # 原始功能
                return self._convert_master_to_premp(revision)
            elif process_type == 'premp_vs_mp':  # 原始功能
                return self._convert_premp_to_mp(revision)
            elif process_type == 'mp_vs_mpbackup':  # 原始功能
                return self._convert_mp_to_mpbackup(revision)
            # tvconfig 功能的轉換類型
            elif process_type == 'master_to_premp':  # tvconfig
                return self._convert_master_to_premp(revision)
            elif process_type == 'master_to_mp':  # tvconfig
                return self._convert_master_to_wave(revision)
            elif process_type == 'master_to_mpbackup':  # tvconfig
                return self._convert_master_to_wavebackup(revision)
            
            # 如果沒有匹配的處理類型，返回原值
            return revision
            
        except Exception as e:
            self.logger.error(f"轉換 revision 失敗: {revision}, 錯誤: {str(e)}")
            return revision

    def _should_skip_project_conversion(self, project_name: str, process_type: str, is_tvconfig: bool = False) -> bool:
        """
        檢查專案是否應該跳過轉換（修復正則表達式支援）
        
        Args:
            project_name: 專案名稱
            process_type: 處理類型
            is_tvconfig: 是否為 tvconfig 轉換
            
        Returns:
            是否應該跳過轉換
        """
        try:
            import re
            
            # 選擇對應的跳過配置
            if is_tvconfig:
                skip_config = getattr(config, 'TVCONFIG_SKIP_PROJECTS', {})
            else:
                skip_config = getattr(config, 'FEATURE_TWO_SKIP_PROJECTS', {})
            
            # 取得該處理類型的跳過專案列表
            skip_projects = skip_config.get(process_type, [])
            
            if not skip_projects:
                return False
            
            # 檢查專案名稱是否在跳過列表中
            for skip_pattern in skip_projects:
                try:
                    # 🔥 修復：首先嘗試正則表達式匹配
                    if re.search(skip_pattern, project_name):
                        context = "tvconfig" if is_tvconfig else "Feature Two"
                        self.logger.info(f"🚫 {context} 跳過轉換專案: {project_name} (正則匹配: {skip_pattern})")
                        return True
                except re.error as regex_error:
                    # 🔥 如果正則表達式無效，回退到字串包含檢查
                    self.logger.debug(f"正則表達式 '{skip_pattern}' 無效: {str(regex_error)}，回退到字串匹配")
                    if skip_pattern in project_name:
                        context = "tvconfig" if is_tvconfig else "Feature Two"
                        self.logger.info(f"🚫 {context} 跳過轉換專案: {project_name} (字串匹配: {skip_pattern})")
                        return True
                except Exception as match_error:
                    # 🔥 其他匹配錯誤，回退到字串包含檢查
                    self.logger.debug(f"匹配模式 '{skip_pattern}' 時發生錯誤: {str(match_error)}，回退到字串匹配")
                    if skip_pattern in project_name:
                        context = "tvconfig" if is_tvconfig else "Feature Two"
                        self.logger.info(f"🚫 {context} 跳過轉換專案: {project_name} (字串匹配: {skip_pattern})")
                        return True
            
            return False
            
        except Exception as e:
            self.logger.warning(f"檢查跳過專案失敗: {str(e)}")
            return False
            
    def _get_gerrit_base_url(self, remote: str) -> str:
        """根據 remote 取得對應的 Gerrit base URL"""
        try:
            if remote == 'rtk-prebuilt':
                return getattr(config, 'GERRIT_PREBUILT_URL', 'https://mm2sd-git2.rtkbf.com')
            else:
                return getattr(config, 'GERRIT_SORUCE_URL', 'https://mm2sd.rtkbf.com')
        except:
            if remote == 'rtk-prebuilt':
                return 'https://mm2sd-git2.rtkbf.com'
            else:
                return 'https://mm2sd.rtkbf.com'
    
    def _build_gerrit_link_from_dest_branch(self, project_name: str, dest_branch: str, remote: str = '') -> str:
        """根據 dest-branch 建立 Gerrit branch/tag 連結"""
        try:
            if not project_name or not dest_branch:
                return ""
            
            # 根據 remote 決定 base URL
            gerrit_base = self._get_gerrit_base_url(remote)
            base_url = f"{gerrit_base}/gerrit/plugins/gitiles"
            
            # 判斷是 tag 還是 branch
            if dest_branch.startswith('refs/tags/'):
                link = f"{base_url}/{project_name}/+/{dest_branch}"
            elif dest_branch.startswith('refs/heads/'):
                link = f"{base_url}/{project_name}/+/{dest_branch}"
            else:
                link = f"{base_url}/{project_name}/+/refs/heads/{dest_branch}"
            
            self.logger.debug(f"建立 branch_link: {project_name} + {dest_branch} -> {link} (remote: {remote})")
            return link
            
        except Exception as e:
            self.logger.error(f"建立 Gerrit 連結失敗 {project_name}: {str(e)}")
            return ""
    
    def _build_gerrit_link(self, project_name: str, revision: str, target_type: str, remote: str = '') -> str:
        """建立 Gerrit branch/tag 連結 - 🔥 修復：顯示文字使用完整 URL"""
        try:
            if not project_name or not revision:
                return ""
            
            # 根據 remote 決定 base URL
            gerrit_base = self._get_gerrit_base_url(remote)
            base_url = f"{gerrit_base}/gerrit/plugins/gitiles"
            
            # 處理 refs/tags/ 或 refs/heads/ 前綴
            clean_revision = revision
            if revision.startswith('refs/tags/'):
                clean_revision = revision[10:]
                target_type = 'tag'
            elif revision.startswith('refs/heads/'):
                clean_revision = revision[11:]
                target_type = 'branch'
            
            if target_type.lower() == 'tag':
                link_url = f"{base_url}/{project_name}/+/refs/tags/{clean_revision}"
            else:
                link_url = f"{base_url}/{project_name}/+/refs/heads/{clean_revision}"
            
            # 🔥 修復：顯示文字直接使用完整 URL，不再使用簡化的顯示名稱
            hyperlink = f'=HYPERLINK("{link_url}","{link_url}")'
            
            self.logger.debug(f"建立 {target_type} HYPERLINK: {project_name} -> 顯示完整URL (remote: {remote})")
            return hyperlink
            
        except Exception as e:
            self.logger.error(f"建立 Gerrit 連結失敗 {project_name}: {str(e)}")
            return ""
    
    def _determine_revision_type(self, revision: str) -> str:
        """判斷 revision 是 branch 還是 tag - 簡化版：只有 refs/tags/ 開頭才是 tag"""
        if not revision:
            return 'Branch'
        
        # 只有以 refs/tags/ 開頭的才是 Tag，其他都是 Branch
        if revision.startswith('refs/tags/'):
            return 'Tag'
        
        return 'Branch'

    def _renumber_projects(self, projects: List[Dict]) -> List[Dict]:
        """重新編號專案列表的 SN"""
        for i, project in enumerate(projects, 1):
            project['SN'] = i
        return projects
    
    def _add_links_to_projects(self, projects: List[Dict]) -> List[Dict]:
        """為專案添加 branch/tag 連結資訊 - 修正版：處理 open_project_link 為空的問題"""
        projects_with_links = []
        
        revision_count = 0
        dest_branch_count = 0
        hash_revision_count = 0
        branch_revision_count = 0
        upstream_used_count = 0
        
        for project in projects:
            enhanced_project = project.copy()
            
            project_name = project.get('name', '')
            remote = project.get('remote', '')
            
            # 統計原始資料
            revision = project.get('revision', '')
            dest_branch = project.get('dest-branch', '')
            upstream = project.get('upstream', '')
            
            if revision:
                revision_count += 1
                if self._is_revision_hash(revision):
                    hash_revision_count += 1
                else:
                    branch_revision_count += 1
            if dest_branch:
                dest_branch_count += 1
            
            # 使用新邏輯取得用於建立連結的 revision
            link_revision = self._get_effective_revision_for_link(project)
            
            # 記錄是否使用了 upstream
            if self._is_revision_hash(revision) and upstream:
                upstream_used_count += 1
            
            # 建立 branch_link
            if link_revision:
                revision_type = self._determine_revision_type(link_revision)
                branch_link = self._build_gerrit_link(project_name, link_revision, revision_type, remote)
                self.logger.debug(f"為專案 {project_name} 建立 branch_link: {link_revision} -> {revision_type} -> {branch_link[:50]}...")
            else:
                branch_link = ""
                self.logger.debug(f"專案 {project_name} 沒有有效 revision，branch_link 為空")
            
            # 目標 branch 資訊
            target_branch = project.get('target_branch', '')
            target_type = project.get('target_type', 'Branch')
            
            # 建立 target_branch_link
            target_branch_link = self._build_gerrit_link(project_name, target_branch, target_type, remote)
            
            # 建立 target_manifest 連結
            target_manifest = self._build_target_manifest_link(target_branch, remote)

            # 建立 target_open_project_link（使用 target_branch）
            target_open_project_link = self._build_open_project_link(project_name, target_branch, remote, is_target=True)
            
            # 🔥 修正 open_project_link：使用 fallback 邏輯
            open_project_link = self._build_open_project_link_with_fallback(project_name, project, remote)

            # revision_diff 欄位將使用 Excel 公式
            revision_diff = ''
            
            # 添加所有欄位
            enhanced_project['branch_link'] = branch_link
            enhanced_project['target_open_project_link'] = target_open_project_link
            enhanced_project['open_project_link'] = open_project_link  
            enhanced_project['target_branch_link'] = target_branch_link
            enhanced_project['target_manifest'] = target_manifest
            
            projects_with_links.append(enhanced_project)
        
        self.logger.info(f"已為 {len(projects_with_links)} 個專案添加連結資訊")
        self.logger.info(f"連結 branch_link 邏輯: Hash revision 使用 upstream，Branch revision 使用 revision")
        self.logger.info(f"欄位統計:")
        self.logger.info(f"  - revision 欄位有值: {revision_count}")
        self.logger.info(f"  - dest-branch 欄位有值: {dest_branch_count}")
        self.logger.info(f"  - Hash revision: {hash_revision_count}")
        self.logger.info(f"  - Branch revision: {branch_revision_count}")
        self.logger.info(f"  - 使用 upstream 建立連結: {upstream_used_count}")
        
        return projects_with_links

    def _build_open_project_link_with_fallback(self, project_name: str, project: Dict, remote: str = '') -> str:
        """
        建立 open_project_link - 修正版：使用 fallback 邏輯處理空 dest-branch
        
        Args:
            project_name: 專案名稱
            project: 專案字典（包含所有欄位）
            remote: remote 類型
            
        Returns:
            HYPERLINK 函數字串
        """
        try:
            if not project_name:
                return ""
            
            # 取得分支名稱，使用 fallback 邏輯
            branch_name = self._get_branch_for_open_project_link(project)
            
            if not branch_name:
                self.logger.debug(f"專案 {project_name} 無法取得有效分支名稱，open_project_link 為空")
                return ""
            
            # 使用標準的建立邏輯
            return self._build_open_project_link(project_name, branch_name, remote, is_target=False)
            
        except Exception as e:
            self.logger.error(f"建立 open_project_link 失敗 {project_name}: {str(e)}")
            return ""

    def _get_branch_for_open_project_link(self, project: Dict) -> str:
        """
        取得用於建立 open_project_link 的分支名稱 - 使用 fallback 邏輯
        
        Priority:
        1. dest-branch (如果存在且非空)
        2. 如果 dest-branch 為空：
        - 如果 revision 不是 hash (如 realtek/android-14/master)，直接使用 revision
        - 如果 revision 是 hash 且有 upstream，使用 upstream
        - 否則返回空字串
        """
        dest_branch = project.get('dest-branch', '').strip()
        revision = project.get('revision', '').strip()
        upstream = project.get('upstream', '').strip()
        
        # 1. 優先使用 dest-branch
        if dest_branch:
            self.logger.debug(f"使用 dest-branch: {dest_branch}")
            return dest_branch
        
        # 2. dest-branch 為空，檢查 revision
        if revision:
            # 如果 revision 不是 hash（如 realtek/android-14/master），直接使用
            if not self._is_revision_hash(revision):
                self.logger.debug(f"dest-branch 為空，revision 不是 hash，使用 revision: {revision}")
                return revision
            
            # 如果 revision 是 hash，嘗試使用 upstream
            if upstream:
                self.logger.debug(f"dest-branch 為空，revision 是 hash，使用 upstream: {upstream}")
                return upstream
            else:
                self.logger.debug(f"dest-branch 為空，revision 是 hash 但沒有 upstream，open_project_link 為空")
                return ""
        
        # 3. 都沒有有效值
        self.logger.debug(f"無法取得有效的分支名稱，open_project_link 為空")
        return ""
                    
    def _build_open_project_link(self, project_name: str, branch_name: str, remote: str = '', is_target: bool = True) -> str:
        """
        🔥 新方法：建立 Open Project 連結
        
        Args:
            project_name: 專案名稱
            branch_name: 分支名稱
            remote: remote 類型
            is_target: True=target_open_project_link, False=open_project_link
            
        Returns:
            HYPERLINK 函數字串
        """
        try:
            if not project_name or not branch_name:
                return ""
            
            import urllib.parse
            
            # URL 編碼專案名稱和分支名稱
            encoded_project = urllib.parse.quote(project_name, safe='')
            encoded_branch = urllib.parse.quote(branch_name, safe='')
            
            # 🔥 根據需求決定使用哪個服務器
            if is_target:
                # target_open_project_link: 根據 remote 決定服務器
                if remote == 'rtk-prebuilt':
                    base_url = 'https://mm2sd-git2.rtkbf.com'
                else:
                    base_url = 'https://mm2sd.rtkbf.com'
            else:
                # open_project_link: 一律使用 mm2sd.rtkbf.com
                base_url = 'https://mm2sd.rtkbf.com'
            
            # 🔥 建立查詢 URL（添加 is:open 條件）
            query_url = f"{base_url}/gerrit/q/project:{encoded_project}+branch:{encoded_branch}+is:open"
            
            # 建立 HYPERLINK 函數
            hyperlink = f'=HYPERLINK("{query_url}","{query_url}")'
            
            link_type = "target_open_project_link" if is_target else "open_project_link"
            self.logger.debug(f"建立 {link_type}: {project_name}/{branch_name} -> {base_url} (remote: {remote})")
            
            return hyperlink
            
        except Exception as e:
            self.logger.error(f"建立 Open Project 連結失敗 {project_name}: {str(e)}")
            return ""
            
    def _build_target_manifest_link(self, target_branch: str, remote: str = '') -> str:
        """
        🔥 建立 target_manifest 連結 - 顯示文字使用完整 URL
        """
        try:
            if not target_branch:
                return ""
            
            # 根據 remote 決定 base URL
            gerrit_base = self._get_gerrit_base_url(remote)
            
            # 🔥 根據 target_branch 決定 manifest 檔案名稱
            if 'premp.google-refplus' in target_branch:
                if 'wave' in target_branch:
                    if 'backup' in target_branch:
                        manifest_name = 'atv-google-refplus-wave-backup.xml'
                    else:
                        manifest_name = 'atv-google-refplus-wave.xml'
                else:
                    manifest_name = 'atv-google-refplus.xml'
            elif 'mp.google-refplus' in target_branch:
                manifest_name = 'atv-google-refplus-mp.xml'
            else:
                # 預設 manifest
                manifest_name = 'atv-google-refplus.xml'
            
            # 建立完整的 manifest 連結
            manifest_url = f"{gerrit_base}/gerrit/plugins/gitiles/realtek/android/manifest/+/refs/heads/{target_branch}/{manifest_name}"
            
            # 🔥 修復：顯示文字使用完整 URL
            hyperlink = f'=HYPERLINK("{manifest_url}","{manifest_url}")'
            
            self.logger.debug(f"建立 target_manifest 連結: {target_branch} -> 顯示完整URL")
            return hyperlink
            
        except Exception as e:
            self.logger.error(f"建立 target_manifest 連結失敗: {str(e)}")
            return ""

    def _parse_manifest(self, input_file: str) -> List[Dict]:
        """解析 manifest.xml 檔案"""
        try:
            tree = ET.parse(input_file)
            root = tree.getroot()
            
            # 讀取 default 標籤的 remote 和 revision 屬性
            default_remote = ''
            default_revision = ''
            default_element = root.find('default')
            if default_element is not None:
                default_remote = default_element.get('remote', '')
                default_revision = default_element.get('revision', '')
                self.logger.info(f"找到預設 remote: {default_remote}")
                self.logger.info(f"找到預設 revision: {default_revision}")
            
            projects = []
            applied_default_revision_count = 0
            
            for project in root.findall('project'):
                # 取得專案的 remote
                project_remote = project.get('remote', '')
                if not project_remote and default_remote:
                    project_remote = default_remote
                
                # 取得專案的 revision
                project_revision = project.get('revision', '')
                if not project_revision and project_remote == 'rtk' and default_revision:
                    project_revision = default_revision
                    applied_default_revision_count += 1
                
                project_data = {
                    'name': project.get('name', ''),
                    'path': project.get('path', ''),
                    'revision': project_revision,
                    'upstream': project.get('upstream', ''),
                    'dest-branch': project.get('dest-branch', ''),
                    'groups': project.get('groups', ''),
                    'clone-depth': project.get('clone-depth', ''),
                    'remote': project_remote
                }
                projects.append(project_data)
            
            self.logger.info(f"解析完成，共 {len(projects)} 個專案")
            
            if applied_default_revision_count > 0:
                self.logger.info(f"✅ 已為 {applied_default_revision_count} 個 rtk remote 專案應用預設 revision")
            
            return projects
            
        except Exception as e:
            self.logger.error(f"解析 manifest 檔案失敗: {str(e)}")
            return []

    def _is_tag_reference(self, reference: str) -> bool:
        """判斷參考是否為 Tag"""
        if not reference:
            return False
        return reference.startswith('refs/tags/')

    def _auto_detect_remote(self, project: Dict) -> str:
        """
        🔥 修正版：自動偵測專案的 remote
        ⚠️  注意：不能只看專案名稱，要考慮多種因素
        
        Args:
            project: 專案字典
            
        Returns:
            偵測到的 remote 類型
        """
        try:
            project_name = project.get('name', '')
            remote = project.get('remote', '')
            
            # 如果已經有 remote，直接使用（這是最可靠的）
            if remote:
                self.logger.debug(f"專案 {project_name} 已有 remote: {remote}")
                return remote
            
            # 🔥 預設為 rtk（大多數專案都是這個）
            detected_remote = 'rtk'
            self.logger.debug(f"預設 remote: {project_name} -> {detected_remote}")
            return detected_remote
            
        except Exception as e:
            self.logger.warning(f"自動偵測 remote 失敗: {str(e)}")
            return 'rtk'  # 預設值

    def _get_prebuilt_gerrit_manager(self):
        """取得或建立 rtk-prebuilt 專用的 GerritManager"""
        if not hasattr(self, '_prebuilt_gerrit_manager'):
            from gerrit_manager import GerritManager
            self._prebuilt_gerrit_manager = GerritManager()
            
            prebuilt_base = self._get_gerrit_base_url('rtk-prebuilt')
            self._prebuilt_gerrit_manager.base_url = prebuilt_base
            self._prebuilt_gerrit_manager.api_url = f"{prebuilt_base}/a"
            
            self.logger.info(f"建立 rtk-prebuilt 專用 GerritManager: {prebuilt_base}")
        
        return self._prebuilt_gerrit_manager

    def _check_target_tag_exists(self, project_name: str, target_tag: str, remote: str = '') -> Dict[str, str]:
        """檢查目標 Tag 是否存在 - 🔥 確保返回完整 revision"""
        result = {
            'exists_status': 'N',
            'revision': ''
        }
        
        try:
            if not project_name or not target_tag:
                return result
            
            tag_name = target_tag
            if tag_name.startswith('refs/tags/'):
                tag_name = tag_name[10:]
            
            # 根據 remote 選擇 GerritManager
            if remote == 'rtk-prebuilt':
                temp_gerrit = self._get_prebuilt_gerrit_manager()
            else:
                temp_gerrit = self.gerrit_manager
            
            tag_info = temp_gerrit.query_tag(project_name, tag_name)
            
            if tag_info['exists']:
                result['exists_status'] = 'Y'
                # 🔥 確保返回完整 revision，不截斷
                full_revision = tag_info['revision']
                result['revision'] = full_revision if full_revision else ''
                self.logger.debug(f"✅ Tag 查詢成功: {project_name}/{tag_name} -> 完整版本: {full_revision}")
            else:
                self.logger.debug(f"❌ Tag 不存在: {project_name}/{tag_name}")
            
        except Exception as e:
            self.logger.debug(f"檢查 Tag 失敗: {project_name} - {target_tag}: {str(e)}")
        
        return result

    def _check_target_branch_exists(self, project_name: str, target_branch: str, remote: str = '') -> Dict[str, str]:
        """
        🔥 修正版：檢查目標分支是否存在 - 根據 remote 欄位直接決定 Gerrit 服務器
        """
        result = {
            'exists_status': 'N',
            'revision': ''
        }
        
        try:
            if not project_name or not target_branch:
                self.logger.debug(f"參數不完整: project_name={project_name}, target_branch={target_branch}")
                return result
            
            # 🔥 修正：根據 remote 欄位直接決定使用哪個 Gerrit 服務器
            if remote == 'rtk-prebuilt':
                # 使用 rtk-prebuilt Gerrit 服務器
                gerrit_server = self._get_gerrit_base_url('rtk-prebuilt')
                self.logger.debug(f"使用 rtk-prebuilt Gerrit 服務器: {gerrit_server}")
                branch_result = self._test_branch_with_remote(project_name, target_branch, 'rtk-prebuilt')
                
                if branch_result['exists']:
                    self.logger.debug(f"✅ 在 rtk-prebuilt 找到分支: {project_name}/{target_branch}")
                    return {
                        'exists_status': 'Y',
                        'revision': branch_result['revision']
                    }
                else:
                    self.logger.debug(f"❌ 在 rtk-prebuilt 未找到分支: {project_name}/{target_branch}")
                    self.logger.debug(f"  錯誤: {branch_result.get('error', '未知')}")
                    return result
            
            else:
                # 使用預設 rtk Gerrit 服務器（包括 remote='' 或 remote='rtk' 的情況）
                gerrit_server = self._get_gerrit_base_url('')
                self.logger.debug(f"使用 rtk Gerrit 服務器: {gerrit_server}")
                branch_result = self._test_branch_with_remote(project_name, target_branch, 'rtk')
                
                if branch_result['exists']:
                    self.logger.debug(f"✅ 在 rtk 找到分支: {project_name}/{target_branch}")
                    return {
                        'exists_status': 'Y',
                        'revision': branch_result['revision']
                    }
                else:
                    self.logger.debug(f"❌ 在 rtk 未找到分支: {project_name}/{target_branch}")
                    self.logger.debug(f"  錯誤: {branch_result.get('error', '未知')}")
                    return result
            
        except Exception as e:
            self.logger.warning(f"檢查分支存在性異常: {project_name}/{target_branch}: {str(e)}")
            import traceback
            self.logger.debug(f"異常詳情: {traceback.format_exc()}")
        
        return result

    def _test_branch_with_remote(self, project_name: str, target_branch: str, remote: str) -> Dict[str, Any]:
        """
        🔥 修正版：使用指定的 remote 測試分支存在性
        
        Args:
            project_name: 專案名稱
            target_branch: 目標分支
            remote: remote 類型 ('rtk' 或 'rtk-prebuilt')
            
        Returns:
            測試結果
        """
        try:
            gerrit_server = self._get_gerrit_base_url(remote)
            self.logger.debug(f"測試分支 {project_name}/{target_branch} 在 {remote}: {gerrit_server}")
            
            # 使用增強版查詢方法
            branch_info = self._query_branch_direct_enhanced(project_name, target_branch, remote)
            
            return branch_info
            
        except Exception as e:
            self.logger.debug(f"測試 remote {remote} 時發生異常: {str(e)}")
            return {
                'exists': False,
                'revision': '',
                'server': remote,
                'error': f'測試 {remote} 異常: {str(e)}'
            }

    def _query_branch_direct_enhanced(self, project_name: str, branch_name: str, remote: str = '') -> Dict[str, Any]:
        """
        🔥 新方法：增強版直接查詢分支 - 根據 remote 選擇正確的 Gerrit 服務器
        """
        try:
            import urllib.parse
            
            encoded_project = urllib.parse.quote(project_name, safe='')
            encoded_branch = urllib.parse.quote(f"refs/heads/{branch_name}", safe='')
            
            # 🔥 根據 remote 選擇對應的 Gerrit 服務器和 GerritManager
            if remote == 'rtk-prebuilt':
                temp_gerrit = self._get_prebuilt_gerrit_manager()
                gerrit_base = self._get_gerrit_base_url('rtk-prebuilt')
                server_type = 'rtk-prebuilt'
            else:
                temp_gerrit = self.gerrit_manager
                gerrit_base = self._get_gerrit_base_url('')
                server_type = 'rtk'
            
            api_url = f"{gerrit_base}/gerrit/a/projects/{encoded_project}/branches/{encoded_branch}"
            
            self.logger.debug(f"查詢分支: {project_name}/{branch_name} 在 {server_type} 服務器")
            self.logger.debug(f"API URL: {api_url}")
            
            response = temp_gerrit._make_request(api_url, timeout=10)
            
            if response.status_code == 200:
                content = response.text
                if content.startswith(")]}'\n"):
                    content = content[5:]
                
                import json
                branch_info = json.loads(content)
                revision = branch_info.get('revision', '')
                
                self.logger.debug(f"✅ 分支查詢成功: {project_name}/{branch_name} -> 完整版本: {revision}")
                
                # 在 return 前加入：
                import copy
                result = {
                    'exists': True,
                    'revision': revision,
                    'server': server_type,
                    'full_revision': revision
                }
                return copy.deepcopy(result)  # 深拷貝避免引用共享
            elif response.status_code == 404:
                self.logger.debug(f"❌ 分支不存在: {project_name}/{branch_name} 在 {server_type}")
                return {
                    'exists': False, 
                    'revision': '',
                    'server': server_type,
                    'error': f'分支不存在於 {server_type} 服務器'
                }
            else:
                error_msg = f"HTTP {response.status_code}: {response.text[:100]}"
                self.logger.debug(f"❌ 查詢失敗: {project_name}/{branch_name} - {error_msg}")
                return {
                    'exists': False, 
                    'revision': '',
                    'server': server_type,
                    'error': error_msg
                }
                
        except Exception as e:
            self.logger.debug(f"❌ 查詢分支異常: {project_name}/{branch_name} - {str(e)}")
            return {
                'exists': False, 
                'revision': '',
                'server': remote or 'unknown',
                'error': f'查詢異常: {str(e)}'
            }

    def _handle_duplicates(self, projects: List[Dict], remove_duplicates: bool) -> tuple:
        """處理重複資料 - 使用 name + path 作為唯一識別"""
        if not remove_duplicates:
            return projects, []
        
        # 使用 name + path 作為主要唯一識別
        check_fields = ['name', 'path', 'revision', 'upstream', 'dest-branch', 'target_branch']
        
        seen = set()
        unique_projects = []
        duplicate_projects = []
        
        for project in projects:
            check_values = tuple(project.get(field, '') for field in check_fields)
            
            if check_values in seen:
                duplicate_projects.append(project)
            else:
                seen.add(check_values)
                unique_projects.append(project)
        
        self.logger.info(f"去重複後：保留 {len(unique_projects)} 個，重複 {len(duplicate_projects)} 個")
        
        return unique_projects, duplicate_projects

    # ============================================
    # 🔥 分支建立相關方法（保持原狀）
    # ============================================

    def _create_branches(self, projects: List[Dict], output_file: str, output_folder: str = None, 
            force_update: bool = False) -> List[Dict]:
        """
        建立分支並返回結果 - 修復版（正確的跳過邏輯 + 分支名稱檢查）
        """
        try:
            self.logger.info("開始建立分支...")
            self.logger.info("目標建立邏輯：")
            self.logger.info("1. 跳過 Tag 類型")
            self.logger.info("2. 跳過來源和目標分支名稱相同的情況")
            self.logger.info("3. 只有當來源和目標版本不同時才建立/更新分支（比較完整 hash）")
            self.logger.info(f"強制更新模式: {'啟用' if force_update else '停用'}")
            
            branch_results = []
            skipped_tags = 0
            skipped_same_branch_name = 0  # 🔥 新增：分支名稱相同跳過計數
            skipped_same_version = 0
            updated_branches = 0
            delete_recreate_count = 0
            prebuilt_count = 0
            normal_count = 0
            data_quality_issues = 0
            
            for project in projects:
                project_name = project.get('name', '')
                target_branch = project.get('target_branch', '')
                target_type = project.get('target_type', 'Branch')
                revision = project.get('revision', '')  # 來源 revision
                target_branch_revision = project.get('target_branch_revision', '')  # 目標分支 revision
                branch_revision = project.get('branch_revision', '-')  # 來源分支的實際 hash
                
                # 使用項目中已設定的 remote
                remote = project.get('remote', '')
                if not remote:
                    remote = self._auto_detect_remote(project)
                
                # 檢查必要資訊
                if not all([project_name, target_branch]):
                    self.logger.debug(f"跳過專案 {project_name}：缺少必要資訊")
                    continue
                
                # 🔥 跳過條件 1: Tag 類型的專案
                if target_type == 'Tag' or self._is_tag_reference(target_branch):
                    skipped_tags += 1
                    branch_result = {
                        'SN': len(branch_results) + 1,
                        'Project': project_name,
                        'revision': revision,
                        'branch_revision': branch_revision,
                        'target_branch': target_branch,
                        'target_type': 'Tag',
                        'target_branch_link': project.get('target_branch_link', ''),
                        'target_branch_revision': revision,
                        'Status': '跳過',
                        'Message': 'Tag 類型不建立分支',
                        'Already_Exists': '-',
                        'Force_Update': '-',
                        'Remote': remote,
                        'Gerrit_Server': self._get_gerrit_base_url(remote)
                    }
                    branch_results.append(branch_result)
                    continue
                
                # 🔥 跳過條件 2: 來源和目標分支名稱相同
                branch_name_check = self._should_skip_same_branch_name(project)
                if branch_name_check['should_skip']:
                    skipped_same_branch_name += 1
                    
                    branch_result = {
                        'SN': len(branch_results) + 1,
                        'Project': project_name,
                        'revision': revision,
                        'branch_revision': branch_revision,
                        'target_branch': target_branch,
                        'target_type': 'Branch',
                        'target_branch_link': project.get('target_branch_link', ''),
                        'target_branch_revision': target_branch_revision,
                        'Status': '跳過',
                        'Message': f'同根生分支，無需對齊 ({branch_name_check["reason"]})',
                        'Already_Exists': '同分支',
                        'Force_Update': '-',
                        'Remote': remote,
                        'Gerrit_Server': self._get_gerrit_base_url(remote)
                    }
                    branch_results.append(branch_result)
                    # self.logger.info(f"跳過 {project_name}：{branch_name_check['reason']}")
                    continue
                
                # 數據品質診斷
                self._diagnose_project_data(project, project_name)
                
                # 🔥 跳過條件 3: 修復版本比較邏輯
                revision_diff = self._calculate_revision_diff_fixed(
                    revision,                # source_revision (可能不是hash)
                    target_branch_revision,  # target_revision (應該是hash)
                    branch_revision,         # branch_revision (來源分支的真實hash)
                    project_name            # 用於日誌
                )
                
                # 如果版本相同且目標分支已存在，跳過建立
                if revision_diff == "N":
                    skipped_same_version += 1
                    
                    # 決定顯示哪個來源版本
                    source_display = branch_revision if branch_revision != "-" and self._is_revision_hash(branch_revision) else revision
                    source_short = source_display[:8] if len(source_display) >= 8 else source_display
                    target_short = target_branch_revision[:8] if target_branch_revision and len(target_branch_revision) >= 8 else target_branch_revision or "N/A"
                    
                    branch_result = {
                        'SN': len(branch_results) + 1,
                        'Project': project_name,
                        'revision': revision,
                        'branch_revision': branch_revision,
                        'target_branch': target_branch,
                        'target_type': 'Branch',
                        'target_branch_link': project.get('target_branch_link', ''),
                        'target_branch_revision': target_branch_revision,
                        'Status': '跳過',
                        'Message': f'Hash 相同，無需更新 (來源: {source_short}, 目標: {target_short})',
                        'Already_Exists': '是',
                        'Force_Update': '否',
                        'Remote': remote,
                        'Gerrit_Server': self._get_gerrit_base_url(remote)
                    }
                    branch_results.append(branch_result)
                    # self.logger.info(f"跳過 {project_name}：Hash 相同 (來源: {source_short}, 目標: {target_short})")
                    continue
                
                # 🔥 執行分支建立/更新（只有通過所有跳過檢查的才會執行）
                source_display = branch_revision if branch_revision != "-" and self._is_revision_hash(branch_revision) else revision
                source_short = source_display[:8] if len(source_display) >= 8 else source_display
                
                # 顯示來源和目標分支名稱資訊
                source_branch = branch_name_check.get('source_branch', 'N/A')
                self.logger.info(f"需要更新分支 {project_name}:")
                self.logger.info(f"  來源分支: {source_branch}")
                self.logger.info(f"  目標分支: {target_branch}")
                self.logger.info(f"  版本更新: {source_short} → 目標分支最新版本")
                
                # 根據 remote 選擇正確的 GerritManager
                if remote == 'rtk-prebuilt':
                    temp_gerrit = self._get_prebuilt_gerrit_manager()
                    prebuilt_count += 1
                    gerrit_server = self._get_gerrit_base_url('rtk-prebuilt')
                else:
                    temp_gerrit = self.gerrit_manager
                    normal_count += 1
                    gerrit_server = self._get_gerrit_base_url('')
                
                # 執行分支建立/更新
                success, branch_result = self._create_or_update_branch_with_retry(
                    temp_gerrit, project_name, target_branch, revision, remote, 
                    gerrit_server, force_update, len(branch_results) + 1
                )
                
                # 確保 branch_result 包含 branch_revision 資訊
                branch_result['branch_revision'] = branch_revision
                
                if success:
                    updated_branches += 1
                    if "刪除後重建" in branch_result.get('Message', ''):
                        delete_recreate_count += 1
                
                branch_results.append(branch_result)
                
                # 進度報告
                if len(branch_results) % 10 == 0:
                    success_count = len([r for r in branch_results if r['Status'] == '成功'])
                    self.logger.info(f"已處理 {len(branch_results)} 個分支，成功 {success_count} 個")
            
            # 🔥 修改：最終統計（包含分支名稱相同跳過統計）
            success_count = len([r for r in branch_results if r['Status'] == '成功'])
            failure_count = len([r for r in branch_results if r['Status'] == '失敗'])
            
            self.logger.info(f"分支建立完成，共處理 {len(branch_results)} 個專案")
            self.logger.info(f"  - 成功更新: {success_count} 個")
            self.logger.info(f"  - 失敗: {failure_count} 個")
            self.logger.info(f"  - 跳過 Tag: {skipped_tags} 個")
            self.logger.info(f"  - 跳過同根生分支: {skipped_same_branch_name} 個")  # 🔥 新增統計
            self.logger.info(f"  - 跳過版本相同: {skipped_same_version} 個")
            if delete_recreate_count > 0:
                self.logger.info(f"  - 刪除後重建: {delete_recreate_count} 個")
            
            return branch_results
            
        except Exception as e:
            self.logger.error(f"建立分支失敗: {str(e)}")
            return []

    def _diagnose_project_data(self, project: Dict, project_name: str) -> None:
        """診斷專案數據品質"""
        revision = project.get('revision', '')
        branch_revision = project.get('branch_revision', '-')
        target_branch_revision = project.get('target_branch_revision', '')
        
        self.logger.debug(f"專案 {project_name} 數據診斷:")
        self.logger.debug(f"  revision: '{revision}' (is_hash: {self._is_revision_hash(revision)})")
        self.logger.debug(f"  branch_revision: '{branch_revision}' (is_hash: {self._is_revision_hash(branch_revision) if branch_revision != '-' else False})")
        self.logger.debug(f"  target_branch_revision: '{target_branch_revision}' (is_hash: {self._is_revision_hash(target_branch_revision)})")
        
        # 檢查潛在問題
        if revision and not self._is_revision_hash(revision) and (not branch_revision or branch_revision == '-'):
            self.logger.warning(f"專案 {project_name}: revision 不是 hash 且沒有有效的 branch_revision")
        
        if not target_branch_revision or target_branch_revision == '-':
            self.logger.warning(f"專案 {project_name}: 缺少 target_branch_revision")

    def _calculate_revision_diff_fixed(self, source_revision: str, target_revision: str, 
                                    branch_revision: str = None, project_name: str = '') -> str:
        """
        修復版：計算 revision 差異 - 正確處理 hash 比較
        
        Args:
            source_revision: 原始 revision（可能是分支名稱或hash）
            target_revision: 目標 revision（應該是hash）
            branch_revision: 來源分支的實際hash（如果有的話）
            project_name: 專案名稱（用於日誌）
            
        Returns:
            "N": 版本相同，不需更新
            "Y": 版本不同，需要更新
        """
        try:
            if not source_revision:
                self.logger.debug(f"{project_name}: 無來源版本，需要更新")
                return "Y"
            
            if not target_revision or target_revision == "-":
                self.logger.debug(f"{project_name}: 無目標版本，需要更新")
                return "Y"
            
            # 選擇正確的比較來源
            is_source_hash = self._is_revision_hash(source_revision)
            
            if is_source_hash:
                # 如果 source_revision 是 hash，直接使用它
                compare_source = source_revision.strip()
                compare_type = "source_revision"
            else:
                # 如果 source_revision 不是 hash，檢查 branch_revision
                if branch_revision and branch_revision != "-" and self._is_revision_hash(branch_revision):
                    compare_source = branch_revision.strip()
                    compare_type = "branch_revision"
                else:
                    # 都不是有效的 hash，無法準確比較，當作需要更新
                    self.logger.warning(f"{project_name}: 無有效的 hash 進行比較 (source: '{source_revision}', branch: '{branch_revision}')，當作需要更新")
                    return "Y"
            
            # 確保目標也是hash
            if not self._is_revision_hash(target_revision):
                self.logger.warning(f"{project_name}: 目標版本不是有效hash: '{target_revision}'，當作需要更新")
                return "Y"
            
            target_clean = target_revision.strip()
            
            # 比較完整的 hash 值
            if compare_source == target_clean:
                self.logger.debug(f"{project_name}: Hash 相同，跳過更新")
                self.logger.debug(f"  使用 {compare_type}: {compare_source[:8]} == {target_clean[:8]}")
                return "N"  # 相同，不需更新
            else:
                self.logger.debug(f"{project_name}: Hash 不同，需要更新")
                self.logger.debug(f"  使用 {compare_type}: {compare_source[:8]} != {target_clean[:8]}")
                return "Y"  # 不同，需要更新
                
        except Exception as e:
            self.logger.debug(f"{project_name}: 計算 revision_diff 失敗: {str(e)}，當作需要更新")
            return "Y"
            
    def _create_or_update_branch_with_retry(self, gerrit_manager, project_name: str, 
                                        target_branch: str, revision: str, remote: str,
                                        gerrit_server: str, force_update: bool, sn: int) -> tuple:
        """
        🔥 改進版：建立或更新分支，優先使用安全的更新方法
        
        流程：
        1. 檢查分支是否存在
        2. 如果不存在 → 建立新分支
        3. 如果存在 → 使用 update_branch（更安全，有備份機制）
        4. 只有在 update 失敗時才回退到刪除重建
        
        Returns:
            (success: bool, branch_result: dict)
        """
        try:
            self.logger.debug(f"處理分支: {project_name}/{target_branch}")
            
            # 🔥 步驟 1: 先檢查分支是否存在
            branch_info = gerrit_manager.check_branch_exists_and_get_revision(project_name, target_branch)
            branch_exists = branch_info.get('exists', False)
            current_revision = branch_info.get('revision', '')
            
            if not branch_exists:
                # 🔥 情況 1: 分支不存在 → 直接建立新分支
                self.logger.info(f"分支不存在，建立新分支: {project_name}/{target_branch}")
                result = gerrit_manager.create_branch(project_name, target_branch, revision)
                
                if result.get('success', False):
                    return True, {
                        'SN': sn,
                        'Project': project_name,
                        'revision': revision,
                        'target_branch': target_branch,
                        'target_type': 'Branch',
                        'target_branch_link': '',
                        'target_branch_revision': revision,
                        'Status': '成功',
                        'Message': f"成功建立新分支：{result.get('message', '')}",
                        'Already_Exists': '否',
                        'Force_Update': '否',
                        'Remote': remote,
                        'Gerrit_Server': gerrit_server
                    }
                else:
                    return False, {
                        'SN': sn,
                        'Project': project_name,
                        'revision': revision,
                        'target_branch': target_branch,
                        'target_type': 'Branch',
                        'target_branch_link': '',
                        'target_branch_revision': revision,
                        'Status': '失敗',
                        'Message': f"建立新分支失敗：{result.get('message', '')}",
                        'Already_Exists': '否',
                        'Force_Update': '否',
                        'Remote': remote,
                        'Gerrit_Server': gerrit_server
                    }
            
            else:
                # 🔥 情況 2: 分支已存在 → 使用更安全的 update_branch
                self.logger.info(f"分支已存在，使用 update_branch: {project_name}/{target_branch}")
                self.logger.info(f"  當前版本: {current_revision}")
                self.logger.info(f"  目標版本: {revision[:8]}")
                
                # 🔥 使用 update_branch（有備份機制，更安全）
                update_result = gerrit_manager.update_branch(
                    project_name, target_branch, revision, force=force_update
                )
                
                if update_result.get('success', False):
                    return True, {
                        'SN': sn,
                        'Project': project_name,
                        'revision': revision,
                        'target_branch': target_branch,
                        'target_type': 'Branch',
                        'target_branch_link': '',
                        'target_branch_revision': revision,
                        'Status': '成功',
                        'Message': f"成功更新分支：{update_result.get('message', '')}",
                        'Already_Exists': '是',
                        'Force_Update': '是' if force_update else '否',
                        'Remote': remote,
                        'Gerrit_Server': gerrit_server
                    }
                
                elif not force_update:
                    # 🔥 如果快進式更新失敗且不是強制模式，直接返回失敗
                    return False, {
                        'SN': sn,
                        'Project': project_name,
                        'revision': revision,
                        'target_branch': target_branch,
                        'target_type': 'Branch',
                        'target_branch_link': '',
                        'target_branch_revision': revision,
                        'Status': '失敗',
                        'Message': f"更新失敗（需要強制更新）：{update_result.get('message', '')}",
                        'Already_Exists': '是',
                        'Force_Update': '否',
                        'Remote': remote,
                        'Gerrit_Server': gerrit_server
                    }
                
                else:
                    # 🔥 強制模式下 update_branch 也失敗，最後手段：刪除重建
                    self.logger.warning(f"update_branch 失敗，使用最後手段刪除重建: {update_result.get('message', '')}")
                    
                    fallback_result = self._fallback_delete_and_recreate(
                        gerrit_manager, project_name, target_branch, revision, 
                        remote, gerrit_server, sn
                    )
                    
                    # 標記這是使用了刪除重建的方法
                    if fallback_result[0]:  # success
                        fallback_result[1]['Message'] = f"透過刪除重建成功：{fallback_result[1]['Message']}"
                    
                    return fallback_result
        
        except Exception as e:
            return False, {
                'SN': sn,
                'Project': project_name,
                'revision': revision,
                'target_branch': target_branch,
                'target_type': 'Branch',
                'target_branch_link': '',
                'target_branch_revision': revision,
                'Status': '失敗',
                'Message': f"建立/更新分支異常：{str(e)}",
                'Already_Exists': '-',
                'Force_Update': '是' if force_update else '否',
                'Remote': remote,
                'Gerrit_Server': gerrit_server
            }

    def _fallback_delete_and_recreate(self, gerrit_manager, project_name: str, 
                                    target_branch: str, revision: str, 
                                    remote: str, gerrit_server: str, sn: int) -> tuple:
        """
        🔥 最後手段：刪除後重建（保留原邏輯作為 fallback）
        只有在 update_branch 完全失敗時才使用
        """
        try:
            self.logger.warning(f"⚠️  使用最後手段：刪除後重建 {project_name}/{target_branch}")
            
            # 刪除分支
            delete_result = gerrit_manager.delete_branch(project_name, target_branch)
            
            if delete_result.get('success', False):
                self.logger.info(f"✅ 成功刪除分支: {project_name}/{target_branch}")
                
                # 重新建立分支
                recreate_result = gerrit_manager.create_branch(project_name, target_branch, revision)
                
                if recreate_result.get('success', False):
                    return True, {
                        'SN': sn,
                        'Project': project_name,
                        'revision': revision,
                        'target_branch': target_branch,
                        'target_type': 'Branch',
                        'target_branch_link': '',
                        'target_branch_revision': revision,
                        'Status': '成功',
                        'Message': f"刪除後重建成功：{recreate_result.get('message', '')}",
                        'Already_Exists': '否',
                        'Force_Update': '是',
                        'Remote': remote,
                        'Gerrit_Server': gerrit_server
                    }
                else:
                    return False, {
                        'SN': sn,
                        'Project': project_name,
                        'revision': revision,
                        'target_branch': target_branch,
                        'target_type': 'Branch',
                        'target_branch_link': '',
                        'target_branch_revision': revision,
                        'Status': '失敗',
                        'Message': f"刪除成功但重建失敗：{recreate_result.get('message', '')}",
                        'Already_Exists': '否',
                        'Force_Update': '是',
                        'Remote': remote,
                        'Gerrit_Server': gerrit_server
                    }
            else:
                return False, {
                    'SN': sn,
                    'Project': project_name,
                    'revision': revision,
                    'target_branch': target_branch,
                    'target_type': 'Branch',
                    'target_branch_link': '',
                    'target_branch_revision': revision,
                    'Status': '失敗',
                    'Message': f"刪除分支失敗：{delete_result.get('message', '')}",
                    'Already_Exists': '是',
                    'Force_Update': '是',
                    'Remote': remote,
                    'Gerrit_Server': gerrit_server
                }
                
        except Exception as e:
            return False, {
                'SN': sn,
                'Project': project_name,
                'revision': revision,
                'target_branch': target_branch,
                'target_type': 'Branch',
                'target_branch_link': '',
                'target_branch_revision': revision,
                'Status': '失敗',
                'Message': f"刪除重建異常：{str(e)}",
                'Already_Exists': '-',
                'Force_Update': '是',
                'Remote': remote,
                'Gerrit_Server': gerrit_server
            }        

    def process_tvconfig_alignment(self, output_folder: str = './output') -> bool:
        """
        處理 Master Tvconfig 對齊功能 - 修正版：使用正確的轉換命名和跳過邏輯 + 保留檔案
        """
        try:
            self.logger.info("=== 開始執行對齊 Master Tvconfig 功能 ===")
            self.logger.info(f"使用 Android 版本: {self.current_android_version}")
            
            # 確保輸出資料夾存在
            utils.ensure_dir(output_folder)
            
            # 🔥 新增：保存檔案列表，用於最終檢查
            saved_files = []
            
            # 步驟 1: 選擇 manifest 來源
            manifest_info = self._choose_tvconfig_manifest_source()
            if not manifest_info:
                return False
            
            manifest_path, is_from_gerrit, original_manifest_content = manifest_info
            
            # 🔥 步驟 1.5: 保存從 Gerrit 下載的原始檔案
            if is_from_gerrit and original_manifest_content:
                gerrit_original_path = self._save_gerrit_manifest_file(
                    original_manifest_content, "atv-google-refplus.xml", output_folder
                )
                if gerrit_original_path:
                    saved_files.append(gerrit_original_path)
            
            # 步驟 2: 選擇處理類型
            process_type = self._choose_tvconfig_process_type()
            if not process_type:
                return False
            
            # 步驟 3: 詢問分支建立選項
            branch_options = self._ask_tvconfig_branch_options()
            if not branch_options:
                return False
            
            create_branches, check_branch_exists, confirmed = branch_options
            if not confirmed:
                return False
            
            # 步驟 4: 備份原始檔案到輸出資料夾
            backup_info = self._backup_tvconfig_manifest_files(
                manifest_path, output_folder, is_from_gerrit, original_manifest_content
            )
            
            # 步驟 5: 展開 manifest（如果需要且是從 gerrit 下載的）
            processed_manifest_path = manifest_path
            expanded_manifest_path = None
            
            if is_from_gerrit and original_manifest_content:
                if self._has_tvconfig_include_tags(original_manifest_content):
                    self.logger.info("檢測到 include 標籤，準備展開 manifest...")
                    expanded_result = self._expand_tvconfig_manifest(output_folder)
                    if expanded_result:
                        expanded_manifest_path, expanded_content = expanded_result
                        processed_manifest_path = expanded_manifest_path
                        self.logger.info(f"✅ 使用展開後的檔案: {expanded_manifest_path}")
                        
                        # 🔥 步驟 5.5: 保存展開檔案
                        saved_expanded_path = self._save_expanded_manifest_file(
                            expanded_content, "atv-google-refplus.xml", output_folder
                        )
                        if saved_expanded_path:
                            saved_files.append(saved_expanded_path)
                    else:
                        self.logger.warning("⚠️ Manifest 展開失敗，使用原始檔案")
            
            # 步驟 6: 解析並過濾專案
            all_projects = self._parse_manifest(processed_manifest_path)
            if not all_projects:
                self.logger.error("無法解析 manifest 檔案或檔案為空")
                return False
                
            tvconfig_projects = self._filter_tvconfigs_projects(all_projects)
            
            if not tvconfig_projects:
                self.logger.error("沒有找到 tvconfigs_prebuilt 相關的專案")
                return False
            
            self.logger.info(f"找到 {len(tvconfig_projects)} 個 tvconfigs_prebuilt 專案")
            
            # 步驟 7: 提取來源 manifest 檔名
            source_manifest_name = self._extract_tvconfig_manifest_filename(processed_manifest_path)
            
            # 步驟 8: 轉換專案（使用現有邏輯，支援新的轉換類型，添加 tvconfig 標記）
            converted_projects = self._convert_projects(
                tvconfig_projects, process_type, check_branch_exists, source_manifest_name, 
                is_tvconfig=True  # 標記為 tvconfig 轉換，會使用 TVCONFIG_SKIP_PROJECTS 配置
            )
            
            # 步驟 9: 添加連結資訊（使用現有邏輯）
            projects_with_links = self._add_links_to_projects(converted_projects)
            
            # 步驟 10: 處理重複（不去重）
            unique_projects, duplicate_projects = projects_with_links, []
            
            # 步驟 11: 重新編號
            unique_projects = self._renumber_projects(unique_projects)
            
            # 🔥 步驟 11.5: 生成轉換後的 manifest 檔案
            converted_manifest_path = self._generate_converted_manifest(
                unique_projects, processed_manifest_path, output_folder, process_type
            )
            if converted_manifest_path:
                saved_files.append(converted_manifest_path)
            
            # 步驟 12: 生成輸出檔案名
            output_filename = f"{process_type}_tvconfigs_prebuilt_prebuild.xlsx"
            
            # 步驟 13: 寫入 Excel（使用現有邏輯）
            self._write_excel_unified_basic(unique_projects, duplicate_projects, output_filename, output_folder)
            
            # 步驟 14: 如果選擇建立分支，執行分支建立
            if create_branches:
                self.logger.info("開始執行分支建立流程...")
                branch_results = self._create_branches(unique_projects, output_filename, output_folder, False)
                self._add_branch_status_sheet_with_revision(output_filename, output_folder, branch_results)
                self.logger.info("✅ 分支建立流程完成")
            
            # 🔥 步驟 15: 最終檔案檢查報告
            excel_path = os.path.join(output_folder, output_filename)
            saved_files.append(excel_path)
            self._final_file_report(output_folder, saved_files)
            
            self.logger.info(f"=== 對齊 Master Tvconfig 功能執行完成，Excel 檔案：{excel_path} ===")
            return True
            
        except Exception as e:
            self.logger.error(f"對齊 Master Tvconfig 功能執行失敗: {str(e)}")
            import traceback
            self.logger.error(f"錯誤詳情: {traceback.format_exc()}")
            return False

    def _choose_tvconfig_manifest_source(self) -> Optional[tuple]:
        """選擇 manifest 來源"""
        print("\n請選擇 manifest.xml 來源：")
        print("[1] 從 Gerrit 自動下載 master 分支的 manifest.xml")
        print("[2] 使用本地 manifest.xml 檔案")
        print("[0] 返回上層選單")
        
        choice = input("請選擇 (1-2，預設: 1): ").strip() or "1"
        
        if choice == "0":
            return None
        elif choice == "1":
            # 從 Gerrit 下載
            result = self._download_tvconfig_master_manifest()
            if result:
                manifest_path, original_content = result
                return manifest_path, True, original_content
            else:
                return None
        elif choice == "2":
            # 使用本地檔案
            manifest_path = input("請輸入 manifest.xml 檔案路徑: ").strip()
            if not os.path.exists(manifest_path):
                print(f"錯誤：檔案不存在 - {manifest_path}")
                return None
            
            # 讀取本地檔案內容
            try:
                with open(manifest_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                return manifest_path, False, content
            except Exception as e:
                print(f"錯誤：無法讀取檔案 - {str(e)}")
                return None
        else:
            print("無效的選擇")
            return None

    def _choose_tvconfig_process_type(self) -> Optional[str]:
        """選擇處理類型 - 修正版：使用正確的 tvconfig 命名"""
        print("\n請選擇目的 code line:")
        print("[1] master_to_premp (master → premp)")
        print("[2] master_to_mp (master → mp)")  
        print("[3] master_to_mpbackup (master → mpbackup)")
        print("[0] 返回上層選單")
        
        choice = input("請選擇 (1-3): ").strip()
        
        if choice == "0":
            return None
        elif choice == "1":
            return "master_to_premp"
        elif choice == "2":
            return "master_to_mp"
        elif choice == "3":
            return "master_to_mpbackup"
        else:
            print("無效的選擇")
            return None

    def _ask_tvconfig_branch_options(self) -> Optional[tuple]:
        """詢問分支建立選項 - 修正版：獨立的分支存在性檢查選項"""
        print("\n分支建立選項:")
        
        # 詢問是否建立分支
        create_branches_input = input("是否建立分支？ (y/N): ").strip().lower()
        create_branches = create_branches_input == 'y'
        
        # 🔥 修正：無論是否建立分支，都詢問是否檢查分支存在性
        check_exists_input = input("是否檢查分支存在性？(會比較慢) (y/N): ").strip().lower()
        check_branch_exists = check_exists_input == 'y'
        
        # 顯示設定摘要
        print(f"\n設定摘要:")
        print(f"- 建立分支: {'是' if create_branches else '否'}")
        print(f"- 檢查分支存在性: {'是' if check_branch_exists else '否'}")
        
        # 最終確認
        confirm_input = input("\n是否確認執行？ (Y/n): ").strip().lower()
        confirmed = confirm_input != 'n'
        
        if not confirmed:
            print("取消執行")
            return None
        
        return create_branches, check_branch_exists, confirmed

    def _download_tvconfig_master_manifest(self) -> Optional[tuple]:
        """從 Gerrit 下載 master 分支的 manifest.xml"""
        try:
            # 🔥 使用動態 Android 版本
            master_branch = config.get_default_android_master_branch()
            manifest_filename = "atv-google-refplus.xml"  # 預設使用這個檔案
            
            gerrit_url = f"https://mm2sd.rtkbf.com/gerrit/plugins/gitiles/realtek/android/manifest/+/refs/heads/{master_branch}/{manifest_filename}"
            
            self.logger.info(f"正在從 Gerrit 下載 master manifest...")
            self.logger.info(f"URL: {gerrit_url}")
            self.logger.info(f"分支: {master_branch}")
            self.logger.info(f"檔案: {manifest_filename}")
            
            # 使用臨時檔案下載
            import tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xml', prefix='tvconfig_master_') as temp_file:
                temp_path = temp_file.name
            
            try:
                success = self.gerrit_manager.download_file_from_link(gerrit_url, temp_path)
                
                if success and os.path.exists(temp_path):
                    with open(temp_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    
                    self.logger.info(f"✅ 成功下載 master manifest: {len(content)} 字符")
                    return temp_path, content
                else:
                    self.logger.error("❌ 下載 master manifest 失敗")
                    return None
                    
            except Exception as e:
                self.logger.error(f"下載過程發生錯誤: {str(e)}")
                return None
            finally:
                # 注意：這裡不清理臨時檔案，因為後續還需要使用
                pass
                
        except Exception as e:
            self.logger.error(f"下載 master manifest 異常: {str(e)}")
            return None

    def _backup_tvconfig_manifest_files(self, manifest_path: str, output_folder: str, 
                            is_from_gerrit: bool, original_content: str) -> Dict[str, str]:
        """備份 manifest 檔案到輸出資料夾 - 修正版：改進檔案名稱處理"""
        backup_info = {}
        
        try:
            if is_from_gerrit:
                # 🔥 如果是從 gerrit 下載的，使用 gerrit_ 前綴 + 原始檔名
                # 但不使用複雜的分支名稱，直接使用簡潔的檔名
                backup_filename = "gerrit_atv-google-refplus.xml"
            else:
                # 🔥 如果是本地檔案，直接使用原檔名（不加 backup_ 前綴）
                original_filename = os.path.basename(manifest_path)
                backup_filename = original_filename
            
            backup_path = os.path.join(output_folder, backup_filename)
            
            # 🔥 檢查檔案是否已存在且內容相同
            should_save = True
            if os.path.exists(backup_path):
                try:
                    with open(backup_path, 'r', encoding='utf-8') as f:
                        existing_content = f.read()
                    
                    if existing_content == original_content:
                        self.logger.info(f"✅ 檔案已存在且內容相同，跳過保存: {backup_filename}")
                        backup_info['original_backup'] = backup_path
                        return backup_info
                    else:
                        self.logger.info(f"⚠️ 檔案已存在但內容不同，將覆蓋: {backup_filename}")
                except Exception as e:
                    self.logger.warning(f"檢查現有檔案失敗，將覆蓋: {str(e)}")
            
            # 寫入備份檔案
            with open(backup_path, 'w', encoding='utf-8') as f:
                f.write(original_content)
            
            backup_info['original_backup'] = backup_path
            
            self.logger.info(f"✅ 已備份 manifest: {backup_filename}")
            
            # 驗證備份檔案
            if os.path.exists(backup_path):
                file_size = os.path.getsize(backup_path)
                self.logger.info(f"✅ 備份檔案驗證成功: {backup_filename} ({file_size} bytes)")
            
            return backup_info
            
        except Exception as e:
            self.logger.error(f"備份 manifest 檔案失敗: {str(e)}")
            return backup_info

    def _has_tvconfig_include_tags(self, xml_content: str) -> bool:
        """檢查 XML 內容是否包含 include 標籤 - 參考 feature_three.py"""
        try:
            import re
            
            # 使用正則表達式檢查 include 標籤
            include_pattern = r'<include\s+name\s*=\s*["\'][^"\']*["\'][^>]*/?>'
            matches = re.findall(include_pattern, xml_content, re.IGNORECASE)
            
            if matches:
                self.logger.info(f"🔍 發現 {len(matches)} 個 include 標籤:")
                for i, match in enumerate(matches, 1):
                    self.logger.info(f"  {i}. {match}")
                return True
            else:
                self.logger.info("ℹ️ 未發現 include 標籤")
                return False
                
        except Exception as e:
            self.logger.error(f"檢查 include 標籤時發生錯誤: {str(e)}")
            return False

    def _expand_tvconfig_manifest(self, output_folder: str) -> Optional[tuple]:
        """
        展開包含 include 的 manifest - 修正版：安全的檔案路徑處理
        """
        import subprocess
        import tempfile
        import shutil
        
        try:
            # 🔥 使用動態分支
            repo_url = "ssh://mm2sd.rtkbf.com:29418/realtek/android/manifest"
            branch = config.get_default_android_master_branch()
            source_filename = "atv-google-refplus.xml"
            
            # 🔥 修正：生成安全的展開檔案名稱
            safe_branch_name = branch.replace('/', '_')
            expanded_filename = f"gerrit_atv-google-refplus_{safe_branch_name}_expanded.xml"
            final_expanded_path = os.path.abspath(os.path.join(output_folder, expanded_filename))
            
            self.logger.info(f"🎯 準備展開 tvconfig manifest...")
            self.logger.info(f"🎯 源檔案: {source_filename}")
            self.logger.info(f"🎯 使用分支: {branch}")
            self.logger.info(f"🎯 展開檔案名: {expanded_filename}")
            
            # 確保輸出資料夾存在
            utils.ensure_dir(output_folder)
            
            # 檢查 repo 命令是否可用
            try:
                repo_check = subprocess.run(
                    ["repo", "--version"], 
                    capture_output=True, 
                    text=True, 
                    timeout=10
                )
                if repo_check.returncode != 0:
                    self.logger.error(f"❌ repo 工具檢查失敗: {repo_check.stderr}")
                    return None
            except FileNotFoundError:
                self.logger.error("❌ repo 命令未找到，請確認已安裝 repo 工具")
                return None
            
            # 建立臨時工作目錄
            temp_work_dir = tempfile.mkdtemp(prefix='tvconfig_repo_expand_')
            self.logger.info(f"📁 建立臨時工作目錄: {temp_work_dir}")
            
            original_cwd = os.getcwd()
            
            try:
                # 切換到臨時目錄
                os.chdir(temp_work_dir)
                
                # repo init
                init_cmd = [
                    "repo", "init", 
                    "-u", repo_url,
                    "-b", branch,
                    "-m", source_filename
                ]
                
                init_result = subprocess.run(
                    init_cmd,
                    capture_output=True,
                    text=True,
                    timeout=120
                )
                
                if init_result.returncode != 0:
                    self.logger.error(f"❌ repo init 失敗: {init_result.stderr}")
                    return None
                
                # repo manifest 展開
                manifest_result = subprocess.run(
                    ["repo", "manifest"],
                    capture_output=True,
                    text=True,
                    timeout=60
                )
                
                if manifest_result.returncode != 0:
                    self.logger.error(f"❌ repo manifest 失敗: {manifest_result.stderr}")
                    return None
                
                expanded_content = manifest_result.stdout
                
                if not expanded_content.strip():
                    self.logger.error("❌ repo manifest 返回空內容")
                    return None
                
                self.logger.info(f"✅ repo manifest 成功，內容長度: {len(expanded_content)} 字符")
                
                # 保存展開檔案到輸出資料夾
                with open(final_expanded_path, 'w', encoding='utf-8') as f:
                    f.write(expanded_content)
                
                # 驗證檔案
                if os.path.exists(final_expanded_path):
                    file_size = os.path.getsize(final_expanded_path)
                    self.logger.info(f"✅ 展開檔案保存成功: {expanded_filename} ({file_size} bytes)")
                    return final_expanded_path, expanded_content
                else:
                    self.logger.error(f"❌ 展開檔案保存失敗: {final_expanded_path}")
                    return None
                    
            finally:
                # 恢復原始工作目錄
                os.chdir(original_cwd)
                
                # 清理臨時目錄
                try:
                    shutil.rmtree(temp_work_dir)
                    self.logger.info(f"🗑️ 清理臨時目錄成功")
                except Exception as e:
                    self.logger.warning(f"⚠️ 清理臨時目錄失敗: {str(e)}")
                    
        except Exception as e:
            self.logger.error(f"❌ 展開 tvconfig manifest 時發生錯誤: {str(e)}")
            return None

    def _save_expanded_tvconfig_manifest(self, expanded_content: str, output_folder: str):
        """保存展開後的 manifest 檔案 - 修正版：安全的檔案名稱"""
        try:
            branch = config.get_default_android_master_branch()
            # 🔥 修正：安全的檔案名稱處理
            safe_branch_name = branch.replace('/', '_')
            expanded_filename = f"gerrit_atv-google-refplus_{safe_branch_name}_expanded.xml"
            expanded_path = os.path.join(output_folder, expanded_filename)
            
            with open(expanded_path, 'w', encoding='utf-8') as f:
                f.write(expanded_content)
            
            if os.path.exists(expanded_path):
                file_size = os.path.getsize(expanded_path)
                self.logger.info(f"✅ 展開檔案已保存: {expanded_filename} ({file_size} bytes)")
            
        except Exception as e:
            self.logger.error(f"保存展開檔案失敗: {str(e)}")

    def _filter_tvconfigs_projects(self, projects: List[Dict]) -> List[Dict]:
        """過濾只保留 name 包含 tvconfigs_prebuilt 的專案"""
        try:
            tvconfig_projects = []
            
            for project in projects:
                project_name = project.get('name', '')
                
                # 檢查 name 是否包含 tvconfigs_prebuilt
                if 'tvconfigs_prebuilt' in project_name:
                    tvconfig_projects.append(project)
                    self.logger.debug(f"✅ 保留專案: {project_name}")
                else:
                    self.logger.debug(f"⏭️ 跳過專案: {project_name}")
            
            self.logger.info(f"過濾完成: 原始 {len(projects)} 個專案 → 保留 {len(tvconfig_projects)} 個 tvconfigs_prebuilt 專案")
            
            return tvconfig_projects
            
        except Exception as e:
            self.logger.error(f"過濾 tvconfigs_prebuilt 專案失敗: {str(e)}")
            return []

    def _extract_tvconfig_manifest_filename(self, manifest_path: str) -> str:
        """提取 manifest 檔案名稱 - 修正版：安全的檔案名稱處理"""
        try:
            if manifest_path:
                filename = os.path.basename(manifest_path)
                # 如果是展開檔案，使用原始檔案名
                if 'expanded' in filename:
                    return "atv-google-refplus.xml"
                else:
                    return filename
            else:
                return "atv-google-refplus.xml"
        except Exception as e:
            self.logger.error(f"提取檔案名失敗: {str(e)}")
            return "atv-google-refplus.xml"

    def _save_original_manifest_file(self, input_file: str, output_folder: str) -> str:
        """
        保存原始 manifest 檔案到輸出資料夾 - 修正版：直接使用原始檔名
        
        Args:
            input_file: 原始輸入檔案路徑
            output_folder: 輸出資料夾
            
        Returns:
            保存的檔案路徑
        """
        try:
            original_filename = os.path.basename(input_file)
            # 🔥 修改：直接使用原始檔名，不加 original_ 前綴
            backup_path = os.path.join(output_folder, original_filename)
            
            # 🔥 檢查檔案是否已存在，如果存在且內容相同則跳過
            should_copy = True
            if os.path.exists(backup_path):
                try:
                    with open(input_file, 'r', encoding='utf-8') as f1:
                        source_content = f1.read()
                    with open(backup_path, 'r', encoding='utf-8') as f2:
                        existing_content = f2.read()
                    
                    if source_content == existing_content:
                        self.logger.info(f"✅ 檔案已存在且內容相同，跳過複製: {original_filename}")
                        return backup_path
                    else:
                        self.logger.info(f"⚠️ 檔案已存在但內容不同，將覆蓋: {original_filename}")
                except Exception as e:
                    self.logger.warning(f"檢查現有檔案失敗，將覆蓋: {str(e)}")
            
            if should_copy:
                # 讀取原始檔案內容
                with open(input_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                # 寫入備份檔案
                with open(backup_path, 'w', encoding='utf-8') as f:
                    f.write(content)
            
            # 驗證備份檔案
            if os.path.exists(backup_path):
                file_size = os.path.getsize(backup_path)
                self.logger.info(f"✅ 已保存原始 manifest: {original_filename} ({file_size} bytes)")
            
            return backup_path
            
        except Exception as e:
            self.logger.error(f"保存原始 manifest 檔案失敗: {str(e)}")
            return ""

    def _save_gerrit_manifest_file(self, content: str, filename: str, output_folder: str) -> str:
        """
        保存從 Gerrit 下載的 manifest 檔案 - 參考 feature_three.py
        
        Args:
            content: 檔案內容
            filename: 原始檔案名
            output_folder: 輸出資料夾
            
        Returns:
            保存的檔案路徑
        """
        try:
            gerrit_filename = f"gerrit_{filename}"
            gerrit_path = os.path.join(output_folder, gerrit_filename)
            
            # 確保輸出資料夾存在
            utils.ensure_dir(output_folder)
            
            self.logger.info(f"準備保存 Gerrit 檔案到: {gerrit_path}")
            self.logger.info(f"檔案內容長度: {len(content)} 字符")
            
            # 寫入檔案
            with open(gerrit_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            # 驗證檔案是否成功保存
            if os.path.exists(gerrit_path):
                file_size = os.path.getsize(gerrit_path)
                self.logger.info(f"✅ Gerrit 檔案已成功保存: {gerrit_filename} ({file_size} bytes)")
                
                # 再次確認檔案內容
                with open(gerrit_path, 'r', encoding='utf-8') as f:
                    saved_content = f.read()
                    if len(saved_content) == len(content):
                        self.logger.info(f"✅ 檔案內容驗證成功: {len(saved_content)} 字符")
                    else:
                        self.logger.warning(f"⚠️ 檔案內容長度不匹配: 原始 {len(content)}, 保存 {len(saved_content)}")
                
                return gerrit_path
            else:
                raise Exception(f"檔案保存後不存在: {gerrit_path}")
                
        except Exception as e:
            self.logger.error(f"保存 Gerrit 檔案失敗: {str(e)}")
            return ""

    def _save_expanded_manifest_file(self, content: str, original_filename: str, output_folder: str) -> str:
        """
        保存展開後的 manifest 檔案 - 參考 feature_three.py
        
        Args:
            content: 展開後的內容
            original_filename: 原始檔案名
            output_folder: 輸出資料夾
            
        Returns:
            保存的檔案路徑
        """
        try:
            # 生成展開檔案名稱
            base_name = os.path.splitext(original_filename)[0]
            expanded_filename = f"gerrit_{base_name}_expanded.xml"
            expanded_path = os.path.join(output_folder, expanded_filename)
            
            # 確保輸出資料夾存在
            utils.ensure_dir(output_folder)
            
            self.logger.info(f"準備保存展開檔案到: {expanded_path}")
            self.logger.info(f"檔案內容長度: {len(content)} 字符")
            
            # 寫入檔案
            with open(expanded_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            # 驗證檔案
            if os.path.exists(expanded_path):
                file_size = os.path.getsize(expanded_path)
                self.logger.info(f"✅ 展開檔案已成功保存: {expanded_filename} ({file_size} bytes)")
                
                # 驗證內容
                with open(expanded_path, 'r', encoding='utf-8') as f:
                    saved_content = f.read()
                    project_count = saved_content.count('<project ')
                    self.logger.info(f"✅ 檔案內容驗證成功: {len(saved_content)} 字符, {project_count} 個專案")
                
                return expanded_path
            else:
                raise Exception(f"展開檔案保存後不存在: {expanded_path}")
                
        except Exception as e:
            self.logger.error(f"保存展開檔案失敗: {str(e)}")
            return ""

    def _generate_converted_manifest(self, projects: List[Dict], original_manifest_path: str, 
                                output_folder: str, process_type: str) -> str:
        """
        生成轉換後的 manifest 檔案 - 修正版：使用正確的目標檔案名
        """
        try:
            # 讀取原始 manifest 檔案
            with open(original_manifest_path, 'r', encoding='utf-8') as f:
                original_content = f.read()
            
            # 解析原始 XML
            import xml.etree.ElementTree as ET
            root = ET.fromstring(original_content)
            
            # 建立專案名稱到轉換資訊的映射
            project_mapping = {}
            for proj in projects:
                project_name = proj.get('name', '')
                target_branch = proj.get('target_branch', '')
                if project_name and target_branch:
                    project_mapping[project_name] = target_branch
            
            # 轉換 XML 內容
            converted_content = self._convert_xml_content_with_projects(
                original_content, project_mapping, process_type
            )
            
            # 🔥 修改：根據 process_type 生成正確的目標檔案名
            converted_filename = self._get_target_manifest_filename(process_type)
            converted_path = os.path.join(output_folder, converted_filename)
            
            # 保存轉換後的檔案
            with open(converted_path, 'w', encoding='utf-8') as f:
                f.write(converted_content)
            
            # 驗證檔案
            if os.path.exists(converted_path):
                file_size = os.path.getsize(converted_path)
                self.logger.info(f"✅ 轉換後 manifest 已成功保存: {converted_filename} ({file_size} bytes)")
                
                # 統計轉換項目
                converted_count = len([proj for proj in projects if proj.get('target_branch', '') != proj.get('revision', '')])
                self.logger.info(f"✅ 已轉換 {converted_count} 個專案的分支")
                
                return converted_path
            else:
                raise Exception(f"轉換後檔案保存失敗: {converted_path}")
                
        except Exception as e:
            self.logger.error(f"生成轉換後 manifest 失敗: {str(e)}")
            return ""

    def _get_target_manifest_filename(self, process_type: str) -> str:
        """
        🔥 新方法：根據處理類型取得正確的目標檔案名
        
        Args:
            process_type: 處理類型
            
        Returns:
            目標檔案名
        """
        filename_mapping = {
            # 功能二的處理類型
            'master_vs_premp': 'atv-google-refplus-premp.xml',
            'premp_vs_mp': 'atv-google-refplus-wave.xml', 
            'mp_vs_mpbackup': 'atv-google-refplus-wave-backup.xml',
            
            # tvconfig 功能的處理類型  
            'master_to_premp': 'atv-google-refplus-premp.xml',
            'master_to_mp': 'atv-google-refplus-wave.xml',
            'master_to_mpbackup': 'atv-google-refplus-wave-backup.xml'
        }
        
        target_filename = filename_mapping.get(process_type)
        
        if target_filename:
            self.logger.info(f"✅ 使用目標檔案名: {process_type} → {target_filename}")
            return target_filename
        else:
            # 如果沒有預定義的映射，使用原來的邏輯作為備案
            self.logger.warning(f"⚠️ 未找到 {process_type} 的檔案名映射，使用預設格式")
            return f"converted_manifest_{process_type}.xml"
            
    def _convert_xml_content_with_projects(self, xml_content: str, project_mapping: Dict[str, str], 
                                        process_type: str) -> str:
        """
        使用專案映射表轉換 XML 內容中的 revision
        
        Args:
            xml_content: 原始 XML 內容
            project_mapping: 專案名稱到目標分支的映射
            process_type: 處理類型
            
        Returns:
            轉換後的 XML 內容
        """
        try:
            converted_content = xml_content
            conversion_count = 0
            
            # 解析 XML
            import xml.etree.ElementTree as ET
            root = ET.fromstring(xml_content)
            
            # 遍歷所有 project 元素
            for project in root.findall('project'):
                project_name = project.get('name', '')
                original_revision = project.get('revision', '')
                
                # 檢查是否需要轉換
                if project_name in project_mapping:
                    target_branch = project_mapping[project_name]
                    
                    # 檢查是否應該跳過轉換
                    if self._should_skip_project_conversion(project_name, process_type, False):
                        self.logger.debug(f"跳過專案 {project_name} 的 XML 轉換")
                        continue
                    
                    if target_branch and target_branch != original_revision:
                        # 進行字串替換
                        old_pattern = f'name="{project_name}"[^>]*revision="{original_revision}"'
                        new_revision_attr = f'revision="{target_branch}"'
                        
                        # 使用安全的替換方法
                        success = self._safe_replace_project_revision_in_xml(
                            converted_content, project_name, original_revision, target_branch
                        )
                        
                        if success:
                            converted_content = success
                            conversion_count += 1
                            self.logger.debug(f"XML 轉換: {project_name} - {original_revision} → {target_branch}")
            
            self.logger.info(f"XML 內容轉換完成，共轉換 {conversion_count} 個專案")
            return converted_content
            
        except Exception as e:
            self.logger.error(f"轉換 XML 內容失敗: {str(e)}")
            return xml_content

    def _safe_replace_project_revision_in_xml(self, xml_content: str, project_name: str, 
                                            old_revision: str, new_revision: str) -> str:
        """
        安全地替換 XML 中特定專案的 revision - 參考 feature_three.py
        """
        try:
            lines = xml_content.split('\n')
            modified = False
            
            for i, line in enumerate(lines):
                # 檢查這一行是否包含目標專案
                if f'name="{project_name}"' in line and 'revision=' in line:
                    # 找到目標行，進行替換
                    if f'revision="{old_revision}"' in line:
                        lines[i] = line.replace(f'revision="{old_revision}"', f'revision="{new_revision}"')
                        modified = True
                        self.logger.debug(f"✅ XML 替換成功: {project_name}")
                        break
                    elif f"revision='{old_revision}'" in line:
                        lines[i] = line.replace(f"revision='{old_revision}'", f"revision='{new_revision}'")
                        modified = True
                        self.logger.debug(f"✅ XML 替換成功 (單引號): {project_name}")
                        break
            
            if modified:
                return '\n'.join(lines)
            else:
                self.logger.warning(f"❌ 未找到匹配的 XML 替換: {project_name} - {old_revision}")
                return xml_content
                
        except Exception as e:
            self.logger.error(f"安全 XML 替換失敗: {str(e)}")
            return xml_content

    def _final_file_report(self, output_folder: str, saved_files: List[str]):
        """
        最終檔案檢查和報告 - 修正版：移除 process_type 依賴
        
        Args:
            output_folder: 輸出資料夾
            saved_files: 已保存的檔案列表
        """
        try:
            self.logger.info("🔍 最終檔案檢查報告:")
            self.logger.info(f"📂 輸出資料夾: {output_folder}")
            
            all_files_exist = True
            
            for file_path in saved_files:
                if file_path and os.path.exists(file_path):
                    file_size = os.path.getsize(file_path)
                    filename = os.path.basename(file_path)
                    self.logger.info(f"  ✅ {filename} ({file_size} bytes)")
                else:
                    filename = os.path.basename(file_path) if file_path else "未知檔案"
                    self.logger.error(f"  ❌ {filename} (檔案不存在)")
                    all_files_exist = False
            
            # 檢查輸出資料夾中的所有 XML 檔案
            self.logger.info(f"\n📋 Output 資料夾中的所有 XML 檔案:")
            xml_files_found = []
            try:
                for filename in os.listdir(output_folder):
                    if filename.lower().endswith('.xml'):
                        file_path = os.path.join(output_folder, filename)
                        file_size = os.path.getsize(file_path)
                        xml_files_found.append((filename, file_size))
                        self.logger.info(f"  📄 {filename} ({file_size} bytes)")
                
                if not xml_files_found:
                    self.logger.warning("  ⚠️ 沒有找到任何 XML 檔案")
                else:
                    self.logger.info(f"\n📊 XML 檔案統計:")
                    
                    gerrit_files = [f for f in xml_files_found if f[0].startswith('gerrit_')]
                    # 🔥 修改：由於現在轉換檔案使用目標檔名，檢查 atv-google-refplus 系列檔案
                    target_manifest_files = [f for f in xml_files_found 
                                        if f[0].startswith('atv-google-refplus') and not f[0].startswith('gerrit_')]
                    converted_files = [f for f in xml_files_found if f[0].startswith('converted_')]  # 保留舊格式檢查
                    
                    # 🔥 修改：原始檔案分類邏輯
                    original_files = [f for f in xml_files_found 
                                    if not f[0].startswith('gerrit_') 
                                    and not f[0].startswith('converted_')
                                    and not f[0].startswith('atv-google-refplus')]
                    
                    if original_files:
                        self.logger.info(f"  🟡 原始/來源檔案: {len(original_files)} 個")
                        for filename, size in original_files:
                            self.logger.info(f"    - {filename} ({size} bytes)")
                    
                    if gerrit_files:
                        self.logger.info(f"  🔵 Gerrit 檔案: {len(gerrit_files)} 個")
                        for filename, size in gerrit_files:
                            file_type = "(展開檔案)" if "_expanded" in filename else "(下載檔案)"
                            self.logger.info(f"    - {filename} ({size} bytes) {file_type}")
                    
                    # 🔥 新增：目標 manifest 檔案分類
                    if target_manifest_files:
                        self.logger.info(f"  🟢 目標 manifest 檔案: {len(target_manifest_files)} 個")
                        for filename, size in target_manifest_files:
                            self.logger.info(f"    - {filename} ({size} bytes)")
                    
                    if converted_files:
                        self.logger.info(f"  🟠 轉換檔案 (舊格式): {len(converted_files)} 個")
                        for filename, size in converted_files:
                            self.logger.info(f"    - {filename} ({size} bytes)")
                    
            except Exception as e:
                self.logger.error(f"  ❌ 無法列出資料夾內容: {str(e)}")
            
            # 總結
            if all_files_exist:
                self.logger.info(f"\n✅ 所有檔案都已成功保存")
                self.logger.info(f"🎯 檔案命名規則:")
                self.logger.info(f"   - 原始檔案: 保持原始檔名")
                self.logger.info(f"   - Gerrit 檔案: gerrit_*.xml")
                self.logger.info(f"   - 展開檔案: gerrit_*_expanded.xml")
                self.logger.info(f"   - 目標 manifest: atv-google-refplus-*.xml")  # 🔥 修正：移除 process_type
            else:
                self.logger.warning(f"\n⚠️ 部分檔案可能保存失敗，請檢查上述報告")
                
        except Exception as e:
            self.logger.error(f"檔案檢查報告失敗: {str(e)}")

    def _get_project_path_for_conversion(self, project_name: str, process_type: str) -> str:
        """
        取得專案的 path 屬性用於自定義轉換規則檢查 - Feature Two 版本
        
        Args:
            project_name: 專案名稱
            process_type: 處理類型
            
        Returns:
            專案的 path 屬性，如果找不到則返回空字串
        """
        try:
            # 🔥 從當前正在處理的專案列表中查找
            if hasattr(self, '_current_projects'):
                for project_info in self._current_projects:
                    if project_info.get('name') == project_name:
                        return project_info.get('path', '')
            
            self.logger.debug(f"無法找到專案 {project_name} 的 path 屬性")
            return ''
            
        except Exception as e:
            self.logger.error(f"取得專案 path 時發生錯誤: {str(e)}")
            return ''      

    def _should_skip_same_branch_name(self, project: Dict) -> Dict[str, Any]:
        """
        檢查來源分支和目標分支名稱是否相同，如果相同則跳過建立
        
        Args:
            project: 專案資訊字典
            
        Returns:
            字典包含 should_skip 和相關資訊
        """
        try:
            project_name = project.get('name', '')
            revision = project.get('revision', '')
            upstream = project.get('upstream', '')
            target_branch = project.get('target_branch', '')
            
            # 取得來源分支名稱
            source_branch = self._get_effective_source_branch_name(project)
            
            if not source_branch or not target_branch:
                return {
                    'should_skip': False,
                    'reason': '',
                    'source_branch': source_branch or 'N/A',
                    'target_branch': target_branch or 'N/A'
                }
            
            # 比較分支名稱
            if source_branch == target_branch:
                return {
                    'should_skip': True,
                    'reason': f'來源和目標分支相同: {source_branch}',
                    'source_branch': source_branch,
                    'target_branch': target_branch
                }
            
            return {
                'should_skip': False,
                'reason': '',
                'source_branch': source_branch,
                'target_branch': target_branch
            }
            
        except Exception as e:
            self.logger.error(f"檢查分支名稱相同性失敗 {project_name}: {str(e)}")
            return {
                'should_skip': False,
                'reason': f'檢查失敗: {str(e)}',
                'source_branch': 'Error',
                'target_branch': 'Error'
            }

    def _get_effective_source_branch_name(self, project: Dict) -> str:
        """
        取得有效的來源分支名稱
        
        Args:
            project: 專案資訊字典
            
        Returns:
            來源分支名稱
        """
        try:
            revision = project.get('revision', '')
            upstream = project.get('upstream', '')
            
            # 如果 revision 不是 hash，直接使用它作為分支名稱
            if revision and not self._is_revision_hash(revision):
                return revision.strip()
            
            # 如果 revision 是 hash，使用 upstream 作為分支名稱
            if self._is_revision_hash(revision) and upstream:
                return upstream.strip()
            
            # 如果都沒有有效值，返回空字串
            return ''
            
        except Exception as e:
            self.logger.error(f"取得來源分支名稱失敗: {str(e)}")
            return ''              