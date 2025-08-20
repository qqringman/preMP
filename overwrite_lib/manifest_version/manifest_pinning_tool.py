#!/usr/bin/env python3
"""
Manifest Pinning Tool - 自動化定版工具 (改進版 + Manifest 比較功能)
用於從 SFTP 下載 manifest 檔案並執行 repo 定版操作
改進版本：簡化 SFTP、改進報告格式、正常日誌輸出、即時 manifest 比較
"""

from enum import Enum
import os
import sys
import argparse
import pandas as pd
import paramiko
import subprocess
import json
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import time
import logging
from dataclasses import dataclass, field, asdict
import signal
import atexit
from contextlib import contextmanager
from collections import defaultdict
import random
import socket
from threading import Semaphore, Lock
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
from xml.dom import minidom
import difflib

# =====================================
# ===== 版本資訊 =====
# =====================================
__version__ = '2.2.0'
__author__ = 'Vince Lin'
__date__ = '2024-12-20'

# =====================================
# ===== XML 比較工具 =====
# =====================================

class ManifestComparator:
    """Manifest XML 檔案比較工具"""
    
    def __init__(self):
        self.logger = self._setup_logger()
    
    def _setup_logger(self):
        """設置獨立的 logger 避免重複日誌"""
        logger = logging.getLogger(f"{__name__}.ManifestComparator")
        # 清除已存在的 handlers 避免重複
        if logger.handlers:
            logger.handlers.clear()
        return logger
    
    def normalize_xml_element(self, element: ET.Element) -> dict:
        """
        標準化 XML 元素，將屬性轉換為有序字典
        用於比較時忽略屬性順序
        """
        result = {
            'tag': element.tag,
            'text': (element.text or '').strip(),
            'tail': (element.tail or '').strip(),
            'attrib': dict(sorted(element.attrib.items())),  # 排序屬性
            'children': []
        }
        
        # 遞迴處理子元素
        for child in element:
            result['children'].append(self.normalize_xml_element(child))
        
        # 按照標籤名和主要屬性排序子元素，確保比較的一致性
        result['children'].sort(key=lambda x: (
            x['tag'], 
            x['attrib'].get('name', ''),
            x['attrib'].get('path', ''),
            x['attrib'].get('revision', '')
        ))
        
        return result
    
    def compare_manifests(self, original_manifest: str, exported_manifest: str) -> dict:
        """
        比較兩個 manifest 檔案
        
        Args:
            original_manifest: 原始 manifest 檔案路徑
            exported_manifest: 導出的 manifest 檔案路徑
            
        Returns:
            比較結果字典，包含 is_identical, differences, summary 等信息
        """
        try:
            if not os.path.exists(original_manifest):
                return {
                    'is_identical': False,
                    'error': f'原始 manifest 檔案不存在: {original_manifest}',
                    'summary': '原始檔案不存在'
                }
            
            if not os.path.exists(exported_manifest):
                return {
                    'is_identical': False,
                    'error': f'導出的 manifest 檔案不存在: {exported_manifest}',
                    'summary': '導出檔案不存在'
                }
            
            # 解析 XML 檔案
            try:
                tree1 = ET.parse(original_manifest)
                root1 = tree1.getroot()
                
                tree2 = ET.parse(exported_manifest)
                root2 = tree2.getroot()
            except ET.ParseError as e:
                return {
                    'is_identical': False,
                    'error': f'XML 解析失敗: {str(e)}',
                    'summary': 'XML 格式錯誤'
                }
            
            # 標準化兩個 XML 根元素
            norm1 = self.normalize_xml_element(root1)
            norm2 = self.normalize_xml_element(root2)
            
            # 比較標準化後的結構
            differences = []
            is_identical = self._deep_compare(norm1, norm2, differences, '')
            
            # 生成統計摘要
            if is_identical:
                summary = "✅ 完全相同"
            else:
                diff_types = set()
                for diff in differences:
                    if 'project' in diff['path'].lower():
                        diff_types.add('專案差異')
                    elif 'remote' in diff['path'].lower():
                        diff_types.add('遠端設定差異')
                    elif 'default' in diff['path'].lower():
                        diff_types.add('預設設定差異')
                    else:
                        diff_types.add('其他差異')
                
                summary = f"❌ 有差異 ({len(differences)} 處): {', '.join(diff_types)}"
            
            # 讀取檔案大小信息
            original_size = os.path.getsize(original_manifest)
            exported_size = os.path.getsize(exported_manifest)
            
            result = {
                'is_identical': is_identical,
                'differences': differences,
                'summary': summary,
                'original_manifest': original_manifest,
                'exported_manifest': exported_manifest,
                'original_size': original_size,
                'exported_size': exported_size,
                'difference_count': len(differences)
            }
            
            # 記錄比較結果
            orig_name = os.path.basename(original_manifest)
            exp_name = os.path.basename(exported_manifest)
            self.logger.info(f"Manifest 比較完成: {orig_name} vs {exp_name}")
            self.logger.info(f"結果: {summary}")
            
            if not is_identical and len(differences) <= 10:  # 只顯示少量差異的詳細信息
                for i, diff in enumerate(differences[:5], 1):
                    self.logger.debug(f"差異 {i}: {diff['path']} - {diff['type']}")
                    if len(differences) > 5:
                        self.logger.debug(f"... 還有 {len(differences)-5} 個差異")
            
            return result
            
        except Exception as e:
            error_msg = f"Manifest 比較過程發生錯誤: {str(e)}"
            self.logger.error(error_msg)
            return {
                'is_identical': False,
                'error': error_msg,
                'summary': '比較過程出錯'
            }
    
    def _deep_compare(self, obj1: dict, obj2: dict, differences: list, path: str = '') -> bool:
        """
        深度比較兩個標準化的 XML 物件
        
        Args:
            obj1, obj2: 要比較的物件
            differences: 用於收集差異的列表
            path: 當前比較的路徑（用於錯誤報告）
            
        Returns:
            是否完全相同
        """
        is_identical = True
        
        # 比較標籤名
        if obj1['tag'] != obj2['tag']:
            differences.append({
                'path': path,
                'type': 'tag_different',
                'expected': obj1['tag'],
                'actual': obj2['tag']
            })
            is_identical = False
        
        # 比較文本內容
        if obj1['text'] != obj2['text']:
            differences.append({
                'path': f"{path}/text",
                'type': 'text_different',
                'expected': obj1['text'],
                'actual': obj2['text']
            })
            is_identical = False
        
        # 比較屬性
        if obj1['attrib'] != obj2['attrib']:
            # 檢查缺失的屬性
            for key in obj1['attrib']:
                if key not in obj2['attrib']:
                    differences.append({
                        'path': f"{path}/@{key}",
                        'type': 'attribute_missing',
                        'expected': obj1['attrib'][key],
                        'actual': None
                    })
                    is_identical = False
                elif obj1['attrib'][key] != obj2['attrib'][key]:
                    differences.append({
                        'path': f"{path}/@{key}",
                        'type': 'attribute_different',
                        'expected': obj1['attrib'][key],
                        'actual': obj2['attrib'][key]
                    })
                    is_identical = False
            
            # 檢查額外的屬性
            for key in obj2['attrib']:
                if key not in obj1['attrib']:
                    differences.append({
                        'path': f"{path}/@{key}",
                        'type': 'attribute_extra',
                        'expected': None,
                        'actual': obj2['attrib'][key]
                    })
                    is_identical = False
        
        # 比較子元素數量
        if len(obj1['children']) != len(obj2['children']):
            differences.append({
                'path': f"{path}/children_count",
                'type': 'children_count_different',
                'expected': len(obj1['children']),
                'actual': len(obj2['children'])
            })
            is_identical = False
        
        # 比較子元素（配對比較）
        min_children = min(len(obj1['children']), len(obj2['children']))
        for i in range(min_children):
            child1 = obj1['children'][i]
            child2 = obj2['children'][i]
            
            # 構建子路徑
            child_path = f"{path}/{child1['tag']}"
            if 'name' in child1['attrib']:
                child_path += f"[@name='{child1['attrib']['name']}']"
            elif 'path' in child1['attrib']:
                child_path += f"[@path='{child1['attrib']['path']}']"
            else:
                child_path += f"[{i}]"
            
            # 遞迴比較子元素
            child_identical = self._deep_compare(child1, child2, differences, child_path)
            if not child_identical:
                is_identical = False
        
        return is_identical
    
    def generate_diff_report(self, comparison_result: dict, output_file: str = None) -> str:
        """
        生成詳細的差異報告
        
        Args:
            comparison_result: compare_manifests 的結果
            output_file: 輸出檔案路徑（可選）
            
        Returns:
            差異報告的文字內容
        """
        if comparison_result.get('error'):
            report = f"比較失敗: {comparison_result['error']}"
            if output_file:
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write(report)
            return report
        
        lines = []
        lines.append("="*80)
        lines.append("Manifest 比較報告")
        lines.append("="*80)
        lines.append(f"原始檔案: {comparison_result['original_manifest']}")
        lines.append(f"導出檔案: {comparison_result['exported_manifest']}")
        lines.append(f"原始大小: {comparison_result['original_size']:,} bytes")
        lines.append(f"導出大小: {comparison_result['exported_size']:,} bytes")
        lines.append(f"比較時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")
        
        if comparison_result['is_identical']:
            lines.append("✅ 結果: 兩個 manifest 檔案完全相同")
        else:
            lines.append(f"❌ 結果: 發現 {len(comparison_result['differences'])} 處差異")
            lines.append("")
            lines.append("差異詳情:")
            lines.append("-" * 60)
            
            for i, diff in enumerate(comparison_result['differences'], 1):
                lines.append(f"{i:3d}. 路徑: {diff['path']}")
                lines.append(f"     類型: {diff['type']}")
                lines.append(f"     預期: {diff.get('expected', 'N/A')}")
                lines.append(f"     實際: {diff.get('actual', 'N/A')}")
                lines.append("")
        
        lines.append("="*80)
        
        report = "\n".join(lines)
        
        if output_file:
            try:
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write(report)
                self.logger.info(f"差異報告已保存至: {output_file}")
            except Exception as e:
                self.logger.error(f"保存差異報告失敗: {e}")
        
        return report

# =====================================
# ===== 配置管理器 =====
# =====================================

class ConfigManager:
    """配置管理器 - 統一管理所有配置的優先級和來源"""
    
    def __init__(self):
        # 基礎配置
        self.sftp_config = {
            'host': 'mmsftpx.realtek.com',
            'port': 22,
            'username': 'lgwar_user',
            'password': 'Ab!123456',
            'timeout': 30,
            'retry_count': 3,
            'retry_delay': 2
        }
        
        self.jira_config = {
            'site': 'jira.realtek.com',
            'username': 'vince_lin',
            'password': 'Amon200!Amon200!',
            'api_url': 'https://jira.realtek.com/rest/api/2'
        }
        
        self.path_config = {
            'default_output_dir': './DB-source',
            'default_mapping_table': '../all_chip_mapping_table.xlsx',
            'manifest_pattern': r'manifest_(\d+)\.xml',
            'report_filename': 'pinning_report.xlsx'
        }
        
        self.repo_config = {
            'repo_command': 'repo',
            'sync_jobs': 8,
            'sync_retry': 2,
            'init_timeout': 300,
            'sync_timeout': 7200,
            'async_sync': True,
            'show_sync_output': False,
            # 🔥 新增清理相關配置
            'check_clean_before_sync': True,  # 是否在sync前檢查乾淨狀態
            'auto_clean_workspace': True,     # 是否自動清理工作空間
            'backup_local_changes': True,     # 是否備份本地修改
            'force_clean_on_dirty': True     # 遇到髒repo是否強制清理
        }
        
        self.parallel_config = {
            'max_workers': 4,
            'enable_parallel': True,
        }
        
        self.log_config = {
            'level': logging.INFO,
            'format': '%(asctime)s - [%(levelname)s] - %(name)s - %(message)s',
            'date_format': '%Y-%m-%d %H:%M:%S'
        }
        
        # 預設執行配置
        self.default_execution_config = {
            'mapping_table': os.path.join(os.path.dirname(__file__), '..', 'all_chip_mapping_table.xlsx'),
            'db_type': 'all',
            'selected_dbs': ['DB2145', 'DB2858', 'DB2575', 'DB2919'],
            'db_versions': {},
            'output_dir': './DB-source',
            'auto_confirm': False,
            'sftp_override': {}
        }
        
        # 配置優先級追蹤
        self.config_sources = {}
        
    def apply_overrides(self, overrides: Dict[str, Any], source: str = 'user'):
        """應用配置覆蓋並記錄來源"""
        for key, value in overrides.items():
            if value is not None:
                self.config_sources[key] = source
                # 根據 key 更新對應的配置
                if key.startswith('sftp_'):
                    sftp_key = key.replace('sftp_', '')
                    if sftp_key in self.sftp_config:
                        self.sftp_config[sftp_key] = value
                elif key.startswith('repo_'):
                    repo_key = key.replace('repo_', '')
                    if repo_key in self.repo_config:
                        self.repo_config[repo_key] = value
                elif key.startswith('parallel_'):
                    parallel_key = key.replace('parallel_', '')
                    if parallel_key in self.parallel_config:
                        self.parallel_config[parallel_key] = value

# 全域配置管理器實例
config_manager = ConfigManager()

# =====================================
# ===== JIRA API 客戶端 =====
# =====================================

from jira import JIRA
import requests
from urllib.parse import quote
import urllib3

# 關閉 SSL 警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class JiraAPIClient:
    """簡化版 JIRA API 客戶端"""
    
    def __init__(self):
        self.logger = setup_logger(self.__class__.__name__)
        self.session = None
        self._connected = False
        self.base_url = f"https://{config_manager.jira_config['site']}"
        
    def connect(self) -> bool:
        """連接到 JIRA"""
        try:
            username = config_manager.jira_config['username']
            password = config_manager.jira_config['password']
            
            self.session = requests.Session()
            self.session.auth = (username, password)
            self.session.headers.update({
                'Accept': 'application/json',
                'Content-Type': 'application/json'
            })
            self.session.verify = False
            
            # 測試連接
            test_url = f"{self.base_url}/rest/api/2/myself"
            response = self.session.get(test_url, timeout=30)
            
            if response.status_code == 200:
                self._connected = True
                return True
            else:
                self.logger.error(f"JIRA 連接失敗: HTTP {response.status_code}")
                return False
                
        except Exception as e:
            self.logger.error(f"JIRA 連接失敗: {e}")
            return False

    def search_db_ticket(self, db_name: str, module: str = None) -> Optional[str]:
        """根據 DB 命名慣例搜尋對應的 JIRA ticket"""
        try:
            if not self._connected:
                if not self.connect():
                    return None
            
            # 根據命名慣例直接構建 ticket key
            db_number = db_name.replace('DB', '')
            possible_tickets = [
                f"MMQCDB-{db_number}",
                f"LGSWRD-{db_number}",
                f"RTK-{db_number}",
                f"DB-{db_number}",
            ]
            
            for ticket_key in possible_tickets:
                if self._check_ticket_exists(ticket_key):
                    return ticket_key
            
            return None
            
        except Exception as e:
            self.logger.error(f"搜尋 JIRA ticket 失敗: {e}")
            return None

    def _check_ticket_exists(self, ticket_key: str) -> bool:
        """檢查指定的 ticket 是否存在"""
        try:
            url = f"{self.base_url}/rest/api/2/issue/{ticket_key}"
            response = self.session.get(url, timeout=30)
            return response.status_code == 200
        except:
            return False

    def get_source_command_from_ticket(self, ticket_key: str) -> Optional[str]:
        """從 JIRA ticket 中提取 source command"""
        try:
            if not self._connected:
                if not self.connect():
                    return None
            
            url = f"{self.base_url}/rest/api/2/issue/{ticket_key}"
            response = self.session.get(url, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                fields = data.get('fields', {})
                
                # 檢查描述欄位
                description = fields.get('description', '')
                if description:
                    cmd = self._extract_repo_command(description)
                    if cmd:
                        return cmd
                
                # 檢查評論
                comments_data = fields.get('comment', {})
                comments = comments_data.get('comments', [])
                for comment in comments:
                    body = comment.get('body', '')
                    if body:
                        cmd = self._extract_repo_command(body)
                        if cmd:
                            return cmd
                
                return None
            
            return None
                
        except Exception as e:
            self.logger.error(f"從 ticket {ticket_key} 獲取 source command 失敗: {e}")
            return None

    def _extract_repo_command(self, text: str) -> Optional[str]:
        """提取 repo init 命令 - 優先選擇有 -m 參數的指令"""
        if not text or 'repo init' not in text:
            return None
        
        lines = text.split('\n')
        found_commands = []
        
        for line in lines:
            line = line.strip()
            
            # 尋找包含 repo init 的行
            if 'repo init' in line:
                repo_start = line.find('repo init')
                if repo_start != -1:
                    cmd = line[repo_start:].strip()
                    cmd = self._clean_command(cmd)
                    if self._is_valid_repo_command(cmd):
                        found_commands.append(cmd)
        
        if not found_commands:
            return None
        
        # 🔥 優先選擇有 "-m" 參數的指令
        commands_with_m = [cmd for cmd in found_commands if '-m ' in cmd]
        if commands_with_m:
            # 如果有多個，取第一個
            return commands_with_m[0]
        
        # 如果沒有 -m 參數的，返回第一個有效的
        return found_commands[0]

    def _clean_command(self, cmd: str) -> str:
        """清理命令字符串"""
        # 移除常見的前綴
        prefixes = ['$', '>', '#', '//', '*', '-', '=>']
        for prefix in prefixes:
            if cmd.startswith(prefix):
                cmd = cmd[len(prefix):].strip()
        
        cmd = cmd.rstrip('\n\r\\')
        return cmd.strip()

    def _is_valid_repo_command(self, cmd: str) -> bool:
        """驗證是否為有效的 repo init 命令"""
        if not cmd or not cmd.startswith('repo init'):
            return False
        
        required_parts = ['-u', '-b']
        return all(part in cmd for part in required_parts)

    def disconnect(self):
        """斷開 JIRA 連接"""
        self._connected = False
        if self.session:
            self.session.close()
            self.session = None

# =====================================
# ===== 日誌設定函式 =====
# =====================================

def setup_logger(name: str = __name__) -> logging.Logger:
    """設定日誌記錄器 - 修復重複日誌問題"""
    logger = logging.getLogger(name)
    
    # 🔥 修復：如果 logger 已經有 handlers，先清除避免重複
    if logger.handlers:
        logger.handlers.clear()
    
    logger.setLevel(config_manager.log_config['level'])
    
    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(config_manager.log_config['level'])
    
    # Formatter
    formatter = logging.Formatter(
        config_manager.log_config['format'],
        datefmt=config_manager.log_config['date_format']
    )
    console_handler.setFormatter(formatter)
    
    logger.addHandler(console_handler)
    
    # 🔥 防止日誌向上傳播造成重複
    logger.propagate = False
    
    return logger

logger = setup_logger(__name__)

# =====================================
# ===== 進度狀態定義 =====
# =====================================

class DBStatus(Enum):
    """DB 處理狀態"""
    PENDING = "等待中"
    DOWNLOADING_MANIFEST = "下載 manifest"
    CHECKING_REPO = "檢查 repo 狀態"
    REPO_INIT = "執行 repo init"
    REPO_SYNC = "執行 repo sync"
    EXPORTING = "導出版本"
    COMPARING = "比較 manifest"  # 🔥 新增比較狀態
    SUCCESS = "✅ 完成"
    SUCCESS_WITH_DIFF = "✅ 完成(有差異)"  # 🔥 新增有差異但成功的狀態
    FAILED = "❌ 失敗"
    SKIPPED = "⭐️ 跳過"

# =====================================
# ===== 資料結構定義 =====
# =====================================

@dataclass
class DBInfo:
    """DB 資訊資料結構 - 新增 manifest 比較相關欄位"""
    sn: int
    module: str
    db_type: str
    db_info: str
    db_folder: str
    sftp_path: str
    version: Optional[str] = None
    jira_link: Optional[str] = None
    source_command: Optional[str] = None
    manifest_file: Optional[str] = None
    manifest_full_path: Optional[str] = None
    local_path: Optional[str] = None
    has_existing_repo: bool = False
    status: DBStatus = DBStatus.PENDING
    error_message: Optional[str] = None
    start_time: Optional[datetime] = None
    end_time: Optional[datetime] = None
    sync_process: Optional[subprocess.Popen] = None
    actual_source_cmd: Optional[str] = None
    sync_log_path: Optional[str] = None
    
    # 🔥 新增 manifest 比較相關欄位
    exported_manifest_path: Optional[str] = None
    manifest_comparison_result: Optional[dict] = None
    manifest_is_identical: Optional[bool] = None
    manifest_differences_count: Optional[int] = None
    manifest_comparison_summary: Optional[str] = None

    def to_dict(self) -> dict:
        """轉換為字典格式"""
        result = asdict(self)
        
        # 處理 Enum 和 datetime
        if isinstance(result['status'], DBStatus):
            result['status'] = result['status'].value
        if result['start_time']:
            result['start_time'] = result['start_time'].strftime('%Y-%m-%d %H:%M:%S')
        if result['end_time']:
            result['end_time'] = result['end_time'].strftime('%Y-%m-%d %H:%M:%S')
        
        # 移除無法序列化的物件
        result.pop('sync_process', None)
        result.pop('manifest_comparison_result', None)  # 比較結果太複雜，不序列化
        
        return result

@dataclass
class PinningReport:
    """定版報告資料結構"""
    total_dbs: int = 0
    successful_dbs: int = 0
    failed_dbs: int = 0
    skipped_dbs: int = 0
    dbs_with_differences: int = 0  # 🔥 新增：有差異的 DB 數量
    db_details: List[DBInfo] = field(default_factory=list)
    start_time: datetime = field(default_factory=datetime.now)
    end_time: Optional[datetime] = None
    
    def add_db(self, db_info: DBInfo):
        # 🔥 檢查是否已存在相同的記錄
        for i, existing_db in enumerate(self.db_details):
            if (existing_db.db_info == db_info.db_info and 
                existing_db.version == db_info.version and
                existing_db.db_type == db_info.db_type and
                existing_db.module == db_info.module):
                # 更新現有記錄而不是添加新的
                self.db_details[i] = db_info
                self._recalculate_stats()
                return
        
        # 如果不存在，才添加新記錄
        self.db_details.append(db_info)
        self._update_stats(db_info)

    def _recalculate_stats(self):
        """重新計算統計數據"""
        self.successful_dbs = sum(1 for db in self.db_details 
                                if db.status in [DBStatus.SUCCESS, DBStatus.SUCCESS_WITH_DIFF])
        self.failed_dbs = sum(1 for db in self.db_details if db.status == DBStatus.FAILED)
        self.skipped_dbs = sum(1 for db in self.db_details if db.status == DBStatus.SKIPPED)
        self.dbs_with_differences = sum(1 for db in self.db_details if db.status == DBStatus.SUCCESS_WITH_DIFF)

    def _update_stats(self, db_info: DBInfo):
        """更新統計數據（新增記錄時）"""
        if db_info.status == DBStatus.SUCCESS:
            self.successful_dbs += 1
        elif db_info.status == DBStatus.SUCCESS_WITH_DIFF:
            self.successful_dbs += 1
            self.dbs_with_differences += 1
        elif db_info.status == DBStatus.FAILED:
            self.failed_dbs += 1
        elif db_info.status == DBStatus.SKIPPED:
            self.skipped_dbs += 1
                
    def finalize(self):
        """完成報告"""
        self.end_time = datetime.now()
        self.total_dbs = len(self.db_details)

# =====================================
# ===== 資源管理器 =====
# =====================================

class ResourceManager:
    """統一管理所有系統資源"""
    
    def __init__(self):
        self.active_processes = {}
        self.sftp_connections = []
        self.lock = threading.Lock()
        self.logger = setup_logger(self.__class__.__name__)
        
        atexit.register(self.cleanup_all)
        signal.signal(signal.SIGINT, self._signal_handler)
        signal.signal(signal.SIGTERM, self._signal_handler)
    
    def _signal_handler(self, signum, frame):
        """處理中斷信號"""
        print(f"\n🛑 收到中斷信號，清理所有進程...")
        
        # 原有的清理
        self.cleanup_all()
        
        # 🔥 新增：系統級強制清理
        print("🚨 執行系統級清理...")
        os.system("pkill -TERM -f 'repo sync' 2>/dev/null || true")
        os.system("pkill -TERM -f 'unbuffer.*repo' 2>/dev/null || true")
        time.sleep(2)
        os.system("pkill -KILL -f 'repo sync' 2>/dev/null || true")
        os.system("pkill -KILL -f 'unbuffer.*repo' 2>/dev/null || true")
        
        print("✅ 清理完成")
        sys.exit(0)
    
    def register_process(self, name: str, process: subprocess.Popen):
        """註冊新的子進程"""
        with self.lock:
            self.active_processes[name] = process
    
    def unregister_process(self, name: str):
        """取消註冊子進程"""
        with self.lock:
            if name in self.active_processes:
                del self.active_processes[name]
    
    def register_sftp(self, connection):
        """註冊 SFTP 連線"""
        with self.lock:
            self.sftp_connections.append(connection)
    
    def cleanup_all(self):
        """清理所有資源"""
        with self.lock:
            # 終止所有子進程
            for name, process in self.active_processes.items():
                try:
                    if process.poll() is None:
                        process.terminate()
                        process.wait(timeout=5)
                except:
                    try:
                        process.kill()
                    except:
                        pass
            
            # 關閉所有 SFTP 連線
            for conn in self.sftp_connections:
                try:
                    if hasattr(conn, 'disconnect'):
                        conn.disconnect()
                    elif hasattr(conn, 'close'):
                        conn.close()
                except:
                    pass
            
            self.active_processes.clear()
            self.sftp_connections.clear()

# 全域資源管理器
resource_manager = ResourceManager()

# =====================================
# ===== 改進版 SFTP 管理器（修復 Garbage packet 問題）=====
# =====================================

class SFTPManager:
    """改進版 SFTP 管理器 - 修復 type 3 unimplemented 錯誤"""
    
    def __init__(self):
        self.logger = setup_logger(self.__class__.__name__)
        self.config = config_manager.sftp_config
        self.sftp = None
        self.transport = None
        self.connected = False
        self._connection_lock = threading.Lock()
        # 快取伺服器能力
        self._server_capabilities = {
            'supports_listdir_attr': None,
            'supports_stat': None,
            'checked': False
        }
        
    def connect(self) -> bool:
        """建立 SFTP 連線"""
        with self._connection_lock:
            try:
                # 如果已連線，先清理
                if self.connected:
                    self.disconnect()
                
                self.logger.info(f"建立 SFTP 連線: {self.config['host']}:{self.config['port']}")
                
                # 重試連線
                for attempt in range(self.config['retry_count']):
                    try:
                        # 建立 Socket 連線
                        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                        sock.settimeout(self.config['timeout'])
                        
                        # 設定 socket 選項避免 Garbage packet
                        sock.setsockopt(socket.SOL_SOCKET, socket.SO_KEEPALIVE, 1)
                        sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
                        
                        # 連接到伺服器
                        sock.connect((self.config['host'], self.config['port']))
                        
                        # 建立 Transport（修復 Garbage packet 的關鍵）
                        self.transport = paramiko.Transport(sock)
                        
                        # 設定較保守的參數避免協議問題
                        self.transport.set_keepalive(30)
                        self.transport.use_compression(False)  # 關閉壓縮
                        
                        # 開始 SSH 握手
                        self.transport.start_client()
                        
                        # 認證
                        self.transport.auth_password(
                            self.config['username'],
                            self.config['password']
                        )
                        
                        # 建立 SFTP 客戶端
                        self.sftp = paramiko.SFTPClient.from_transport(self.transport)
                        
                        # 測試連線並檢測伺服器能力
                        self._detect_server_capabilities()
                        
                        self.connected = True
                        resource_manager.register_sftp(self)
                        self.logger.info(f"SFTP 連線成功 (嘗試 {attempt + 1})")
                        return True
                        
                    except Exception as e:
                        self.logger.warning(f"SFTP 連線嘗試 {attempt + 1} 失敗: {e}")
                        self._cleanup_failed_connection()
                        
                        if attempt < self.config['retry_count'] - 1:
                            time.sleep(self.config['retry_delay'])
                
                self.logger.error("SFTP 連線失敗，已達最大重試次數")
                return False
                
            except Exception as e:
                self.logger.error(f"SFTP 連線過程發生錯誤: {e}")
                self._cleanup_failed_connection()
                return False
    
    def _detect_server_capabilities(self):
        """檢測 SFTP 伺服器支援的功能"""
        try:
            self.logger.debug("檢測 SFTP 伺服器能力...")
            
            # 測試基本 listdir
            try:
                self.sftp.listdir('.')
                self.logger.debug("✅ 基本 listdir 支援")
            except Exception as e:
                self.logger.warning(f"❌ 基本 listdir 不支援: {e}")
                raise Exception("伺服器不支援基本 SFTP 操作")
            
            # 測試 listdir_attr
            try:
                self.sftp.listdir_attr('.')
                self._server_capabilities['supports_listdir_attr'] = True
                self.logger.debug("✅ listdir_attr 支援")
            except Exception as e:
                self._server_capabilities['supports_listdir_attr'] = False
                self.logger.debug(f"❌ listdir_attr 不支援: {e}")
            
            # 測試 stat
            try:
                self.sftp.stat('.')
                self._server_capabilities['supports_stat'] = True
                self.logger.debug("✅ stat 支援")
            except Exception as e:
                self._server_capabilities['supports_stat'] = False
                self.logger.debug(f"❌ stat 不支援: {e}")
            
            self._server_capabilities['checked'] = True
            self.logger.info(f"伺服器能力檢測完成: listdir_attr={self._server_capabilities['supports_listdir_attr']}, stat={self._server_capabilities['supports_stat']}")
            
        except Exception as e:
            self.logger.warning(f"伺服器能力檢測失敗: {e}")
            # 設定為最保守的模式
            self._server_capabilities = {
                'supports_listdir_attr': False,
                'supports_stat': False,
                'checked': True
            }
    
    def _cleanup_failed_connection(self):
        """清理失敗的連線"""
        try:
            if self.sftp:
                self.sftp.close()
                self.sftp = None
            if self.transport:
                self.transport.close()
                self.transport = None
            self.connected = False
            self._server_capabilities['checked'] = False
        except:
            pass
    
    def disconnect(self):
        """安全斷開 SFTP 連線"""
        with self._connection_lock:
            try:
                if self.sftp:
                    self.sftp.close()
                    self.sftp = None
                if self.transport:
                    self.transport.close()
                    self.transport = None
                self.connected = False
                self._server_capabilities['checked'] = False
                self.logger.info("SFTP 連線已關閉")
            except Exception as e:
                self.logger.warning(f"關閉 SFTP 連線時發生錯誤: {e}")
    
    def _ensure_connected(self) -> bool:
        """確保連線有效"""
        if not self.connected or not self.sftp or not self.transport:
            return self.connect()
        
        # 測試連線是否仍然有效
        try:
            self.sftp.listdir('.')
            return True
        except:
            self.logger.info("SFTP 連線已斷開，重新連線...")
            return self.connect()
    
    def get_latest_version(self, sftp_path: str) -> Tuple[str, str, str]:
        """
        取得最新版本資訊（版號最大的）
        
        Args:
            sftp_path: SFTP 路徑
            
        Returns:
            (db_folder, db_version, full_path)
        """
        try:
            if not self._ensure_connected():
                raise Exception("無法建立 SFTP 連線")
            
            # 從路徑中提取基礎資訊
            path_parts = sftp_path.rstrip('/').split('/')
            db_folder = path_parts[-1] if path_parts else ''
            
            # 列出目錄內容
            try:
                items = self._safe_listdir(sftp_path)
            except:
                self.logger.warning(f"無法列出目錄: {sftp_path}")
                return db_folder, '', sftp_path
            
            # 過濾版本資料夾（格式: 數字開頭，如 536_all_202507312300）
            version_folders = []
            for item in items:
                # 檢查是否為版本資料夾格式
                match = re.match(r'^(\d+)', item)
                if match:
                    version_num = int(match.group(1))
                    version_folders.append((version_num, item))
            
            if not version_folders:
                self.logger.warning(f"找不到版本資料夾: {sftp_path}")
                return db_folder, '', sftp_path
            
            # 按版本號排序，取最大的（最新的）
            version_folders.sort(key=lambda x: x[0], reverse=True)
            latest_version = version_folders[0][1]
            full_path = f"{sftp_path}/{latest_version}"
            
            self.logger.info(f"找到最新版本: {latest_version} (版號: {version_folders[0][0]}) in {sftp_path}")
            return db_folder, latest_version, full_path
            
        except Exception as e:
            self.logger.error(f"取得版本資訊失敗: {str(e)}")
            return '', '', sftp_path
    
    def get_specific_version(self, sftp_path: str, db_info: str) -> Tuple[str, str, str]:
        """
        取得特定版本資訊
        
        Args:
            sftp_path: SFTP 基礎路徑
            db_info: DB資訊 (格式: DB2302#196)
            
        Returns:
            (db_folder, db_version, full_path)
        """
        try:
            if not self._ensure_connected():
                raise Exception("無法建立 SFTP 連線")
            
            # 解析 db_info
            if '#' not in db_info:
                self.logger.warning(f"DB資訊格式錯誤: {db_info}")
                return self.get_latest_version(sftp_path)
            
            db_number, version_prefix = db_info.split('#')
            
            # 列出目錄找到對應的 DB 資料夾
            parent_path = '/'.join(sftp_path.rstrip('/').split('/')[:-1])
            items = self._safe_listdir(parent_path)
            
            db_folder = None
            for item in items:
                if item.startswith(f"{db_number}_"):
                    db_folder = item
                    break
            
            if not db_folder:
                self.logger.warning(f"找不到 DB 資料夾: {db_number}")
                return '', '', sftp_path
            
            # 建構完整路徑
            db_path = f"{parent_path}/{db_folder}"
            
            # 列出版本資料夾
            version_items = self._safe_listdir(db_path)
            
            # 找到對應版本
            for version_item in version_items:
                if version_item.startswith(f"{version_prefix}_"):
                    full_path = f"{db_path}/{version_item}"
                    self.logger.info(f"找到指定版本: {version_item}")
                    return db_folder, version_item, full_path
            
            self.logger.warning(f"找不到指定版本: {version_prefix}")
            return db_folder, '', db_path
            
        except Exception as e:
            self.logger.error(f"取得特定版本失敗: {str(e)}")
            return '', '', sftp_path
    
    def find_latest_manifest(self, path: str, db_name: str = None, target_version: str = None) -> Optional[Tuple[str, str]]:
        """改進版：快速搜尋 manifest 檔案（強化錯誤處理）"""
        try:
            if not self._ensure_connected():
                raise Exception("無法建立 SFTP 連線")
            
            if target_version:
                self.logger.info(f"快速搜尋指定版本 {target_version}: {path}")
                return self._find_specific_version_manifest(path, target_version)
            else:
                self.logger.info(f"快速搜尋最新版本: {path}")
                return self._find_latest_version_manifest(path)
                
        except Exception as e:
            self.logger.error(f"搜索 manifest 失敗: {e}")
            return None

    def _find_latest_version_manifest(self, base_path: str) -> Optional[Tuple[str, str]]:
        """搜尋最新版本的 manifest - 強化 type 3 錯誤處理"""
        try:
            # 列出目錄
            version_dirs = []
            try:
                items = self._safe_listdir_with_details(base_path)
                
                for item_info in items:
                    try:
                        if item_info.get('is_dir', False):
                            dir_name = item_info['name']
                            version_num = self._extract_version_number(dir_name)
                            if version_num:
                                version_dirs.append({
                                    'name': dir_name,
                                    'version': int(version_num),
                                    'mtime': item_info.get('mtime', 0)
                                })
                    except Exception as e:
                        self.logger.debug(f"跳過問題目錄: {item_info.get('name', 'unknown')}, 錯誤: {e}")
                        continue
                        
            except Exception as e:
                self.logger.error(f"列出目錄失敗: {e}")
                return None
            
            if not version_dirs:
                raise Exception(f"在 {base_path} 中沒有找到版本目錄")
            
            # 按版本號降序排列（最新版本在前）
            version_dirs.sort(key=lambda x: (x['version'], x['mtime']), reverse=True)
            self.logger.info(f"找到 {len(version_dirs)} 個版本目錄，最新版本: {version_dirs[0]['version']}")
            
            # 只檢查前3個最新版本
            for version_dir in version_dirs[:3]:
                self.logger.info(f"檢查版本 {version_dir['version']} ({version_dir['name']})")
                manifest_path = self._try_get_manifest_from_dir(
                    base_path, 
                    version_dir['name'], 
                    str(version_dir['version'])
                )
                if manifest_path:
                    return manifest_path
            
            raise Exception("在最新版本目錄中找不到有效的 manifest")
            
        except Exception as e:
            self.logger.error(f"搜尋最新版本失敗: {e}")
            return None
        
    def _find_specific_version_manifest(self, base_path: str, version: str) -> Optional[Tuple[str, str]]:
        """搜尋指定版本的 manifest - 強化 type 3 錯誤處理"""
        try:
            self.logger.info(f"搜尋指定版本: {version}")
            
            # 嘗試列出目錄，如果失敗就用直接路徑方式
            matching_dirs = []
            try:
                items = self._safe_listdir_with_details(base_path)
                
                for item_info in items:
                    try:
                        # 檢查是否為目錄
                        if item_info.get('is_dir', False):
                            dir_name = item_info['name']
                            if self._matches_version(dir_name, version):
                                matching_dirs.append({
                                    'name': dir_name,
                                    'mtime': item_info.get('mtime', 0)
                                })
                                self.logger.debug(f"找到匹配版本目錄: {dir_name}")
                    except Exception as e:
                        self.logger.debug(f"跳過問題項目: {item_info.get('name', 'unknown')}, 錯誤: {e}")
                        continue
                        
            except Exception as e:
                self.logger.warning(f"列出目錄失敗，嘗試直接路徑: {e}")
                return self._try_direct_version_paths(base_path, version)
            
            if not matching_dirs:
                # 如果沒找到，嘗試直接構建路徑
                return self._try_direct_version_paths(base_path, version)
            
            # 按時間排序，取最新的
            matching_dirs.sort(key=lambda x: x['mtime'], reverse=True)
            
            # 嘗試每個匹配的版本目錄
            for dir_info in matching_dirs:
                manifest_path = self._try_get_manifest_from_dir(base_path, dir_info['name'], version)
                if manifest_path:
                    return manifest_path
            
            # 如果匹配目錄都失敗，嘗試直接路徑
            return self._try_direct_version_paths(base_path, version)
            
        except Exception as e:
            self.logger.error(f"搜尋指定版本失敗: {e}")
            return None

    def _safe_listdir(self, path: str) -> list:
        """最安全的列目錄方法 - 只使用基本 listdir"""
        try:
            return self.sftp.listdir(path)
        except Exception as e:
            self.logger.error(f"列出目錄失敗: {path}, 錯誤: {e}")
            raise
    
    def _safe_listdir_with_details(self, path: str) -> list:
        """安全的列出目錄內容，根據伺服器能力選擇方法"""
        try:
            result = []
            
            # 如果支援 listdir_attr，優先使用
            if self._server_capabilities.get('supports_listdir_attr', False):
                try:
                    items = self.sftp.listdir_attr(path)
                    for item in items:
                        result.append({
                            'name': item.filename,
                            'is_dir': self._is_directory_from_stat(item),
                            'mtime': getattr(item, 'st_mtime', 0),
                            'size': getattr(item, 'st_size', 0)
                        })
                    return result
                except Exception as e:
                    error_msg = str(e).lower()
                    if 'unimplemented' in error_msg or 'type 3' in error_msg:
                        self.logger.warning(f"listdir_attr 不支援，切換到基本模式: {e}")
                        self._server_capabilities['supports_listdir_attr'] = False
                    else:
                        raise
            
            # 回退到基本 listdir + 個別 stat
            self.logger.debug("使用基本 listdir 模式")
            filenames = self.sftp.listdir(path)
            
            for filename in filenames:
                item_info = {
                    'name': filename,
                    'is_dir': False,  # 預設為檔案
                    'mtime': 0,
                    'size': 0
                }
                
                # 嘗試獲取詳細資訊
                if self._server_capabilities.get('supports_stat', True):
                    try:
                        full_path = f"{path}/{filename}"
                        stat_info = self.sftp.stat(full_path)
                        item_info.update({
                            'is_dir': self._is_directory_from_stat(stat_info),
                            'mtime': getattr(stat_info, 'st_mtime', 0),
                            'size': getattr(stat_info, 'st_size', 0)
                        })
                    except Exception as e:
                        error_msg = str(e).lower()
                        if 'unimplemented' in error_msg or 'type 3' in error_msg:
                            self.logger.debug(f"stat 不支援: {filename}")
                            self._server_capabilities['supports_stat'] = False
                        else:
                            self.logger.debug(f"stat 失敗: {filename}, {e}")
                        
                        # 使用啟發式方法判斷是否為目錄
                        item_info['is_dir'] = self._guess_is_directory(filename)
                else:
                    # 使用啟發式方法
                    item_info['is_dir'] = self._guess_is_directory(filename)
                
                result.append(item_info)
            
            return result
            
        except Exception as e:
            self.logger.error(f"列出目錄詳細資訊失敗: {path}, 錯誤: {e}")
            raise
    
    def _is_directory_from_stat(self, stat_obj) -> bool:
        """從 stat 物件判斷是否為目錄"""
        try:
            return bool(stat_obj.st_mode & 0o40000)
        except:
            return False
    
    def _guess_is_directory(self, filename: str) -> bool:
        """啟發式方法猜測是否為目錄"""
        # 如果檔名包含明顯的副檔名，可能是檔案
        common_extensions = ['.xml', '.txt', '.log', '.zip', '.tar', '.gz', '.json', '.csv']
        filename_lower = filename.lower()
        
        for ext in common_extensions:
            if filename_lower.endswith(ext):
                return False
        
        # 如果看起來像版本目錄格式，假設是目錄
        if re.match(r'^\d+(_all)?(_\d{12})?(_.*)?$', filename):
            return True
        
        # 預設假設是目錄（在搜尋 manifest 的情境下比較安全）
        return True

    def _try_direct_version_paths(self, base_path: str, version: str) -> Optional[Tuple[str, str]]:
        """當找不到版本目錄時，嘗試直接構建可能的路徑"""
        try:
            import datetime
            today = datetime.datetime.now()
            
            self.logger.info(f"嘗試直接路徑搜尋版本 {version}")
            
            # 生成可能的時間戳（最近30天，每3天一個）
            for days_back in range(0, 30, 3):
                date = today - datetime.timedelta(days=days_back)
                
                # 生成多個可能的時間格式
                timestamps = [
                    date.strftime("%Y%m%d0000"),  # 202508170000
                    date.strftime("%Y%m%d1200"),  # 202508171200
                    date.strftime("%Y%m%d1800"),  # 202508171800
                    date.strftime("%Y%m%d2300"),  # 202508172300
                ]
                
                for timestamp in timestamps:
                    # 嘗試不同的目錄格式
                    possible_dirs = [
                        f"{version}_{timestamp}",
                        f"{version}_all_{timestamp}",
                        f"{version}_all_{timestamp}_no-release",
                        f"{version}_all_{timestamp}_backup",
                        f"{version}_{timestamp}_backup",
                    ]
                    
                    for dir_name in possible_dirs:
                        try:
                            manifest_path = self._try_get_manifest_from_dir(base_path, dir_name, version)
                            if manifest_path:
                                self.logger.info(f"✅ 透過直接路徑找到: {dir_name}")
                                return manifest_path
                        except Exception as e:
                            self.logger.debug(f"直接路徑失敗: {dir_name}, {e}")
                            continue
            
            return None
            
        except Exception as e:
            self.logger.debug(f"直接路徑搜尋失敗: {e}")
            return None
            
    def _try_get_manifest_from_dir(self, base_path: str, dir_name: str, version: str) -> Optional[Tuple[str, str]]:
        """嘗試從指定目錄獲取 manifest - 使用最基本的方法"""
        try:
            version_path = f"{base_path}/{dir_name}"
            
            # 直接構建可能的 manifest 檔案名
            possible_manifests = [
                f"manifest_{version}.xml",
                "manifest.xml",
                f"default_{version}.xml", 
                "default.xml",
                f"{version}.xml"
            ]
            
            for manifest_name in possible_manifests:
                manifest_full_path = f"{version_path}/{manifest_name}"
                
                try:
                    # 使用最基本的檢查方法
                    if self._file_exists_and_valid(manifest_full_path):
                        self.logger.info(f"✅ 找到有效 manifest: {manifest_name} (路徑: {dir_name})")
                        return manifest_full_path, manifest_name
                        
                except Exception as e:
                    self.logger.debug(f"檢查檔案失敗: {manifest_name}, {e}")
                    continue
            
            return None
            
        except Exception as e:
            self.logger.debug(f"檢查目錄失敗: {dir_name}, {e}")
            return None

    def _file_exists_and_valid(self, file_path: str) -> bool:
        """最基本的檔案存在性和有效性檢查"""
        try:
            # 如果 stat 可用，使用 stat
            if self._server_capabilities.get('supports_stat', True):
                try:
                    stat_info = self.sftp.stat(file_path)
                    file_size = getattr(stat_info, 'st_size', 0)
                    return file_size > 1000  # manifest 檔案應該大於 1KB
                except Exception as e:
                    error_msg = str(e).lower()
                    if 'unimplemented' in error_msg or 'type 3' in error_msg:
                        self._server_capabilities['supports_stat'] = False
                    elif 'no such file' in error_msg or 'not found' in error_msg:
                        return False
                    else:
                        raise
            
            # 如果 stat 不可用，嘗試打開檔案來檢查
            try:
                with self.sftp.open(file_path, 'r') as f:
                    # 讀取前幾個字節檢查是否像 XML
                    header = f.read(100).decode('utf-8', errors='ignore')
                    return '<?xml' in header or '<manifest' in header
            except Exception:
                return False
                
        except Exception:
            return False
                    
    def _matches_version(self, dir_name: str, target_version: str) -> bool:
        """檢查目錄名是否匹配指定版本"""
        patterns = [
            rf'^{target_version}_\d{{12}}$',                    # 5_202508170000
            rf'^{target_version}_all_\d{{12}}$',                # 702_all_202508092300  
            rf'^{target_version}_all_\d{{12}}_.*$',             # 669_all_202507142300_no-release
            rf'^{target_version}_\d{{12}}_.*$',                 # 5_202508170000_backup
        ]
        
        for pattern in patterns:
            if re.match(pattern, dir_name):
                return True
        return False
    
    def _extract_version_number(self, name: str) -> Optional[str]:
        """提取版本號"""
        patterns = [
            r'^(\d+)_all_\d{12}',      # 702_all_202508092300
            r'^(\d+)_all_\d{12}_',     # 669_all_202507142300_no-release
            r'^(\d+)_\d{12}',          # 5_202508170000
            r'^(\d+)_\d{12}_',         # 5_202508170000_backup
            r'^(\d+)$',                # 5
            r'^v(\d+)',                # v5
        ]
        
        for pattern in patterns:
            match = re.search(pattern, name)
            if match:
                return match.group(1)
        
        return None
    
    def download_file(self, remote_path: str, local_path: str) -> bool:
        """下載檔案"""
        try:
            if not self._ensure_connected():
                return False
            
            os.makedirs(os.path.dirname(local_path), exist_ok=True)
            
            filename = os.path.basename(remote_path)
            self.logger.info(f"開始下載: {filename}")
            
            # 獲取檔案大小（如果支援的話）
            file_size = 0
            if self._server_capabilities.get('supports_stat', True):
                try:
                    file_stat = self.sftp.stat(remote_path)
                    file_size = file_stat.st_size
                    self.logger.debug(f"檔案大小: {file_size} bytes")
                except Exception as e:
                    error_msg = str(e).lower()
                    if 'unimplemented' in error_msg or 'type 3' in error_msg:
                        self._server_capabilities['supports_stat'] = False
                        self.logger.debug("stat 不支援，跳過檔案大小檢查")
                    else:
                        self.logger.debug(f"無法獲取檔案大小: {e}")
            
            # 下載檔案
            self.sftp.get(remote_path, local_path)
            
            # 驗證下載
            if os.path.exists(local_path) and os.path.getsize(local_path) > 0:
                actual_size = os.path.getsize(local_path)
                self.logger.info(f"下載完成: {filename} ({actual_size} bytes)")
                
                # 驗證檔案大小是否一致（如果有的話）
                if file_size > 0 and actual_size != file_size:
                    self.logger.warning(f"檔案大小不一致: 預期 {file_size}, 實際 {actual_size}")
                
                return True
            else:
                self.logger.error(f"下載的檔案無效或為空: {local_path}")
                return False
                
        except Exception as e:
            self.logger.error(f"下載檔案失敗: {e}")
            return False

# =====================================
# ===== Source Command 管理器 =====
# =====================================

class SourceCommandManager:
    """管理不同 DB 的 source command"""
    
    def __init__(self):
        self.logger = setup_logger(self.__class__.__name__)
        self.jira_client = JiraAPIClient()
        self.cache = {}
        
    def get_source_command(self, db_info: DBInfo, mapping_df: pd.DataFrame = None) -> Optional[str]:
        """根據 DB 資訊獲取對應的 source command"""
        db_name = db_info.db_info
        
        # 檢查快取
        if db_name in self.cache:
            self.logger.info(f"使用快取的 source command for {db_name}")
            return self.cache[db_name]
        
        # 從 JIRA 搜尋
        self.logger.info(f"從 JIRA 搜尋 {db_name} 的 source command")
        ticket_key = self.jira_client.search_db_ticket(db_name, db_info.module)
        if ticket_key:
            # 更新 jira_link
            db_info.jira_link = f"https://{config_manager.jira_config['site']}/browse/{ticket_key}"
            self.logger.info(f"找到 JIRA ticket: {ticket_key}")
            
            cmd = self.jira_client.get_source_command_from_ticket(ticket_key)
            if cmd:
                self.cache[db_name] = cmd
                self.logger.info(f"成功從 JIRA 獲取 source command")
                return cmd
        
        self.logger.warning(f"無法從 JIRA 獲取 {db_name} 的 source command")
        return None
    
    def clear_cache(self):
        """清除所有快取"""
        self.cache.clear()

# =====================================
# ===== Mapping Table 讀取器 =====
# =====================================

class MappingTableReader:
    """讀取和解析 mapping table 的類別"""
    
    def __init__(self):
        self.logger = setup_logger(self.__class__.__name__)
        self.df = None
        
    def load_excel(self, file_path: str) -> bool:
        """載入 Excel 檔案"""
        try:
            self.df = pd.read_excel(file_path)
            self.logger.info(f"成功載入 mapping table: {len(self.df)} 筆資料")
            return True
        except Exception as e:
            self.logger.error(f"載入 Excel 失敗: {str(e)}")
            return False
    
    def get_db_info_list(self, db_type: str = 'all') -> List[DBInfo]:
        """取得 DB 資訊列表"""
        db_list = []
        
        if self.df is None:
            return db_list
        
        type_columns = {
            'master': ('DB_Type', 'DB_Info', 'DB_Folder', 'SftpPath'),
            'premp': ('premp_DB_Type', 'premp_DB_Info', 'premp_DB_Folder', 'premp_SftpPath'),
            'mp': ('mp_DB_Type', 'mp_DB_Info', 'mp_DB_Folder', 'mp_SftpPath'),
            'mpbackup': ('mpbackup_DB_Type', 'mpbackup_DB_Info', 'mpbackup_DB_Folder', 'mpbackup_SftpPath')
        }
        
        types_to_process = type_columns.keys() if db_type == 'all' else [db_type]
        
        for idx, row in self.df.iterrows():
            for dtype in types_to_process:
                if dtype not in type_columns:
                    continue
                    
                cols = type_columns[dtype]
                db_info_col = cols[1]
                
                if db_info_col in row and pd.notna(row[db_info_col]):
                    db_info = DBInfo(
                        sn=row['SN'] if 'SN' in row else idx + 1,
                        module=row['Module'] if 'Module' in row else '',
                        db_type=dtype,
                        db_info=str(row[db_info_col]),
                        db_folder=str(row[cols[2]]) if cols[2] in row and pd.notna(row[cols[2]]) else '',
                        sftp_path=str(row[cols[3]]) if cols[3] in row and pd.notna(row[cols[3]]) else ''
                    )
                    db_list.append(db_info)
        
        return db_list

# =====================================
# ===== Repo 管理器 =====
# =====================================

class RepoManager:
    """Repo 指令管理器"""
    
    def __init__(self):
        self.logger = setup_logger(self.__class__.__name__)
        self.lock = threading.Lock()
    
    def check_repo_exists(self, work_dir: str) -> bool:
        """檢查 .repo 目錄是否存在"""
        repo_dir = os.path.join(work_dir, '.repo')
        exists = os.path.exists(repo_dir)
        self.logger.info(f"檢查 .repo 目錄: {work_dir} -> {'存在' if exists else '不存在'}")
        return exists

    # 🔥 新增：快速檢查方法
    def _quick_repo_status_check(self, work_dir: str) -> dict:
        """
        快速檢查repo狀態（適用於大型repo）
        """
        try:
            self.logger.debug(f"嘗試快速repo狀態檢查...")
            
            # 檢查是否有 .repo/project.list 來快速判斷
            repo_dir = os.path.join(work_dir, '.repo')
            project_list_file = os.path.join(repo_dir, 'project.list')
            
            if not os.path.exists(project_list_file):
                return None
            
            # 快速檢查是否有明顯的修改
            quick_cmd = f"{config_manager.repo_config['repo_command']} status --porcelain"
            success, output = self.run_command(quick_cmd, cwd=work_dir, timeout=60)
            
            if success:
                # 如果輸出為空或很少，可能是乾淨的
                lines = [line.strip() for line in output.split('\n') if line.strip()]
                
                if len(lines) == 0:
                    self.logger.info(f"✅ 快速檢查：工作空間乾淨")
                    return {
                        'is_clean': True,
                        'modified_files': [],
                        'untracked_files': [],
                        'staged_files': [],
                        'details': '快速檢查：工作空間乾淨'
                    }
                elif len(lines) < 50:  # 如果修改較少，繼續完整檢查
                    self.logger.debug(f"快速檢查：發現 {len(lines)} 行變更，執行完整檢查")
                    return None
                else:
                    # 太多修改，直接返回髒狀態
                    self.logger.warning(f"快速檢查：發現大量變更 ({len(lines)} 行)")
                    return {
                        'is_clean': False,
                        'modified_files': ['多個文件'],
                        'untracked_files': [],
                        'staged_files': [],
                        'details': f'快速檢查：發現大量變更 ({len(lines)} 行)'
                    }
            
            return None
            
        except Exception as e:
            self.logger.debug(f"快速檢查失敗: {e}")
            return None
    
    # 🔥 新增：基於git的備用檢查
    def _fallback_git_status_check(self, work_dir: str) -> dict:
        """
        當repo status失敗時的備用檢查方法
        """
        try:
            self.logger.info(f"執行基於git的備用狀態檢查...")
            
            # 掃描可能的git目錄
            modified_files = []
            untracked_files = []
            
            # 在工作目錄下查找git專案
            for root, dirs, files in os.walk(work_dir):
                if '.git' in dirs:
                    git_dir = root
                    rel_path = os.path.relpath(git_dir, work_dir)
                    
                    try:
                        # 檢查這個git專案的狀態
                        git_status_cmd = "git status --porcelain"
                        success, output = self.run_command(git_status_cmd, cwd=git_dir, timeout=30)
                        
                        if success and output.strip():
                            for line in output.split('\n'):
                                if line.strip():
                                    status_code = line[:2]
                                    file_path = line[3:].strip()
                                    full_path = f"{rel_path}/{file_path}" if rel_path != '.' else file_path
                                    
                                    if status_code.strip() in ['M', 'MM', 'AM', 'AD']:
                                        modified_files.append(full_path)
                                    elif status_code.strip() in ['??']:
                                        untracked_files.append(full_path)
                                        
                    except Exception as e:
                        self.logger.debug(f"檢查git目錄 {git_dir} 失敗: {e}")
                        continue
                    
                    # 避免遞歸進入.git目錄
                    dirs.remove('.git')
                
                # 限制掃描深度避免太慢
                if root.count(os.sep) - work_dir.count(os.sep) > 3:
                    dirs.clear()
            
            is_clean = len(modified_files) == 0 and len(untracked_files) == 0
            
            details_parts = []
            if modified_files:
                details_parts.append(f"修改文件: {len(modified_files)}個")
            if untracked_files:
                details_parts.append(f"未追蹤文件: {len(untracked_files)}個")
            
            details = '; '.join(details_parts) if details_parts else '備用檢查：工作空間乾淨'
            
            result = {
                'is_clean': is_clean,
                'modified_files': modified_files,
                'untracked_files': untracked_files,
                'staged_files': [],
                'details': f"備用檢查：{details}"
            }
            
            self.logger.info(f"備用檢查完成: {details}")
            return result
            
        except Exception as e:
            self.logger.error(f"備用檢查失敗: {e}")
            return {
                'is_clean': False,
                'modified_files': [],
                'untracked_files': [],
                'staged_files': [],
                'details': f'備用檢查失敗: {str(e)}'
            }

    # 🔥 新增：檢查repo工作空間是否乾淨
    def check_repo_clean_status(self, work_dir: str) -> dict:
        """
        檢查repo工作空間是否乾淨 - 🔥 超快版本
        """
        try:
            if not self.check_repo_exists(work_dir):
                return {
                    'is_clean': True,
                    'modified_files': [],
                    'untracked_files': [],
                    'staged_files': [],
                    'details': '.repo目錄不存在，視為乾淨狀態'
                }
            
            self.logger.info(f"快速檢查repo工作空間狀態: {work_dir}")
            
            # 🔥 超快檢查：只看是否有明顯的 .git/index.lock 文件
            has_obvious_changes = False
            
            # 檢查常見的髒文件指標
            common_dirty_indicators = [
                '.repo/project.list.lock',
                '.repo/repo.lock',
            ]
            
            for indicator in common_dirty_indicators:
                if os.path.exists(os.path.join(work_dir, indicator)):
                    has_obvious_changes = True
                    break
            
            if has_obvious_changes:
                return {
                    'is_clean': False,
                    'modified_files': ['detected lock files'],
                    'untracked_files': [],
                    'staged_files': [],
                    'details': '檢測到鎖定文件，可能有操作進行中'
                }
            else:
                # 假設乾淨
                self.logger.info(f"✅ 快速檢查：未發現明顯髒狀態指標")
                return {
                    'is_clean': True,
                    'modified_files': [],
                    'untracked_files': [],
                    'staged_files': [],
                    'details': '快速檢查：假設工作空間乾淨'
                }
                
        except Exception as e:
            self.logger.warning(f"快速檢查失敗，假設乾淨: {e}")
            return {
                'is_clean': True,
                'modified_files': [],
                'untracked_files': [],
                'staged_files': [],
                'details': f'檢查失敗，假設乾淨: {str(e)}'
            }
    
    # 🔥 新增：備份本地修改
    def backup_local_changes(self, work_dir: str, db_name: str) -> bool:
        """
        備份本地修改到備份目錄 - 🔥 修復路徑和指令問題
        
        Args:
            work_dir: 工作目錄
            db_name: DB名稱（用於備份目錄命名）
            
        Returns:
            是否備份成功
        """
        try:
            if not config_manager.repo_config['backup_local_changes']:
                self.logger.info(f"跳過備份本地修改（配置已關閉）")
                return True
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_dir = os.path.join(work_dir, 'backups', f'local_changes_{timestamp}')
            os.makedirs(backup_dir, exist_ok=True)
            
            self.logger.info(f"備份本地修改到: {backup_dir}")
            
            # 🔥 修復：使用絕對路徑和正確的重定向方式
            diff_file = os.path.join(backup_dir, 'repo_diff.patch')
            
            # 確保diff文件存在
            with open(diff_file, 'w') as f:
                f.write("")  # 創建空文件
            
            # 🔥 修復指令：使用python方式處理重定向
            diff_cmd = f"{config_manager.repo_config['repo_command']} diff"
            success, diff_output = self.run_command(diff_cmd, cwd=work_dir, timeout=120)
            
            if success:
                # 直接寫入文件而不是使用shell重定向
                with open(diff_file, 'w', encoding='utf-8') as f:
                    f.write(diff_output)
                
                self.logger.info(f"✅ 差異已備份到: {diff_file}")
                
                # 🔥 新增：同時備份status信息
                status_file = os.path.join(backup_dir, 'repo_status.txt')
                status_cmd = f"{config_manager.repo_config['repo_command']} status"
                status_success, status_output = self.run_command(status_cmd, cwd=work_dir, timeout=60)
                
                if status_success:
                    with open(status_file, 'w', encoding='utf-8') as f:
                        f.write(status_output)
                    self.logger.debug(f"狀態已備份到: {status_file}")
                
                # 寫入備份說明
                info_file = os.path.join(backup_dir, 'backup_info.txt')
                with open(info_file, 'w', encoding='utf-8') as f:
                    f.write(f"備份時間: {datetime.now()}\n")
                    f.write(f"DB名稱: {db_name}\n")
                    f.write(f"工作目錄: {work_dir}\n")
                    f.write(f"備份原因: 定版前自動備份\n")
                    f.write(f"diff檔案: {diff_file}\n")
                    f.write(f"status檔案: {status_file}\n")
                    f.write(f"差異大小: {len(diff_output)} 字元\n")
                
                self.logger.info(f"✅ 備份完成，差異大小: {len(diff_output)} 字元")
                return True
            else:
                self.logger.warning(f"備份repo diff失敗: {diff_output}")
                
                # 🔥 嘗試基於git的備份
                return self._backup_using_git(work_dir, backup_dir, db_name)
                
        except Exception as e:
            self.logger.error(f"備份本地修改失敗: {e}")
            return False

    # 🔥 新增：基於git的備份方法
    def _backup_using_git(self, work_dir: str, backup_dir: str, db_name: str) -> bool:
        """
        使用git指令備份修改
        """
        try:
            self.logger.info(f"嘗試使用git方式備份...")
            
            backup_count = 0
            
            # 掃描git目錄並備份
            for root, dirs, files in os.walk(work_dir):
                if '.git' in dirs:
                    git_dir = root
                    rel_path = os.path.relpath(git_dir, work_dir)
                    
                    try:
                        # 為每個git項目創建diff
                        project_name = rel_path.replace(os.sep, '_') if rel_path != '.' else 'root'
                        diff_file = os.path.join(backup_dir, f'git_diff_{project_name}.patch')
                        
                        git_diff_cmd = "git diff HEAD"
                        success, diff_output = self.run_command(git_diff_cmd, cwd=git_dir, timeout=60)
                        
                        if success and diff_output.strip():
                            with open(diff_file, 'w', encoding='utf-8') as f:
                                f.write(f"# Git diff for: {rel_path}\n")
                                f.write(f"# Generated at: {datetime.now()}\n\n")
                                f.write(diff_output)
                            
                            backup_count += 1
                            self.logger.debug(f"備份git項目: {rel_path}")
                        
                    except Exception as e:
                        self.logger.debug(f"備份git項目 {git_dir} 失敗: {e}")
                        continue
                    
                    dirs.remove('.git')
                
                # 限制掃描深度
                if root.count(os.sep) - work_dir.count(os.sep) > 3:
                    dirs.clear()
            
            if backup_count > 0:
                self.logger.info(f"✅ 使用git方式備份了 {backup_count} 個項目")
                return True
            else:
                self.logger.warning(f"⚠️ 沒有找到需要備份的git修改")
                return True  # 沒有修改也算成功
                
        except Exception as e:
            self.logger.error(f"git備份失敗: {e}")
            return False

    def clean_repo_workspace(self, work_dir: str, force: bool = False) -> bool:
        """
        清理repo工作空間，還原到乾淨狀態 - 🔥 優化清理步驟和超時處理
        
        Args:
            work_dir: 工作目錄
            force: 是否強制清理（會刪除所有本地修改）
            
        Returns:
            是否清理成功
        """
        try:
            if not self.check_repo_exists(work_dir):
                self.logger.info(f"repo不存在，跳過清理: {work_dir}")
                return True
            
            self.logger.info(f"開始清理repo工作空間: {work_dir}")
            
            # 🔥 優化：分步驟執行清理，每步都有合理的超時時間
            clean_steps = [
                {
                    'name': '重置所有修改',
                    'cmd': f"{config_manager.repo_config['repo_command']} forall -c 'git reset --hard HEAD'",
                    'timeout': 600  # 10分鐘
                },
                {
                    'name': '清理未追蹤檔案',
                    'cmd': f"{config_manager.repo_config['repo_command']} forall -c 'git clean -fd'",
                    'timeout': 300  # 5分鐘
                },
                {
                    'name': '確保正確分支',
                    'cmd': f"{config_manager.repo_config['repo_command']} forall -c 'git checkout .'",
                    'timeout': 300  # 5分鐘
                }
            ]
            
            success_count = 0
            
            for i, step in enumerate(clean_steps, 1):
                self.logger.info(f"執行清理步驟 {i}/{len(clean_steps)}: {step['name']}")
                
                success, output = self.run_command(
                    step['cmd'], 
                    cwd=work_dir, 
                    timeout=step['timeout']
                )
                
                if success:
                    success_count += 1
                    self.logger.info(f"✅ 清理步驟 {i} 完成")
                else:
                    if force:
                        self.logger.warning(f"⚠️ 清理步驟 {i} 失敗但繼續執行: {output[:200]}")
                    else:
                        self.logger.error(f"❌ 清理步驟 {i} 失敗: {output[:200]}")
                        
                        # 🔥 如果是超時，嘗試更溫和的方式
                        if "timeout" in output.lower() or "Command timeout" in output:
                            self.logger.info(f"清理步驟超時，嘗試基於git的清理...")
                            git_clean_success = self._clean_using_git(work_dir)
                            if git_clean_success:
                                success_count += 1
                                continue
                        
                        return False
            
            # 檢查清理結果
            if success_count >= 2:  # 至少2個步驟成功
                # 再次快速檢查狀態
                try:
                    clean_status = self._quick_repo_status_check(work_dir)
                    if clean_status and clean_status['is_clean']:
                        self.logger.info(f"✅ 工作空間清理完成，狀態乾淨")
                        return True
                    elif clean_status and not clean_status['is_clean']:
                        if force:
                            self.logger.warning(f"⚠️ 強制模式：工作空間仍有變更但繼續執行")
                            return True
                        else:
                            self.logger.warning(f"⚠️ 清理後仍有變更: {clean_status['details']}")
                            return False
                    else:
                        # 快速檢查失敗，但清理步驟成功，假設OK
                        self.logger.info(f"✅ 清理步驟完成，假設工作空間已清理")
                        return True
                except Exception as e:
                    self.logger.warning(f"檢查清理結果失敗: {e}")
                    if force or success_count == len(clean_steps):
                        return True
                    else:
                        return False
            else:
                self.logger.error(f"❌ 清理失敗，只有 {success_count}/{len(clean_steps)} 步驟成功")
                return False
                
        except Exception as e:
            self.logger.error(f"清理工作空間失敗: {e}")
            return False
            
    # 🔥 新增：清理工作空間
    def clean_repo_workspace(self, work_dir: str, force: bool = False) -> bool:
        """
        清理repo工作空間，還原到乾淨狀態
        
        Args:
            work_dir: 工作目錄
            force: 是否強制清理（會刪除所有本地修改）
            
        Returns:
            是否清理成功
        """
        try:
            if not self.check_repo_exists(work_dir):
                self.logger.info(f"repo不存在，跳過清理: {work_dir}")
                return True
            
            self.logger.info(f"開始清理repo工作空間: {work_dir}")
            
            # 執行清理命令序列
            clean_commands = [
                # 1. 重置所有修改
                f"{config_manager.repo_config['repo_command']} forall -c 'git reset --hard HEAD'",
                # 2. 清理未追蹤的檔案
                f"{config_manager.repo_config['repo_command']} forall -c 'git clean -fd'",
                # 3. 確保在正確的分支上
                f"{config_manager.repo_config['repo_command']} forall -c 'git checkout .'",
            ]
            
            for i, cmd in enumerate(clean_commands, 1):
                self.logger.info(f"執行清理步驟 {i}/{len(clean_commands)}: {cmd.split(' forall')[0]}...")
                
                success, output = self.run_command(cmd, cwd=work_dir, timeout=300)
                
                if not success:
                    if force:
                        self.logger.warning(f"清理步驟 {i} 失敗但繼續執行: {output}")
                    else:
                        self.logger.error(f"清理步驟 {i} 失敗: {output}")
                        return False
                else:
                    self.logger.debug(f"清理步驟 {i} 完成")
            
            # 再次檢查狀態
            clean_status = self.check_repo_clean_status(work_dir)
            
            if clean_status['is_clean']:
                self.logger.info(f"✅ 工作空間清理完成，狀態乾淨")
                return True
            else:
                if force:
                    self.logger.warning(f"⚠️ 強制模式：工作空間仍有變更但繼續執行")
                    return True
                else:
                    self.logger.error(f"❌ 清理後工作空間仍不乾淨: {clean_status['details']}")
                    return False
                
        except Exception as e:
            self.logger.error(f"清理工作空間失敗: {e}")
            return False
    
    # 🔥 新增：處理髒repo的策略
    def handle_dirty_repo(self, work_dir: str, db_name: str, clean_status: dict) -> bool:
        """
        處理有本地修改的repo - 🔥 添加用戶選擇和跳過選項
        
        Args:
            work_dir: 工作目錄
            db_name: DB名稱
            clean_status: repo狀態檢查結果
            
        Returns:
            是否處理成功
        """
        try:
            self.logger.warning(f"發現本地修改: {clean_status['details']}")
            
            # 🔥 如果是檢查超時，提供快速處理選項
            if "timeout" in clean_status['details'].lower() or "Command timeout" in clean_status['details']:
                self.logger.warning(f"⚠️ repo status 檢查超時，可能因為工作空間太大")
                
                if config_manager.repo_config['skip_clean_on_timeout']:
                    self.logger.info(f"📋 配置為跳過超時清理，假設工作空間可用")
                    return True
                
                # 詢問用戶是否要強制清理
                if config_manager.repo_config['ask_user_on_dirty']:
                    print(f"\n⚠️ {db_name}: repo status 檢查超時")
                    print("可能原因：工作空間很大或有大量修改")
                    print("選項：")
                    print("  1. 跳過清理，直接使用現有工作空間 (推薦)")
                    print("  2. 強制清理 (可能很慢)")
                    print("  3. 跳過這個DB")
                    
                    choice = input("請選擇 (1/2/3) [1]: ").strip() or "1"
                    
                    if choice == "1":
                        self.logger.info(f"👌 用戶選擇跳過清理，直接使用工作空間")
                        return True
                    elif choice == "2":
                        self.logger.info(f"🧹 用戶選擇強制清理")
                        return self._force_clean_large_repo(work_dir, db_name)
                    else:
                        self.logger.info(f"⏭️ 用戶選擇跳過此DB")
                        return False
                else:
                    # 非互動模式，根據配置決定
                    if config_manager.repo_config['force_clean_on_dirty']:
                        return self._force_clean_large_repo(work_dir, db_name)
                    else:
                        self.logger.info(f"⏭️ 自動跳過清理（非互動模式）")
                        return True
            
            # 顯示修改詳情（只顯示前幾個）
            if clean_status['modified_files']:
                self.logger.info(f"修改的文件:")
                for file_path in clean_status['modified_files'][:5]:  # 只顯示前5個
                    self.logger.info(f"  - {file_path}")
                if len(clean_status['modified_files']) > 5:
                    self.logger.info(f"  ... 還有 {len(clean_status['modified_files']) - 5} 個文件")
            
            # 🔥 根據配置決定處理策略
            if config_manager.repo_config['ask_user_on_dirty']:
                # 互動模式：詢問用戶
                print(f"\n⚠️ {db_name}: 發現本地修改")
                print(f"詳情: {clean_status['details']}")
                print("選項：")
                print("  1. 跳過清理，直接使用 (快速)")
                print("  2. 備份並清理 (較慢)")
                print("  3. 跳過這個DB")
                
                choice = input("請選擇 (1/2/3) [1]: ").strip() or "1"
                
                if choice == "1":
                    self.logger.info(f"👌 用戶選擇跳過清理")
                    return True
                elif choice == "2":
                    return self._backup_and_clean(work_dir, db_name)
                else:
                    self.logger.info(f"⏭️ 用戶選擇跳過此DB")
                    return False
                    
            elif config_manager.repo_config['auto_clean_workspace']:
                # 自動清理模式
                self.logger.info(f"🔄 自動清理模式：備份並清理本地修改")
                return self._backup_and_clean(work_dir, db_name)
            else:
                # 不清理模式
                self.logger.warning(f"⚠️ 發現本地修改但未啟用自動清理，直接使用現有工作空間")
                return True
                    
        except Exception as e:
            self.logger.error(f"處理髒repo失敗: {e}")
            return False

    # 🔥 新增：針對大型repo的強制清理
    def _force_clean_large_repo(self, work_dir: str, db_name: str) -> bool:
        """
        針對大型repo的強制清理策略
        """
        try:
            self.logger.info(f"🧹 開始強制清理大型repo: {db_name}")
            
            if config_manager.repo_config['quick_clean_only']:
                # 只做最基本的清理
                self.logger.info(f"執行快速清理...")
                
                # 簡單粗暴：刪除 .repo/project-objects 強制重新sync
                project_objects = os.path.join(work_dir, '.repo', 'project-objects')
                if os.path.exists(project_objects):
                    import shutil
                    shutil.rmtree(project_objects)
                    self.logger.info(f"✅ 已清理項目緩存")
                
                return True
            else:
                # 執行完整清理（會很慢）
                return self.clean_repo_workspace(work_dir, force=True)
                
        except Exception as e:
            self.logger.error(f"強制清理失敗: {e}")
            return False

    # 🔥 新增：備份並清理的組合方法
    def _backup_and_clean(self, work_dir: str, db_name: str) -> bool:
        """
        備份並清理的組合方法
        """
        try:
            # 先備份
            if config_manager.repo_config['backup_local_changes']:
                backup_success = self.backup_local_changes(work_dir, db_name)
                if not backup_success:
                    self.logger.warning(f"備份失敗，但繼續清理")
            
            # 再清理
            if config_manager.repo_config['quick_clean_only']:
                return self._force_clean_large_repo(work_dir, db_name)
            else:
                return self.clean_repo_workspace(work_dir, force=True)
                
        except Exception as e:
            self.logger.error(f"備份並清理失敗: {e}")
            return False
            
    def run_command(self, cmd: str, cwd: str = None, timeout: int = None) -> Tuple[bool, str]:
        """同步執行指令"""
        try:
            self.logger.debug(f"執行指令: {cmd}")
            result = subprocess.run(
                cmd,
                shell=True,
                cwd=cwd,
                capture_output=True,
                text=True,
                timeout=timeout
            )
            
            if result.returncode == 0:
                self.logger.debug(f"指令執行成功")
                return True, result.stdout
            else:
                self.logger.warning(f"指令執行失敗 (返回碼: {result.returncode})")
                return False, result.stderr
                
        except subprocess.TimeoutExpired:
            self.logger.error(f"指令執行超時")
            return False, "Command timeout"
        except Exception as e:
            self.logger.error(f"指令執行異常: {e}")
            return False, str(e)
    
    def repo_init(self, work_dir: str, init_cmd: str) -> bool:
        """執行 repo init - 增強清理和錯誤恢復"""
        # 🔥 徹底清理可能存在的舊 .repo 目錄和相關檔案
        repo_dir = os.path.join(work_dir, '.repo')
        if os.path.exists(repo_dir):
            self.logger.info(f"發現舊的 .repo 目錄，執行徹底清理: {repo_dir}")
            import shutil
            try:
                # 先嘗試正常刪除
                shutil.rmtree(repo_dir)
                self.logger.info("✅ 舊 .repo 目錄清理成功")
            except Exception as e:
                self.logger.warning(f"正常清理失敗，嘗試強制清理: {e}")
                # 強制清理
                try:
                    import subprocess
                    if os.name == 'posix':  # Linux/Unix
                        subprocess.run(['rm', '-rf', repo_dir], check=True)
                        self.logger.info("✅ 強制清理成功")
                    else:
                        raise Exception("無法強制清理")
                except Exception as e2:
                    self.logger.error(f"❌ 強制清理也失敗: {e2}")
                    return False
        
        # 🔥 清理可能存在的其他相關檔案
        cleanup_patterns = [
            '.repo_*',
            'repo_*',
            '.repopickle_*'
        ]
        
        for pattern in cleanup_patterns:
            try:
                import glob
                for item in glob.glob(os.path.join(work_dir, pattern)):
                    if os.path.isfile(item):
                        os.remove(item)
                    elif os.path.isdir(item):
                        shutil.rmtree(item)
                    self.logger.debug(f"清理: {item}")
            except Exception as e:
                self.logger.debug(f"清理 {pattern} 時發生錯誤: {e}")
        
        # 🔥 增加重試機制和更長的超時時間
        max_retries = 3
        timeout_values = [180, 300, 600]  # 3分鐘、5分鐘、10分鐘
        for attempt in range(max_retries):
            timeout = timeout_values[attempt]
            self.logger.info(f"執行 repo init (嘗試 {attempt + 1}/{max_retries}，超時: {timeout}秒): {init_cmd}")
            
            success, output = self.run_command(
                init_cmd,
                cwd=work_dir,
                timeout=timeout
            )
            
            if success:
                self.logger.info("✅ Repo init 執行成功")
                return True
            else:
                self.logger.warning(f"❌ Repo init 嘗試 {attempt + 1} 失敗: {output}")
                
                # 如果不是最後一次嘗試，清理並重試
                if attempt < max_retries - 1:
                    self.logger.info(f"準備重試，先清理環境...")
                    repo_dir = os.path.join(work_dir, '.repo')
                    if os.path.exists(repo_dir):
                        try:
                            import shutil
                            shutil.rmtree(repo_dir)
                        except:
                            pass
                    time.sleep(15)  # 等待 15 秒後重試
        
        self.logger.error("❌ Repo init 所有重試都失敗")
        return False
    
    def apply_manifest(self, work_dir: str, manifest_file: str) -> bool:
        """應用 manifest 檔案 - 改進版本"""
        try:
            manifest_name = os.path.basename(manifest_file)
            self.logger.info(f"應用 manifest: {manifest_name}")
            
            repo_dir = os.path.join(work_dir, '.repo')
            manifests_dir = os.path.join(repo_dir, 'manifests')
            
            if not os.path.exists(manifests_dir):
                self.logger.error(f"Manifests 目錄不存在: {manifests_dir}")
                return False
            
            # 複製 manifest 檔案
            dest_file = os.path.join(manifests_dir, manifest_name)
            import shutil
            shutil.copy2(manifest_file, dest_file)
            self.logger.debug(f"複製 manifest: {manifest_file} -> {dest_file}")
            
            # 🔥 修改：增加超時時間並添加重試機制
            switch_cmd = f"{config_manager.repo_config['repo_command']} init -m {manifest_name}"
            self.logger.info(f"切換 manifest: {switch_cmd}")
            
            # 🔥 嘗試多次，增加超時時間
            max_attempts = 3
            timeout_values = [120, 180, 300]  # 2分鐘、3分鐘、5分鐘
            
            for attempt in range(max_attempts):
                timeout = timeout_values[attempt]
                self.logger.info(f"嘗試 {attempt + 1}/{max_attempts}，超時設定: {timeout}秒")
                
                success, output = self.run_command(
                    switch_cmd,
                    cwd=work_dir,
                    timeout=timeout
                )
                
                if success:
                    self.logger.info(f"✅ 成功切換到 manifest: {manifest_name}")
                    return True
                else:
                    self.logger.warning(f"❌ 嘗試 {attempt + 1} 失敗: {output}")
                    
                    # 如果不是最後一次，等待一下再重試
                    if attempt < max_attempts - 1:
                        import time
                        wait_time = 10 * (attempt + 1)  # 10秒、20秒
                        self.logger.info(f"等待 {wait_time} 秒後重試...")
                        time.sleep(wait_time)
            
            self.logger.error(f"❌ 所有嘗試都失敗，無法切換 manifest")
            return False
            
        except Exception as e:
            self.logger.error(f"應用 manifest 失敗: {str(e)}")
            return False
    
    def start_repo_sync_async(self, work_dir: str, db_name: str) -> subprocess.Popen:
        """🎯 修復版本 - 使用 unbuffer 確保實時輸出"""
        try:
            self.logger.info(f"{db_name}: 啟動 unbuffer 版本 repo sync")
            
            # 檢查工作目錄
            if not os.path.exists(os.path.join(work_dir, '.repo')):
                raise Exception(f"工作目錄沒有 .repo: {work_dir}")
            
            # 檢查 unbuffer 是否可用
            try:
                subprocess.run(['which', 'unbuffer'], check=True, capture_output=True, timeout=5)
                use_unbuffer = True
            except:
                use_unbuffer = False
                self.logger.warning(f"{db_name}: unbuffer 不可用，使用 script 方法")
            
            # 建立日誌
            log_dir = os.path.join(work_dir, 'logs')
            os.makedirs(log_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            method_name = "unbuffer" if use_unbuffer else "script"
            log_file = os.path.join(log_dir, f'repo_sync_{method_name}_{timestamp}.log')
            
            # 寫入初始信息
            with open(log_file, 'w', encoding='utf-8') as f:
                f.write(f"=== Fixed Repo Sync Log for {db_name} ===\n")
                f.write(f"開始時間: {datetime.now()}\n")
                f.write(f"工作目錄: {work_dir}\n")
                f.write(f"使用方法: {method_name}\n\n")
            
            # 準備命令
            repo_cmd = config_manager.repo_config['repo_command']
            jobs = min(config_manager.repo_config['sync_jobs'], 4)
            
            if use_unbuffer:
                # 🎯 使用 unbuffer 方法（已驗證有效）
                cmd_parts = [
                    'unbuffer',
                    repo_cmd, 'sync', 
                    f'-j{jobs}', 
                    '--verbose', 
                    '--force-sync'
                ]
                
                process = subprocess.Popen(
                    cmd_parts,
                    cwd=work_dir,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    bufsize=0,
                    preexec_fn=os.setsid
                )
            else:
                # 備用：使用 script 方法
                shell_cmd = f"""
    cd "{work_dir}"
    echo "[SCRIPT] 進程啟動，PID: $$" >> "{log_file}"
    script -fq /dev/null -c "{repo_cmd} sync -j{jobs} --verbose --force-sync" | tee -a "{log_file}"
    echo "[SCRIPT] 進程結束，時間: $(date)" >> "{log_file}"
    """
                
                process = subprocess.Popen(
                    ['bash', '-c', shell_cmd],
                    cwd=work_dir,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True
                )
            
            # 🎯 創建實時日誌寫入線程（關鍵改進）
            def log_writer():
                try:
                    with open(log_file, 'a', encoding='utf-8', buffering=1) as f:
                        f.write(f"[UNBUFFER] 進程啟動，PID: {process.pid}\n")
                        f.write(f"[UNBUFFER] 開始時間: {datetime.now()}\n\n")
                        f.flush()
                        
                        # 🔥 進度追蹤變數
                        last_reported_progress = -1
                        last_report_time = datetime.now()
                        message_count = 0
                        
                        if use_unbuffer:
                            while True:
                                line = process.stdout.readline()
                                if line:
                                    # 🔍 所有內容都寫入文件（不變）
                                    f.write(line)
                                    f.flush()
                                    
                                    # 🔥 智能過濾 console 輸出
                                    line_clean = line.strip()
                                    message_count += 1
                                    
                                    # ✅ 只報告重要的進度變化
                                    if "Syncing:" in line_clean and "%" in line_clean:
                                        import re
                                        progress_match = re.search(r'Syncing:\s*(\d+)%\s*\((\d+)/(\d+)\)', line_clean)
                                        if progress_match:
                                            current_progress = int(progress_match.group(1))
                                            current_count = int(progress_match.group(2))
                                            total_count = int(progress_match.group(3))
                                            
                                            # 🎯 只在以下情況報告進度：
                                            should_report = (
                                                last_reported_progress == -1 or  # 第一次
                                                current_progress - last_reported_progress >= 5 or  # 進度增加 5%以上
                                                current_progress % 10 == 0 or  # 每10%里程碑
                                                current_progress >= 95  # 接近完成時
                                            )
                                            
                                            if should_report:
                                                # 計算速度
                                                current_time = datetime.now()
                                                elapsed = (current_time - last_report_time).total_seconds()
                                                
                                                speed_info = ""
                                                if last_reported_progress > 0 and elapsed > 0:
                                                    progress_diff = current_progress - last_reported_progress
                                                    speed = progress_diff / (elapsed / 60)  # %/分鐘
                                                    if speed > 0:
                                                        remaining_time = (100 - current_progress) / speed
                                                        speed_info = f" (預計剩餘: {remaining_time:.0f}分鐘)"
                                                
                                                # 📊 簡潔的進度報告
                                                # self.logger.info(
                                                #    f"{db_name}: {current_progress}% "
                                                #    f"({current_count}/{total_count}){speed_info}"
                                                #)
                                                
                                                last_reported_progress = current_progress
                                                last_report_time = current_time
                                    
                                    # ⚠️ 報告錯誤和警告
                                    elif any(keyword in line_clean.lower() for keyword in 
                                        ['error:', 'fatal:', 'failed', 'timeout', 'exception', 'abort']):
                                        self.logger.warning(f"{db_name}: {line_clean}")
                                    
                                    # 🎉 報告重要里程碑
                                    elif any(phrase in line_clean.lower() for phrase in
                                        ['sync has finished', 'completed successfully', 'repo sync complete']):
                                        self.logger.info(f"{db_name}: 同步完成！")
                                    
                                    # 🚫 不報告的內容：
                                    # - "Skipped fetching project"
                                    # - "fetching project" 
                                    # - "..working.."
                                    # - 重複的進度信息
                                    
                                elif process.poll() is not None:
                                    break
                        
                        return_code = process.poll()
                        f.write(f"\n[UNBUFFER] 進程結束，返回碼: {return_code}\n")
                        f.write(f"[UNBUFFER] 結束時間: {datetime.now()}\n")
                        f.write(f"[UNBUFFER] 總處理消息數: {message_count}\n")
                        f.flush()
                        
                        # 📈 最終報告
                        if return_code == 0:
                            self.logger.info(f"{db_name}: ✅ 同步成功完成")
                        else:
                            self.logger.error(f"{db_name}: ❌ 同步失敗 (返回碼: {return_code})")
                        
                except Exception as e:
                    self.logger.error(f"{db_name}: 日誌寫入錯誤: {e}")
            
            # 啟動日誌線程
            if use_unbuffer:  # 只有 unbuffer 需要日誌線程
                log_thread = threading.Thread(target=log_writer, daemon=True)
                log_thread.start()
                process._log_thread = log_thread
            
            # 驗證啟動
            time.sleep(2)
            if process.poll() is not None:
                raise Exception(f"進程立即失敗，返回碼: {process.poll()}")
            
            # 保存進程信息
            process._log_file_path = log_file
            process._db_name = db_name
            process._start_time = datetime.now()
            
            resource_manager.register_process(db_name, process)
            self.logger.info(f"{db_name}: repo sync 啟動成功 (PID: {process.pid})")
            
            return process
            
        except Exception as e:
            self.logger.error(f"{db_name}: 啟動失敗: {e}")
            return None
    
    def check_process_status(self, db_name: str, process: subprocess.Popen) -> Optional[int]:
        """改進的進程狀態檢查"""
        with self.lock:
            if process:
                try:
                    poll = process.poll()
                    
                    # 🔥 加強日誌
                    self.logger.debug(f"{db_name}: 檢查進程狀態 PID={process.pid}, poll={poll}")
                    
                    if poll is not None:
                        # 🔥 進程已結束，記錄詳細信息
                        self.logger.info(f"{db_name}: 進程已結束，返回碼={poll}")
                        
                        # 正確關閉文件句柄
                        if hasattr(process, '_log_file_handle') and process._log_file_handle:
                            try:
                                process._log_file_handle.flush()
                                process._log_file_handle.close()
                                self.logger.debug(f"{db_name}: 日誌文件句柄已關閉")
                            except Exception as e:
                                self.logger.warning(f"{db_name}: 關閉日誌句柄失敗: {e}")
                        
                        resource_manager.unregister_process(db_name)
                        
                        # 🔥 寫入完成標記到日誌
                        if hasattr(process, '_log_file_path'):
                            try:
                                with open(process._log_file_path, 'a') as f:
                                    f.write(f"\n=== 進程結束 ===\n")
                                    f.write(f"返回碼: {poll}\n")
                                    f.write(f"結束時間: {datetime.now()}\n")
                            except:
                                pass
                        
                        return poll
                    else:
                        # 🔥 進程仍在運行，但驗證 PID 是否真的存在
                        try:
                            os.kill(process.pid, 0)  # 檢查進程是否存在
                            self.logger.debug(f"{db_name}: 進程 {process.pid} 確實在運行")
                        except OSError:
                            self.logger.warning(f"{db_name}: 進程 {process.pid} 不存在但 poll() 返回 None")
                            return -1  # 進程異常消失
                    
                    return None
                    
                except Exception as e:
                    self.logger.error(f"{db_name}: 檢查進程狀態時出錯: {e}")
                    return -1
            
            return None
    
    def export_manifest(self, work_dir: str, output_file: str = "vp_manifest.xml") -> bool:
        """導出 manifest - 🔥 修改：增加重試機制和更詳細的驗證"""
        cmd = f"{config_manager.repo_config['repo_command']} manifest -r -o {output_file}"
        self.logger.info(f"導出 manifest: {cmd}")
        
        # 🔥 增加超時時間到更長，並添加重試機制
        max_attempts = 5  # 增加到5次
        timeout_values = [120, 180, 240, 300, 600]  # 逐步增加超時時間
        
        for attempt in range(max_attempts):
            timeout = timeout_values[attempt]
            self.logger.info(f"導出 manifest 嘗試 {attempt + 1}/{max_attempts}，超時設定: {timeout}秒")
            
            success, output = self.run_command(cmd, cwd=work_dir, timeout=timeout)
            
            if success:
                output_path = os.path.join(work_dir, output_file)
                if os.path.exists(output_path):
                    file_size = os.path.getsize(output_path)
                    
                    # 🔥 更嚴格的檔案有效性檢查
                    if file_size > 1000:  # 至少要有 1KB
                        # 檢查檔案是否為有效的 XML
                        try:
                            with open(output_path, 'r', encoding='utf-8') as f:
                                first_line = f.readline().strip()
                                if first_line.startswith('<?xml') or '<manifest' in first_line:
                                    self.logger.info(f"✅ 成功導出有效的  manifest: {output_path} ({file_size} bytes)")
                                    return True
                                else:
                                    self.logger.warning(f"❌ 導出檔案格式無效: {first_line[:50]}")
                                    continue
                        except Exception as e:
                            self.logger.warning(f"❌ 無法驗證導出檔案: {e}")
                            continue
                    else:
                        self.logger.warning(f"❌ 導出檔案太小: {file_size} bytes")
                        continue
                else:
                    self.logger.warning(f"❌ 導出檔案不存在: {output_path}")
            else:
                self.logger.warning(f"❌ 導出嘗試 {attempt + 1} 失敗: {output}")
                
                # 如果不是最後一次，等待後重試
                if attempt < max_attempts - 1:
                    wait_time = 10 * (attempt + 1)
                    self.logger.info(f"等待 {wait_time} 秒後重試...")
                    time.sleep(wait_time)
        
        self.logger.error(f"❌ 所有導出嘗試都失敗")
        return False

# =====================================
# ===== 主要處理類別 =====
# =====================================

class ManifestPinningTool:
    """Manifest 定版工具（改進版 + 零失敗機制 + Manifest 比較功能）"""

    def __init__(self):
        self.logger = setup_logger(self.__class__.__name__)
        self.repo_manager = RepoManager()
        self.mapping_reader = MappingTableReader()
        self.source_cmd_manager = SourceCommandManager()
        self.manifest_comparator = ManifestComparator()  
        self.report = PinningReport()
        self.output_dir = config_manager.path_config['default_output_dir']
        self.dry_run = False
        self.zero_fail_mode = False
        
        # 即時報告更新的鎖
        self._report_update_lock = threading.Lock()
        
        # 線程安全鎖
        self._sftp_lock = threading.Lock()
        
        # 🔥 新增：防止重複處理的屬性
        self._processed_dbs = set()
        self._db_processing_lock = threading.Lock()

    def load_mapping_table(self, file_path: str) -> bool:
        """載入 mapping table"""
        return self.mapping_reader.load_excel(file_path)

    def get_all_dbs(self, db_type: str = 'all') -> List[DBInfo]:
        """取得所有 DB 資訊"""
        return self.mapping_reader.get_db_info_list(db_type)

    def process_db_phase1(self, db_info: DBInfo) -> DBInfo:
        """
        改進版 Phase 1 處理 - 線程安全 + 🔥 防止重複處理 + 跳過清理檢查
        """
        db_info.start_time = datetime.now()
        local_sftp_manager = None
        
        # 🔥 檢查是否已經在處理中，防止重複
        db_key = f"{db_info.db_info}_{db_info.db_type}_{db_info.module}"
        
        with self._db_processing_lock:
            if db_key in self._processed_dbs:
                self.logger.warning(f"{db_info.db_info}: 已在處理中，跳過重複處理")
                return db_info
            
            self._processed_dbs.add(db_key)
        
        try:
            self.logger.info(f"開始處理 {db_info.db_info} (Phase 1)")
            
            # 建立本地目錄
            local_path = os.path.join(self.output_dir, db_info.module, db_info.db_info)
            os.makedirs(local_path, exist_ok=True)
            db_info.local_path = local_path
            
            # 🔥 檢查磁碟空間（至少需要 15GB）
            import shutil
            free_space = shutil.disk_usage(local_path).free
            required_space = 15 * 1024 * 1024 * 1024  # 15GB
            
            if free_space < required_space:
                raise Exception(f"磁碟空間不足: 可用 {free_space/1024/1024/1024:.1f}GB，建議至少 15GB")
            
            self.logger.debug(f"{db_info.db_info}: 磁碟空間檢查通過 ({free_space/1024/1024/1024:.1f}GB 可用)")
            
            # 🔥 檢查文件系統類型和硬鏈接支持
            try:
                # 檢查是否支持硬鏈接
                test_file1 = os.path.join(local_path, f'.test_hardlink_src_{os.getpid()}')
                test_file2 = os.path.join(local_path, f'.test_hardlink_dst_{os.getpid()}')
                
                try:
                    with open(test_file1, 'w') as f:
                        f.write('test')
                    os.link(test_file1, test_file2)
                    os.unlink(test_file1)
                    os.unlink(test_file2)
                    self.logger.debug(f"{db_info.db_info}: 文件系統支持硬鏈接")
                except Exception as e:
                    self.logger.warning(f"{db_info.db_info}: 文件系統可能不支持硬鏈接: {e}")
                    # 清理測試文件
                    for test_file in [test_file1, test_file2]:
                        try:
                            if os.path.exists(test_file):
                                os.unlink(test_file)
                        except:
                            pass
                            
            except Exception as e:
                self.logger.debug(f"文件系統檢查失敗: {e}")
            
            # Step 1: SFTP 操作（線程安全）
            with self._sftp_lock:
                self.logger.info(f"{db_info.db_info}: 快速搜尋 manifest (線程安全)")
                
                local_sftp_manager = SFTPManager()
                
                if not local_sftp_manager.connect():
                    raise Exception("無法建立 SFTP 連線")
                
                target_version = db_info.version
                result = local_sftp_manager.find_latest_manifest(
                    db_info.sftp_path, 
                    db_info.db_info,
                    target_version
                )
                
                if not result:
                    raise Exception("找不到 manifest 檔案")
                
                manifest_full_path, manifest_name = result
                db_info.manifest_full_path = manifest_full_path
                db_info.manifest_file = manifest_name
                
                if not db_info.version:
                    match = re.match(r'manifest_(\d+)\.xml', manifest_name)
                    if match:
                        db_info.version = match.group(1)
                        self.logger.info(f"{db_info.db_info}: 檢測到版本 {db_info.version}")
                
                local_manifest = os.path.join(local_path, manifest_name)
                if not local_sftp_manager.download_file(manifest_full_path, local_manifest):
                    raise Exception("下載 manifest 失敗")
                
                self.logger.info(f"{db_info.db_info}: manifest 下載完成: {manifest_name}")
                
                local_sftp_manager.disconnect()
                self.logger.info(f"{db_info.db_info}: ✅ SFTP 連線已立即斷開")
                local_sftp_manager = None
            
            # Step 2: 檢查 repo 狀態
            db_info.has_existing_repo = self.repo_manager.check_repo_exists(local_path)
            
            # 🔥 Step 2.5: 清理檢查（根據配置決定是否執行）
            if (db_info.has_existing_repo and 
                config_manager.repo_config.get('check_clean_before_sync', False)):
                
                self.logger.info(f"{db_info.db_info}: 檢查工作空間是否乾淨")
                
                clean_status = self.repo_manager.check_repo_clean_status(local_path)
                
                if not clean_status['is_clean']:
                    self.logger.warning(f"{db_info.db_info}: 發現本地修改: {clean_status['details']}")
                    
                    # 處理髒repo
                    handle_success = self.repo_manager.handle_dirty_repo(
                        local_path, 
                        db_info.db_info, 
                        clean_status
                    )
                    
                    if not handle_success:
                        raise Exception(f"工作空間有未提交的修改且無法自動清理: {clean_status['details']}")
                    
                    self.logger.info(f"{db_info.db_info}: ✅ 工作空間清理完成")
                else:
                    self.logger.info(f"{db_info.db_info}: ✅ 工作空間狀態乾淨")
            else:
                if db_info.has_existing_repo:
                    self.logger.info(f"{db_info.db_info}: ⏭️ 跳過工作空間清理檢查（根據配置）")
                else:
                    self.logger.info(f"{db_info.db_info}: 新的工作空間，無需清理")
            
            # Step 3: 獲取 source command
            self.logger.info(f"{db_info.db_info}: 獲取 source command")
            source_cmd = self.source_cmd_manager.get_source_command(db_info, self.mapping_reader.df)
            if not source_cmd:
                raise Exception("無法取得 source command")
            
            db_info.actual_source_cmd = source_cmd
            self.logger.info(f"{db_info.db_info}: source command 獲取成功")
            
            # Step 4: 執行 repo init（如果需要）
            self.logger.info(f"{db_info.db_info}: 執行 repo 初始化")
            if not db_info.has_existing_repo:
                self.logger.info(f"{db_info.db_info}: .repo 目錄不存在，執行完整 repo init")
                if not self.repo_manager.repo_init(local_path, source_cmd):
                    raise Exception("Repo init 失敗")
            else:
                self.logger.info(f"{db_info.db_info}: .repo 目錄存在，跳過 repo init")
            
            # Step 5: 應用 manifest（這是關鍵步驟）
            self.logger.info(f"{db_info.db_info}: 開始應用 manifest（可能需要較長時間）")
            if not self.repo_manager.apply_manifest(local_path, local_manifest):
                raise Exception("套用 manifest 失敗")
            
            self.logger.info(f"{db_info.db_info}: ✅ repo 初始化完成")
            
            # Step 6: 啟動 repo sync
            if not self.dry_run:
                self.logger.info(f"{db_info.db_info}: 啟動 repo sync")
                process = self.repo_manager.start_repo_sync_async(local_path, db_info.db_info)
                if not process:
                    raise Exception("啟動 repo sync 失敗")
                
                db_info.sync_process = process
                self.logger.info(f"{db_info.db_info}: repo sync 已啟動 (PID: {process.pid})")
            else:
                db_info.status = DBStatus.SKIPPED
                db_info.end_time = datetime.now()
                self.logger.info(f"{db_info.db_info}: 測試模式 - 跳過 repo sync")
            
            self.logger.info(f"{db_info.db_info}: ✅ Phase 1 完成")
            
        except Exception as e:
            db_info.status = DBStatus.FAILED
            db_info.error_message = str(e)
            db_info.end_time = datetime.now()
            self.logger.error(f"{db_info.db_info}: ❌ Phase 1 失敗 - {str(e)}")
            
            # 🔥 處理失敗時移除處理標記，允許重試
            with self._db_processing_lock:
                self._processed_dbs.discard(db_key)
                
        finally:
            # 確保 SFTP 連線一定會被斷開
            if local_sftp_manager:
                try:
                    local_sftp_manager.disconnect()
                    self.logger.debug(f"{db_info.db_info}: 確保 SFTP 連線已斷開")
                except:
                    pass
        
        return db_info

    def process_db_phase2(self, db_info: DBInfo) -> DBInfo:
        """🔥 新版 Phase 2：Sync 完成後立即比較 manifest 並更新報告"""
        try:
            # 🔥 防止重複處理：如果已經有最終狀態，直接返回
            if db_info.status in [DBStatus.SUCCESS, DBStatus.SUCCESS_WITH_DIFF, DBStatus.FAILED, DBStatus.SKIPPED]:
                self.logger.debug(f"{db_info.db_info}: 已處理完成，狀態: {db_info.status.value}")
                return db_info
            
            self.logger.info(f"{db_info.db_info}: 開始 Phase 2")
            
            if self.dry_run:
                db_info.status = DBStatus.SKIPPED
                db_info.end_time = datetime.now()
                self.logger.info(f"{db_info.db_info}: 測試模式完成")
                return db_info
            
            # 檢查 sync 進程狀態
            sync_result = None
            if db_info.sync_process:
                poll = self.repo_manager.check_process_status(
                    db_info.db_info, 
                    db_info.sync_process
                )
                
                if poll is None:
                    self.logger.debug(f"{db_info.db_info}: repo sync 仍在執行中")
                    return db_info  # 還在執行中
                
                sync_result = poll
                
                # 記錄 sync log 路徑
                if hasattr(db_info.sync_process, '_log_file_path'):
                    db_info.sync_log_path = db_info.sync_process._log_file_path
            
            # 🔥 先檢查 sync 結果，只有成功的才嘗試導出 manifest
            if sync_result is None:
                # 沒有 sync 進程，可能是 dry run 或其他情況
                db_info.status = DBStatus.SKIPPED
                db_info.end_time = datetime.now()
                self.logger.info(f"{db_info.db_info}: 沒有 sync 進程，跳過")
                return db_info
            
            elif sync_result == 0:
                # 🎉 完全成功的情況
                self.logger.info(f"{db_info.db_info}: ✅ Sync 完成，立即導出版本資訊 ...")
                
                # Step 1: 立即導出 vp_manifest.xml
                exported_manifest_path = os.path.join(db_info.local_path, 'vp_manifest.xml')
                export_success = self.repo_manager.export_manifest(db_info.local_path, 'vp_manifest.xml')
                
                if export_success and os.path.exists(exported_manifest_path):
                    db_info.exported_manifest_path = exported_manifest_path
                    self.logger.info(f"{db_info.db_info}: ✅ manifest 導出成功")
                    
                    # Step 2: 立即比較 manifest
                    self.logger.info(f"{db_info.db_info}: 🔍 開始比較 manifest...")
                    db_info.status = DBStatus.COMPARING
                    
                    original_manifest_path = os.path.join(db_info.local_path, db_info.manifest_file)
                    comparison_result = self.manifest_comparator.compare_manifests(
                        original_manifest_path, 
                        exported_manifest_path
                    )
                    
                    # 保存比較結果
                    db_info.manifest_comparison_result = comparison_result
                    db_info.manifest_is_identical = comparison_result.get('is_identical', False)
                    db_info.manifest_differences_count = comparison_result.get('difference_count', 0)
                    db_info.manifest_comparison_summary = comparison_result.get('summary', '比較失敗')
                    
                    # 生成差異報告（如果有差異）
                    if not db_info.manifest_is_identical:
                        diff_report_path = os.path.join(db_info.local_path, f'manifest_diff_report.txt')
                        self.manifest_comparator.generate_diff_report(comparison_result, diff_report_path)
                        self.logger.info(f"{db_info.db_info}: 📄 差異報告已保存: {diff_report_path}")
                    
                    self.logger.info(f"{db_info.db_info}: 🔍 Manifest 比較完成: {db_info.manifest_comparison_summary}")
                    
                    # 設定最終狀態
                    if db_info.manifest_is_identical:
                        db_info.status = DBStatus.SUCCESS
                    else:
                        db_info.status = DBStatus.SUCCESS_WITH_DIFF
                        
                else:
                    self.logger.warning(f"{db_info.db_info}: ⚠️ manifest 導出失敗")
                    db_info.manifest_comparison_summary = "導出失敗，無法比較"
                    db_info.status = DBStatus.SUCCESS  # sync 成功但導出失敗，還是算成功
                
                db_info.end_time = datetime.now()
                elapsed = db_info.end_time - db_info.start_time
                self.logger.info(f"{db_info.db_info}: ✅ 完全成功 (耗時: {elapsed})")
            
            elif sync_result == 1:
                # 🔄 可能的部分成功情況
                self.logger.info(f"{db_info.db_info}: 🔍 分析部分成功結果...")
                success_rate, failed_projects = self._analyze_sync_result(db_info)
                
                if success_rate >= 95.0:  # 🔥 95% 以上成功率算成功
                    self.logger.info(f"{db_info.db_info}: ✅ 部分成功 ({success_rate:.1f}%)，嘗試導出版本資訊...")
                    
                    # 嘗試導出 manifest
                    exported_manifest_path = os.path.join(db_info.local_path, 'vp_manifest.xml')
                    export_success = self.repo_manager.export_manifest(db_info.local_path, 'vp_manifest.xml')
                    
                    if export_success and os.path.exists(exported_manifest_path):
                        # 有成功導出，進行比較
                        db_info.exported_manifest_path = exported_manifest_path
                        self.logger.info(f"{db_info.db_info}: 🔍 開始比較 manifest...")
                        
                        original_manifest_path = os.path.join(db_info.local_path, db_info.manifest_file)
                        comparison_result = self.manifest_comparator.compare_manifests(
                            original_manifest_path, 
                            exported_manifest_path
                        )
                        
                        db_info.manifest_comparison_result = comparison_result
                        db_info.manifest_is_identical = comparison_result.get('is_identical', False)
                        db_info.manifest_differences_count = comparison_result.get('difference_count', 0)
                        db_info.manifest_comparison_summary = comparison_result.get('summary', '比較失敗')
                        
                        if db_info.manifest_is_identical:
                            db_info.status = DBStatus.SUCCESS
                        else:
                            db_info.status = DBStatus.SUCCESS_WITH_DIFF
                    else:
                        # 導出失敗但 sync 部分成功
                        db_info.status = DBStatus.SUCCESS
                        db_info.manifest_comparison_summary = "部分成功但導出失敗"
                    
                    warning_msg = f"部分成功 ({success_rate:.1f}%)，失敗項目: {len(failed_projects)} 個"
                    if failed_projects:
                        warning_msg += f" - {', '.join(failed_projects[:3])}"
                        if len(failed_projects) > 3:
                            warning_msg += f" 等 {len(failed_projects)} 個"
                    
                    db_info.error_message = warning_msg
                else:
                    # 成功率太低，算失敗
                    raise Exception(f"同步成功率太低 ({success_rate:.1f}%)，失敗項目: {len(failed_projects)} 個")
                
                db_info.end_time = datetime.now()
                elapsed = db_info.end_time - db_info.start_time
                self.logger.info(f"{db_info.db_info}: ✅ 部分成功完成 (耗時: {elapsed})")
            
            else:
                # 🚫 其他錯誤碼，直接算失敗，不嘗試導出
                if sync_result == -15:
                    error_msg = "Sync 被終止 (SIGTERM)"
                elif sync_result == -9:
                    error_msg = "Sync 被強制終止 (SIGKILL)"
                elif sync_result < 0:
                    error_msg = f"Sync 被信號終止 (信號: {-sync_result})"
                else:
                    error_msg = f"Sync 失敗 (返回碼: {sync_result})"
                
                raise Exception(error_msg)
            
            # Step 3: 🔥 立即更新報告到 Excel（不等最後）
            self._update_report_immediately(db_info)
            
        except Exception as e:
            db_info.status = DBStatus.FAILED
            db_info.error_message = str(e)
            db_info.end_time = datetime.now()
            self.logger.error(f"{db_info.db_info}: Phase 2 失敗 - {str(e)}")
            
            # 失敗的也要立即更新報告
            self._update_report_immediately(db_info)
        
        return db_info

    def _update_report_immediately(self, db_info: DBInfo):
        """🔥 修復版本：即時更新單個 DB 的報告到 Excel - 防止重複"""
        try:
            with self._report_update_lock:
                self.logger.info(f"{db_info.db_info}: 📊 立即更新報告...")
                
                # 🔥 更精確的查找邏輯 - 使用多個條件
                found_index = -1
                for i, existing_db in enumerate(self.report.db_details):
                    # 🔥 使用更嚴格的比較條件
                    if (existing_db.db_info == db_info.db_info and 
                        existing_db.db_type == db_info.db_type and
                        existing_db.module == db_info.module):
                        found_index = i
                        break
                
                if found_index >= 0:
                    # 🔥 更新現有記錄 - 保留較新的資料
                    existing_db = self.report.db_details[found_index]
                    
                    # 只更新有值的欄位，避免覆蓋有用資料
                    if db_info.status != DBStatus.PENDING:
                        existing_db.status = db_info.status
                    if db_info.end_time:
                        existing_db.end_time = db_info.end_time
                    if db_info.error_message:
                        existing_db.error_message = db_info.error_message
                    if db_info.sync_log_path:
                        existing_db.sync_log_path = db_info.sync_log_path
                    if db_info.exported_manifest_path:
                        existing_db.exported_manifest_path = db_info.exported_manifest_path
                    if db_info.manifest_comparison_result:
                        existing_db.manifest_comparison_result = db_info.manifest_comparison_result
                        existing_db.manifest_is_identical = db_info.manifest_is_identical
                        existing_db.manifest_differences_count = db_info.manifest_differences_count
                        existing_db.manifest_comparison_summary = db_info.manifest_comparison_summary
                    
                    self.logger.debug(f"{db_info.db_info}: 更新現有記錄 (索引: {found_index})")
                else:
                    # 🔥 添加新記錄前再次確認不重複
                    duplicate_check = any(
                        existing.db_info == db_info.db_info and 
                        existing.db_type == db_info.db_type and
                        existing.module == db_info.module
                        for existing in self.report.db_details
                    )
                    
                    if not duplicate_check:
                        self.report.db_details.append(db_info)
                        self.logger.debug(f"{db_info.db_info}: 添加新記錄")
                    else:
                        self.logger.warning(f"{db_info.db_info}: 發現重複，跳過添加")
                
                # 🔥 清理重複資料
                self._remove_duplicates()
                
                # 重新計算統計
                self.report.total_dbs = len(self.report.db_details)
                self.report.successful_dbs = sum(1 for db in self.report.db_details 
                                            if db.status in [DBStatus.SUCCESS, DBStatus.SUCCESS_WITH_DIFF])
                self.report.failed_dbs = sum(1 for db in self.report.db_details if db.status == DBStatus.FAILED)
                self.report.skipped_dbs = sum(1 for db in self.report.db_details if db.status == DBStatus.SKIPPED)
                self.report.dbs_with_differences = sum(1 for db in self.report.db_details if db.status == DBStatus.SUCCESS_WITH_DIFF)
                
                # 立即生成並更新 Excel 報告
                report_path = os.path.join(self.output_dir, config_manager.path_config['report_filename'])
                self._generate_partial_report(report_path)
                
                self.logger.info(f"{db_info.db_info}: ✅ 報告已立即更新")
                
        except Exception as e:
            self.logger.warning(f"{db_info.db_info}: 立即更新報告失敗: {e}")
            
    # 🔥 新增：清理重複資料的方法
    def _remove_duplicates(self):
        """清理報告中的重複資料"""
        try:
            seen = set()
            unique_dbs = []
            
            for db in self.report.db_details:
                # 建立唯一識別符
                key = (db.db_info, db.db_type, db.module)
                
                if key not in seen:
                    seen.add(key)
                    unique_dbs.append(db)
                else:
                    self.logger.debug(f"移除重複記錄: {db.db_info} ({db.db_type}, {db.module})")
            
            # 如果發現重複，更新列表
            if len(unique_dbs) != len(self.report.db_details):
                removed_count = len(self.report.db_details) - len(unique_dbs)
                self.report.db_details = unique_dbs
                self.logger.info(f"清理了 {removed_count} 筆重複記錄")
                
        except Exception as e:
            self.logger.warning(f"清理重複資料失敗: {e}")
                        
    def _update_report_immediately(self, db_info: DBInfo):
        """🔥 新增：立即更新單個 DB 的報告到 Excel"""
        try:
            with self._report_update_lock:
                self.logger.info(f"{db_info.db_info}: 📊 立即更新報告...")
                
                # 更新或添加到報告中
                found = False
                for i, existing_db in enumerate(self.report.db_details):
                    if existing_db.db_info == db_info.db_info and existing_db.db_type == db_info.db_type:
                        self.report.db_details[i] = db_info
                        found = True
                        break
                
                if not found:
                    self.report.db_details.append(db_info)
                
                # 重新計算統計
                self.report.total_dbs = len(self.report.db_details)
                self.report.successful_dbs = sum(1 for db in self.report.db_details 
                                               if db.status in [DBStatus.SUCCESS, DBStatus.SUCCESS_WITH_DIFF])
                self.report.failed_dbs = sum(1 for db in self.report.db_details if db.status == DBStatus.FAILED)
                self.report.skipped_dbs = sum(1 for db in self.report.db_details if db.status == DBStatus.SKIPPED)
                self.report.dbs_with_differences = sum(1 for db in self.report.db_details if db.status == DBStatus.SUCCESS_WITH_DIFF)
                
                # 立即生成並更新 Excel 報告
                report_path = os.path.join(self.output_dir, config_manager.path_config['report_filename'])
                self._generate_partial_report(report_path)
                
                self.logger.info(f"{db_info.db_info}: ✅ 報告已立即更新")
                
        except Exception as e:
            self.logger.warning(f"{db_info.db_info}: 立即更新報告失敗: {e}")

    def _wait_and_process_syncs_enhanced(self, db_results: List[DBInfo]):
        """🔥 新版：等待同步完成並即時處理，避免重複處理"""
        max_wait_time = config_manager.repo_config['sync_timeout']
        start_wait = time.time()
        
        active_syncs = [db for db in db_results if db.sync_process and db.status != DBStatus.FAILED]
        self.logger.info(f"監控 {len(active_syncs)} 個活躍的 repo sync 進程")
        
        # 初始化進度追蹤
        progress_tracker = {}
        processed_dbs = set()  # 🔥 追蹤已處理的 DB，避免重複處理
        
        for db_info in active_syncs:
            progress_tracker[db_info.db_info] = {
                'start_time': db_info.start_time or datetime.now(),
                'last_log_size': 0,
                'estimated_progress': 0,
                'current_activity': '初始化中...',
                'log_file': self._get_sync_log_file(db_info),
                'last_check_time': datetime.now(),
                'error_count': 0,
                'critical_errors': []
            }
        
        check_interval = 30  # 30秒檢查一次
        
        while True:
            all_complete = True
            elapsed = int(time.time() - start_wait)
            current_time = time.time()
            
            print("\n" + "="*100)
            print(f"📊 Repo Sync 進度監控 - 已等待 {elapsed}s")
            print("="*100)
            
            current_failed_count = 0
            current_running_count = 0
            current_completed_count = 0
            
            for db_info in active_syncs:
                db_name = db_info.db_info
                
                # 🔥 如果已經處理過，跳過
                if db_name in processed_dbs:
                    current_completed_count += 1
                    continue
                
                if db_info.status == DBStatus.FAILED:
                    current_failed_count += 1
                    processed_dbs.add(db_name)
                    continue
                
                tracker = progress_tracker[db_name]
                
                # 構建包含版本信息的顯示名稱
                manifest_info = ""
                if db_info.manifest_file:
                    manifest_info = f" ({db_info.manifest_file})"
                elif db_info.version:
                    manifest_info = f" (v{db_info.version})"
                
                display_name = f"{db_name}{manifest_info}"
                
                if db_info.sync_process:
                    try:
                        poll = db_info.sync_process.poll()
                        
                        if poll is None:  # 仍在運行
                            all_complete = False
                            current_running_count += 1
                            
                            # 更新進度信息
                            self._update_progress_info(db_info, tracker)
                            
                            runtime = datetime.now() - tracker['start_time']
                            runtime_str = f"{int(runtime.total_seconds()//60)}:{int(runtime.total_seconds()%60):02d}"
                            
                            progress = tracker['estimated_progress']
                            bar_length = 20
                            filled = int(bar_length * progress / 100)
                            bar = "█" * filled + "▒" * (bar_length - filled)
                            
                            activity = tracker.get('current_project', '').split('/')[-1] or tracker.get('current_activity', '同步中')
                            if len(activity) > 15:
                                activity = activity[:12] + "..."
                            
                            # 顯示錯誤狀態
                            status_info = ""
                            if tracker['critical_errors']:
                                status_info = f" ⚠️{len(tracker['critical_errors'])}"
                            
                            print(f"🔄 {display_name:30s} │{bar}│ {progress:3d}% │ {runtime_str} │ {activity}{status_info}")
                            
                            # 檢查超時
                            if time.time() - start_wait > max_wait_time:
                                self.logger.warning(f"{db_name}: repo sync 超時，強制終止")
                                try:
                                    db_info.sync_process.terminate()
                                    db_info.sync_process.wait(timeout=5)
                                except:
                                    try:
                                        db_info.sync_process.kill()
                                    except:
                                        pass
                                db_info.status = DBStatus.FAILED
                                db_info.error_message = "Sync 超時"
                                current_failed_count += 1
                                processed_dbs.add(db_name)
                                print(f"⏰ {display_name:30s} │ 超時終止")
                                
                        else:  # 進程已結束
                            # 🔥 立即處理這個 DB，只處理一次
                            if db_name not in processed_dbs:
                                processed_dbs.add(db_name)
                                
                                # 在後台線程中處理 Phase 2，避免阻塞監控
                                def process_phase2_async(db):
                                    try:
                                        processed_db = self.process_db_phase2(db)
                                    except Exception as e:
                                        self.logger.error(f"{db.db_info}: 後台處理 Phase 2 失敗: {e}")
                                
                                # 啟動後台處理
                                import threading
                                phase2_thread = threading.Thread(target=process_phase2_async, args=(db_info,), daemon=True)
                                phase2_thread.start()
                                
                                runtime = datetime.now() - tracker['start_time']
                                runtime_str = f"{int(runtime.total_seconds()//60)}:{int(runtime.total_seconds()%60):02d}"
                                
                                if poll == 0:  # 成功完成
                                    current_completed_count += 1
                                    bar = "█" * 20
                                    print(f"✅ {display_name:30s} │{bar}│ 100% │ {runtime_str} │ 完成+處理中")
                                else:  # 失敗
                                    current_failed_count += 1
                                    error_msg = f"Sync 失敗 (返回碼: {poll})"
                                    print(f"❌ {display_name:30s} │{'':20s}│   0% │ {runtime_str} │ {error_msg[:30]}")
                            else:
                                # 已經處理過的
                                current_completed_count += 1
                            
                    except Exception as e:
                        self.logger.error(f"{db_name}: 檢查進程狀態失敗: {e}")
                        if db_name not in processed_dbs:
                            processed_dbs.add(db_name)
                            db_info.status = DBStatus.FAILED
                            db_info.error_message = f"進程監控失敗: {e}"
                            self.report.add_db(db_info)
                            current_failed_count += 1
                            print(f"⚠️  {display_name:30s} │ 監控錯誤 │   0% │ {str(e)[:30]}")
            
            # 顯示總體統計
            print("-"*100)
            print(f"📈 總計: 運行中 {current_running_count} | 完成 {current_completed_count} | 失敗 {current_failed_count}")
            
            if all_complete or (time.time() - start_wait) > max_wait_time:
                break
            
            # 等待下次檢查
            time.sleep(check_interval)
        
        # 等待所有後台 Phase 2 處理完成
        self.logger.info("等待所有後台處理完成...")
        time.sleep(5)  # 給後台線程一些時間完成
        
        # 最終統計
        completed = len([db for db in active_syncs if db.db_info in processed_dbs and 
                        db.status in [DBStatus.SUCCESS, DBStatus.SUCCESS_WITH_DIFF]])
        failed = len([db for db in active_syncs if db.status == DBStatus.FAILED])
        
        print(f"\n📋 Repo sync 最終統計:")
        print(f"   ✅ 成功: {completed}")
        print(f"   ❌ 失敗: {failed}")
        
        self.logger.info(f"📋 Repo sync 完成統計: 成功 {completed}, 失敗 {failed}")
        
    def _generate_partial_report(self, output_file: str):
        """🔥 新增：生成部分報告（用於即時更新）"""
        try:
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            
            report_data = []
            for i, db in enumerate(self.report.db_details, 1):
                # 手動構建字典，確保所有欄位正確
                db_dict = {
                    'sn': i,
                    'module': db.module,
                    'db_type': db.db_type,
                    'db_info': db.db_info,
                    'status': db.status.value if hasattr(db.status, 'value') else str(db.status),
                    'version': db.version or '未指定',
                    'manifest_file': db.manifest_file or '',
                    'start_time': db.start_time.strftime('%Y-%m-%d %H:%M:%S') if db.start_time else '',
                    'end_time': db.end_time.strftime('%Y-%m-%d %H:%M:%S') if db.end_time else '',
                    'error_message': db.error_message or '',
                    'sync_log_path': db.sync_log_path or '',
                    'sftp_path': db.sftp_path,
                    'local_path': db.local_path or '',
                    'has_existing_repo': '是' if db.has_existing_repo else '否',
                    # 'jira_link': db.jira_link or '未找到',
                    # 🔥 新增 manifest 比較相關欄位
                    'manifest_比較結果': db.manifest_comparison_summary or '',
                    'manifest_是否相同': '是' if db.manifest_is_identical else '否' if db.manifest_is_identical is False else '',
                    'manifest_差異數量': db.manifest_differences_count or 0,
                    'exported_manifest_path': db.exported_manifest_path or '',
                }
                
                # 重新命名欄位
                db_dict['完整_JIRA_連結'] = db.jira_link or '未找到'
                db_dict['完整_repo_init_指令'] = db.actual_source_cmd or '未記錄'
                
                report_data.append(db_dict)
            
            df = pd.DataFrame(report_data)
            
            # 重新排列欄位順序
            important_columns = [
                'sn', 'module', 'db_type', 'db_info', 'status', 'version', 'manifest_file',
                'manifest_比較結果', 'manifest_是否相同', 'manifest_差異數量',
                '完整_JIRA_連結', '完整_repo_init_指令',
                'start_time', 'end_time', 'sync_log_path', 'error_message'
            ]
            
            existing_columns = [col for col in important_columns if col in df.columns]
            other_columns = [col for col in df.columns if col not in important_columns]
            df = df[existing_columns + other_columns]
            
            # 🔥 修正：重新計算統計，基於實際的 status 值
            status_counts = df['status'].value_counts()
            successful_count = status_counts.get('✅ 完成', 0) + status_counts.get('✅ 完成(有差異)', 0)
            failed_count = status_counts.get('❌ 失敗', 0)
            skipped_count = status_counts.get('⭐️ 跳過', 0)
            dbs_with_diff = status_counts.get('✅ 完成(有差異)', 0)
            
            # 建立摘要
            current_time = datetime.now()
            summary = {
                '項目': ['總 DB 數', '成功', '失敗', '跳過', '有差異', '最後更新時間'],
                '數值': [
                    len(self.report.db_details),
                    successful_count,
                    failed_count,
                    skipped_count,
                    dbs_with_diff,
                    current_time.strftime('%Y-%m-%d %H:%M:%S')
                ]
            }
            summary_df = pd.DataFrame(summary)
            
            # 寫入改進版 Excel
            self._write_enhanced_excel(df, summary_df, output_file)
            
        except Exception as e:
            self.logger.warning(f"生成部分報告失敗: {str(e)}")

    def _analyze_sync_result(self, db_info: DBInfo) -> tuple:
        """分析 sync 結果，精確提取失敗項目 - 清理版本"""
        try:
            log_file = db_info.sync_log_path
            if not log_file:
                log_file = self._get_sync_log_file(db_info)
            
            if not log_file or not os.path.exists(log_file):
                return 0.0, []
            
            total_projects = 0
            failed_projects = set()
            
            with open(log_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                
                # 🔥 先清理 ANSI 轉義字符
                import re
                clean_content = re.sub(r'\x1b\[[0-9;]*[mK]', '', content)
                
                # 提取同步統計
                sync_pattern = r'Syncing:\s*\d+%\s*\((\d+)/(\d+)\)'
                matches = re.findall(sync_pattern, clean_content)
                
                if matches:
                    last_match = matches[-1]
                    completed = int(last_match[0])
                    total_projects = int(last_match[1])
                    self.logger.debug(f"{db_info.db_info}: 同步統計 - 完成 {completed}/{total_projects}")
                
                # 🔥 精確提取失敗項目 - 只從特定錯誤部分提取
                
                # 方式1：從 "The following projects failed" 到下一個錯誤部分
                failed_section_match = re.search(
                    r'The following projects failed[^:]*:\s*(.*?)(?=\n\s*error:|$)', 
                    clean_content, 
                    re.DOTALL | re.IGNORECASE
                )
                
                if failed_section_match:
                    failed_text = failed_section_match.group(1)
                    # 提取有效的項目路徑
                    project_lines = [line.strip() for line in failed_text.split('\n') if line.strip()]
                    for line in project_lines:
                        # 只接受以 kernel/ 開頭的有效路徑
                        if line.startswith('kernel/') and '/' in line and len(line) > 10:
                            # 清理可能的尾部字符
                            clean_project = re.sub(r'[^\w/\-_.].*$', '', line)
                            if clean_project and clean_project.count('/') >= 2:  # 至少包含 kernel/xxx/yyy
                                failed_projects.add(clean_project)
                
                # 方式2：從具體的 GitCommandError 消息中提取
                git_error_pattern = r"GitCommandError:.*?on\s+(kernel/[^\s'\"]+)\s+failed"
                git_errors = re.findall(git_error_pattern, clean_content, re.IGNORECASE)
                for project in git_errors:
                    # 清理項目路徑
                    clean_project = re.sub(r'[^\w/\-_.].*$', '', project)
                    if clean_project and clean_project.count('/') >= 2:
                        failed_projects.add(clean_project)
                
                # 方式3：從 "fatal: not a git repository" 錯誤推斷項目
                repo_error_pattern = r"fatal: not a git repository.*?/([^/\s'\"]+)\.git"
                repo_errors = re.findall(repo_error_pattern, clean_content)
                
                # 根據項目名稱推斷完整路徑
                known_project_mappings = {
                    'vts': 'kernel/android/U/test/vts',
                    'hal': 'kernel/android/U/test/vts-testcase/hal',
                    'hal-trace': 'kernel/android/U/test/vts-testcase/hal-trace'
                }
                
                for project_name in repo_errors:
                    if project_name in known_project_mappings:
                        failed_projects.add(known_project_mappings[project_name])
            
            # 🔥 清理和驗證失敗項目列表
            valid_failed_projects = []
            for project in failed_projects:
                # 驗證項目路徑格式
                if (project.startswith('kernel/') and 
                    project.count('/') >= 2 and 
                    len(project) > 10 and
                    not any(char in project for char in ['\x1b', '"', "'", ':', ' ']) and
                    re.match(r'^kernel/[a-zA-Z0-9_\-/]+$', project)):
                    valid_failed_projects.append(project)
            
            # 去重並排序
            valid_failed_projects = sorted(list(set(valid_failed_projects)))
            
            # 計算成功率
            if total_projects > 0:
                success_count = total_projects - len(valid_failed_projects)
                success_rate = (success_count / total_projects) * 100
                
                self.logger.info(f"{db_info.db_info}: 清理後分析 - 總計:{total_projects}, 成功:{success_count}, 失敗:{len(valid_failed_projects)}, 成功率:{success_rate:.1f}%")
                self.logger.info(f"{db_info.db_info}: 有效失敗項目: {valid_failed_projects}")
                
                return success_rate, valid_failed_projects
            
            return 0.0, []
            
        except Exception as e:
            self.logger.debug(f"分析同步結果失敗: {e}")
            return 0.0, []

    def _get_sync_log_file(self, db_info: DBInfo) -> str:
        """獲取 sync 日誌文件路徑 - 優先使用 unbuffer 版本"""
        try:
            log_dir = os.path.join(db_info.local_path, 'logs')
            if os.path.exists(log_dir):
                # 尋找最新的日誌文件，優先選擇 unbuffer 版本
                log_files = []
                for f in os.listdir(log_dir):
                    if f.startswith('repo_sync_') and f.endswith('.log'):
                        file_path = os.path.join(log_dir, f)
                        mtime = os.path.getmtime(file_path)
                        
                        # 🔥 給不同類型的日誌不同優先級
                        priority = 0
                        if 'unbuffer' in f:
                            priority = 100  # 最高優先級
                        elif 'script' in f:
                            priority = 50
                        elif 'hotfix' in f or 'fixed' in f:
                            priority = 25
                        # 舊版本的日誌優先級為 0
                        
                        log_files.append((priority, mtime, file_path))
                
                if log_files:
                    # 按優先級和修改時間排序
                    log_files.sort(key=lambda x: (x[0], x[1]), reverse=True)
                    latest_log = log_files[0][2]
                    self.logger.debug(f"{db_info.db_info}: 使用日誌文件: {os.path.basename(latest_log)}")
                    return latest_log
            
            self.logger.debug(f"{db_info.db_info}: 日誌目錄不存在: {log_dir}")
        except Exception as e:
            self.logger.debug(f"{db_info.db_info}: 獲取日誌文件失敗: {e}")
        
        return ""

    def _wait_for_all_syncs_enhanced(self, db_results: List[DBInfo]):
        """完整版進度監控 - 🔥 修復重複日誌問題"""
        max_wait_time = config_manager.repo_config['sync_timeout']
        start_wait = time.time()
        
        active_syncs = [db for db in db_results if db.sync_process and db.status != DBStatus.FAILED]
        self.logger.info(f"監控 {len(active_syncs)} 個活躍的 repo sync 進程")
        
        # 初始化進度追蹤
        progress_tracker = {}
        for db_info in active_syncs:
            progress_tracker[db_info.db_info] = {
                'start_time': db_info.start_time or datetime.now(),
                'last_log_size': 0,
                'estimated_progress': 0,
                'current_activity': '初始化中...',
                'log_file': self._get_sync_log_file(db_info),
                'last_check_time': datetime.now(),
                'error_count': 0,
                'critical_errors': []  # 🔥 記錄嚴重錯誤
            }
        
        check_interval = 30  # 30秒檢查一次
        
        while True:
            all_complete = True
            elapsed = int(time.time() - start_wait)
            current_time = time.time()
            
            print("\n" + "="*100)
            print(f"📊 Repo Sync 進度監控 - 已等待 {elapsed}s")
            print("="*100)
            
            current_failed_count = 0
            current_running_count = 0
            current_completed_count = 0
            
            for db_info in active_syncs:
                if db_info.status == DBStatus.FAILED:
                    current_failed_count += 1
                    continue
                
                db_name = db_info.db_info
                tracker = progress_tracker[db_name]
                
                # 構建包含版本信息的顯示名稱
                manifest_info = ""
                if db_info.manifest_file:
                    manifest_info = f" ({db_info.manifest_file})"
                elif db_info.version:
                    manifest_info = f" (v{db_info.version})"
                
                display_name = f"{db_name}{manifest_info}"
                
                if db_info.sync_process:
                    try:
                        poll = db_info.sync_process.poll()
                        
                        if poll is None:  # 仍在運行
                            all_complete = False
                            current_running_count += 1
                            
                            # 更新進度信息
                            self._update_progress_info(db_info, tracker)
                            
                            runtime = datetime.now() - tracker['start_time']
                            runtime_str = f"{int(runtime.total_seconds()//60)}:{int(runtime.total_seconds()%60):02d}"
                            
                            progress = tracker['estimated_progress']
                            bar_length = 20
                            filled = int(bar_length * progress / 100)
                            bar = "█" * filled + "▒" * (bar_length - filled)
                            
                            activity = tracker.get('current_project', '').split('/')[-1] or tracker.get('current_activity', '同步中')
                            if len(activity) > 15:
                                activity = activity[:12] + "..."
                            
                            # 🔥 顯示錯誤狀態
                            status_info = ""
                            if tracker['critical_errors']:
                                status_info = f" ⚠️{len(tracker['critical_errors'])}"
                            
                            print(f"🔄 {display_name:30s} │{bar}│ {progress:3d}% │ {runtime_str} │ {activity}{status_info}")
                            
                            # 檢查超時
                            if time.time() - start_wait > max_wait_time:
                                self.logger.warning(f"{db_name}: repo sync 超時，強制終止")
                                try:
                                    db_info.sync_process.terminate()
                                    db_info.sync_process.wait(timeout=5)
                                except:
                                    try:
                                        db_info.sync_process.kill()
                                    except:
                                        pass
                                db_info.status = DBStatus.FAILED
                                db_info.error_message = "Sync 超時"
                                current_failed_count += 1
                                print(f"⏰ {display_name:30s} │ 超時終止")
                                
                        elif poll == 0:  # 成功完成
                            current_completed_count += 1
                            runtime = datetime.now() - tracker['start_time']
                            runtime_str = f"{int(runtime.total_seconds()//60)}:{int(runtime.total_seconds()%60):02d}"
                            
                            bar = "█" * 20
                            print(f"✅ {display_name:30s} │{bar}│ 100% │ {runtime_str} │ 完成+已導出")
                            
                        else:  # 失敗 (poll != 0)
                            current_failed_count += 1
                            runtime = datetime.now() - tracker['start_time']
                            runtime_str = f"{int(runtime.total_seconds()//60)}:{int(runtime.total_seconds()%60):02d}"
                            
                            # 🔥 分析失敗原因
                            error_msg = f"Sync 失敗 (返回碼: {poll})"
                            
                            # 檢查是否為部分失敗
                            if poll == 1:  # 部分失敗，可能可以救活
                                success_rate, failed_projects = self._analyze_sync_result(db_info)
                                if failed_projects:
                                    error_msg += f" - {len(failed_projects)} 個項目失敗"
                            
                            db_info.status = DBStatus.FAILED
                            db_info.error_message = error_msg
                            print(f"❌ {display_name:30s} │{'':20s}│   0% │ {runtime_str} │ {error_msg[:30]}")
                            
                    except Exception as e:
                        self.logger.error(f"{db_name}: 檢查進程狀態失敗: {e}")
                        db_info.status = DBStatus.FAILED
                        db_info.error_message = f"進程監控失敗: {e}"
                        current_failed_count += 1
                        print(f"⚠️  {display_name:30s} │ 監控錯誤 │   0% │ {str(e)[:30]}")
            
            # 顯示總體統計
            print("-"*100)
            print(f"📈 總計: 運行中 {current_running_count} | 完成 {current_completed_count} | 失敗 {current_failed_count}")
            
            if all_complete or (time.time() - start_wait) > max_wait_time:
                break
            
            # 等待下次檢查
            time.sleep(check_interval)
        
        # 最終統計
        completed = sum(1 for db in active_syncs if db.sync_process and db.sync_process.poll() == 0)
        failed = sum(1 for db in active_syncs if db.status == DBStatus.FAILED)
        
        print(f"\n📋 Repo sync 最終統計:")
        print(f"   ✅ 成功: {completed}")
        print(f"   ❌ 失敗: {failed}")
        
        self.logger.info(f"📋 Repo sync 完成統計: 成功 {completed}, 失敗 {failed}")

    def _update_progress_info(self, db_info: DBInfo, tracker: dict):
        """更新進度信息 - 專門優化 unbuffer 輸出解析"""
        try:
            log_file = tracker.get('log_file')
            
            # 🔥 每次都重新獲取日誌文件，確保使用最新的 unbuffer 日誌
            current_log_file = self._get_sync_log_file(db_info)
            if current_log_file and current_log_file != log_file:
                tracker['log_file'] = current_log_file
                log_file = current_log_file
                self.logger.debug(f"{db_info.db_info}: 切換到新日誌文件: {os.path.basename(log_file)}")
            
            if not log_file or not os.path.exists(log_file):
                tracker['current_activity'] = '等待日誌...'
                tracker['estimated_progress'] = self._get_time_based_progress(tracker)
                return
            
            # 🔥 優化的日誌解析 - 專門處理 unbuffer 格式
            try:
                file_size = os.path.getsize(log_file)
                
                # 只讀取最後 2KB 避免處理大文件
                read_size = min(file_size, 2048)
                
                with open(log_file, 'r', encoding='utf-8', errors='ignore') as f:
                    f.seek(max(0, file_size - read_size))
                    content = f.read()
                    lines = content.split('\n')[-15:]  # 最後15行
                
                # 解析最新的同步狀態
                latest_progress = 0
                latest_activity = "同步中..."
                current_project = ""
                total_projects = 0
                current_count = 0
                
                # 🔥 重點：解析 unbuffer 輸出的特定格式
                for line in reversed(lines):
                    line = line.strip()
                    if not line:
                        continue
                    
                    # 解析 "Syncing: X% (current/total) time | jobs | project" 格式
                    if "Syncing:" in line and "%" in line:
                        import re
                        
                        # 匹配進度百分比和計數
                        progress_match = re.search(r'Syncing:\s*(\d+)%\s*\((\d+)/(\d+)\)', line)
                        if progress_match:
                            latest_progress = int(progress_match.group(1))
                            current_count = int(progress_match.group(2))
                            total_projects = int(progress_match.group(3))
                            
                            # 提取當前處理的項目
                            # 格式通常是: "... | X jobs | time project_path @ ..."
                            if '|' in line:
                                parts = line.split('|')
                                for part in parts:
                                    part = part.strip()
                                    # 尋找包含項目路徑的部分
                                    if '@' in part:
                                        project_info = part.split('@')[-1].strip()
                                    elif '/' in part and 'job' not in part and ':' not in part:
                                        project_info = part
                                    else:
                                        continue
                                    
                                    # 提取項目名稱
                                    if '/' in project_info:
                                        current_project = project_info.split('/')[-1]
                                    else:
                                        current_project = project_info
                                    break
                            
                            # 構建活動描述
                            latest_activity = f"同步: {current_count}/{total_projects}"
                            if current_project:
                                # 限制項目名稱長度
                                project_name = current_project[:20] + "..." if len(current_project) > 20 else current_project
                                latest_activity += f" - {project_name}"
                            
                            break
                    
                    # 🔥 解析其他狀態信息
                    elif "Fetching project" in line:
                        project_match = re.search(r'Fetching project\s+([^\s]+)', line)
                        if project_match:
                            project_path = project_match.group(1)
                            current_project = project_path.split('/')[-1]
                            latest_activity = f"獲取: {current_project}"
                    
                    elif "Skipped fetching project" in line:
                        project_match = re.search(r'Skipped fetching project\s+([^\s]+)', line)
                        if project_match:
                            project_path = project_match.group(1)
                            current_project = project_path.split('/')[-1]
                            latest_activity = f"跳過: {current_project}"
                
                # 🔥 更新追蹤信息
                if latest_progress > 0:
                    tracker['estimated_progress'] = latest_progress
                else:
                    # 如果沒有解析到進度，使用時間估算
                    tracker['estimated_progress'] = self._get_time_based_progress(tracker)
                
                tracker['current_activity'] = latest_activity
                tracker['current_project'] = current_project
                tracker['total_projects'] = total_projects
                tracker['current_count'] = current_count
                tracker['last_update'] = datetime.now()
                
                # 🔥 調試信息（可選）
                if latest_progress > 0:
                    self.logger.debug(f"{db_info.db_info}: 解析成功 - {latest_progress}% ({current_count}/{total_projects}) {current_project}")
                    
            except Exception as e:
                self.logger.debug(f"{db_info.db_info}: 解析日誌失敗: {e}")
                tracker['current_activity'] = '解析失敗'
                tracker['estimated_progress'] = self._get_time_based_progress(tracker)
                
        except Exception as e:
            self.logger.debug(f"{db_info.db_info}: 進度更新失敗: {e}")
            tracker['current_activity'] = '更新失敗'
            tracker['estimated_progress'] = self._get_time_based_progress(tracker)

    def _get_time_based_progress(self, tracker: dict) -> int:
        """基於時間的進度估算"""
        runtime_minutes = (datetime.now() - tracker['start_time']).total_seconds() / 60
        # 每分鐘約1.5%的進度，最多95%
        return min(int(runtime_minutes * 1.5), 95)

    def process_dbs_async(self, db_list: List[str], db_versions: Dict[str, str] = None):
        """異步處理多個 DB"""
        db_versions = db_versions or {}
        db_infos = []
        
        # 準備 DB 資訊
        all_db_infos = self.get_all_dbs('all')
        
        for db_name in db_list:
            if '#' in db_name:
                db_name, version = db_name.split('#', 1)
            else:
                version = db_versions.get(db_name)
            
            for db_info in all_db_infos:
                if db_info.db_info == db_name:
                    db_info.version = version
                    db_infos.append(db_info)
                    break
        
        if not db_infos:
            self.logger.error("沒有找到要處理的 DB")
            return

        self.logger.info(f"開始處理 {len(db_infos)} 個 DB")
        
        try:
            # Phase 1: 準備和啟動 sync
            self.logger.info("執行 Phase 1: 準備工作和啟動同步")
            
            with ThreadPoolExecutor(max_workers=config_manager.parallel_config['max_workers']) as executor:
                futures = {executor.submit(self.process_db_phase1, db_info): db_info for db_info in db_infos}
                
                phase1_results = []
                for future in as_completed(futures):
                    try:
                        result = future.result(timeout=300)
                        phase1_results.append(result)
                    except Exception as e:
                        db_info = futures[future]
                        self.logger.error(f"{db_info.db_info}: Phase 1 異常 - {e}")
                        db_info.status = DBStatus.FAILED
                        db_info.error_message = str(e)
                        db_info.end_time = datetime.now()
                        phase1_results.append(db_info)
            
            # 🔥 修復：使用自定義的等待和處理邏輯，避免重複調用 Phase 2
            if not self.dry_run:
                self.logger.info("等待所有 repo sync 完成並即時處理...")
                self._wait_and_process_syncs_enhanced(phase1_results)
            else:
                # 測試模式直接處理
                for db_info in phase1_results:
                    self.report.add_db(db_info)
            
            self.logger.info("所有 DB 處理完成")
            
        except Exception as e:
            self.logger.error(f"處理過程發生錯誤: {e}")

    def process_selected_dbs(self, db_list: List[str], db_versions: Dict[str, str] = None):
        """處理選定的 DB"""
        if config_manager.repo_config['async_sync'] and not self.dry_run:
            self.process_dbs_async(db_list, db_versions)
        else:
            # 同步處理邏輯
            for db_name in db_list:
                if '#' in db_name:
                    db_name, version = db_name.split('#', 1)
                else:
                    version = db_versions.get(db_name) if db_versions else None
                
                for db_info in self.get_all_dbs('all'):
                    if db_info.db_info == db_name:
                        db_info.version = version
                        
                        db_info = self.process_db_phase1(db_info)
                        
                        if not self.dry_run and db_info.sync_process:
                            self.logger.info(f"等待 {db_name} sync 完成...")
                            db_info.sync_process.wait()
                        
                        db_info = self.process_db_phase2(db_info)
                        self.report.add_db(db_info)
                        break

    def generate_report(self, output_file: str = None):
        """產生改進版報告 - 🔥 修復重複問題"""
        self.report.finalize()
        
        if not output_file:
            output_file = os.path.join(
                self.output_dir, 
                config_manager.path_config['report_filename']
            )
        
        try:
            self.logger.info("開始產生 Excel 報告")
            
            # 🔥 在生成報告前先清理重複
            self._remove_duplicates()
            
            report_data = []
            seen_combinations = set()
            
            for i, db in enumerate(self.report.db_details, 1):
                # 🔥 再次確保不重複
                key = (db.db_info, db.db_type, db.module)
                if key in seen_combinations:
                    self.logger.warning(f"跳過重複記錄: {db.db_info} ({db.db_type}, {db.module})")
                    continue
                
                seen_combinations.add(key)
                
                # 確保 status 有值並轉換為字符串
                status_value = db.status
                if hasattr(status_value, 'value'):
                    status_str = status_value.value
                elif isinstance(status_value, str):
                    status_str = status_value
                else:
                    status_str = str(status_value)
                
                # 手動構建字典，確保所有欄位正確
                db_dict = {
                    'sn': len(report_data) + 1,
                    'module': db.module or '',
                    'db_type': db.db_type or '',
                    'db_info': db.db_info or '',
                    'status': status_str,
                    'version': db.version or '未指定',
                    'manifest_file': db.manifest_file or '',
                    'start_time': db.start_time.strftime('%Y-%m-%d %H:%M:%S') if db.start_time else '',
                    'end_time': db.end_time.strftime('%Y-%m-%d %H:%M:%S') if db.end_time else '',
                    'error_message': db.error_message or '',
                    'sync_log_path': db.sync_log_path or '',
                    'sftp_path': db.sftp_path or '',
                    'local_path': db.local_path or '',
                    'has_existing_repo': '是' if db.has_existing_repo else '否',
                    # 🚫 移除這行：'jira_link': db.jira_link or '未找到',
                    # manifest 比較相關欄位
                    'manifest_比較結果': getattr(db, 'manifest_comparison_summary', '') or '',
                    'manifest_是否相同': '是' if getattr(db, 'manifest_is_identical', None) else '否' if getattr(db, 'manifest_is_identical', None) is False else '',
                    'manifest_差異數量': getattr(db, 'manifest_differences_count', 0) or 0,
                    'exported_manifest_path': getattr(db, 'exported_manifest_path', '') or '',
                }
                
                # 重新命名欄位
                db_dict['完整_JIRA_連結'] = db.jira_link or '未找到'  # ✅ 保留這個
                db_dict['完整_repo_init_指令'] = getattr(db, 'actual_source_cmd', '') or '未記錄'
                
                report_data.append(db_dict)
            
            self.logger.info(f"報告資料：去重複後共 {len(report_data)} 筆記錄")
            
            if not report_data:
                self.logger.warning("沒有資料可以產生報告")
                return
            
            # ... 其餘代碼保持不變
            
        except Exception as e:
            self.logger.error(f"產生報告失敗: {str(e)}")
            import traceback
            traceback.print_exc()

    def _write_enhanced_excel(self, main_df: pd.DataFrame, summary_df: pd.DataFrame, output_file: str):
        """寫入改進版 Excel（自動適寬、表頭藍底白字）"""
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 寫入主要資料
                main_df.to_excel(writer, sheet_name='詳細資訊', index=False)
                summary_df.to_excel(writer, sheet_name='摘要', index=False)
                
                # 格式化工作表
                for sheet_name in ['詳細資訊', '摘要']:
                    worksheet = writer.sheets[sheet_name]
                    self._format_worksheet_enhanced(worksheet)
            
            self.logger.info(f"成功寫入增強版 Excel 檔案: {output_file}")
            
        except Exception as e:
            self.logger.error(f"寫入 Excel 檔案失敗: {str(e)}")
            raise

    def _format_worksheet_enhanced(self, worksheet):
        """格式化工作表（自動適寬、表頭藍底白字）"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # 設定表頭格式（藍底白字）
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 應用到第一行（表頭）
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # 自動調整欄寬
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # 設定最小寬度和最大寬度
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 設定行高
            for row in worksheet.iter_rows():
                worksheet.row_dimensions[row[0].row].height = 20
            
            # 凍結第一行
            worksheet.freeze_panes = 'A2'
            
        except Exception as e:
            self.logger.warning(f"格式化工作表時發生錯誤: {e}")

# =====================================
# ===== 互動式介面 =====
# =====================================

class InteractiveUI:
    """互動式使用者介面"""
    
    def __init__(self):
        self.tool = ManifestPinningTool()
        self.logger = setup_logger(self.__class__.__name__)
        self.selected_dbs = []
        self.db_versions = {}
        self.selected_db_type = 'all'
        
        self._load_default_config()
    
    def _load_default_config(self):
        """載入預設配置"""
        if config_manager.default_execution_config.get('sftp_override'):
            config_manager.apply_overrides(
                config_manager.default_execution_config['sftp_override'],
                source='default_config'
            )
        
        if config_manager.default_execution_config.get('output_dir'):
            self.tool.output_dir = config_manager.default_execution_config['output_dir']
    
    def setup_sftp(self):
        """設定 SFTP 連線資訊"""
        print("\n目前 SFTP 設定:")
        print(f"  Host: {config_manager.sftp_config['host']}")
        print(f"  Port: {config_manager.sftp_config['port']}")
        print(f"  Username: {config_manager.sftp_config['username']}")
        
        if input("\n是否要修改設定? (y/N): ").strip().lower() == 'y':
            host = input(f"Host [{config_manager.sftp_config['host']}]: ").strip()
            if host:
                config_manager.sftp_config['host'] = host
                
            port = input(f"Port [{config_manager.sftp_config['port']}]: ").strip()
            if port:
                config_manager.sftp_config['port'] = int(port)
                
            username = input(f"Username [{config_manager.sftp_config['username']}]: ").strip()
            if username:
                config_manager.sftp_config['username'] = username
                
            password = input("Password (留空保持原值): ").strip()
            if password:
                config_manager.sftp_config['password'] = password
            
            print("✅ SFTP 設定已更新")
    
    def display_current_settings(self):
        """顯示目前設定"""
        print("\n目前設定:")
        print(f"  Mapping table: {'已載入' if self.tool.mapping_reader.df is not None else '未載入'}")
        print(f"  DB 類型: {self.selected_db_type}")
        print(f"  選擇的 DB: {len(self.selected_dbs)} 個")
        print(f"  設定版本的 DB: {len(self.db_versions)} 個")
        print(f"  輸出目錄: {self.tool.output_dir}")
        print(f"  平行處理: {'啟用' if config_manager.parallel_config['enable_parallel'] else '關閉'}")
        print(f"  Worker 數量: {config_manager.parallel_config['max_workers']}")
    
    def load_mapping_table(self):
        """載入 mapping table"""
        default_mapping = config_manager.default_execution_config.get('mapping_table')
        
        if default_mapping and os.path.exists(default_mapping):
            print(f"\n📌 找到預設 mapping table: {default_mapping}")
            if input("是否使用預設檔案? (Y/n): ").strip().lower() != 'n':
                file_path = default_mapping
            else:
                file_path = input(
                    f"請輸入 mapping table 路徑 [{config_manager.path_config['default_mapping_table']}]: "
                ).strip()
                if not file_path:
                    file_path = config_manager.path_config['default_mapping_table']
        else:
            default_path = config_manager.path_config['default_mapping_table']
            file_path = input(f"請輸入 mapping table 路徑 [{default_path}]: ").strip()
            if not file_path:
                file_path = default_path
        
        if not os.path.exists(file_path):
            print(f"❌ 檔案不存在: {file_path}")
            return
        
        if self.tool.load_mapping_table(file_path):
            print(f"✅ 成功載入 mapping table，共 {len(self.tool.mapping_reader.df)} 筆資料")
            all_dbs = self.tool.get_all_dbs()
            unique_db_names = list(set([db.db_info for db in all_dbs]))
            print(f"   找到 {len(unique_db_names)} 個不重複的 DB")
            
            if config_manager.default_execution_config.get('db_type'):
                self.selected_db_type = config_manager.default_execution_config['db_type']
                print(f"   📌 使用預設 DB 類型: {self.selected_db_type}")
        else:
            print("❌ 載入失敗")
    
    def select_db_type(self):
        """選擇 DB 類型"""
        default_type = config_manager.default_execution_config.get('db_type')
        
        if default_type and default_type in ['all', 'master', 'premp', 'mp', 'mpbackup']:
            print(f"\n📌 找到預設 DB 類型: {default_type}")
            if input("是否使用預設值? (Y/n): ").strip().lower() != 'n':
                self.selected_db_type = default_type
                print(f"✅ 已選擇 DB 類型: {self.selected_db_type}")
                return self.selected_db_type
        
        print("\n選擇 DB 類型:")
        print("1. All (所有類型)")
        print("2. Master")
        print("3. PreMP")
        print("4. MP")
        print("5. MP Backup")
        
        choice = input("請選擇: ").strip()
        
        type_map = {
            '1': 'all',
            '2': 'master',
            '3': 'premp',
            '4': 'mp',
            '5': 'mpbackup'
        }
        
        self.selected_db_type = type_map.get(choice, 'all')
        print(f"✅ 已選擇 DB 類型: {self.selected_db_type}")
        return self.selected_db_type
    
    def select_dbs(self):
        """選擇要定版的 DB"""
        if not self.tool.mapping_reader or self.tool.mapping_reader.df is None:
            print("❌ 請先載入 mapping table")
            return []
        
        all_db_infos = self.tool.get_all_dbs(self.selected_db_type)
        if not all_db_infos:
            print(f"❌ 沒有找到 {self.selected_db_type} 類型的 DB")
            return []
        
        # 取得不重複的 DB 列表
        unique_dbs = {}
        for db_info in all_db_infos:
            if db_info.db_info not in unique_dbs:
                unique_dbs[db_info.db_info] = db_info
        
        db_list = list(unique_dbs.keys())
        db_list.sort()
        
        # 檢查預設 DB 列表
        default_dbs = config_manager.default_execution_config.get('selected_dbs')
        
        if default_dbs:
            if default_dbs in ['all', '*']:
                print(f"\n📌 預設配置為選擇所有 {self.selected_db_type} 類型的 DB")
                if input(f"是否選擇全部 {len(db_list)} 個 DB? (Y/n): ").strip().lower() != 'n':
                    self.selected_dbs = db_list
                    print(f"✅ 已選擇所有 {len(db_list)} 個 DB")
                    return db_list
            
            elif isinstance(default_dbs, list) and len(default_dbs) > 0:
                parsed_dbs = []
                for db_spec in default_dbs:
                    if '#' in db_spec:
                        db_name, version = db_spec.split('#', 1)
                        if db_name in db_list:
                            parsed_dbs.append(db_name)
                            self.db_versions[db_name] = version
                    else:
                        if db_spec in db_list:
                            parsed_dbs.append(db_spec)
                
                if parsed_dbs:
                    print(f"\n📌 找到預設 DB 列表: {', '.join(default_dbs)}")
                    print(f"   其中 {len(parsed_dbs)} 個 DB 存在於當前 mapping table")
                    if input("是否使用預設 DB 列表? (Y/n): ").strip().lower() != 'n':
                        self.selected_dbs = parsed_dbs
                        print(f"✅ 已選擇 {len(parsed_dbs)} 個 DB")
                        return parsed_dbs
        
        # 顯示 DB 列表和選擇邏輯
        print(f"\n找到 {len(db_list)} 個不重複的 DB (類型: {self.selected_db_type})")
        
        print("\nDB 列表:")
        for i, db in enumerate(db_list, 1):
            db_info = unique_dbs[db]
            print(f"{i:3d}. {db:10s} - {db_info.module:10s} ({db_info.db_type})")
        
        print("\n選擇方式:")
        print("1. 全選")
        print("2. 輸入編號範圍 (如: 1-5,7,9-12)")
        print("3. 輸入 DB 名稱列表 (逗號分隔)")
        
        choice = input("請選擇: ").strip()
        
        selected = []
        
        if choice == '1':
            selected = db_list
        elif choice == '2':
            indices_input = input("請輸入編號範圍: ").strip()
            try:
                indices = self._parse_number_range(indices_input, len(db_list))
                selected = [db_list[i-1] for i in indices if 1 <= i <= len(db_list)]
            except Exception as e:
                print(f"❌ 無效的編號範圍: {e}")
                return []
        elif choice == '3':
            db_names = input("請輸入 DB 名稱 (如: DB2302,DB2575): ").strip()
            input_dbs = [db.strip() for db in db_names.split(',')]
            for db in input_dbs:
                if db in db_list:
                    selected.append(db)
                else:
                    print(f"⚠️ DB {db} 不存在，已跳過")
        
        self.selected_dbs = selected
        print(f"✅ 已選擇 {len(selected)} 個 DB")
        return selected
    
    def _parse_number_range(self, range_str: str, max_num: int) -> List[int]:
        """解析數字範圍字串"""
        result = []
        parts = range_str.split(',')
        
        for part in parts:
            part = part.strip()
            if '-' in part:
                start, end = part.split('-', 1)
                start = int(start.strip())
                end = int(end.strip())
                
                if start > end:
                    start, end = end, start
                
                for i in range(start, min(end + 1, max_num + 1)):
                    if i > 0:
                        result.append(i)
            else:
                num = int(part)
                if 1 <= num <= max_num:
                    result.append(num)
        
        return sorted(list(set(result)))
    
    def setup_db_versions(self):
        """設定 DB 版本"""
        if not self.selected_dbs:
            print("❌ 請先選擇 DB")
            return
        
        default_versions = config_manager.default_execution_config.get('db_versions', {})
        
        if default_versions:
            applicable_versions = {
                db: ver for db, ver in default_versions.items() 
                if db in self.selected_dbs
            }
            
            if applicable_versions:
                print(f"\n📌 找到預設版本設定:")
                for db, ver in applicable_versions.items():
                    print(f"   {db}: 版本 {ver}")
                
                if input("是否使用預設版本設定? (Y/n): ").strip().lower() != 'n':
                    self.db_versions.update(applicable_versions)
                    print(f"✅ 已套用 {len(applicable_versions)} 個 DB 的預設版本")
                    
                    unset_dbs = [db for db in self.selected_dbs if db not in self.db_versions]
                    if unset_dbs:
                        print(f"\n還有 {len(unset_dbs)} 個 DB 未設定版本:")
                        for db in unset_dbs:
                            version = input(f"{db} 的版本 [最新]: ").strip()
                            if version:
                                self.db_versions[db] = version
                    return
        
        print("\n設定 DB 版本 (留空使用最新版本)")
        for db in self.selected_dbs:
            if db in self.db_versions:
                print(f"{db}: 已設定版本 {self.db_versions[db]}")
                continue
            
            version = input(f"{db} 的版本 [最新]: ").strip()
            if version:
                self.db_versions[db] = version
        
        if self.db_versions:
            print(f"✅ 已設定 {len(self.db_versions)} 個 DB 的版本")
    
    def execute_pinning(self):
        """執行定版"""
        if not self.selected_dbs:
            print("❌ 請先選擇要定版的 DB")
            return
        
        if not self.tool.mapping_reader or self.tool.mapping_reader.df is None:
            print("❌ 請先載入 mapping table")
            return
        
        print("\n" + "="*60)
        print("準備執行定版")
        print("="*60)
        print(f"📌 DB 數量: {len(self.selected_dbs)}")
        print(f"📂 輸出目錄: {self.tool.output_dir}")
        
        # 詢問輸出目錄
        default_output = config_manager.default_execution_config.get('output_dir')
        if default_output:
            print(f"📌 找到預設輸出目錄: {default_output}")
            if input("是否使用預設輸出目錄? (Y/n): ").strip().lower() != 'n':
                self.tool.output_dir = default_output
            else:
                output_dir = input(f"輸出目錄 [{self.tool.output_dir}]: ").strip()
                if output_dir:
                    self.tool.output_dir = output_dir
        else:
            output_dir = input(f"輸出目錄 [{self.tool.output_dir}]: ").strip()
            if output_dir:
                self.tool.output_dir = output_dir
        
        # 確認執行
        if config_manager.default_execution_config.get('auto_confirm'):
            print("📌 自動確認執行（根據預設配置）")
        else:
            if input("\n確認開始執行? (Y/n): ").strip().lower() == 'n':
                print("❌ 使用者取消操作")
                return
        
        print("\n🚀 開始執行定版...")
        
        print("🌐 準備 SFTP 連線（每個 DB 使用獨立連線）...")
        
        try:
            # 執行定版
            self.tool.process_selected_dbs(self.selected_dbs, self.db_versions)
            
            # 產生報告
            print("\n📊 產生報告...")
            report_path = os.path.join(
                self.tool.output_dir, 
                config_manager.path_config['report_filename']
            )
            self.tool.generate_report(report_path)
            
            print("\n✨ 定版完成！")
            
        finally:
            resource_manager.cleanup_all()
    
    def quick_execute_with_defaults(self):
        """使用預設配置快速執行"""
        print("\n" + "="*60)
        print("快速執行模式 - 使用預設配置")
        print("="*60)
        
        # 自動載入 mapping table
        if config_manager.default_execution_config.get('mapping_table'):
            file_path = config_manager.default_execution_config['mapping_table']
            if os.path.exists(file_path):
                print(f"📌 載入預設 mapping table: {file_path}")
                if self.tool.load_mapping_table(file_path):
                    print(f"✅ 成功載入")
                else:
                    print("❌ 載入失敗")
                    return
            else:
                print(f"❌ 預設 mapping table 不存在: {file_path}")
                return
        else:
            print("❌ 未設定預設 mapping table")
            return
        
        # 設定 DB 類型
        if config_manager.default_execution_config.get('db_type'):
            self.selected_db_type = config_manager.default_execution_config['db_type']
            print(f"📌 使用預設 DB 類型: {self.selected_db_type}")
        
        # 選擇 DB
        default_dbs = config_manager.default_execution_config.get('selected_dbs')
        all_db_infos = self.tool.get_all_dbs(self.selected_db_type)
        
        if default_dbs:
            if default_dbs in ['all', '*']:
                unique_dbs = list(set([db.db_info for db in all_db_infos]))
                self.selected_dbs = unique_dbs
                print(f"📌 選擇所有 {self.selected_db_type} 類型的 DB: {len(unique_dbs)} 個")
            
            elif isinstance(default_dbs, list) and len(default_dbs) > 0:
                parsed_dbs = []
                for db_spec in default_dbs:
                    if '#' in db_spec:
                        db_name, version = db_spec.split('#', 1)
                        parsed_dbs.append(db_name)
                        self.db_versions[db_name] = version
                    else:
                        parsed_dbs.append(db_spec)
                self.selected_dbs = parsed_dbs
                print(f"📌 使用預設 DB 列表: {len(parsed_dbs)} 個")
        else:
            print("⚠️ 預設配置未指定 DB 列表，無法自動執行")
            return
        
        # 設定版本
        if config_manager.default_execution_config.get('db_versions'):
            self.db_versions.update(config_manager.default_execution_config['db_versions'])
            print(f"📌 套用預設版本設定: {len(self.db_versions)} 個")
        
        # 設定輸出目錄
        if config_manager.default_execution_config.get('output_dir'):
            self.tool.output_dir = config_manager.default_execution_config['output_dir']
            print(f"📌 使用預設輸出目錄: {self.tool.output_dir}")
        
        # 顯示摘要
        print("\n執行摘要:")
        print(f"  DB 類型: {self.selected_db_type}")
        print(f"  DB 數量: {len(self.selected_dbs)}")
        print(f"  設定版本: {len(self.db_versions)} 個")
        print(f"  輸出目錄: {self.tool.output_dir}")
        
        # 確認執行
        if not config_manager.default_execution_config.get('auto_confirm'):
            if input("\n確認執行? (Y/n): ").strip().lower() == 'n':
                print("❌ 使用者取消")
                return
        
        # 執行定版
        self.execute_pinning()
    
    def test_jira_connection(self):
        """測試 JIRA 連線"""
        print("\n測試 JIRA 連線...")
        
        if not hasattr(self.tool, 'source_cmd_manager'):
            self.tool.source_cmd_manager = SourceCommandManager()
        
        if self.tool.source_cmd_manager.jira_client.connect():
            print("✅ JIRA 連線成功！")
            
            if input("\n是否要測試查詢 DB 的 source command? (y/N): ").strip().lower() == 'y':
                db_name = input("請輸入 DB 名稱 (例如: DB2302): ").strip()
                if db_name:
                    print(f"\n查詢 {db_name} 的 source command...")
                    
                    test_db_info = DBInfo(
                        sn=0,
                        module="Test",
                        db_type="master",
                        db_info=db_name,
                        db_folder="",
                        sftp_path=""
                    )
                    
                    cmd = self.tool.source_cmd_manager.get_source_command(
                        test_db_info,
                        self.tool.mapping_reader.df if self.tool.mapping_reader else None
                    )
                    
                    if cmd:
                        print(f"\n找到的 source command:")
                        print(f"  {cmd}")
                    else:
                        print(f"\n未找到 {db_name} 的 source command")
        else:
            print("❌ JIRA 連線失敗！")
    
    def test_sftp_connection(self):
        """測試 SFTP 連線"""
        print("\n測試 SFTP 連線...")
        print(f"伺服器: {config_manager.sftp_config['host']}:{config_manager.sftp_config['port']}")
        print(f"用戶: {config_manager.sftp_config['username']}")
        
        # 創建臨時 SFTP 管理器進行測試
        test_sftp_manager = SFTPManager()
        
        if test_sftp_manager.connect():
            print("✅ SFTP 連線成功！")
            
            if input("\n是否要測試路徑存取? (y/N): ").strip().lower() == 'y':
                path = input("請輸入要測試的路徑: ").strip()
                if path:
                    try:
                        result = test_sftp_manager.find_latest_manifest(path)
                        if result:
                            print(f"✅ 找到 manifest: {result[1]}")
                        else:
                            print("❌ 沒有找到 manifest")
                    except Exception as e:
                        print(f"❌ 測試失敗: {e}")
            
            test_sftp_manager.disconnect()
        else:
            print("❌ SFTP 連線失敗！")

    # 🔥 新增：設定清理選項
    def setup_clean_options(self):
        """設定清理相關選項 - 🔥 添加更多實用選項"""
        print("\n" + "="*50)
        print("🧹 清理選項設定")
        print("="*50)
        
        print(f"目前設定:")
        print(f"  檢查乾淨狀態: {'是' if config_manager.repo_config['check_clean_before_sync'] else '否'}")
        print(f"  自動清理工作空間: {'是' if config_manager.repo_config['auto_clean_workspace'] else '否'}")
        print(f"  備份本地修改: {'是' if config_manager.repo_config['backup_local_changes'] else '否'}")
        print(f"  🔥 檢查超時時跳過: {'是' if config_manager.repo_config['skip_clean_on_timeout'] else '否'}")
        print(f"  🔥 只做快速清理: {'是' if config_manager.repo_config['quick_clean_only'] else '否'}")
        print(f"  🔥 髒repo時詢問用戶: {'是' if config_manager.repo_config['ask_user_on_dirty'] else '否'}")
        
        if input("\n是否要修改設定? (y/N): ").strip().lower() == 'y':
            
            # 檢查乾淨狀態
            check_clean = input(f"檢查乾淨狀態 [{'是' if config_manager.repo_config['check_clean_before_sync'] else '否'}] (y/n/enter跳過): ").strip().lower()
            if check_clean == 'y':
                config_manager.repo_config['check_clean_before_sync'] = True
            elif check_clean == 'n':
                config_manager.repo_config['check_clean_before_sync'] = False
            
            # 🔥 新增：檢查超時時跳過
            skip_timeout = input(f"檢查超時時跳過清理 [{'是' if config_manager.repo_config['skip_clean_on_timeout'] else '否'}] (y/n/enter跳過): ").strip().lower()
            if skip_timeout == 'y':
                config_manager.repo_config['skip_clean_on_timeout'] = True
            elif skip_timeout == 'n':
                config_manager.repo_config['skip_clean_on_timeout'] = False
            
            # 🔥 新增：髒repo時詢問用戶
            ask_user = input(f"發現髒repo時詢問用戶 [{'是' if config_manager.repo_config['ask_user_on_dirty'] else '否'}] (y/n/enter跳過): ").strip().lower()
            if ask_user == 'y':
                config_manager.repo_config['ask_user_on_dirty'] = True
            elif ask_user == 'n':
                config_manager.repo_config['ask_user_on_dirty'] = False
            
            # 自動清理
            auto_clean = input(f"自動清理工作空間 [{'是' if config_manager.repo_config['auto_clean_workspace'] else '否'}] (y/n/enter跳過): ").strip().lower()
            if auto_clean == 'y':
                config_manager.repo_config['auto_clean_workspace'] = True
            elif auto_clean == 'n':
                config_manager.repo_config['auto_clean_workspace'] = False
            
            print("✅ 清理選項已更新")
            print("\n💡 建議配置（針對大型repo）:")
            print("  - 檢查超時時跳過: 是")
            print("  - 髒repo時詢問用戶: 是") 
            print("  - 自動清理工作空間: 否")
    
    # 🔥 新增：手動清理工作空間
    def manual_clean_workspace(self):
        """手動清理工作空間"""
        print("\n" + "="*50)
        print("🧹 手動清理工作空間")
        print("="*50)
        
        if not self.tool.mapping_reader or self.tool.mapping_reader.df is None:
            print("❌ 請先載入 mapping table")
            return
        
        # 選擇要清理的DB
        all_db_infos = self.tool.get_all_dbs(self.selected_db_type)
        unique_dbs = list(set([db.db_info for db in all_db_infos]))
        unique_dbs.sort()
        
        print(f"\n找到 {len(unique_dbs)} 個 DB:")
        for i, db in enumerate(unique_dbs, 1):
            local_path = os.path.join(self.tool.output_dir, 'unknown', db)  # 簡化路徑檢查
            status = "🟢" if not self.tool.repo_manager.check_repo_exists(local_path) else "🔶"
            print(f"{i:3d}. {status} {db}")
        
        print("\n🟢 = 無repo目錄  🔶 = 有repo目錄")
        
        choice = input("\n選擇要清理的DB (輸入編號，多個用逗號分隔，或輸入'all'): ").strip()
        
        if choice.lower() == 'all':
            selected_dbs = unique_dbs
        else:
            try:
                indices = [int(x.strip()) for x in choice.split(',')]
                selected_dbs = [unique_dbs[i-1] for i in indices if 1 <= i <= len(unique_dbs)]
            except:
                print("❌ 無效的輸入")
                return
        
        if not selected_dbs:
            print("❌ 沒有選擇任何DB")
            return
        
        print(f"\n準備清理 {len(selected_dbs)} 個DB:")
        for db in selected_dbs:
            print(f"  - {db}")
        
        if input("\n確認執行清理? (y/N): ").strip().lower() != 'y':
            print("❌ 用戶取消")
            return
        
        # 執行清理
        success_count = 0
        for db in selected_dbs:
            try:
                # 構建可能的路徑
                possible_paths = []
                for db_info in all_db_infos:
                    if db_info.db_info == db:
                        local_path = os.path.join(self.tool.output_dir, db_info.module, db_info.db_info)
                        if local_path not in possible_paths:
                            possible_paths.append(local_path)
                
                cleaned_any = False
                for local_path in possible_paths:
                    if self.tool.repo_manager.check_repo_exists(local_path):
                        print(f"\n🧹 清理 {db} ({local_path})")
                        
                        # 檢查狀態
                        clean_status = self.tool.repo_manager.check_repo_clean_status(local_path)
                        print(f"   狀態: {clean_status['details']}")
                        
                        # 備份並清理
                        if not clean_status['is_clean']:
                            if config_manager.repo_config['backup_local_changes']:
                                self.tool.repo_manager.backup_local_changes(local_path, db)
                            
                            if self.tool.repo_manager.clean_repo_workspace(local_path, force=True):
                                print(f"   ✅ 清理成功")
                                cleaned_any = True
                            else:
                                print(f"   ❌ 清理失敗")
                        else:
                            print(f"   ✅ 已經是乾淨狀態")
                            cleaned_any = True
                
                if cleaned_any:
                    success_count += 1
                    
            except Exception as e:
                print(f"   ❌ 清理 {db} 失敗: {e}")
        
        print(f"\n🎉 清理完成: {success_count}/{len(selected_dbs)} 個DB清理成功")

    def display_menu(self) -> str:
        """顯示主選單"""
        print("\n" + "="*60)
        print("Manifest 定版工具 - 主選單 (改進版 + Manifest 比較)")
        
        has_defaults = any([
            config_manager.default_execution_config.get('mapping_table'),
            config_manager.default_execution_config.get('db_type'),
            config_manager.default_execution_config.get('selected_dbs'),
            config_manager.default_execution_config.get('db_versions'),
            config_manager.default_execution_config.get('output_dir')
        ])
        
        if has_defaults:
            print("(📌 已載入預設配置)")
        
        print("="*60)
        print("1. 載入 mapping table")
        print("2. 設定 SFTP 連線資訊")
        print("3. 選擇 DB 類型 (master/premp/mp/mpbackup/all)")
        print("4. 選擇要定版的 DB")
        print("5. 設定 DB 版本")
        print("6. 開始執行定版")
        print("7. 顯示目前設定")
        print("8. 快速執行（使用所有預設值）")
        print("9. 測試 JIRA 連線")
        print("10. 測試 SFTP 連線")
        print("11. 🔥 設定清理選項")  # 新增
        print("12. 🔥 手動清理工作空間")  # 新增
        print("0. 結束程式")
        print("="*60)
        
        return input("請選擇功能: ").strip()
    
    def run_interactive(self):
        """執行互動式介面"""
        print("\n歡迎使用 Manifest 定版工具！")
        print(f"版本: {__version__} (改進版 + Manifest 比較功能)")
        print("改進內容: 修復 SFTP Garbage packet 問題、改善日誌輸出、即時 manifest 比較、🔥 新增工作空間清理功能")
        
        # 檢查是否有預設配置
        has_complete_defaults = all([
            config_manager.default_execution_config.get('mapping_table'),
            os.path.exists(config_manager.default_execution_config.get('mapping_table', ''))
        ])
        
        if has_complete_defaults:
            print("\n📌 偵測到完整的預設配置")
            if input("是否要使用預設配置快速執行? (y/N): ").strip().lower() == 'y':
                self.quick_execute_with_defaults()
                return
        
        while True:
            try:
                choice = self.display_menu()
                
                if choice == '0':
                    print("\n👋 再見！")
                    break
                elif choice == '1':
                    self.load_mapping_table()
                elif choice == '2':
                    self.setup_sftp()
                elif choice == '3':
                    self.select_db_type()
                elif choice == '4':
                    self.select_dbs()
                elif choice == '5':
                    self.setup_db_versions()
                elif choice == '6':
                    self.execute_pinning()
                elif choice == '7':
                    self.display_current_settings()
                elif choice == '8':
                    self.quick_execute_with_defaults()
                elif choice == '9':
                    self.test_jira_connection()
                elif choice == '10':
                    self.test_sftp_connection()
                elif choice == '11':  # 🔥 新增
                    self.setup_clean_options()
                elif choice == '12':  # 🔥 新增
                    self.manual_clean_workspace()
                else:
                    print("❌ 無效的選擇，請重新輸入")
                    
            except KeyboardInterrupt:
                print("\n\n⚠️ 使用者中斷")
                if input("確定要結束程式嗎? (Y/n): ").strip().lower() != 'n':
                    break
            except Exception as e:
                self.logger.error(f"發生錯誤: {str(e)}")
                print(f"❌ 發生錯誤: {str(e)}")
                if input("是否繼續? (Y/n): ").strip().lower() == 'n':
                    break
        
        resource_manager.cleanup_all()

# =====================================
# ===== 主程式 =====
# =====================================

def main():
    """主程式入口"""
    parser = argparse.ArgumentParser(
        description='Manifest 定版工具 - 自動化 repo 定版處理 (改進版 + Manifest 比較功能)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
版本: {__version__}
作者: {__author__}
日期: {__date__}

改進內容:
- 修復 SFTP "Garbage packet received" 錯誤
- 改為正常順序日誌輸出，便於 debug
- 改進 Excel 報告格式（自動適寬、表頭藍底白字）
- 記錄完整 repo init 指令和 JIRA 連結
- ✨ 新增 Manifest 比較功能：自動比較原始 manifest 和導出的 vp_manifest.xml
- 🔥 即時報告更新：每個 DB 完成後立即更新 Excel 報告
- 修復重複日誌輸出問題

範例:
  # 使用互動式介面
  python {sys.argv[0]}
  
  # 處理指定的 DB
  python {sys.argv[0]} -m mapping.xlsx -d DB2302,DB2575 -o ./output
  
  # 測試模式
  python {sys.argv[0]} -m mapping.xlsx --dry-run
        """
    )
    
    # 基本參數
    parser.add_argument('-m', '--mapping', 
                        type=str,
                        help='Mapping table Excel 檔案路徑')
    
    parser.add_argument('-o', '--output', 
                        type=str,
                        help=f'輸出目錄 (預設: {config_manager.path_config["default_output_dir"]})')
    
    parser.add_argument('-t', '--type',
                        choices=['all', 'master', 'premp', 'mp', 'mpbackup'],
                        default='all',
                        help='要處理的 DB 類型 (預設: all)')
    
    parser.add_argument('-d', '--dbs', 
                        type=str,
                        help='要處理的 DB 列表，逗號分隔 (可包含版本，如: DB2302#3,DB2575)')
    
    parser.add_argument('-v', '--versions', 
                        type=str,
                        help='DB 版本設定，格式: DB2302#3,DB2575#186')
    
    parser.add_argument('--dry-run', 
                        action='store_true',
                        help='測試模式，只顯示將要執行的動作，不實際執行')
    
    parser.add_argument('--debug', 
                        action='store_true',
                        help='啟用 debug 模式，顯示詳細日誌')
    
    parser.add_argument('--version', 
                        action='version',
                        version=f'%(prog)s {__version__}')
    
    # 解析參數
    args = parser.parse_args()
    
    # 設定日誌等級
    if args.debug:
        config_manager.log_config['level'] = logging.DEBUG
        logging.getLogger().setLevel(logging.DEBUG)
        for handler in logging.getLogger().handlers:
            handler.setLevel(logging.DEBUG)
        print("🔍 Debug 模式已啟用")
    
    # 檢查是否為測試模式
    if args.dry_run:
        print("\n" + "="*60)
        print("🧪 測試模式 (Dry Run) - 不會實際執行任何操作")
        print("="*60)
    
    # 決定執行模式
    if args.mapping:
        # 命令列模式
        print("\n" + "="*60)
        print(f"📋 Manifest 定版工具 v{__version__} - 命令列模式 (改進版 + Manifest 比較)")
        print("="*60)
        
        tool = ManifestPinningTool()
        
        if args.dry_run:
            tool.dry_run = True
        
        try:
            # 載入 mapping table
            print(f"\n📂 載入 mapping table: {args.mapping}")
            if not os.path.exists(args.mapping):
                print(f"❌ 檔案不存在: {args.mapping}")
                sys.exit(1)
            
            if not tool.load_mapping_table(args.mapping):
                print("❌ 無法載入 mapping table")
                sys.exit(1)
            
            print(f"✅ 成功載入 mapping table")
            
            # 設定輸出目錄
            tool.output_dir = args.output or config_manager.path_config['default_output_dir']
            os.makedirs(tool.output_dir, exist_ok=True)
            print(f"📂 輸出目錄: {tool.output_dir}")

            print(f"\n🌐 準備 SFTP 連線: {config_manager.sftp_config['host']}")
            print("ℹ️  每個 DB 將使用獨立的 SFTP 連線")

            try:
                # 決定要處理的 DB 列表
                db_list = []
                db_versions = {}
                
                if args.dbs:
                    db_specs = [db.strip() for db in args.dbs.split(',')]
                    
                    for db_spec in db_specs:
                        if '#' in db_spec:
                            db_name, version = db_spec.split('#', 1)
                            db_list.append(db_name)
                            db_versions[db_name] = version
                        else:
                            db_list.append(db_spec)
                    
                    print(f"\n📌 使用指定的 DB 列表: {', '.join(db_list)}")
                else:
                    all_db_infos = tool.get_all_dbs(args.type)
                    db_list = list(set([db.db_info for db in all_db_infos]))
                    
                    if args.type == 'all':
                        print(f"\n📌 使用所有 DB，共 {len(db_list)} 個")
                    else:
                        print(f"\n📌 使用所有 {args.type} 類型的 DB，共 {len(db_list)} 個")
                
                # 處理額外的版本設定
                if args.versions:
                    version_specs = [v.strip() for v in args.versions.split(',')]
                    for version_spec in version_specs:
                        if '#' in version_spec:
                            db_name, version = version_spec.split('#', 1)
                            db_versions[db_name] = version
                    
                    print(f"📌 設定了 {len(db_versions)} 個 DB 的版本")
                
                # 確認處理資訊
                print("\n" + "-"*40)
                print("📋 準備處理以下 DB:")
                for i, db in enumerate(db_list, 1):
                    version_info = f" (版本: {db_versions[db]})" if db in db_versions else " (最新版本)"
                    print(f"  {i:3d}. {db}{version_info}")
                print("-"*40)
                
                if not db_list:
                    print("❌ 沒有找到要處理的 DB")
                    sys.exit(1)
                
                # 詢問確認（除非是測試模式）
                if not args.dry_run:
                    if sys.stdin.isatty():
                        confirm = input(f"\n確認要處理 {len(db_list)} 個 DB? (Y/n): ").strip().lower()
                        if confirm == 'n':
                            print("❌ 使用者取消操作")
                            sys.exit(0)
                
                # 開始處理
                print("\n" + "="*60)
                if args.dry_run:
                    print("🧪 開始測試執行（不會實際執行操作）")
                else:
                    print("🚀 開始執行定版處理")
                print("="*60)
                
                start_time = datetime.now()
                
                # 執行處理
                tool.process_selected_dbs(db_list, db_versions)
                
                end_time = datetime.now()
                elapsed_time = end_time - start_time
                
                # 產生報告
                if not args.dry_run:
                    print("\n📊 產生處理報告...")
                    report_path = os.path.join(
                        tool.output_dir, 
                        config_manager.path_config['report_filename']
                    )
                    tool.generate_report(report_path)
                
                # 顯示結果摘要
                print("\n" + "="*60)
                print("✨ 處理完成！")
                print("="*60)
                print(f"📊 總 DB 數: {tool.report.total_dbs}")
                print(f"✅ 成功: {tool.report.successful_dbs}")
                print(f"❌ 失敗: {tool.report.failed_dbs}")
                print(f"⭐ 跳過: {tool.report.skipped_dbs}")
                print(f"🔍 有差異: {tool.report.dbs_with_differences}")
                print(f"⏱️ 總耗時: {elapsed_time}")
                print(f"📂 輸出目錄: {tool.output_dir}")
                if not args.dry_run:
                    print(f"📊 報告檔案: {report_path}")
                print("="*60)
                
                # 如果有失敗的項目，顯示詳細資訊
                if tool.report.failed_dbs > 0:
                    print("\n❌ 失敗的 DB:")
                    for db in tool.report.db_details:
                        if db.status == DBStatus.FAILED:
                            print(f"  - {db.module}/{db.db_info}: {db.error_message}")
                
                # 如果有差異的項目，顯示信息
                if tool.report.dbs_with_differences > 0:
                    print(f"\n🔍 有 {tool.report.dbs_with_differences} 個 DB 的 manifest 有差異:")
                    for db in tool.report.db_details:
                        if db.status == DBStatus.SUCCESS_WITH_DIFF:
                            print(f"  - {db.module}/{db.db_info}: {db.manifest_comparison_summary}")
                
            finally:
                print("\n📌 清理資源...")
                resource_manager.cleanup_all()
                
        except KeyboardInterrupt:
            print("\n🛑 收到 Ctrl+C，清理所有進程...")
            resource_manager.cleanup_all()
            sys.exit(0)
            
        except Exception as e:
            print(f"\n❌ 發生錯誤: {str(e)}")
            if args.debug:
                import traceback
                traceback.print_exc()
            sys.exit(1)
        
        finally:
            resource_manager.cleanup_all()
    
    else:
        # 互動式模式
        print("\n" + "="*60)
        print(f"🎮 Manifest 定版工具 v{__version__} - 互動式介面 (改進版 + Manifest 比較)")
        print("="*60)
        print("改進內容: 修復 SFTP Garbage packet 問題、改善日誌輸出、即時 manifest 比較")
        print("提示: 使用 -h 參數查看命令列選項")
        print("="*60)
        
        try:
            ui = InteractiveUI()
            
            if args.output:
                ui.tool.output_dir = args.output
            
            if args.dry_run:
                ui.tool.dry_run = True
                print("🧪 測試模式已啟用")
            
            ui.run_interactive()
            
        except KeyboardInterrupt:
            print("\n\n👋 再見！")
            sys.exit(0)
            
        except Exception as e:
            print(f"\n❌ 發生錯誤: {str(e)}")
            if args.debug:
                import traceback
                traceback.print_exc()
            sys.exit(1)
        
        finally:
            resource_manager.cleanup_all()

if __name__ == "__main__":
    main()