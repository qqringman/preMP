#!/usr/bin/env python3
"""
測試 Master to PreMP Manifest 轉換規則
比對轉換結果與正確版 PreMP，輸出差異報告
"""

import os
import sys
import xml.etree.ElementTree as ET
import pandas as pd
from typing import Dict, List, Tuple, Optional
import argparse
from datetime import datetime
import logging

# 添加專案路徑
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

import config
from overwrite_lib.feature_three import FeatureThree
from excel_handler import ExcelHandler
import utils

# 設定日誌
logger = utils.setup_logger(__name__)

class ManifestConversionTester:
    """Manifest 轉換規則測試器"""
    
    def __init__(self):
        self.feature_three = FeatureThree()
        self.excel_handler = ExcelHandler()
        self.logger = logger
        
        # 統計資料
        self.stats = {
            'total_projects': 0,
            'matched': 0,
            'mismatched': 0,
            'not_found_in_premp': 0,
            'extra_in_premp': 0
        }
        
    def parse_manifest(self, file_path: str) -> Dict[str, Dict]:
        """
        解析 manifest.xml 檔案
        
        Args:
            file_path: manifest.xml 檔案路徑
            
        Returns:
            字典，key 是專案名稱，value 是專案屬性
        """
        try:
            self.logger.info(f"解析 manifest 檔案: {file_path}")
            
            # 檢查檔案是否存在
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"檔案不存在: {file_path}")
            
            # 解析 XML
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # 提取所有專案
            projects = {}
            for project in root.findall('project'):
                name = project.get('name', '')
                if not name:
                    continue
                    
                projects[name] = {
                    'name': name,
                    'path': project.get('path', ''),
                    'revision': project.get('revision', ''),
                    'upstream': project.get('upstream', ''),
                    'dest-branch': project.get('dest-branch', ''),
                    'groups': project.get('groups', ''),
                    'clone-depth': project.get('clone-depth', ''),
                    'remote': project.get('remote', ''),
                }
            
            self.logger.info(f"成功解析 {len(projects)} 個專案")
            return projects
            
        except Exception as e:
            self.logger.error(f"解析 manifest 檔案失敗: {str(e)}")
            raise
    
    def convert_revision(self, revision: str) -> str:
        """
        使用 feature_three 的轉換邏輯轉換 revision
        
        Args:
            revision: 原始 revision
            
        Returns:
            轉換後的 revision
        """
        try:
            return self.feature_three._convert_master_to_premp(revision)
        except Exception as e:
            self.logger.error(f"轉換 revision 失敗: {revision}, 錯誤: {str(e)}")
            return revision
    
    def compare_manifests(self, master_projects: Dict, premp_projects: Dict) -> List[Dict]:
        """
        比對 master 轉換後與 premp 的差異
        
        Args:
            master_projects: master manifest 的專案
            premp_projects: premp manifest 的專案
            
        Returns:
            差異列表
        """
        differences = []
        
        # 統計
        self.stats['total_projects'] = len(master_projects)
        
        # 比對 master 中的每個專案
        for name, master_proj in master_projects.items():
            # 轉換 master revision
            master_revision = master_proj['revision']
            converted_revision = self.convert_revision(master_revision)
            
            # 在 premp 中查找對應專案
            if name in premp_projects:
                premp_proj = premp_projects[name]
                premp_revision = premp_proj['revision']
                
                # 比對轉換後的 revision 與 premp revision
                if converted_revision == premp_revision:
                    self.stats['matched'] += 1
                    status = '✅ 匹配'
                else:
                    self.stats['mismatched'] += 1
                    status = '❌ 不匹配'
                    
                    # 記錄差異
                    differences.append({
                        'SN': len(differences) + 1,
                        '專案名稱': name,
                        '專案路徑': master_proj['path'],
                        'Master Revision': master_revision,
                        '轉換後 Revision': converted_revision,
                        'PreMP Revision (正確版)': premp_revision,
                        '狀態': status,
                        '轉換是否正確': '否',
                        '差異說明': f'期望: {premp_revision}, 實際: {converted_revision}',
                        'Upstream': master_proj.get('upstream', ''),
                        'Dest-Branch': master_proj.get('dest-branch', ''),
                        'Groups': master_proj.get('groups', ''),
                        'Remote': master_proj.get('remote', '')
                    })
            else:
                # 在 premp 中找不到對應專案
                self.stats['not_found_in_premp'] += 1
                status = '⚠️ PreMP中不存在'
                
                differences.append({
                    'SN': len(differences) + 1,
                    '專案名稱': name,
                    '專案路徑': master_proj['path'],
                    'Master Revision': master_revision,
                    '轉換後 Revision': converted_revision,
                    'PreMP Revision (正確版)': 'N/A (專案不存在)',
                    '狀態': status,
                    '轉換是否正確': 'N/A',
                    '差異說明': '專案在 PreMP manifest 中不存在',
                    'Upstream': master_proj.get('upstream', ''),
                    'Dest-Branch': master_proj.get('dest-branch', ''),
                    'Groups': master_proj.get('groups', ''),
                    'Remote': master_proj.get('remote', '')
                })
        
        # 找出 premp 中有但 master 中沒有的專案
        for name in premp_projects:
            if name not in master_projects:
                self.stats['extra_in_premp'] += 1
                differences.append({
                    'SN': len(differences) + 1,
                    '專案名稱': name,
                    '專案路徑': premp_projects[name]['path'],
                    'Master Revision': 'N/A (專案不存在)',
                    '轉換後 Revision': 'N/A',
                    'PreMP Revision (正確版)': premp_projects[name]['revision'],
                    '狀態': '🔶 僅存在於PreMP',
                    '轉換是否正確': 'N/A',
                    '差異說明': '專案僅存在於 PreMP manifest 中',
                    'Upstream': premp_projects[name].get('upstream', ''),
                    'Dest-Branch': premp_projects[name].get('dest-branch', ''),
                    'Groups': premp_projects[name].get('groups', ''),
                    'Remote': premp_projects[name].get('remote', '')
                })
        
        return differences
    
    def generate_excel_report(self, differences: List[Dict], output_file: str, 
                            master_file: str, premp_file: str) -> bool:
        """
        生成 Excel 測試報告
        
        Args:
            differences: 差異列表
            output_file: 輸出檔案路徑
            master_file: master manifest 檔案路徑
            premp_file: premp manifest 檔案路徑
            
        Returns:
            是否成功生成報告
        """
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 頁籤 1: 測試摘要
                summary_data = [{
                    '測試時間': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'Master Manifest': os.path.basename(master_file),
                    'PreMP Manifest (正確版)': os.path.basename(premp_file),
                    '總專案數': self.stats['total_projects'],
                    '✅ 匹配數': self.stats['matched'],
                    '❌ 不匹配數': self.stats['mismatched'],
                    '⚠️ PreMP中不存在': self.stats['not_found_in_premp'],
                    '🔶 僅存在於PreMP': self.stats['extra_in_premp'],
                    '匹配率': f"{(self.stats['matched'] / self.stats['total_projects'] * 100):.2f}%" if self.stats['total_projects'] > 0 else '0%'
                }]
                
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='測試摘要', index=False)
                
                # 頁籤 2: 差異詳情（只包含有問題的項目）
                if differences:
                    df_diff = pd.DataFrame(differences)
                    
                    # 只保留有問題的項目（排除匹配的）
                    df_diff_filtered = df_diff[df_diff['狀態'] != '✅ 匹配']
                    
                    if not df_diff_filtered.empty:
                        df_diff_filtered.to_excel(writer, sheet_name='差異詳情', index=False)
                    else:
                        # 如果沒有差異，創建一個說明頁籤
                        df_no_diff = pd.DataFrame([{
                            '結果': '✅ 所有轉換規則測試通過！',
                            '說明': '所有 Master revision 轉換後都與 PreMP 正確版完全匹配'
                        }])
                        df_no_diff.to_excel(writer, sheet_name='測試結果', index=False)
                
                # 頁籤 3: 所有專案對照表
                all_comparisons = []
                for diff in differences:
                    all_comparisons.append({
                        'SN': diff['SN'],
                        '專案名稱': diff['專案名稱'],
                        'Master Revision': diff['Master Revision'],
                        '轉換後 Revision': diff['轉換後 Revision'],
                        'PreMP Revision (正確版)': diff['PreMP Revision (正確版)'],
                        '匹配結果': '✅' if diff['狀態'] == '✅ 匹配' else '❌'
                    })
                
                df_all = pd.DataFrame(all_comparisons)
                df_all.to_excel(writer, sheet_name='所有專案對照', index=False)
                
                # 頁籤 4: 轉換規則統計
                rule_stats = self._analyze_conversion_rules(differences)
                if rule_stats:
                    df_rules = pd.DataFrame(rule_stats)
                    df_rules.to_excel(writer, sheet_name='轉換規則統計', index=False)
                
                # 格式化所有工作表
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    self._format_worksheet(worksheet, sheet_name)
            
            self.logger.info(f"✅ 成功生成測試報告: {output_file}")
            return True
            
        except Exception as e:
            self.logger.error(f"生成 Excel 報告失敗: {str(e)}")
            return False
    
    def _analyze_conversion_rules(self, differences: List[Dict]) -> List[Dict]:
        """分析轉換規則的使用情況"""
        rule_usage = {}
        
        for diff in differences:
            if diff['Master Revision'] == 'N/A (專案不存在)':
                continue
                
            # 分析使用了哪種轉換規則
            master_rev = diff['Master Revision']
            converted_rev = diff['轉換後 Revision']
            
            # 判斷規則類型
            rule_type = self._identify_rule_type(master_rev, converted_rev)
            
            if rule_type not in rule_usage:
                rule_usage[rule_type] = {
                    '規則類型': rule_type,
                    '使用次數': 0,
                    '成功次數': 0,
                    '失敗次數': 0,
                    '範例': []
                }
            
            rule_usage[rule_type]['使用次數'] += 1
            
            if diff['狀態'] == '✅ 匹配':
                rule_usage[rule_type]['成功次數'] += 1
            else:
                rule_usage[rule_type]['失敗次數'] += 1
                
                # 記錄失敗範例（最多3個）
                if len(rule_usage[rule_type]['範例']) < 3:
                    rule_usage[rule_type]['範例'].append(f"{master_rev} → {converted_rev}")
        
        # 轉換為列表並加入成功率
        result = []
        for rule_type, stats in rule_usage.items():
            stats['成功率'] = f"{(stats['成功次數'] / stats['使用次數'] * 100):.1f}%" if stats['使用次數'] > 0 else '0%'
            stats['失敗範例'] = '\n'.join(stats['範例']) if stats['範例'] else 'N/A'
            del stats['範例']
            result.append(stats)
        
        return result
    
    def _identify_rule_type(self, master_rev: str, converted_rev: str) -> str:
        """識別使用的轉換規則類型"""
        # 檢查是否使用精確匹配
        if master_rev in config.MASTER_TO_PREMP_EXACT_MAPPING:
            return "精確匹配"
        
        # 檢查是否保持不變
        if master_rev == converted_rev:
            return "保持不變"
        
        # 檢查是否是晶片轉換
        for chip in config.CHIP_TO_RTD_MAPPING.keys():
            if f'/{chip}/' in master_rev:
                return f"晶片轉換 ({chip})"
        
        # 檢查是否是 upgrade 版本轉換
        if 'upgrade' in master_rev or 'upgrade' in converted_rev:
            return "Upgrade版本轉換"
        
        # 檢查是否是 kernel 版本轉換
        if 'linux-' in master_rev:
            return "Kernel版本轉換"
        
        # 檢查是否是 mp 到 premp 轉換
        if 'mp.google-refplus' in master_rev and 'premp.google-refplus' in converted_rev:
            return "MP到PreMP轉換"
        
        # 預設
        return "智能推斷或預設"
    
    def _format_worksheet(self, worksheet, sheet_name: str):
        """格式化 Excel 工作表"""
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        
        # 定義顏色
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # 差異顏色
        red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        green_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
        
        # 設定標題格式
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 根據頁籤設定特定格式
        if sheet_name == '差異詳情':
            # 為不同狀態設定背景色
            for row in range(2, worksheet.max_row + 1):
                status_cell = worksheet[f'G{row}']  # 狀態欄位
                if status_cell.value:
                    if '不匹配' in str(status_cell.value):
                        for col in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row, column=col).fill = red_fill
                    elif '不存在' in str(status_cell.value):
                        for col in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row, column=col).fill = yellow_fill
        
        # 自動調整欄寬
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def test_conversion(self, master_file: str, premp_file: str, output_file: str) -> bool:
        """
        執行轉換測試
        
        Args:
            master_file: master manifest.xml 檔案路徑
            premp_file: premp manifest.xml 檔案路徑
            output_file: 輸出 Excel 檔案路徑
            
        Returns:
            測試是否全部通過
        """
        try:
            self.logger.info("="*80)
            self.logger.info("開始測試 Master to PreMP 轉換規則")
            self.logger.info("="*80)
            
            # 步驟 1: 解析 manifest 檔案
            self.logger.info("\n📋 步驟 1: 解析 manifest 檔案")
            master_projects = self.parse_manifest(master_file)
            premp_projects = self.parse_manifest(premp_file)
            
            # 步驟 2: 比對轉換結果
            self.logger.info("\n🔍 步驟 2: 比對轉換結果")
            differences = self.compare_manifests(master_projects, premp_projects)
            
            # 步驟 3: 生成報告
            self.logger.info("\n📊 步驟 3: 生成測試報告")
            self.generate_excel_report(differences, output_file, master_file, premp_file)
            
            # 步驟 4: 顯示測試結果
            self.logger.info("\n📈 測試結果統計:")
            self.logger.info(f"  總專案數: {self.stats['total_projects']}")
            self.logger.info(f"  ✅ 匹配: {self.stats['matched']}")
            self.logger.info(f"  ❌ 不匹配: {self.stats['mismatched']}")
            self.logger.info(f"  ⚠️ PreMP中不存在: {self.stats['not_found_in_premp']}")
            self.logger.info(f"  🔶 僅存在於PreMP: {self.stats['extra_in_premp']}")
            
            # 計算測試是否通過
            all_passed = (self.stats['mismatched'] == 0 and 
                         self.stats['not_found_in_premp'] == 0)
            
            if all_passed:
                self.logger.info("\n✅ 所有轉換規則測試通過！")
            else:
                self.logger.warning(f"\n⚠️ 發現 {self.stats['mismatched']} 個轉換錯誤")
                self.logger.info(f"詳細差異請查看: {output_file}")
            
            self.logger.info("="*80)
            return all_passed
            
        except Exception as e:
            self.logger.error(f"測試執行失敗: {str(e)}")
            return False


def main():
    """主函數"""
    parser = argparse.ArgumentParser(description='測試 Master to PreMP Manifest 轉換規則')
    parser.add_argument('master_file', help='Master manifest.xml 檔案路徑')
    parser.add_argument('premp_file', help='PreMP manifest.xml 檔案路徑（正確版）')
    parser.add_argument('-o', '--output', default='conversion_test_report.xlsx',
                       help='輸出 Excel 檔案名稱（預設: conversion_test_report.xlsx）')
    
    args = parser.parse_args()
    
    # 確保輸出目錄存在
    output_dir = os.path.dirname(args.output) or '.'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # 執行測試
    tester = ManifestConversionTester()
    success = tester.test_conversion(args.master_file, args.premp_file, args.output)
    
    # 返回狀態碼
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()