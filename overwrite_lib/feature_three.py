"""
功能三：Manifest 轉換工具 - 微調版本
從 Gerrit 下載源檔案，進行 revision 轉換，並與目標檔案比較差異
微調：確保 Gerrit 檔案正確保存，增加 revision 比較資訊，標題格式化
修正：確保展開檔案正確保存到 output 資料夾
修改：改進特殊項目處理邏輯，通用檢查master和premp是否相同
"""
import os
import xml.etree.ElementTree as ET
import pandas as pd
import re
import tempfile
from typing import Dict, List, Any, Optional, Tuple
import utils
import config  # 🔥 新增：導入 config 模組
from excel_handler import ExcelHandler
from gerrit_manager import GerritManager

logger = utils.setup_logger(__name__)

class FeatureThree:
    """功能三：Manifest 轉換工具 - 微調版本"""
    
    def __init__(self):
        self.logger = logger
        self.excel_handler = ExcelHandler()
        self.gerrit_manager = GerritManager()
        
        # Gerrit 基礎 URL 模板
        android_master_branch = config.get_default_android_master_branch()
        self.gerrit_base_url = f"https://mm2sd.rtkbf.com/gerrit/plugins/gitiles/realtek/android/manifest/+/refs/heads/{android_master_branch}"
        
        # 檔案映射表
        self.source_files = {
            'master_to_premp': 'atv-google-refplus.xml',
            'premp_to_mp': 'atv-google-refplus-premp.xml',
            'mp_to_mpbackup': 'atv-google-refplus-wave.xml'
        }
        
        # 輸出檔案映射表
        self.output_files = {
            'master_to_premp': 'atv-google-refplus-premp.xml',
            'premp_to_mp': 'atv-google-refplus-wave.xml',
            'mp_to_mpbackup': 'atv-google-refplus-wave-backup.xml'
        }
        
        # 目標檔案映射表（用於比較）
        self.target_files = {
            'master_to_premp': 'atv-google-refplus-premp.xml',
            'premp_to_mp': 'atv-google-refplus-wave.xml', 
            'mp_to_mpbackup': 'atv-google-refplus-wave-backup.xml'
        }
    
    def process(self, overwrite_type: str, output_folder: str, 
                excel_filename: Optional[str] = None, push_to_gerrit: bool = False,
                rddb_number: Optional[str] = None, chip_status: Optional[Dict[str, str]] = None) -> bool:
        """
        處理功能三的主要邏輯
        
        Args:
            overwrite_type: 轉換類型
            output_folder: 輸出資料夾
            excel_filename: 自定義 Excel 檔名
            push_to_gerrit: 是否推送到 Gerrit
            rddb_number: RDDB 號碼（可選，預設使用 config.py 的設定）
            chip_status: 晶片狀態字典（可選，如 {'Mac7p': 'Y', 'Merlin7': 'Y'}）
            
        Returns:
            是否處理成功
        """
        # 將 rddb_number 儲存為實例變數，供後續使用
        self.rddb_number = rddb_number
        self.chip_status = chip_status
        
        try:
            self.logger.info("=== 開始執行功能三：Manifest 轉換工具 (微調版本 + include 展開) ===")
            self.logger.info(f"轉換類型: {overwrite_type}")
            self.logger.info(f"輸出資料夾: {output_folder}")
            self.logger.info(f"推送到 Gerrit: {'是' if push_to_gerrit else '否'}")
            
            # 驗證參數
            if overwrite_type not in self.source_files:
                self.logger.error(f"不支持的轉換類型: {overwrite_type}")
                return False
            
            # 確保輸出資料夾存在
            utils.ensure_dir(output_folder)
            
            # 記錄下載狀態
            source_download_success = False
            target_download_success = False
            
            # 步驟 1: 從 Gerrit 下載源檔案
            source_content = self._download_source_file(overwrite_type)
            if source_content:
                source_download_success = True
                self.logger.info("✅ 源檔案下載成功")
            else:
                self.logger.error("❌ 下載源檔案失敗")
                # 仍然繼續執行，生成錯誤報告
            
            # 步驟 1.5: 保存源檔案到 output 資料夾（加上 gerrit_ 前綴）
            source_file_path = None
            if source_content:
                source_file_path = self._save_source_file(source_content, overwrite_type, output_folder)
            
            # 🆕 步驟 1.6: 檢查是否有 include 標籤，如果有則展開
            expanded_content = None
            expanded_file_path = None
            use_expanded = False
            
            if source_content and self._has_include_tags(source_content):
                self.logger.info("🔍 檢測到 include 標籤，準備展開 manifest...")
                expanded_content, expanded_file_path = self._expand_manifest_with_repo_fixed(
                    overwrite_type, output_folder
                )
                if expanded_content and expanded_file_path:
                    use_expanded = True
                    self.logger.info("✅ Manifest 展開成功，將使用展開後的檔案進行轉換")
                    self.logger.info(f"✅ 展開檔案已保存到: {expanded_file_path}")
                    
                    # 🆕 驗證展開檔案是否真的存在
                    if os.path.exists(expanded_file_path):
                        file_size = os.path.getsize(expanded_file_path)
                        self.logger.info(f"✅ 展開檔案驗證成功: {os.path.basename(expanded_file_path)} ({file_size} bytes)")
                    else:
                        self.logger.error(f"❌ 展開檔案不存在: {expanded_file_path}")
                        use_expanded = False
                else:
                    self.logger.warning("⚠️ Manifest 展開失敗，將使用原始檔案")
            else:
                self.logger.info("ℹ️ 未檢測到 include 標籤，使用原始檔案")
            
            # 決定要使用的內容進行轉換
            content_for_conversion = expanded_content if use_expanded else source_content
            
            # 步驟 2: 進行 revision 轉換
            if content_for_conversion:
                converted_content, conversion_info = self._convert_revisions(content_for_conversion, overwrite_type)
            else:
                # 如果沒有源檔案，建立空的轉換結果
                converted_content = ""
                conversion_info = []
            
            # 步驟 3: 保存轉換後的檔案
            output_file_path = None
            if converted_content:
                output_file_path = self._save_converted_file(converted_content, overwrite_type, output_folder)
            
            # 步驟 4: 從 Gerrit 下載目標檔案進行比較
            target_content = self._download_target_file(overwrite_type)
            target_file_path = None
            if target_content:
                target_download_success = True
                target_file_path = self._save_target_file(target_content, overwrite_type, output_folder)
                self.logger.info("✅ 目標檔案下載成功")
            else:
                self.logger.warning("⚠️ 無法下載目標檔案，將跳過差異比較")
            
            # 步驟 5: 進行差異分析
            diff_analysis = self._analyze_differences(
                converted_content, target_content, overwrite_type, conversion_info
            )
            
            # 步驟 6: 推送到 Gerrit（如果需要）
            push_result = None
            if push_to_gerrit and converted_content:
                self.logger.info("🚀 開始推送到 Gerrit...")
                push_result = self._push_to_gerrit(overwrite_type, converted_content, target_content, output_folder)
            else:
                self.logger.info("⭐ 跳過 Gerrit 推送")
            
            # 步驟 7: 產生 Excel 報告（包含展開檔案資訊）
            excel_file = self._generate_excel_report_safe(
                overwrite_type, source_file_path, output_file_path, target_file_path, 
                diff_analysis, output_folder, excel_filename, source_download_success, 
                target_download_success, push_result, expanded_file_path, use_expanded
            )
            
            # 最終檔案檢查和報告
            self._final_file_report_complete(
                output_folder, source_file_path, output_file_path, target_file_path, 
                excel_file, source_download_success, target_download_success, expanded_file_path
            )
            
            self.logger.info(f"=== 功能三執行完成，Excel 報告：{excel_file} ===")
            return True
            
        except Exception as e:
            self.logger.error(f"功能三執行失敗: {str(e)}")
            
            # 即使失敗也嘗試生成錯誤報告
            try:
                error_excel = self._generate_error_report(output_folder, overwrite_type, str(e))
                if error_excel:
                    self.logger.info(f"已生成錯誤報告: {error_excel}")
            except:
                pass
            
            return False

    def _has_include_tags(self, xml_content: str) -> bool:
        """
        檢查 XML 內容是否包含 include 標籤
        
        Args:
            xml_content: XML 檔案內容
            
        Returns:
            是否包含 include 標籤
        """
        try:
            import re
            
            # 使用正則表達式檢查 include 標籤
            include_pattern = r'<include\s+name\s*=\s*["\'][^"\']*["\'][^>]*/?>'
            matches = re.findall(include_pattern, xml_content, re.IGNORECASE)
            
            if matches:
                self.logger.info(f"🔍 發現 {len(matches)} 個 include 標籤:")
                for i, match in enumerate(matches, 1):
                    self.logger.info(f"  {i}. {match}")
                return True
            else:
                self.logger.info("ℹ️ 未發現 include 標籤")
                return False
                
        except Exception as e:
            self.logger.error(f"檢查 include 標籤時發生錯誤: {str(e)}")
            return False
    
    def _expand_manifest_with_repo_fixed(self, overwrite_type: str, output_folder: str) -> tuple:
        """
        使用 repo 命令展開包含 include 的 manifest - 修正版本，同時保存到臨時目錄和輸出目錄
        """
        import subprocess
        import tempfile
        import shutil
        
        try:
            # 取得相關參數
            source_filename = self.source_files[overwrite_type]
            repo_url = "ssh://mm2sd.rtkbf.com:29418/realtek/android/manifest"
            # 🔥 修改：使用動態分支
            branch = config.get_default_android_master_branch()
            
            # 生成展開檔案名稱 - 使用絕對路徑解決臨時目錄問題
            expanded_filename = f"gerrit_{source_filename.replace('.xml', '_expand.xml')}"
            # 關鍵修正：轉為絕對路徑，避免在臨時目錄中誤保存
            final_expanded_path = os.path.abspath(os.path.join(output_folder, expanded_filename))
            
            self.logger.info(f"🎯 準備展開 manifest...")
            self.logger.info(f"🎯 源檔案: {source_filename}")
            self.logger.info(f"🎯 使用分支: {branch}")  # 🔥 新增：顯示使用的動態分支
            self.logger.info(f"🎯 展開檔案名: {expanded_filename}")
            self.logger.info(f"🎯 目標絕對路徑: {final_expanded_path}")
            
            # 在切換目錄前確保輸出資料夾存在
            utils.ensure_dir(output_folder)
            abs_output_folder = os.path.abspath(output_folder)
            self.logger.info(f"🎯 輸出資料夾絕對路徑: {abs_output_folder}")
            
            # 檢查 repo 命令是否可用
            try:
                repo_check = subprocess.run(
                    ["repo", "--version"], 
                    capture_output=True, 
                    text=True, 
                    timeout=10
                )
                if repo_check.returncode == 0:
                    self.logger.info(f"✅ repo 工具可用: {repo_check.stdout.strip()}")
                else:
                    self.logger.error(f"❌ repo 工具檢查失敗: {repo_check.stderr}")
                    return None, None
            except FileNotFoundError:
                self.logger.error("❌ repo 命令未找到，請確認已安裝 repo 工具")
                self.logger.error("安裝方法: curl https://storage.googleapis.com/git-repo-downloads/repo > ~/.local/bin/repo && chmod a+x ~/.local/bin/repo")
                return None, None
            except Exception as e:
                self.logger.error(f"❌ repo 工具檢查異常: {str(e)}")
                return None, None
            
            # 建立臨時工作目錄
            temp_work_dir = tempfile.mkdtemp(prefix='repo_expand_')
            self.logger.info(f"📁 建立臨時工作目錄: {temp_work_dir}")
            
            original_cwd = os.getcwd()
            
            try:
                # 切換到臨時目錄
                os.chdir(temp_work_dir)
                self.logger.info(f"📂 切換到臨時目錄: {temp_work_dir}")
                
                # 🔥 步驟 1: repo init - 使用動態分支
                self.logger.info(f"📄 執行 repo init...")
                init_cmd = [
                    "repo", "init", 
                    "-u", repo_url,
                    "-b", branch,  # 🔥 使用動態分支
                    "-m", source_filename
                ]
                
                self.logger.info(f"🎯 Init 指令: {' '.join(init_cmd)}")
                
                init_result = subprocess.run(
                    init_cmd,
                    capture_output=True,
                    text=True,
                    timeout=120  # 2分鐘超時
                )
                
                self.logger.info(f"🔍 repo init 返回碼: {init_result.returncode}")
                if init_result.stdout:
                    self.logger.info(f"🔍 repo init stdout: {init_result.stdout}")
                if init_result.stderr:
                    self.logger.info(f"🔍 repo init stderr: {init_result.stderr}")
                
                if init_result.returncode != 0:
                    self.logger.error(f"❌ repo init 失敗 (返回碼: {init_result.returncode})")
                    return None, None
                
                self.logger.info("✅ repo init 成功")
                
                # 檢查 .repo 目錄是否存在
                repo_dir = os.path.join(temp_work_dir, ".repo")
                if os.path.exists(repo_dir):
                    self.logger.info(f"✅ .repo 目錄已建立: {repo_dir}")
                else:
                    self.logger.error(f"❌ .repo 目錄不存在: {repo_dir}")
                    return None, None
                
                # 步驟 2: repo manifest 展開
                self.logger.info(f"📄 執行 repo manifest 展開...")
                
                manifest_cmd = ["repo", "manifest"]
                self.logger.info(f"🎯 Manifest 指令: {' '.join(manifest_cmd)}")
                
                manifest_result = subprocess.run(
                    manifest_cmd,
                    capture_output=True,
                    text=True,
                    timeout=60
                )
                
                self.logger.info(f"🔍 repo manifest 返回碼: {manifest_result.returncode}")
                if manifest_result.stderr:
                    self.logger.info(f"🔍 repo manifest stderr: {manifest_result.stderr}")
                
                if manifest_result.returncode != 0:
                    self.logger.error(f"❌ repo manifest 失敗 (返回碼: {manifest_result.returncode})")
                    return None, None
                
                expanded_content = manifest_result.stdout
                
                if not expanded_content.strip():
                    self.logger.error("❌ repo manifest 返回空內容")
                    return None, None
                
                self.logger.info(f"✅ repo manifest 成功，內容長度: {len(expanded_content)} 字符")
                
                # 檢查展開內容的基本特徵
                project_count = expanded_content.count('<project ')
                include_count = expanded_content.count('<include ')
                self.logger.info(f"🔍 展開內容分析:")
                self.logger.info(f"   - Project 標籤數量: {project_count}")
                self.logger.info(f"   - Include 標籤數量: {include_count}")
                
                # 步驟 3A: 在臨時目錄保存一份展開檔案
                temp_expanded_path = os.path.join(temp_work_dir, expanded_filename)
                self.logger.info(f"📁 在臨時目錄保存展開檔案: {temp_expanded_path}")
                
                try:
                    with open(temp_expanded_path, 'w', encoding='utf-8') as f:
                        f.write(expanded_content)
                    self.logger.info(f"✅ 臨時目錄檔案保存成功")
                    
                    # 驗證臨時檔案
                    if os.path.exists(temp_expanded_path):
                        temp_file_size = os.path.getsize(temp_expanded_path)
                        self.logger.info(f"✅ 臨時檔案驗證: {temp_file_size} bytes")
                    
                except Exception as temp_write_error:
                    self.logger.error(f"❌ 臨時目錄檔案保存失敗: {str(temp_write_error)}")
                    return None, None
                
                # 步驟 3B: 同時複製到輸出資料夾（使用絕對路徑）
                self.logger.info(f"📁 複製展開檔案到輸出資料夾...")
                self.logger.info(f"📁 目標絕對路徑: {final_expanded_path}")
                self.logger.info(f"📁 當前工作目錄: {os.getcwd()}")
                
                # 關鍵：確保目標資料夾存在（使用絕對路徑）
                target_dir = os.path.dirname(final_expanded_path)
                utils.ensure_dir(target_dir)
                self.logger.info(f"✅ 目標資料夾確認存在: {target_dir}")
                
                # 複製檔案到輸出目錄（使用絕對路徑）
                try:
                    shutil.copy2(temp_expanded_path, final_expanded_path)
                    self.logger.info(f"✅ 檔案複製完成（臨時→輸出）")
                except Exception as copy_error:
                    self.logger.error(f"❌ 檔案複製失敗: {str(copy_error)}")
                    self.logger.error(f"❌ 源路徑: {temp_expanded_path}")
                    self.logger.error(f"❌ 目標路徑: {final_expanded_path}")
                    return None, None
                
                # 步驟 4: 驗證兩個位置的檔案都存在
                self.logger.info(f"🔍 驗證檔案保存狀態...")
                
                # 驗證臨時檔案
                if os.path.exists(temp_expanded_path):
                    temp_size = os.path.getsize(temp_expanded_path)
                    self.logger.info(f"✅ 臨時檔案存在: {temp_expanded_path} ({temp_size} bytes)")
                else:
                    self.logger.error(f"❌ 臨時檔案不存在: {temp_expanded_path}")
                
                # 驗證輸出檔案
                if os.path.exists(final_expanded_path):
                    file_size = os.path.getsize(final_expanded_path)
                    self.logger.info(f"✅ 輸出檔案存在: {final_expanded_path} ({file_size} bytes)")
                    
                    # 驗證檔案內容一致性
                    try:
                        with open(final_expanded_path, 'r', encoding='utf-8') as f:
                            saved_content = f.read()
                            
                        if len(saved_content) == len(expanded_content):
                            self.logger.info(f"✅ 檔案內容驗證成功 ({len(saved_content)} 字符)")
                        else:
                            self.logger.warning(f"⚠️ 檔案內容長度不匹配: 原始 {len(expanded_content)}, 保存 {len(saved_content)}")
                            
                        # 驗證專案數量
                        saved_project_count = saved_content.count('<project ')
                        self.logger.info(f"✅ 保存檔案專案數量: {saved_project_count}")
                        
                    except Exception as read_error:
                        self.logger.error(f"❌ 檔案內容驗證失敗: {str(read_error)}")
                        return None, None
                    
                    # 成功返回
                    self.logger.info(f"🎉 展開檔案處理完成!")
                    self.logger.info(f"   📁 臨時位置: {temp_expanded_path}")
                    self.logger.info(f"   📁 輸出位置: {final_expanded_path}")
                    self.logger.info(f"   📊 檔案大小: {file_size} bytes")
                    self.logger.info(f"   📊 專案數量: {project_count}")
                    
                    return expanded_content, final_expanded_path
                else:
                    self.logger.error(f"❌ 輸出檔案不存在: {final_expanded_path}")
                    
                    # 檢查輸出目錄狀態
                    if os.path.exists(abs_output_folder):
                        files_in_output = os.listdir(abs_output_folder)
                        self.logger.error(f"❌ 輸出目錄內容: {files_in_output}")
                    else:
                        self.logger.error(f"❌ 輸出目錄不存在: {abs_output_folder}")
                    
                    return None, None
                
            finally:
                # 在清理前顯示臨時目錄內容
                self.logger.info(f"🔍 清理前臨時目錄內容:")
                try:
                    temp_files = os.listdir(temp_work_dir)
                    for filename in temp_files[:10]:  # 只顯示前10個檔案
                        filepath = os.path.join(temp_work_dir, filename)
                        if os.path.isfile(filepath):
                            filesize = os.path.getsize(filepath)
                            self.logger.info(f"   📄 {filename} ({filesize} bytes)")
                        else:
                            self.logger.info(f"   📁 {filename} (目錄)")
                except Exception as e:
                    self.logger.warning(f"⚠️ 無法列出臨時目錄內容: {str(e)}")
                
                # 恢復原始工作目錄
                os.chdir(original_cwd)
                self.logger.info(f"📂 恢復原始工作目錄: {original_cwd}")
                
                # 延遲清理臨時目錄（可選：保留一段時間供調試）
                # 注意：這裡我們還是清理，但添加了更多日誌
                try:
                    shutil.rmtree(temp_work_dir)
                    self.logger.info(f"🗑️  清理臨時目錄成功: {temp_work_dir}")
                except Exception as e:
                    self.logger.warning(f"⚠️ 清理臨時目錄失敗: {str(e)}")
                
        except subprocess.TimeoutExpired:
            self.logger.error("❌ repo 命令執行超時")
            return None, None
        except Exception as e:
            self.logger.error(f"❌ 展開 manifest 時發生錯誤: {str(e)}")
            import traceback
            self.logger.error(f"❌ 錯誤詳情: {traceback.format_exc()}")
            return None, None
            
    def _download_source_file(self, overwrite_type: str) -> Optional[str]:
        """從 Gerrit 下載源檔案"""
        try:
            source_filename = self.source_files[overwrite_type]
            source_url = f"{self.gerrit_base_url}/{source_filename}"
            
            self.logger.info(f"下載源檔案: {source_filename}")
            self.logger.info(f"URL: {source_url}")
            
            # 使用臨時檔案下載
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xml') as temp_file:
                temp_path = temp_file.name
            
            try:
                success = self.gerrit_manager.download_file_from_link(source_url, temp_path)
                
                if success and os.path.exists(temp_path):
                    with open(temp_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    
                    self.logger.info(f"成功下載源檔案: {len(content)} 字符")
                    return content
                else:
                    self.logger.error("下載源檔案失敗")
                    return None
                    
            finally:
                # 清理臨時檔案
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
                    
        except Exception as e:
            self.logger.error(f"下載源檔案異常: {str(e)}")
            return None
    
    def _download_target_file(self, overwrite_type: str) -> Optional[str]:
        """從 Gerrit 下載目標檔案進行比較"""
        try:
            target_filename = self.target_files[overwrite_type]
            target_url = f"{self.gerrit_base_url}/{target_filename}"
            
            self.logger.info(f"下載目標檔案: {target_filename}")
            self.logger.info(f"URL: {target_url}")
            
            # 使用臨時檔案下載
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xml') as temp_file:
                temp_path = temp_file.name
            
            try:
                success = self.gerrit_manager.download_file_from_link(target_url, temp_path)
                
                if success and os.path.exists(temp_path):
                    with open(temp_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    
                    self.logger.info(f"成功下載目標檔案: {len(content)} 字符")
                    self.logger.info(f"臨時檔案位置: {temp_path}")
                    return content
                else:
                    self.logger.warning("下載目標檔案失敗，將只進行轉換不做比較")
                    return None
                    
            finally:
                # 清理臨時檔案
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
                    self.logger.info(f"已清理臨時檔案: {temp_path}")
                    
        except Exception as e:
            self.logger.warning(f"下載目標檔案異常: {str(e)}")
            return None
    
    # 🔥 完全重寫 _convert_revisions 方法，移除有問題的正則表達式
    def _convert_revisions(self, xml_content: str, overwrite_type: str) -> Tuple[str, List[Dict]]:
        """
        根據轉換類型進行 revision 轉換 - 修正正則表達式錯誤版本
        確保儲存 default revision 供 source_link 生成使用
        """
        try:
            self.logger.info(f"開始進行 revision 轉換: {overwrite_type}")
            
            # 先解析 XML 取得 default 資訊
            temp_root = ET.fromstring(xml_content)
            
            # 🆕 讀取 default 標籤的 remote 和 revision 屬性
            default_remote = ''
            default_revision = ''
            default_element = temp_root.find('default')
            if default_element is not None:
                default_remote = default_element.get('remote', '')
                default_revision = default_element.get('revision', '')
                self.logger.info(f"找到預設 remote: {default_remote}, revision: {default_revision}")
            
            # 🆕 儲存為實例變數供 source_link 生成使用
            self.default_remote = default_remote
            self.default_revision = default_revision
            
            conversion_info = []
            conversion_count = 0
            skipped_no_revision = 0  # 🔥 統計沒有 revision 而跳過的專案
            hash_revision_count = 0
            branch_revision_count = 0
            upstream_used_count = 0
            
            # 建立轉換後的內容（從原始字串開始）
            converted_content = xml_content
            
            # 遍歷所有 project 元素以記錄轉換資訊
            for project in temp_root.findall('project'):
                project_name = project.get('name', '')
                project_remote = project.get('remote', '') or default_remote
                original_revision = project.get('revision', '')  # 🔥 只使用原始的 revision
                upstream = project.get('upstream', '')
                
                # 🔥 添加調試：記錄所有處理的專案
                composite_key = f"{project_name}|{project.get('path', '')}"
                self.logger.debug(f"處理專案: {composite_key}, 原始 revision: '{original_revision}'")
                
                # 🔥 如果沒有 revision，記錄但跳過轉換
                if not original_revision:
                    skipped_no_revision += 1
                    self.logger.debug(f"跳過沒有 revision 的專案: {project_name}")
                    
                    # 🔥 重要：即使沒有 revision，也要加入 conversion_info
                    conversion_info.append({
                        'name': project_name,
                        'path': project.get('path', ''),
                        'original_revision': '',
                        'effective_revision': '',
                        'converted_revision': '',
                        'upstream': upstream,
                        'dest-branch': project.get('dest-branch', ''),
                        'groups': project.get('groups', ''),
                        'clone-depth': project.get('clone-depth', ''),
                        'remote': project.get('remote', ''),
                        'original_remote': project.get('remote', ''),
                        'changed': False,
                        'used_default_revision': False,
                        'used_upstream_for_conversion': False
                    })
                    continue
                
                # 使用新邏輯取得用於轉換的有效 revision
                effective_revision = self._get_effective_revision_for_conversion(project)
                
                # 統計 revision 類型
                if self._is_revision_hash(original_revision):
                    hash_revision_count += 1
                    if upstream:
                        upstream_used_count += 1
                elif original_revision:
                    branch_revision_count += 1
                
                if not effective_revision:
                    self.logger.debug(f"專案 {project_name} 沒有有效的轉換 revision，但仍記錄")
                    # 🔥 沒有有效 revision 也要記錄
                    conversion_info.append({
                        'name': project_name,
                        'path': project.get('path', ''),
                        'original_revision': original_revision,
                        'effective_revision': '',
                        'converted_revision': original_revision,  # 保持原值
                        'upstream': upstream,
                        'dest-branch': project.get('dest-branch', ''),
                        'groups': project.get('groups', ''),
                        'clone-depth': project.get('clone-depth', ''),
                        'remote': project.get('remote', ''),
                        'original_remote': project.get('remote', ''),
                        'changed': False,
                        'used_default_revision': False,
                        'used_upstream_for_conversion': False
                    })
                    continue
                
                old_revision = effective_revision
                new_revision = self._convert_single_revision(effective_revision, overwrite_type)
                
                # 🔥 增強除錯 - MP to MPBackup 專用
                if overwrite_type == 'mp_to_mpbackup':
                    self.logger.debug(f"🔍 MP to MPBackup 轉換除錯:")
                    self.logger.debug(f"  專案: {project_name}")
                    self.logger.debug(f"  原始 revision: {original_revision}")
                    self.logger.debug(f"  有效 revision: {effective_revision}")
                    self.logger.debug(f"  轉換結果: {new_revision}")
                    self.logger.debug(f"  是否改變: {new_revision != old_revision}")
                
                # 🔥 重要：所有專案都要記錄到 conversion_info 中
                conversion_info.append({
                    'name': project_name,
                    'path': project.get('path', ''),
                    'original_revision': original_revision,
                    'effective_revision': effective_revision,
                    'converted_revision': new_revision,
                    'upstream': upstream,
                    'dest-branch': project.get('dest-branch', ''),
                    'groups': project.get('groups', ''),
                    'clone-depth': project.get('clone-depth', ''),
                    'remote': project.get('remote', ''),
                    'original_remote': project.get('remote', ''),  # 🔥 保存原始 remote
                    'changed': new_revision != old_revision,
                    'used_default_revision': False,  # 🔥 不再插入 default revision
                    'used_upstream_for_conversion': self._is_revision_hash(original_revision) and upstream
                })
                
                # 如果需要轉換，在字串中直接替換
                if new_revision != old_revision:
                    # 🔥 使用安全的替換方法
                    replacement_success = self._safe_replace_revision_in_xml(
                        converted_content, project_name, old_revision, new_revision
                    )
                    
                    if replacement_success:
                        converted_content = replacement_success
                        conversion_count += 1
                        self.logger.debug(f"字串替換成功: {project_name} - {old_revision} → {new_revision}")
                        
                        # 🔥 MP to MPBackup 特別記錄
                        if overwrite_type == 'mp_to_mpbackup':
                            self.logger.info(f"✅ MP to MPBackup 轉換成功: {project_name}")
                            self.logger.info(f"  {old_revision} → {new_revision}")
            
            self.logger.info(f"revision 轉換完成，共轉換 {conversion_count} 個專案")
            self.logger.info(f"📊 處理統計:")
            self.logger.info(f"  - ⭐ 跳過沒有 revision 的專案: {skipped_no_revision} 個")
            self.logger.info(f"  - 📸 Hash revision: {hash_revision_count} 個")
            self.logger.info(f"  - 📹 Branch revision: {branch_revision_count} 個")
            self.logger.info(f"  - ⬆️ 使用 upstream 進行轉換: {upstream_used_count} 個")
            self.logger.info(f"  - 📋 總記錄專案數: {len(conversion_info)} 個")
            self.logger.info("✅ 保留了所有原始格式：XML 宣告、註解、空格、換行等")
            
            # 🔥 特別檢查 MP to MPBackup 轉換效果
            if overwrite_type == 'mp_to_mpbackup':
                self._verify_mp_to_mpbackup_conversion(converted_content, xml_content)
            
            return converted_content, conversion_info
            
        except Exception as e:
            self.logger.error(f"revision 轉換失敗: {str(e)}")
            import traceback
            self.logger.error(f"錯誤詳情: {traceback.format_exc()}")
            return xml_content, []

    def _verify_mp_to_mpbackup_conversion(self, converted_content: str, original_content: str):
        """驗證 MP to MPBackup 轉換是否成功"""
        try:
            # 統計轉換前後的 revision 差異
            original_wave_count = original_content.count('mp.google-refplus.wave')
            original_backup_count = original_content.count('mp.google-refplus.wave.backup')
            
            converted_wave_count = converted_content.count('mp.google-refplus.wave')
            converted_backup_count = converted_content.count('mp.google-refplus.wave.backup')
                
        except Exception as e:
            self.logger.error(f"驗證 MP to MPBackup 轉換時發生錯誤: {str(e)}")
            
    def _safe_insert_revision(self, xml_content: str, project_name: str, revision: str) -> str:
        """
        安全地為專案插入 revision 屬性
        """
        try:
            lines = xml_content.split('\n')
            
            for i, line in enumerate(lines):
                if f'name="{project_name}"' in line and 'revision=' not in line:
                    # 找到沒有 revision 的專案行，插入 revision
                    if line.strip().endswith('/>'):
                        # 單行標籤
                        new_line = line.replace('/>', f' revision="{revision}"/>')
                        lines[i] = new_line
                        self.logger.debug(f"✅ 插入 revision: {project_name}")
                        break
                    elif line.strip().endswith('>'):
                        # 多行標籤的開始
                        new_line = line.replace('>', f' revision="{revision}">')
                        lines[i] = new_line
                        self.logger.debug(f"✅ 插入 revision (多行): {project_name}")
                        break
            
            return '\n'.join(lines)
            
        except Exception as e:
            self.logger.error(f"插入 revision 失敗: {str(e)}")
            return xml_content
            
    def _safe_replace_revision_in_xml(self, xml_content: str, project_name: str, 
                                 old_revision: str, new_revision: str) -> str:
        """
        安全的 XML 字串替換 - 避免有問題的正則表達式
        """
        try:
            # 🔥 使用簡單的字串搜尋和替換，避免複雜的正則表達式
            lines = xml_content.split('\n')
            modified = False
            
            for i, line in enumerate(lines):
                # 檢查這一行是否包含目標專案
                if f'name="{project_name}"' in line and 'revision=' in line:
                    # 找到目標行，進行替換
                    if f'revision="{old_revision}"' in line:
                        lines[i] = line.replace(f'revision="{old_revision}"', f'revision="{new_revision}"')
                        modified = True
                        self.logger.debug(f"✅ 替換成功: {project_name}")
                        break
                    elif f"revision='{old_revision}'" in line:
                        lines[i] = line.replace(f"revision='{old_revision}'", f"revision='{new_revision}'")
                        modified = True
                        self.logger.debug(f"✅ 替換成功 (單引號): {project_name}")
                        break
            
            if modified:
                return '\n'.join(lines)
            else:
                self.logger.warning(f"❌ 未找到匹配的替換: {project_name} - {old_revision}")
                return xml_content
                
        except Exception as e:
            self.logger.error(f"安全替換失敗: {str(e)}")
            return xml_content
            
    def _convert_single_revision(self, revision: str, overwrite_type: str) -> str:
        """轉換單一 revision"""
        if overwrite_type == 'master_to_premp':
            return self._convert_master_to_premp(revision)
        elif overwrite_type == 'premp_to_mp':
            return self._convert_premp_to_mp(revision)
        elif overwrite_type == 'mp_to_mpbackup':
            return self._convert_mp_to_mpbackup(revision)
        else:
            return revision
    
    def _convert_master_to_premp(self, revision: str) -> str:
        """
        master → premp 轉換規則 - 使用動態 Android 版本，動態 kernel 版本匹配
        """
        if not revision:
            return revision
        
        original_revision = revision.strip()
        
        # 跳過 Google 開頭的項目
        if original_revision.startswith('google/'):
            self.logger.debug(f"跳過 Google 項目: {original_revision}")
            return original_revision
        
        # 跳過特殊項目
        if self._should_skip_revision_conversion(original_revision):
            return original_revision
        
        # 🔥 修改：精確匹配轉換規則 - 使用動態版本（移除預定義 kernel 版本）
        exact_mappings = {
            # 基本 master 分支轉換
            'realtek/master': config.get_default_premp_branch(),
            'realtek/gaia': config.get_default_premp_branch(),
            'realtek/gki/master': config.get_default_premp_branch(),
            
            # Android master 分支
            config.get_default_android_master_branch(): config.get_default_premp_branch(),
            
            # mp.google-refplus 轉換
            'realtek/mp.google-refplus': config.get_default_premp_branch(),
            config.get_android_path('realtek/android-{android_version}/mp.google-refplus'): config.get_default_premp_branch(),
        }
        
        # 檢查精確匹配
        if original_revision in exact_mappings:
            result = exact_mappings[original_revision]
            self.logger.debug(f"精確匹配轉換: {original_revision} → {result}")
            return result
        
        # 🔥 修改：模式匹配轉換規則 - 完全使用正則表達式動態匹配
        import re
        
        # vX.X.X 版本轉換 - 保留版本號
        pattern_version = r'realtek/(v\d+\.\d+(?:\.\d+)?)/master$'
        match_version = re.match(pattern_version, original_revision)
        if match_version:
            version = match_version.group(1)
            result = f'realtek/{version}/premp.google-refplus'
            self.logger.debug(f"版本格式轉換: {original_revision} → {result}")
            return result
        
        # 規則 1: mp.google-refplus.upgrade-11.rtdXXXX → premp.google-refplus.upgrade-11.rtdXXXX
        pattern1 = r'realtek/android-(\d+)/mp\.google-refplus\.upgrade-(\d+)\.(rtd\w+)'
        match1 = re.match(pattern1, original_revision)
        if match1:
            android_ver, upgrade_ver, rtd_chip = match1.groups()
            if android_ver == config.get_current_android_version():
                result = config.get_premp_branch_with_upgrade(upgrade_ver, rtd_chip)
            else:
                # 如果是不同的 Android 版本，保持原版本
                result = f'realtek/android-{android_ver}/premp.google-refplus.upgrade-{upgrade_ver}.{rtd_chip}'
            self.logger.debug(f"模式1轉換: {original_revision} → {result}")
            return result
        
        # 規則 2: mp.google-refplus.upgrade-11 → premp.google-refplus.upgrade-11
        pattern2 = r'realtek/android-(\d+)/mp\.google-refplus\.upgrade-(\d+)$'
        match2 = re.match(pattern2, original_revision)
        if match2:
            android_ver, upgrade_ver = match2.groups()
            if android_ver == config.get_current_android_version():
                result = config.get_premp_branch_with_upgrade(upgrade_ver)
            else:
                result = f'realtek/android-{android_ver}/premp.google-refplus.upgrade-{upgrade_ver}'
            self.logger.debug(f"模式2轉換: {original_revision} → {result}")
            return result
        
        # 🔥 規則 3: linux-X.X/master → linux-X.X/android-{current_version}/premp.google-refplus（完全動態）
        pattern3 = r'realtek/linux-([\d.]+)/master$'
        match3 = re.match(pattern3, original_revision)
        if match3:
            linux_ver = match3.group(1)
            result = config.get_linux_android_path(
                linux_ver, 'realtek/linux-{linux_ver}/android-{android_version}/premp.google-refplus'
            )
            self.logger.debug(f"模式3轉換（動態 kernel 版本）: {original_revision} → {result}")
            return result
        
        # 🔥 規則 4: linux-X.X/android-Y/master → linux-X.X/android-{current_version}/premp.google-refplus（完全動態）
        pattern4 = r'realtek/linux-([\d.]+)/android-(\d+)/master$'
        match4 = re.match(pattern4, original_revision)
        if match4:
            linux_ver, android_ver = match4.groups()
            # 自動升級到當前 Android 版本
            result = config.get_linux_android_path(
                linux_ver, 'realtek/linux-{linux_ver}/android-{android_version}/premp.google-refplus'
            )
            self.logger.debug(f"模式4轉換（動態 kernel，升級 Android）: {original_revision} → {result}")
            return result
        
        # 🔥 規則 5: linux-X.X/android-Y/mp.google-refplus → linux-X.X/android-{current_version}/premp.google-refplus（完全動態）
        pattern5 = r'realtek/linux-([\d.]+)/android-(\d+)/mp\.google-refplus$'
        match5 = re.match(pattern5, original_revision)
        if match5:
            linux_ver, android_ver = match5.groups()
            result = config.get_linux_android_path(
                linux_ver, 'realtek/linux-{linux_ver}/android-{android_version}/premp.google-refplus'
            )
            self.logger.debug(f"模式5轉換（動態 kernel）: {original_revision} → {result}")
            return result
        
        # 🔥 規則 6: linux-X.X/android-Y/mp.google-refplus.rtdXXXX → linux-X.X/android-{current_version}/premp.google-refplus.rtdXXXX（完全動態）
        pattern6 = r'realtek/linux-([\d.]+)/android-(\d+)/mp\.google-refplus\.(rtd\w+)'
        match6 = re.match(pattern6, original_revision)
        if match6:
            linux_ver, android_ver, rtd_chip = match6.groups()
            base_path = config.get_linux_android_path(
                linux_ver, 'realtek/linux-{linux_ver}/android-{android_version}/premp.google-refplus'
            )
            result = f"{base_path}.{rtd_chip}"
            self.logger.debug(f"模式6轉換（動態 kernel）: {original_revision} → {result}")
            return result
        
        # 規則 7: android-Y/mp.google-refplus → android-{current_version}/premp.google-refplus
        pattern7 = r'realtek/android-(\d+)/mp\.google-refplus$'
        match7 = re.match(pattern7, original_revision)
        if match7:
            android_ver = match7.group(1)
            result = config.get_default_premp_branch()
            self.logger.debug(f"模式7轉換（升級到當前版本）: {original_revision} → {result}")
            return result
        
        # 規則 8: android-Y/mp.google-refplus.rtdXXXX → android-{current_version}/premp.google-refplus.rtdXXXX
        pattern8 = r'realtek/android-(\d+)/mp\.google-refplus\.(rtd\w+)'
        match8 = re.match(pattern8, original_revision)
        if match8:
            android_ver, rtd_chip = match8.groups()
            result = config.get_premp_branch_with_chip(rtd_chip)
            self.logger.debug(f"模式8轉換（升級到當前版本）: {original_revision} → {result}")
            return result
        
        # 規則 9: 晶片特定的 master 分支 → premp.google-refplus.rtdXXXX（使用當前 Android 版本）
        for chip, rtd_model in config.CHIP_TO_RTD_MAPPING.items():
            if f'realtek/{chip}/master' == original_revision:
                result = config.get_premp_branch_with_chip(rtd_model)
                self.logger.debug(f"晶片轉換（當前 Android 版本）: {original_revision} → {result}")
                return result
        
        # 智能轉換備案
        smart_result = self._smart_conversion_fallback(original_revision)
        self.logger.debug(f"智能轉換: {original_revision} → {smart_result}")
        return smart_result

    def _should_skip_revision_conversion(self, revision: str) -> bool:
        """
        判斷是否應該跳過 revision 轉換 - 與測試模組完全同步
        
        Args:
            revision: 原始 revision
            
        Returns:
            是否應該跳過轉換
        """
        if not revision:
            return True
        
        # 🆕 跳過 Google 開頭的項目
        if revision.startswith('google/'):
            return True
        
        # 跳過 refs/tags/
        if revision.startswith('refs/tags/'):
            return True
        
        return False

    def _smart_conversion_fallback(self, revision: str) -> str:
        """
        智能轉換備案 - 使用動態 Android 版本
        """
        # 如果包含 mp.google-refplus，嘗試替換為 premp.google-refplus
        if 'mp.google-refplus' in revision:
            result = revision.replace('mp.google-refplus', 'premp.google-refplus')
            self.logger.debug(f"智能替換 mp→premp: {revision} → {result}")
            return result
        
        # 如果是 master 但沒有匹配到特定規則
        if '/master' in revision and 'realtek/' in revision:
            # 🔥 修改：使用動態 Android 版本提取
            import re
            android_match = re.search(r'android-(\d+)', revision)
            if android_match:
                # 保持原 Android 版本或升級到當前版本（可選）
                result = config.get_default_premp_branch()  # 使用當前版本
                self.logger.debug(f"智能Android版本轉換（升級到當前）: {revision} → {result}")
                return result
            else:
                result = config.get_default_premp_branch()
                self.logger.debug(f"智能預設轉換: {revision} → {result}")
                return result
        
        # 如果完全沒有匹配，返回當前版本的預設值
        result = config.get_default_premp_branch()
        self.logger.debug(f"備案預設轉換（當前版本）: {revision} → {result}")
        return result
            
    def _convert_premp_to_mp(self, revision: str) -> str:
        """premp → mp 轉換規則"""
        # 將 premp.google-refplus 關鍵字替換為 mp.google-refplus.wave
        return revision.replace('premp.google-refplus', 'mp.google-refplus.wave')
    
    def _convert_mp_to_mpbackup(self, revision: str) -> str:
        """mp → mpbackup 轉換規則 - 修正正則表達式錯誤"""
        if not revision:
            return revision
        
        original_revision = revision.strip()
        
        # 記錄轉換前的狀態
        self.logger.debug(f"MP to MPBackup 轉換輸入: {original_revision}")
        
        # 檢查是否已經是 backup 格式
        if 'mp.google-refplus.wave.backup' in original_revision:
            self.logger.debug(f"已經是 backup 格式，不需轉換: {original_revision}")
            return original_revision
        
        # 🔥 主要轉換邏輯 - 簡化版，避免複雜正則表達式
        if 'mp.google-refplus.wave' in original_revision and 'backup' not in original_revision:
            result = original_revision.replace('mp.google-refplus.wave', 'mp.google-refplus.wave.backup')
            self.logger.debug(f"標準轉換: {original_revision} → {result}")
            return result
        
        # 🔥 處理以 .wave 結尾但沒有 backup 的情況
        if original_revision.endswith('.wave') and 'mp.google-refplus' in original_revision and 'backup' not in original_revision:
            result = original_revision + '.backup'
            self.logger.debug(f"後綴轉換: {original_revision} → {result}")
            return result
        
        # 🔥 處理包含 wave 但格式特殊的情況 - 使用安全的字串操作
        if 'mp.google-refplus' in original_revision and 'wave' in original_revision and 'backup' not in original_revision:
            # 找到 wave 的位置，在後面加 .backup
            wave_index = original_revision.find('wave')
            if wave_index != -1:
                # 檢查 wave 後面是否直接結束或跟著其他字符
                after_wave = original_revision[wave_index + 4:]  # wave 有4個字符
                if not after_wave or after_wave.startswith('.') or after_wave.startswith('/'):
                    # 在 wave 後面插入 .backup
                    result = original_revision[:wave_index + 4] + '.backup' + after_wave
                    self.logger.debug(f"插入轉換: {original_revision} → {result}")
                    return result
        
        # 如果沒有匹配，返回原值
        self.logger.debug(f"MP to MPBackup 轉換無變化: {original_revision}")
        return original_revision
    
    def _save_source_file(self, content: str, overwrite_type: str, output_folder: str) -> str:
        """保存源檔案（從 Gerrit 下載的源檔案） - 新增方法"""
        try:
            # 使用源檔案名並加上 gerrit_ 前綴
            source_filename = self.source_files[overwrite_type]
            gerrit_source_filename = f"gerrit_{source_filename}"
            source_path = os.path.join(output_folder, gerrit_source_filename)
            
            # 確保輸出資料夾存在
            utils.ensure_dir(output_folder)
            
            self.logger.info(f"準備保存源檔案到: {source_path}")
            self.logger.info(f"檔案內容長度: {len(content)} 字符")
            
            # 寫入檔案
            with open(source_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            # 驗證檔案是否成功保存
            if os.path.exists(source_path):
                file_size = os.path.getsize(source_path)
                self.logger.info(f"✅ Gerrit 源檔案已成功保存: {source_path}")
                self.logger.info(f"✅ 檔案大小: {file_size} bytes")
                self.logger.info(f"✅ 檔案名格式: gerrit_{source_filename}")
                
                # 再次確認檔案內容
                with open(source_path, 'r', encoding='utf-8') as f:
                    saved_content = f.read()
                    if len(saved_content) == len(content):
                        self.logger.info(f"✅ 源檔案內容驗證成功: {len(saved_content)} 字符")
                    else:
                        self.logger.warning(f"⚠️ 源檔案內容長度不匹配: 原始 {len(content)}, 保存 {len(saved_content)}")
                
                return source_path
            else:
                raise Exception(f"源檔案保存後不存在: {source_path}")
                
        except Exception as e:
            self.logger.error(f"保存源檔案失敗: {str(e)}")
            self.logger.error(f"目標路徑: {source_path}")
            self.logger.error(f"輸出資料夾: {output_folder}")
            self.logger.error(f"檔案名: {gerrit_source_filename}")
            raise
    
    def _save_converted_file(self, content: str, overwrite_type: str, output_folder: str) -> str:
        """保存轉換後的檔案 - 增強版本，確保檔案正確保存"""
        try:
            output_filename = self.output_files[overwrite_type]
            output_path = os.path.join(output_folder, output_filename)
            
            # 確保輸出資料夾存在
            utils.ensure_dir(output_folder)
            
            self.logger.info(f"準備保存轉換檔案到: {output_path}")
            self.logger.info(f"檔案內容長度: {len(content)} 字符")
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            # 驗證檔案是否成功保存
            if os.path.exists(output_path):
                file_size = os.path.getsize(output_path)
                self.logger.info(f"✅ 轉換後檔案已成功保存: {output_path}")
                self.logger.info(f"✅ 檔案大小: {file_size} bytes")
                
                # 再次確認檔案內容
                with open(output_path, 'r', encoding='utf-8') as f:
                    saved_content = f.read()
                    if len(saved_content) == len(content):
                        self.logger.info(f"✅ 檔案內容驗證成功: {len(saved_content)} 字符")
                    else:
                        self.logger.warning(f"⚠️ 檔案內容長度不匹配: 原始 {len(content)}, 保存 {len(saved_content)}")
                
                return output_path
            else:
                raise Exception(f"檔案保存後不存在: {output_path}")
            
        except Exception as e:
            self.logger.error(f"保存轉換檔案失敗: {str(e)}")
            self.logger.error(f"目標路徑: {output_path}")
            self.logger.error(f"輸出資料夾: {output_folder}")
            self.logger.error(f"檔案名: {output_filename}")
            raise
    
    def _save_target_file(self, content: str, overwrite_type: str, output_folder: str) -> str:
        """保存目標檔案（從 Gerrit 下載的） - 修正版本，確保檔案正確保存"""
        try:
            # 使用正確的目標檔名並加上 gerrit_ 前綴
            original_target_filename = self.target_files[overwrite_type]
            target_filename = f"gerrit_{original_target_filename}"
            target_path = os.path.join(output_folder, target_filename)
            
            # 確保輸出資料夾存在
            utils.ensure_dir(output_folder)
            
            self.logger.info(f"準備保存目標檔案到: {target_path}")
            self.logger.info(f"檔案內容長度: {len(content)} 字符")
            
            # 寫入檔案
            with open(target_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            # 驗證檔案是否成功保存
            if os.path.exists(target_path):
                file_size = os.path.getsize(target_path)
                self.logger.info(f"✅ Gerrit 目標檔案已成功保存: {target_path}")
                self.logger.info(f"✅ 檔案大小: {file_size} bytes")
                self.logger.info(f"✅ 檔案名格式: gerrit_{original_target_filename}")
                
                # 再次確認檔案內容
                with open(target_path, 'r', encoding='utf-8') as f:
                    saved_content = f.read()
                    if len(saved_content) == len(content):
                        self.logger.info(f"✅ 檔案內容驗證成功: {len(saved_content)} 字符")
                    else:
                        self.logger.warning(f"⚠️ 檔案內容長度不匹配: 原始 {len(content)}, 保存 {len(saved_content)}")
                
                return target_path
            else:
                raise Exception(f"檔案保存後不存在: {target_path}")
                
        except Exception as e:
            self.logger.error(f"保存目標檔案失敗: {str(e)}")
            self.logger.error(f"目標路徑: {target_path}")
            self.logger.error(f"輸出資料夾: {output_folder}")
            self.logger.error(f"檔案名: {target_filename}")
            raise
    
    def _get_source_and_target_filenames(self, overwrite_type: str) -> tuple:
        """取得來源和目標檔案名稱"""
        source_filename = self.output_files.get(overwrite_type, 'unknown.xml')
        target_filename = f"gerrit_{self.target_files.get(overwrite_type, 'unknown.xml')}"
        return source_filename, target_filename
    
    def _build_project_line_content(self, project: Dict, use_converted_revision: bool = False) -> str:
        """根據專案資訊建立完整的 project 行內容 - 徹底修正remote問題"""
        try:
            # 建立基本的 project 標籤
            project_line = "<project"
            
            # 標準屬性順序
            attrs_order = ['groups', 'name', 'path', 'revision', 'upstream', 'dest-branch', 'clone-depth', 'remote']
            
            for attr in attrs_order:
                value = project.get(attr, '')
                
                # 特殊處理 revision
                if attr == 'revision' and use_converted_revision:
                    value = project.get('converted_revision', project.get('revision', ''))
                
                # 🔥 徹底修正 remote 屬性問題
                if attr == 'remote':
                    # 檢查專案的原始資料中是否有明確的 remote 屬性
                    original_remote = project.get('original_remote', None)  # 使用原始remote
                    if original_remote is None or original_remote == '':
                        # 如果原始專案沒有remote，完全跳過
                        self.logger.debug(f"專案 {project.get('name', 'unknown')} 原始沒有remote屬性，跳過添加")
                        continue
                    value = original_remote
                
                # 只添加非空值
                if value and value.strip():
                    project_line += f' {attr}="{value}"'
            
            # 🔥 重要修改：不要自動添加 />，保持原始格式
            project_line += ">"
            
            return project_line
            
        except Exception as e:
            self.logger.error(f"建立 project 行內容失敗: {str(e)}")
            return f"<project name=\"{project.get('name', 'unknown')}\" ... >"
    
    def _analyze_differences(self, converted_content: str, target_content: Optional[str], 
                    overwrite_type: str, conversion_info: List[Dict]) -> Dict[str, Any]:
        """分析轉換檔案與目標檔案的差異 - 修正版本，基於差異頁籤重新計算統計"""
        
        # 檔案來源確認日誌
        self.logger.info(f"🔍 差異分析檔案確認:")
        self.logger.info(f"   轉換類型: {overwrite_type}")
        self.logger.info(f"   來源檔案: {self.source_files.get(overwrite_type, 'unknown')}")
        self.logger.info(f"   輸出檔案: {self.output_files.get(overwrite_type, 'unknown')}")
        self.logger.info(f"   比較目標檔案: {self.target_files.get(overwrite_type, 'unknown')}")
        self.logger.info(f"   轉換後內容長度: {len(converted_content) if converted_content else 0}")
        self.logger.info(f"   目標內容長度: {len(target_content) if target_content else 0}")
        
        analysis = {
            'has_target': target_content is not None,
            'converted_projects': conversion_info,
            'target_projects': [],
            'differences': [],
            'summary': {}
        }
        
        try:
            if target_content:
                # 解析目標檔案
                target_root = ET.fromstring(target_content)
                target_projects = self._extract_projects_with_line_numbers(target_content)
                analysis['target_projects'] = target_projects
                
                # 進行差異比較（使用修正後的邏輯）
                differences = self._compare_projects_with_conversion_info(
                    conversion_info, target_projects, overwrite_type
                )
                analysis['differences'] = differences
                
                # 🔥 修正統計摘要 - 基於差異頁籤的實際結果重新計算
                total_projects = len(conversion_info)
                converted_projects = sum(1 for proj in conversion_info if proj.get('changed', False))
                unchanged_projects = total_projects - converted_projects
                
                # 🔥 重新統計：基於差異頁籤的實際結果
                same_count = sum(1 for diff in differences if diff.get('comparison_status') == '✔️ 相同')
                different_count = sum(1 for diff in differences if diff.get('comparison_status') not in ['✔️ 相同'])
                
                self.logger.info(f"🔍 差異頁籤統計詳情:")
                self.logger.info(f"   差異頁籤總項目數: {len(differences)}")
                for status in ['✔️ 相同', '❌ 不同', '🆕 新增', '🗑️ 刪除']:
                    count = sum(1 for diff in differences if diff.get('comparison_status') == status)
                    self.logger.info(f"   {status}: {count}")
                
                analysis['summary'] = {
                    'converted_count': total_projects,  # 總專案數
                    'target_count': len(target_projects),
                    'actual_conversion_count': converted_projects,  # 實際轉換數
                    'unchanged_count': unchanged_projects,  # 未轉換數
                    'differences_count': different_count,  # 🔥 修正：差異數量（不是相同的數量）
                    'identical_converted_count': same_count,  # 🔥 修正：相同數量（✔️ 相同的數量）
                    'conversion_match_rate': f"{(same_count / max(len(differences), 1) * 100):.1f}%" if len(differences) > 0 else "N/A"
                }
                
                self.logger.info(f"差異分析完成:")
                self.logger.info(f"  📋 總專案數: {total_projects}")
                self.logger.info(f"  📄 實際轉換專案: {converted_projects}")
                self.logger.info(f"  ⭕ 未轉換專案: {unchanged_projects}")
                self.logger.info(f"  ✔️ 轉換後相同: {same_count}")
                self.logger.info(f"  ❌ 轉換後有差異: {different_count}")
                self.logger.info(f"  📊 基於差異頁籤的匹配率: {analysis['summary']['conversion_match_rate']}")
            else:
                analysis['summary'] = {
                    'converted_count': len(conversion_info),
                    'target_count': 0,
                    'actual_conversion_count': sum(1 for proj in conversion_info if proj.get('changed', False)),
                    'unchanged_count': len(conversion_info) - sum(1 for proj in conversion_info if proj.get('changed', False)),
                    'differences_count': 0,
                    'identical_converted_count': 0,
                    'conversion_match_rate': "N/A (無目標檔案)"
                }
                self.logger.info("沒有目標檔案，跳過差異比較")
            
            return analysis
            
        except Exception as e:
            self.logger.error(f"差異分析失敗: {str(e)}")
            return analysis
    
    def _extract_projects_with_line_numbers(self, xml_content: str) -> List[Dict[str, Any]]:
        """提取專案資訊並記錄行號 - 改進版本，提取完整的project行內容"""
        projects = []
        lines = xml_content.split('\n')
        
        try:
            root = ET.fromstring(xml_content)
            
            # 為每個 project 找到對應的完整行內容
            for project in root.findall('project'):
                project_name = project.get('name', '')
                
                # 在原始內容中尋找對應的行號和完整內容
                line_number, full_line = self._find_project_line_and_content(lines, project_name)
                
                project_info = {
                    'line_number': line_number,
                    'name': project.get('name', ''),
                    'path': project.get('path', ''),
                    'revision': project.get('revision', ''),
                    'upstream': project.get('upstream', ''),
                    'dest-branch': project.get('dest-branch', ''),
                    'groups': project.get('groups', ''),
                    'clone-depth': project.get('clone-depth', ''),
                    'remote': project.get('remote', ''),
                    'full_line': full_line  # 完整的project行內容
                }
                projects.append(project_info)
            
            return projects
            
        except Exception as e:
            self.logger.error(f"提取專案資訊失敗: {str(e)}")
            return []

    def _find_project_line_and_content(self, lines: List[str], project_name: str) -> tuple:
        """尋找專案在 XML 中的行號和完整內容 - 修正版本，只抓 project 標籤本身"""
        line_number = 0
        full_content = ""
        
        try:
            import re
            
            for i, line in enumerate(lines, 1):
                stripped_line = line.strip()
                
                # 檢查是否包含該專案名稱
                if f'name="{project_name}"' in line:
                    line_number = i
                    
                    # 🆕 使用正則表達式只抓取 project 標籤本身
                    if stripped_line.startswith('<project') and stripped_line.endswith('/>'):
                        # 單行 project 標籤
                        full_content = stripped_line
                    elif stripped_line.startswith('<project'):
                        # 多行 project 標籤，需要找到結束位置
                        if '/>' in stripped_line:
                            # project 標籤在同一行結束
                            project_match = re.search(r'<project[^>]*/?>', stripped_line)
                            if project_match:
                                full_content = project_match.group(0)
                        else:
                            # project 標籤跨多行，找到 > 結束
                            project_content = stripped_line
                            for j in range(i, len(lines)):
                                next_line = lines[j].strip()
                                if j > i - 1:  # 不重複第一行
                                    project_content += " " + next_line
                                
                                if next_line.endswith('>'):
                                    # 🆕 使用正則表達式提取完整的 project 標籤
                                    project_match = re.search(r'<project[^>]*>', project_content)
                                    if project_match:
                                        full_content = project_match.group(0)
                                    break
                    else:
                        # 如果不是以<project開始，往前找
                        for k in range(i-2, -1, -1):
                            prev_line = lines[k].strip()
                            if prev_line.startswith('<project'):
                                # 組合完整內容，然後用正則表達式提取
                                combined_content = prev_line
                                for m in range(k+1, i):
                                    combined_content += " " + lines[m].strip()
                                combined_content += " " + stripped_line
                                
                                project_match = re.search(r'<project[^>]*/?>', combined_content)
                                if project_match:
                                    full_content = project_match.group(0)
                                    line_number = k + 1
                                break
                    
                    break
            
            # 清理多餘的空格
            full_content = ' '.join(full_content.split())
            
            self.logger.debug(f"找到專案 {project_name} 在第 {line_number} 行: {full_content[:100]}...")
            
            return line_number, full_content
            
        except Exception as e:
            self.logger.error(f"尋找專案行失敗 {project_name}: {str(e)}")
            return 0, f"<project name=\"{project_name}\" ... />"
    
    def _compare_projects_with_conversion_info(self, converted_projects: List[Dict], 
                                    target_projects: List[Dict], overwrite_type: str) -> List[Dict]:
        """使用轉換資訊比較專案差異 - 修正版本，比較所有專案並修正統計"""
        differences = []
        
        # 🔥 修改：建立目標專案的索引 - 使用 name+path 組合作為 key
        target_index = {}
        for proj in target_projects:
            name = proj['name']
            path = proj['path']
            composite_key = f"{name}|{path}"
            target_index[composite_key] = proj
        
        # 取得正確的檔案名稱
        source_file, gerrit_source_file = self._get_source_and_target_filenames(overwrite_type)
        
        # 🔥 統計計數器
        total_compared = 0
        same_count = 0
        different_count = 0
        new_count = 0
        
        for conv_proj in converted_projects:
            project_name = conv_proj['name']
            project_path = conv_proj['path']
            conv_composite_key = f"{project_name}|{project_path}"
            has_conversion = conv_proj.get('changed', False)
            
            # 🔥 移除轉換過濾條件 - 比較所有專案，不管是否有轉換
            total_compared += 1
            
            # 檢查專案是否在目標中存在
            if conv_composite_key not in target_index:
                # 專案在轉換後存在，但在 Gerrit 中不存在 - 新增
                new_count += 1
                difference = {
                    'SN': len(differences) + 1,
                    'source_file': source_file,
                    'content': conv_proj.get('content', self._build_project_line_content(conv_proj, use_converted_revision=True)),
                    'name': conv_proj['name'],
                    'path': conv_proj['path'],
                    'revision': conv_proj['converted_revision'],
                    'upstream': conv_proj['upstream'],
                    'dest-branch': conv_proj['dest-branch'],
                    'groups': conv_proj['groups'],
                    'clone-depth': conv_proj['clone-depth'],
                    'remote': conv_proj['remote'],
                    'source_link': self._generate_source_link(conv_proj['name'], conv_proj['converted_revision'], conv_proj['remote']),
                    'gerrit_source_file': gerrit_source_file,
                    'gerrit_content': 'N/A (專案不存在)',
                    'gerrit_name': 'N/A',
                    'gerrit_path': 'N/A',
                    'gerrit_revision': 'N/A',
                    'gerrit_upstream': 'N/A',
                    'gerrit_dest-branch': 'N/A',
                    'gerrit_groups': 'N/A',
                    'gerrit_clone-depth': 'N/A',
                    'gerrit_remote': 'N/A',
                    'gerrit_source_link': 'N/A',
                    'comparison_status': '🆕 新增',
                    'comparison_result': '僅存在於轉換後',
                    'status_color': 'yellow'
                }
                differences.append(difference)
                continue
            
            # 使用 composite key 取得目標專案
            target_proj = target_index[conv_composite_key]
            
            # 🔥 修正比較邏輯：取得詳細差異
            diff_details = self._get_detailed_differences_between_projects(conv_proj, target_proj, use_converted_revision=True)
            is_identical = len(diff_details) == 0

            # 判斷比較狀態並計數
            if is_identical:
                comparison_status = '✔️ 相同'
                comparison_result = '轉換後與 Gerrit 完全一致'
                status_color = 'green'
                same_count += 1
            else:
                comparison_status = '❌ 不同'
                # 🔥 使用詳細的差異摘要
                comparison_result = self._format_difference_summary(diff_details)
                status_color = 'red'
                different_count += 1
            
            # 記錄所有比較結果（包含相同的）
            difference = {
                'SN': len(differences) + 1,
                'source_file': source_file,
                'content': conv_proj.get('content', self._build_project_line_content(conv_proj, use_converted_revision=True)),
                'name': conv_proj['name'],
                'path': conv_proj['path'],
                'revision': conv_proj['converted_revision'],
                'upstream': conv_proj['upstream'],
                'dest-branch': conv_proj['dest-branch'],
                'groups': conv_proj['groups'],
                'clone-depth': conv_proj['clone-depth'],
                'remote': conv_proj['remote'],
                'source_link': self._generate_source_link(conv_proj['name'], conv_proj['converted_revision'], conv_proj['remote']),
                'gerrit_source_file': gerrit_source_file,
                'gerrit_content': target_proj.get('full_line', target_proj['full_line']),
                'gerrit_name': target_proj['name'],
                'gerrit_path': target_proj['path'],
                'gerrit_revision': target_proj['revision'],
                'gerrit_upstream': target_proj['upstream'],
                'gerrit_dest-branch': target_proj['dest-branch'],
                'gerrit_groups': target_proj['groups'],
                'gerrit_clone-depth': target_proj['clone-depth'],
                'gerrit_remote': target_proj['remote'],
                'gerrit_source_link': self._generate_source_link(target_proj['name'], target_proj['revision'], target_proj['remote']),
                'comparison_status': comparison_status,
                'comparison_result': comparison_result,
                'status_color': status_color
            }
            differences.append(difference)
        
        # 檢查 Gerrit 中存在但轉換後不存在的專案（刪除）
        converted_composite_keys = set()
        for proj in converted_projects:
            composite_key = f"{proj['name']}|{proj['path']}"
            converted_composite_keys.add(composite_key)

        deleted_count = 0
        for composite_key, target_proj in target_index.items():
            if composite_key not in converted_composite_keys:
                deleted_count += 1
                difference = {
                    'SN': len(differences) + 1,
                    'source_file': source_file,
                    'content': 'N/A (專案已刪除)',
                    'name': target_proj['name'],
                    'path': target_proj['path'],
                    'revision': 'N/A',
                    'upstream': 'N/A',
                    'dest-branch': 'N/A',
                    'groups': 'N/A',
                    'clone-depth': 'N/A',
                    'remote': 'N/A',
                    'source_link': 'N/A',
                    'gerrit_source_file': gerrit_source_file,
                    'gerrit_content': target_proj.get('full_line', target_proj['full_line']),
                    'gerrit_name': target_proj['name'],
                    'gerrit_path': target_proj['path'],
                    'gerrit_revision': target_proj['revision'],
                    'gerrit_upstream': target_proj['upstream'],
                    'gerrit_dest-branch': target_proj['dest-branch'],
                    'gerrit_groups': target_proj['groups'],
                    'gerrit_clone-depth': target_proj['clone-depth'],
                    'gerrit_remote': target_proj['remote'],
                    'gerrit_source_link': self._generate_source_link(target_proj['name'], target_proj['revision'], target_proj['remote']),
                    'comparison_status': '🗑️ 刪除',
                    'comparison_result': '僅存在於 Gerrit',
                    'status_color': 'orange'
                }
                differences.append(difference)
        
        # 🔥 添加詳細統計日誌
        self.logger.info(f"🔍 差異比較詳細統計:")
        self.logger.info(f"   總比較專案數: {total_compared}")
        self.logger.info(f"   ✔️ 相同: {same_count}")
        self.logger.info(f"   ❌ 不同: {different_count}")
        self.logger.info(f"   🆕 新增: {new_count}")
        self.logger.info(f"   🗑️ 刪除: {deleted_count}")
        self.logger.info(f"   📋 差異頁籤總項目: {len(differences)}")
        
        return differences
    
    def _get_detailed_differences_between_projects(self, source_proj: Dict, target_proj: Dict, use_converted_revision: bool = True) -> List[Dict]:
        """
        取得兩個專案之間的詳細差異列表
        
        Args:
            source_proj: 來源專案資訊
            target_proj: 目標專案資訊
            use_converted_revision: 是否使用轉換後的 revision
            
        Returns:
            差異詳情列表
        """
        differences = []
        
        try:
            # 要比較的屬性列表
            attrs_to_compare = ['name', 'path', 'revision', 'upstream', 'dest-branch', 'groups', 'clone-depth', 'remote']
            
            # 逐一比較每個屬性
            for attr in attrs_to_compare:
                if attr == 'revision' and use_converted_revision:
                    source_val = source_proj.get('converted_revision', '').strip()
                else:
                    source_val = source_proj.get(attr, '').strip()
                
                target_val = target_proj.get(attr, '').strip()
                
                # 如果不同，記錄差異
                if source_val != target_val:
                    diff_info = {
                        'attribute': attr,
                        'source_value': source_val,
                        'target_value': target_val
                    }
                    differences.append(diff_info)
            
            return differences
            
        except Exception as e:
            self.logger.error(f"取得專案間詳細差異失敗: {str(e)}")
            return []

    def _format_difference_summary(self, diff_details: List[Dict]) -> str:
        """
        格式化差異摘要（與 manifest_conversion.py 完全相同的邏輯）
        
        Args:
            diff_details: 差異詳情列表
            
        Returns:
            格式化的差異摘要字串
        """
        try:
            if not diff_details:
                return "無差異"
            
            # 按屬性重要性排序
            attr_priority = {'revision': 1, 'name': 2, 'path': 3, 'upstream': 4, 'dest-branch': 5, 
                            'groups': 6, 'clone-depth': 7, 'remote': 8}
            
            diff_details.sort(key=lambda x: attr_priority.get(x['attribute'], 99))
            
            # 格式化差異說明
            diff_parts = []
            for diff in diff_details[:3]:  # 最多顯示前3個差異
                attr = diff['attribute']
                source_val = diff['source_value'] or '(空)'
                target_val = diff['target_value'] or '(空)'
                
                # 特殊處理不同屬性的顯示
                if attr == 'revision':
                    diff_parts.append(f"版本號[{source_val} ≠ {target_val}]")
                elif attr == 'upstream':
                    diff_parts.append(f"上游分支[{source_val} ≠ {target_val}]")
                elif attr == 'dest-branch':
                    diff_parts.append(f"目標分支[{source_val} ≠ {target_val}]")
                elif attr == 'groups':
                    diff_parts.append(f"群組[{source_val} ≠ {target_val}]")
                elif attr == 'clone-depth':
                    diff_parts.append(f"克隆深度[{source_val} ≠ {target_val}]")
                elif attr == 'remote':
                    diff_parts.append(f"遠端[{source_val} ≠ {target_val}]")
                else:
                    diff_parts.append(f"{attr}[{source_val} ≠ {target_val}]")
            
            # 如果差異超過3個，加上省略號
            if len(diff_details) > 3:
                diff_parts.append(f"等{len(diff_details)}項差異")
            
            return "、".join(diff_parts)
            
        except Exception as e:
            self.logger.error(f"格式化差異摘要失敗: {str(e)}")
            return "差異格式化失敗"

    def _generate_source_link(self, project_name: str, revision: str, remote: str = '') -> str:
        """
        根據專案名稱、revision 和 remote 生成 gerrit source link - 新版本
        
        Args:
            project_name: 專案名稱
            revision: revision 字串
            remote: remote 名稱
            
        Returns:
            完整的 gerrit 連結 URL
        """
        try:
            if not project_name:
                return 'N/A'
            
            # 🆕 如果 revision 為空，使用 default revision
            if not revision or revision.strip() == '':
                revision = getattr(self, 'default_revision', '')
                if not revision:
                    self.logger.warning(f"專案 {project_name} 沒有 revision 且無法取得 default revision")
                    return 'N/A'
                self.logger.debug(f"專案 {project_name} 使用 default revision: {revision}")
            
            revision = revision.strip()
            
            # 🆕 根據 remote 決定 base URL
            if remote == 'rtk-prebuilt':
                base_url = "https://mm2sd-git2.rtkbf.com/gerrit/plugins/gitiles"
            else:  # rtk 或空值或其他
                base_url = "https://mm2sd.rtkbf.com/gerrit/plugins/gitiles"
            
            # 🆕 判斷 revision 類型並生成相應的連結
            if self._is_revision_hash(revision):
                # Hash 格式：直接使用 hash
                link = f"{base_url}/{project_name}/+/{revision}"
                self.logger.debug(f"生成 hash 連結: {project_name} → {link}")
                
            elif revision.startswith('refs/tags/'):
                # Tag 格式：直接使用完整的 refs/tags/xxx
                link = f"{base_url}/{project_name}/+/{revision}"
                self.logger.debug(f"生成 tag 連結: {project_name} → {link}")
                
            elif revision.startswith('refs/heads/'):
                # Branch 格式：直接使用完整的 refs/heads/xxx
                link = f"{base_url}/{project_name}/+/{revision}"
                self.logger.debug(f"生成 branch 連結（完整路徑）: {project_name} → {link}")
                
            else:
                # 🆕 其他情況：判斷是否為 branch name，加上 refs/heads/ 前綴
                if '/' in revision and not revision.startswith('refs/'):
                    # 看起來像是 branch path，加上 refs/heads/
                    full_branch_path = f"refs/heads/{revision}"
                    link = f"{base_url}/{project_name}/+/{full_branch_path}"
                    self.logger.debug(f"生成 branch 連結（補充前綴）: {project_name} → {link}")
                else:
                    # 無法確定類型，嘗試當作 branch 處理
                    link = f"{base_url}/{project_name}/+/refs/heads/{revision}"
                    self.logger.warning(f"無法確定 revision 類型，當作 branch 處理: {project_name} - {revision}")
            
            return link
            
        except Exception as e:
            self.logger.error(f"生成 source link 失敗: {project_name} - {revision} - {str(e)}")
            return 'N/A'
                    
    def _push_to_gerrit(self, overwrite_type: str, converted_content: str, 
                       target_content: Optional[str], output_folder: str) -> Dict[str, Any]:
        """推送轉換後的檔案到 Gerrit 服務器"""
        push_result = {
            'success': False,
            'message': '',
            'need_push': False,
            'commit_id': '',
            'review_url': ''
        }
        
        try:
            self.logger.info("🚀 開始推送到 Gerrit 服務器...")
            
            # 判斷是否需要推送
            push_result['need_push'] = self._should_push_to_gerrit(converted_content, target_content)
            
            if not push_result['need_push']:
                push_result['success'] = True
                push_result['message'] = "目標檔案與轉換結果相同，無需推送"
                self.logger.info("✅ " + push_result['message'])
                return push_result
            
            # 執行 Git 操作
            git_result = self._execute_git_push(overwrite_type, converted_content, output_folder)
            
            push_result.update(git_result)
            
            if push_result['success']:
                self.logger.info(f"✅ 成功推送到 Gerrit: {push_result['review_url']}")
            else:
                self.logger.error(f"❌ 推送失敗: {push_result['message']}")
            
            return push_result
            
        except Exception as e:
            push_result['message'] = f"推送過程發生錯誤: {str(e)}"
            self.logger.error(push_result['message'])
            return push_result
    
    def _should_push_to_gerrit(self, converted_content: str, target_content: Optional[str]) -> bool:
        """判斷是否需要推送到 Gerrit"""
        if target_content is None:
            self.logger.info("🎯 目標檔案不存在，需要推送新檔案")
            return True
        
        # 簡單的內容比較（忽略空白差異）
        def normalize_content(content):
            return ''.join(content.split()) if content else ''
        
        converted_normalized = normalize_content(converted_content)
        target_normalized = normalize_content(target_content)
        
        if converted_normalized == target_normalized:
            self.logger.info("📋 轉換結果與目標檔案相同，無需推送")
            return False
        else:
            self.logger.info("📄 轉換結果與目標檔案不同，需要推送更新")
            return True
    
    def _execute_git_push(self, overwrite_type: str, converted_content: str, output_folder: str) -> Dict[str, Any]:
        """執行 Git clone, commit, push 操作 - 使用動態 Android 版本"""
        import subprocess
        import tempfile
        import shutil
        
        result = {
            'success': False,
            'message': '',
            'commit_id': '',
            'review_url': ''
        }
        
        # 建立臨時 Git 工作目錄
        temp_git_dir = None
        
        try:
            temp_git_dir = tempfile.mkdtemp(prefix='gerrit_push_')
            self.logger.info(f"📁 建立臨時 Git 目錄: {temp_git_dir}")
            
            # 🔥 修改：Git 設定 - 使用動態版本
            repo_url = "ssh://mm2sd.rtkbf.com:29418/realtek/android/manifest"
            branch = config.get_default_android_master_branch()  # 使用動態分支
            target_filename = self.output_files[overwrite_type]
            source_filename = self.source_files[overwrite_type]
            
            # 步驟 1: Clone repository
            self.logger.info(f"📥 Clone repository: {repo_url}")
            clone_cmd = ["git", "clone", "-b", branch, repo_url, temp_git_dir]
            
            subprocess.run(clone_cmd, check=True, capture_output=True, text=True, timeout=60)
            
            # 步驟 2: 切換到 Git 目錄並設定環境
            original_cwd = os.getcwd()
            os.chdir(temp_git_dir)
            
            # 步驟 2.5: 安裝 commit-msg hook
            self.logger.info(f"🔧 安裝 commit-msg hook...")
            try:
                hook_install_cmd = "gitdir=$(git rev-parse --git-dir); scp -p -P 29418 vince_lin@mm2sd.rtkbf.com:hooks/commit-msg ${gitdir}/hooks/"
                
                result_hook = subprocess.run(
                    hook_install_cmd,
                    shell=True,
                    capture_output=True,
                    text=True,
                    timeout=30
                )
                
                if result_hook.returncode == 0:
                    self.logger.info(f"✅ commit-msg hook 安裝成功")
                    
                    # 驗證 hook 是否存在
                    gitdir_result = subprocess.run(["git", "rev-parse", "--git-dir"], 
                                                capture_output=True, text=True, check=True)
                    git_dir = gitdir_result.stdout.strip()
                    hook_path = os.path.join(git_dir, "hooks", "commit-msg")
                    
                    if os.path.exists(hook_path):
                        self.logger.info(f"✅ hook 檔案確認存在: {hook_path}")
                        
                        # 確保 hook 有執行權限
                        import stat
                        os.chmod(hook_path, stat.S_IRWXU | stat.S_IRGRP | stat.S_IROTH)
                        self.logger.info(f"✅ hook 執行權限設定完成")
                    else:
                        self.logger.warning(f"⚠️ hook 檔案不存在: {hook_path}")
                        
                else:
                    self.logger.warning(f"⚠️ commit-msg hook 安裝失敗: {result_hook.stderr}")
                    
            except Exception as e:
                self.logger.warning(f"⚠️ 安裝 commit-msg hook 異常: {str(e)}")
            
            # 步驟 3: 寫入轉換後的檔案
            target_file_path = os.path.join(temp_git_dir, target_filename)
            with open(target_file_path, 'w', encoding='utf-8') as f:
                f.write(converted_content)
            
            self.logger.info(f"📁 寫入檔案: {target_filename}")
            
            # 步驟 4: 檢查 Git 狀態
            status_result = subprocess.run(
                ["git", "status", "--porcelain"], 
                capture_output=True, text=True, check=True
            )
            
            if not status_result.stdout.strip():
                result['success'] = True
                result['message'] = "檔案內容無變化，無需推送"
                self.logger.info("✅ " + result['message'])
                return result
            
            # 步驟 5: Add 檔案
            subprocess.run(["git", "add", target_filename], check=True)
            
            # 步驟 6: 生成 Commit Message（從 config.py）
            commit_message = self._generate_commit_message(
                overwrite_type=overwrite_type,
                source_file=source_filename,
                target_file=target_filename
            )
            
            self.logger.info("📝 使用 config.py 的 commit message 模板")
            self.logger.debug(f"Commit message:\n{commit_message[:200]}...")  # 只顯示前200字符
            
            # 不手動添加 Change-Id，讓 commit-msg hook 處理
            commit_result = subprocess.run(
                ["git", "commit", "-m", commit_message],
                capture_output=True, text=True, check=True
            )
            
            # 取得 commit ID
            commit_id_result = subprocess.run(
                ["git", "rev-parse", "HEAD"],
                capture_output=True, text=True, check=True
            )
            result['commit_id'] = commit_id_result.stdout.strip()[:8]
            
            self.logger.info(f"📝 創建 commit: {result['commit_id']}")
            
            # 檢查 commit message 是否包含 Change-Id（由 hook 添加）
            commit_msg_result = subprocess.run(
                ["git", "log", "-1", "--pretty=format:%B"],
                capture_output=True, text=True, check=True
            )
            
            if "Change-Id:" in commit_msg_result.stdout:
                self.logger.info(f"✅ Change-Id 已由 hook 自動添加")
            else:
                self.logger.warning(f"⚠️ commit message 中沒有 Change-Id，推送可能失敗")
            
            # 步驟 7: Push to Gerrit for review
            push_cmd = ["git", "push", "origin", f"HEAD:refs/for/{branch}"]
            
            self.logger.info(f"🚀 推送到 Gerrit: {' '.join(push_cmd)}")
            
            push_result = subprocess.run(
                push_cmd, capture_output=True, text=True, check=True, timeout=60
            )
            
            # 解析 Gerrit review URL
            output_lines = push_result.stderr.split('\n')
            review_url = ""
            for line in output_lines:
                if 'https://' in line and 'gerrit' in line:
                    import re
                    url_match = re.search(r'https://[^\s]+', line)
                    if url_match:
                        review_url = url_match.group(0)
                        break
            
            result['success'] = True
            result['message'] = f"成功推送到 Gerrit，Commit ID: {result['commit_id']}"
            result['review_url'] = review_url or f"https://mm2sd.rtkbf.com/gerrit/#/q/commit:{result['commit_id']}"
            
            self.logger.info(f"🎉 推送成功！Review URL: {result['review_url']}")
            
            return result
            
        except subprocess.TimeoutExpired:
            result['message'] = "Git 操作逾時，請檢查網路連線"
            return result
        except subprocess.CalledProcessError as e:
            error_msg = e.stderr if e.stderr else str(e)
            
            # 特別處理常見錯誤
            if "missing Change-Id" in error_msg or "invalid Change-Id" in error_msg:
                result['message'] = f"Change-Id 問題: {error_msg}"
                self.logger.error(f"Change-Id 錯誤: {error_msg}")
            elif "Permission denied" in error_msg or "publickey" in error_msg:
                result['message'] = f"SSH 認證失敗: {error_msg}"
                self.logger.error(f"SSH 認證問題: {error_msg}")
            elif "remote rejected" in error_msg:
                result['message'] = f"Gerrit 拒絕推送: {error_msg}"
                self.logger.error(f"Gerrit 拒絕: {error_msg}")
            else:
                result['message'] = f"Git 命令失敗: {error_msg}"
                
            return result
        except Exception as e:
            result['message'] = f"Git 操作異常: {str(e)}"
            return result
        finally:
            # 恢復原始工作目錄
            if 'original_cwd' in locals():
                os.chdir(original_cwd)
            
            # 清理臨時目錄
            if temp_git_dir and os.path.exists(temp_git_dir):
                try:
                    shutil.rmtree(temp_git_dir)
                    self.logger.info(f"🗑️  清理臨時目錄: {temp_git_dir}")
                except Exception as e:
                    self.logger.warning(f"清理臨時目錄失敗: {str(e)}")

    def _generate_commit_message(self, overwrite_type: str, source_file: str = None, 
                                target_file: str = None, rddb_number: str = None) -> str:
        """
        從 config.py 生成 commit message
        
        Args:
            overwrite_type: 轉換類型
            source_file: 源檔案名稱
            target_file: 目標檔案名稱
            rddb_number: RDDB 號碼（可選）
            
        Returns:
            格式化的 commit message
        """
        try:
            import config
            
            # 取得 RDDB 號碼
            if not rddb_number:
                rddb_number = getattr(config, 'DEFAULT_RDDB_NUMBER', 'RDDB-923')
            
            # 檢查是否使用詳細模板
            use_detailed = getattr(config, 'USE_DETAILED_COMMIT_MESSAGE', True)
            
            if use_detailed:
                # 使用詳細模板
                template = getattr(config, 'COMMIT_MESSAGE_TEMPLATE', '')
                
                if not template:
                    # 如果 config 中沒有模板，使用預設
                    self.logger.warning("config.py 中沒有 COMMIT_MESSAGE_TEMPLATE，使用預設模板")
                    template = self._get_default_detailed_template()
                
                # 格式化模板
                commit_message = template.format(
                    rddb_number=rddb_number,
                    overwrite_type=overwrite_type
                )
                
                self.logger.info(f"使用詳細 commit message 模板 (RDDB: {rddb_number})")
                
            else:
                # 使用簡單模板
                template = getattr(config, 'COMMIT_MESSAGE_SIMPLE', '')
                
                if not template:
                    # 如果沒有簡單模板，使用預設
                    template = """Auto-generated manifest update: {overwrite_type}

    Generated by manifest conversion tool
    Source: {source_file}
    Target: {target_file}"""
                
                # 格式化模板
                commit_message = template.format(
                    overwrite_type=overwrite_type,
                    source_file=source_file or self.source_files.get(overwrite_type, 'unknown'),
                    target_file=target_file or self.output_files.get(overwrite_type, 'unknown')
                )
                
                self.logger.info("使用簡單 commit message 模板")
            
            return commit_message
            
        except Exception as e:
            self.logger.error(f"生成 commit message 失敗: {str(e)}")
            
            # 返回預設的簡單訊息
            default_message = f"""Auto-generated manifest update: {overwrite_type}

    Generated by manifest conversion tool
    Source: {source_file or self.source_files.get(overwrite_type, 'unknown')}
    Target: {target_file or self.output_files.get(overwrite_type, 'unknown')}"""
            
            self.logger.warning("使用預設 commit message")
            return default_message

    def _get_default_detailed_template(self) -> str:
        """取得預設的詳細 commit message 模板"""
        return """[{rddb_number}][Manifest][Feat]: manifest update: {overwrite_type}
    
    * Problem: for google request
    * Root Cause: change branches every three months
    * Solution: manifest update: {overwrite_type}
    * Dependency: N
    * Testing Performed:
    - ac/dc on/off: ok
    * RD-CodeSync: Y
    * PL-CodeSync: N
    - Mac7p: N
    - Mac8q: N
    - Mac9q: N
    - Merlin5: N
    - Merlin7: N
    - Merlin8: N
    - Merlin8p: N
    - Merlin9: N
    - Dante: N
    - Dora: N
    - MatriX: N
    - Midas: N
    - AN11: N
    - AN12: N
    - AN14: N
    - AN16: N
    - R2U: N
    - Kernel_Only: N
    * Supplementary:
    
    * Issue:
    - {rddb_number}"""
        
    def _generate_excel_report(self, overwrite_type: str, source_file_path: Optional[str],
                    output_file_path: Optional[str], target_file_path: Optional[str], 
                    diff_analysis: Dict, output_folder: str, 
                    excel_filename: Optional[str], source_download_success: bool,
                    target_download_success: bool, push_result: Optional[Dict[str, Any]] = None,
                    expanded_file_path: Optional[str] = None, use_expanded: bool = False) -> str:
        """產生 Excel 報告 - 修正版本，新的頁籤順序和底色"""
        try:
            if excel_filename:
                excel_file = os.path.join(output_folder, excel_filename)
            else:
                default_name = f"{overwrite_type}_conversion_report.xlsx"
                excel_file = os.path.join(output_folder, default_name)
            
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                # 頁籤 1: 轉換摘要（淺藍色底色）
                summary_data = [{
                    'SN': 1,
                    '轉換類型': overwrite_type,
                    'Gerrit 源檔案': os.path.basename(source_file_path) if source_file_path else '無',
                    '源檔案下載狀態': '成功' if source_download_success else '失敗',
                    '源檔案': self.source_files.get(overwrite_type, ''),  # 這個會被轉為連結
                    '包含 include 標籤': '是' if use_expanded else '否',
                    'Gerrit 展開檔案': os.path.basename(expanded_file_path) if expanded_file_path else '無',
                    '使用展開檔案轉換': '是' if use_expanded else '否',
                    '輸出檔案': os.path.basename(output_file_path) if output_file_path else '',
                    'Gerrit 目標檔案': os.path.basename(target_file_path) if target_file_path else '無',
                    '目標檔案下載狀態': '成功' if target_download_success else '失敗 (檔案不存在)',
                    '目標檔案': self.target_files.get(overwrite_type, ''),  # 這個會被轉為連結
                    '📊 總專案數': diff_analysis['summary'].get('converted_count', 0),
                    '🔄 實際轉換專案數': diff_analysis['summary'].get('actual_conversion_count', 0),
                    '⭕ 未轉換專案數': diff_analysis['summary'].get('unchanged_count', 0),
                    '🎯 目標檔案專案數': diff_analysis['summary'].get('target_count', 0),
                    '❌ 轉換後與 Gerrit Manifest 差異數': diff_analysis['summary'].get('differences_count', 0),
                    '✔️ 轉換後與 Gerrit Manifest 相同數': diff_analysis['summary'].get('identical_converted_count', 0)
                }]

                if push_result:
                    summary_data[0].update({
                        '推送狀態': '成功' if push_result['success'] else '失敗',
                        '推送結果': push_result['message'],
                        'Commit ID': push_result.get('commit_id', ''),
                        'Review URL': push_result.get('review_url', '')
                    })
                else:
                    summary_data[0].update({
                        '推送狀態': '未執行',
                        '推送結果': '未執行推送',
                        'Commit ID': '',
                        'Review URL': ''
                    })

                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='轉換摘要', index=False)

                # 頁籤 2: 轉換後專案（淺藍色底色）
                if diff_analysis['converted_projects']:
                    converted_data = []
                    for i, proj in enumerate(diff_analysis['converted_projects'], 1):
                        has_conversion = proj.get('changed', False)
                        if has_conversion:
                            conversion_status = '🔄 已轉換'
                            status_description = f"{proj['original_revision']} → {proj['converted_revision']}"
                        else:
                            conversion_status = '⭕ 未轉換'
                            status_description = f"保持原值: {proj['original_revision']}"
                        
                        converted_data.append({
                            'SN': i,
                            '專案名稱': proj['name'],
                            '專案路徑': proj['path'],
                            '轉換狀態': conversion_status,
                            '來源檔案': f"gerrit_{self.source_files.get(overwrite_type, 'unknown.xml')}",
                            '原始 Revision': proj['original_revision'],
                            '轉換後檔案': self.output_files.get(overwrite_type, 'unknown.xml'),
                            '轉換後 Revision': proj['converted_revision'],
                            'Revision 是否相等': '',
                            '轉換說明': status_description,
                            'Upstream': proj['upstream'],
                            'Dest-Branch': proj['dest-branch'],
                            'Groups': proj['groups'],
                            'Clone-Depth': proj['clone-depth'],
                            'Remote': proj['remote']
                        })
                    
                    df_converted = pd.DataFrame(converted_data)
                    df_converted.to_excel(writer, sheet_name='轉換後專案', index=False)
                
                # 頁籤 3: 轉換後與 Gerrit manifest 的差異（淺紅色底色）
                if diff_analysis['has_target'] and diff_analysis['differences']:
                    diff_sheet_name = "轉換後與 Gerrit manifest 的差異"
                    df_diff = pd.DataFrame(diff_analysis['differences'])
                    
                    diff_columns = [
                        'SN', 'comparison_status', 'comparison_result',
                        'source_file', 'content', 'name', 'path', 
                        'revision',
                        'upstream', 'dest-branch', 'groups', 'clone-depth', 'remote', 'source_link',
                        'gerrit_source_file', 'gerrit_content', 'gerrit_name', 
                        'gerrit_path', 'gerrit_revision', 'gerrit_upstream', 
                        'gerrit_dest-branch', 'gerrit_groups', 'gerrit_clone-depth', 'gerrit_remote', 'gerrit_source_link'
                    ]
                    
                    available_columns = [col for col in diff_columns if col in df_diff.columns]
                    df_diff = df_diff[available_columns]
                    
                    df_diff.to_excel(writer, sheet_name=diff_sheet_name, index=False)
                
                # 🔥 頁籤 4: 來源 gerrit manifest（修改名稱）
                if diff_analysis['converted_projects']:
                    source_data = []
                    for i, proj in enumerate(diff_analysis['converted_projects'], 1):
                        source_link = self._generate_source_link(proj['name'], proj['original_revision'], proj['remote'])
                        gerrit_source_filename = f"gerrit_{self.source_files.get(overwrite_type, 'unknown.xml')}"
                        
                        source_data.append({
                            'SN': i,
                            'source_file': gerrit_source_filename,
                            'name': proj['name'],
                            'path': proj['path'],
                            'revision': proj['original_revision'],
                            'upstream': proj['upstream'],
                            'dest-branch': proj['dest-branch'],
                            'groups': proj['groups'],
                            'clone-depth': proj['clone-depth'],
                            'remote': proj['remote'],
                            'source_link': source_link
                        })
                    
                    df_source = pd.DataFrame(source_data)
                    df_source.to_excel(writer, sheet_name='來源 gerrit manifest', index=False)  # 🔥 修改名稱
                
                # 頁籤 5: 轉換後的 manifest（淺綠色底色）
                if diff_analysis['converted_projects']:
                    converted_manifest_data = []
                    for i, proj in enumerate(diff_analysis['converted_projects'], 1):
                        source_link = self._generate_source_link(proj['name'], proj['converted_revision'], proj['remote'])
                        output_filename = self.output_files.get(overwrite_type, 'unknown.xml')
                        
                        converted_manifest_data.append({
                            'SN': i,
                            'source_file': output_filename,
                            'name': proj['name'],
                            'path': proj['path'],
                            'revision': proj['converted_revision'],
                            'upstream': proj['upstream'],
                            'dest-branch': proj['dest-branch'],
                            'groups': proj['groups'],
                            'clone-depth': proj['clone-depth'],
                            'remote': proj['remote'],
                            'source_link': source_link
                        })
                    
                    df_converted_manifest = pd.DataFrame(converted_manifest_data)
                    df_converted_manifest.to_excel(writer, sheet_name='轉換後的 manifest', index=False)
                
                # 🔥 頁籤 6: 目的 gerrit manifest（修改名稱）
                if diff_analysis['has_target'] and diff_analysis['target_projects']:
                    gerrit_data = []
                    for i, proj in enumerate(diff_analysis['target_projects'], 1):
                        source_link = self._generate_source_link(proj['name'], proj['revision'], proj['remote'])
                        gerrit_target_filename = f"gerrit_{self.target_files.get(overwrite_type, 'unknown.xml')}"
                        
                        gerrit_data.append({
                            'SN': i,
                            'source_file': gerrit_target_filename,
                            'name': proj['name'],
                            'path': proj['path'],
                            'revision': proj['revision'],
                            'upstream': proj['upstream'],
                            'dest-branch': proj['dest-branch'],
                            'groups': proj['groups'],
                            'clone-depth': proj['clone-depth'],
                            'remote': proj['remote'],
                            'source_link': source_link
                        })
                    
                    df_gerrit = pd.DataFrame(gerrit_data)
                    df_gerrit.to_excel(writer, sheet_name='目的 gerrit manifest', index=False)  # 🔥 修改名稱
                
                # 🔥 所有格式化都在 ExcelWriter context 內完成
                workbook = writer.book
                
                # 格式化所有工作表
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # 應用格式化
                    self._format_worksheet_with_background_colors(worksheet, sheet_name)
                    
                    # 為相關頁籤添加超連結
                    if sheet_name in ['來源 gerrit manifest', '轉換後的 manifest', '目的 gerrit manifest', '轉換後與 Gerrit manifest 的差異']:
                        self._add_manifest_hyperlinks(worksheet, sheet_name)
                    
                    # 為轉換後專案頁籤添加特殊的 Gerrit 連結
                    if sheet_name == '轉換後專案':
                        self._add_converted_projects_hyperlinks(worksheet, overwrite_type)
                        self._add_revision_comparison_formula_converted_projects(worksheet)
                    
                    # 為轉換摘要頁籤添加超連結
                    if sheet_name == '轉換摘要':
                        self._add_summary_hyperlinks(worksheet, overwrite_type)
                        self._format_summary_content_backgrounds(worksheet)
                
                # 自動調整所有頁籤的欄寬
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    self._auto_adjust_column_widths_enhanced(worksheet, sheet_name)
            
            self.logger.info(f"成功產生 Excel 報告: {excel_file}")
            return excel_file
            
        except Exception as e:
            self.logger.error(f"產生 Excel 報告失敗: {str(e)}")
            raise

    def _add_converted_projects_hyperlinks(self, worksheet, overwrite_type: str):
        """
        為轉換後專案頁籤添加來源檔案的 Gerrit 超連結
        
        Args:
            worksheet: Excel 工作表
            overwrite_type: 轉換類型
        """
        try:
            # 找到來源檔案欄位的位置
            source_file_col = None
            
            for col_num, cell in enumerate(worksheet[1], 1):  # 表頭行
                header_value = str(cell.value) if cell.value else ''
                if header_value == '來源檔案':
                    source_file_col = col_num
                    break
            
            # 為來源檔案欄位添加連結
            if source_file_col:
                source_filename = self.source_files.get(overwrite_type, '')
                if source_filename:
                    gerrit_url = self._generate_gerrit_manifest_link(source_filename)
                    display_text = f"gerrit_{source_filename}"
                    
                    # 在數據行添加超連結（從第2行開始）
                    for row_num in range(2, worksheet.max_row + 1):
                        self._add_hyperlink_to_cell(worksheet, row_num, source_file_col, gerrit_url, display_text)
                    
                    self.logger.info(f"已為轉換後專案添加來源檔案連結: {display_text}")
            
        except Exception as e:
            self.logger.error(f"添加轉換後專案超連結失敗: {str(e)}")

    def _add_manifest_hyperlinks(self, worksheet, sheet_name: str):
        """
        為 manifest 相關頁籤添加 source_link 欄位的正確 gerrit 連結
        
        Args:
            worksheet: Excel 工作表
            sheet_name: 頁籤名稱
        """
        try:
            # 找到需要處理的欄位
            source_file_col = None
            gerrit_source_file_col = None
            source_link_col = None
            gerrit_source_link_col = None
            
            for col_num, cell in enumerate(worksheet[1], 1):  # 表頭行
                header_value = str(cell.value) if cell.value else ''
                
                if header_value == 'source_file':
                    source_file_col = col_num
                elif header_value == 'gerrit_source_file':
                    gerrit_source_file_col = col_num
                elif header_value == 'source_link':
                    source_link_col = col_num
                elif header_value == 'gerrit_source_link':
                    gerrit_source_link_col = col_num
            
            # 🔥 只有特定頁籤的 source_file 欄位需要添加連結
            source_file_need_link = sheet_name in ['來源 gerrit manifest', '目的 gerrit manifest']
            
            # 為 source_file 欄位添加連結（僅限指定頁籤）
            if source_file_col and source_file_need_link:
                for row_num in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_num, column=source_file_col)
                    filename = str(cell.value) if cell.value else ''
                    
                    if filename and filename not in ['', 'N/A']:
                        gerrit_url = self._generate_gerrit_manifest_link(filename)
                        self._add_hyperlink_formula_to_cell(worksheet, row_num, source_file_col, gerrit_url, filename)
            
            # 為 gerrit_source_file 欄位添加連結
            if gerrit_source_file_col:
                for row_num in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_num, column=gerrit_source_file_col)
                    filename = str(cell.value) if cell.value else ''
                    
                    if filename and filename not in ['', 'N/A']:
                        gerrit_url = self._generate_gerrit_manifest_link(filename)
                        self._add_hyperlink_formula_to_cell(worksheet, row_num, gerrit_source_file_col, gerrit_url, filename)
            
            # 🆕 為 source_link 欄位添加正確的專案連結（重點修改）
            if source_link_col:
                for row_num in range(2, worksheet.max_row + 1):
                    # 取得該行的專案資訊
                    name_cell = self._find_cell_value_in_row(worksheet, row_num, ['name'])
                    revision_cell = self._find_cell_value_in_row(worksheet, row_num, ['revision'])
                    remote_cell = self._find_cell_value_in_row(worksheet, row_num, ['remote'])
                    
                    if name_cell:
                        project_name = str(name_cell)
                        revision = str(revision_cell) if revision_cell else ''
                        remote = str(remote_cell) if remote_cell else ''
                        
                        # 🆕 使用新的生成邏輯
                        gerrit_project_url = self._generate_source_link(project_name, revision, remote)
                        
                        if gerrit_project_url and gerrit_project_url != 'N/A':
                            # 🔥 使用 HYPERLINK 格式，顯示連結本身
                            self._add_hyperlink_formula_to_cell(worksheet, row_num, source_link_col, gerrit_project_url, gerrit_project_url)
            
            # 🆕 為 gerrit_source_link 欄位添加正確的專案連結（重點修改）
            if gerrit_source_link_col:
                for row_num in range(2, worksheet.max_row + 1):
                    # 取得該行的 Gerrit 專案資訊
                    gerrit_name_cell = self._find_cell_value_in_row(worksheet, row_num, ['gerrit_name'])
                    gerrit_revision_cell = self._find_cell_value_in_row(worksheet, row_num, ['gerrit_revision'])
                    gerrit_remote_cell = self._find_cell_value_in_row(worksheet, row_num, ['gerrit_remote'])
                    
                    if gerrit_name_cell:
                        project_name = str(gerrit_name_cell)
                        revision = str(gerrit_revision_cell) if gerrit_revision_cell else ''
                        remote = str(gerrit_remote_cell) if gerrit_remote_cell else ''
                        
                        # 🆕 使用新的生成邏輯
                        gerrit_project_url = self._generate_source_link(project_name, revision, remote)
                        
                        if gerrit_project_url and gerrit_project_url != 'N/A':
                            # 🔥 使用 HYPERLINK 格式
                            self._add_hyperlink_formula_to_cell(worksheet, row_num, gerrit_source_link_col, gerrit_project_url, gerrit_project_url)
            
            # 記錄處理結果
            if source_file_col and source_file_need_link:
                self.logger.info(f"✅ 已為 {sheet_name} 添加 source_file 欄位連結")
            
            if gerrit_source_file_col:
                self.logger.info(f"✅ 已為 {sheet_name} 添加 gerrit_source_file 欄位連結")
            
            if source_link_col:
                self.logger.info(f"✅ 已為 {sheet_name} 添加 source_link 欄位正確的專案連結")
            
            if gerrit_source_link_col:
                self.logger.info(f"✅ 已為 {sheet_name} 添加 gerrit_source_link 欄位正確的專案連結")
            
        except Exception as e:
            self.logger.error(f"添加 {sheet_name} 超連結失敗: {str(e)}")

    def _add_hyperlink_formula_to_cell(self, worksheet, row: int, col: int, url: str, display_text: str = None):
        """
        為 Excel 單元格添加 HYPERLINK 函數格式的超連結 - 改進版本
        
        Args:
            worksheet: Excel 工作表
            row: 行號
            col: 列號  
            url: 連結 URL
            display_text: 顯示文字（可選）
        """
        try:
            from openpyxl.styles import Font
            
            cell = worksheet.cell(row=row, column=col)
            
            # 🆕 清理 URL 中的特殊字符，避免 Excel 公式錯誤
            clean_url = url.replace('"', '""')  # 轉義雙引號
            
            # 🔥 使用 HYPERLINK 函數格式
            if display_text and display_text != url:
                # 如果有不同的顯示文字
                clean_display_text = str(display_text).replace('"', '""')  # 轉義顯示文字中的雙引號
                cell.value = f'=HYPERLINK("{clean_url}","{clean_display_text}")'
            else:
                # 如果沒有特別的顯示文字，就顯示 URL 本身
                cell.value = f'=HYPERLINK("{clean_url}")'
            
            # 🔥 設定藍色超連結樣式
            cell.font = Font(color="0000FF", underline="single")
            
            self.logger.debug(f"添加 HYPERLINK 函數: {display_text or url} → {url}")
            
        except Exception as e:
            self.logger.error(f"添加 HYPERLINK 函數失敗: {str(e)}")
            # 備用方案：只顯示文字
            cell = worksheet.cell(row=row, column=col)
            cell.value = display_text or url
            cell.font = Font(color="0000FF")

    def _find_cell_value_in_row(self, worksheet, row_num: int, header_names: List[str]):
        """
        在指定行中尋找指定表頭名稱對應的值
        
        Args:
            worksheet: Excel 工作表
            row_num: 行號
            header_names: 要尋找的表頭名稱列表
            
        Returns:
            找到的值，如果沒找到則返回 None
        """
        try:
            # 先找到表頭對應的欄位號
            for col_num, header_cell in enumerate(worksheet[1], 1):  # 表頭行
                header_value = str(header_cell.value) if header_cell.value else ''
                
                if header_value in header_names:
                    # 找到對應欄位，取得該行該欄的值
                    cell = worksheet.cell(row=row_num, column=col_num)
                    return cell.value
            
            return None
            
        except Exception as e:
            self.logger.error(f"尋找行中欄位值失敗: {str(e)}")
            return None
                    
    def _add_summary_hyperlinks(self, worksheet, overwrite_type: str):
        """
        為轉換摘要頁籤添加 Gerrit 超連結
        
        Args:
            worksheet: Excel 工作表
            overwrite_type: 轉換類型
        """
        try:
            # 找到需要添加連結的欄位
            target_columns = {
                '源檔案': self.source_files.get(overwrite_type, ''),
                '目標檔案': self.target_files.get(overwrite_type, '')
            }
            
            # 🔥 新增：為 Gerrit 檔案欄位添加連結的映射
            gerrit_file_columns = ['Gerrit 源檔案', 'Gerrit 目標檔案']
            
            # 為每個目標欄位添加連結
            for col_num, cell in enumerate(worksheet[1], 1):  # 表頭行
                header_value = str(cell.value) if cell.value else ''
                
                # 處理源檔案和目標檔案
                if header_value in target_columns:
                    filename = target_columns[header_value]
                    if filename and filename != '':
                        gerrit_url = self._generate_gerrit_manifest_link(filename)
                        
                        # 在數據行添加超連結（第2行）
                        self._add_hyperlink_to_cell(worksheet, 2, col_num, gerrit_url, filename)
                        
                        self.logger.info(f"已為轉換摘要添加 {header_value} 連結: {filename}")
                
                # 🔥 新增：處理 Gerrit 檔案欄位
                elif header_value in gerrit_file_columns:
                    # 取得該欄位第2行的值（實際的檔案名）
                    data_cell = worksheet.cell(row=2, column=col_num)
                    filename = str(data_cell.value) if data_cell.value else ''
                    
                    if filename and filename not in ['', '無']:
                        # 生成 Gerrit 連結
                        gerrit_url = self._generate_gerrit_manifest_link(filename)
                        
                        # 添加超連結到該單元格
                        self._add_hyperlink_to_cell(worksheet, 2, col_num, gerrit_url, filename)
                        
                        self.logger.info(f"已為轉換摘要添加 {header_value} 連結: {filename}")
            
        except Exception as e:
            self.logger.error(f"添加轉換摘要超連結失敗: {str(e)}")
            
    def _add_revision_comparison_formula_converted_projects(self, worksheet):
        """為轉換後專案頁籤添加真正的動態條件格式 - 修改置中對齊，不要粗體"""
        try:
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
            from openpyxl.formatting.rule import Rule
            from openpyxl.styles.differential import DifferentialStyle
            
            # 找到相關欄位的位置
            original_revision_col = None
            converted_revision_col = None
            comparison_col = None
            
            for col_num, cell in enumerate(worksheet[1], 1):
                header_value = str(cell.value) if cell.value else ''
                if header_value == '原始 Revision':
                    original_revision_col = col_num
                elif header_value == '轉換後 Revision':
                    converted_revision_col = col_num
                elif header_value == 'Revision 是否相等':
                    comparison_col = col_num
            
            if not all([original_revision_col, converted_revision_col, comparison_col]):
                self.logger.warning(f"轉換後專案頁籤：無法找到所需的欄位位置")
                return
            
            # 取得欄位字母
            original_col_letter = get_column_letter(original_revision_col)
            converted_col_letter = get_column_letter(converted_revision_col)
            comparison_col_letter = get_column_letter(comparison_col)
            
            # 🔥 只添加 Excel 公式，不手動設定顏色
            for row_num in range(2, worksheet.max_row + 1):
                formula = f'=IF({original_col_letter}{row_num}={converted_col_letter}{row_num},"Y","N")'
                cell = worksheet[f"{comparison_col_letter}{row_num}"]
                cell.value = formula
                # 🔥 設定置中對齊，不要粗體
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 🔥 設定真正的動態條件格式 - 不要粗體
            green_font = Font(color="00B050", bold=False)  # 🔥 修改：不要粗體
            red_font = Font(color="FF0000", bold=False)    # 🔥 修改：不要粗體
            
            # 條件格式範圍
            range_string = f"{comparison_col_letter}2:{comparison_col_letter}{worksheet.max_row}"
            
            # 🔥 為 "Y" 值設定綠色字體（使用文字比較）
            green_rule = Rule(
                type="containsText",
                operator="containsText",
                text="Y",
                dxf=DifferentialStyle(font=green_font)
            )
            green_rule.formula = [f'NOT(ISERROR(SEARCH("Y",{comparison_col_letter}2)))']
            
            # 🔥 為 "N" 值設定紅色字體（使用文字比較）
            red_rule = Rule(
                type="containsText", 
                operator="containsText",
                text="N",
                dxf=DifferentialStyle(font=red_font)
            )
            red_rule.formula = [f'NOT(ISERROR(SEARCH("N",{comparison_col_letter}2)))']
            
            # 添加條件格式規則
            worksheet.conditional_formatting.add(range_string, green_rule)
            worksheet.conditional_formatting.add(range_string, red_rule)
            
            self.logger.info("✅ 已添加真正的動態條件格式（置中對齊，不粗體）")
            
        except Exception as e:
            self.logger.error(f"添加動態條件格式失敗: {str(e)}")

    def _generate_gerrit_manifest_link(self, filename: str) -> str:
        """
        生成 Gerrit manifest 檔案的連結 - 使用動態 Android 版本
        
        Args:
            filename: manifest 檔案名稱
            
        Returns:
            Gerrit 連結 URL
        """
        try:
            if not filename or filename == '無':
                return '無'
            
            # 移除 gerrit_ 前綴（如果有的話）
            clean_filename = filename.replace('gerrit_', '') if filename.startswith('gerrit_') else filename
            
            # 🔥 修改：構建 Gerrit 連結 - 使用動態分支
            master_branch = config.get_default_android_master_branch()
            base_url = f"https://mm2sd.rtkbf.com/gerrit/plugins/gitiles/realtek/android/manifest/+/refs/heads/{master_branch}"
            gerrit_link = f"{base_url}/{clean_filename}"
            
            self.logger.debug(f"生成 Gerrit 連結（動態版本）: {clean_filename} → {gerrit_link}")
            return gerrit_link
            
        except Exception as e:
            self.logger.error(f"生成 Gerrit 連結失敗: {str(e)}")
            return filename  # 返回原始檔名作為備用

    def _add_hyperlink_to_cell(self, worksheet, row: int, col: int, url: str, display_text: str):
        """
        為 Excel 單元格添加超連結 - 改進版本，確保藍色樣式
        
        Args:
            worksheet: Excel 工作表
            row: 行號
            col: 列號  
            url: 連結 URL
            display_text: 顯示文字
        """
        try:
            from openpyxl.worksheet.hyperlink import Hyperlink
            from openpyxl.styles import Font
            
            cell = worksheet.cell(row=row, column=col)
            
            # 🆕 方案1: 使用完整的 HYPERLINK 函數格式
            try:
                # 使用 Excel 的 HYPERLINK 函數，這樣 Excel 會更友善地處理
                cell.value = f'=HYPERLINK("{url}","{display_text}")'
                # 🔥 設定藍色超連結樣式
                cell.font = Font(color="0000FF", underline="single")
                self.logger.debug(f"添加 HYPERLINK 函數: {display_text} → {url}")
                return
            except Exception as e:
                self.logger.warning(f"HYPERLINK 函數失敗，嘗試標準超連結: {str(e)}")
            
            # 🆕 方案2: 標準超連結（備用）
            cell.value = display_text
            cell.hyperlink = Hyperlink(ref=f"{cell.coordinate}", target=url)
            # 🔥 設定藍色超連結樣式
            cell.font = Font(color="0000FF", underline="single")
            
            self.logger.debug(f"添加標準超連結: {display_text} → {url}")
            
        except Exception as e:
            self.logger.error(f"添加超連結失敗: {str(e)}")
            # 備用方案：顯示文字 + URL 備註
            cell = worksheet.cell(row=row, column=col)
            cell.value = f"{display_text}"
            # 🔥 即使是備用方案，也設定藍色字體
            cell.font = Font(color="0000FF")
            
            # 在註解中添加 URL
            try:
                from openpyxl.comments import Comment
                cell.comment = Comment(f"Gerrit 連結:\n{url}", "System")
            except:
                pass
            
    def _format_summary_content_backgrounds(self, worksheet):
        """為轉換摘要頁籤的統計欄位內容設定底色"""
        try:
            from openpyxl.styles import PatternFill
            from openpyxl.utils import get_column_letter
            
            # 🔥 定義內容底色
            light_blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")    # 淺藍底
            light_red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")     # 淺紅底
            light_green_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")   # 淺綠底
            lighter_blue_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")  # 更淺藍底
            
            # 🔥 定義需要設定內容底色的欄位
            content_background_fields = {
                '📊 總專案數': light_blue_fill,
                '🔄 實際轉換專案數': light_green_fill,
                '⭕ 未轉換專案數': light_red_fill,
                '🎯 目標檔案專案數': light_blue_fill,
                '❌ 轉換後與 Gerrit Manifest 差異數': light_red_fill,
                '✔️ 轉換後與 Gerrit Manifest 相同數': light_green_fill
            }
            
            # 找到需要設定底色的欄位並應用
            for col_num, cell in enumerate(worksheet[1], 1):
                header_value = str(cell.value) if cell.value else ''
                
                if header_value in content_background_fields:
                    fill_color = content_background_fields[header_value]
                    
                    # 為內容行（第2行開始）設定底色
                    for row_num in range(2, worksheet.max_row + 1):
                        content_cell = worksheet.cell(row=row_num, column=col_num)
                        content_cell.fill = fill_color
                    
                    self.logger.debug(f"已為 {header_value} 的內容設定底色")
            
            self.logger.info("✅ 已設定轉換摘要統計欄位的內容底色")
            
        except Exception as e:
            self.logger.error(f"設定轉換摘要內容底色失敗: {str(e)}")
            
    def _center_sn_columns(self, worksheet):
        """將所有 SN 欄位的內容置中"""
        try:
            from openpyxl.styles import Alignment
            
            # 找到 SN 欄位
            sn_columns = []
            for col_num, cell in enumerate(worksheet[1], 1):
                header_value = str(cell.value) if cell.value else ''
                if header_value == 'SN':
                    sn_columns.append(col_num)
            
            # 為所有 SN 欄位的內容設定置中對齊
            for col_num in sn_columns:
                for row_num in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if sn_columns:
                self.logger.debug(f"已將 {len(sn_columns)} 個 SN 欄位設定為置中對齊")
            
        except Exception as e:
            self.logger.error(f"設定 SN 欄位置中失敗: {str(e)}")
            
    def _auto_adjust_column_widths_enhanced(self, worksheet, sheet_name: str):
        """自動調整欄寬以適應內容 - 增強版本，確保所有內容都能完整顯示"""
        try:
            from openpyxl.utils import get_column_letter
            
            # 特別處理轉換摘要頁籤
            if sheet_name == "轉換摘要":
                self._adjust_summary_column_widths_enhanced(worksheet)
                return
            
            # 遍歷所有欄位
            for col in worksheet.columns:
                max_content_width = 0
                header_width = 0
                column = col[0].column_letter
                
                # 🔥 計算表頭的顯示寬度
                header_cell = worksheet[f"{column}1"]
                if header_cell.value:
                    header_value = str(header_cell.value)
                    header_width = self._calculate_display_width(header_value)
                    self.logger.debug(f"欄位 {column} 表頭 '{header_value}' 顯示寬度: {header_width}")
                
                # 🔥 計算內容的最大顯示寬度（檢查所有行）
                for cell in col[1:]:  # 跳過表頭行
                    try:
                        if cell.value:
                            cell_content = str(cell.value)
                            # 🆕 特別處理 HYPERLINK 函數內容
                            if cell_content.startswith('=HYPERLINK('):
                                # 從 HYPERLINK 函數中提取顯示文字
                                import re
                                match = re.search(r'=HYPERLINK\("[^"]*","([^"]*)"', cell_content)
                                if match:
                                    display_text = match.group(1)
                                    cell_width = self._calculate_display_width(display_text)
                                else:
                                    cell_width = self._calculate_display_width(cell_content)
                            else:
                                cell_width = self._calculate_display_width(cell_content)
                            
                            if cell_width > max_content_width:
                                max_content_width = cell_width
                    except:
                        pass
                
                # 🔥 取表頭寬度和內容寬度的較大值，加上足夠的邊距
                required_width = max(header_width, max_content_width) + 5  # 增加邊距
                
                # 🔥 設定特殊欄位的最小寬度
                if header_cell.value:
                    header_value = str(header_cell.value)
                    
                    if 'revision' in header_value.lower():
                        min_width = 40  # 增加 revision 欄位寬度
                    elif 'content' in header_value:
                        min_width = 80  # 增加 content 欄位寬度
                    elif header_value in ['name', 'gerrit_name', '專案名稱']:
                        min_width = 30  # 增加專案名稱欄位寬度
                    elif header_value in ['path', '專案路徑']:
                        min_width = 35  # 增加路徑欄位寬度
                    elif 'source_link' in header_value or 'gerrit_source_link' in header_value:
                        min_width = 60  # 增加連結欄位寬度
                    elif header_value in ['groups']:
                        min_width = 45  # 增加 groups 欄位寬度
                    elif header_value == 'SN':
                        min_width = 8
                    elif header_value in ['comparison_status', 'comparison_result']:
                        min_width = 25  # 增加比較狀態欄位寬度
                    elif 'upstream' in header_value.lower():
                        min_width = 25  # 增加 upstream 欄位寬度
                    elif 'dest-branch' in header_value.lower():
                        min_width = 25  # 增加 dest-branch 欄位寬度
                    elif 'clone-depth' in header_value.lower():
                        min_width = 15  # clone-depth 欄位寬度
                    elif 'remote' in header_value.lower():
                        min_width = 15  # remote 欄位寬度
                    else:
                        # 🔥 一般欄位最小寬度 = max(表頭寬度 + 邊距, 15)
                        min_width = max(header_width + 5, 15)
                    
                    final_width = max(required_width, min_width)
                else:
                    final_width = max(required_width, 15)
                
                # 🔥 設定最大寬度限制（增加到 120）
                final_width = min(final_width, 120)
                
                # 應用欄寬
                worksheet.column_dimensions[column].width = final_width
                
                self.logger.debug(f"欄位 {column} 最終寬度: {final_width} (表頭:{header_width}, 內容:{max_content_width})")
            
            self.logger.debug(f"已自動調整 {sheet_name} 的欄寬（增強版，確保所有內容完整顯示）")
            
        except Exception as e:
            self.logger.error(f"自動調整欄寬失敗 {sheet_name}: {str(e)}")

    def _adjust_summary_column_widths_enhanced(self, worksheet):
        """專門調整轉換摘要頁籤的欄寬 - 增強版本，確保所有內容完整顯示"""
        try:
            from openpyxl.utils import get_column_letter
            
            # 🔥 動態計算每個欄位的適當寬度
            for col_num, cell in enumerate(worksheet[1], 1):
                col_letter = get_column_letter(col_num)
                header_value = str(cell.value) if cell.value else ''
                
                if header_value:
                    # 🔥 使用精確的寬度計算
                    header_display_width = self._calculate_display_width(header_value)
                    
                    # 🔥 計算該欄位內容的最大寬度
                    max_content_width = 0
                    for row_num in range(2, worksheet.max_row + 1):
                        content_cell = worksheet.cell(row=row_num, column=col_num)
                        if content_cell.value:
                            content_str = str(content_cell.value)
                            # 特別處理 HYPERLINK 函數
                            if content_str.startswith('=HYPERLINK('):
                                import re
                                match = re.search(r'=HYPERLINK\("[^"]*","([^"]*)"', content_str)
                                if match:
                                    display_text = match.group(1)
                                    content_width = self._calculate_display_width(display_text)
                                else:
                                    content_width = self._calculate_display_width(content_str)
                            else:
                                content_width = self._calculate_display_width(content_str)
                            
                            if content_width > max_content_width:
                                max_content_width = content_width
                    
                    # 🔥 根據欄位類型設定基礎寬度
                    if header_value == 'SN':
                        base_width = 8
                    elif 'revision' in header_value.lower():
                        base_width = 40
                    elif 'content' in header_value:
                        base_width = 80
                    elif 'URL' in header_value:
                        base_width = 60
                    elif '❌' in header_value or '✅' in header_value or '✔️' in header_value:
                        base_width = 40  # 長的統計欄位
                    elif '📊' in header_value or '🔄' in header_value or '⭕' in header_value or '🎯' in header_value:
                        base_width = max(header_display_width + 5, 20)  # 根據實際內容調整
                    elif '狀態' in header_value:
                        base_width = max(header_display_width + 5, 20)
                    elif 'ID' in header_value:
                        base_width = 15
                    elif '檔案' in header_value:
                        base_width = max(max_content_width + 5, 25)  # 根據檔名長度調整
                    else:
                        base_width = max(header_display_width + 5, 15)  # 確保表頭能完整顯示
                    
                    # 🔥 確保寬度足夠顯示所有內容
                    final_width = max(base_width, header_display_width + 5, max_content_width + 5)
                    
                    # 🔥 設定合理的最大寬度限制
                    final_width = min(final_width, 100)
                    
                else:
                    final_width = 15  # 空表頭的預設寬度
                
                # 設定欄寬
                worksheet.column_dimensions[col_letter].width = final_width
                
                self.logger.debug(f"轉換摘要欄位 '{header_value}' 計算寬度: {final_width}")
            
            self.logger.info("✅ 已調整轉換摘要頁籤的欄寬（增強版，確保所有內容完整顯示）")
            
        except Exception as e:
            self.logger.error(f"調整轉換摘要欄寬失敗: {str(e)}")
            
    def _calculate_display_width(self, text: str) -> float:
        """
        計算文字的顯示寬度
        中文字符通常需要2個單位寬度，英文字符需要1個單位寬度
        
        Args:
            text: 要計算的文字
            
        Returns:
            顯示寬度
        """
        if not text:
            return 0
        
        width = 0
        for char in str(text):
            # 判斷是否為中文字符、全形字符或特殊符號
            if ord(char) > 127:  # 非 ASCII 字符
                if ord(char) >= 0x4e00 and ord(char) <= 0x9fff:  # 中文字符
                    width += 2
                elif ord(char) >= 0xff00 and ord(char) <= 0xffef:  # 全形字符
                    width += 2
                elif char in '📊🔄⭕🎯❌✅✔️':  # emoji 符號
                    width += 2.5
                else:
                    width += 1.5  # 其他特殊字符
            else:
                width += 1  # ASCII 字符
        
        return width

    def _set_conversion_status_colors_v3(self, worksheet):
        """設定轉換狀態的文字顏色 - 修正版本，不要粗體"""
        try:
            from openpyxl.styles import Font
            
            blue_font = Font(color="0070C0", bold=False)   # 🔥 修改：藍字，不要粗體
            gray_font = Font(color="808080", bold=False)   # 🔥 修改：灰字，不要粗體
            
            # 只找轉換狀態欄位
            status_column = None
            
            for col_num, cell in enumerate(worksheet[1], 1):
                header_value = str(cell.value) if cell.value else ''
                if '轉換狀態' in header_value:
                    status_column = col_num
                    break
            
            # 設定轉換狀態顏色
            if status_column:
                for row_num in range(2, worksheet.max_row + 1):
                    status_cell = worksheet.cell(row=row_num, column=status_column)
                    status_value = str(status_cell.value) if status_cell.value else ''
                    
                    if '🔄 已轉換' in status_value:
                        status_cell.font = blue_font  # 🔥 不要粗體
                    elif '⭕ 未轉換' in status_value:
                        status_cell.font = gray_font  # 🔥 不要粗體
            
        except Exception as e:
            self.logger.error(f"設定轉換狀態顏色失敗: {str(e)}")
                        
    def _format_worksheet_with_background_colors(self, worksheet, sheet_name: str):
        """格式化工作表 - 修正版本，設定Excel頁籤標籤顏色和新的表頭顏色"""
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
            
            # 🆕 設定Excel頁籤標籤顏色
            if sheet_name in ['轉換摘要', '轉換後專案']:
                # 淺藍色頁籤
                worksheet.sheet_properties.tabColor = "ADD8E6"  # Light Blue
            elif sheet_name in ['來源 gerrit manifest', '轉換後的 manifest', '目的 gerrit manifest']:
                # 淺綠色頁籤
                worksheet.sheet_properties.tabColor = "90EE90"  # Light Green
            elif sheet_name in ['轉換後與 Gerrit manifest 的差異']:
                # 淺紅色頁籤
                worksheet.sheet_properties.tabColor = "FFB6C1"  # Light Pink
            
            # 🆕 新增顏色定義
            blue_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # 藍底
            green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")        # 綠底
            red_fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")         # 紅底
            orange_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")      # 橘底
            purple_fill = PatternFill(start_color="8A2BE2", end_color="8A2BE2", fill_type="solid")      # 紫底
            dark_cyan_fill = PatternFill(start_color="008B8B", end_color="008B8B", fill_type="solid")   # 藍深青色
            teal_fill = PatternFill(start_color="20B2AA", end_color="20B2AA", fill_type="solid")        # 青藍色
            lighter_teal_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid") # 更淺的藍色
            link_blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")   # 🔥 新增：連結藍色背景
            
            white_font = Font(color="FFFFFF", bold=True)    # 白字
            blue_font = Font(color="0070C0", bold=True)     # 藍字
            gray_font = Font(color="808080", bold=True)     # 灰字
            
            # 🔥 修改特殊顏色的欄位定義
            orange_header_fields = ["推送狀態", "推送結果", "Commit ID", "Review URL", "轉換說明", 
                                "comparison_status", "comparison_result"]
            green_header_fields = ["Gerrit 源檔案", "Gerrit 展開檔案", "Gerrit 目標檔案",
                                "gerrit_content", "gerrit_name", "gerrit_path", "gerrit_upstream", 
                                "gerrit_dest-branch", "gerrit_groups", "gerrit_clone-depth", 
                                "gerrit_remote"]  # 🔥 移除 gerrit_source_link
            purple_header_fields = ["源檔案", "輸出檔案", "目標檔案", "來源檔案", "轉換後檔案",
                                "source_file", "gerrit_source_file"]
            dark_cyan_header_fields = ["🎯 目標檔案專案數", "❌ 轉換後與 Gerrit Manifest 差異數", "✔️ 轉換後與 Gerrit Manifest 相同數"]
            link_blue_header_fields = ["source_link", "gerrit_source_link", "📊 總專案數", "🔄 實際轉換專案數", "⭕ 未轉換專案數"]  # 🔥 新增藍色背景欄位
            
            # 設定表頭和欄寬
            for col_num, cell in enumerate(worksheet[1], 1):
                col_letter = get_column_letter(col_num)
                header_value = str(cell.value) if cell.value else ''
                
                # 🆕 根據欄位名稱設定特殊顏色
                if header_value in orange_header_fields:
                    cell.fill = orange_fill
                    cell.font = white_font
                    self.logger.debug(f"設定橘底白字表頭: {header_value}")
                elif header_value in green_header_fields:
                    cell.fill = green_fill
                    cell.font = white_font
                    self.logger.debug(f"設定綠底白字表頭: {header_value}")
                elif header_value in purple_header_fields:
                    cell.fill = purple_fill
                    cell.font = white_font
                    self.logger.debug(f"設定紫底白字表頭: {header_value}")
                elif header_value in dark_cyan_header_fields:
                    cell.fill = dark_cyan_fill
                    cell.font = white_font
                    self.logger.debug(f"設定藍深青色白字表頭: {header_value}")
                elif header_value in link_blue_header_fields:  # 🔥 新增藍色背景表頭
                    cell.fill = link_blue_fill
                    cell.font = white_font
                    self.logger.debug(f"設定藍色白字表頭: {header_value}")
                else:
                    # 預設所有其他表頭都是藍底白字
                    cell.fill = blue_header_fill
                    cell.font = white_font
                
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 特殊處理轉換後專案頁籤
                if sheet_name == "轉換後專案":
                    if header_value == '原始 Revision':
                        cell.fill = red_fill
                        cell.font = white_font
                        worksheet.column_dimensions[col_letter].width = 35
                    elif header_value == '轉換後 Revision':
                        cell.fill = red_fill
                        cell.font = white_font
                        worksheet.column_dimensions[col_letter].width = 35
                    elif header_value == 'Revision 是否相等':
                        cell.fill = red_fill
                        cell.font = white_font
                        worksheet.column_dimensions[col_letter].width = 15
                    elif 'revision' in header_value.lower():
                        worksheet.column_dimensions[col_letter].width = 35

                # 特殊處理差異頁籤
                elif sheet_name == "轉換後與 Gerrit manifest 的差異":
                    # revision 和 gerrit_revision 都用紅底白字
                    if header_value in ['revision', 'gerrit_revision']:
                        cell.fill = red_fill
                        cell.font = white_font
                    
                    # 設定欄寬
                    if 'content' in header_value or 'gerrit_content' in header_value:
                        worksheet.column_dimensions[col_letter].width = 80
                    elif 'revision' in header_value.lower():
                        worksheet.column_dimensions[col_letter].width = 35
                    elif 'comparison' in header_value:
                        worksheet.column_dimensions[col_letter].width = 20
                    elif header_value in ['name', 'gerrit_name']:
                        worksheet.column_dimensions[col_letter].width = 25
                    
                    # 根據比較狀態設定行的背景色
                    self._set_comparison_row_colors(worksheet, col_num, header_value)
                
                # manifest 相關頁籤的處理
                elif sheet_name in ['來源 gerrit manifest', '轉換後的 manifest', '目的 gerrit manifest']:
                    if 'revision' in header_value.lower():
                        worksheet.column_dimensions[col_letter].width = 35
                    elif header_value in ['name']:
                        worksheet.column_dimensions[col_letter].width = 25
                    elif header_value in ['path']:
                        worksheet.column_dimensions[col_letter].width = 30
                    elif header_value in ['groups']:
                        worksheet.column_dimensions[col_letter].width = 40
                    elif header_value == 'source_link':
                        worksheet.column_dimensions[col_letter].width = 60
                    elif header_value == 'source_file':
                        worksheet.column_dimensions[col_letter].width = 30
                
                # 其他頁籤的一般欄寬調整
                else:
                    if 'revision' in header_value.lower():
                        worksheet.column_dimensions[col_letter].width = 35
                    elif '名稱' in header_value or 'name' in header_value:
                        worksheet.column_dimensions[col_letter].width = 25
                    elif '路徑' in header_value or 'path' in header_value:
                        worksheet.column_dimensions[col_letter].width = 30
            
            # 🔥 添加 SN 欄位置中功能
            self._center_sn_columns(worksheet)
            
            # 設定轉換後專案頁籤的轉換狀態顏色
            if sheet_name == "轉換後專案":
                self._set_conversion_status_colors_v3(worksheet)  # 使用新版本，不要粗體
            
            self.logger.debug(f"已格式化工作表: {sheet_name}")
            
        except Exception as e:
            self.logger.error(f"格式化工作表失敗 {sheet_name}: {str(e)}")
            
    def _set_comparison_row_colors(self, worksheet, status_col_num: int, header_value: str):
        """設定比較狀態的行顏色 - 修正版本"""
        try:
            from openpyxl.styles import PatternFill
            
            # 找到比較狀態欄位
            if header_value != 'comparison_status':
                return
            
            # 定義狀態顏色（覆蓋頁籤底色）
            status_colors = {
                '✔️ 相同': PatternFill(start_color="D4FFCD", end_color="D4FFCD", fill_type="solid"),     # 深一點的綠
                '❌ 不同': PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),     # 深一點的紅
                '🆕 新增': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),     # 深一點的黃
                '🗑️ 刪除': PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")      # 深一點的橘
            }
            
            # 設定每一行的背景色
            for row_num in range(2, worksheet.max_row + 1):
                status_cell = worksheet.cell(row=row_num, column=status_col_num)
                status_value = str(status_cell.value) if status_cell.value else ''
                
                # 根據狀態設定整行背景色
                for status, fill_color in status_colors.items():
                    if status in status_value:
                        for col in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row_num, column=col).fill = fill_color
                        break
            
        except Exception as e:
            self.logger.error(f"設定比較狀態行顏色失敗: {str(e)}")
            
    def _generate_excel_report_safe(self, overwrite_type: str, source_file_path: Optional[str],
                                  output_file_path: Optional[str], target_file_path: Optional[str], 
                                  diff_analysis: Dict, output_folder: str, 
                                  excel_filename: Optional[str], source_download_success: bool,
                                  target_download_success: bool, push_result: Optional[Dict[str, Any]] = None,
                                  expanded_file_path: Optional[str] = None, use_expanded: bool = False) -> str:
        """安全的 Excel 報告生成 - 包含展開檔案資訊"""
        try:
            return self._generate_excel_report(
                overwrite_type=overwrite_type,
                source_file_path=source_file_path,
                output_file_path=output_file_path,
                target_file_path=target_file_path,
                diff_analysis=diff_analysis,
                output_folder=output_folder,
                excel_filename=excel_filename,
                source_download_success=source_download_success,
                target_download_success=target_download_success,
                push_result=push_result,
                expanded_file_path=expanded_file_path,
                use_expanded=use_expanded
            )
        except Exception as e:
            self.logger.error(f"標準 Excel 報告生成失敗: {str(e)}")
            self.logger.info("嘗試生成基本錯誤報告...")
            return self._generate_error_report(output_folder, overwrite_type, str(e))
    
    def _generate_error_report(self, output_folder: str, overwrite_type: str, error_message: str) -> str:
        """生成基本錯誤報告"""
        try:
            excel_filename = f"{overwrite_type}_error_report.xlsx"
            excel_file = os.path.join(output_folder, excel_filename)
            
            error_data = [{
                'SN': 1,
                '轉換類型': overwrite_type,
                '處理狀態': '失敗',
                '錯誤訊息': error_message,
                '時間': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'),
                '建議': '請檢查網路連線和 Gerrit 認證設定'
            }]
            
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                df_error = pd.DataFrame(error_data)
                df_error.to_excel(writer, sheet_name='錯誤報告', index=False)
                
                # 格式化
                worksheet = writer.sheets['錯誤報告']
                self.excel_handler._format_worksheet(worksheet)
            
            self.logger.info(f"已生成基本錯誤報告: {excel_file}")
            return excel_file
            
        except Exception as e:
            self.logger.error(f"連基本錯誤報告都無法生成: {str(e)}")
            return ""
    
    def _final_file_report_complete(self, output_folder: str, source_file_path: Optional[str], 
                          output_file_path: Optional[str], target_file_path: Optional[str], 
                          excel_file: str, source_download_success: bool, target_download_success: bool,
                          expanded_file_path: Optional[str] = None):
        """最終檔案檢查和報告 - 增強版本，包含展開檔案"""
        try:
            self.logger.info("🔍 最終檔案檢查報告:")
            self.logger.info(f"📂 輸出資料夾: {output_folder}")
            
            # 檢查所有應該存在的檔案
            files_to_check = []
            
            if source_file_path:
                status = "✅" if source_download_success else "❌"
                files_to_check.append((f"Gerrit 源檔案 {status}", source_file_path))
            
            if expanded_file_path:
                files_to_check.append(("Gerrit 展開檔案 ✅", expanded_file_path))
            
            if output_file_path:
                files_to_check.append(("轉換後檔案 ✅", output_file_path))
            
            if target_file_path:
                status = "✅" if target_download_success else "❌"
                files_to_check.append((f"Gerrit 目標檔案 {status}", target_file_path))
            
            files_to_check.append(("Excel 報告 ✅", excel_file))
            
            # 逐一檢查檔案
            all_files_exist = True
            for file_type, file_path in files_to_check:
                if os.path.exists(file_path):
                    file_size = os.path.getsize(file_path)
                    filename = os.path.basename(file_path)
                    self.logger.info(f"  {file_type}: {filename} ({file_size} bytes)")
                else:
                    self.logger.error(f"  {file_type}: {file_path} (檔案不存在)")
                    all_files_exist = False
            
            # 下載狀態統計
            self.logger.info(f"\n📊 Gerrit 下載狀態統計:")
            self.logger.info(f"  🔵 源檔案下載: {'✅ 成功' if source_download_success else '❌ 失敗'}")
            self.logger.info(f"  🟡 目標檔案下載: {'✅ 成功' if target_download_success else '❌ 失敗 (檔案不存在)'}")
            
            if expanded_file_path:
                self.logger.info(f"  🟢 Manifest 展開: {'✅ 成功' if os.path.exists(expanded_file_path) else '❌ 失敗'}")
            
            if not target_download_success:
                self.logger.info(f"  💡 提示: 目標檔案在 Gerrit 上不存在是正常情況")
                self.logger.info(f"       這表示該 manifest 檔案尚未在 master 分支上建立")
            
            # 額外檢查 output 資料夾中的所有 XML 檔案
            self.logger.info(f"\n📋 Output 資料夾中的所有 XML 檔案:")
            xml_files_found = []
            try:
                for filename in os.listdir(output_folder):
                    if filename.lower().endswith('.xml'):
                        file_path = os.path.join(output_folder, filename)
                        file_size = os.path.getsize(file_path)
                        xml_files_found.append((filename, file_size))
                        self.logger.info(f"  📄 {filename} ({file_size} bytes)")
                
                if not xml_files_found:
                    self.logger.warning("  ⚠️ 沒有找到任何 XML 檔案")
                else:
                    self.logger.info(f"\n📊 XML 檔案統計:")
                    gerrit_files = [f for f in xml_files_found if f[0].startswith('gerrit_')]
                    converted_files = [f for f in xml_files_found if not f[0].startswith('gerrit_')]
                    
                    self.logger.info(f"  🔵 Gerrit 原始檔案: {len(gerrit_files)} 個")
                    for filename, size in gerrit_files:
                        is_expanded = "_expand.xml" in filename
                        file_type = "(展開檔案)" if is_expanded else "(原始檔案)"
                        self.logger.info(f"    - {filename} ({size} bytes) {file_type}")
                    
                    self.logger.info(f"  🟢 轉換後檔案: {len(converted_files)} 個")
                    for filename, size in converted_files:
                        self.logger.info(f"    - {filename} ({size} bytes)")
                    
            except Exception as e:
                self.logger.error(f"  ❌ 無法列出資料夾內容: {str(e)}")
            
            # 總結
            if all_files_exist:
                self.logger.info(f"\n✅ 所有檔案都已成功保存")
                if source_file_path:
                    source_filename = os.path.basename(source_file_path)
                    self.logger.info(f"🎯 重點提醒: 已保存 Gerrit 源檔案為 {source_filename}")
                if expanded_file_path:
                    expanded_filename = os.path.basename(expanded_file_path)
                    self.logger.info(f"🎯 重點提醒: 已保存展開檔案為 {expanded_filename}")
                    self.logger.info(f"🎯 轉換使用的是展開後的檔案")
                if not target_download_success:
                    self.logger.info(f"📋 Excel 報告中已記錄目標檔案下載失敗狀態（紅字標示）")
            else:
                self.logger.warning(f"\n⚠️ 部分檔案可能保存失敗，請檢查上述報告")
                
        except Exception as e:
            self.logger.error(f"檔案檢查報告失敗: {str(e)}")

    def _is_revision_hash(self, revision: str) -> bool:
        """
        判斷 revision 是否為 commit hash - 改進版本
        
        Args:
            revision: revision 字串
            
        Returns:
            True 如果是 hash，False 如果是 branch/tag name
        """
        if not revision:
            return False
        
        revision = revision.strip()
        
        # 🆕 明確排除 refs/ 開頭的（這些是 branch 或 tag）
        if revision.startswith('refs/'):
            return False
        
        # 🆕 Hash 特徵：純十六進制字串
        # 40 字符的完整 hash
        if len(revision) == 40 and all(c in '0123456789abcdefABCDEF' for c in revision):
            return True
        
        # 7-12 字符的短 hash
        if 7 <= len(revision) <= 12 and all(c in '0123456789abcdefABCDEF' for c in revision):
            return True
        
        # 🆕 包含非十六進制字符的一定不是 hash
        if any(c not in '0123456789abcdefABCDEF' for c in revision):
            return False
        
        # 其他情況當作 branch name 處理
        return False

    def _get_effective_revision_for_conversion(self, project_element) -> str:
        """
        取得用於轉換的有效 revision（XML 元素版本）
        
        邏輯：
        - 如果 revision 是 hash → 使用 upstream
        - 如果 revision 是 branch name → 使用 revision
        - 如果都沒有 → 使用 default revision（如果 remote=rtk）
        
        Args:
            project_element: XML project 元素
            
        Returns:
            用於轉換的 revision 字串
        """
        revision = project_element.get('revision', '')
        upstream = project_element.get('upstream', '')
        remote = project_element.get('remote', '') or self.default_remote
        
        # 如果 revision 是 hash，使用 upstream
        if self._is_revision_hash(revision):
            if upstream:
                self.logger.debug(f"專案 {project_element.get('name', '')} revision 是 hash，使用 upstream: {upstream}")
                return upstream
            else:
                self.logger.warning(f"專案 {project_element.get('name', '')} revision 是 hash 但沒有 upstream")
                return ''
        
        # 如果 revision 是 branch name，直接使用
        if revision:
            self.logger.debug(f"專案 {project_element.get('name', '')} revision 是 branch name: {revision}")
            return revision
        
        # 如果沒有 revision，返回空字串（會由其他邏輯處理 default revision）
        self.logger.debug(f"專案 {project_element.get('name', '')} 沒有 revision")
        return ''    