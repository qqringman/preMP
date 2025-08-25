"""
檔案比較模組（增強版 - 支援 Mapping Table）
處理 manifest.xml, F_Version.txt, Version.txt 的差異比較
支援一次執行所有比對情境
支援 Mapping Table 機制進行精確比對
"""
import os
import re
import glob
import xml.etree.ElementTree as ET
from typing import List, Dict, Any, Tuple, Set, Optional
import pandas as pd
import utils
import config
from excel_handler import ExcelHandler

logger = utils.setup_logger(__name__)

class FileComparator:
    """檔案比較器類別"""
    
    def __init__(self):
        self.logger = logger
        self.excel_handler = ExcelHandler()
        self.base_url_prebuilt = config.GERRIT_BASE_URL_PREBUILT
        self.base_url_normal = config.GERRIT_BASE_URL_NORMAL
        self.mapping_tables = {}  # 儲存載入的 mapping tables
        
    def _shorten_hash(self, hash_str: str) -> str:
        """將 hash code 縮短為前 7 個字元"""
        if hash_str and len(hash_str) >= 7:
            return hash_str[:7]
        return hash_str
    
    def _generate_link(self, project_info: Dict[str, str]) -> str:
        """根據 project 資訊生成 Gerrit link"""
        name = project_info.get('name', '')
        upstream = project_info.get('upstream', '')
        dest_branch = project_info.get('dest-branch', '')
        
        # 優先使用 upstream，如果沒有則使用 dest-branch
        branch = upstream if upstream else dest_branch
        
        if name and branch:
            # 判斷是否包含 prebuilt 或 prebuild，選擇對應的 base URL
            if 'prebuilt' in name.lower() or 'prebuild' in name.lower():
                base_url = self.base_url_prebuilt
            else:
                base_url = self.base_url_normal
            
            # 處理特殊的 branch 格式
            if branch.startswith('refs/'):
                # 如果 branch 已經包含 refs/，直接使用
                # 例如：refs/tags/u-tv-keystone-rtk-refplus-wave3-release-UKR8.20250727
                return f"{base_url}{name}/+log/{branch}"
            else:
                # 一般情況，加上 refs/heads/
                return f"{base_url}{name}/+log/refs/heads/{branch}"
        return ""
    
    def _extract_simple_module_name(self, full_module: str) -> str:
        """從完整模組路徑提取簡單模組名稱"""
        # PrebuildFW/bootcode -> bootcode
        # DailyBuild/Merlin7 -> Merlin7
        if '/' in full_module:
            return full_module.split('/')[-1]
        return full_module
        
    def _parse_manifest_xml(self, file_path: str) -> Dict[str, Dict[str, str]]:
        """
        解析 manifest.xml 檔案
        
        Args:
            file_path: XML 檔案路徑
            
        Returns:
            專案資訊字典 {key: project_info}
        """
        projects = {}
        
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # 尋找所有 project 元素
            for project in root.findall('.//project'):
                name = project.get('name', '')
                path = project.get('path', '')
                key = f"{name}||{path}"  # 使用 name 和 path 作為唯一鍵
                
                projects[key] = {
                    'name': name,
                    'path': path,
                    'revision': project.get('revision', ''),
                    'upstream': project.get('upstream', ''),
                    'dest-branch': project.get('dest-branch', ''),
                    'groups': project.get('groups', ''),
                    'clone-depth': project.get('clone-depth', ''),
                    'remote': project.get('remote', ''),
                    'element': ET.tostring(project, encoding='unicode').strip()
                }
                
            self.logger.info(f"成功解析 {file_path}，找到 {len(projects)} 個專案")
            
        except Exception as e:
            self.logger.error(f"解析 XML 檔案失敗 {file_path}: {str(e)}")
            
        return projects
        
    def _compare_manifest_files(self, file1: str, file2: str, module: str, base_folder: str = None, compare_folder: str = None, module_path: str = None, compare_mode: str = None) -> Tuple[List[Dict], List[Dict], List[Dict]]:
        """
        比較兩個 manifest.xml 檔案（修改後的邏輯）
        """
        # 解析兩個檔案
        base_projects = self._parse_manifest_xml(file1)
        compare_projects = self._parse_manifest_xml(file2)
        
        # 簡化模組名稱
        simple_module = self._extract_simple_module_name(module)
        
        # 1. 比較 revision 差異
        revision_diff = []
        sn = 1
        
        for key, base_proj in base_projects.items():
            if key in compare_projects:
                compare_proj = compare_projects[key]
                base_revision = base_proj.get('revision', '')
                compare_revision = compare_proj.get('revision', '')
                
                if base_revision != compare_revision or not compare_revision:
                    revision_diff.append({
                        'SN': sn,
                        'module': simple_module,
                        'location_path': module_path if module_path else '',
                        'base_folder': base_folder,
                        'compare_folder': compare_folder,
                        'name': base_proj['name'],
                        'path': base_proj['path'],
                        'base_short': self._shorten_hash(base_revision),
                        'base_revision': base_revision,
                        'compare_short': self._shorten_hash(compare_revision),
                        'compare_revision': compare_revision,
                        'base_upstream': base_proj.get('upstream', ''),
                        'compare_upstream': compare_proj.get('upstream', ''),
                        'base_dest-branch': base_proj.get('dest-branch', ''),
                        'compare_dest-branch': compare_proj.get('dest-branch', ''),
                        'base_link': self._generate_link(base_proj),
                        'compare_link': self._generate_link(compare_proj)
                    })
                    sn += 1
        
        # 2. 檢查分支命名錯誤
        branch_error = []
        sn = 1
        
        # 根據資料夾名稱決定檢查規則
        check_keyword = None
        if base_folder and compare_folder:
            if "-premp" in compare_folder and "-premp" not in base_folder:
                check_keyword = 'premp'
            elif "-wave" in compare_folder and "-wave.backup" not in compare_folder:
                if "-premp" in base_folder:
                    check_keyword = 'wave'
            elif "-wave.backup" in compare_folder:
                check_keyword = 'wave.backup'
        
        if check_keyword:
            for key, compare_proj in compare_projects.items():
                upstream = compare_proj.get('upstream', '')
                dest_branch = compare_proj.get('dest-branch', '')
                revision = compare_proj.get('revision', '')
                
                # 根據檢查關鍵字進行相應的檢查
                should_check = False
                
                if check_keyword == 'premp':
                    # 檢查是否都不包含 'premp'
                    if upstream and dest_branch:
                        if 'premp' not in upstream and 'premp' not in dest_branch:
                            should_check = True
                elif check_keyword == 'wave':
                    # 檢查是否都不包含 'wave'
                    if upstream and dest_branch:
                        if 'wave' not in upstream and 'wave' not in dest_branch:
                            should_check = True
                elif check_keyword == 'wave.backup':
                    # 檢查是否都不包含 'wave.backup'
                    if upstream and dest_branch:
                        if 'wave.backup' not in upstream and 'wave.backup' not in dest_branch:
                            should_check = True
                
                if should_check:
                    # 檢查是否包含 wave
                    has_wave = ('wave' in upstream or 'wave' in dest_branch)
                    
                    # 決定問題描述
                    problem = ""
                    if not has_wave:  # 只有 has_wave = N 時才顯示問題
                        if check_keyword == 'premp':
                            problem = "沒改成 premp"
                        elif check_keyword == 'wave':
                            problem = "沒改成 wave"
                        elif check_keyword == 'wave.backup':
                            problem = "沒改成 wavebackup"
                    
                    branch_error.append({
                        'SN': sn,
                        'module': simple_module,
                        'location_path': module_path if module_path else '',
                        'base_folder': base_folder,
                        'compare_folder': compare_folder,
                        'name': compare_proj['name'],
                        'path': compare_proj['path'],
                        'revision_short': self._shorten_hash(revision),
                        'revision': revision,
                        'upstream': upstream,
                        'dest-branch': dest_branch,
                        'compare_link': self._generate_link(compare_proj),
                        'has_wave': 'Y' if has_wave else 'N',
                        'problem': problem
                    })
                    sn += 1
        
        # 3. 檢查缺少或新增的 project（保持原有邏輯）
        lost_project = []
        sn = 1
        
        # 決定 Base folder 值
        base_folder_value = ""
        if compare_mode == 'master_vs_premp':
            base_folder_value = "premp"
        elif compare_mode == 'premp_vs_wave':
            base_folder_value = "wave"
        elif compare_mode == 'wave_vs_backup':
            base_folder_value = "wavebackup"
        
        # 檢查在 base 檔案中但不在 compare 檔案中的項目（刪除）
        for key, base_proj in base_projects.items():
            if key not in compare_projects:
                revision = base_proj.get('revision', '')
                lost_project.append({
                    'SN': sn,
                    'Base folder': base_folder_value,
                    '狀態': '刪除',
                    'module': simple_module,
                    'location_path': module_path if module_path else '',
                    'folder': base_folder,  # 記錄是哪個資料夾
                    'name': base_proj['name'],
                    'path': base_proj['path'],
                    'upstream': base_proj.get('upstream', ''),
                    'dest-branch': base_proj.get('dest-branch', ''),
                    'revision': revision,
                    'link': self._generate_link(base_proj)
                })
                sn += 1
        
        # 檢查在 compare 檔案中但不在 base 檔案中的項目（新增）
        for key, compare_proj in compare_projects.items():
            if key not in base_projects:
                revision = compare_proj.get('revision', '')
                lost_project.append({
                    'SN': sn,
                    'Base folder': base_folder_value,
                    '狀態': '新增',
                    'module': simple_module,
                    'location_path': module_path if module_path else '',
                    'folder': compare_folder,  # 記錄是哪個資料夾
                    'name': compare_proj['name'],
                    'path': compare_proj['path'],
                    'upstream': compare_proj.get('upstream', ''),
                    'dest-branch': compare_proj.get('dest-branch', ''),
                    'revision': revision,
                    'link': self._generate_link(compare_proj)
                })
                sn += 1
        
        return revision_diff, branch_error, lost_project
        
    def _compare_text_files(self, file1: str, file2: str, file_type: str) -> List[Dict[str, Any]]:
        """
        比較兩個文字檔案（根據檔案類型使用不同的比較規則）
        
        Args:
            file1: 第一個檔案路徑
            file2: 第二個檔案路徑
            file_type: 檔案類型（F_Version.txt 或 Version.txt）
            
        Returns:
            差異列表
        """
        differences = []
        
        try:
            with open(file1, 'r', encoding='utf-8', errors='ignore') as f1:
                content1 = f1.read()
                lines1 = content1.splitlines()
            with open(file2, 'r', encoding='utf-8', errors='ignore') as f2:
                content2 = f2.read()
                lines2 = content2.splitlines()
            
            if file_type.lower() == 'f_version.txt':
                # F_Version.txt 比較規則：只比較 P_GIT_xxx 行
                differences = self._compare_f_version_format(lines1, lines2, content1, content2)
                
            elif file_type.lower() == 'version.txt':
                # Version.txt 可能有多種格式，需要智能判斷
                
                # 檢查是否包含 P_GIT 格式（類似 F_Version.txt）
                has_p_git = any('P_GIT_' in line for line in lines1 + lines2)
                
                if has_p_git:
                    # 如果有 P_GIT 格式，使用 F_Version 的比較方式
                    differences = self._compare_f_version_format(lines1, lines2, content1, content2)
                else:
                    # 使用一般的 key:value 格式比較
                    differences = self._compare_key_value_format(lines1, lines2, content1, content2)
                    
        except Exception as e:
            self.logger.error(f"比較文字檔案失敗: {str(e)}")
            
        return differences

    def _compare_key_value_format(self, lines1: List[str], lines2: List[str],
                                content1: str, content2: str) -> List[Dict[str, Any]]:
        """
        比較 key:value 格式（Version.txt 的一般格式）
        
        Args:
            lines1: 第一個檔案的行列表
            lines2: 第二個檔案的行列表
            content1: 第一個檔案的完整內容
            content2: 第二個檔案的完整內容
            
        Returns:
            差異列表
        """
        differences = []
        
        # 定義需要比對的關鍵字
        important_keys = [
            'GIT_HASH', 'CommitHash', 'CommitCount', 
            'F_HASH', 'P_JIRA_ID', 'P_JIRA_URL',
            'P_CATEGORY', 'P_JKS_BUILD_URL', 'P_REPO_INIT',
            'P_DAILY_BUILD'
        ]
        
        # 提取包含關鍵字的行
        key_lines1 = {}
        key_lines2 = {}
        
        for line in lines1:
            line = line.strip()
            if not line or line.startswith('#') or line.startswith('//'):
                continue
                
            # 尋找包含重要關鍵字的行
            for key in important_keys:
                if key in line:
                    # 嘗試解析 key:value 格式
                    if ':' in line:
                        parts = line.split(':', 1)
                        key_part = parts[0].strip()
                        value_part = parts[1].strip() if len(parts) > 1 else ''
                        
                        # 使用較寬鬆的 key 匹配
                        if key in key_part:
                            key_lines1[key_part] = line
                            break
                    elif ';' in line:
                        # 處理用分號分隔的格式
                        key_lines1[key] = line
                        break
        
        for line in lines2:
            line = line.strip()
            if not line or line.startswith('#') or line.startswith('//'):
                continue
                
            for key in important_keys:
                if key in line:
                    if ':' in line:
                        parts = line.split(':', 1)
                        key_part = parts[0].strip()
                        value_part = parts[1].strip() if len(parts) > 1 else ''
                        
                        if key in key_part:
                            key_lines2[key_part] = line
                            break
                    elif ';' in line:
                        key_lines2[key] = line
                        break
        
        # 比較差異
        all_keys = set(key_lines1.keys()) | set(key_lines2.keys())
        for key in sorted(all_keys):
            line1 = key_lines1.get(key, '')
            line2 = key_lines2.get(key, '')
            
            # 提取值進行比較
            value1 = self._extract_value_from_line(line1)
            value2 = self._extract_value_from_line(line2)
            
            if value1 != value2:
                differences.append({
                    'line': key,
                    'file1': line1 if line1 else '(檔案不存在)',
                    'file2': line2 if line2 else '(檔案不存在)',
                    'content1': content1,
                    'content2': content2
                })
        
        return differences

    def _extract_value_from_line(self, line: str) -> str:
        """
        從行中提取值部分
        
        Args:
            line: 包含 key:value 或其他格式的行
            
        Returns:
            提取的值
        """
        if not line:
            return ''
        
        # 處理 key:value 格式
        if ':' in line:
            parts = line.split(':', 1)
            if len(parts) > 1:
                return parts[1].strip()
        
        # 處理 key=value 格式
        if '=' in line:
            parts = line.split('=', 1)
            if len(parts) > 1:
                return parts[1].strip()
        
        return line
                
    def _compare_f_version_format(self, lines1: List[str], lines2: List[str], 
                                content1: str, content2: str) -> List[Dict[str, Any]]:
        """
        比較 F_Version.txt 格式（P_GIT_xxx 行）
        
        Args:
            lines1: 第一個檔案的行列表
            lines2: 第二個檔案的行列表
            content1: 第一個檔案的完整內容
            content2: 第二個檔案的完整內容
            
        Returns:
            差異列表
        """
        differences = []
        git_lines1 = {}
        git_lines2 = {}
        
        # 提取 P_GIT_xxx 行
        for line in lines1:
            line = line.strip()
            if line.startswith('P_GIT_'):
                # P_GIT_001;realtek/bootcode;realtek/mac7p_64/master;2ef5076;1005445
                parts = line.split(';')
                if len(parts) >= 5:
                    git_id = parts[0]
                    git_lines1[git_id] = {
                        'full_line': line,
                        'parts': parts,
                        'repo': parts[1] if len(parts) > 1 else '',
                        'branch': parts[2] if len(parts) > 2 else '',
                        'commit': parts[3] if len(parts) > 3 else '',  # 這是需要比對的
                        'count': parts[4] if len(parts) > 4 else ''    # 這也需要比對
                    }
        
        for line in lines2:
            line = line.strip()
            if line.startswith('P_GIT_'):
                parts = line.split(';')
                if len(parts) >= 5:
                    git_id = parts[0]
                    git_lines2[git_id] = {
                        'full_line': line,
                        'parts': parts,
                        'repo': parts[1] if len(parts) > 1 else '',
                        'branch': parts[2] if len(parts) > 2 else '',
                        'commit': parts[3] if len(parts) > 3 else '',
                        'count': parts[4] if len(parts) > 4 else ''
                    }
        
        # 比較差異
        all_git_ids = set(git_lines1.keys()) | set(git_lines2.keys())
        for git_id in sorted(all_git_ids):
            line1_data = git_lines1.get(git_id, {})
            line2_data = git_lines2.get(git_id, {})
            
            line1 = line1_data.get('full_line', '')
            line2 = line2_data.get('full_line', '')
            
            # 比較 commit hash (第4個欄位) 和 count (第5個欄位)
            if line1 and line2:
                commit1 = line1_data.get('commit', '')
                commit2 = line2_data.get('commit', '')
                count1 = line1_data.get('count', '')
                count2 = line2_data.get('count', '')
                
                if commit1 != commit2 or count1 != count2:
                    differences.append({
                        'line': git_id,
                        'file1': line1,
                        'file2': line2,
                        'content1': content1,
                        'content2': content2
                    })
            elif line1 or line2:
                # 只有一邊有這個 GIT ID
                differences.append({
                    'line': git_id,
                    'file1': line1,
                    'file2': line2,
                    'content1': content1,
                    'content2': content2
                })
        
        return differences
            
    def compare_all_scenarios(self, source_dir: str, output_dir: str = None) -> Dict[str, Any]:
        """
        執行所有比對情境（支援 mapping table）
        """
        output_dir = output_dir or source_dir

        # 確保輸出目錄存在
        if not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
            self.logger.info(f"建立輸出目錄: {output_dir}")
        
        # 載入 mapping tables
        self.mapping_tables = self._load_mapping_tables(source_dir)
        
        # 如果有 mapping table，優先使用
        if self.mapping_tables:
            self.logger.info(f"找到 {len(self.mapping_tables)} 個 mapping tables，將優先使用")
            return self._compare_with_mapping(source_dir, output_dir)
        else:
            self.logger.info("未找到 mapping tables，使用原有邏輯")
            return self._compare_without_mapping(source_dir, output_dir)

    def _write_total_summary_report(self, all_results, scenario_data, output_file):
        """寫入總摘要報告，包含所有情境的統計"""
        try:
            import pandas as pd
            from openpyxl.styles import PatternFill, Font, Alignment
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 總摘要頁籤 - 這個頁籤必須存在且可見
                summary_rows = []
                
                # 各情境的統計
                scenarios_to_check = ['master_vs_premp', 'premp_vs_wave', 'wave_vs_backup']
                total_success = 0
                total_failed = 0
                
                for scenario_key in scenarios_to_check:
                    if scenario_key in all_results:
                        scenario_result = all_results[scenario_key]
                        scenario_name = self._get_scenario_display_name(scenario_key)
                        
                        # 確保 scenario_result 是字典且有預期的鍵
                        if isinstance(scenario_result, dict):
                            success_count = scenario_result.get('success', 0)
                            failed_count = scenario_result.get('failed', 0)
                            modules_list = scenario_result.get('modules', [])
                            failed_modules_list = scenario_result.get('failed_modules', [])
                            
                            total_success += success_count
                            total_failed += failed_count
                            
                            summary_rows.append({
                                '比對情境': scenario_name,
                                '成功模組數': success_count,
                                '失敗模組數': failed_count,
                                '成功模組清單': ', '.join(modules_list[:5]) + ('...' if len(modules_list) > 5 else ''),
                                '失敗模組清單': ', '.join(failed_modules_list[:5]) + ('...' if len(failed_modules_list) > 5 else '')
                            })
                        else:
                            # 如果不是字典或沒有資料，創建空的記錄
                            scenario_name = self._get_scenario_display_name(scenario_key)
                            summary_rows.append({
                                '比對情境': scenario_name,
                                '成功模組數': 0,
                                '失敗模組數': 0,
                                '成功模組清單': '無',
                                '失敗模組清單': '無'
                            })
                
                # 如果沒有任何情境資料，至少創建一個預設行
                if not summary_rows:
                    summary_rows.append({
                        '比對情境': '無資料',
                        '成功模組數': 0,
                        '失敗模組數': 0,
                        '成功模組清單': '無',
                        '失敗模組清單': '無'
                    })
                
                # 加入總計
                summary_rows.append({
                    '比對情境': '總計',
                    '成功模組數': total_success,
                    '失敗模組數': total_failed,
                    '成功模組清單': f'共 {total_success} 個模組',
                    '失敗模組清單': f'共 {total_failed} 個模組'
                })
                
                # 創建總摘要 DataFrame 並寫入（這個必須存在）
                df_summary = pd.DataFrame(summary_rows)
                df_summary.to_excel(writer, sheet_name='總摘要', index=False)
                
                # 所有差異的統計頁籤
                stats_data = []
                for scenario_key in scenarios_to_check:
                    if scenario_key in scenario_data and isinstance(scenario_data[scenario_key], dict):
                        data = scenario_data[scenario_key]
                        scenario_name = self._get_scenario_display_name(scenario_key)
                        
                        stats_data.append({
                            '情境': scenario_name,
                            'Revision 差異': len(data.get('revision_diff', [])),
                            '分支錯誤': len(data.get('branch_error', [])),
                            '新增/刪除專案': len(data.get('lost_project', [])),
                            '版本檔案差異': len(data.get('version_diff', [])),
                            '無法比對': len(data.get('cannot_compare', []))
                        })
                    else:
                        # 沒有資料的情境也要列出
                        scenario_name = self._get_scenario_display_name(scenario_key)
                        stats_data.append({
                            '情境': scenario_name,
                            'Revision 差異': 0,
                            '分支錯誤': 0,
                            '新增/刪除專案': 0,
                            '版本檔案差異': 0,
                            '無法比對': 0
                        })
                
                # 如果沒有統計資料，創建預設的
                if not stats_data:
                    stats_data.append({
                        '情境': '無資料',
                        'Revision 差異': 0,
                        '分支錯誤': 0,
                        '新增/刪除專案': 0,
                        '版本檔案差異': 0,
                        '無法比對': 0
                    })
                
                df_stats = pd.DataFrame(stats_data)
                df_stats.to_excel(writer, sheet_name='差異統計', index=False)
                
                # 格式化工作表
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    if hasattr(self, 'excel_handler'):
                        self.excel_handler._format_worksheet(worksheet)
                    else:
                        # 基本格式化
                        for cell in worksheet[1]:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            cell.font = Font(color="FFFFFF", bold=True)
            
            self.logger.info(f"成功寫入總摘要報告: {output_file}")
            
        except Exception as e:
            self.logger.error(f"寫入總摘要報告失敗: {str(e)}")
            import traceback
            self.logger.error(traceback.format_exc())
            
            # 如果寫入失敗，創建一個最基本的 Excel 檔案
            try:
                basic_data = pd.DataFrame([{
                    '狀態': '報告生成失敗',
                    '錯誤訊息': str(e),
                    '時間': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
                }])
                basic_data.to_excel(output_file, sheet_name='錯誤報告', index=False)
                self.logger.info(f"創建了基本錯誤報告: {output_file}")
            except Exception as backup_error:
                self.logger.error(f"創建基本報告也失敗: {str(backup_error)}")
                raise
        
    def _write_scenario_summary_report(self, revision_diff, branch_error, lost_project, 
                                    version_diff, cannot_compare, scenario_results,
                                    output_file, scenario_name):
        """為單一情境寫入摘要報表"""
        try:
            import pandas as pd
            from openpyxl.styles import PatternFill, Font, Alignment
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 摘要頁籤
                summary_data = [{
                    '項目': '比對情境',
                    '值': self._get_scenario_display_name(scenario_name)
                }, {
                    '項目': '成功模組數',
                    '值': scenario_results['success']
                }, {
                    '項目': '失敗模組數',
                    '值': scenario_results['failed']
                }, {
                    '項目': '成功模組清單',
                    '值': ', '.join(scenario_results['modules']) if scenario_results['modules'] else '無'
                }, {
                    '項目': '失敗模組清單',
                    '值': ', '.join(scenario_results['failed_modules']) if scenario_results['failed_modules'] else '無'
                }]
                
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='摘要', index=False)
                
                # 各資料表
                if revision_diff:
                    df = pd.DataFrame(revision_diff)
                    df.to_excel(writer, sheet_name='revision_diff', index=False)
                
                if branch_error:
                    df = pd.DataFrame(branch_error)
                    df.to_excel(writer, sheet_name='branch_error', index=False)
                
                if lost_project:
                    df = pd.DataFrame(lost_project)
                    df.to_excel(writer, sheet_name='lost_project', index=False)
                
                if version_diff:
                    df = pd.DataFrame(version_diff)
                    df.to_excel(writer, sheet_name='version_diff', index=False)
                
                if cannot_compare:
                    df = pd.DataFrame(cannot_compare)
                    df.to_excel(writer, sheet_name='無法比對', index=False)
                
                # 格式化所有工作表
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    self.excel_handler._format_worksheet(worksheet)
            
            self.logger.info(f"成功寫入情境摘要報表: {output_file}")
            
        except Exception as e:
            self.logger.error(f"寫入情境摘要報表失敗: {str(e)}")
            raise

    def _get_scenario_display_name(self, scenario):
        """取得情境的顯示名稱"""
        name_map = {
            'master_vs_premp': 'Master vs PreMP',
            'premp_vs_wave': 'PreMP vs Wave',
            'wave_vs_backup': 'Wave vs Backup'
        }
        return name_map.get(scenario, scenario)

    def _compare_with_mapping(self, source_dir: str, output_dir: str) -> Dict[str, Any]:
        """
        使用 mapping table 執行比對
        """
        # 確保輸出目錄存在
        if not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
            self.logger.info(f"建立輸出目錄: {output_dir}")
        
        # 初始化結果
        all_results = {
            'master_vs_premp': {
                'success': 0,
                'failed': 0,
                'modules': [],
                'failed_modules': [],
                'reports': [],
                'summary_report': None
            },
            'premp_vs_wave': {
                'success': 0,
                'failed': 0,
                'modules': [],
                'failed_modules': [],
                'reports': [],
                'summary_report': None
            },
            'wave_vs_backup': {
                'success': 0,
                'failed': 0,
                'modules': [],
                'failed_modules': [],
                'reports': [],
                'summary_report': None
            },
            'failed': 0,
            'failed_modules': [],
            'summary_report': ''
        }
        
        # 為每個情境準備獨立的資料容器
        scenario_data = {
            'master_vs_premp': {
                'revision_diff': [],
                'branch_error': [],
                'lost_project': [],
                'version_diff': [],
                'cannot_compare': []
            },
            'premp_vs_wave': {
                'revision_diff': [],
                'branch_error': [],
                'lost_project': [],
                'version_diff': [],
                'cannot_compare': []
            },
            'wave_vs_backup': {
                'revision_diff': [],
                'branch_error': [],
                'lost_project': [],
                'version_diff': [],
                'cannot_compare': []
            }
        }
        
        try:
            # 處理每個情境
            scenarios = ['master_vs_premp', 'premp_vs_wave', 'wave_vs_backup']
            
            for scenario in scenarios:
                self.logger.info(f"處理情境: {scenario}")
                
                # 尋找適用的 mapping table
                mapping_df = None
                
                # 優先使用情境特定的 mapping table
                if scenario in self.mapping_tables:
                    mapping_df = self.mapping_tables[scenario]
                    self.logger.info(f"使用情境特定的 mapping table: {scenario}")
                # 檢查是否有 DailyBuild mapping
                elif 'dailybuild' in self.mapping_tables:
                    mapping_df = self.mapping_tables['dailybuild']
                    self.logger.info(f"使用 DailyBuild mapping table")
                # 檢查是否有 PrebuildFW mapping
                elif 'prebuild' in self.mapping_tables or 'prebuildffw' in self.mapping_tables:
                    # 修正 DataFrame 判斷邏輯
                    if 'prebuild' in self.mapping_tables:
                        mapping_df = self.mapping_tables['prebuild']
                    else:
                        mapping_df = self.mapping_tables['prebuildffw']
                    self.logger.info(f"使用 PrebuildFW mapping table")
                # 使用通用 mapping
                elif 'general' in self.mapping_tables:
                    mapping_df = self.mapping_tables['general']
                    self.logger.info(f"使用通用 mapping table")
                
                if mapping_df is not None:
                    self.logger.info(f"Mapping table 欄位: {mapping_df.columns.tolist()}")
                    self.logger.info(f"Mapping table 資料筆數: {len(mapping_df)}")
                    
                    # 根據 mapping table 找出比對對
                    comparison_pairs = self._find_comparison_pairs_from_mapping(
                        mapping_df, source_dir, scenario
                    )
                    
                    self.logger.info(f"找到 {len(comparison_pairs)} 對需要比對的資料夾")
                    
                    # 執行比對
                    for pair in comparison_pairs:
                        try:
                            self.logger.info(f"比對: {pair['base_path']} vs {pair['compare_path']}")
                            
                            # 確保路徑存在
                            if not os.path.exists(pair['base_path']) or not os.path.exists(pair['compare_path']):
                                self.logger.warning(f"路徑不存在: {pair['base_path']} 或 {pair['compare_path']}")
                                continue
                            
                            results = self._compare_specific_folders(
                                os.path.dirname(pair['base_path']),
                                os.path.basename(pair['base_path']),
                                os.path.basename(pair['compare_path']),
                                pair['module'],
                                scenario
                            )
                            
                            # 收集資料
                            scenario_data[scenario]['revision_diff'].extend(results['revision_diff'])
                            scenario_data[scenario]['branch_error'].extend(results['branch_error'])
                            scenario_data[scenario]['lost_project'].extend(results['lost_project'])
                            if 'version_diffs' in results:
                                scenario_data[scenario]['version_diff'].extend(results['version_diffs'])
                            
                            # 記錄成功
                            all_results[scenario]['success'] += 1
                            all_results[scenario]['modules'].append(pair['module'])
                            
                            # 在寫入個別報表時
                            if any([results['revision_diff'], results['branch_error'], results['lost_project']]):
                                scenario_dir = os.path.join(output_dir, scenario)
                                module_output_dir = os.path.join(scenario_dir, pair['module'])
                                
                                # 確保目錄存在
                                if not os.path.exists(module_output_dir):
                                    os.makedirs(module_output_dir, exist_ok=True)
                                
                                compare_filename = self._generate_compare_filename(
                                    pair['module'], pair['base_folder'], pair['compare_folder']
                                )
                                
                                report_file = self._write_module_compare_report(
                                    pair['module'], results, module_output_dir, compare_filename
                                )
                                if report_file:
                                    all_results[scenario]['reports'].append(report_file)
                            
                        except Exception as e:
                            self.logger.error(f"比對 {pair['module']} ({scenario}) 失敗: {str(e)}")
                            all_results[scenario]['failed'] += 1
                            all_results[scenario]['failed_modules'].append(pair['module'])
                else:
                    self.logger.warning(f"未找到 {scenario} 的 mapping table")
            
            # 為每個情境生成獨立的 all_scenarios_compare.xlsx
            for scenario_key in scenarios:
                data = scenario_data[scenario_key]
                
                # 重新編號
                for i, item in enumerate(data['revision_diff'], 1):
                    item['SN'] = i
                for i, item in enumerate(data['branch_error'], 1):
                    item['SN'] = i
                for i, item in enumerate(data['lost_project'], 1):
                    item['SN'] = i
                for i, item in enumerate(data['version_diff'], 1):
                    item['SN'] = i
                
                # 建立情境輸出目錄
                scenario_output_dir = os.path.join(output_dir, scenario_key)
                if not os.path.exists(scenario_output_dir):
                    os.makedirs(scenario_output_dir, exist_ok=True)
                
                # 寫入該情境的整合報表
                summary_report_path = os.path.join(scenario_output_dir, 'all_scenarios_compare.xlsx')
                
                self._write_scenario_summary_report(
                    data['revision_diff'],
                    data['branch_error'],
                    data['lost_project'],
                    data['version_diff'],
                    data.get('cannot_compare', []),
                    all_results[scenario_key],
                    summary_report_path,
                    scenario_key
                )
                
                all_results[scenario_key]['summary_report'] = summary_report_path
                self.logger.info(f"為 {scenario_key} 生成摘要報告: {summary_report_path}")
            
            # 生成總摘要
            total_summary_path = os.path.join(output_dir, 'all_scenarios_summary.xlsx')
            self._write_total_summary_report(all_results, scenario_data, total_summary_path)
            all_results['summary_report'] = total_summary_path
            
            return all_results
            
        except Exception as e:
            self.logger.error(f"使用 mapping table 比對失敗: {str(e)}")
            import traceback
            self.logger.error(traceback.format_exc())
            # 如果失敗，改用原有邏輯
            self.logger.info("改用原有比對邏輯")
            return self._compare_without_mapping(source_dir, output_dir)

    def _compare_without_mapping(self, source_dir: str, output_dir: str) -> Dict[str, Any]:
        """
        不使用 mapping table 的原有比對邏輯（修改為與 mapping 版本一致的輸出格式）
        
        Args:
            source_dir: 來源目錄
            output_dir: 輸出目錄
            
        Returns:
            比對結果
        """
        output_dir = output_dir or source_dir
        
        # 初始化結果（與 _compare_with_mapping 一致）
        all_results = {
            'master_vs_premp': {
                'success': 0,
                'failed': 0,
                'modules': [],
                'failed_modules': [],
                'reports': [],
                'summary_report': None
            },
            'premp_vs_wave': {
                'success': 0,
                'failed': 0,
                'modules': [],
                'failed_modules': [],
                'reports': [],
                'summary_report': None
            },
            'wave_vs_backup': {
                'success': 0,
                'failed': 0,
                'modules': [],
                'failed_modules': [],
                'reports': [],
                'summary_report': None
            },
            'failed': 0,
            'failed_modules': [],
            'summary_report': ''
        }
        
        # 為每個情境準備獨立的資料容器（與 _compare_with_mapping 一致）
        scenario_data = {
            'master_vs_premp': {
                'revision_diff': [],
                'branch_error': [],
                'lost_project': [],
                'version_diff': [],
                'cannot_compare': []
            },
            'premp_vs_wave': {
                'revision_diff': [],
                'branch_error': [],
                'lost_project': [],
                'version_diff': [],
                'cannot_compare': []
            },
            'wave_vs_backup': {
                'revision_diff': [],
                'branch_error': [],
                'lost_project': [],
                'version_diff': [],
                'cannot_compare': []
            }
        }
        
        try:
            # 取得所有模組
            actual_modules = self._get_all_modules(source_dir)
            self.logger.info(f"找到 {len(actual_modules)} 個模組")
            
            # 處理每個模組
            for top_dir, module, module_path in actual_modules:
                full_module = f"{top_dir}/{module}" if top_dir else module
                
                # 執行三種比對情境
                scenarios = [
                    ('master_vs_premp', 'Master vs PreMP'),
                    ('premp_vs_wave', 'PreMP vs Wave'),
                    ('wave_vs_backup', 'Wave vs Wave.backup')
                ]
                
                module_has_comparison = False
                module_failures = {}  # 記錄每個情境的失敗原因
                
                for scenario_key, scenario_name in scenarios:
                    base_folder, compare_folder, missing_info = self._find_folders_for_comparison(
                        module_path, scenario_key
                    )
                    
                    if base_folder and compare_folder:
                        # 可以進行比對
                        try:
                            results = self._compare_specific_folders(
                                module_path, base_folder, compare_folder, full_module, scenario_key
                            )
                            
                            # 收集資料到對應的情境容器
                            scenario_data[scenario_key]['revision_diff'].extend(results['revision_diff'])
                            scenario_data[scenario_key]['branch_error'].extend(results['branch_error'])
                            scenario_data[scenario_key]['lost_project'].extend(results['lost_project'])
                            if 'version_diffs' in results:
                                scenario_data[scenario_key]['version_diff'].extend(results['version_diffs'])
                            
                            # 記錄成功
                            all_results[scenario_key]['success'] += 1
                            all_results[scenario_key]['modules'].append(module)
                            module_has_comparison = True
                            
                            # 寫入個別報表
                            if any([results['revision_diff'], results['branch_error'], results['lost_project']]):
                                # 建立輸出路徑
                                scenario_dir = os.path.join(output_dir, scenario_key)
                                if top_dir:
                                    module_output_dir = os.path.join(scenario_dir, top_dir, module)
                                else:
                                    module_output_dir = os.path.join(scenario_dir, module)
                                
                                if not os.path.exists(module_output_dir):
                                    os.makedirs(module_output_dir, exist_ok=True)
                                
                                # 生成檔案名稱
                                compare_filename = self._generate_compare_filename(
                                    module, base_folder, compare_folder
                                )
                                
                                report_file = self._write_module_compare_report(
                                    module, results, module_output_dir, compare_filename
                                )
                                if report_file:
                                    all_results[scenario_key]['reports'].append(report_file)
                                    
                        except Exception as e:
                            self.logger.error(f"比對 {module} ({scenario_name}) 失敗: {str(e)}")
                            all_results[scenario_key]['failed'] += 1
                            all_results[scenario_key]['failed_modules'].append(module)
                            module_failures[scenario_key] = f"比對失敗: {str(e)}"
                    else:
                        # 無法比對 - 記錄到對應情境
                        all_results[scenario_key]['failed'] += 1
                        all_results[scenario_key]['failed_modules'].append(module)
                        module_failures[scenario_key] = missing_info
                        
                        # 將無法比對的模組記錄到對應情境的 cannot_compare
                        folders = [f for f in os.listdir(module_path) 
                                if os.path.isdir(os.path.join(module_path, f))]
                        
                        if folders:
                            full_path = os.path.join(module_path, folders[0])
                        else:
                            full_path = module_path
                        
                        cannot_compare_item = {
                            'SN': len(scenario_data[scenario_key]['cannot_compare']) + 1,
                            'module': module,
                            'location_path': module_path,
                            'folder_count': len(folders),
                            'folders': ', '.join(folders) if folders else '無資料夾',
                            'path': full_path,
                            'reason': missing_info
                        }
                        scenario_data[scenario_key]['cannot_compare'].append(cannot_compare_item)
                
                # 如果這個模組完全無法比對任何情境
                if not module_has_comparison:
                    all_results['failed'] += 1
                    all_results['failed_modules'].append(module)
            
            # 為每個情境生成獨立的 all_scenarios_compare.xlsx
            for scenario_key in ['master_vs_premp', 'premp_vs_wave', 'wave_vs_backup']:
                data = scenario_data[scenario_key]
                
                # 重新編號
                for i, item in enumerate(data['revision_diff'], 1):
                    item['SN'] = i
                for i, item in enumerate(data['branch_error'], 1):
                    item['SN'] = i
                for i, item in enumerate(data['lost_project'], 1):
                    item['SN'] = i
                for i, item in enumerate(data['version_diff'], 1):
                    item['SN'] = i
                for i, item in enumerate(data['cannot_compare'], 1):
                    item['SN'] = i
                
                # 建立情境輸出目錄
                scenario_output_dir = os.path.join(output_dir, scenario_key)
                if not os.path.exists(scenario_output_dir):
                    os.makedirs(scenario_output_dir, exist_ok=True)
                
                # 寫入該情境的整合報表
                summary_report_path = os.path.join(scenario_output_dir, 'all_scenarios_compare.xlsx')
                
                self._write_scenario_summary_report(
                    data['revision_diff'],
                    data['branch_error'],
                    data['lost_project'],
                    data['version_diff'],
                    data.get('cannot_compare', []),
                    all_results[scenario_key],
                    summary_report_path,
                    scenario_key
                )
                
                all_results[scenario_key]['summary_report'] = summary_report_path
                self.logger.info(f"為 {scenario_key} 生成摘要報告: {summary_report_path}")
            
            # 生成總摘要
            total_summary_path = os.path.join(output_dir, 'all_scenarios_summary.xlsx')
            self._write_total_summary_report(all_results, scenario_data, total_summary_path)
            all_results['summary_report'] = total_summary_path
            
            return all_results
            
        except Exception as e:
            self.logger.error(f"執行所有比對情境失敗: {str(e)}")
            raise

    def _generate_compare_filename(self, module: str, base_folder: str, compare_folder: str) -> str:
        """
        生成比較檔案名稱
        例如：bootcode_RDDB-531_RDDB-1046-premp_compare.xlsx
        """
        # 提取資料夾中的 RDDB 或 DB 編號
        base_id = base_folder
        compare_id = compare_folder
        
        # 簡化：去除後綴，只保留編號部分
        if base_folder.startswith('RDDB-'):
            base_id = base_folder.split('-premp')[0].split('-wave')[0]
        elif base_folder.startswith('DB'):
            base_id = base_folder.split('-premp')[0].split('-wave')[0]
            
        filename = f"{module}_{base_id}_{compare_folder}_compare.xlsx"
        return filename
    
    def _get_all_modules(self, source_dir: str) -> List[Tuple[str, str, str]]:
        """取得所有模組"""
        actual_modules = []
        
        # 檢查是否有 PrebuildFW 或 DailyBuild 子目錄
        has_top_dirs = False
        
        # 檢查 PrebuildFW 目錄
        prebuild_path = os.path.join(source_dir, 'PrebuildFW')
        if os.path.exists(prebuild_path) and os.path.isdir(prebuild_path):
            has_top_dirs = True
            # 取得 PrebuildFW 下的所有模組
            for module in os.listdir(prebuild_path):
                module_path = os.path.join(prebuild_path, module)
                if os.path.isdir(module_path):
                    actual_modules.append(('PrebuildFW', module, module_path))
        
        # 檢查 DailyBuild 目錄
        daily_path = os.path.join(source_dir, 'DailyBuild')
        if os.path.exists(daily_path) and os.path.isdir(daily_path):
            has_top_dirs = True
            # 取得 DailyBuild 下的所有模組（平台）
            for platform in os.listdir(daily_path):
                platform_path = os.path.join(daily_path, platform)
                if os.path.isdir(platform_path):
                    actual_modules.append(('DailyBuild', platform, platform_path))
        
        # 如果沒有頂層目錄，使用原本的邏輯
        if not has_top_dirs:
            modules = [d for d in os.listdir(source_dir) 
                      if os.path.isdir(os.path.join(source_dir, d))]
            for module in modules:
                module_path = os.path.join(source_dir, module)
                actual_modules.append(('', module, module_path))
        
        return actual_modules
    
    def _write_all_scenarios_report(self, revision_diff: List[Dict], branch_error: List[Dict],
                               lost_project: List[Dict], version_diff: List[Dict],
                               cannot_compare_modules: List[Dict], all_results: Dict,
                               output_dir: str) -> str:
        """寫入所有比對情境的整合報表"""
        try:
            import pandas as pd
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.worksheet.filters import FilterColumn, Filters, AutoFilter

            # 確保輸出目錄存在（重要修正）
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                self.logger.info(f"建立目錄: {output_dir}")
                            
            output_file = os.path.join(output_dir, "all_scenarios_compare.xlsx")
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 摘要頁籤
                summary_data = []
                
                # 各情境的統計
                scenarios = [
                    ('Master vs PreMP', all_results['master_vs_premp']),
                    ('PreMP vs Wave', all_results['premp_vs_wave']),
                    ('Wave vs Backup', all_results['wave_vs_backup'])
                ]
                
                for scenario_name, scenario_result in scenarios:
                    summary_data.append({
                        '比對情境': scenario_name,
                        '成功模組數': scenario_result['success'],
                        '失敗模組數': scenario_result['failed'],
                        '成功模組清單': ', '.join(scenario_result['modules']) if scenario_result['modules'] else '',
                        '失敗模組清單': ', '.join(scenario_result['failed_modules']) if scenario_result['failed_modules'] else ''
                    })
                
                # 加入總計
                total_modules = len(set(
                    all_results['master_vs_premp']['modules'] +
                    all_results['premp_vs_wave']['modules'] +
                    all_results['wave_vs_backup']['modules']
                ))
                
                total_failed = len(set(
                    all_results['master_vs_premp']['failed_modules'] +
                    all_results['premp_vs_wave']['failed_modules'] +
                    all_results['wave_vs_backup']['failed_modules']
                ))
                
                summary_data.append({
                    '比對情境': '總計',
                    '成功模組數': total_modules,
                    '失敗模組數': total_failed,
                    '成功模組清單': str(total_modules),
                    '失敗模組清單': str(total_failed)
                })
                
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='摘要', index=False)
                
                # 格式化摘要頁籤
                worksheet_summary = writer.sheets['摘要']
                
                # 設定標題列格式（深藍底白字）
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                # 格式化標題列
                for cell in worksheet_summary[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 設定總計列的格式（淺藍底）
                total_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                total_font = Font(bold=True)
                
                # 找到總計列（最後一列）
                total_row = len(df_summary) + 1  # +1 因為第一行是標題
                for col in range(1, len(df_summary.columns) + 1):
                    cell = worksheet_summary.cell(row=total_row, column=col)
                    cell.fill = total_fill
                    cell.font = total_font
                    # 總計列的數字也要置中
                    if col in [2, 3]:  # 成功模組數和失敗模組數欄位
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 設定所有資料列的對齊方式
                for row_idx in range(2, total_row + 1):  # 包含總計列
                    # 比對情境欄位（第1欄）靠左
                    worksheet_summary.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left', vertical='center')
                    
                    # 成功模組數和失敗模組數（第2、3欄）置中
                    worksheet_summary.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center', vertical='center')
                    worksheet_summary.cell(row=row_idx, column=3).alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 成功模組清單和失敗模組清單（第4、5欄）靠左
                    if row_idx < total_row:  # 非總計列
                        worksheet_summary.cell(row=row_idx, column=4).alignment = Alignment(horizontal='left', vertical='center')
                        worksheet_summary.cell(row=row_idx, column=5).alignment = Alignment(horizontal='left', vertical='center')
                    else:  # 總計列的清單欄位也置中
                        worksheet_summary.cell(row=row_idx, column=4).alignment = Alignment(horizontal='center', vertical='center')
                        worksheet_summary.cell(row=row_idx, column=5).alignment = Alignment(horizontal='center', vertical='center')
                
                # 調整欄寬
                worksheet_summary.column_dimensions['A'].width = 20  # 比對情境
                worksheet_summary.column_dimensions['B'].width = 15  # 成功模組數
                worksheet_summary.column_dimensions['C'].width = 15  # 失敗模組數
                worksheet_summary.column_dimensions['D'].width = 60  # 成功模組清單
                worksheet_summary.column_dimensions['E'].width = 60  # 失敗模組清單
                
                # revision_diff 頁籤
                if revision_diff:
                    df = pd.DataFrame(revision_diff)
                    columns_order = ['SN', 'module', 'location_path', 'base_folder', 'compare_folder', 'name', 'path', 
                                'base_short', 'base_revision', 'compare_short', 'compare_revision',
                                'base_upstream', 'compare_upstream', 'base_dest-branch', 'compare_dest-branch',
                                'base_link', 'compare_link']
                    columns_order = [col for col in columns_order if col in df.columns]
                    df = df[columns_order]
                    df.to_excel(writer, sheet_name='revision_diff', index=False)
                
                # branch_error 頁籤
                if branch_error:
                    df = pd.DataFrame(branch_error)
                    columns_order = ['SN', 'module', 'location_path', 'base_folder', 'compare_folder', 'name', 'path', 
                                'revision_short', 'revision', 'upstream', 'dest-branch', 
                                'compare_link', 'problem', 'has_wave']
                    columns_order = [col for col in columns_order if col in df.columns]
                    df = df[columns_order]
                    df_sorted = df.sort_values('has_wave', ascending=True)
                    df_sorted.to_excel(writer, sheet_name='branch_error', index=False)
                
                # lost_project 頁籤
                if lost_project:
                    df = pd.DataFrame(lost_project)
                    columns_order = ['SN', 'Base folder', '狀態', 'module', 'location_path', 'folder', 'name', 'path', 
                                'upstream', 'dest-branch', 'revision', 'link']
                    columns_order = [col for col in columns_order if col in df.columns]
                    df = df[columns_order]
                    df.to_excel(writer, sheet_name='lost_project', index=False)
                
                # version_diff 頁籤
                if version_diff:
                    df = pd.DataFrame(version_diff)
                    # 移除 module_path 欄位
                    if 'module_path' in df.columns:
                        df = df.drop('module_path', axis=1)
                    # 移除 is_different 欄位（如果存在）
                    if 'is_different' in df.columns:
                        df = df.drop('is_different', axis=1)
                    
                    # 確保欄位順序正確
                    columns_order = ['SN', 'module', 'location_path', 'base_folder', 'compare_folder', 'file_type', 
                                'base_content', 'compare_content', 'org_content']
                    
                    # 只保留存在的欄位
                    columns_order = [col for col in columns_order if col in df.columns]
                    
                    # 重新排序欄位
                    df = df.reindex(columns=columns_order)
                    
                    df.to_excel(writer, sheet_name='version_diff', index=False)
                
                # 無法比對頁籤
                if cannot_compare_modules:
                    df = pd.DataFrame(cannot_compare_modules)
                    columns_order = ['SN', 'module', 'location_path', 'folder_count', 'folders', 'path', 'reason']
                    columns_order = [col for col in columns_order if col in df.columns]
                    df = df[columns_order]
                    df.to_excel(writer, sheet_name='無法比對', index=False)
                
                # 格式化所有工作表
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    self.excel_handler._format_worksheet(worksheet)
                
                # 套用特定格式和篩選
                self._apply_special_formatting_and_filters(writer, revision_diff, branch_error, 
                                            lost_project, version_diff)
            
            self.logger.info(f"成功寫入所有情境整合報表: {output_file}")
            return output_file
            
        except Exception as e:
            self.logger.error(f"寫入所有情境整合報表失敗: {str(e)}")
            raise
    
    def _apply_special_formatting_and_filters(self, writer, revision_diff, branch_error, 
                        lost_project, version_diff):
        """套用特定欄位的格式和篩選器"""
        import pandas as pd
        from openpyxl.styles import PatternFill, Font
        from openpyxl.worksheet.filters import FilterColumn, Filters, AutoFilter
        
        # revision_diff 頁籤的特定格式
        if revision_diff and 'revision_diff' in writer.sheets:
            worksheet = writer.sheets['revision_diff']
            df = pd.DataFrame(revision_diff)
            
            # 設定深紅底標題的欄位
            header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            white_font = Font(color="FFFFFF", bold=True)
            red_font = Font(color="FF0000")
            
            # 找到需要格式化的欄位位置
            target_columns = ['base_short', 'base_revision', 'compare_short', 'compare_revision']
            column_indices = {}
            
            for idx, col in enumerate(df.columns):
                if col in target_columns:
                    column_indices[col] = idx + 1
            
            # 設定標題為深紅底白字
            for col_name, col_idx in column_indices.items():
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = white_font
            
            # 設定內容為紅字
            for row in range(2, len(df) + 2):
                for col_name, col_idx in column_indices.items():
                    worksheet.cell(row=row, column=col_idx).font = red_font
        
        # branch_error 頁籤的特定格式和篩選
        if branch_error and 'branch_error' in writer.sheets:
            worksheet = writer.sheets['branch_error']
            df = pd.DataFrame(branch_error)
            
            # 找到 "problem" 和 "has_wave" 欄位的位置
            problem_col = None
            has_wave_col = None
            for idx, col in enumerate(df.columns):
                if col == 'problem':
                    problem_col = idx
                    print("problem_col:", problem_col)
                elif col == 'has_wave':
                    has_wave_col = idx + 2
                    print("has_wave_col:", has_wave_col)
            
            if problem_col:
                # 設定深紅底白字
                header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                white_font = Font(color="FFFFFF", bold=True)
                red_font = Font(color="FF0000")
                
                # 設定標題
                cell = worksheet.cell(row=1, column=problem_col)
                cell.fill = header_fill
                cell.font = white_font
                
                # 設定 problem 欄位有內容的儲存格為紅字
                for row_idx in range(2, worksheet.max_row + 1):  # 使用 worksheet.max_row
                    cell_value = worksheet.cell(row=row_idx, column=problem_col).value
                    if cell_value and str(cell_value).strip():  # 如果有內容
                        worksheet.cell(row=row_idx, column=problem_col).font = red_font
            
            # 設定自動篩選
            worksheet.auto_filter.ref = worksheet.dimensions

            # 設定 has_wave 欄位的篩選條件為只顯示 "N"
            if has_wave_col:
                # 取得所有唯一的 has_wave 值
                has_wave_values = df['has_wave'].unique().tolist()
                
                # 建立篩選器，注意 colId 是從 0 開始的
                has_wave_df_index = df.columns.get_loc('has_wave') + 1
                print("has_wave_df_index:", has_wave_df_index)
                filter_column = FilterColumn(colId=has_wave_df_index)
                filter_column.filters = Filters()
                filter_column.filters.filter = ['N']
                
                # 如果有 'Y' 值，需要將其設為隱藏
                if 'Y' in has_wave_values:
                    # 直接根據工作表中的值來隱藏
                    for row_idx in range(2, worksheet.max_row + 1):
                        if worksheet.cell(row=row_idx, column=has_wave_col).value == 'Y':
                            worksheet.row_dimensions[row_idx].hidden = True
                
                # 將篩選器加入自動篩選
                worksheet.auto_filter.filterColumn.append(filter_column)
        
        # lost_project 頁籤的特定格式
        if lost_project and 'lost_project' in writer.sheets:
            worksheet = writer.sheets['lost_project']
            df = pd.DataFrame(lost_project)
            
            # 定義顏色
            red_font = Font(color="FF0000")  # 紅色
            blue_font = Font(color="0000FF")  # 藍色
            
            # 找到 "Base folder" 和 "狀態" 欄位的位置
            base_folder_col = None
            status_col = None
            for idx, col in enumerate(df.columns):
                if col == 'Base folder':
                    base_folder_col = idx + 1
                    # 設定深紅底白字
                    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                    white_font = Font(color="FFFFFF", bold=True)
                    cell = worksheet.cell(row=1, column=base_folder_col)
                    cell.fill = header_fill
                    cell.font = white_font
                elif col == '狀態':
                    status_col = idx + 1
                    # 設定深紅底白字
                    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                    white_font = Font(color="FFFFFF", bold=True)
                    cell = worksheet.cell(row=1, column=status_col)
                    cell.fill = header_fill
                    cell.font = white_font
            
            # 設定狀態欄位的文字顏色
            if status_col:
                for row_idx in range(2, len(df) + 2):
                    df_row_idx = row_idx - 2
                    if '狀態' in df.columns:
                        status_value = df.iloc[df_row_idx]['狀態']
                        if status_value == '刪除':
                            worksheet.cell(row=row_idx, column=status_col).font = red_font
                        elif status_value == '新增':
                            worksheet.cell(row=row_idx, column=status_col).font = blue_font
        
        # version_diff 頁籤的特定格式
        if version_diff and 'version_diff' in writer.sheets:
            worksheet = writer.sheets['version_diff']
            df = pd.DataFrame(version_diff)
            
            from openpyxl.cell.text import InlineFont
            from openpyxl.cell.rich_text import TextBlock, CellRichText
            
            # 設定深紅底白字的標題
            header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            white_font = Font(color="FFFFFF", bold=True)
            red_font = Font(color="FF0000")
            
            # 找到需要格式化的欄位位置
            target_columns = ['base_content', 'compare_content']
            column_indices = {}
            
            for idx, col in enumerate(df.columns):
                if col in target_columns:
                    column_indices[col] = idx + 2
            
            # 設定標題為深紅底白字（只有 base_content 和 compare_content）
            for col_name, col_idx in column_indices.items():
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = white_font
            
            # 處理內容，根據檔案類型標記不同部分
            for row_idx in range(2, len(df) + 2):
                # 讀取實際的值
                df_row_idx = row_idx - 2  # DataFrame 的索引
                
                # 取得檔案類型
                file_type = ''
                if 'file_type' in df.columns:
                    file_type = df.iloc[df_row_idx]['file_type']
                
                # 取得 base_content 和 compare_content
                base_content = ''
                compare_content = ''
                
                if 'base_content' in df.columns:
                    base_content = df.iloc[df_row_idx]['base_content']
                if 'compare_content' in df.columns:
                    compare_content = df.iloc[df_row_idx]['compare_content']
                
                # 檢查是否為檔案不存在的情況
                if str(base_content) == '(檔案不存在)':
                    # base_content 為紅字
                    if 'base_content' in column_indices:
                        worksheet.cell(row=row_idx, column=column_indices['base_content']).font = red_font
                        worksheet.cell(row=row_idx, column=column_indices['base_content']).value = base_content
                        
                    # 如果 compare_content 是 "(檔案存在)"，保持黑字
                    if 'compare_content' in column_indices and str(compare_content) == '(檔案存在)':
                        worksheet.cell(row=row_idx, column=column_indices['compare_content']).value = compare_content
                        
                elif str(compare_content) == '(檔案不存在)':
                    # compare_content 為紅字
                    if 'compare_content' in column_indices:
                        worksheet.cell(row=row_idx, column=column_indices['compare_content']).font = red_font
                        worksheet.cell(row=row_idx, column=column_indices['compare_content']).value = compare_content
                        
                    # 如果 base_content 是 "(檔案存在)"，保持黑字
                    if 'base_content' in column_indices and str(base_content) == '(檔案存在)':
                        worksheet.cell(row=row_idx, column=column_indices['base_content']).value = base_content
                else:
                    # 根據檔案類型和內容選擇處理方式
                    if str(file_type).lower() == 'f_version.txt' and base_content and compare_content:
                        # F_Version.txt: 處理 P_GIT_xxx 行 - 支援多行處理
                        base_lines = str(base_content).split('\n')
                        compare_lines = str(compare_content).split('\n')
                        
                        # 檢查是否包含 P_GIT_ 行
                        has_p_git_base = any(line.strip().startswith('P_GIT_') for line in base_lines)
                        has_p_git_compare = any(line.strip().startswith('P_GIT_') for line in compare_lines)
                        
                        if has_p_git_base or has_p_git_compare:
                            self._format_f_version_content(worksheet, row_idx, column_indices, base_content, compare_content)
                        else:
                            # 其他行不需要特殊格式化
                            pass
                    elif 'F_HASH:' in str(base_content) or 'F_HASH:' in str(compare_content):
                        # Version.txt with F_HASH
                        self._format_f_hash_content(worksheet, row_idx, column_indices, base_content, compare_content)
                    elif ':' in str(base_content) or ':' in str(compare_content):
                        # Other Version.txt with colon
                        self._format_colon_content(worksheet, row_idx, column_indices, base_content, compare_content)

    def _format_f_version_content(self, worksheet, row_idx, column_indices, base_content, compare_content):
        """格式化 F_Version.txt 內容 - 只標記 git hash 和 svn number，支援多行"""
        from openpyxl.cell.text import InlineFont
        from openpyxl.cell.rich_text import TextBlock, CellRichText
        
        # 處理 base_content - 支援多行
        if 'base_content' in column_indices and base_content:
            base_lines = str(base_content).split('\n')
            compare_lines = str(compare_content).split('\n') if compare_content else []
            
            rich_text_parts = []
            
            for line_idx, base_line in enumerate(base_lines):
                if line_idx > 0:  # 不是第一行就加換行
                    rich_text_parts.append(TextBlock(InlineFont(color="000000"), "\n"))
                
                base_line = base_line.strip()
                if base_line.startswith('P_GIT_'):
                    base_parts = base_line.split(';')
                    compare_parts = []
                    
                    # 找到對應的比較行
                    if line_idx < len(compare_lines):
                        compare_line = compare_lines[line_idx].strip()
                        if compare_line.startswith('P_GIT_'):
                            compare_parts = compare_line.split(';')
                    
                    # 處理每個部分
                    for i, part in enumerate(base_parts):
                        if i > 0:
                            rich_text_parts.append(TextBlock(InlineFont(color="000000"), ";"))
                        
                        # 只有第4和第5部分（索引3和4，即git hash 和 svn number）需要比較
                        if i in [3, 4] and i < len(compare_parts) and part != compare_parts[i]:
                            rich_text_parts.append(TextBlock(InlineFont(color="FF0000"), part))
                        else:
                            rich_text_parts.append(TextBlock(InlineFont(color="000000"), part))
                else:
                    # 非 P_GIT_ 行，保持原樣
                    rich_text_parts.append(TextBlock(InlineFont(color="000000"), base_line))
            
            cell_rich_text = CellRichText(rich_text_parts)
            worksheet.cell(row=row_idx, column=column_indices['base_content']).value = cell_rich_text
        
        # 處理 compare_content - 支援多行
        if 'compare_content' in column_indices and compare_content:
            base_lines = str(base_content).split('\n') if base_content else []
            compare_lines = str(compare_content).split('\n')
            
            rich_text_parts = []
            
            for line_idx, compare_line in enumerate(compare_lines):
                if line_idx > 0:  # 不是第一行就加換行
                    rich_text_parts.append(TextBlock(InlineFont(color="000000"), "\n"))
                
                compare_line = compare_line.strip()
                if compare_line.startswith('P_GIT_'):
                    compare_parts = compare_line.split(';')
                    base_parts = []
                    
                    # 找到對應的基準行
                    if line_idx < len(base_lines):
                        base_line = base_lines[line_idx].strip()
                        if base_line.startswith('P_GIT_'):
                            base_parts = base_line.split(';')
                    
                    # 處理每個部分
                    for i, part in enumerate(compare_parts):
                        if i > 0:
                            rich_text_parts.append(TextBlock(InlineFont(color="000000"), ";"))
                        
                        # 只有第4和第5部分（索引3和4，即git hash 和 svn number）需要比較
                        if i in [3, 4] and i < len(base_parts) and part != base_parts[i]:
                            rich_text_parts.append(TextBlock(InlineFont(color="FF0000"), part))
                        else:
                            rich_text_parts.append(TextBlock(InlineFont(color="000000"), part))
                else:
                    # 非 P_GIT_ 行，保持原樣
                    rich_text_parts.append(TextBlock(InlineFont(color="000000"), compare_line))
            
            cell_rich_text = CellRichText(rich_text_parts)
            worksheet.cell(row=row_idx, column=column_indices['compare_content']).value = cell_rich_text

    def _format_f_hash_content(self, worksheet, row_idx, column_indices, base_content, compare_content):
        """格式化 F_HASH 內容 - 只標記 hash 值"""
        from openpyxl.cell.text import InlineFont
        from openpyxl.cell.rich_text import TextBlock, CellRichText
        
        # 處理 base_content
        if 'base_content' in column_indices and base_content and 'F_HASH:' in str(base_content):
            parts = str(base_content).split('F_HASH:', 1)
            if len(parts) == 2:
                hash1 = parts[1].strip()
                hash2 = ''
                if compare_content and 'F_HASH:' in str(compare_content):
                    compare_parts = str(compare_content).split('F_HASH:', 1)
                    if len(compare_parts) == 2:
                        hash2 = compare_parts[1].strip()
                
                rich_text_parts = [
                    TextBlock(InlineFont(color="000000"), "F_HASH: "),
                    TextBlock(InlineFont(color="FF0000" if hash1 != hash2 else "000000"), hash1)
                ]
                
                cell_rich_text = CellRichText(rich_text_parts)
                worksheet.cell(row=row_idx, column=column_indices['base_content']).value = cell_rich_text
        
        # 處理 compare_content
        if 'compare_content' in column_indices and compare_content and 'F_HASH:' in str(compare_content):
            parts = str(compare_content).split('F_HASH:', 1)
            if len(parts) == 2:
                hash2 = parts[1].strip()
                hash1 = ''
                if base_content and 'F_HASH:' in str(base_content):
                    base_parts = str(base_content).split('F_HASH:', 1)
                    if len(base_parts) == 2:
                        hash1 = base_parts[1].strip()
                
                rich_text_parts = [
                    TextBlock(InlineFont(color="000000"), "F_HASH: "),
                    TextBlock(InlineFont(color="FF0000" if hash1 != hash2 else "000000"), hash2)
                ]
                
                cell_rich_text = CellRichText(rich_text_parts)
                worksheet.cell(row=row_idx, column=column_indices['compare_content']).value = cell_rich_text

    def _format_colon_content(self, worksheet, row_idx, column_indices, base_content, compare_content):
        """格式化包含冒號的內容 - 只標記值的部分"""
        from openpyxl.cell.text import InlineFont
        from openpyxl.cell.rich_text import TextBlock, CellRichText
        
        # 處理 base_content
        if 'base_content' in column_indices and base_content and ':' in str(base_content):
            parts = str(base_content).split(':', 1)
            if len(parts) == 2:
                key = parts[0]
                value1 = parts[1].strip()
                value2 = ''
                
                if compare_content and ':' in str(compare_content):
                    compare_parts = str(compare_content).split(':', 1)
                    if len(compare_parts) == 2 and compare_parts[0] == key:
                        value2 = compare_parts[1].strip()
                
                rich_text_parts = [
                    TextBlock(InlineFont(color="000000"), key + ": "),
                    TextBlock(InlineFont(color="FF0000" if value1 != value2 else "000000"), value1)
                ]
                
                cell_rich_text = CellRichText(rich_text_parts)
                worksheet.cell(row=row_idx, column=column_indices['base_content']).value = cell_rich_text
        
        # 處理 compare_content
        if 'compare_content' in column_indices and compare_content and ':' in str(compare_content):
            parts = str(compare_content).split(':', 1)
            if len(parts) == 2:
                key = parts[0]
                value2 = parts[1].strip()
                value1 = ''
                
                if base_content and ':' in str(base_content):
                    base_parts = str(base_content).split(':', 1)
                    if len(base_parts) == 2 and base_parts[0] == key:
                        value1 = base_parts[1].strip()
                
                rich_text_parts = [
                    TextBlock(InlineFont(color="000000"), key + ": "),
                    TextBlock(InlineFont(color="FF0000" if value1 != value2 else "000000"), value2)
                ]
                
                cell_rich_text = CellRichText(rich_text_parts)
                worksheet.cell(row=row_idx, column=column_indices['compare_content']).value = cell_rich_text
                        
    def compare_module_folders(self, module_path: str, base_folder_suffix: str = None) -> Dict[str, Any]:
        """
        比較模組下的兩個資料夾
        
        Args:
            module_path: 模組路徑
            base_folder_suffix: 指定要作為 base 的資料夾後綴（如 "wave", "premp"）
            
        Returns:
            比較結果
        """
        results = {
            'module': os.path.basename(module_path),
            'revision_diff': [],
            'branch_error': [],
            'lost_project': [],
            'text_file_differences': {},
            'base_folder': '',
            'compare_folder': ''
        }
        
        try:
            # 取得模組下的所有資料夾
            folders = [f for f in os.listdir(module_path) 
                      if os.path.isdir(os.path.join(module_path, f))]
            
            if len(folders) < 2:
                return results
            
            # 根據使用者選擇決定 base 和 compare 資料夾
            base_folder = None
            compare_folder = None
            
            if base_folder_suffix:
                # 尋找符合後綴的資料夾作為 base
                for folder in folders:
                    if folder.endswith(f"-{base_folder_suffix}"):
                        base_folder = folder
                        break
                    elif base_folder_suffix == "default" and not any(folder.endswith(suffix) for suffix in ["-wave", "-premp", "-wave.backup"]):
                        base_folder = folder
                        break
                        
                # 找出 compare 資料夾（另一個資料夾）
                if base_folder:
                    for folder in folders:
                        if folder != base_folder:
                            compare_folder = folder
                            break
            
            # 如果沒有找到或沒有指定，使用前兩個資料夾
            if not base_folder or not compare_folder:
                base_folder = folders[0]
                compare_folder = folders[1]
            
            folder1_path = os.path.join(module_path, base_folder)
            folder2_path = os.path.join(module_path, compare_folder)
            
            self.logger.info(f"比較資料夾: {base_folder} (base) vs {compare_folder} (compare)")
            results['base_folder'] = base_folder
            results['compare_folder'] = compare_folder
            
            # 比較每個目標檔案
            for target_file in config.TARGET_FILES:
                file1 = utils.find_file_case_insensitive(folder1_path, target_file)
                file2 = utils.find_file_case_insensitive(folder2_path, target_file)
                
                if file1 and file2:
                    if target_file.lower() == 'manifest.xml':
                        # 比較 manifest.xml（傳入 module_path）
                        revision_diff, branch_error, lost_project = self._compare_manifest_files(
                            file1, file2, results['module'], base_folder, compare_folder, module_path
                        )
                        results['revision_diff'] = revision_diff
                        results['branch_error'] = branch_error
                        results['lost_project'] = lost_project
                    else:
                        # 比較文字檔案（使用新的比較規則）
                        differences = self._compare_text_files(file1, file2, target_file)
                        if differences:
                            results['text_file_differences'][target_file] = differences
                            
                            # 為整合報表準備版本檔案差異資料
                            if target_file.lower() in ['version.txt', 'f_version.txt']:
                                if 'version_diffs' not in results:
                                    results['version_diffs'] = []
                                for diff in differences:
                                    version_diff_item = {
                                        'module': self._extract_simple_module_name(results['module']),
                                        'location_path': module_path,
                                        'base_folder': base_folder,
                                        'compare_folder': compare_folder,
                                        'file_type': target_file,  # 這裡應該是檔案類型，如 "F_Version.txt" 或 "Version.txt"
                                        'base_content': diff.get('file1', ''),  # 這裡應該是差異內容
                                        'compare_content': diff.get('file2', ''),  # 這裡應該是差異內容
                                        'org_content': diff.get('content1', '')  # 這裡是原始完整內容
                                    }
                                    results['version_diffs'].append(version_diff_item)
                else:
                    self.logger.warning(f"檔案 {target_file} 在一個或兩個資料夾中都不存在")
                    
        except Exception as e:
            self.logger.error(f"比較模組資料夾失敗: {str(e)}")
            
        return results
        
    def compare_all_modules(self, source_dir: str, output_dir: str = None, compare_mode: str = None) -> List[str]:
        """
        比較所有模組
        
        Args:
            source_dir: 來源目錄
            output_dir: 輸出目錄
            compare_mode: 比對模式
            
        Returns:
            產生的比較報表檔案列表
        """
        output_dir = output_dir or source_dir

        # 確保輸出目錄存在（重要修正）
        if not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
            self.logger.info(f"建立輸出目錄: {output_dir}")
            
        compare_files = []
        
        # 用於整合報表的資料
        all_revision_diff = []
        all_branch_error = []
        all_lost_project = []
        all_version_diff = []  # 新增：版本檔案差異
        cannot_compare_modules = []  # 記錄無法比對的模組
        
        try:
            # 取得所有模組
            actual_modules = self._get_all_modules(source_dir)
            self.logger.info(f"找到 {len(actual_modules)} 個模組")
            
            # 處理每個模組
            for top_dir, module, module_path in actual_modules:
                # 組合完整模組名稱
                full_module = f"{top_dir}/{module}" if top_dir else module
                
                # 根據比對模式找出需要比對的資料夾
                base_folder, compare_folder, missing_info = self._find_folders_for_comparison(
                    module_path, compare_mode
                )
                
                if not base_folder or not compare_folder:
                    # 記錄無法比對的原因
                    folders = [f for f in os.listdir(module_path) 
                              if os.path.isdir(os.path.join(module_path, f))]
                    
                    # 如果有資料夾，使用第一個資料夾的完整路徑
                    if folders:
                        full_path = os.path.join(module_path, folders[0])
                    else:
                        full_path = module_path
                        
                    self.logger.warning(f"模組 {module} 無法進行比對")
                    cannot_compare_modules.append({
                        'SN': len(cannot_compare_modules) + 1,
                        'module': module,  # 簡化的模組名稱
                        'location_path': module_path,
                        'folder_count': len(folders),
                        'folders': ', '.join(folders) if folders else '無資料夾',
                        'path': full_path,
                        'reason': missing_info
                    })
                    continue
                
                # 比較模組
                results = self._compare_specific_folders(
                    module_path, base_folder, compare_folder, full_module, compare_mode
                )
                
                # 收集所有資料（模組名稱已經在 _compare_manifest_files 中簡化）
                all_revision_diff.extend(results['revision_diff'])
                all_branch_error.extend(results['branch_error'])
                all_lost_project.extend(results['lost_project'])
                
                # 收集版本檔案差異
                if 'version_diffs' in results:
                    all_version_diff.extend(results['version_diffs'])
                
                # 如果有比較結果，寫入模組報表
                if any([results['revision_diff'], results['branch_error'], results['lost_project']]):
                    # 建立輸出路徑，保持目錄結構
                    if top_dir:
                        module_output_dir = os.path.join(output_dir, top_dir, module)
                    else:
                        module_output_dir = os.path.join(output_dir, module)
                    
                    # 確保目錄存在（使用 exist_ok=True）
                    if not os.path.exists(module_output_dir):
                        os.makedirs(module_output_dir, exist_ok=True)
                    
                    # 生成檔案名稱
                    compare_filename = self._generate_compare_filename(
                        module, base_folder, compare_folder
                    )
                        
                    report_file = self._write_module_compare_report(
                        module, results, module_output_dir, compare_filename
                    )
                    if report_file:
                        compare_files.append(report_file)
                        
            # 寫入整合報表（包含無法比對的模組）
            if any([all_revision_diff, all_branch_error, all_lost_project, all_version_diff, cannot_compare_modules]):
                self._write_all_compare_report(
                    all_revision_diff, all_branch_error, all_lost_project, 
                    all_version_diff, cannot_compare_modules, output_dir
                )
                
            return compare_files
            
        except Exception as e:
            self.logger.error(f"比較所有模組失敗: {str(e)}")
            raise
            
    def _find_folders_for_comparison(self, module_path: str, compare_mode: str) -> Tuple[str, str, str]:
        """
        根據比對模式找出需要比對的資料夾
        
        Args:
            module_path: 模組路徑
            compare_mode: 比對模式
            
        Returns:
            (base_folder, compare_folder, missing_info)
        """
        folders = [f for f in os.listdir(module_path) 
                   if os.path.isdir(os.path.join(module_path, f))]
        
        # 判斷是 RDDB 還是 DB 格式
        is_rddb_format = any(folder.startswith('RDDB-') for folder in folders)
        
        # 根據資料夾名稱分類
        master_folder = None
        premp_folder = None
        wave_folder = None
        backup_folder = None
        
        for folder in folders:
            if folder.endswith('-wave.backup'):
                backup_folder = folder
            elif folder.endswith('-wave'):
                wave_folder = folder
            elif folder.endswith('-premp'):
                premp_folder = folder
            else:
                # 沒有後綴的是 master
                if is_rddb_format and folder.startswith('RDDB-'):
                    master_folder = folder
                elif not is_rddb_format and folder.startswith('DB'):
                    master_folder = folder
        
        # 根據比對模式決定要比對的資料夾
        base_folder = None
        compare_folder = None
        missing_info = ""
        
        if compare_mode == 'master_vs_premp':
            base_folder = master_folder
            compare_folder = premp_folder
            if not master_folder:
                if is_rddb_format:
                    missing_info = "缺少 master 資料夾 (RDDB-XXX)"
                else:
                    missing_info = "缺少 master 資料夾 (DBXXXX)"
            elif not premp_folder:
                if is_rddb_format:
                    missing_info = "缺少 premp 資料夾 (RDDB-XXX-premp)"
                else:
                    missing_info = "缺少 premp 資料夾 (DBXXXX-premp)"
                
        elif compare_mode == 'premp_vs_wave':
            base_folder = premp_folder
            compare_folder = wave_folder
            if not premp_folder:
                if is_rddb_format:
                    missing_info = "缺少 premp 資料夾 (RDDB-XXX-premp)"
                else:
                    missing_info = "缺少 premp 資料夾 (DBXXXX-premp)"
            elif not wave_folder:
                if is_rddb_format:
                    missing_info = "缺少 wave 資料夾 (RDDB-XXX-wave)"
                else:
                    missing_info = "缺少 wave 資料夾 (DBXXXX-wave)"
                
        elif compare_mode == 'wave_vs_backup':
            base_folder = wave_folder
            compare_folder = backup_folder
            if not wave_folder:
                if is_rddb_format:
                    missing_info = "缺少 wave 資料夾 (RDDB-XXX-wave)"
                else:
                    missing_info = "缺少 wave 資料夾 (DBXXXX-wave)"
            elif not backup_folder:
                if is_rddb_format:
                    missing_info = "缺少 wave.backup 資料夾 (RDDB-XXX-wave.backup)"
                else:
                    missing_info = "缺少 wave.backup 資料夾 (DBXXXX-wave.backup)"
                
        elif compare_mode and compare_mode.startswith('manual_'):
            # 手動模式解析
            parts = compare_mode.split('_')
            if len(parts) >= 4:
                base_type = parts[1]
                compare_type = parts[3]
                
                # 根據類型找資料夾
                type_to_folder = {
                    'master': master_folder,
                    'premp': premp_folder,
                    'wave': wave_folder,
                    'wave.backup': backup_folder
                }
                
                base_folder = type_to_folder.get(base_type)
                compare_folder = type_to_folder.get(compare_type)
                
                if not base_folder:
                    missing_info = f"缺少 {base_type} 資料夾"
                elif not compare_folder:
                    missing_info = f"缺少 {compare_type} 資料夾"
        
        return base_folder, compare_folder, missing_info

    def _compare_specific_folders(self, module_path: str, base_folder: str, compare_folder: str, 
                              module: str, compare_mode: str = None) -> Dict[str, Any]:
        """
        比較指定的兩個資料夾
        """
        results = {
            'module': module,
            'revision_diff': [],
            'branch_error': [],
            'lost_project': [],
            'text_file_differences': {},
            'version_diffs': [],
            'base_folder': base_folder,
            'compare_folder': compare_folder
        }
        
        try:
            folder1_path = os.path.join(module_path, base_folder)
            folder2_path = os.path.join(module_path, compare_folder)
            
            self.logger.info(f"比較資料夾: {base_folder} (base) vs {compare_folder} (compare)")
            
            # 判斷是否為 DailyBuild
            is_daily_build = 'DailyBuild' in module_path
            
            # 比較每個目標檔案
            for target_file in config.TARGET_FILES:
                # 如果是 DailyBuild 的 Version.txt，跳過
                if is_daily_build and target_file.lower() == 'version.txt':
                    self.logger.info(f"跳過 DailyBuild 的 Version.txt 比對")
                    continue
                    
                file1 = utils.find_file_case_insensitive(folder1_path, target_file)
                file2 = utils.find_file_case_insensitive(folder2_path, target_file)
                
                if file1 and file2:
                    if target_file.lower() == 'manifest.xml':
                        # 比較 manifest.xml
                        revision_diff, branch_error, lost_project = self._compare_manifest_files(
                            file1, file2, module, base_folder, compare_folder, module_path, compare_mode
                        )
                        results['revision_diff'] = revision_diff
                        results['branch_error'] = branch_error
                        results['lost_project'] = lost_project
                    else:
                        # 比較文字檔案（Version.txt 或 F_Version.txt）
                        differences = self._compare_text_files(file1, file2, target_file)
                        if differences:
                            results['text_file_differences'][target_file] = differences
                            
                            # 為整合報表準備版本檔案差異資料 - 整合同一檔案的所有差異
                            if len(differences) > 0:
                                # 將所有差異整合成一筆資料
                                all_base_lines = []
                                all_compare_lines = []
                                full_content = differences[0].get('content1', '')
                                
                                for diff in differences:
                                    all_base_lines.append(diff.get('file1', ''))
                                    all_compare_lines.append(diff.get('file2', ''))
                                
                                # 建立整合的版本差異項目
                                version_diff_item = {
                                    'module': self._extract_simple_module_name(module),
                                    'location_path': module_path,
                                    'base_folder': base_folder,
                                    'compare_folder': compare_folder,
                                    'file_type': target_file,
                                    'base_content': '\n'.join(all_base_lines),  # 合併所有 base 差異行
                                    'compare_content': '\n'.join(all_compare_lines),  # 合併所有 compare 差異行
                                    'org_content': full_content,  # 完整內容
                                    'diff_count': len(differences)  # 記錄差異行數
                                }
                                results['version_diffs'].append(version_diff_item)
                elif file1 or file2:
                    # 只有一個檔案存在的情況
                    self.logger.warning(f"檔案 {target_file} 只在一個資料夾中存在")
                    
                    # 如果是 DailyBuild 的 Version.txt，跳過
                    if is_daily_build and target_file.lower() == 'version.txt':
                        continue
                        
                    if target_file.lower() in ['version.txt', 'f_version.txt']:
                        # 處理只有一個檔案的情況
                        if file1:
                            try:
                                with open(file1, 'r', encoding='utf-8', errors='ignore') as f:
                                    content1 = f.read()
                                version_diff_item = {
                                    'module': self._extract_simple_module_name(module),
                                    'location_path': module_path,
                                    'base_folder': base_folder,
                                    'compare_folder': compare_folder,
                                    'file_type': target_file,
                                    'base_content': '(檔案存在)',
                                    'compare_content': '(檔案不存在)',
                                    'org_content': content1,
                                    'diff_count': 0
                                }
                                results['version_diffs'].append(version_diff_item)
                            except Exception as e:
                                self.logger.error(f"讀取檔案失敗 {file1}: {str(e)}")
                        elif file2:
                            try:
                                with open(file2, 'r', encoding='utf-8', errors='ignore') as f:
                                    content2 = f.read()
                                version_diff_item = {
                                    'module': self._extract_simple_module_name(module),
                                    'location_path': module_path,
                                    'base_folder': base_folder,
                                    'compare_folder': compare_folder,
                                    'file_type': target_file,
                                    'base_content': '(檔案不存在)',
                                    'compare_content': '(檔案存在)',
                                    'org_content': content2,
                                    'diff_count': 0
                                }
                                results['version_diffs'].append(version_diff_item)
                            except Exception as e:
                                self.logger.error(f"讀取檔案失敗 {file2}: {str(e)}")
                else:
                    self.logger.warning(f"檔案 {target_file} 在兩個資料夾中都不存在")
                        
        except Exception as e:
            self.logger.error(f"比較資料夾失敗: {str(e)}")
            
        return results
        
    def _write_module_compare_report(self, module: str, results: Dict, output_dir: str, filename: str = None) -> str:
        """
        寫入單一模組的比較報表（與 all_compare.xlsx 相同格式）
        """
        try:
            import pandas as pd
            from openpyxl.styles import PatternFill, Font
            from openpyxl.worksheet.filters import FilterColumn, Filters

            # 確保輸出目錄存在（重要修正）
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                self.logger.info(f"建立目錄: {output_dir}")
                            
            # 使用自訂檔名或預設檔名
            output_file = os.path.join(output_dir, filename or f"{module}_compare.xlsx")
            
            # 準備各頁籤的資料
            revision_diff = results.get('revision_diff', [])
            branch_error = results.get('branch_error', [])
            lost_project = results.get('lost_project', [])
            version_diffs = results.get('version_diffs', [])
            
            # 重新編號
            for i, item in enumerate(revision_diff, 1):
                item['SN'] = i
            for i, item in enumerate(branch_error, 1):
                item['SN'] = i
            for i, item in enumerate(lost_project, 1):
                item['SN'] = i
            for i, item in enumerate(version_diffs, 1):
                item['SN'] = i
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # revision_diff 頁籤
                if revision_diff:
                    df = pd.DataFrame(revision_diff)
                    columns_order = ['SN', 'module', 'location_path', 'base_folder', 'compare_folder', 'name', 'path', 
                                'base_short', 'base_revision', 'compare_short', 'compare_revision',
                                'base_upstream', 'compare_upstream', 'base_dest-branch', 'compare_dest-branch',
                                'base_link', 'compare_link']
                    columns_order = [col for col in columns_order if col in df.columns]
                    df = df[columns_order]
                    df.to_excel(writer, sheet_name='revision_diff', index=False)
                else:
                    # 即使沒有資料也建立空的頁籤
                    pd.DataFrame(columns=['SN', 'module', 'location_path', 'base_folder', 'compare_folder', 'name', 'path',
                                        'base_short', 'base_revision', 'compare_short', 'compare_revision']).to_excel(
                        writer, sheet_name='revision_diff', index=False)
                
                # branch_error 頁籤
                if branch_error:
                    df = pd.DataFrame(branch_error)
                    columns_order = ['SN', 'module', 'location_path', 'base_folder', 'compare_folder', 'name', 'path', 
                                'revision_short', 'revision', 'upstream', 'dest-branch', 
                                'compare_link', 'problem', 'has_wave']
                    columns_order = [col for col in columns_order if col in df.columns]
                    df = df[columns_order]
                    
                    # 將 has_wave = N 的資料排在前面
                    df_sorted = df.sort_values('has_wave', ascending=True)
                    df_sorted.to_excel(writer, sheet_name='branch_error', index=False)
                else:
                    pd.DataFrame(columns=['SN', 'module', 'location_path', 'base_folder', 'compare_folder', 'name', 'path',
                                        'problem', 'has_wave']).to_excel(
                        writer, sheet_name='branch_error', index=False)
                
                # lost_project 頁籤
                if lost_project:
                    df = pd.DataFrame(lost_project)
                    columns_order = ['SN', 'Base folder', '狀態', 'module', 'location_path', 'folder', 'name', 'path', 
                                'upstream', 'dest-branch', 'revision', 'link']
                    columns_order = [col for col in columns_order if col in df.columns]
                    df = df[columns_order]
                    df.to_excel(writer, sheet_name='lost_project', index=False)
                else:
                    pd.DataFrame(columns=['SN', 'Base folder', '狀態', 'module', 'location_path', 'folder', 'name', 'path']).to_excel(
                        writer, sheet_name='lost_project', index=False)
                
                # version_diff 頁籤 - 修正：加入 org_content
                if version_diffs:
                    df = pd.DataFrame(version_diffs)
                    # 確保欄位順序正確，org_content 在最後
                    columns_order = ['SN', 'module', 'location_path', 'base_folder', 'compare_folder', 'file_type', 
                                'base_content', 'compare_content', 'org_content']  # 加入 org_content
                    columns_order = [col for col in columns_order if col in df.columns]
                    df = df[columns_order]
                    df.to_excel(writer, sheet_name='version_diff', index=False)
                else:
                    pd.DataFrame(columns=['SN', 'module', 'location_path', 'base_folder', 'compare_folder', 'file_type',
                                        'base_content', 'compare_content', 'org_content']).to_excel(  # 加入 org_content
                        writer, sheet_name='version_diff', index=False)
                
                # 先格式化所有工作表（基本格式）
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    self.excel_handler._format_worksheet(worksheet)
                
                # 套用特定欄位的格式和篩選
                self._apply_special_formatting_and_filters(writer, revision_diff, branch_error, 
                                                    lost_project, version_diffs)
                        
            self.logger.info(f"成功寫入比較報表: {output_file}")
            return output_file
            
        except Exception as e:
            self.logger.error(f"寫入模組比較報表失敗: {str(e)}")
            return None
            
    def _write_all_compare_report(self, revision_diff: List[Dict], branch_error: List[Dict],
                             lost_project: List[Dict], version_diff: List[Dict],
                             cannot_compare_modules: List[Dict], output_dir: str) -> str:
        """寫入整合比較報表（包含所有比較結果）"""
        try:
            import pandas as pd
            from openpyxl.styles import PatternFill, Font
            from openpyxl.worksheet.filters import FilterColumn, Filters
            
            # 確保輸出目錄存在（重要修正）
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                self.logger.info(f"建立目錄: {output_dir}")
                
            output_file = os.path.join(output_dir, "all_compare.xlsx")
            
            # 重新編號
            for i, item in enumerate(revision_diff, 1):
                item['SN'] = i
            for i, item in enumerate(branch_error, 1):
                item['SN'] = i
            for i, item in enumerate(lost_project, 1):
                item['SN'] = i
            for i, item in enumerate(version_diff, 1):
                item['SN'] = i
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # revision_diff 頁籤（只在有資料時產生）
                if revision_diff:
                    df = pd.DataFrame(revision_diff)
                    # 調整欄位順序
                    columns_order = ['SN', 'module', 'location_path', 'base_folder', 'compare_folder', 'name', 'path', 
                                'base_short', 'base_revision', 'compare_short', 'compare_revision',
                                'base_upstream', 'compare_upstream', 'base_dest-branch', 'compare_dest-branch',
                                'base_link', 'compare_link']
                    columns_order = [col for col in columns_order if col in df.columns]
                    df = df[columns_order]
                    df.to_excel(writer, sheet_name='revision_diff', index=False)
                
                # branch_error 頁籤（只在有資料時產生）
                if branch_error:
                    df = pd.DataFrame(branch_error)
                    # 調整欄位順序
                    columns_order = ['SN', 'module', 'location_path', 'base_folder', 'compare_folder', 'name', 'path', 
                                'revision_short', 'revision', 'upstream', 'dest-branch', 
                                'compare_link', 'problem', 'has_wave']
                    columns_order = [col for col in columns_order if col in df.columns]
                    df = df[columns_order]
                    
                    # 將 has_wave = N 的資料排在前面
                    df_sorted = df.sort_values('has_wave', ascending=True)
                    df_sorted.to_excel(writer, sheet_name='branch_error', index=False)
                
                # lost_project 頁籤（只在有資料時產生）
                if lost_project:
                    df = pd.DataFrame(lost_project)
                    # 調整欄位順序
                    columns_order = ['SN', 'Base folder', '狀態', 'module', 'location_path', 'folder', 'name', 'path', 
                                'upstream', 'dest-branch', 'revision', 'link']
                    columns_order = [col for col in columns_order if col in df.columns]
                    df = df[columns_order]
                    df.to_excel(writer, sheet_name='lost_project', index=False)
                
                # version_diff 頁籤
                if version_diff:
                    df = pd.DataFrame(version_diff)
                    # 移除 module_path 欄位
                    if 'module_path' in df.columns:
                        df = df.drop('module_path', axis=1)
                    # 移除 is_different 欄位（如果存在）
                    if 'is_different' in df.columns:
                        df = df.drop('is_different', axis=1)
                    
                    # 確保欄位順序正確
                    columns_order = ['SN', 'module', 'location_path', 'base_folder', 'compare_folder', 'file_type', 
                                'base_content', 'compare_content', 'org_content']
                    
                    # 只保留存在的欄位
                    columns_order = [col for col in columns_order if col in df.columns]
                    
                    # 重新排序欄位
                    df = df.reindex(columns=columns_order)
                    
                    df.to_excel(writer, sheet_name='version_diff', index=False)
                
                # 無法比對頁籤（只在有資料時產生）
                if cannot_compare_modules:
                    df = pd.DataFrame(cannot_compare_modules)
                    columns_order = ['SN', 'module', 'location_path', 'folder_count', 'folders', 'path', 'reason']
                    columns_order = [col for col in columns_order if col in df.columns]
                    df = df[columns_order]
                    df.to_excel(writer, sheet_name='無法比對', index=False)
                
                # 先格式化所有工作表（基本格式）
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    self.excel_handler._format_worksheet(worksheet)
                
                # 套用特定欄位的格式和篩選
                self._apply_special_formatting_and_filters(writer, revision_diff, branch_error, 
                                                       lost_project, version_diff)
                        
            self.logger.info(f"成功寫入整合報表: {output_file}")
            return output_file
            
        except Exception as e:
            self.logger.error(f"寫入整合報表失敗: {str(e)}")
            raise

    def _load_mapping_tables(self, source_dir: str) -> Dict[str, pd.DataFrame]:
        """
        載入所有 mapping table 檔案
        
        Args:
            source_dir: 來源目錄
            
        Returns:
            包含所有 mapping table 的字典
        """
        mapping_tables = {}
        
        try:
            import glob
            import fnmatch
            
            # 搜尋所有可能的 mapping table 檔案模式
            mapping_patterns = [
                '*_mapping.xlsx',
                '*_mapping.xls'
            ]
            
            self.logger.info(f"開始在 {source_dir} 搜尋 mapping tables")
            
            # 收集所有找到的檔案（避免重複）
            found_files = set()
            
            # 使用 os.walk 搜尋所有子目錄
            for root, dirs, files in os.walk(source_dir):
                for pattern in mapping_patterns:
                    for file in files:
                        # 使用 fnmatch 進行模式匹配
                        if fnmatch.fnmatch(file.lower(), pattern.lower()):
                            file_path = os.path.join(root, file)
                            
                            # 避免重複處理
                            if file_path not in found_files:
                                found_files.add(file_path)
                                
                                try:
                                    # 讀取 Excel 檔案
                                    df = pd.read_excel(file_path)
                                    
                                    self.logger.info(f"成功讀取 mapping table: {file}")
                                    self.logger.info(f"  欄位: {df.columns.tolist()}")
                                    self.logger.info(f"  資料筆數: {len(df)}")
                                    
                                    # 根據檔名決定 mapping 類型
                                    file_base = os.path.splitext(file)[0].lower()
                                    
                                    # 判斷 mapping 類型
                                    if 'master_vs_premp' in file_base:
                                        mapping_type = 'master_vs_premp'
                                    elif 'premp_vs_wave' in file_base or 'premp_vs_mp' in file_base:
                                        mapping_type = 'premp_vs_wave'
                                    elif 'wave_vs_backup' in file_base or 'mp_vs_mpbackup' in file_base:
                                        mapping_type = 'wave_vs_backup'
                                    elif 'dailybuild' in file_base:
                                        mapping_type = 'dailybuild'
                                    elif 'prebuildffw' in file_base or 'prebuild' in file_base:
                                        mapping_type = 'prebuild'
                                    else:
                                        mapping_type = 'general'
                                    
                                    # 如果同類型已存在，根據檔名長度決定優先級（更具體的優先）
                                    if mapping_type in mapping_tables:
                                        existing_file_length = len(mapping_tables[mapping_type].attrs.get('source_file', ''))
                                        new_file_length = len(file)
                                        if new_file_length > existing_file_length:
                                            mapping_tables[mapping_type] = df
                                            df.attrs['source_file'] = file
                                            self.logger.info(f"覆蓋 mapping table 類型 {mapping_type}: {file}")
                                    else:
                                        mapping_tables[mapping_type] = df
                                        df.attrs['source_file'] = file
                                        self.logger.info(f"載入 mapping table: {file} (類型: {mapping_type})")
                                    
                                except Exception as e:
                                    self.logger.warning(f"無法讀取 mapping table {file}: {str(e)}")
                                    
        except Exception as e:
            self.logger.error(f"載入 mapping tables 失敗: {str(e)}")
            
        return mapping_tables
    
    def _find_comparison_pairs_from_mapping(self, mapping_df: pd.DataFrame, source_dir: str, scenario: str) -> List[Dict[str, Any]]:
        """
        根據 mapping table 找出需要比對的檔案對
        """
        comparison_pairs = []
        
        try:
            # 預期的欄位名稱（不區分大小寫）- 擴充欄位
            expected_columns = {
                'db_type': ['DB_Type', 'db_type', 'Type', 'type'],
                'db_info': ['DB_Info', 'db_info', 'Info', 'info', 'Module', 'module'],
                'db_folder': ['DB_Folder', 'db_folder', 'Folder', 'folder'],
                'sftp_path': ['SftpPath', 'sftp_path', 'Path', 'path', 'LocalPath', 'local_path'],
                'compare_db_type': ['compare_DB_Type', 'compare_db_type', 'Compare_Type', 'compare_type'],
                'compare_db_info': ['compare_DB_Info', 'compare_db_info', 'Compare_Info', 'compare_info'],
                'compare_db_folder': ['compare_DB_Folder', 'compare_db_folder', 'Compare_Folder', 'compare_folder'],
                'compare_sftp_path': ['compare_SftpPath', 'compare_sftp_path', 'Compare_Path', 'compare_path'],
                'module': ['Module', 'module', 'ModuleName', 'module_name']
            }
            
            # 找出實際存在的欄位名稱
            actual_columns = {}
            for key, possible_names in expected_columns.items():
                for name in possible_names:
                    if name in mapping_df.columns:
                        actual_columns[key] = name
                        break
            
            self.logger.info(f"找到的 mapping 欄位: {actual_columns}")
            self.logger.info(f"正在處理情境: {scenario}")
            
            # 遍歷每一行
            for idx, row in mapping_df.iterrows():
                try:
                    # 取得基礎資訊
                    db_type = row.get(actual_columns.get('db_type', ''), '')
                    db_info = row.get(actual_columns.get('db_info', ''), '')
                    db_folder = row.get(actual_columns.get('db_folder', ''), '')
                    sftp_path = row.get(actual_columns.get('sftp_path', ''), '')
                    module_name = row.get(actual_columns.get('module', ''), '')
                    
                    # 取得比對資訊
                    compare_db_type = row.get(actual_columns.get('compare_db_type', ''), '')
                    compare_db_info = row.get(actual_columns.get('compare_db_info', ''), '')
                    compare_db_folder = row.get(actual_columns.get('compare_db_folder', ''), '')
                    compare_sftp_path = row.get(actual_columns.get('compare_sftp_path', ''), '')
                    
                    self.logger.info(f"行 {idx}: DB_Type='{db_type}', compare_DB_Type='{compare_db_type}'")
                    self.logger.info(f"行 {idx}: DB_Info='{db_info}', DB_Folder='{db_folder}'")
                    self.logger.info(f"行 {idx}: compare_DB_Info='{compare_db_info}', compare_DB_Folder='{compare_db_folder}'")
                    
                    # 檢查是否符合當前情境
                    if not self._match_scenario(db_type, compare_db_type, scenario):
                        self.logger.info(f"行 {idx}: 不符合情境 {scenario}，跳過")
                        continue
                    
                    self.logger.info(f"行 {idx}: 符合情境 {scenario}，開始尋找本地路徑")
                    
                    # 在 source_dir 中尋找對應的本地路徑
                    base_local_path = self._find_local_path(source_dir, sftp_path, db_info, db_folder)
                    compare_local_path = self._find_local_path(source_dir, compare_sftp_path, compare_db_info, compare_db_folder)
                    
                    self.logger.info(f"行 {idx}: base_local_path = {base_local_path}")
                    self.logger.info(f"行 {idx}: compare_local_path = {compare_local_path}")
                    
                    if base_local_path and compare_local_path:
                        # 使用 module_name 或從路徑提取
                        if not module_name:
                            # 從路徑提取模組名稱
                            path_parts = base_local_path.replace('\\', '/').split('/')
                            for part in path_parts:
                                if part in ['bootcode', 'emcu', 'audio_fw', 'video_fw', 'tee', 'bl31', 'dprx_quickshow']:
                                    module_name = part
                                    break
                            
                            if not module_name:
                                module_name = db_info  # 使用 DB_Info 作為後備
                        
                        comparison_pairs.append({
                            'module': module_name,
                            'base_folder': os.path.basename(base_local_path),
                            'compare_folder': os.path.basename(compare_local_path),
                            'base_path': base_local_path,
                            'compare_path': compare_local_path,
                            'db_type': db_type,
                            'compare_db_type': compare_db_type
                        })
                        self.logger.info(f"成功建立比對對: {module_name} - {db_info} vs {compare_db_info}")
                    else:
                        self.logger.warning(f"行 {idx}: 無法找到完整的路徑對 - Base: {base_local_path}, Compare: {compare_local_path}")
                        
                except Exception as e:
                    self.logger.warning(f"處理 mapping 行 {idx} 時出錯: {str(e)}")
                    
        except Exception as e:
            self.logger.error(f"解析 mapping table 失敗: {str(e)}")
            
        self.logger.info(f"總共找到 {len(comparison_pairs)} 對需要比對的資料")
        return comparison_pairs
    
    def _match_scenario(self, db_type: str, compare_db_type: str, scenario: str) -> bool:
        """
        檢查 DB 類型是否符合比對情境
        """
        db_type = str(db_type).lower().strip()
        compare_db_type = str(compare_db_type).lower().strip()
        
        scenario_map = {
            'master_vs_premp': [
                ('master', 'premp'),
                ('main', 'premp'),
                ('release', 'premp')
            ],
            'premp_vs_wave': [
                ('premp', 'wave'),
                ('premp', 'mp'),
                ('pre-mp', 'wave'),
                ('pre-mp', 'mp')
            ],
            'wave_vs_backup': [
                ('wave', 'backup'),
                ('wave', 'wavebackup'),
                ('wave', 'wave.backup'),
                ('mp', 'mpbackup'),
                ('mp', 'mp.backup')
            ]
        }
        
        if scenario == 'all':
            # 檢查所有情境
            for scenario_pairs in scenario_map.values():
                for valid_pair in scenario_pairs:
                    if (db_type == valid_pair[0] and compare_db_type == valid_pair[1]):
                        return True
        elif scenario in scenario_map:
            # 檢查特定情境
            for valid_pair in scenario_map[scenario]:
                if (db_type == valid_pair[0] and compare_db_type == valid_pair[1]):
                    return True
                    
        return False
    
    def _find_local_path(self, source_dir: str, sftp_path: str, db_info: str, db_folder: str = None) -> Optional[str]:
        """
        根據 SFTP 路徑、DB 資訊和 DB Folder 找出本地路徑
        """
        try:
            # 策略 0: 如果有 db_folder，優先使用它來匹配
            if db_folder:
                self.logger.info(f"嘗試使用 DB_Folder 匹配: {db_folder}")
                for root, dirs, files in os.walk(source_dir):
                    for dir_name in dirs:
                        # 完全匹配 db_folder
                        if dir_name == db_folder:
                            dir_path = os.path.join(root, dir_name)
                            if self._has_target_files(dir_path):
                                self.logger.info(f"通過 DB_Folder 完全匹配找到路徑: {dir_path}")
                                return dir_path
                        # 部分匹配 - 檢查 db_folder 是否包含在 dir_name 中
                        if db_folder in dir_name or dir_name in db_folder:
                            dir_path = os.path.join(root, dir_name)
                            if self._has_target_files(dir_path):
                                self.logger.info(f"通過 DB_Folder 部分匹配找到路徑: {dir_path}")
                                return dir_path
            
            # 策略 1: 直接匹配 DB 資訊
            if db_info:
                self.logger.info(f"嘗試使用 DB_Info 匹配: {db_info}")
                for root, dirs, files in os.walk(source_dir):
                    for dir_name in dirs:
                        # 檢查目錄名稱是否包含 DB 資訊
                        if db_info in dir_name:
                            dir_path = os.path.join(root, dir_name)
                            if self._has_target_files(dir_path):
                                self.logger.info(f"通過 DB_Info 找到路徑: {dir_path}")
                                return dir_path
            
            # 策略 2: 從 SFTP 路徑提取關鍵資訊
            if sftp_path:
                self.logger.info(f"嘗試從 SFTP 路徑匹配: {sftp_path}")
                # 處理 Windows 和 Linux 路徑分隔符
                sftp_path = sftp_path.replace('\\', '/')
                path_parts = sftp_path.strip('/').split('/')
                
                # 嘗試匹配路徑中的各個部分
                for i in range(len(path_parts) - 1, -1, -1):
                    path_part = path_parts[i]
                    
                    # 跳過太短的部分
                    if len(path_part) < 3:
                        continue
                        
                    for root, dirs, files in os.walk(source_dir):
                        for dir_name in dirs:
                            # 檢查是否包含路徑部分
                            if path_part in dir_name or dir_name in path_part:
                                dir_path = os.path.join(root, dir_name)
                                if self._has_target_files(dir_path):
                                    self.logger.info(f"通過 SFTP 路徑部分 '{path_part}' 找到路徑: {dir_path}")
                                    return dir_path
            
            self.logger.warning(f"無法找到本地路徑 - DB_Info: {db_info}, DB_Folder: {db_folder}, SFTP: {sftp_path}")
            
        except Exception as e:
            self.logger.warning(f"尋找本地路徑失敗: {str(e)}")
            
        return None
    
    def _has_target_files(self, dir_path: str) -> bool:
        """
        檢查目錄是否包含目標檔案
        
        Args:
            dir_path: 目錄路徑
            
        Returns:
            是否包含目標檔案
        """
        target_files = ['manifest.xml', 'version.txt', 'f_version.txt']
        
        try:
            files = os.listdir(dir_path)
            for file in files:
                if file.lower() in target_files:
                    return True
        except:
            pass
            
        return False            